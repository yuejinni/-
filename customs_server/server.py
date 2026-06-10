"""
本地报关文件生成服务
端口：5008
Chrome 扩展通过 POST http://localhost:5008/generate 调用
"""
import sys
import os

# PyInstaller 打包后资源路径兼容
if getattr(sys, 'frozen', False):
    _BASE = sys._MEIPASS
    # 优先从 exe 同目录加载 .py 文件（方便热更新，无需重新打包）
    _EXE_DIR = os.path.dirname(sys.executable)
    sys.path.insert(0, _EXE_DIR)
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))
    _EXE_DIR = _BASE

_DATA_BASE = _EXE_DIR if getattr(sys, 'frozen', False) else _BASE

sys.path.insert(0, _BASE)

from flask import Flask, request, jsonify, render_template_string, send_from_directory, send_file
from flask_cors import CORS
import traceback
import json
import sqlite3
import hashlib
import secrets
import hmac
import html
import subprocess
import time
import shutil
import zipfile
import re
import math
import urllib.request
from urllib.parse import urlparse
from collections import Counter

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from excel_gen import generate_all, process_tax_pdf
from ai_identify import (
    identify_and_save, identify_from_jdy,
    get_config_for_ui, save_config, get_existing_codes,
    preview_from_license, write_confirmed_license,
)
import jdy_api
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from flask import session, redirect, url_for
from werkzeug.security import check_password_hash

app = Flask(__name__)
APP_BUILD = '20260603-reorder-mvp'

_AUTH_USER = os.environ.get('QIHANG_LEGACY_USER', 'qh0001')
_AUTH_PASS = os.environ.get('QIHANG_LEGACY_PASS', 'CHANGE_ME_LEGACY_PASS')
_AUTH_SECRET_FILE = os.path.join(_DATA_BASE, 'server_secret.key')
_AUTH_DB_FILE = os.path.join(_DATA_BASE, 'auth_users.sqlite3')
_ADMIN_USER = os.environ.get('QIHANG_ADMIN_USER', 'admin')
_ADMIN_PASS = os.environ.get('QIHANG_ADMIN_PASS', 'CHANGE_ME_ADMIN_PASS')


def _load_server_secret():
    try:
        if os.path.exists(_AUTH_SECRET_FILE):
            with open(_AUTH_SECRET_FILE, 'r', encoding='utf-8') as f:
                secret = f.read().strip()
                if secret:
                    return secret
        os.makedirs(os.path.dirname(_AUTH_SECRET_FILE), exist_ok=True)
        secret = secrets.token_urlsafe(48)
        with open(_AUTH_SECRET_FILE, 'w', encoding='utf-8') as f:
            f.write(secret)
        return secret
    except Exception:
        return 'qihang-local-fallback-secret'


app.secret_key = _load_server_secret()
app.permanent_session_lifetime = timedelta(days=365)
app.config.update(
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_HTTPONLY=True,
)


def _hash_password(password, salt=''):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac('sha256', (password or '').encode('utf-8'), salt.encode('utf-8'), 120000)
    return f'pbkdf2_sha256${salt}${digest.hex()}'


def _verify_password(password, stored):
    try:
        stored = str(stored or '')
        parts = str(stored or '').split('$')
        if len(parts) == 3 and parts[0] == 'pbkdf2_sha256':
            expected = _hash_password(password, parts[1])
            return hmac.compare_digest(expected, stored)
        return check_password_hash(stored, password or '')
    except Exception:
        return False


def _auth_conn():
    os.makedirs(os.path.dirname(_AUTH_DB_FILE), exist_ok=True)
    conn = sqlite3.connect(_AUTH_DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name TEXT NOT NULL DEFAULT '',
            role TEXT NOT NULL DEFAULT 'user',
            status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL DEFAULT '',
            approved_at TEXT NOT NULL DEFAULT ''
        )
    ''')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    admin = conn.execute('SELECT id FROM users WHERE username = ?', (_ADMIN_USER,)).fetchone()
    admin_hash = _hash_password(_ADMIN_PASS)
    if admin:
        conn.execute(
            '''UPDATE users
               SET password_hash = COALESCE(NULLIF(password_hash, ''), ?), name = ?, role = 'admin', status = 'active', approved_at = COALESCE(NULLIF(approved_at, ''), ?)
               WHERE username = ?''',
            (admin_hash, '管理员', now, _ADMIN_USER),
        )
    else:
        conn.execute(
            '''INSERT INTO users (username, password_hash, name, role, status, created_at, approved_at)
               VALUES (?, ?, ?, 'admin', 'active', ?, ?)''',
            (_ADMIN_USER, admin_hash, '管理员', now, now),
        )
    conn.commit()
    return conn


def _row_to_user(row):
    if not row:
        return None
    return {
        'id': row['id'],
        'username': row['username'],
        'name': row['name'],
        'role': row['role'],
        'status': row['status'],
        'created_at': row['created_at'],
        'approved_at': row['approved_at'],
    }


def _current_user():
    if not _is_logged_in():
        return None
    return {
        'username': session.get('auth_user') or '',
        'name': session.get('auth_name') or '',
        'role': session.get('auth_role') or 'user',
        'is_admin': session.get('auth_role') == 'admin',
    }


def _is_admin():
    return _is_logged_in() and session.get('auth_role') == 'admin'


def _admin_path_required():
    admin_exact = {
        '/ai-config', '/browse-folder', '/browse-file',
        '/jdy-config', '/jdy-refresh-auth', '/jdy-test',
        '/system-update/upload', '/system-update/backups', '/system-update/rollback',
        '/system-logs', '/sales-sync-config', '/sales-sync/status', '/sales-sync-run',
        '/jdy-webhook/status', '/jdy-webhook/worker-control', '/jdy-webhook/recover-stale',
        '/jdy-webhook/failures', '/jdy-webhook/retry-failed-dry-run', '/jdy-webhook/retry-failed',
        '/jdy-webhook/resolve-recorded-only-dry-run', '/jdy-webhook/resolve-recorded-only',
        '/jdy-webhook/reclassify-retry-pending-dry-run', '/jdy-webhook/reclassify-retry-pending',
        '/jdy-webhook/archive-failed-dry-run', '/jdy-webhook/archive-failed',
        '/jdy-webhook/cleanup-old-dry-run', '/jdy-webhook/cleanup-old',
        '/jdy-webhook/accessory-diagnose', '/jdy-webhook/accessory-diagnose-live',
        '/cache/catalog-status', '/cache/product-refresh-dry-run', '/cache/product-refresh',
        '/cache/product-refresh/status', '/cache/product-refresh/cancel',
        '/sync/coverage-status', '/sync/realtime-health',
        '/supplier-cache/status', '/supplier-cache/refresh-dry-run', '/supplier-cache/refresh',
        '/sales-cache/status', '/sales-cache-refresh', '/sales-cache-cleanup',
        '/clear-supplier-cache', '/admin/users',
        '/attachments/status', '/attachments/refresh-dry-run', '/attachments/refresh',
        '/attachments/history-backfill-dry-run', '/attachments/history-backfill',
        '/customs-products/import-confirm',
    }
    if request.path in admin_exact:
        return True
    if request.path.startswith('/customs-products/') and request.method in ('PATCH', 'PUT', 'DELETE'):
        return True
    return request.path.startswith('/admin/users/')

_LOG_DIR = os.path.join(_DATA_BASE, 'logs')
_APP_LOG_FILE = os.path.join(_LOG_DIR, 'server.log')
_GENERATED_DOWNLOAD_DIR = os.path.join(_DATA_BASE, '_generated_downloads')
_GENERATED_WORK_DIR = os.path.join(_DATA_BASE, '_generated_work')


def _log_event(tag, message):
    try:
        os.makedirs(_LOG_DIR, exist_ok=True)
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(_APP_LOG_FILE, 'a', encoding='utf-8', errors='replace') as f:
            f.write(f'[{ts}] [{tag}] {message}\n')
    except Exception:
        pass


def _tail_text_file(path, max_bytes=131072):
    if not path or not os.path.exists(path):
        return ''
    max_bytes = max(1024, min(int(max_bytes or 131072), 1024 * 1024))
    with open(path, 'rb') as f:
        f.seek(0, os.SEEK_END)
        size = f.tell()
        f.seek(max(0, size - max_bytes), os.SEEK_SET)
        data = f.read()
    return data.decode('utf-8', errors='replace')


def _system_log_files():
    def latest_update_log():
        update_dir = os.path.join(_DATA_BASE, '_updates')
        names = ['apply_update.log']
        try:
            names.extend(
                name for name in os.listdir(update_dir)
                if name.lower().startswith('update_') and name.lower().endswith('.log')
            )
        except Exception:
            pass
        paths = [os.path.join(update_dir, name) for name in names]
        paths = [path for path in paths if os.path.exists(path)]
        if not paths:
            return os.path.join(update_dir, 'apply_update.log')
        return max(paths, key=lambda path: os.path.getmtime(path))

    candidates = {
        'server': ('业务日志 server.log', os.path.join(_LOG_DIR, 'server.log')),
        'watchdog': ('守护进程 watchdog.log', os.path.join(_LOG_DIR, 'watchdog.log')),
        'stdout': ('服务输出 server_stdout.log', os.path.join(_LOG_DIR, 'server_stdout.log')),
        'stderr': ('服务错误 server_stderr.log', os.path.join(_LOG_DIR, 'server_stderr.log')),
        'update': ('在线更新日志', latest_update_log()),
        'runtime_out': ('本地运行输出 server_runtime.out.log', os.path.join(_DATA_BASE, 'server_runtime.out.log')),
        'runtime_err': ('本地运行错误 server_runtime.err.log', os.path.join(_DATA_BASE, 'server_runtime.err.log')),
        'sales_sync_out': ('销售同步输出 sales_quantity_sync.out.log', os.path.join(_DATA_BASE, 'sales_quantity_sync.out.log')),
        'sales_sync_err': ('销售同步错误 sales_quantity_sync.err.log', os.path.join(_DATA_BASE, 'sales_quantity_sync.err.log')),
    }
    items = []
    for key, (label, path) in candidates.items():
        exists = os.path.exists(path)
        try:
            stat = os.stat(path) if exists else None
            size = stat.st_size if stat else 0
            mtime = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S') if stat else ''
        except Exception:
            size = 0
            mtime = ''
        items.append({'key': key, 'label': label, 'exists': exists, 'size': size, 'mtime': mtime})
    return items, {key: path for key, (_, path) in candidates.items()}


def _is_logged_in():
    if session.get('auth_ok') is not True:
        return False
    username = session.get('auth_user') or ''
    if not username:
        session.clear()
        return False
    try:
        with _auth_conn() as conn:
            row = conn.execute('SELECT username, name, role, status FROM users WHERE username = ?', (username,)).fetchone()
        if not row or row['status'] != 'active':
            session.clear()
            return False
        session['auth_name'] = row['name']
        session['auth_role'] = row['role']
        return True
    except Exception:
        return False


def _wants_json():
    accept = request.headers.get('Accept', '')
    return (
        request.path.startswith('/api/')
        or request.is_json
        or 'application/json' in accept
        or request.path not in ('/', '/jdy', '/login')
    )


@app.before_request
def require_login():
    if request.method == 'OPTIONS':
        return None
    public_paths = {'/health', '/login', '/register', '/favicon.ico', '/jdy-webhook'}
    if request.path in public_paths:
        return None
    if _is_logged_in():
        if _admin_path_required() and not _is_admin():
            if _wants_json():
                error = '只有管理员可以导入或维护报关商品资料' if request.path.startswith('/customs-products/') else '只有管理员可以进入设置'
                return jsonify({'success': False, 'error': error}), 403
            return redirect('/jdy')
        return None
    if _wants_json():
        return jsonify({'success': False, 'error': '请先登录'}), 401
    return redirect(url_for('login_page', next=request.full_path or '/jdy'))


@app.route('/')
def home():
    return redirect('/jdy')


def _legacy_login_page_unused():
    error = ''
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        remember = (request.form.get('remember') or '1') in ('1', 'on', 'true', 'yes')
        if hmac.compare_digest(username, _AUTH_USER) and hmac.compare_digest(password, _AUTH_PASS):
            session.clear()
            session.permanent = remember
            session['auth_ok'] = True
            session['auth_user'] = username
            nxt = request.args.get('next') or '/jdy'
            if not nxt.startswith('/'):
                nxt = '/jdy'
            return redirect(nxt)
        error = '账号或密码错误'
    return f'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>祺航本地项目登录</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;background:#eef3f7;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;color:#1f2d3d}}
.box{{width:min(92vw,380px);background:#fff;border:1px solid #dde6ee;border-radius:14px;box-shadow:0 18px 45px rgba(30,55,90,.14);padding:28px}}
h1{{font-size:24px;margin:0 0 8px;text-align:center}}p{{margin:0 0 22px;text-align:center;color:#788797;font-size:14px}}
label{{display:block;font-size:13px;color:#5f6f7f;margin:14px 0 6px}}input{{width:100%;height:44px;border:1px solid #cfd9e3;border-radius:8px;padding:0 12px;font-size:16px;outline:none}}input:focus{{border-color:#0cbfbd;box-shadow:0 0 0 3px rgba(12,191,189,.14)}}
.password-wrap{{position:relative}}.password-wrap input{{padding-right:48px}}.eye-btn{{position:absolute;right:6px;top:6px;width:34px;height:32px;margin:0;border:0;border-radius:6px;background:#eef3f7;color:#465565;font-size:15px;cursor:pointer}}
.remember{{display:flex;align-items:center;gap:8px;margin-top:14px;color:#5f6f7f;font-size:14px}}.remember input{{width:18px;height:18px;padding:0}}
button{{width:100%;height:44px;margin-top:20px;border:0;border-radius:8px;background:#0cbfbd;color:#fff;font-size:16px;font-weight:700;cursor:pointer}}
.err{{min-height:20px;margin-top:12px;color:#d93025;text-align:center;font-size:13px}}
</style>
</head>
<body>
  <form class="box" method="post">
    <h1>祺航本地项目</h1>
    <p>请输入账号密码后继续访问</p>
    <label>账号</label>
    <input name="username" autocomplete="username" autofocus>
    <label>密码</label>
    <div class="password-wrap">
      <input id="password" name="password" type="password" autocomplete="current-password">
      <button class="eye-btn" type="button" onclick="const p=document.getElementById('password');p.type=p.type==='password'?'text':'password';this.textContent=p.type==='password'?'👁':'🙈';">👁</button>
    </div>
    <label class="remember"><input name="remember" type="checkbox" value="1" checked> 保持登录</label>
    <button type="submit">登录</button>
    <div class="err">{error}</div>
  </form>
</body>
</html>'''


@app.route('/logout', methods=['POST', 'GET'])
def logout_page():
    session.clear()
    return redirect('/login')


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    error = ''
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        remember = (request.form.get('remember') or '1') in ('1', 'on', 'true', 'yes')
        with _auth_conn() as conn:
            row = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if row and _verify_password(password, row['password_hash']):
            if row['status'] != 'active':
                error = '账号已提交注册申请，请等待管理员通过'
            else:
                session.clear()
                session.permanent = remember
                session['auth_ok'] = True
                session['auth_user'] = row['username']
                session['auth_name'] = row['name']
                session['auth_role'] = row['role']
                nxt = request.args.get('next') or '/jdy'
                if not nxt.startswith('/'):
                    nxt = '/jdy'
                return redirect(nxt)
        else:
            if hmac.compare_digest(username, _AUTH_USER) and hmac.compare_digest(password, _AUTH_PASS):
                session.clear()
                session.permanent = remember
                session['auth_ok'] = True
                session['auth_user'] = username
                session['auth_name'] = username
                session['auth_role'] = 'admin' if username == _ADMIN_USER else 'user'
                nxt = request.args.get('next') or '/jdy'
                if not nxt.startswith('/'):
                    nxt = '/jdy'
                return redirect(nxt)
            error = '账号或密码错误'
    return f'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>祺航本地项目登录</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;background:#eef3f7;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;color:#1f2d3d}}
.box{{width:min(92vw,380px);background:#fff;border:1px solid #dde6ee;border-radius:14px;box-shadow:0 18px 45px rgba(30,55,90,.14);padding:28px}}
h1{{font-size:24px;margin:0 0 8px;text-align:center}}p{{margin:0 0 22px;text-align:center;color:#788797;font-size:14px}}
label{{display:block;font-size:13px;color:#5f6f7f;margin:14px 0 6px}}input{{width:100%;height:44px;border:1px solid #cfd9e3;border-radius:8px;padding:0 12px;font-size:16px;outline:none}}input:focus{{border-color:#0cbfbd;box-shadow:0 0 0 3px rgba(12,191,189,.14)}}
.password-wrap{{position:relative}}.password-wrap input{{padding-right:52px}}.eye-btn{{position:absolute;right:6px;top:6px;width:42px;height:32px;margin:0;border:0;border-radius:6px;background:#eef3f7;color:#465565;font-size:13px;cursor:pointer}}
.remember{{display:flex;align-items:center;gap:8px;margin-top:14px;color:#5f6f7f;font-size:14px}}.remember input{{width:18px;height:18px;padding:0}}
button{{width:100%;height:44px;margin-top:20px;border:0;border-radius:8px;background:#0cbfbd;color:#fff;font-size:16px;font-weight:700;cursor:pointer}}
.err{{min-height:20px;margin-top:12px;color:#d93025;text-align:center;font-size:13px}}.link{{display:block;text-align:center;margin-top:14px;color:#0b6fbd;text-decoration:none;font-size:14px}}
</style>
</head>
<body>
  <form class="box" method="post">
    <h1>祺航本地项目</h1>
    <p>请输入账号密码后继续访问</p>
    <label>账号</label>
    <input name="username" autocomplete="username" autofocus>
    <label>密码</label>
    <div class="password-wrap">
      <input id="password" name="password" type="password" autocomplete="current-password">
      <button class="eye-btn" type="button" onclick="const p=document.getElementById('password');p.type=p.type==='password'?'text':'password';this.textContent=p.type==='password'?'显示':'隐藏';">显示</button>
    </div>
    <label class="remember"><input name="remember" type="checkbox" value="1" checked> 保持登录</label>
    <button type="submit">登录</button>
    <div class="err">{error}</div>
    <a class="link" href="/register">申请注册账号</a>
  </form>
</body>
</html>'''

@app.route('/register', methods=['GET', 'POST'])
def register_page():
    error = ''
    ok = ''
    username = ''
    name = ''
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        name = (request.form.get('name') or '').strip()
        password = request.form.get('password') or ''
        confirm = request.form.get('confirm') or ''
        if not username or not name or not password:
            error = '请完整填写登录账户、姓名和密码'
        elif password != confirm:
            error = '两次输入的密码不一致'
        elif len(password) < 6:
            error = '密码至少需要 6 位'
        else:
            try:
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                with _auth_conn() as conn:
                    conn.execute(
                        '''INSERT INTO users (username, password_hash, name, role, status, created_at, approved_at)
                           VALUES (?, ?, ?, 'user', 'pending', ?, '')''',
                        (username, _hash_password(password), name, now),
                    )
                    conn.commit()
                ok = '注册申请已提交，请等待管理员在设置页审核通过'
                username = ''
                name = ''
            except sqlite3.IntegrityError:
                error = '这个登录账户已经存在'
            except Exception as e:
                error = f'注册失败：{e}'
    return f'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>申请注册账号</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;background:#eef3f7;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;color:#1f2d3d}}
.box{{width:min(92vw,420px);background:#fff;border:1px solid #dde6ee;border-radius:14px;box-shadow:0 18px 45px rgba(30,55,90,.14);padding:28px}}
h1{{font-size:24px;margin:0 0 8px;text-align:center}}p{{margin:0 0 22px;text-align:center;color:#788797;font-size:14px}}
label{{display:block;font-size:13px;color:#5f6f7f;margin:14px 0 6px}}input{{width:100%;height:44px;border:1px solid #cfd9e3;border-radius:8px;padding:0 12px;font-size:16px;outline:none}}input:focus{{border-color:#0cbfbd;box-shadow:0 0 0 3px rgba(12,191,189,.14)}}
button{{width:100%;height:44px;margin-top:20px;border:0;border-radius:8px;background:#0cbfbd;color:#fff;font-size:16px;font-weight:700;cursor:pointer}}
.err{{min-height:20px;margin-top:12px;color:#d93025;text-align:center;font-size:13px}}.ok{{min-height:20px;margin-top:12px;color:#19a453;text-align:center;font-size:13px}}.link{{display:block;text-align:center;margin-top:14px;color:#0b6fbd;text-decoration:none;font-size:14px}}
</style>
</head>
<body>
  <form class="box" method="post">
    <h1>申请注册账号</h1>
    <p>提交后需要管理员在设置页审核通过</p>
    <label>登录账户</label>
    <input name="username" value="{username}" autocomplete="username" autofocus>
    <label>姓名</label>
    <input name="name" value="{name}" autocomplete="name">
    <label>密码</label>
    <input name="password" type="password" autocomplete="new-password">
    <label>确认密码</label>
    <input name="confirm" type="password" autocomplete="new-password">
    <button type="submit">提交注册申请</button>
    <div class="err">{error}</div>
    <div class="ok">{ok}</div>
    <a class="link" href="/login">返回登录</a>
  </form>
</body>
</html>'''

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# API 结果缓存（按 code，24小时过期）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _register_page_v2():
    error = ''
    ok = ''
    username = ''
    name = ''
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        name = (request.form.get('name') or '').strip()
        password = request.form.get('password') or ''
        confirm = request.form.get('confirm') or ''
        if not username or not name or not password:
            error = '请完整填写登录账户、密码、确认密码和姓名'
        elif password != confirm:
            error = '两次输入的密码不一致'
        elif len(password) < 6:
            error = '密码至少需要 6 位'
        else:
            try:
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                with _auth_conn() as conn:
                    conn.execute(
                        '''INSERT INTO users (username, password_hash, name, role, status, created_at, approved_at)
                           VALUES (?, ?, ?, 'user', 'pending', ?, '')''',
                        (username, _hash_password(password), name, now),
                    )
                    conn.commit()
                ok = '申请已提交，等待管理员审核。审核通过后就可以登录使用。'
                username = ''
                name = ''
            except sqlite3.IntegrityError:
                error = '这个登录账户已经存在，请换一个账户名'
            except Exception as e:
                error = f'注册失败：{e}'
    success_block = f'<div class="ok">{html.escape(ok)}</div>' if ok else ''
    error_block = f'<div class="err">{html.escape(error)}</div>' if error else ''
    return f'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>申请注册账号</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;background:#eef3f7;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;color:#1f2d3d}}
.box{{width:min(92vw,420px);background:#fff;border:1px solid #dde6ee;border-radius:14px;box-shadow:0 18px 45px rgba(30,55,90,.14);padding:28px}}
h1{{font-size:24px;margin:0 0 8px;text-align:center}}p{{margin:0 0 22px;text-align:center;color:#788797;font-size:14px}}
label{{display:block;font-size:13px;color:#5f6f7f;margin:14px 0 6px}}input{{width:100%;height:44px;border:1px solid #cfd9e3;border-radius:8px;padding:0 12px;font-size:16px;outline:none}}input:focus{{border-color:#0cbfbd;box-shadow:0 0 0 3px rgba(12,191,189,.14)}}
button{{width:100%;height:44px;margin-top:20px;border:0;border-radius:8px;background:#0cbfbd;color:#fff;font-size:16px;font-weight:700;cursor:pointer}}
.err{{margin-top:12px;color:#d93025;text-align:center;font-size:14px}}.ok{{margin-top:14px;padding:12px;border-radius:8px;background:#e9f8ef;color:#16823c;text-align:center;font-size:14px;line-height:1.6}}.link{{display:block;text-align:center;margin-top:14px;color:#0b6fbd;text-decoration:none;font-size:14px}}
</style>
</head>
<body>
  <form class="box" method="post">
    <h1>申请注册账号</h1>
    <p>提交后需要管理员在设置页审核通过</p>
    <label>登录账户</label>
    <input name="username" value="{html.escape(username)}" autocomplete="username" autofocus>
    <label>姓名</label>
    <input name="name" value="{html.escape(name)}" autocomplete="name">
    <label>密码</label>
    <input name="password" type="password" autocomplete="new-password">
    <label>确认密码</label>
    <input name="confirm" type="password" autocomplete="new-password">
    <button type="submit">提交注册申请</button>
    {error_block}
    {success_block}
    <a class="link" href="/login">返回登录</a>
  </form>
</body>
</html>'''


app.view_functions['register_page'] = _register_page_v2


@app.route('/auth/me', methods=['GET'])
def auth_me():
    user = _current_user()
    return jsonify({'success': True, 'user': user, 'is_admin': bool(user and user.get('is_admin'))})


@app.route('/admin/users', methods=['GET'])
def admin_users():
    with _auth_conn() as conn:
        rows = conn.execute(
            '''SELECT id, username, name, role, status, created_at, approved_at
               FROM users
               ORDER BY CASE status WHEN 'pending' THEN 0 WHEN 'active' THEN 1 ELSE 2 END, id DESC'''
        ).fetchall()
    return jsonify({'success': True, 'list': [_row_to_user(row) for row in rows]})


@app.route('/admin/users/<int:user_id>/approve', methods=['POST'])
def admin_user_approve(user_id):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _auth_conn() as conn:
        row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': '用户不存在'}), 404
        if row['role'] == 'admin':
            return jsonify({'success': False, 'error': '管理员账号不需要审核'}), 400
        conn.execute(
            "UPDATE users SET status = 'active', approved_at = ? WHERE id = ?",
            (now, user_id),
        )
        conn.commit()
    return jsonify({'success': True})


@app.route('/admin/users/<int:user_id>/reject', methods=['POST'])
def admin_user_reject(user_id):
    with _auth_conn() as conn:
        row = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if not row:
            return jsonify({'success': False, 'error': '用户不存在'}), 404
        if row['role'] == 'admin':
            return jsonify({'success': False, 'error': '不能拒绝管理员账号'}), 400
        conn.execute("UPDATE users SET status = 'rejected' WHERE id = ?", (user_id,))
        conn.commit()
    return jsonify({'success': True})


_CACHE_FILE  = os.path.join(_DATA_BASE, '_api_cache', 'code_cache.json')
_CACHE_TTL_H = 24
_cache_lock   = threading.Lock()
_history_lock = threading.Lock()   # generation_history.json 写锁

def _load_code_cache():
    try:
        if os.path.exists(_CACHE_FILE):
            with open(_CACHE_FILE, encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_code_cache(cache):
    os.makedirs(os.path.dirname(_CACHE_FILE), exist_ok=True)
    with _cache_lock:
        with open(_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)

def _cache_valid(entry):
    """检查缓存条目是否在 TTL 内有效"""
    ts = entry.get('fetched_at', '')
    if not ts:
        return False
    try:
        return datetime.now() - datetime.strptime(ts, '%Y-%m-%d %H:%M:%S') < timedelta(hours=_CACHE_TTL_H)
    except Exception:
        return False
CORS(app, origins='*')   # 允许 Chrome 扩展跨域调用

@app.after_request
def allow_private_network(response):
    # Chrome Private Network Access：允许公网页面访问本地服务
    response.headers['Access-Control-Allow-Private-Network'] = 'true'
    return response

def _env_int(name, default):
    try:
        return int(str(os.environ.get(name, '')).strip() or default)
    except Exception:
        return default


def _workers_disabled():
    return str(os.environ.get('QIHANG_DISABLE_WORKERS', '')).strip().lower() in ('1', 'true', 'yes', 'on')


PORT = _env_int('QIHANG_PORT', 5008)


@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'port': PORT, 'build': APP_BUILD})


def _fetch_safe_exchange_rate():
    """
    从国家外汇管理局获取本月第一个工作日的美元汇率中间价。
    返回 (rate: float, date_str: str) 或 (None, None)。
    """
    from datetime import date, timedelta
    import urllib.request as _req
    from urllib.parse import urlencode
    import ssl as _ssl
    import re as _re

    today = date.today()
    d = date(today.year, today.month, 1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    end_d = min(d + timedelta(days=13), today)
    payload = urlencode({
        'startDate': d.strftime('%Y-%m-%d'),
        'endDate':   end_d.strftime('%Y-%m-%d'),
        'queryYN':   'true',
    }).encode('utf-8')

    _ctx = _ssl.create_default_context()
    _ctx.check_hostname = False
    _ctx.verify_mode    = _ssl.CERT_NONE

    req = _req.Request(
        'https://www.safe.gov.cn/AppStructured/hlw/RMBQuery.do',
        data=payload,
        headers={
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent':   'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Referer':      'https://www.safe.gov.cn/safe/rmbhlzjj/index.html',
        }
    )
    with _req.urlopen(req, timeout=12, context=_ctx) as r:
        html = r.read().decode('utf-8', errors='replace')

    rows = _re.findall(r'<tr[^>]*>(.*?)</tr>', html, _re.DOTALL | _re.IGNORECASE)
    for row in rows:
        cells = _re.findall(r'<td[^>]*>(.*?)</td>', row, _re.DOTALL | _re.IGNORECASE)
        cells = [_re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        if len(cells) < 2:
            continue
        date_cell = cells[0]
        usd_cell  = cells[1]
        if not _re.match(r'\d{4}[-/]\d{2}[-/]\d{2}', date_cell):
            continue
        try:
            usd_val  = float(usd_cell)
            rate     = round(usd_val / 100, 4)
            row_date = _re.sub(r'[-/]', '-', date_cell[:10])
            return rate, row_date
        except ValueError:
            continue
    return None, None


@app.route('/get-exchange-rate', methods=['GET'])
def get_exchange_rate():
    """
    从国家外汇管理局获取本月第一个工作日（Mon-Fri）的美元汇率中间价。
    接口：POST https://www.safe.gov.cn/AppStructured/hlw/RMBQuery.do
    返回 HTML 表格，美元列值为"100美元兑人民币"，除以 100 得实际汇率。
    """
    try:
        rate, row_date = _fetch_safe_exchange_rate()
        if rate:
            print(f'[RATE] SAFE 汇率 {row_date}: {rate}')
            return jsonify({'success': True, 'rate': rate,
                            'date': row_date, 'source': 'SAFE外汇局'})
        from datetime import date
        d = date.today().replace(day=1)
        return jsonify({'success': False,
                        'error': f'外汇局未找到 {d} 起的汇率数据，请手动填写'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/browse-folder', methods=['GET'])
def browse_folder():
    """调用系统原生文件夹选择对话框（macOS 用 osascript，Windows 用 tkinter）"""
    import platform
    system = platform.system()

    if system == 'Darwin':  # macOS：用 AppleScript，不依赖 tkinter
        try:
            import subprocess
            script = 'POSIX path of (choose folder with prompt "选择报关文件保存文件夹")'
            result = subprocess.run(
                ['osascript', '-e', script],
                capture_output=True, text=True, timeout=120
            )
            folder = result.stdout.strip().rstrip('/')
            if result.returncode == 0 and folder:
                return jsonify({'success': True, 'path': folder})
            return jsonify({'success': False, 'path': ''})
        except subprocess.TimeoutExpired:
            return jsonify({'success': False, 'path': ''})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    else:  # Windows：用 tkinter
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            folder = filedialog.askdirectory(title='选择报关文件保存文件夹')
            root.destroy()
            if folder:
                return jsonify({'success': True, 'path': folder})
            return jsonify({'success': False, 'path': ''})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})


@app.route('/browse-file', methods=['GET'])
def browse_file():
    """调用系统原生文件选择对话框（macOS 用 osascript，Windows 用 tkinter）"""
    import platform
    system = platform.system()

    if system == 'Darwin':
        try:
            import subprocess
            script = 'POSIX path of (choose file with prompt "选择退税联 PDF 文件" of type {"pdf"})'
            result = subprocess.run(
                ['osascript', '-e', script],
                capture_output=True, text=True, timeout=120
            )
            path = result.stdout.strip()
            if result.returncode == 0 and path:
                return jsonify({'success': True, 'path': path})
            return jsonify({'success': False, 'path': ''})
        except subprocess.TimeoutExpired:
            return jsonify({'success': False, 'path': ''})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    else:  # Windows
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            path = filedialog.askopenfilename(
                title='选择退税联 PDF 文件',
                filetypes=[('PDF 文件', '*.pdf'), ('所有文件', '*.*')]
            )
            root.destroy()
            if path:
                return jsonify({'success': True, 'path': path})
            return jsonify({'success': False, 'path': ''})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})


def _cleanup_old_generated_downloads(max_age_hours=24):
    now = time.time()
    for folder in (_GENERATED_DOWNLOAD_DIR, _GENERATED_WORK_DIR):
        try:
            if not os.path.isdir(folder):
                continue
            for name in os.listdir(folder):
                path = os.path.join(folder, name)
                try:
                    if now - os.path.getmtime(path) < max_age_hours * 3600:
                        continue
                    if os.path.isdir(path):
                        shutil.rmtree(path, ignore_errors=True)
                    else:
                        os.remove(path)
                except Exception:
                    pass
        except Exception:
            pass


def _zip_generated_folder(folder, invoice_no=''):
    os.makedirs(_GENERATED_DOWNLOAD_DIR, exist_ok=True)
    _cleanup_old_generated_downloads()
    safe_invoice = ''.join(c if c.isalnum() or c in ('-', '_') else '_' for c in str(invoice_no or 'generated'))[:60]
    filename = f'报关单_{safe_invoice}_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{secrets.token_hex(4)}.zip'
    zip_path = os.path.join(_GENERATED_DOWNLOAD_DIR, filename)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder):
            for file in files:
                full = os.path.join(root, file)
                rel = os.path.relpath(full, folder)
                zf.write(full, rel)
    return filename, zip_path


@app.route('/generated-download/<path:filename>', methods=['GET'])
def generated_download(filename):
    safe_name = os.path.basename(filename)
    path = os.path.join(_GENERATED_DOWNLOAD_DIR, safe_name)
    if not os.path.isfile(path):
        return jsonify({'success': False, 'error': '下载文件不存在或已过期，请重新生成'}), 404
    return send_file(path, as_attachment=True, download_name=safe_name, mimetype='application/zip')


@app.route('/generate', methods=['POST'])
def generate():
    """
    请求体 JSON：
    {
        "items": [
            {"code": "T.26431T-24", "unit": "PAC", "qty": 1.0, "rmb_price": 89.0},
            ...
        ],
        "exchange_rate": 7.2,
        "invoice_no": "LK00022",
        "target_total_usd": null,      // 可选，目标总金额
        "output_path": "C:\\Users\\xxx\\Desktop\\export"
    }
    """
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'success': False, 'error': '无效的请求体'}), 400

        required = ['items', 'exchange_rate']
        for field in required:
            if field not in data:
                return jsonify({'success': False, 'error': f'缺少字段：{field}'}), 400

        output_path = str(data.get('output_path') or '').strip()
        download_mode = bool(data.get('download_mode')) or not output_path
        if download_mode:
            os.makedirs(_GENERATED_WORK_DIR, exist_ok=True)
            work_name = f'gen_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{secrets.token_hex(4)}'
            output_path = os.path.join(_GENERATED_WORK_DIR, work_name)
            data['output_path'] = output_path

        # 调试日志：记录前端传来的原始数据
        items = data.get('items', [])
        if items:
            print(f'[DEBUG] 收到 {len(items)} 个商品，前3个: ')
            for it in items[:3]:
                print(f'  code={it.get("code")}, qty={it.get("qty")}, unit={it.get("unit")}, rmb_price={it.get("rmb_price")}, boxes={it.get("boxes")}')
            # 找 F.3153F-296
            for it in items:
                if '3153F' in str(it.get('code', '')):
                    print(f'  [TARGET] code={it.get("code")}, qty={it.get("qty")}, unit={it.get("unit")}, boxes={it.get("boxes")}')

        # 提取 order_refs 信息（单号/备注/账套，用于子文件夹命名 + 限定查询账套）
        order_refs = data.get('order_refs') or []
        first_ref = order_refs[0] if order_refs else {}
        data['order_no']     = (first_ref.get('number') or first_ref.get('id') or '').strip()
        data['order_remark'] = (first_ref.get('remark') or '').strip()

        # 根据调拨单账套决定只查对应的 JDY 账套，不允许跨账套查任何信息
        order_account = (first_ref.get('account') or '').strip()
        _cfg2 = _load_jdy_config2()
        if order_account and order_account == _cfg2.get('name'):
            preferred_cli_fn = _ensure_jdy_client2
        elif order_account:
            preferred_cli_fn = _ensure_jdy_client
        else:
            preferred_cli_fn = None  # 未知账套，仍查两个
        _cli_label = 'cli2' if preferred_cli_fn is _ensure_jdy_client2 else ('cli1' if preferred_cli_fn else '两个')
        print(f'[SUPPLIER] 调拨单账套={order_account!r}, 使用客户端={_cli_label}')

        # 多账套混选时，每个 item 携带自己的 account，按 code 建立精确账套映射
        code_account_map = {}
        for it in items:
            _code = (it.get('code') or '').strip()
            _acct = (it.get('account') or '').strip()
            if _code and _acct:
                if _acct == _cfg2.get('name'):
                    code_account_map[_code] = _ensure_jdy_client2
                else:
                    code_account_map[_code] = _ensure_jdy_client
        if code_account_map:
            _accts = set(fn.__name__ for fn in code_account_map.values())
            print(f'[SUPPLIER] per-code 账套映射: {len(code_account_map)} 个商品, 账套: {_accts}')

        # ── 手工填写的 proLicense（格式：品名*材料*分类简称，来自前端弹窗）──
        manual_licenses = data.get('manual_licenses') or []
        if manual_licenses:
            from ai_identify import fill_from_license, _load_config, _find_excel
            from jdy_api import parse_dimensions as _parse_dims
            _ai_cfg     = _load_ai_config()
            _excel_path = _find_excel(_ai_cfg)
            for ml in manual_licenses:
                _code    = (ml.get('code') or '').strip()
                _raw     = (ml.get('pro_license') or '').strip()   # 格式：品名*材料*分类简称
                if not (_code and _raw):
                    continue
                # 拆分：至少需要 品名 和 材料
                _parts   = [p.strip() for p in _raw.split('*')]
                _name    = _parts[0] if len(_parts) > 0 else ''
                _mat     = _parts[1] if len(_parts) > 1 else ''
                _cat_str = _parts[2] if len(_parts) > 2 else ''
                if not (_name and _mat):
                    print(f'[WARN] manual_license {_code}: 格式错误 "{_raw}"，需要 品名*材料')
                    continue
                # 从 JDY 取 product id / dimensions 等额外字段，并回写 proLicense
                _extra = {}
                _prod_id = None
                try:
                    _cli_fn = (code_account_map or {}).get(_code) or preferred_cli_fn
                    _cli = _cli_fn() if _cli_fn else (_ensure_jdy_client() or _ensure_jdy_client2())
                    if _cli:
                        _prod = _cli.get_product_by_code(_code)
                        if _prod:
                            _prod_id = _prod.get('id')
                            _cfg2  = _load_jdy_config2()
                            _acct  = _cfg2['name'] if _cli is _ensure_jdy_client2() else _load_jdy_config()['name']
                            _extra = {
                                'category':   _cat_str or _prod.get('categoryName') or '',
                                'dimensions': _parse_dims(_prod.get('registrationNo') or ''),
                                'account':    _acct,
                            }
                            # 回写 proLicense 到精斗云
                            try:
                                _cli.update_product_pro_license(_prod_id, _raw)
                                print(f'[LICENSE MANUAL] {_code} proLicense 已回写 JDY: {_raw!r}')
                            except Exception as _we:
                                print(f'[WARN] {_code} 回写 JDY 失败: {_we}')
                except Exception as _e:
                    print(f'[WARN] manual_license extra fields {_code}: {_e}')
                if _excel_path:
                    try:
                        fill_from_license(_excel_path, _code, _name, _mat,
                                          extra_fields=_extra, cat_str=_cat_str or None)
                        print(f'[LICENSE MANUAL] {_code} 手工补全完成: {_name}/{_mat}' +
                              (f', 分类简称: {_cat_str}' if _cat_str else ''))
                    except Exception as _e:
                        print(f'[WARN] manual_license fill {_code}: {_e}')

        # ── 生成前自动档案补全：缺失商品先用 proLicense 填表（限定同账套）──
        auto_filled, no_license_items = _auto_license_fill(
            items, preferred_cli_fn=preferred_cli_fn, code_account_map=code_account_map)
        if auto_filled:
            print(f'[LICENSE AUTO] 自动补全 {len(auto_filled)} 个缺失商品: {auto_filled}')

        # proLicense 为空的商品 → 让前端弹窗让用户手工填写，再次提交
        if no_license_items and not manual_licenses:
            return jsonify({
                'success': True,
                'needs_manual_license': True,
                'no_license_items': no_license_items,
            })

        # 查询供应商映射（用于生成采购合同）
        supplier_overrides = data.get('supplier_overrides') or {}
        try:
            supplier_map, needs_select = _lookup_suppliers(
                items,
                preferred_cli_fn=preferred_cli_fn,
                supplier_overrides=supplier_overrides,
                code_account_map=code_account_map,
            )
            data['supplier_map'] = supplier_map
            print(f'[SUPPLIER] 查到 {len(supplier_map)} 个商品的供应商，待选 {len(needs_select)} 个')
        except Exception as e:
            print(f'[WARN] 供应商查询失败: {e}')
            data['supplier_map'] = {}
            needs_select = {}

        # 若有商品供应商待选，先返回给前端弹窗确认
        if needs_select:
            return jsonify({
                'success': True,
                'needs_supplier_select': True,
                'select_items': needs_select,
            })

        # 批量查 API 尺寸（registrationNo），缓存命中直接跳过
        try:
            cache = _load_code_cache()
            all_codes = list({(it.get('code') or '').strip() for it in items if it.get('code')})

            # 区分：缓存有效 vs 需要重新拉取
            api_dims_map = {}
            for c in all_codes:
                if _cache_valid(cache.get(c, {})) and 'dims' in cache.get(c, {}):
                    api_dims_map[c] = cache[c]['dims']   # 直接用缓存

            codes_need_dims = [c for c in all_codes if c not in api_dims_map]

            if codes_need_dims:
                _dim_fn_to_codes = {}
                _dim_no_acct = []
                for _code in codes_need_dims:
                    _fn = code_account_map.get(_code) or preferred_cli_fn
                    if _fn:
                        _dim_fn_to_codes.setdefault(_fn, []).append(_code)
                    else:
                        _dim_no_acct.append(_code)
                for _fn2 in [_ensure_jdy_client, _ensure_jdy_client2]:
                    _dim_fn_to_codes.setdefault(_fn2, []).extend(_dim_no_acct)

                now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for _fn, _codes in _dim_fn_to_codes.items():
                    if not _codes:
                        continue
                    try:
                        cli = _fn()
                        if not cli:
                            continue
                        for i in range(0, len(_codes), 30):
                            batch = cli.get_products_by_codes(_codes[i:i+30])
                            for prod in (batch or []):
                                c   = prod.get('number') or prod.get('code', '')
                                reg = prod.get('registrationNo', '')
                                if c and reg:
                                    d = jdy_api.parse_dimensions(reg)
                                    if d:
                                        api_dims_map[c] = d
                                        # 写入磁盘缓存
                                        if c not in cache:
                                            cache[c] = {}
                                        cache[c]['dims'] = d
                                        cache[c].setdefault('fetched_at', now_str)
                    except Exception as e:
                        print(f'[WARN] api_dims_map {_fn.__name__}: {e}')
                _save_code_cache(cache)

            data['api_dims_map'] = api_dims_map
            print(f'[DIMS] API 尺寸 {len(api_dims_map)} 个（其中 {len(all_codes)-len(codes_need_dims)} 个命中缓存）')
        except Exception as e:
            print(f'[WARN] api_dims 失败: {e}')
            data['api_dims_map'] = {}

        # total_cbm：前端可传入，默认满柜 68CBM
        total_cbm = data.get('total_cbm')
        if total_cbm is None:
            total_cbm = 68.0
        data['total_cbm'] = float(total_cbm)

        result = generate_all(data)
        result['auto_filled'] = auto_filled   # 返回给前端提示
        # 记录哪些调拨单被用于生成报关单
        order_refs = data.get('order_refs') or []
        if order_refs:
            _record_gen_history(order_refs)
        if download_mode:
            zip_name, zip_path = _zip_generated_folder(output_path, data.get('invoice_no') or '')
            result['download_url'] = url_for('generated_download', filename=zip_name)
            result['download_name'] = zip_name
            result['download_size'] = os.path.getsize(zip_path) if os.path.exists(zip_path) else 0
        return jsonify({'success': True, **result})

    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/ai-config', methods=['GET'])
def ai_config_get():
    try:
        return jsonify({'success': True, **get_config_for_ui()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/ai-config', methods=['POST'])
def ai_config_set():
    try:
        data = request.get_json(force=True) or {}
        save_config(data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/ai-codes', methods=['GET'])
def ai_codes():
    """返回 Excel B 列所有已识别的编码集合，供前端批量预判断"""
    try:
        from ai_identify import _load_config, _find_excel
        config = _load_config()
        excel_path = _find_excel(config)
        if not excel_path:
            return jsonify({'success': True, 'codes': []})
        codes = get_existing_codes(excel_path)
        return jsonify({'success': True, 'codes': codes})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/ai-identify', methods=['POST'])
def ai_identify():
    """
    请求体 JSON：
    {
        "code":      "T.F240T-308",
        "name":      "荔枝纹斜挎包",    // 可选
        "category":  "Lady Bag",        // 可选
        "image_url": "https://...",     // 与 image_b64 二选一
        "image_b64": "..."
    }
    """
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'success': False, 'error': '无效请求体'}), 400
        result = identify_and_save(data)
        return jsonify({'success': True, **result})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 精斗云 (JDY) 相关端点
# ──────────────────────────────────────────────────────────────────────────────

def _runtime_config_candidates():
    candidates = []
    env_path = os.environ.get('QIHANG_CONFIG_PATH') or os.environ.get('QIHANG_CONFIG_FILE')
    if env_path:
        candidates.append(env_path)
    candidates.extend([
        os.path.join(os.getcwd(), 'ai_config.json'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ai_config.json'),
        os.path.join(_DATA_BASE, 'ai_config.json'),
        os.path.join(_EXE_DIR, 'ai_config.json'),
    ])
    try:
        from ai_identify import _CONFIG_FILE
        candidates.append(_CONFIG_FILE)
    except Exception:
        pass
    candidates.append(os.path.join(os.path.dirname(_DATA_BASE), 'ai_config.json'))
    result = []
    seen = set()
    for path in candidates:
        if not path:
            continue
        full = os.path.abspath(path)
        key = os.path.normcase(full)
        if key not in seen:
            seen.add(key)
            result.append(full)
    return result


def _runtime_config_file():
    for path in _runtime_config_candidates():
        if os.path.exists(path):
            return path
    return os.path.join(_DATA_BASE, 'ai_config.json')


_RUNTIME_CONFIG_LOAD_ERROR = ''


def _load_runtime_config():
    global _RUNTIME_CONFIG_LOAD_ERROR
    _RUNTIME_CONFIG_LOAD_ERROR = ''
    path = _runtime_config_file()
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8-sig') as f:
                return json.load(f)
        except Exception as exc:
            _RUNTIME_CONFIG_LOAD_ERROR = f'{type(exc).__name__}: {exc}'
    return {}


def _runtime_config_debug():
    path = _runtime_config_file()
    cfg = _load_runtime_config()
    keys = {
        'jdy_app_key': bool(cfg.get('jdy_app_key')),
        'jdy_client_id': bool(cfg.get('jdy_client_id')),
        'jdy_db_id': bool(cfg.get('jdy_db_id')),
        'jdy_domain': bool(cfg.get('jdy_domain')),
        'jdy_app_secret': bool(cfg.get('jdy_app_secret')),
        'jdy_client_secret': bool(cfg.get('jdy_client_secret')),
        'jdy2_app_key': bool(cfg.get('jdy2_app_key')),
        'jdy2_client_id': bool(cfg.get('jdy2_client_id')),
        'jdy2_db_id': bool(cfg.get('jdy2_db_id')),
        'jdy2_domain': bool(cfg.get('jdy2_domain')),
        'jdy2_app_secret': bool(cfg.get('jdy2_app_secret')),
        'jdy2_client_secret': bool(cfg.get('jdy2_client_secret')),
    }
    return {
        'path': path,
        'exists': os.path.exists(path),
        'cwd': os.getcwd(),
        'has_keys': keys,
        'load_error': _RUNTIME_CONFIG_LOAD_ERROR,
    }


def _save_runtime_config(cfg):
    path = _runtime_config_file()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _load_jdy_config():
    """从 ai_config.json 读取账套1 JDY 配置"""
    cfg = _load_runtime_config()
    return {
        'name':          cfg.get('jdy_name', '饰品'),
        'client_id':     cfg.get('jdy_client_id', ''),
        'app_key':       cfg.get('jdy_app_key', ''),
        'app_secret':    cfg.get('jdy_app_secret', ''),
        'db_id':         cfg.get('jdy_db_id', ''),
        'outer_instance_id': cfg.get('jdy_outer_instance_id', ''),
        'domain':        cfg.get('jdy_domain', ''),
        'app_signature':  cfg.get('jdy_app_signature', ''),
        'client_secret':  cfg.get('jdy_client_secret', ''),
    }


def _load_jdy_config2():
    """从 ai_config.json 读取账套2 JDY 配置"""
    cfg = _load_runtime_config()
    return {
        'name':          cfg.get('jdy2_name', '箱包'),
        'client_id':     cfg.get('jdy2_client_id', ''),
        'app_key':       cfg.get('jdy2_app_key', ''),
        'app_secret':    cfg.get('jdy2_app_secret', ''),
        'db_id':         cfg.get('jdy2_db_id', ''),
        'outer_instance_id': cfg.get('jdy2_outer_instance_id', ''),
        'domain':        cfg.get('jdy2_domain', ''),
        'app_signature':  cfg.get('jdy2_app_signature', ''),
        'client_secret':  cfg.get('jdy2_client_secret', ''),
    }


def _save_jdy_config(updates):
    """将 JDY 配置合并写回 ai_config.json"""
    cfg = _load_runtime_config()
    secret_keys = {
        'jdy_app_secret', 'jdy_client_secret',
        'jdy2_app_secret', 'jdy2_client_secret',
    }
    for k, v in updates.items():
        if k in secret_keys and isinstance(v, str) and not v.strip():
            continue
        if v is not None:
            cfg[k] = v
    _save_runtime_config(cfg)


def _ensure_jdy_client():
    """返回已初始化的账套1 JDYClient，或根据配置新建"""
    cli = jdy_api.get_client()
    if cli:
        return cli
    cfg = _load_jdy_config()
    if not all([cfg['client_id'], cfg['app_key'], cfg['db_id']]):
        return None
    if not cfg['app_secret'] and not cfg['client_secret']:
        return None
    if not cfg['app_secret']:
        cfg['app_secret'] = cfg['client_secret']
    return jdy_api.init_client(
        cfg['client_id'], cfg['app_key'], cfg['app_secret'],
        cfg['db_id'], cfg.get('domain', ''), cfg.get('app_signature', ''),
        cfg.get('client_secret', '')
    )


def _ensure_jdy_client2():
    """返回已初始化的账套2 JDYClient，或根据配置新建（未配置返回 None）"""
    cli = jdy_api.get_client_2()
    if cli:
        return cli
    cfg = _load_jdy_config2()
    # app_secret 可以用 client_secret 代替（正式版无轮换密钥时）
    if not all([cfg['client_id'], cfg['app_key'], cfg['db_id']]):
        return None
    if not cfg['app_secret'] and not cfg['client_secret']:
        return None
    # 若 app_secret 为空，用 client_secret 填充
    if not cfg['app_secret']:
        cfg['app_secret'] = cfg['client_secret']
    return jdy_api.init_client_2(
        cfg['client_id'], cfg['app_key'], cfg['app_secret'],
        cfg['db_id'], cfg.get('domain', ''), cfg.get('app_signature', ''),
        cfg.get('client_secret', '')
    )


def _jdy_domain_candidates(primary=''):
    candidates = [
        primary,
        'https://vip2-hz.jdy.com',
        'https://vip1-gd.jdy.com',
        'https://vip2-gd.jdy.com',
        'https://vip1-hz.jdy.com',
    ]
    result = []
    for item in candidates:
        val = str(item or '').strip().rstrip('/')
        if val and val not in result:
            result.append(val)
    return result


def _probe_jdy_business_route(cli):
    """Token success is not enough; probe a tiny business API to validate IDC routing."""
    try:
        result = cli._request(
            'POST',
            '/jdyscm/purchaseOrder/list',
            body={'filter': {'page': 1, 'pageSize': 1}},
            query=cli._api_query(),
            timeout=30,
        )
        code = result.get('errcode') or result.get('code')
        if code in (3000002000, 5001, 5003):
            return False, result.get('msg') or result.get('description_cn') or result.get('description') or str(code)
        return True, ''
    except Exception as e:
        return False, _short_sync_error(e) if '_short_sync_error' in globals() else str(e)[:200]


def _init_jdy_client_with_working_domain(idx, client_id, app_key, app_secret, db_id, domain, client_secret):
    init_fn = jdy_api.init_client_2 if int(idx) == 2 else jdy_api.init_client
    last_error = ''
    for candidate in _jdy_domain_candidates(domain):
        cli = init_fn(client_id, app_key, app_secret, db_id, candidate, '', client_secret)
        test = cli.test_connection()
        if not test.get('ok'):
            last_error = test.get('msg') or 'token test failed'
            continue
        ok, probe_error = _probe_jdy_business_route(cli)
        if ok:
            return cli, candidate, test.get('msg', '')
        last_error = probe_error
    return None, '', last_error or 'no working JDY business route'


def _refresh_jdy_auth_for_idx(idx):
    cfg_all = _load_runtime_config()
    pfx = '' if int(idx) == 1 else '2'
    client_id = cfg_all.get(f'jdy{pfx}_client_id', '')
    client_secret = cfg_all.get(f'jdy{pfx}_client_secret', '')
    outer_id = cfg_all.get(f'jdy{pfx}_outer_instance_id', '')
    if not all([client_id, client_secret, outer_id]):
        return {'success': False, 'error': f'账套{idx}缺少 client_id / client_secret / outer_instance_id'}
    result = jdy_api.push_app_authorize(client_id, client_secret, outer_id)
    items = result.get('data') or []
    if not items:
        return {'success': False, 'error': f'账套{idx}无授权数据：请确认 outerInstanceId 是企业授权实例ID，不能用 dbId 代替'}
    item = items[0]
    app_key = item.get('appKey', '')
    app_secret = item.get('appSecret', '')
    domain = item.get('domain', '')
    account_id = str(item.get('accountId', ''))
    if not app_secret:
        return {'success': False, 'error': f'账套{idx}返回中无 appSecret'}
    updates = {
        f'jdy{pfx}_app_key': app_key,
        f'jdy{pfx}_app_secret': app_secret,
    }
    if domain:
        updates[f'jdy{pfx}_domain'] = domain
    if account_id:
        updates[f'jdy{pfx}_db_id'] = account_id
    cfg_all.update(updates)
    _save_runtime_config(cfg_all)
    cli, working_domain, msg = _init_jdy_client_with_working_domain(
        idx,
        client_id, app_key, app_secret,
        account_id or cfg_all.get(f'jdy{pfx}_db_id', ''),
        domain or cfg_all.get(f'jdy{pfx}_domain', ''),
        client_secret,
    )
    if not cli:
        return {'success': False, 'error': f'账套{idx}业务接口路由不可用：{msg}'}
    if working_domain and working_domain != cfg_all.get(f'jdy{pfx}_domain', ''):
        cfg_all[f'jdy{pfx}_domain'] = working_domain
        _save_runtime_config(cfg_all)
    return {'success': True, 'client': cli, 'message': msg, 'domain': working_domain}


def _refresh_jdy_auth_for_account_name(account_name):
    cfg1 = _load_jdy_config()
    cfg2 = _load_jdy_config2()
    if account_name == cfg2.get('name'):
        return _refresh_jdy_auth_for_idx(2)
    return _refresh_jdy_auth_for_idx(1)


def _client_for_sync_account(account_name, cli_fn):
    """同步任务默认使用现有授权，只有本地凭证不完整时才尝试刷新授权。"""
    cli = cli_fn()
    if cli:
        return cli
    auth = _refresh_jdy_auth_for_account_name(account_name)
    if auth.get('success'):
        return auth.get('client') or cli_fn()
    raise RuntimeError(auth.get('error') or 'JDY API not configured')


def _mask(s):
    s = s or ''
    return ('*' * max(len(s) - 4, 0) + s[-4:]) if len(s) >= 4 else '*' * len(s)


def _do_license_fill_one(code, account='', preferred_cli_fn=None):
    """
    查询单个商品的 proLicense，解析后写入 Excel。
    preferred_cli_fn：若提供，严格只查该账套，不跨账套兜底。
    返回 dict（含 status）或 None（跳过）。
    """
    from ai_identify import parse_prolicense, fill_from_license, _load_config, _find_excel
    from jdy_api import parse_dimensions

    cfg2 = _load_jdy_config2()
    if preferred_cli_fn:
        # 调拨单账套已知 → 严格限制，不跨账套
        cli = preferred_cli_fn()
    else:
        cli = _ensure_jdy_client2() if account == cfg2.get('name') else _ensure_jdy_client()
        if not cli:
            cli = _ensure_jdy_client() or _ensure_jdy_client2()
    if not cli:
        return {'code': code, 'status': 'no_client'}

    product = cli.get_product_by_code(code)
    if not product and not preferred_cli_fn:
        # 未指定账套时才允许跨账套兜底
        other = _ensure_jdy_client2() if cli is _ensure_jdy_client() else _ensure_jdy_client()
        if other:
            product = other.get_product_by_code(code)
            if product:
                cli = other
    if not product:
        return {'code': code, 'status': 'not_found'}

    pro_license = product.get('proLicense') or ''
    name_part, material_part = parse_prolicense(pro_license)
    if not name_part or not material_part:
        return {'code': code, 'status': 'no_license', 'proLicense': pro_license,
                'cn_name': (product.get('name') or '').strip()}

    cfg1   = _load_jdy_config()
    actual = cfg2.get('name') if cli is _ensure_jdy_client2() else cfg1.get('name')
    remark = product.get('remark') or ''
    try:
        conversion = float(remark.strip()) if remark.strip() else None
    except ValueError:
        conversion = None

    ai_cfg     = _load_ai_config()
    excel_path = _find_excel(ai_cfg)
    if not excel_path:
        return {'code': code, 'status': 'no_excel'}

    r = fill_from_license(
        excel_path, code, name_part, material_part,
        extra_fields={
            'category':   product.get('categoryName') or '',
            'dimensions': parse_dimensions(product.get('registrationNo') or ''),
            'conversion': conversion,
            'account':    actual,
        },
    )
    return {'status': 'ok', **r}


def _auto_license_fill(items, preferred_cli_fn=None, code_account_map=None):
    """
    生成前检查哪些商品不在基础资料里，自动用 proLicense 补全。
    code_account_map：{code: cli_fn}，优先级高于 preferred_cli_fn。
    返回 (filled_codes, no_license_items)
      filled_codes:    成功补全的 code 列表
      no_license_items: proLicense 为空、需要手工填写的 [{code, cn_name}]
    """
    try:
        from excel_gen import load_base_data
        base   = load_base_data()
        codes  = [it['code'] for it in items if it.get('code')]
        missing = [c for c in codes if c not in base]
        filled, no_license_items = [], []
        for code in missing:
            code_cli_fn = (code_account_map or {}).get(code) or preferred_cli_fn
            r = _do_license_fill_one(code, preferred_cli_fn=code_cli_fn)
            if r and r.get('status') == 'ok':
                filled.append(code)
            elif r and r.get('status') == 'no_license':
                no_license_items.append({'code': code, 'cn_name': r.get('cn_name', '')})
        return filled, no_license_items
    except Exception as e:
        print(f'[WARN] _auto_license_fill: {e}')
        return [], []


# 模块级缓存：服务器运行期间，同一商品编码不重复查询
# code -> {'supplierName', 'supplierNumber', 'origin_city'}  或  None（表示已查过但无结果）
_supplier_lookup_cache = {}
_city_lookup_cache     = {}  # supplierNumber -> origin_city


def _lookup_suppliers(items, preferred_cli_fn=None, supplier_overrides=None, code_account_map=None):
    """
    查询供应商信息。缓存（_api_cache/code_cache.json）命中直接返回，
    未命中的码用 5 线程并发调取 JDY API，结果写入磁盘缓存（24h 有效）。
    返回 (result_dict, needs_select_dict)
    """
    _SKIP_KW = ('国内展厅', '国内采购', '盘存')
    now_str   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    supplier_overrides = supplier_overrides or {}
    cache = _load_code_cache()

    # 去重
    seen, codes = set(), []
    for it in items:
        c = (it.get('code') or '').strip()
        if c and c not in seen:
            seen.add(c); codes.append(c)

    # 哪些码需要调 API（未缓存 或 已过期）
    def _po_cached(code):
        e = cache.get(code, {})
        return _cache_valid(e) and 'po_result' in e

    to_fetch = [c for c in codes if c not in supplier_overrides and not _po_cached(c)]

    # ── 并发抓取 ──────────────────────────────────────────────────────────
    def _fetch_one(code):
        import time as _t
        _code_cli = (code_account_map or {}).get(code) or preferred_cli_fn
        _cli_fns  = [_code_cli] if _code_cli else [_ensure_jdy_client, _ensure_jdy_client2]

        po_result   = None   # None=未找到, dict=找到, {'needs_select':[...]}=多供应商
        found_cli   = None
        found_snum  = None

        for cli_fn in _cli_fns:
            po = None
            for attempt in range(3):
                try:
                    cli = cli_fn()
                    if not cli:
                        break
                    po = cli.get_purchase_orders_by_product(code, page_size=10)
                    break
                except Exception as e:
                    err = str(e)
                    if any(k in err for k in ('Connection reset', 'Errno 54', 'Errno 104',
                                              'RemoteDisconnected', 'Connection aborted')) \
                            and attempt < 2:
                        _t.sleep(0.5)
                        continue
                    po = None
                    break

            if po is None:
                continue

            orders = [o for o in ((po or {}).get('list') or [])
                      if not any(kw in (o.get('supplierName') or '') for kw in _SKIP_KW)]
            if not orders:
                continue

            # 收集所有唯一编号（保序，名称去重前），供后续城市遍历用
            all_snums_ordered = []
            _seen_snum = set()
            for o in orders:
                snum = (o.get('supplierNumber') or '').strip()
                if snum and snum not in _seen_snum:
                    _seen_snum.add(snum)
                    all_snums_ordered.append(snum)

            unique = {}
            for o in orders:
                snum = (o.get('supplierNumber') or '').strip()
                sname = (o.get('supplierName') or '').strip()
                if snum and snum not in unique:
                    unique[snum] = sname

            # 按名称再去重：同名供应商只保留第一条（不同 supplierNumber 也算同一家）
            seen_names = set()
            deduped = {}
            for snum, sname in unique.items():
                if sname not in seen_names:
                    seen_names.add(sname)
                    deduped[snum] = sname
            unique = deduped

            if len(unique) > 1:
                # 多供应商时先过滤掉禁用的（status=0查不到 → 禁用）
                enabled_unique = {}
                for _k, _v in unique.items():
                    try:
                        if cli.get_supplier_by_number(_k, status=0):
                            enabled_unique[_k] = _v
                        else:
                            print(f'[SUPPLIER] {_k}({_v}) 已禁用，从候选列表移除')
                    except Exception:
                        enabled_unique[_k] = _v   # 查询异常时保留，不误删
                unique = enabled_unique

            if len(unique) > 1:
                po_result = {'needs_select': [{'supplierNumber': k, 'supplierName': v}
                                              for k, v in unique.items()]}
            elif len(unique) == 1:
                # 过滤后只剩一个，直接用，不用弹窗
                found_snum = list(unique.keys())[0]
                # 从 orders 里找这个 snum 对应的最新一条
                _matched = next((o for o in orders
                                 if (o.get('supplierNumber') or '').strip() == found_snum), orders[0])
                po_result  = {'supplierName':   (_matched.get('supplierName') or '').strip(),
                              'supplierNumber': found_snum}
                found_cli  = cli
            else:
                first = orders[0]
                po_result  = {'supplierName':   (first.get('supplierName') or '').strip(),
                              'supplierNumber': (first.get('supplierNumber') or '').strip()}
                found_cli  = cli
                found_snum = po_result['supplierNumber']

                # 若首选供应商已禁用（status=0查不到），从同名的其他编号里找启用的替换
                try:
                    _sup_check = found_cli.get_supplier_by_number(found_snum, status=0)
                    if _sup_check is None:   # status=0查不到 → 禁用
                        print(f'[SUPPLIER] {found_snum} 已禁用，尝试其他编号')
                        for _snum2 in all_snums_ordered:
                            if _snum2 == found_snum:
                                continue
                            _sup2 = found_cli.get_supplier_by_number(_snum2, status=0)
                            if _sup2:   # status=0能查到 → 启用
                                found_snum = _snum2
                                po_result['supplierNumber'] = _snum2
                                po_result['supplierName']   = (_sup2.get('name') or po_result['supplierName']).strip()
                                print(f'[SUPPLIER] 改用启用编号 {_snum2}')
                                break
                except Exception:
                    pass
            break

        # 查城市：遍历所有供应商编号，status=0只查启用的，找到有城市信息的为止
        origin_city = ''
        if found_cli and all_snums_ordered:
            for _snum_try in all_snums_ordered:
                try:
                    sup = found_cli.get_supplier_by_number(_snum_try, status=0)
                    if not sup:   # 禁用或不存在，跳过
                        continue
                    for c in (sup.get('contacts') or []):
                        city   = (c.get('city') or '').strip()
                        county = (c.get('county') or '').strip()
                        # 优先用县级市（更精确），fallback 到地级市
                        if county: origin_city = county; break
                        if city:   origin_city = city;   break
                    if origin_city:
                        break
                except Exception:
                    pass

        # 注意：不再 fallback '义乌市'，空串时让 process_items 用基础资料 H列的 origin
        return code, {
            'fetched_at':   now_str,
            'po_result':    po_result,
            'origin_city':  origin_city,
        }

    if to_fetch:
        print(f'[CACHE] {len(to_fetch)} 个码调取 API，'
              f'{len(codes) - len(to_fetch)} 个命中缓存，5线程并发')
        with ThreadPoolExecutor(max_workers=5) as pool:
            futs = {pool.submit(_fetch_one, c): c for c in to_fetch}
            for fut in as_completed(futs):
                try:
                    code, entry = fut.result()
                    cache[code] = entry
                except Exception as e:
                    print(f'[CACHE] fetch error: {e}')
        _save_code_cache(cache)
    else:
        print(f'[CACHE] 全部 {len(codes)} 个码命中缓存')

    # ── 构建返回值 ─────────────────────────────────────────────────────────
    supplier_order, counter = {}, [1]
    def _get_idx(name):
        if name not in supplier_order:
            supplier_order[name] = counter[0]; counter[0] += 1
        return supplier_order[name]

    result, needs_select = {}, {}
    for code in codes:
        # 用户已选定
        if code in supplier_overrides:
            ov = supplier_overrides[code]
            sname  = (ov.get('supplierName') or '未知供应商').strip()
            snumber = (ov.get('supplierNumber') or '').strip()
            # 城市：优先缓存
            city_entry = cache.get(code, {})
            origin_city = city_entry.get('origin_city', '义乌市') if _po_cached(code) else '义乌市'
            result[code] = {'supplierName': sname, 'supplierNumber': snumber,
                            'origin_city': origin_city, 'supplierIdx': _get_idx(sname)}
            continue

        entry = cache.get(code, {})
        po    = entry.get('po_result')

        if po is None:
            needs_select[code] = []
            print(f'[SUPPLIER] code={code} 未找到采购单')
            continue

        if 'needs_select' in po:
            needs_select[code] = po['needs_select']
            print(f'[SUPPLIER] code={code} 多供应商，需选择')
            continue

        sname = po.get('supplierName') or '未知供应商'
        result[code] = {
            'supplierName':   sname,
            'supplierNumber': po.get('supplierNumber', ''),
            'origin_city':    entry.get('origin_city', '义乌市'),
            'supplierIdx':    _get_idx(sname),
        }

    return result, needs_select


def _load_ai_config():
    return _load_runtime_config()


def _extract_supplier_address(supplier):
    """
    从供应商首要联系人中提取地址信息，返回 (district, address_detail)。
      district      → H 列（contacts[].county，即区）
      address_detail → R 列（contacts[].address，即详细地址）

    实测 JDY supplier/list 联系人字段：
      province / city / county（区）/ address（详细地址）/ isPrimary
    """
    contacts = supplier.get('contacts') or []
    if not contacts:
        return '', ''

    # 找首要联系人（isPrimary=true），否则取第一个
    primary = next((c for c in contacts if c.get('isPrimary')), contacts[0])

    district = (primary.get('county') or '').strip()
    detail   = (primary.get('address') or '').strip()
    return district, detail


# ── 生成历史记录（generation_history.json 与 ai_config.json 同目录）────────
_GEN_HISTORY_FILE = None

def _gh_file():
    global _GEN_HISTORY_FILE
    if _GEN_HISTORY_FILE is None:
        from ai_identify import _CONFIG_FILE
        _GEN_HISTORY_FILE = os.path.join(
            os.path.dirname(os.path.abspath(_CONFIG_FILE)), 'generation_history.json')
    return _GEN_HISTORY_FILE

def _load_gen_history():
    try:
        p = _gh_file()
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as fp:
                return json.load(fp)
    except Exception:
        pass
    return {}

def _record_gen_history(order_refs):
    """order_refs: [{id, account}, ...]  —  生成成功后调用"""
    if not order_refs:
        return
    import datetime
    with _history_lock:
        h = _load_gen_history()
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for ref in order_refs:
            oid = str(ref.get('id', '')).strip()
            acc = str(ref.get('account', '')).strip()
            if not oid:
                continue
            key = f'{acc}::{oid}'
            if key not in h:
                h[key] = {'id': oid, 'account': acc, 'count': 0, 'last_at': ''}
            h[key]['count'] += 1
            h[key]['last_at'] = now
        with open(_gh_file(), 'w', encoding='utf-8') as fp:
            json.dump(h, fp, ensure_ascii=False, indent=2)
    print(f'[GEN] 已记录 {len(order_refs)} 张单的生成历史')


@app.route('/jdy-config', methods=['GET'])
def jdy_config_get():
    """返回两套 JDY 配置（敏感字段打码）"""
    try:
        cfg  = _load_jdy_config()
        cfg2 = _load_jdy_config2()
        return jsonify({
            'success': True,
            # 账套 1
            'name':       cfg.get('name', '饰品'),
            'client_id':  cfg.get('client_id', ''),
            'app_key':    cfg.get('app_key', ''),
            'db_id':      cfg.get('db_id', ''),
            'outer_instance_id': cfg.get('outer_instance_id', ''),
            'domain':     cfg.get('domain', ''),
            'app_secret_masked':    _mask(cfg.get('app_secret', '')),
            'client_secret_masked': _mask(cfg.get('client_secret', '')),
            'has_config': bool(cfg.get('client_id') and cfg.get('app_key')
                               and cfg.get('app_secret') and cfg.get('db_id')),
            # 账套 2
            'name2':       cfg2.get('name', '箱包'),
            'client_id2':  cfg2.get('client_id', ''),
            'app_key2':    cfg2.get('app_key', ''),
            'db_id2':      cfg2.get('db_id', ''),
            'outer_instance_id2': cfg2.get('outer_instance_id', ''),
            'domain2':     cfg2.get('domain', ''),
            'app_secret_masked2':    _mask(cfg2.get('app_secret', '')),
            'client_secret_masked2': _mask(cfg2.get('client_secret', '')),
            'has_config2': bool(cfg2.get('client_id') and cfg2.get('app_key')
                                and cfg2.get('app_secret') and cfg2.get('db_id')),
            'config_debug': _runtime_config_debug(),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-config', methods=['POST'])
def jdy_config_set():
    """保存 JDY 配置（idx=1 或 idx=2 区分账套）"""
    try:
        data = request.get_json(force=True) or {}
        idx  = int(data.get('idx', 1))   # 1=饰品, 2=箱包
        pfx  = '' if idx == 1 else '2'   # json key 前缀
        critical_keys = ('client_id', 'app_key', 'app_secret', 'client_secret', 'db_id', 'outer_instance_id', 'domain')
        if not any(str(data.get(k) or '').strip() for k in critical_keys):
            return jsonify({'success': False, 'error': '配置未加载，禁止保存空配置'}), 400
        updates = {}
        name = str(data.get('name') or '').strip()
        client_id = str(data.get('client_id') or '').strip()
        app_key = str(data.get('app_key') or '').strip()
        db_id = str(data.get('db_id') or '').strip()
        outer_instance_id = str(data.get('outer_instance_id') or '').strip()
        domain = str(data.get('domain') or '').strip()
        if name:
            updates[f'jdy{pfx}_name']          = name
        if client_id:
            updates[f'jdy{pfx}_client_id']     = client_id
        if app_key:
            updates[f'jdy{pfx}_app_key']       = app_key
        app_secret = str(data.get('app_secret') or '').strip()
        client_secret = str(data.get('client_secret') or '').strip()
        if app_secret:
            updates[f'jdy{pfx}_app_secret']    = app_secret
        if client_secret:
            updates[f'jdy{pfx}_client_secret'] = client_secret
        if db_id:
            updates[f'jdy{pfx}_db_id']         = db_id
        if outer_instance_id:
            updates[f'jdy{pfx}_outer_instance_id'] = outer_instance_id
        if domain:
            updates[f'jdy{pfx}_domain']        = domain
        _save_jdy_config(updates)
        # 重新初始化对应客户端
        if idx == 1:
            cfg = _load_jdy_config()
            if cfg['client_id'] and cfg['app_key'] and cfg['app_secret'] and cfg['db_id']:
                jdy_api.init_client(
                    cfg['client_id'], cfg['app_key'], cfg['app_secret'],
                    cfg['db_id'], cfg.get('domain', ''), cfg.get('app_signature', ''),
                    cfg.get('client_secret', '')
                )
        else:
            cfg2 = _load_jdy_config2()
            if cfg2['client_id'] and cfg2['app_key'] and cfg2['app_secret'] and cfg2['db_id']:
                jdy_api.init_client_2(
                    cfg2['client_id'], cfg2['app_key'], cfg2['app_secret'],
                    cfg2['db_id'], cfg2.get('domain', ''), cfg2.get('app_signature', ''),
                    cfg2.get('client_secret', '')
                )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook', methods=['GET', 'POST'])
def jdy_webhook():
    """
    接收精斗云实时授权推送（appKey / appSecret / domain / accountId 自动更新）
    配置地址：open.jdy.com → 应用详情 → 沙箱调试/企业授权 → 消息订阅
    请求格式参见官方文档：bizType=app_authorize
    """
    try:
        if request.method == 'GET':
            return jsonify({'success': True, 'message': 'JDY webhook endpoint is ready'})

        biz_type = (request.args.get('bizType') or '').strip()
        data = request.get_json(force=True) or {}
        if not biz_type:
            biz_type = data.get('bizType', '')

        print(f'[JDY Webhook] bizType={biz_type}, body={json.dumps(data, ensure_ascii=False)[:300]}')

        # app_authorize updates credentials; business messages are queued for slow local sync.
        if biz_type != 'app_authorize':
            queued = _enqueue_jdy_webhook_events(biz_type, data)
            return jsonify({'errcode': '0', 'description': 'ok',
                            'data': {'status': '0', 'msg': f'queued {queued}', 'type': biz_type}})

        items = data.get('data') or []
        if not items:
            return jsonify({'errcode': '0', 'description': 'ok',
                            'data': {'status': '0', 'msg': 'empty data', 'type': biz_type}})

        cfg1 = _load_jdy_config()
        cfg2 = _load_jdy_config2()

        for item in items:
            incoming_key = str(item.get('appKey', ''))
            account_id   = str(item.get('accountId', ''))

            # 按 appKey 精确匹配账套（最可靠），其次按 accountId
            if   incoming_key and incoming_key == cfg2.get('app_key', ''):
                pfx = '2'
            elif incoming_key and incoming_key == cfg1.get('app_key', ''):
                pfx = ''
            elif account_id and account_id == str(cfg2.get('db_id', '')):
                pfx = '2'
            elif account_id and account_id == str(cfg1.get('db_id', '')):
                pfx = ''
            else:
                # 无法匹配任何已知账套 → 跳过，避免覆盖错误账套
                print(f'[JDY Webhook] 跳过无法匹配的推送: appKey={incoming_key} accountId={account_id}')
                continue

            # 只更新会轮换的凭证（appKey / appSecret），不动 domain 和 dbId
            updates = {}
            if item.get('appKey'):    updates[f'jdy{pfx}_app_key']    = item['appKey']
            if item.get('appSecret'): updates[f'jdy{pfx}_app_secret'] = item['appSecret']

            if updates:
                _save_jdy_config(updates)
                print(f'[JDY Webhook] 账套{pfx or "1"}密钥更新: appKey={incoming_key[:8]}… accountId={account_id}')

        # 重新初始化两个客户端
        jdy_api._client_instance   = None
        jdy_api._client_instance_2 = None
        _ensure_jdy_client()
        _ensure_jdy_client2()

        return jsonify({'errcode': '0', 'description': 'success',
                        'data': {'status': '0', 'msg': 'success', 'type': biz_type}})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] jdy_webhook: {tb}')
        return jsonify({'errcode': '1', 'description': str(e),
                        'data': {'status': '1', 'msg': str(e), 'type': ''}}), 200


def _write_update_apply_script(script_path):
    ps = r'''
param(
  [string]$ZipPath,
  [string]$AppDir
)
$ErrorActionPreference = "Stop"
$port = 5008
$timestamp = Get-Date -Format "yyyyMMdd_HHmmss"
$updateDir = Join-Path $AppDir "_updates"
$log = Join-Path $updateDir "apply_update.log"
New-Item -ItemType Directory -Force -Path $updateDir | Out-Null
function Log([string]$msg) {
  ("[{0}] {1}" -f (Get-Date -Format "yyyy-MM-dd HH:mm:ss"), $msg) | Out-File -FilePath $log -Append -Encoding UTF8
}
function Stop-Watchdog() {
  try { schtasks /End /TN "QiHangLocalProjectWatchdog" /F | Out-Null } catch {}
  try { schtasks /Delete /TN "QiHangLocalProjectWatchdog" /F | Out-Null } catch {}
  try {
    $watchers = Get-CimInstance Win32_Process | Where-Object {
      ($_.Name -match "cmd|powershell") -and ($_.CommandLine -like "*--watch*")
    }
    foreach ($p in $watchers) {
      Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue
      Log "Stopped watchdog PID $($p.ProcessId)"
    }
  } catch {
    Log "Stop watchdog scan failed: $($_.Exception.Message)"
  }
}
function Stop-Port([int]$p) {
  $conns = Get-NetTCPConnection -LocalPort $p -ErrorAction SilentlyContinue | Where-Object { $_.State -eq "Listen" }
  foreach ($c in $conns) {
    try {
      Stop-Process -Id $c.OwningProcess -Force -ErrorAction SilentlyContinue
      Log "Stopped PID $($c.OwningProcess) on port $p"
    } catch {
      Log "Failed to stop PID $($c.OwningProcess): $($_.Exception.Message)"
    }
  }
}
try {
  "`n--- update $timestamp ---" | Out-File -FilePath $log -Append -Encoding UTF8
  Log "Update started. zip=$ZipPath appDir=$AppDir"
  Start-Sleep -Seconds 2
  Stop-Watchdog
  Stop-Port $port
  Start-Sleep -Seconds 2

  $tmp = Join-Path $updateDir ("extract_" + $timestamp)
  if (Test-Path $tmp) { Remove-Item -LiteralPath $tmp -Recurse -Force }
  New-Item -ItemType Directory -Force -Path $tmp | Out-Null
  Expand-Archive -LiteralPath $ZipPath -DestinationPath $tmp -Force
  Log "Archive extracted to $tmp"

  $serverExeInZip = Get-ChildItem -LiteralPath $tmp -Filter "server.exe" -File -Recurse |
    Sort-Object { $_.FullName.Length } |
    Select-Object -First 1
  if ($serverExeInZip) {
    $src = $serverExeInZip.DirectoryName
  } else {
    throw "server.exe not found in update archive"
  }
  Log "Using update source: $src"

  $backupRoot = Join-Path $updateDir "backups"
  $backup = Join-Path $backupRoot $timestamp
  New-Item -ItemType Directory -Force -Path $backup | Out-Null
  $skip = @("ai_config.json", "server_secret.key", "_sales_cache", "_api_cache", "_items_store", "_updates", "logs", "generation_history.json")

  foreach ($item in Get-ChildItem -LiteralPath $AppDir -Force) {
    if ($skip -contains $item.Name) { continue }
    try {
      Copy-Item -LiteralPath $item.FullName -Destination (Join-Path $backup $item.Name) -Recurse -Force
    } catch {
      Log "Backup skip $($item.Name): $($_.Exception.Message)"
    }
  }
  Log "Backup created: $backup"

  foreach ($item in Get-ChildItem -LiteralPath $src -Force) {
    if ($skip -contains $item.Name) {
      Log "Skip runtime item from update: $($item.Name)"
      continue
    }
    $dst = Join-Path $AppDir $item.Name
    $copied = $false
    for ($i = 1; $i -le 5; $i++) {
      try {
        Copy-Item -LiteralPath $item.FullName -Destination $dst -Recurse -Force
        $copied = $true
        break
      } catch {
        Log "Copy retry $i for $($item.Name): $($_.Exception.Message)"
        Stop-Watchdog
        Stop-Port $port
        Start-Sleep -Seconds 2
      }
    }
    if (-not $copied) { throw "Failed to update $($item.Name)" }
    Log "Updated $($item.Name)"
  }

  $watchBat = $null
  foreach ($bat in Get-ChildItem -LiteralPath $AppDir -Filter "*.bat" -File -ErrorAction SilentlyContinue) {
    if (Select-String -LiteralPath $bat.FullName -Pattern "--watch","QiHangLocalProjectWatchdog" -Quiet -ErrorAction SilentlyContinue) {
      $watchBat = $bat.FullName
      break
    }
  }
  $exe = Join-Path $AppDir "server.exe"
  if (Test-Path $watchBat) {
    Start-Process -FilePath $watchBat -WorkingDirectory $AppDir
    Log "Started watchdog bat"
  } elseif (Test-Path $exe) {
    Start-Process -FilePath $exe -WorkingDirectory $AppDir
    Log "Started server.exe"
  } else {
    throw "server.exe not found after update"
  }
  Log "Update completed"
} catch {
  Log ("Update failed: " + $_.Exception.Message)
  $exe = Join-Path $AppDir "server.exe"
  if (Test-Path $exe) { Start-Process -FilePath $exe -WorkingDirectory $AppDir }
  exit 1
}
'''
    os.makedirs(os.path.dirname(script_path), exist_ok=True)
    with open(script_path, 'w', encoding='utf-8') as f:
        f.write(ps.strip() + '\n')


def _list_update_backups():
    backup_root = os.path.join(_DATA_BASE, '_updates', 'backups')
    rows = []
    if not os.path.isdir(backup_root):
        return rows
    for name in os.listdir(backup_root):
        path = os.path.join(backup_root, name)
        if not os.path.isdir(path):
            continue
        try:
            stat = os.stat(path)
            rows.append({
                'name': name,
                'path': path,
                'mtime': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            })
        except Exception:
            rows.append({'name': name, 'path': path, 'mtime': ''})
    rows.sort(key=lambda x: x.get('name', ''), reverse=True)
    return rows


def _write_update_rollback_script(script_path):
    ps = r'''
param(
  [string]$BackupDir,
  [string]$AppDir
)
$ErrorActionPreference = "Stop"
$port = 5008
$timestamp = Get-Date -Format "yyyyMMdd_HHmmss"
$updateDir = Join-Path $AppDir "_updates"
$log = Join-Path $updateDir "apply_update.log"
New-Item -ItemType Directory -Force -Path $updateDir | Out-Null
function Log([string]$msg) {
  ("[{0}] {1}" -f (Get-Date -Format "yyyy-MM-dd HH:mm:ss"), $msg) | Out-File -FilePath $log -Append -Encoding UTF8
}
function Stop-Watchdog() {
  try { schtasks /End /TN "QiHangLocalProjectWatchdog" /F | Out-Null } catch {}
  try { schtasks /Delete /TN "QiHangLocalProjectWatchdog" /F | Out-Null } catch {}
  try {
    $watchers = Get-CimInstance Win32_Process | Where-Object {
      ($_.Name -match "cmd|powershell") -and ($_.CommandLine -like "*--watch*")
    }
    foreach ($p in $watchers) {
      Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue
      Log "Stopped watchdog PID $($p.ProcessId)"
    }
  } catch {
    Log "Stop watchdog scan failed: $($_.Exception.Message)"
  }
}
function Stop-Port([int]$p) {
  $conns = Get-NetTCPConnection -LocalPort $p -ErrorAction SilentlyContinue | Where-Object { $_.State -eq "Listen" }
  foreach ($c in $conns) {
    try {
      Stop-Process -Id $c.OwningProcess -Force -ErrorAction SilentlyContinue
      Log "Stopped PID $($c.OwningProcess) on port $p"
    } catch {
      Log "Failed to stop PID $($c.OwningProcess): $($_.Exception.Message)"
    }
  }
}
try {
  "`n--- rollback $timestamp ---" | Out-File -FilePath $log -Append -Encoding UTF8
  Log "Rollback started. backup=$BackupDir appDir=$AppDir"
  if (-not (Test-Path $BackupDir)) { throw "backup not found: $BackupDir" }
  Start-Sleep -Seconds 2
  Stop-Watchdog
  Stop-Port $port
  Start-Sleep -Seconds 2

  foreach ($item in Get-ChildItem -LiteralPath $BackupDir -Force) {
    $dst = Join-Path $AppDir $item.Name
    $copied = $false
    for ($i = 1; $i -le 5; $i++) {
      try {
        Copy-Item -LiteralPath $item.FullName -Destination $dst -Recurse -Force
        $copied = $true
        break
      } catch {
        Log "Rollback copy retry $i for $($item.Name): $($_.Exception.Message)"
        Stop-Watchdog
        Stop-Port $port
        Start-Sleep -Seconds 2
      }
    }
    if (-not $copied) { throw "Failed to restore $($item.Name)" }
    Log "Restored $($item.Name)"
  }

  $watchBat = $null
  foreach ($bat in Get-ChildItem -LiteralPath $AppDir -Filter "*.bat" -File -ErrorAction SilentlyContinue) {
    if (Select-String -LiteralPath $bat.FullName -Pattern "--watch","QiHangLocalProjectWatchdog" -Quiet -ErrorAction SilentlyContinue) {
      $watchBat = $bat.FullName
      break
    }
  }
  $exe = Join-Path $AppDir "server.exe"
  if (Test-Path $watchBat) {
    Start-Process -FilePath $watchBat -WorkingDirectory $AppDir
    Log "Started watchdog bat"
  } elseif (Test-Path $exe) {
    Start-Process -FilePath $exe -WorkingDirectory $AppDir
    Log "Started server.exe"
  } else {
    throw "server.exe not found after rollback"
  }
  Log "Rollback completed"
} catch {
  Log ("Rollback failed: " + $_.Exception.Message)
  $exe = Join-Path $AppDir "server.exe"
  if (Test-Path $exe) { Start-Process -FilePath $exe -WorkingDirectory $AppDir }
  exit 1
}
'''
    os.makedirs(os.path.dirname(script_path), exist_ok=True)
    with open(script_path, 'w', encoding='utf-8') as f:
        f.write(ps.strip() + '\n')


@app.route('/system-update/upload', methods=['POST'])
def system_update_upload():
    """
    上传新版 zip 到服务器并触发离线覆盖更新。
    zip 可包含顶层目录，内部有“服务端”目录时优先使用该目录。
    """
    try:
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'success': False, 'error': '请选择 zip 更新包'}), 400
        if not f.filename.lower().endswith('.zip'):
            return jsonify({'success': False, 'error': '只支持 zip 更新包'}), 400

        update_dir = os.path.join(_DATA_BASE, '_updates')
        os.makedirs(update_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_path = os.path.join(update_dir, f'pending_update_{ts}.zip')
        f.save(zip_path)

        script_path = os.path.join(update_dir, 'apply_update.ps1')
        _write_update_apply_script(script_path)

        powershell = os.path.join(os.environ.get('SystemRoot', r'C:\Windows'),
                                  r'System32\WindowsPowerShell\v1.0\powershell.exe')
        if not os.path.exists(powershell):
            powershell = 'powershell.exe'

        subprocess.Popen(
            [powershell, '-NoProfile', '-ExecutionPolicy', 'Bypass',
             '-File', script_path, zip_path, _EXE_DIR],
            cwd=_EXE_DIR,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0),
        )
        return jsonify({
            'success': True,
            'message': '更新包已上传，服务器将在几秒后自动重启并应用更新',
            'zip_path': zip_path,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] system_update_upload: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/system-update/backups', methods=['GET'])
def system_update_backups():
    try:
        return jsonify({'success': True, 'backups': _list_update_backups()})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] system_update_backups: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/system-update/rollback', methods=['POST'])
def system_update_rollback():
    try:
        data = request.get_json(force=True) or {}
        name = str(data.get('backup') or '').strip()
        if not name or any(ch in name for ch in ('/', '\\', ':', '..')):
            return jsonify({'success': False, 'error': '请选择有效备份'}), 400
        backup_root = os.path.join(_DATA_BASE, '_updates', 'backups')
        backup_dir = os.path.abspath(os.path.join(backup_root, name))
        if not backup_dir.startswith(os.path.abspath(backup_root)) or not os.path.isdir(backup_dir):
            return jsonify({'success': False, 'error': '备份不存在'}), 404

        update_dir = os.path.join(_DATA_BASE, '_updates')
        os.makedirs(update_dir, exist_ok=True)
        script_path = os.path.join(update_dir, 'rollback_update.ps1')
        _write_update_rollback_script(script_path)
        powershell = os.path.join(os.environ.get('SystemRoot', r'C:\Windows'),
                                  r'System32\WindowsPowerShell\v1.0\powershell.exe')
        if not os.path.exists(powershell):
            powershell = 'powershell.exe'
        subprocess.Popen(
            [powershell, '-NoProfile', '-ExecutionPolicy', 'Bypass',
             '-File', script_path, backup_dir, _EXE_DIR],
            cwd=_EXE_DIR,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0),
        )
        return jsonify({'success': True, 'message': '回滚已开始，服务器将在几秒后重启', 'backup': name})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] system_update_rollback: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/system-logs', methods=['GET'])
def system_logs():
    try:
        items, paths = _system_log_files()
        key = request.args.get('file') or 'server'
        if key not in paths:
            key = 'server'
        try:
            max_kb = int(request.args.get('max_kb') or 128)
        except Exception:
            max_kb = 128
        content = _tail_text_file(paths.get(key), max_bytes=max_kb * 1024)
        return jsonify({
            'success': True,
            'files': items,
            'current': key,
            'content': content,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] system_logs: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-refresh-auth', methods=['POST'])
def jdy_refresh_auth():
    """
    主动调 push_app_authorize 获取正式版 appKey/appSecret/domain
    body: {idx: 1 或 2}
    成功后自动写入 ai_config.json 并重新初始化 client
    """
    try:
        data = request.get_json(force=True) or {}
        idx  = int(data.get('idx', 1))
        pfx  = '' if idx == 1 else '2'

        cfg_all = _load_runtime_config()
        client_id     = cfg_all.get(f'jdy{pfx}_client_id', '')
        client_secret = cfg_all.get(f'jdy{pfx}_client_secret', '')
        outer_id      = cfg_all.get(f'jdy{pfx}_outer_instance_id', '')

        if not all([client_id, client_secret, outer_id]):
            return jsonify({'success': False,
                            'error': f'账套{idx}缺少 client_id / client_secret / outer_instance_id'})

        auth = _refresh_jdy_auth_for_idx(idx)
        if not auth.get('success'):
            return jsonify({'success': False, 'error': auth.get('error') or '刷新授权失败'})
        cfg_all = _load_runtime_config()
        app_key = cfg_all.get(f'jdy{pfx}_app_key', '')
        app_secret = cfg_all.get(f'jdy{pfx}_app_secret', '')
        domain = cfg_all.get(f'jdy{pfx}_domain', '')
        account_id = cfg_all.get(f'jdy{pfx}_db_id', '')

        return jsonify({'success': True,
                        'app_key': app_key,
                        'app_secret_masked': app_secret[:6] + '…',
                        'domain': domain,
                        'account_id': account_id})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] jdy_refresh_auth: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-test', methods=['POST'])
def jdy_test():
    """测试 JDY API 连通性（idx=1 或 2 指定账套）"""
    try:
        data = request.get_json(force=True) or {}
        idx  = int(data.get('idx', 1))
        cli  = _ensure_jdy_client2() if idx == 2 else _ensure_jdy_client()
        if not cli:
            return jsonify({'success': False, 'error': f'请先填写账套{idx} JDY API 配置'})
        result = cli.test_connection()
        if result['ok']:
            return jsonify({'success': True, 'message': result['msg']})
        else:
            return jsonify({'success': False, 'error': result['msg']})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/purchase-orders', methods=['GET'])
def purchase_orders_list():
    """
    合并两账套购货入库单列表（并行查询）
    ?search=xxx&begin_date=2026-04-01&end_date=2026-05-15
    每条记录附加 _account 字段
    """
    from concurrent.futures import ThreadPoolExecutor
    try:
        search     = request.args.get('search', '')
        begin_date = request.args.get('begin_date', '')
        end_date   = request.args.get('end_date', '')

        cfg1 = _load_jdy_config();  name1 = cfg1.get('name', '饰品')
        cfg2 = _load_jdy_config2(); name2 = cfg2.get('name', '箱包')
        cli1 = _ensure_jdy_client()
        cli2 = _ensure_jdy_client2()

        if not cli1 and not cli2:
            return jsonify({'success': False, 'error': '请先配置 JDY API'})

        def _fetch(cli, name):
            r     = cli.get_purchase_orders(page=1, page_size=200,
                                            search=search,
                                            begin_date=begin_date, end_date=end_date)
            items = r.get('list') or []
            for item in items:
                item['_account'] = name
                # 整理 entries：只保留前端需要的字段
                item['_product_codes'] = [
                    e.get('productNumber', '') for e in (item.get('entries') or [])
                    if e.get('productNumber')
                ]
            print(f'[{name}] 购货单共 {len(items)} 条')
            return items

        all_items, errors = [], []
        with ThreadPoolExecutor(max_workers=2) as exe:
            tasks = {}
            if cli1: tasks[exe.submit(_fetch, cli1, name1)] = name1
            if cli2: tasks[exe.submit(_fetch, cli2, name2)] = name2
            for fut, name in tasks.items():
                try:
                    all_items.extend(fut.result())
                except Exception as e:
                    errors.append(f'{name}: {e}')

        all_items.sort(key=lambda x: x.get('date') or '', reverse=True)
        return jsonify({'success': True, 'list': all_items,
                        'total': len(all_items), 'errors': errors})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/transfer-orders', methods=['GET'])
def transfer_orders():
    """
    合并两个账套的调拨单列表（两账套并行查询）
    ?page=1&page_size=20&search=xxx&begin_date=2026-04-01&end_date=2026-05-15
    每条记录附加 _account 字段（账套名称）
    """
    from concurrent.futures import ThreadPoolExecutor
    try:
        page       = int(request.args.get('page', 1))
        page_size  = int(request.args.get('page_size', 20))
        search     = request.args.get('search', '')
        begin_date = request.args.get('begin_date', '')
        end_date   = request.args.get('end_date', '')
        source     = request.args.get('source', 'cache')

        cfg1 = _load_jdy_config()
        cfg2 = _load_jdy_config2()
        name1 = cfg1.get('name', '饰品')
        name2 = cfg2.get('name', '箱包')

        if source != 'live':
            all_items = _read_cached_transfer_orders(begin_date, end_date, 'all', search.strip().lower())
            return jsonify({
                'success': True,
                'list': all_items,
                'total': len(all_items),
                'account_names': [name1, name2],
                'errors': [],
                'cache': True,
                'cache_stats': _transfer_cache_stats(),
            })

        cli1 = _ensure_jdy_client()
        cli2 = _ensure_jdy_client2()

        if not cli1 and not cli2:
            return jsonify({'success': False, 'error': '请先配置 JDY API'})

        def _fetch(cli, name):
            r     = cli.get_transfer_orders(page=1, page_size=200, search=search,
                                            begin_date=begin_date, end_date=end_date)
            items = r.get('list') or []
            all_out_locs = set()
            items = [_normalize_transfer_order(item, name) for item in items]
            for item in items:
                all_out_locs.update(item.get('_outLocNames') or [])
            print(f'[{name}] 共 {len(items)} 条，全部调出库位: {all_out_locs}')
            return items

        all_items = []
        errors    = []
        tasks     = {}
        with ThreadPoolExecutor(max_workers=2) as exe:
            if cli1: tasks[exe.submit(_fetch, cli1, name1)] = name1
            if cli2: tasks[exe.submit(_fetch, cli2, name2)] = name2
            for fut, name in tasks.items():
                try:
                    all_items.extend(fut.result())
                except Exception as e:
                    errors.append(f'{name}: {e}')
                    print(f'[ERROR] {name} transfer_orders: {e}')

        # 按日期降序排列
        all_items.sort(key=lambda x: x.get('date') or x.get('createTime') or '', reverse=True)

        # 不做服务端分页，全量返回给前端（前端在过滤后再分页）
        return jsonify({'success': True, 'list': all_items, 'total': len(all_items),
                        'account_names': [name1, name2], 'errors': errors, 'cache': False})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


def _fetch_live_transfer_orders_all(cli, account_name, begin_date='', end_date='', search=''):
    page = 1
    page_size = 20
    rows = []
    total = None
    while True:
        result = cli.get_transfer_orders(
            page=page,
            page_size=page_size,
            search=search,
            begin_date=begin_date,
            end_date=end_date,
        )
        batch = result.get('list') or []
        rows.extend(_normalize_transfer_order(row, account_name) for row in batch)
        total = result.get('total') or total
        if not batch or len(batch) < page_size:
            break
        if total and len(rows) >= int(total):
            break
        page += 1
        if page > 80:
            break
        time.sleep(0.8)
    return rows


def _fetch_live_transfer_orders_by_day(cli, account_name, begin_date='', end_date='', search=''):
    if not begin_date or not end_date:
        return _fetch_live_transfer_orders_all(cli, account_name, begin_date, end_date, search)
    rows = []
    seen_numbers = set()
    for day in _iter_date_strings(begin_date, end_date):
        day_rows = _fetch_live_transfer_orders_all(cli, account_name, day, day, search)
        _log_event('TRANSFER_SYNC', f'{account_name}: {day} fetched {len(day_rows)} rows')
        for row in day_rows:
            key = str(row.get('number') or row.get('billNo') or row.get('id') or '')
            if key and key in seen_numbers:
                continue
            if key:
                seen_numbers.add(key)
            rows.append(row)
    return rows


def _refresh_transfer_cache(begin_date='', end_date='', account='all', search=''):
    started_at = datetime.now()
    seen = 0
    new_count = 0
    changed_count = 0
    unchanged_count = 0
    errors = []
    _log_event('TRANSFER_SYNC', f'start account={account} begin={begin_date} end={end_date} search={search}')
    for cli_fn, name in _sales_sources_for_account(account):
        try:
            cli = _client_for_sync_account(name, cli_fn)
            if not cli:
                errors.append(f'{name}: JDY API not configured')
                _log_event('TRANSFER_SYNC', f'{name}: JDY API not configured')
                continue
            rows = _fetch_live_transfer_orders_by_day(cli, name, begin_date, end_date, search)
            _log_event('TRANSFER_SYNC', f'{name}: fetched {len(rows)} rows from JDY')
            for attempt in range(3):
                try:
                    with _sales_cache_conn() as conn:
                        for row in rows:
                            number = row.get('number') or row.get('billNo') or row.get('id') or ''
                            new_hash = _transfer_order_signature(row)
                            old_hash = _cached_transfer_hash(conn, name, number)
                            if old_hash is None:
                                new_count += 1
                            elif old_hash != new_hash:
                                changed_count += 1
                            else:
                                unchanged_count += 1
                            _cache_upsert_transfer_order(conn, row)
                        conn.commit()
                    break
                except sqlite3.OperationalError as e:
                    if 'locked' not in str(e).lower() or attempt >= 2:
                        raise
                    _log_event('TRANSFER_SYNC', f'{name}: database locked, retry {attempt + 1}/3')
                    time.sleep(1.5 * (attempt + 1))
            seen += len(rows)
            msg = f'{name}: cached={len(rows)} window={begin_date}~{end_date}'
            print(f'[TRANSFER SYNC] {msg}')
            _log_event('TRANSFER_SYNC', msg)
        except Exception as e:
            err = f'{name}: {_short_sync_error(e)}'
            errors.append(err)
            print(f'[TRANSFER SYNC ERROR] {err}')
            _log_event('TRANSFER_SYNC_ERROR', err)
    result = {
        'seen': seen,
        'new_count': new_count,
        'changed_count': changed_count,
        'unchanged_count': unchanged_count,
        'errors': errors,
        'begin_date': begin_date,
        'end_date': end_date,
        'duration_seconds': round((datetime.now() - started_at).total_seconds(), 2),
        'stats': _transfer_cache_stats(),
    }
    _log_event('TRANSFER_SYNC', f'finish seen={seen} new={new_count} changed={changed_count} unchanged={unchanged_count} errors={errors}')
    return result


@app.route('/transfer-cache-refresh', methods=['POST'])
def transfer_cache_refresh():
    try:
        body = request.get_json(silent=True) or {}
        begin_date = request.args.get('begin_date', '') or body.get('begin_date', '')
        end_date = request.args.get('end_date', '') or body.get('end_date', '')
        account = request.args.get('account', 'all') or body.get('account', 'all')
        search = request.args.get('search', '') or body.get('search', '')
        result = _refresh_transfer_cache(begin_date, end_date, account, search)
        return jsonify({'success': not bool(result.get('errors')), **result})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] transfer_cache_refresh: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/transfer-order/<order_id>', methods=['GET'])
def transfer_order_detail(order_id):
    """获取调拨单详情，?account=账套名 路由到对应 client"""
    try:
        account = request.args.get('account', '')
        source = request.args.get('source', 'cache')
        if source != 'live':
            cached = _read_cached_transfer_detail(order_id, account)
            if cached:
                return jsonify({'success': True, 'data': cached, 'cache': True})
            return jsonify({'success': False, 'error': '本地没有这张调拨单，请先同步调拨单缓存'}), 404

        cfg1 = _load_jdy_config()
        cfg2 = _load_jdy_config2()
        if account == cfg2.get('name', '箱包'):
            cli = _ensure_jdy_client2()
        else:
            cli = _ensure_jdy_client()
        if not cli:
            return jsonify({'success': False, 'error': '请先配置 JDY API'})
        detail = cli.get_transfer_order_detail(order_id)
        return jsonify({'success': True, 'data': _normalize_transfer_order(detail, account), 'cache': False})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-auto-fill', methods=['POST'])
def jdy_auto_fill():
    """
    自动补全单个商品基础资料

    请求体:
    {
        "code":            "T.F240T-307",
        "account":         "祺航饰品",          // 可选，指定账套
        "supplier_number": "SUP001",            // 可选，用户选定的供应商编号
        "confirmed":       null 或 {...}        // 用户批量确认后传回
    }

    返回状态:
        needs_supplier_select=True  → 有多条购货单，前端弹窗选供应商
        needs_confirm=True          → AI 与历史冲突，前端加入批量确认队列
        否则                        → 直接写入 Excel
    """
    try:
        data            = request.get_json(force=True) or {}
        code            = (data.get('code') or '').strip()
        account         = (data.get('account') or '').strip()
        supplier_number = (data.get('supplier_number') or '').strip()
        confirmed       = data.get('confirmed')

        if not code:
            return jsonify({'success': False, 'error': '缺少 code 参数'})

        # 按账套选 client
        cfg2 = _load_jdy_config2()
        cli  = _ensure_jdy_client2() if account == cfg2.get('name') else _ensure_jdy_client()
        if not cli:
            # 尝试另一个
            cli = _ensure_jdy_client() or _ensure_jdy_client2()
        if not cli:
            return jsonify({'success': False, 'error': '请先配置 JDY API'})

        # ① 查商品详情
        product = cli.get_product_by_code(code)
        if not product:
            # 未在当前账套找到，换另一个账套试试
            other = _ensure_jdy_client2() if cli is _ensure_jdy_client() else _ensure_jdy_client()
            if other:
                product = other.get_product_by_code(code)
                if product:
                    cli = other
        if not product:
            return jsonify({'success': False, 'error': f'精斗云中未找到商品: {code}'})

        # ② 从商品档案中取字段
        from jdy_api import parse_dimensions
        category   = product.get('categoryName') or ''
        reg_no     = product.get('registrationNo') or ''
        remark     = product.get('remark') or ''
        name       = product.get('productName') or ''
        dimensions = parse_dimensions(reg_no)
        try:
            conversion = float(remark.strip()) if remark.strip() else None
        except ValueError:
            conversion = None

        # ③ 取商品第一张图
        imgs      = product.get('multiImg') or []
        image_url = imgs[0].get('url') if isinstance(imgs, list) and imgs else ''
        if not image_url:
            for sku in (product.get('invSku') or []):
                image_url = sku.get('skuImg') or ''
                if image_url:
                    break

        # ④ 购货入库单 → 供应商编号
        #    若调用方已传入 supplier_number（用户选了），直接用
        #    否则查购货单：单条自动取；多条返回列表让前端选
        origin = ''
        if not supplier_number:
            po_result = cli.get_purchase_orders_by_product(code, page_size=20)
            po_list   = po_result.get('list', [])
            if len(po_list) == 1:
                supplier_number = po_list[0].get('supplierNumber', '')
            elif len(po_list) > 1:
                # 多条购货单，前端需要弹窗让用户选
                return jsonify({
                    'success': True,
                    'needs_supplier_select': True,
                    'code': code,
                    'purchase_orders': [
                        {
                            'number':         r.get('number', ''),
                            'date':           r.get('date', ''),
                            'supplierNumber': r.get('supplierNumber', ''),
                            'supplierName':   r.get('supplierName', ''),
                            'totalQty':       r.get('totalQty', 0),
                        }
                        for r in po_list
                    ],
                })
            # len == 0：无购货单，origin 留空

        address_detail = ''
        if supplier_number:
            supplier = cli.get_supplier_by_number(supplier_number)
            if supplier:
                origin, address_detail = _extract_supplier_address(supplier)

        # ⑤ 调 AI 识别 + 写 Excel
        payload = {
            'code':      code,
            'name':      name,
            'category':  category,
            'image_url': image_url,
        }
        # 记录实际使用的账套名（cli 可能被换成另一账套）
        cfg1 = _load_jdy_config()
        actual_account = cfg2.get('name') if cli is _ensure_jdy_client2() else cfg1.get('name')

        extra_fields = {
            'category':        category,
            'origin':          origin,
            'address_detail':  address_detail,
            'dimensions':      dimensions,
            'conversion':      conversion,
            'supplier_number': supplier_number,
            'account':         actual_account,
        }

        result = identify_from_jdy(payload, extra_fields=extra_fields, confirmed=confirmed)
        return jsonify({'success': True, **result})

    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/purchase-orders-by-product', methods=['GET'])
def purchase_orders_by_product():
    """
    查询某商品的购货入库单列表（供选择供应商用）
    ?code=商品编号&account=账套名
    返回: [{number, date, supplierNumber, supplierName, totalQty, totalAmount}, ...]
    """
    try:
        code    = request.args.get('code', '').strip()
        account = request.args.get('account', '').strip()
        if not code:
            return jsonify({'success': False, 'error': '缺少 code 参数'})

        cfg2 = _load_jdy_config2()
        cli  = _ensure_jdy_client2() if account == cfg2.get('name') else _ensure_jdy_client()
        if not cli:
            return jsonify({'success': False, 'error': '请先配置 JDY API'})

        result = cli.get_purchase_orders_by_product(code, page_size=20)
        # 精简返回字段（只前端需要的）
        items = [
            {
                'number':         r.get('number', ''),
                'date':           r.get('date', ''),
                'supplierNumber': r.get('supplierNumber', ''),
                'supplierName':   r.get('supplierName', ''),
                'totalQty':       r.get('totalQty', 0),
                'totalAmount':    r.get('totalAmount', 0),
                'checkStatus':    r.get('checkStatus', False),
            }
            for r in result.get('list', [])
        ]
        return jsonify({'success': True, 'list': items, 'total': result.get('total', 0)})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] purchase_orders_by_product: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/compare-products-batch', methods=['POST'])
def compare_products_batch():
    """
    批量对比：Excel 已有商品 vs JDY 当前档案，找出需要处理的商品
    请求体: { "codes": ["T.xxx", ...] }   // 前端分批传，每次最多30个
    返回每个商品的状态:
        status: "ok"        — 数据完整，无需处理
                "missing"   — Excel 中不存在
                "incomplete"— 存在但有列为空(E/F/H/I/J/T)
                "error"     — JDY 查询失败
    """
    try:
        from ai_identify import _load_config, _find_excel, _EXCEL_LOCK
        import openpyxl

        data  = request.get_json(force=True) or {}
        codes = [c.strip() for c in (data.get('codes') or []) if c.strip()]
        if not codes:
            return jsonify({'success': False, 'error': '缺少 codes 参数'})

        config     = _load_config()
        excel_path = _find_excel(config)
        if not excel_path:
            return jsonify({'success': False, 'error': '未找到基础资料 Excel'})

        # 读 Excel 中已有数据（B,E,F,H,I,J,T 列）
        excel_data = {}   # code → {e,f,h,i,j,t}
        with _EXCEL_LOCK:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
            try:
                for row in ws.iter_rows(values_only=True):
                    if len(row) < 2 or row[1] is None:
                        continue
                    b = str(row[1]).strip()
                    excel_data[b] = {
                        'e': row[4]  if len(row) > 4  else None,   # E 海关编码
                        'f': row[5]  if len(row) > 5  else None,   # F 商品名称
                        'h': row[7]  if len(row) > 7  else None,   # H 货源地
                        'i': row[8]  if len(row) > 8  else None,   # I 材料
                        'j': row[9]  if len(row) > 9  else None,   # J 用途
                        't': row[19] if len(row) > 19 else None,   # T 供应商编号
                    }
            finally:
                wb.close()

        # 用两个账套批量查 JDY（每账套最多30个）
        jdy_map = {}   # code → product dict
        cli1, cli2 = _ensure_jdy_client(), _ensure_jdy_client2()
        for cli in [c for c in [cli1, cli2] if c]:
            try:
                prods = cli.get_products_by_codes(codes)
                for p in prods:
                    pn = str(p.get('productNumber', '')).strip()
                    if pn and pn not in jdy_map:
                        jdy_map[pn] = p
            except Exception as ex:
                print(f'[WARN] compare_products_batch JDY error: {ex}')

        results = []
        for code in codes:
            ex_row = excel_data.get(code)
            jdy_p  = jdy_map.get(code)

            if ex_row is None:
                results.append({'code': code, 'status': 'missing',
                                 'jdy_name': jdy_p.get('productName', '') if jdy_p else ''})
                continue

            # 检查关键列是否有空
            empty_cols = [col for col, val in [
                ('E海关编码', ex_row['e']), ('F商品名称', ex_row['f']),
                ('H货源地',   ex_row['h']), ('I材料',     ex_row['i']),
                ('J用途',     ex_row['j']), ('T供应商',   ex_row['t']),
            ] if not val]

            if empty_cols:
                results.append({'code': code, 'status': 'incomplete',
                                 'empty_cols': empty_cols,
                                 'jdy_name': jdy_p.get('productName', '') if jdy_p else ''})
            else:
                results.append({'code': code, 'status': 'ok'})

        summary = {s: sum(1 for r in results if r['status'] == s)
                   for s in ('ok', 'missing', 'incomplete', 'error')}
        return jsonify({'success': True, 'results': results, 'summary': summary})

    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] compare_products_batch: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/license-fill', methods=['POST'])
def license_fill():
    """
    档案补全：用商品档案生产许可证 (proLicense) 字段补全 Excel。
    body: {"codes": ["T.xxx", ...], "account": "饰品"}
    或单个: {"code": "T.xxx", "account": "饰品"}
    """
    try:
        data    = request.get_json(force=True) or {}
        codes   = data.get('codes') or ([data['code']] if data.get('code') else [])
        account = (data.get('account') or '').strip()
        if not codes:
            return jsonify({'success': False, 'error': '缺少 codes 参数'})

        results = []
        for code in codes:
            r = _do_license_fill_one(code, account)
            results.append(r or {'code': code, 'status': 'skipped'})

        ok    = [r for r in results if r.get('status') == 'ok']
        other = [r for r in results if r.get('status') != 'ok']
        return jsonify({'success': True, 'results': results,
                        'ok_count': len(ok), 'skip': other})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] license_fill: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/clear-supplier-cache', methods=['POST'])
def clear_supplier_cache():
    """清空指定商品或全部商品的供应商查询缓存。"""
    try:
        body = request.get_json(silent=True) or {}
        codes = [str(c).strip() for c in (body.get('codes') or []) if str(c).strip()]
        cache = _load_code_cache()
        total_before = len(cache)
        if codes:
            cleared = 0
            for code in codes:
                if code in cache:
                    del cache[code]
                    cleared += 1
        else:
            cleared = total_before
            cache = {}
        _save_code_cache(cache)
        _supplier_lookup_cache.clear()
        _city_lookup_cache.clear()
        return jsonify({'success': True, 'cleared': cleared, 'total_before': total_before})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _do_license_preview_one(code, account, existing_codes, cli1, cli2, cfg1_name, cfg2_name):
    """
    查询单个商品 proLicense，返回预览数据（不写 Excel）。
    existing_codes: 预先读取的 Excel 编码 set，避免重复 IO。
    cli1/cli2: 预先初始化好的 JDY 客户端，避免并发重复认证。
    """
    from ai_identify import parse_prolicense
    from jdy_api import parse_dimensions

    cli = cli2 if account == cfg2_name else cli1
    if not cli:
        cli = cli1 or cli2
    if not cli:
        return {'code': code, 'status': 'no_client'}

    product = cli.get_product_by_code(code)
    if not product:
        other = cli2 if cli is cli1 else cli1
        if other:
            product = other.get_product_by_code(code)
            if product:
                cli = other
    if not product:
        return {'code': code, 'status': 'not_found'}

    in_excel    = code in existing_codes
    pro_license = product.get('proLicense') or ''
    name_part, material_part = parse_prolicense(pro_license)
    if not name_part or not material_part:
        return {'code': code, 'status': 'no_license', 'in_excel': in_excel, 'proLicense': pro_license}

    actual = cfg2_name if cli is cli2 else cfg1_name
    remark = product.get('remark') or ''
    try:
        conversion = float(remark.strip()) if remark.strip() else None
    except ValueError:
        conversion = None

    multi_img = product.get('multiImg') or []
    img_url   = multi_img[0].get('url', '') if multi_img else ''

    preview = preview_from_license(code, name_part, material_part, extra_fields={
        'category':   product.get('categoryName') or '',
        'dimensions': parse_dimensions(product.get('registrationNo') or ''),
        'conversion': conversion,
        'account':    actual,
    })
    return {'status': 'ok', 'in_excel': in_excel, 'img_url': img_url, **preview}


@app.route('/license-preview', methods=['POST'])
def license_preview():
    """
    档案补全预览：返回各商品的预览数据，不写入 Excel。
    body: {"codes": [...], "account": "饰品"}
    """
    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from ai_identify import _find_excel, get_existing_codes

        data    = request.get_json(force=True) or {}
        codes   = data.get('codes') or ([data['code']] if data.get('code') else [])
        account = (data.get('account') or '').strip()
        if not codes:
            return jsonify({'success': False, 'error': '缺少 codes 参数'})

        # 预加载：只读一次 Excel + 只初始化一次 JDY 客户端
        ai_cfg        = _load_ai_config()
        excel_path    = _find_excel(ai_cfg)
        existing_codes = set()
        if excel_path:
            try:
                existing_codes = get_existing_codes(excel_path)
            except Exception:
                pass

        cli1      = _ensure_jdy_client()
        cli2      = _ensure_jdy_client2()
        cfg1_name = _load_jdy_config().get('name', '')
        cfg2_name = _load_jdy_config2().get('name', '')

        # 并发查询（IO 密集，线程池加速）
        results_map = {}
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = {
                pool.submit(_do_license_preview_one,
                            code, account, existing_codes,
                            cli1, cli2, cfg1_name, cfg2_name): code
                for code in codes
            }
            for fut in as_completed(futures):
                code = futures[fut]
                try:
                    results_map[code] = fut.result()
                except Exception as ex:
                    results_map[code] = {'code': code, 'status': 'error', 'error': str(ex)}

        # 保持原始顺序
        results = [results_map.get(c, {'code': c, 'status': 'skipped'}) for c in codes]
        ok    = [r for r in results if r.get('status') == 'ok']
        other = [r for r in results if r.get('status') != 'ok']
        return jsonify({'success': True, 'results': results,
                        'ok_count': len(ok), 'skip': other})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] license_preview: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/license-confirm-write', methods=['POST'])
def license_confirm_write():
    """
    档案补全写入：将用户在确认弹窗中确认（可能修改过）的数据写入 Excel。
    body: {"items": [{"code","product_name","material","hs_code","usage","account","category",...}]}
    """
    try:
        data  = request.get_json(force=True) or {}
        items = data.get('items') or []
        if not items:
            return jsonify({'success': False, 'error': '缺少 items 参数'})

        from ai_identify import _load_config, _find_excel
        ai_cfg     = _load_ai_config()
        excel_path = _find_excel(ai_cfg)
        if not excel_path:
            return jsonify({'success': False, 'error': '未找到 Excel 文件，请先在设置中配置'})

        results = []
        for item in items:
            code = item.get('code', '')
            try:
                r = write_confirmed_license(
                    excel_path,
                    code,
                    product_name=item.get('product_name') or '',
                    material=item.get('material') or '',
                    hs_code=item.get('hs_code') or '',
                    usage=item.get('usage') or '',
                    extra_fields={
                        'account':   item.get('account') or '',
                        'category':  item.get('category') or '',
                    },
                )
                results.append({'code': code, 'status': 'ok', 'row': r.get('row')})
            except Exception as ex:
                results.append({'code': code, 'status': 'error', 'error': str(ex)})

        ok   = sum(1 for r in results if r['status'] == 'ok')
        fail = len(results) - ok
        return jsonify({'success': True, 'results': results, 'ok': ok, 'fail': fail})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] license_confirm_write: {tb}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-product', methods=['GET'])
def jdy_product():
    """根据编号查询精斗云商品原始数据（调试用），自动尝试两个账套"""
    try:
        code = request.args.get('code', '').strip()
        if not code:
            return jsonify({'success': False, 'error': '缺少 code 参数'})
        # 先查账套1，没找到再查账套2
        cli1 = _ensure_jdy_client()
        product = cli1.get_product_by_code(code) if cli1 else None
        account = _load_jdy_config().get('name', '饰品')
        if not product:
            cli2 = _ensure_jdy_client2()
            if cli2:
                product = cli2.get_product_by_code(code)
                account = _load_jdy_config2().get('name', '箱包')
        return jsonify({'success': True, 'data': product, 'account': account})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── PyWebView 桌面 UI ─────────────────────────────────────────────────────────

@app.route('/generation-history', methods=['GET'])
def generation_history():
    """返回所有调拨单的生成次数历史"""
    return jsonify({'success': True, 'data': _load_gen_history()})


@app.route('/list-items-json', methods=['GET'])
def list_items_json():
    """列出 _items_store/ 目录下所有可用的 items.json 存档，供 Stage 2 下拉选择。"""
    try:
        import glob as _glob
        from excel_gen import _ITEMS_STORE
        files = sorted(
            _glob.glob(os.path.join(_ITEMS_STORE, '*_items.json')),
            key=os.path.getmtime, reverse=True
        )
        result = []
        for fpath in files:
            try:
                with open(fpath, encoding='utf-8') as fp:
                    d = json.load(fp)
                result.append({
                    'filename':     os.path.basename(fpath),
                    'order_no':     d.get('order_no', ''),
                    'invoice_no':   d.get('invoice_no', ''),
                    'output_path':  d.get('output_path', ''),
                    'generated_at': d.get('generated_at', ''),
                    'item_count':   len(d.get('items', [])),
                })
            except Exception:
                pass
        return jsonify({'success': True, 'files': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/process-pdf', methods=['POST'])
def process_pdf():
    """
    阶段二：处理退税联 PDF，生成发票导入模板、更新基础资料、生成凌航发票。
    请求体：{items_filename, pdf_path}
      items_filename: _items_store/ 中的文件名（如 DB20260512002_items.json）
    """
    try:
        body = request.get_json(force=True) or {}
        pdf_path       = (body.get('pdf_path') or '').strip()
        items_filename = (body.get('items_filename') or '').strip()
        export_date    = (body.get('export_date') or '').strip()
        customs_no     = (body.get('customs_no') or '').strip()
        bl_no          = (body.get('bl_no') or '').strip()

        if not items_filename:
            return jsonify({'success': False, 'error': '请选择要处理的报关批次'}), 400
        if pdf_path and not os.path.exists(pdf_path):
            return jsonify({'success': False, 'error': f'PDF 文件不存在: {pdf_path}'}), 400

        # 从 _items_store 读取存档
        from excel_gen import _ITEMS_STORE
        items_json_path = os.path.join(_ITEMS_STORE, items_filename)
        if not os.path.exists(items_json_path):
            return jsonify({'success': False,
                            'error': f'找不到存档文件: {items_filename}'}), 400
        print(f'[PDF] 使用存档文件: {items_json_path}')

        with open(items_json_path, encoding='utf-8') as f:
            saved = json.load(f)

        # 从存档读取 invoice_no / order_no / output_path
        invoice_no  = saved.get('invoice_no', '')
        order_no    = saved.get('order_no', '')
        output_path = saved.get('output_path', '')
        if not output_path or not os.path.isdir(output_path):
            return jsonify({'success': False,
                            'error': f'存档中的输出目录不存在或已移动: {output_path}'}), 400

        # 供应商映射：优先从 items.json 读（Stage 1 已存），避免重复调 API
        supplier_map = saved.get('supplier_map') or {}
        if supplier_map:
            print(f'[PDF] 使用 items.json 缓存的 supplier_map（{len(supplier_map)} 个）')
        else:
            print('[PDF] items.json 无 supplier_map，降级调 API...')
            try:
                supplier_map, _ = _lookup_suppliers(saved.get('items', []))
            except Exception as e:
                print(f'[WARN] 退税联供应商查询失败: {e}')
                supplier_map = {}

        # 从外管局获取本月汇率（与汇率按钮逻辑相同），用于发票 RMB 金额换算
        try:
            inv_rate, inv_rate_date = _fetch_safe_exchange_rate()
            if inv_rate:
                print(f'[RATE] 发票汇率（外管局）{inv_rate_date}: {inv_rate}')
            else:
                inv_rate = float(saved.get('exchange_rate') or 0)
                print(f'[RATE] 外管局无数据，使用存档汇率: {inv_rate}')
        except Exception as e:
            inv_rate = float(saved.get('exchange_rate') or 0)
            print(f'[WARN] 获取发票汇率失败，使用存档汇率 {inv_rate}: {e}')

        # 回写发票汇率到 items.json，供凌航发票等直接读取
        if inv_rate and inv_rate != float(saved.get('invoice_exchange_rate') or 0):
            saved['invoice_exchange_rate'] = inv_rate
            with open(items_json_path, 'w', encoding='utf-8') as _f:
                json.dump(saved, _f, ensure_ascii=False, indent=2)
            print(f'[RATE] 发票汇率 {inv_rate} 已写入 items.json')

        result = process_tax_pdf({
            'invoice_no':            invoice_no,
            'order_no':              order_no,
            'pdf_path':              pdf_path,
            'output_path':           output_path,
            'items_json_path':       items_json_path,
            'supplier_map':          supplier_map,
            'export_date':           export_date,
            'customs_no':            customs_no,
            'bl_no':                 bl_no,
            'invoice_exchange_rate': inv_rate,
        })
        return jsonify({'success': True, **result})

    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] /process-pdf: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


def _mock_sales_orders(date_str):
    """销货单页面骨架数据。后续替换为精斗云销货单/销售出库单接口。"""
    date_str = date_str or datetime.now().strftime('%Y-%m-%d')
    base_products = [
        {
            'code': 'C.D767C-46',
            'name': '白钻石手表手镯',
            'spec': '10 Ai06',
            'barcode': '240705C120021',
            'qty': 20,
            'unit': 'DZ',
            'amount': 1980,
            'warehouses': [
                {'name': '新大仓库', 'qty': 166},
                {'name': '在途', 'qty': 54},
                {'name': '工厂', 'qty': 120},
            ],
            'purchase': {'ordered': 80, 'received': 26, 'pending': 54},
        },
        {
            'code': 'C.D767C-47',
            'name': '水晶手表手镯',
            'spec': '10 Aj01',
            'barcode': '240705C120022',
            'qty': 10,
            'unit': 'DZ',
            'amount': 990,
            'warehouses': [
                {'name': '新大仓库', 'qty': 74},
                {'name': '在途', 'qty': 36},
                {'name': '工厂', 'qty': 78},
            ],
            'purchase': {'ordered': 60, 'received': 24, 'pending': 36},
        },
        {
            'code': 'C.D767C-44',
            'name': '五排钻石手表 + 手镯',
            'spec': '10 Aj02',
            'barcode': '240705C120019',
            'qty': 20,
            'unit': 'DZ',
            'amount': 1980,
            'warehouses': [
                {'name': '新大仓库', 'qty': 58},
                {'name': '在途', 'qty': 36},
                {'name': '工厂', 'qty': 96},
            ],
            'purchase': {'ordered': 48, 'received': 12, 'pending': 36},
        },
    ]
    return [
        {
            'number': 'XH20260525008',
            'date': date_str,
            'customerName': 'AL KABAYEL DISCOUNT CENTER',
            'account': '祺航箱包',
            'totalQty': 40,
            'totalAmount': 1081.50,
            'checkStatusName': '未审核',
            'entries': [base_products[2], base_products[0]],
        },
        {
            'number': 'XH20260525007',
            'date': date_str,
            'customerName': 'DRAGON GIFT CENTER',
            'account': '祺航饰品',
            'totalQty': 170,
            'totalAmount': 16670.00,
            'checkStatusName': '未审核',
            'entries': base_products,
        },
        {
            'number': 'XH20260525006',
            'date': date_str,
            'customerName': 'CASH',
            'account': '祺航饰品',
            'totalQty': 19,
            'totalAmount': 1199.10,
            'checkStatusName': '未审核',
            'entries': [base_products[1]],
        },
    ]


_SALES_CACHE_DIR = os.path.join(_DATA_BASE, '_sales_cache')
_SALES_CACHE_DB = os.path.join(_SALES_CACHE_DIR, 'sales_cache.sqlite3')
_LOCAL_JSON_CACHE = {}


def _sales_cache_conn():
    os.makedirs(_SALES_CACHE_DIR, exist_ok=True)
    conn = sqlite3.connect(_SALES_CACHE_DB, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA busy_timeout=30000')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_orders (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            date TEXT,
            customer_name TEXT,
            total_qty REAL DEFAULT 0,
            total_amount REAL DEFAULT 0,
            check_status TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_details (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            date TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS transfer_orders (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            date TEXT,
            out_location TEXT,
            check_status TEXT,
            remark TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS transfer_details (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            date TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sales_product_quantities (
            account TEXT NOT NULL,
            code TEXT NOT NULL,
            stock_new REAL DEFAULT 0,
            stock_transit REAL DEFAULT 0,
            stock_local REAL DEFAULT 0,
            stock_factory REAL DEFAULT 0,
            factory_qty REAL DEFAULT 0,
            updated_at TEXT NOT NULL,
            error TEXT DEFAULT '',
            PRIMARY KEY (account, code)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS accessory_purchase_orders (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            date TEXT,
            supplier_name TEXT,
            supplier_number TEXT,
            total_qty REAL DEFAULT 0,
            total_amount REAL DEFAULT 0,
            bill_status TEXT,
            bill_status_name TEXT,
            entries_count INTEGER DEFAULT 0,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS jdy_suppliers (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            name TEXT,
            category_text TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS accessory_products (
            account TEXT NOT NULL,
            code TEXT NOT NULL,
            name TEXT,
            category TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, code)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS jdy_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT,
            product_id TEXT,
            product_number TEXT,
            product_name TEXT,
            spec TEXT,
            barcode TEXT,
            category_id TEXT,
            category_name TEXT,
            unit_id TEXT,
            unit_name TEXT,
            default_supplier_id TEXT,
            default_supplier_number TEXT,
            default_supplier_name TEXT,
            image_url TEXT,
            status TEXT,
            data_json TEXT,
            created_at TEXT,
            updated_at TEXT,
            last_seen_at TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS webhook_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT,
            biz_type TEXT,
            resource TEXT,
            bill_no TEXT,
            action TEXT,
            status TEXT NOT NULL DEFAULT 'pending',
            attempts INTEGER NOT NULL DEFAULT 0,
            payload_json TEXT NOT NULL,
            error TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            processed_at TEXT DEFAULT ''
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS webhook_runtime_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            processing_mode TEXT NOT NULL DEFAULT 'manual',
            auto_enabled INTEGER NOT NULL DEFAULT 0,
            paused INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS reorder_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT,
            supplier_number TEXT,
            supplier_name TEXT,
            code TEXT NOT NULL,
            name TEXT,
            spec TEXT,
            barcode TEXT,
            unit TEXT,
            image_url TEXT,
            source_type TEXT,
            source_number TEXT,
            source_date TEXT,
            source_customer TEXT,
            source_user TEXT,
            sales_qty_60 REAL DEFAULT 0,
            stock_new REAL DEFAULT 0,
            stock_transit REAL DEFAULT 0,
            stock_local REAL DEFAULT 0,
            stock_factory REAL DEFAULT 0,
            factory_qty REAL DEFAULT 0,
            suggested_qty REAL DEFAULT 0,
            confirmed_qty REAL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'pending',
            note TEXT DEFAULT '',
            created_by TEXT,
            created_by_name TEXT,
            reorder_price REAL DEFAULT 0,
            generated_batch_no TEXT,
            generated_at TEXT,
            sync_status TEXT DEFAULT 'not_synced',
            jdy_order_no TEXT,
            synced_at TEXT,
            sync_error TEXT,
            data_json TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS purchase_inbounds (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            date TEXT,
            supplier_number TEXT,
            supplier_name TEXT,
            total_qty REAL DEFAULT 0,
            total_amount REAL DEFAULT 0,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS purchase_inbound_attachments (
            account TEXT NOT NULL,
            number TEXT NOT NULL,
            attachment_key TEXT NOT NULL,
            name TEXT,
            url TEXT,
            local_path TEXT,
            data_json TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account, number, attachment_key)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS bill_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT,
            source_type TEXT,
            source_number TEXT,
            source_date TEXT,
            bill_type TEXT,
            attachment_key TEXT,
            name TEXT,
            url TEXT,
            local_path TEXT,
            file_mime TEXT,
            file_size INTEGER DEFAULT 0,
            download_status TEXT DEFAULT '',
            download_error TEXT DEFAULT '',
            data_json TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_sales_orders_date ON sales_orders(date)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_sales_orders_customer ON sales_orders(customer_name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_transfer_orders_date ON transfer_orders(date)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_transfer_orders_out_location ON transfer_orders(out_location)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_accessory_po_supplier ON accessory_purchase_orders(supplier_name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_accessory_po_date ON accessory_purchase_orders(date)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_jdy_suppliers_account_name ON jdy_suppliers(account, name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_accessory_products_category ON accessory_products(category)')
    _ensure_jdy_product_columns(conn)
    conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_jdy_products_unique ON jdy_products(account, product_number)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_jdy_products_name ON jdy_products(account, product_name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_jdy_products_category ON jdy_products(account, category_name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_jdy_products_updated ON jdy_products(account, updated_at)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_jdy_products_seen ON jdy_products(account, last_seen_at)')
    _ensure_customs_product_tables(conn)
    webhook_cols = {str(row['name']).lower() for row in conn.execute('PRAGMA table_info(webhook_events)').fetchall()}
    webhook_additions = [
        ('event_id', 'TEXT'),
        ('payload_hash', 'TEXT'),
        ('updated_at', 'TEXT'),
    ]
    for name, ddl in webhook_additions:
        key = name.lower()
        if key not in webhook_cols:
            conn.execute(f'ALTER TABLE webhook_events ADD COLUMN {name} {ddl}')
            webhook_cols.add(key)
    conn.execute('CREATE INDEX IF NOT EXISTS idx_webhook_events_status ON webhook_events(status, id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_webhook_events_bill ON webhook_events(account, bill_no, resource)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_webhook_events_event_id ON webhook_events(event_id, account, resource, bill_no)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_webhook_events_payload_hash ON webhook_events(payload_hash)')
    _ensure_reorder_item_columns(conn)
    conn.execute('CREATE INDEX IF NOT EXISTS idx_reorder_items_supplier ON reorder_items(status, supplier_name, supplier_number)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_reorder_items_code ON reorder_items(account, code, status)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_reorder_items_picker ON reorder_items(status, created_by_name, created_by)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_reorder_items_source ON reorder_items(account, source_number, code)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_reorder_items_batch ON reorder_items(generated_batch_no, sync_status, generated_at)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_purchase_inbounds_code_json ON purchase_inbounds(account, number)')
    conn.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_bill_attachments_unique
        ON bill_attachments(account, source_type, source_number, attachment_key)
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_bill_attachments_source
        ON bill_attachments(account, source_type, source_number)
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_bill_attachments_date
        ON bill_attachments(source_date)
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS purchase_order_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT NOT NULL,
            order_number TEXT NOT NULL,
            order_date TEXT NOT NULL,
            supplier_name TEXT,
            supplier_number TEXT,
            product_number TEXT NOT NULL,
            price REAL NOT NULL,
            qty REAL DEFAULT 0,
            unit TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_pop_product
        ON purchase_order_prices(account, product_number)
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_bill_attachments_status
        ON bill_attachments(download_status)
    ''')
    cols = {str(row['name']).lower() for row in conn.execute('PRAGMA table_info(sales_product_quantities)').fetchall()}
    if 'stock_local' not in cols:
        conn.execute('ALTER TABLE sales_product_quantities ADD COLUMN stock_local REAL DEFAULT 0')
        cols.add('stock_local')
    if 'stock_factory' not in cols:
        conn.execute('ALTER TABLE sales_product_quantities ADD COLUMN stock_factory REAL DEFAULT 0')
        cols.add('stock_factory')
    _ensure_reorder_item_columns(conn)
    return conn


def _ensure_reorder_item_columns(conn):
    cols = {str(row['name']).lower() for row in conn.execute('PRAGMA table_info(reorder_items)').fetchall()}
    additions = [
        ('created_by', 'TEXT'),
        ('created_by_name', 'TEXT'),
        ('reorder_price', 'REAL DEFAULT 0'),
        ('generated_batch_no', 'TEXT'),
        ('generated_at', 'TEXT'),
        ('sync_status', "TEXT DEFAULT 'not_synced'"),
        ('jdy_order_no', 'TEXT'),
        ('synced_at', 'TEXT'),
        ('sync_error', 'TEXT'),
    ]
    for name, ddl in additions:
        key = name.lower()
        if key not in cols:
            conn.execute(f'ALTER TABLE reorder_items ADD COLUMN {name} {ddl}')
            cols.add(key)


def _ensure_jdy_product_columns(conn):
    cols = {str(row['name']).lower() for row in conn.execute('PRAGMA table_info(jdy_products)').fetchall()}
    additions = [
        ('account', 'TEXT'),
        ('product_id', 'TEXT'),
        ('product_number', 'TEXT'),
        ('product_name', 'TEXT'),
        ('spec', 'TEXT'),
        ('barcode', 'TEXT'),
        ('category_id', 'TEXT'),
        ('category_name', 'TEXT'),
        ('unit_id', 'TEXT'),
        ('unit_name', 'TEXT'),
        ('default_supplier_id', 'TEXT'),
        ('default_supplier_number', 'TEXT'),
        ('default_supplier_name', 'TEXT'),
        ('image_url', 'TEXT'),
        ('status', 'TEXT'),
        ('data_json', 'TEXT'),
        ('created_at', 'TEXT'),
        ('updated_at', 'TEXT'),
        ('last_seen_at', 'TEXT'),
    ]
    for name, ddl in additions:
        key = name.lower()
        if key not in cols:
            conn.execute(f'ALTER TABLE jdy_products ADD COLUMN {name} {ddl}')
            cols.add(key)


def _cache_upsert_sales_order(conn, order):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT OR REPLACE INTO sales_orders
        (account, number, date, customer_name, total_qty, total_amount, check_status, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        order.get('account') or '',
        order.get('number') or '',
        order.get('date') or '',
        order.get('customerName') or '',
        _num(order.get('totalQty')),
        _num(order.get('totalAmount')),
        str(order.get('checkStatusName') or ''),
        json.dumps(order, ensure_ascii=False),
        now,
    ))


def _cache_upsert_sales_detail(conn, order):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT OR REPLACE INTO sales_details
        (account, number, date, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        order.get('account') or '',
        order.get('number') or '',
        order.get('date') or '',
        json.dumps(order, ensure_ascii=False),
        now,
    ))


def _cache_upsert_sales_product_quantity(conn, account, code, stock, factory_qty, error=''):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT OR REPLACE INTO sales_product_quantities
        (account, code, stock_new, stock_transit, stock_local, stock_factory, factory_qty, updated_at, error)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        account or '',
        code or '',
        _num((stock or {}).get('新大仓库')),
        _num((stock or {}).get('在途')),
        _num((stock or {}).get('金华/本地')),
        _num((stock or {}).get('工厂订单')),
        _num(factory_qty),
        now,
        str(error or ''),
    ))


def _read_cached_sales_product_quantities(account, codes):
    codes = [str(c).strip() for c in codes if c]
    if not codes:
        return {}
    result = {}
    with _sales_cache_conn() as conn:
        for i in range(0, len(codes), 800):
            batch = codes[i:i + 800]
            placeholders = ','.join('?' for _ in batch)
            rows = conn.execute(f'''
                SELECT code, stock_new, stock_transit, stock_local, stock_factory, factory_qty, updated_at, error
                FROM sales_product_quantities
                WHERE account = ? AND code IN ({placeholders})
            ''', [account, *batch]).fetchall()
            for row in rows:
                result[row['code']] = {
                    'stock': {
                        '新大仓库': _num(row['stock_new']),
                        '在途': _num(row['stock_transit']),
                        '金华/本地': _num(row['stock_local']),
                        '工厂订单': _num(row['stock_factory']),
                    },
                    'factory_qty': _num(row['factory_qty']),
                    'updated_at': row['updated_at'] or '',
                    'error': row['error'] or '',
                }
    return result


def _sales_order_list_projection(order):
    return {
        'number': order.get('number') or '',
        'date': order.get('date') or '',
        'customerName': order.get('customerName') or '',
        'account': order.get('account') or '',
        'totalQty': order.get('totalQty') or 0,
        'totalAmount': order.get('totalAmount') or 0,
        'checkStatusName': order.get('checkStatusName') or '',
        'updatedAt': order.get('updatedAt') or order.get('updated_at') or '',
    }


def _read_cached_sales_orders(date_str, account='all', search=''):
    with _sales_cache_conn() as conn:
        clauses = ['date = ?']
        params = [date_str]
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if search:
            clauses.append('(LOWER(number) LIKE ? OR LOWER(customer_name) LIKE ? OR LOWER(data_json) LIKE ?)')
            kw = f'%{search.lower()}%'
            params.extend([kw, kw, kw])
        rows = conn.execute(f'''
            SELECT data_json, updated_at FROM sales_orders
            WHERE {' AND '.join(clauses)}
            ORDER BY number DESC
        ''', params).fetchall()
    items = []
    for row in rows:
        order = json.loads(row['data_json'])
        order.setdefault('updatedAt', row['updated_at'] if 'updated_at' in row.keys() else '')
        items.append(_sales_order_list_projection(order))
    return items


def _read_cached_sales_orders_by_updated(account='all', search='', updated_from='', updated_to='', limit=100):
    limit = max(1, min(int(limit or 100), 500))
    conn = _sales_readonly_conn()
    if not conn:
        return []
    try:
        clauses = ['1 = 1']
        params = []
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if updated_from:
            clauses.append('updated_at >= ?')
            params.append(str(updated_from)[:19])
        if updated_to:
            clauses.append('updated_at <= ?')
            params.append(str(updated_to)[:19])
        if search:
            clauses.append('(LOWER(number) LIKE ? OR LOWER(customer_name) LIKE ? OR LOWER(data_json) LIKE ?)')
            kw = f'%{search.lower()}%'
            params.extend([kw, kw, kw])
        rows = conn.execute(f'''
            SELECT data_json, updated_at
            FROM sales_orders
            WHERE {' AND '.join(clauses)}
            ORDER BY updated_at DESC, number DESC
            LIMIT ?
        ''', [*params, limit]).fetchall()
    finally:
        conn.close()
    items = []
    for row in rows:
        order = json.loads(row['data_json'])
        order['updatedAt'] = row['updated_at'] or ''
        items.append(_sales_order_list_projection(order))
    return items


def _read_cached_sales_detail(order_no, account=''):
    with _sales_cache_conn() as conn:
        if account and account != 'all':
            row = conn.execute('''
                SELECT data_json FROM sales_details
                WHERE account = ? AND number = ?
            ''', (account, order_no)).fetchone()
        else:
            row = conn.execute('''
                SELECT data_json FROM sales_details
                WHERE number = ?
                ORDER BY updated_at DESC
                LIMIT 1
            ''', (order_no,)).fetchone()
    return json.loads(row['data_json']) if row else None


def _sales_readonly_conn():
    if not os.path.exists(_SALES_CACHE_DB):
        return None
    uri = 'file:' + os.path.abspath(_SALES_CACHE_DB).replace('\\', '/') + '?mode=ro'
    conn = sqlite3.connect(uri, uri=True, timeout=10)
    conn.row_factory = sqlite3.Row
    return conn


CUSTOMS_PRODUCT_COLUMNS = [
    'account', 'product_code', 'product_name', 'category_name', 'spec', 'unit', 'barcode',
    'jdy_product_id', 'customs_name_cn', 'customs_name_en', 'hs_code', 'customs_unit',
    'origin', 'material', 'usage', 'tax_refund_rate', 'is_tax_refund',
    'gross_weight_per_pkg', 'net_weight_per_pkg', 'carton_length', 'carton_width',
    'carton_height', 'carton_volume', 'pcs_per_pkg', 'supplier_number', 'tax_code',
    'tax_category_short_name', 'source', 'source_excel_path', 'source_row_no',
    'created_at', 'updated_at', 'updated_by',
]
CUSTOMS_PRODUCT_NUMERIC_FIELDS = {
    'tax_refund_rate', 'gross_weight_per_pkg', 'net_weight_per_pkg', 'carton_length',
    'carton_width', 'carton_height', 'carton_volume', 'pcs_per_pkg',
}
CUSTOMS_PRODUCT_REQUIRED_FIELDS = [
    'hs_code', 'customs_name_cn', 'customs_unit', 'origin', 'material', 'usage',
]
CUSTOMS_PRODUCT_MUTABLE_FIELDS = [
    'account', 'product_name', 'category_name', 'spec', 'unit', 'barcode', 'jdy_product_id',
    'customs_name_cn', 'customs_name_en', 'hs_code', 'customs_unit', 'origin', 'material',
    'usage', 'tax_refund_rate', 'is_tax_refund', 'gross_weight_per_pkg', 'net_weight_per_pkg',
    'carton_length', 'carton_width', 'carton_height', 'carton_volume', 'pcs_per_pkg',
    'supplier_number', 'tax_code', 'tax_category_short_name',
]


def _ensure_customs_product_tables(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS customs_product_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT,
            product_code TEXT NOT NULL UNIQUE,
            product_name TEXT,
            category_name TEXT,
            spec TEXT,
            unit TEXT,
            barcode TEXT,
            jdy_product_id TEXT,
            customs_name_cn TEXT,
            customs_name_en TEXT,
            hs_code TEXT,
            customs_unit TEXT,
            origin TEXT,
            material TEXT,
            usage TEXT,
            tax_refund_rate REAL,
            is_tax_refund TEXT,
            gross_weight_per_pkg REAL,
            net_weight_per_pkg REAL,
            carton_length REAL,
            carton_width REAL,
            carton_height REAL,
            carton_volume REAL,
            pcs_per_pkg REAL,
            supplier_number TEXT,
            tax_code TEXT,
            tax_category_short_name TEXT,
            source TEXT,
            source_excel_path TEXT,
            source_row_no INTEGER,
            created_at TEXT,
            updated_at TEXT,
            updated_by TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS customs_product_master_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_code TEXT NOT NULL,
            changed_at TEXT NOT NULL,
            changed_by TEXT,
            change_source TEXT,
            old_json TEXT,
            new_json TEXT,
            created_at TEXT NOT NULL
        )
    ''')
    cols = {
        str(row['name']).lower()
        for row in conn.execute('PRAGMA table_info(customs_product_master)').fetchall()
    }
    additions = [
        ('account', 'TEXT'),
        ('product_code', 'TEXT'),
        ('product_name', 'TEXT'),
        ('category_name', 'TEXT'),
        ('spec', 'TEXT'),
        ('unit', 'TEXT'),
        ('barcode', 'TEXT'),
        ('jdy_product_id', 'TEXT'),
        ('customs_name_cn', 'TEXT'),
        ('customs_name_en', 'TEXT'),
        ('hs_code', 'TEXT'),
        ('customs_unit', 'TEXT'),
        ('origin', 'TEXT'),
        ('material', 'TEXT'),
        ('usage', 'TEXT'),
        ('tax_refund_rate', 'REAL'),
        ('is_tax_refund', 'TEXT'),
        ('gross_weight_per_pkg', 'REAL'),
        ('net_weight_per_pkg', 'REAL'),
        ('carton_length', 'REAL'),
        ('carton_width', 'REAL'),
        ('carton_height', 'REAL'),
        ('carton_volume', 'REAL'),
        ('pcs_per_pkg', 'REAL'),
        ('supplier_number', 'TEXT'),
        ('tax_code', 'TEXT'),
        ('tax_category_short_name', 'TEXT'),
        ('source', 'TEXT'),
        ('source_excel_path', 'TEXT'),
        ('source_row_no', 'INTEGER'),
        ('created_at', 'TEXT'),
        ('updated_at', 'TEXT'),
        ('updated_by', 'TEXT'),
    ]
    for name, ddl in additions:
        if name.lower() not in cols:
            conn.execute(f'ALTER TABLE customs_product_master ADD COLUMN {name} {ddl}')
            cols.add(name.lower())
    hcols = {
        str(row['name']).lower()
        for row in conn.execute('PRAGMA table_info(customs_product_master_history)').fetchall()
    }
    hadditions = [
        ('product_code', 'TEXT'),
        ('changed_at', 'TEXT'),
        ('changed_by', 'TEXT'),
        ('change_source', 'TEXT'),
        ('old_json', 'TEXT'),
        ('new_json', 'TEXT'),
        ('created_at', 'TEXT'),
    ]
    for name, ddl in hadditions:
        if name.lower() not in hcols:
            conn.execute(f'ALTER TABLE customs_product_master_history ADD COLUMN {name} {ddl}')
            hcols.add(name.lower())
    conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_customs_product_code ON customs_product_master(product_code)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_customs_product_account ON customs_product_master(account)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_customs_product_source ON customs_product_master(source)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_customs_product_updated ON customs_product_master(updated_at)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_customs_product_history_code ON customs_product_master_history(product_code, changed_at)')


def _customs_product_row_to_dict(row):
    if not row:
        return {}
    data = {}
    for key in row.keys():
        data[key] = row[key]
    return data


def _clean_excel_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_excel_number(value):
    if value in (None, ''):
        return None
    try:
        return float(str(value).replace(',', '').strip())
    except Exception:
        return None


def _customs_json_for_compare(data):
    return {
        key: data.get(key)
        for key in CUSTOMS_PRODUCT_COLUMNS
        if key not in ('id', 'created_at', 'updated_at', 'updated_by')
    }


def _customs_missing_fields(data):
    return [
        field for field in CUSTOMS_PRODUCT_REQUIRED_FIELDS
        if str((data or {}).get(field) or '').strip() == ''
    ]


def _customs_product_source_path(explicit_path=''):
    explicit_path = str(explicit_path or '').strip().strip('"')
    if explicit_path and os.path.exists(explicit_path):
        return explicit_path
    try:
        from ai_identify import _find_excel
        path = _find_excel(_load_ai_config())
        if path and os.path.exists(path):
            return path
    except Exception:
        pass
    candidates = [
        os.path.join(_DATA_BASE, '报关产品基础资料（智谱）.xlsx'),
        os.path.join(_BASE, '报关产品基础资料（智谱）.xlsx'),
        os.path.join(os.path.dirname(_DATA_BASE), '报关产品基础资料（智谱）.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return ''


def _customs_load_existing_map(conn):
    if not conn:
        return {}
    try:
        exists = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='customs_product_master'"
        ).fetchone()
        if not exists:
            return {}
        rows = conn.execute('SELECT * FROM customs_product_master').fetchall()
        return {
            str(row['product_code'] or '').strip(): _customs_product_row_to_dict(row)
            for row in rows
            if str(row['product_code'] or '').strip()
        }
    except Exception:
        return {}


def _customs_parse_excel(excel_path):
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [_clean_excel_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
        header_counts = Counter([h for h in headers if h])
        duplicate_headers = [
            {'field': key, 'count': count}
            for key, count in header_counts.items()
            if count > 1
        ]
        rows = []
        blank_code_rows = []
        total_rows = 0
        code_counts = Counter()
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, '') for value in row):
                continue
            total_rows += 1
            product_code = _clean_excel_text(row[1] if len(row) > 1 else '')
            if not product_code:
                blank_code_rows.append(row_no)
                continue
            code_counts[product_code] += 1
            item = {
                'account': _clean_excel_text(row[20] if len(row) > 20 else ''),
                'product_code': product_code,
                'product_name': _clean_excel_text(row[2] if len(row) > 2 else ''),
                'category_name': _clean_excel_text(row[3] if len(row) > 3 else ''),
                'spec': '',
                'unit': '',
                'barcode': '',
                'jdy_product_id': '',
                'customs_name_cn': _clean_excel_text(row[5] if len(row) > 5 else ''),
                'customs_name_en': _clean_excel_text(row[22] if len(row) > 22 else ''),
                'hs_code': _clean_excel_text(row[4] if len(row) > 4 else ''),
                'customs_unit': _clean_excel_text(row[6] if len(row) > 6 else ''),
                'origin': _clean_excel_text(row[7] if len(row) > 7 else ''),
                'material': _clean_excel_text(row[8] if len(row) > 8 else ''),
                'usage': _clean_excel_text(row[9] if len(row) > 9 else ''),
                'tax_refund_rate': _clean_excel_number(row[10] if len(row) > 10 else None),
                'is_tax_refund': _clean_excel_text(row[17] if len(row) > 17 else ''),
                'gross_weight_per_pkg': _clean_excel_number(row[11] if len(row) > 11 else None),
                'net_weight_per_pkg': _clean_excel_number(row[12] if len(row) > 12 else None),
                'carton_length': _clean_excel_number(row[13] if len(row) > 13 else None),
                'carton_width': _clean_excel_number(row[14] if len(row) > 14 else None),
                'carton_height': _clean_excel_number(row[15] if len(row) > 15 else None),
                'carton_volume': _clean_excel_number(row[16] if len(row) > 16 else None),
                'pcs_per_pkg': _clean_excel_number(row[18] if len(row) > 18 else None),
                'supplier_number': _clean_excel_text(row[19] if len(row) > 19 else ''),
                'tax_code': _clean_excel_text(row[21] if len(row) > 21 else ''),
                'tax_category_short_name': _clean_excel_text(row[23] if len(row) > 23 else ''),
                'source': 'zhipu_excel',
                'source_excel_path': excel_path,
                'source_row_no': row_no,
            }
            rows.append(item)
        duplicate_codes = [
            {'product_code': code, 'count': count}
            for code, count in code_counts.items()
            if count > 1
        ]
        duplicate_set = {item['product_code'] for item in duplicate_codes}
        valid_rows = [item for item in rows if item['product_code'] not in duplicate_set]
        missing_counts = {}
        for item in rows:
            for field in CUSTOMS_PRODUCT_REQUIRED_FIELDS:
                if str(item.get(field) or '').strip() == '':
                    missing_counts[field] = missing_counts.get(field, 0) + 1
        return {
            'excel_path': excel_path,
            'sheet_name': ws.title,
            'headers': headers,
            'duplicate_headers': duplicate_headers,
            'total_rows': total_rows,
            'valid_code_count': len(rows),
            'blank_code_count': len(blank_code_rows),
            'blank_code_rows': blank_code_rows[:100],
            'duplicate_code_count': len(duplicate_codes),
            'duplicate_codes': duplicate_codes[:200],
            'rows': rows,
            'valid_rows': valid_rows,
            'conflict_rows': [item for item in rows if item['product_code'] in duplicate_set],
            'missing_field_counts': missing_counts,
        }
    finally:
        wb.close()


def _customs_import_preview(excel_path, conn=None):
    parsed = _customs_parse_excel(excel_path)
    existing = _customs_load_existing_map(conn)
    insert_count = update_count = unchanged_count = 0
    samples = []
    for item in parsed['valid_rows']:
        old = existing.get(item['product_code'])
        if not old:
            insert_count += 1
        else:
            old_cmp = _customs_json_for_compare(old)
            new_cmp = _customs_json_for_compare(item)
            if old_cmp == new_cmp:
                unchanged_count += 1
            else:
                update_count += 1
        if len(samples) < 10:
            samples.append(item)
    skipped = parsed['blank_code_count'] + len(parsed['conflict_rows']) + unchanged_count
    return {
        'success': True,
        'dry_run': True,
        'would_write': False,
        'called_jdy': False,
        'excel_path': parsed['excel_path'],
        'sheet_name': parsed['sheet_name'],
        'headers': parsed['headers'],
        'duplicate_headers': parsed['duplicate_headers'],
        'total_rows': parsed['total_rows'],
        'valid_code_count': parsed['valid_code_count'],
        'blank_code_count': parsed['blank_code_count'],
        'blank_code_rows': parsed['blank_code_rows'],
        'duplicate_code_count': parsed['duplicate_code_count'],
        'duplicate_codes': parsed['duplicate_codes'],
        'field_missing_counts': parsed['missing_field_counts'],
        'estimated_insert': insert_count,
        'estimated_update': update_count,
        'estimated_skip': skipped,
        'unchanged_count': unchanged_count,
        'conflict_count': len(parsed['conflict_rows']),
        'sample_rows': samples,
    }


def _customs_history_insert(conn, product_code, old_data, new_data, change_source, changed_by, changed_at):
    conn.execute('''
        INSERT INTO customs_product_master_history
        (product_code, changed_at, changed_by, change_source, old_json, new_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        product_code,
        changed_at,
        changed_by or '',
        change_source or '',
        json.dumps(old_data or {}, ensure_ascii=False, sort_keys=True),
        json.dumps(new_data or {}, ensure_ascii=False, sort_keys=True),
        changed_at,
    ))


def _customs_upsert_product(conn, item, updated_by='', change_source='zhipu_excel', now=None):
    now = now or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    product_code = str(item.get('product_code') or '').strip()
    if not product_code:
        return 'skipped'
    old_row = conn.execute(
        'SELECT * FROM customs_product_master WHERE product_code = ?',
        (product_code,),
    ).fetchone()
    values = {key: item.get(key) for key in CUSTOMS_PRODUCT_COLUMNS if key not in ('created_at', 'updated_at', 'updated_by')}
    values['product_code'] = product_code
    values['source'] = change_source or values.get('source') or 'local_edit'
    values['updated_at'] = now
    values['updated_by'] = updated_by or ''
    if old_row:
        old_data = _customs_product_row_to_dict(old_row)
        old_cmp = _customs_json_for_compare(old_data)
        new_cmp = _customs_json_for_compare(values)
        if old_cmp == new_cmp:
            return 'unchanged'
        set_fields = [key for key in CUSTOMS_PRODUCT_COLUMNS if key not in ('created_at',)]
        conn.execute(
            f'''UPDATE customs_product_master
                   SET {', '.join([field + '=?' for field in set_fields])}
                 WHERE product_code = ?''',
            [values.get(field) for field in set_fields] + [product_code],
        )
        new_row = conn.execute('SELECT * FROM customs_product_master WHERE product_code = ?', (product_code,)).fetchone()
        _customs_history_insert(
            conn,
            product_code,
            old_data,
            _customs_product_row_to_dict(new_row),
            change_source,
            updated_by,
            now,
        )
        return 'updated'
    insert_fields = [key for key in CUSTOMS_PRODUCT_COLUMNS if key != 'id']
    values['created_at'] = now
    conn.execute(
        f'''INSERT INTO customs_product_master
            ({', '.join(insert_fields)})
            VALUES ({', '.join(['?'] * len(insert_fields))})''',
        [values.get(field) for field in insert_fields],
    )
    return 'inserted'


_CUSTOMS_ZHIPU_CODE_CACHE = {
    'key': None,
    'codes': set(),
    'meta': {},
}


def _customs_zhipu_code_index():
    """Read the default Zhipu Excel once per file version for list status badges."""
    excel_path = _customs_product_source_path('')
    if not excel_path:
        return set(), {'available': False, 'excel_path': '', 'error': 'not_found'}
    try:
        stat = os.stat(excel_path)
        cache_key = (os.path.abspath(excel_path), stat.st_mtime, stat.st_size)
        if _CUSTOMS_ZHIPU_CODE_CACHE.get('key') == cache_key:
            return _CUSTOMS_ZHIPU_CODE_CACHE.get('codes') or set(), _CUSTOMS_ZHIPU_CODE_CACHE.get('meta') or {}
        parsed = _customs_parse_excel(excel_path)
        codes = {
            str(item.get('product_code') or '').strip()
            for item in parsed.get('valid_rows') or []
            if str(item.get('product_code') or '').strip()
        }
        meta = {
            'available': True,
            'excel_path': excel_path,
            'sheet_name': parsed.get('sheet_name') or '',
            'code_count': len(codes),
            'duplicate_code_count': parsed.get('duplicate_code_count') or 0,
            'loaded_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        _CUSTOMS_ZHIPU_CODE_CACHE.update({'key': cache_key, 'codes': codes, 'meta': meta})
        return codes, meta
    except Exception as e:
        meta = {'available': False, 'excel_path': excel_path, 'error': _short_sync_error(e)}
        _CUSTOMS_ZHIPU_CODE_CACHE.update({'key': None, 'codes': set(), 'meta': meta})
        return set(), meta


def _customs_master_missing_fields(jdy):
    missing = []
    for key in ('product_name', 'spec', 'unit_name'):
        if str((jdy or {}).get(key) or '').strip() == '':
            missing.append(key)
    return missing


def _customs_grid_value(row, key):
    c = row.get('customs') or {}
    j = row.get('jdy') or {}
    status = row.get('status') or {}
    computed = row.get('computed') or {}
    values = {
        'product_code': row.get('product_code') or '',
        'product_name': j.get('product_name') or c.get('product_name') or '',
        'jdy_product_name': j.get('product_name') or c.get('product_name') or '',
        'spec': j.get('spec') or c.get('spec') or '',
        'jdy_spec': j.get('spec') or c.get('spec') or '',
        'unit': j.get('unit_name') or c.get('unit') or '',
        'jdy_unit': j.get('unit_name') or c.get('unit') or '',
        'default_supplier': j.get('default_supplier_name') or j.get('default_supplier_number') or '',
        'jdy_supplier': j.get('default_supplier_name') or j.get('default_supplier_number') or '',
        'hs_code': c.get('hs_code') or '',
        'customs_name_cn': c.get('customs_name_cn') or '',
        'material': c.get('material') or '',
        'usage': c.get('usage') or '',
        'customs_name_en': c.get('customs_name_en') or '',
        'customs_unit': c.get('customs_unit') or '',
        'updated_at': c.get('updated_at') or j.get('updated_at') or '',
        'maintenance_status': status.get('label') or '',
        'status': status.get('label') or '',
        'zhipu_status': status.get('zhipu_label') or '',
        'source': c.get('source') or '',
        'production_license_preview': computed.get('production_license_preview') or '',
    }
    if key in values:
        return values[key]
    if key.startswith('jdy_'):
        return j.get(key[4:]) or ''
    return c.get(key) or ''


def _customs_product_status(customs, zhipu_has_match=False, zhipu_available=False, master_missing_fields=None):
    has_data = bool(customs and customs.get('product_code'))
    missing = _customs_missing_fields(customs or {}) if has_data else CUSTOMS_PRODUCT_REQUIRED_FIELDS[:]
    maintained = has_data and not missing
    source = str((customs or {}).get('source') or '')
    if maintained:
        label = '已有报关资料'
    else:
        label = '缺少报关资料'
    if not zhipu_available:
        zhipu_label = '未读取智谱'
        zhipu_status = 'unknown'
    elif zhipu_has_match and not maintained:
        zhipu_label = '智谱可补全'
        zhipu_status = 'match_fillable'
    elif zhipu_has_match:
        zhipu_label = '智谱有匹配'
        zhipu_status = 'match'
    else:
        zhipu_label = '智谱无匹配'
        zhipu_status = 'no_match'
    master_missing_fields = master_missing_fields or []
    return {
        'has_customs_data': has_data,
        'is_maintained': maintained,
        'missing_fields': missing,
        'source': source,
        'label': label,
        'maintenance_status': 'complete' if maintained else ('incomplete' if has_data else 'missing'),
        'zhipu_match': bool(zhipu_has_match),
        'zhipu_status': zhipu_status,
        'zhipu_label': zhipu_label,
        'master_missing_fields': master_missing_fields,
        'has_master_blank': bool(master_missing_fields),
        'is_new_or_missing': (not maintained) or bool(master_missing_fields) or (zhipu_available and not zhipu_has_match),
    }


def _customs_product_list_response():
    page = _bounded_query_int(request.args.get('page'), 1, 1, 1000000)
    page_size = _bounded_query_int(request.args.get('page_size') or request.args.get('limit'), 50, 1, 500)
    account = str(request.args.get('account') or 'all').strip() or 'all'
    q = str(request.args.get('q') or request.args.get('search') or '').strip().lower()
    status = str(request.args.get('status') or 'all').strip().lower()
    has_customs_data = str(request.args.get('has_customs_data') or '').strip().lower()
    sort_by = str(request.args.get('sort_by') or 'product_code').strip()
    sort_dir = str(request.args.get('sort_dir') or 'asc').strip().lower()
    blank_fields = {
        x.strip()
        for x in str(request.args.get('blank_fields') or '').split(',')
        if x.strip()
    }
    filter_args = {
        key[7:]: str(value or '').strip().lower()
        for key, value in request.args.items()
        if key.startswith('filter_') and str(value or '').strip()
    }
    zhipu_codes, zhipu_meta = _customs_zhipu_code_index()
    zhipu_available = bool(zhipu_meta.get('available'))
    conn = _sales_readonly_conn()
    base = {
        'success': True,
        'local_only': True,
        'called_jdy': False,
        'source': 'jdy_products LEFT JOIN customs_product_master',
        'total': 0,
        'page': page,
        'page_size': page_size,
        'items': [],
        'zhipu': zhipu_meta,
    }
    if not conn:
        return {**base, 'message': '本地 SQLite 尚不存在，请先导入或刷新本地商品主档。'}, 200
    try:
        has_customs = _cache_sqlite_table_count(conn, 'customs_product_master').get('exists')
        has_jdy = _cache_sqlite_table_count(conn, 'jdy_products').get('exists')
        if not has_jdy and not has_customs:
            return {**base, 'message': '本地商品主档和报关资料表均不存在。'}, 200
        customs_rows = conn.execute('SELECT * FROM customs_product_master').fetchall() if has_customs else []
        customs_map = {
            str(row['product_code'] or '').strip(): _customs_product_row_to_dict(row)
            for row in customs_rows
            if str(row['product_code'] or '').strip()
        }
        product_map = {}
        if has_jdy:
            clauses, params = ['COALESCE(product_number, "") != ""'], []
            if account != 'all':
                clauses.append('account = ?')
                params.append(account)
            rows = conn.execute(f'''
                SELECT id, account, product_id, product_number, product_name, spec, barcode,
                       category_name, unit_name, default_supplier_number, default_supplier_name,
                       image_url, status, data_json, updated_at
                FROM jdy_products
                WHERE {' AND '.join(clauses)}
                ORDER BY account ASC, product_number ASC, id ASC
            ''', params).fetchall()
            for row in rows:
                code = str(row['product_number'] or '').strip()
                if code and code not in product_map:
                    product_map[code] = _customs_product_row_to_dict(row)
        all_codes = sorted(product_map.keys()) if has_jdy else sorted(customs_map.keys())
        filtered = []
        counts = {
            'complete': 0,
            'missing_customs': 0,
            'zhipu_fillable': 0,
            'zhipu_no_match': 0,
            'master_blank': 0,
            'new_or_missing': 0,
        }
        for code in all_codes:
            jdy = product_map.get(code) or {}
            customs = customs_map.get(code) or {}
            if account != 'all':
                acct = str(jdy.get('account') or customs.get('account') or '')
                if acct and acct != account:
                    continue
            master_missing = _customs_master_missing_fields(jdy)
            st = _customs_product_status(
                {**customs, 'product_code': customs.get('product_code') or code} if customs else {},
                zhipu_has_match=code in zhipu_codes,
                zhipu_available=zhipu_available,
                master_missing_fields=master_missing,
            )
            cn = str(customs.get('customs_name_cn') or '')
            material = str(customs.get('material') or '')
            row = {
                'product_code': code,
                'jdy': {
                    'id': jdy.get('id') or '',
                    'account': jdy.get('account') or '',
                    'product_id': jdy.get('product_id') or '',
                    'product_name': jdy.get('product_name') or '',
                    'spec': jdy.get('spec') or '',
                    'unit_name': jdy.get('unit_name') or '',
                    'barcode': jdy.get('barcode') or '',
                    'category_name': jdy.get('category_name') or '',
                    'image_url': jdy.get('image_url') or '',
                    'default_supplier_number': jdy.get('default_supplier_number') or '',
                    'default_supplier_name': jdy.get('default_supplier_name') or '',
                    'status': jdy.get('status') or '',
                    'updated_at': jdy.get('updated_at') or '',
                },
                'customs': customs,
                'computed': {
                    'production_license_preview': f'{cn}*{material}' if cn or material else '',
                    'pro_license_preview': f'{cn}*{material}' if cn or material else '',
                },
                'status': st,
            }
            search_text = ' '.join([
                str(_customs_grid_value(row, key) or '')
                for key in (
                    'product_code', 'product_name', 'spec', 'jdy_barcode', 'material',
                    'customs_name_cn', 'customs_name_en', 'default_supplier', 'hs_code',
                )
            ]).lower()
            if q and q not in search_text:
                continue
            if has_customs_data in ('1', 'true', 'yes', 'y') and not st['has_customs_data']:
                continue
            if has_customs_data in ('0', 'false', 'no', 'n') and st['has_customs_data']:
                continue
            if status in ('maintained', 'complete', 'has_customs', '已有报关资料') and not st['is_maintained']:
                continue
            if status in ('missing', 'incomplete', 'missing_customs', '缺少报关资料') and st['is_maintained']:
                continue
            if status in ('zhipu_fillable', 'zhipu_match_fillable', '智谱可补全') and st.get('zhipu_status') != 'match_fillable':
                continue
            if status in ('zhipu_no_match', 'new_product', '智谱无匹配') and st.get('zhipu_status') != 'no_match':
                continue
            if status in ('master_blank', '主档字段为空') and not st.get('has_master_blank'):
                continue
            if status in ('new_missing', 'new_or_missing', '新增/缺失') and not st.get('is_new_or_missing'):
                continue
            if status in ('zhipu', 'zhipu_excel') and st.get('source') != 'zhipu_excel':
                continue
            if status in ('local', 'local_edit') and st.get('source') != 'local_edit':
                continue
            if any(str(_customs_grid_value(row, field) or '').strip() for field in blank_fields):
                continue
            filter_miss = False
            for key, needle in filter_args.items():
                if needle not in str(_customs_grid_value(row, key) or '').lower():
                    filter_miss = True
                    break
            if filter_miss:
                continue
            if st['is_maintained']:
                counts['complete'] += 1
            else:
                counts['missing_customs'] += 1
            if st.get('zhipu_status') == 'match_fillable':
                counts['zhipu_fillable'] += 1
            if st.get('zhipu_status') == 'no_match':
                counts['zhipu_no_match'] += 1
            if st.get('has_master_blank'):
                counts['master_blank'] += 1
            if st.get('is_new_or_missing'):
                counts['new_or_missing'] += 1
            filtered.append(row)
        sortable = {
            'product_code', 'product_name', 'jdy_product_name', 'updated_at',
            'status', 'maintenance_status',
        }
        if sort_by not in sortable:
            sort_by = 'product_code'
        reverse = sort_dir == 'desc'
        filtered.sort(
            key=lambda item: str(_customs_grid_value(item, sort_by) or '').lower(),
            reverse=reverse,
        )
        total = len(filtered)
        start = (page - 1) * page_size
        return {
            **base,
            'total': total,
            'page': page,
            'page_size': page_size,
            'items': filtered[start:start + page_size],
            'summary': {
                'customs_table_exists': bool(has_customs),
                'jdy_products_exists': bool(has_jdy),
                'customs_count': len(customs_map),
                'jdy_count': len(product_map),
                'filtered_counts': counts,
                'sort_by': sort_by,
                'sort_dir': 'desc' if reverse else 'asc',
            },
        }, 200
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _customs_product_status_response():
    account = str(request.args.get('account') or 'all').strip() or 'all'
    conn = _sales_readonly_conn()
    base = {
        'success': True,
        'local_only': True,
        'called_jdy': False,
        'source': 'customs_product_master+jdy_products',
        'account': account,
        'customs_table_exists': False,
        'jdy_products_exists': False,
        'customs_count': 0,
        'jdy_count': 0,
        'linked_count': 0,
        'customs_only_count': 0,
        'jdy_only_count': 0,
        'maintained_count': 0,
        'missing_count': 0,
        'incomplete_count': 0,
        'zhipu_import_count': 0,
        'local_edit_count': 0,
        'latest_import_at': '',
        'latest_local_edit_at': '',
        'latest_updated_at': '',
        'required_fields': CUSTOMS_PRODUCT_REQUIRED_FIELDS,
    }
    if not conn:
        return {**base, 'message': '本地 SQLite 尚不存在。'}, 200
    try:
        has_customs = _cache_sqlite_table_count(conn, 'customs_product_master').get('exists')
        has_jdy = _cache_sqlite_table_count(conn, 'jdy_products').get('exists')
        customs_map = {}
        product_map = {}
        if has_customs:
            rows = conn.execute('SELECT * FROM customs_product_master').fetchall()
            for row in rows:
                data = _customs_product_row_to_dict(row)
                code = str(data.get('product_code') or '').strip()
                if code:
                    customs_map[code] = data
        if has_jdy:
            clauses, params = ['COALESCE(product_number, "") != ""'], []
            if account != 'all':
                clauses.append('account = ?')
                params.append(account)
            rows = conn.execute(f'''
                SELECT account, product_number
                FROM jdy_products
                WHERE {' AND '.join(clauses)}
            ''', params).fetchall()
            for row in rows:
                code = str(row['product_number'] or '').strip()
                if code and code not in product_map:
                    product_map[code] = _customs_product_row_to_dict(row)
        all_codes = set(product_map.keys()) | set(customs_map.keys())
        maintained = missing = incomplete = zhipu = local = linked = customs_only = jdy_only = 0
        latest_import_at = ''
        latest_local_edit_at = ''
        latest_updated_at = ''
        for code in all_codes:
            customs = customs_map.get(code) or {}
            jdy = product_map.get(code) or {}
            if account != 'all':
                acct = str(jdy.get('account') or customs.get('account') or '')
                if acct and acct != account:
                    continue
            if customs and jdy:
                linked += 1
            elif customs:
                customs_only += 1
            else:
                jdy_only += 1
            st = _customs_product_status(customs)
            if not st['has_customs_data']:
                missing += 1
            elif st['is_maintained']:
                maintained += 1
            else:
                incomplete += 1
            updated_at = str(customs.get('updated_at') or '')
            if updated_at and updated_at > latest_updated_at:
                latest_updated_at = updated_at
            if st.get('source') == 'zhipu_excel':
                zhipu += 1
                if updated_at and updated_at > latest_import_at:
                    latest_import_at = updated_at
            elif st.get('source') == 'local_edit':
                local += 1
                if updated_at and updated_at > latest_local_edit_at:
                    latest_local_edit_at = updated_at
        return {
            **base,
            'customs_table_exists': bool(has_customs),
            'jdy_products_exists': bool(has_jdy),
            'customs_count': len(customs_map),
            'jdy_count': len(product_map),
            'linked_count': linked,
            'customs_only_count': customs_only,
            'jdy_only_count': jdy_only,
            'maintained_count': maintained,
            'missing_count': missing,
            'incomplete_count': incomplete,
            'zhipu_import_count': zhipu,
            'local_edit_count': local,
            'latest_import_at': latest_import_at,
            'latest_local_edit_at': latest_local_edit_at,
            'latest_updated_at': latest_updated_at,
        }, 200
    finally:
        try:
            conn.close()
        except Exception:
            pass


JDY_SUPPLIER_EXTRA_COLUMNS = {
    'status_code': 'TEXT',
    'status': 'TEXT',
    'status_name': 'TEXT',
    'contact': 'TEXT',
    'phone': 'TEXT',
    'last_seen_at': 'TEXT',
    'last_seen_enabled_at': 'TEXT',
    'disabled_at': 'TEXT',
    'manual_category_text': 'TEXT',
    'manual_tags': 'TEXT',
    'manual_note': 'TEXT',
    'accessory_override': 'TEXT',
}


def _jdy_supplier_columns(conn):
    try:
        return [str(row['name'] if isinstance(row, sqlite3.Row) else row[1]) for row in conn.execute('PRAGMA table_info(jdy_suppliers)').fetchall()]
    except Exception:
        return []


def _missing_jdy_supplier_columns(columns):
    existing = {str(c).lower() for c in (columns or [])}
    return [name for name in JDY_SUPPLIER_EXTRA_COLUMNS if name.lower() not in existing]


def _ensure_jdy_suppliers_schema(conn):
    columns = _jdy_supplier_columns(conn)
    if not columns:
        return []
    added = []
    for name in _missing_jdy_supplier_columns(columns):
        conn.execute(f'ALTER TABLE jdy_suppliers ADD COLUMN {name} {JDY_SUPPLIER_EXTRA_COLUMNS[name]}')
        added.append(name)
        columns.append(name)
    return added


def _backup_sales_cache_db(reason='supplier_refresh'):
    if not os.path.exists(_SALES_CACHE_DB):
        raise FileNotFoundError(_SALES_CACHE_DB)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_reason = ''.join(ch if ch.isalnum() or ch in ('_', '-') else '_' for ch in str(reason or 'backup'))
    backup_path = os.path.join(_SALES_CACHE_DIR, f'sales_cache.sqlite3.bak_{stamp}_{safe_reason}')
    shutil.copy2(_SALES_CACHE_DB, backup_path)
    if not os.path.exists(backup_path) or os.path.getsize(backup_path) != os.path.getsize(_SALES_CACHE_DB):
        raise RuntimeError('sales_cache backup failed or size mismatch')
    return backup_path


WEBHOOK_RATE_NORMAL_PER_MIN = 300
WEBHOOK_RATE_BOOST_PER_MIN = 450
WEBHOOK_RATE_HARD_MAX_PER_MIN = 450
WEBHOOK_BOOST_PENDING_THRESHOLD = 3000
WEBHOOK_BOOST_ETA_MINUTES = 20
WEBHOOK_BOOST_COOLDOWN_PENDING = 1000
WEBHOOK_BACKOFF_PER_MIN = 120
WEBHOOK_ERROR_BACKOFF_SECONDS = 300


def _webhook_backlog_estimate(pending_total):
    pending_total = max(0, int(pending_total or 0))
    normal = min(WEBHOOK_RATE_NORMAL_PER_MIN, WEBHOOK_RATE_HARD_MAX_PER_MIN)
    boost = min(WEBHOOK_RATE_BOOST_PER_MIN, WEBHOOK_RATE_HARD_MAX_PER_MIN)
    return {
        'pending_total': pending_total,
        'normal_rate_per_min': normal,
        'boost_rate_per_min': boost,
        'eta_minutes_normal': round(pending_total / normal, 2) if normal else 0,
        'eta_minutes_boost': round(pending_total / boost, 2) if boost else 0,
    }


def _webhook_backoff_active():
    until = float(_webhook_state.get('backoff_until') or 0)
    return until > time.time()


def _webhook_rate_policy(pending_total=None):
    pending_total = _webhook_pending_count() if pending_total is None else int(pending_total or 0)
    estimate = _webhook_backlog_estimate(pending_total)
    processing_mode = str(_webhook_state.get('processing_mode') or 'manual').strip().lower()
    if _webhook_state.get('paused'):
        return {
            'current_rate_per_min': 0,
            'mode': 'paused' if processing_mode == 'paused' else 'manual',
            'reason': _webhook_state.get('pause_reason') or 'manual processing required',
            'processing_mode': processing_mode,
            'hard_max_per_min': WEBHOOK_RATE_HARD_MAX_PER_MIN,
        }
    if _webhook_backoff_active():
        return {
            'current_rate_per_min': min(WEBHOOK_BACKOFF_PER_MIN, WEBHOOK_RATE_HARD_MAX_PER_MIN),
            'mode': 'backoff',
            'reason': _webhook_state.get('backoff_reason') or 'recent API/network errors',
            'processing_mode': processing_mode,
            'hard_max_per_min': WEBHOOK_RATE_HARD_MAX_PER_MIN,
        }
    if (
        pending_total >= WEBHOOK_BOOST_PENDING_THRESHOLD
        or estimate['eta_minutes_normal'] > WEBHOOK_BOOST_ETA_MINUTES
        or (_webhook_state.get('mode') == 'boost' and pending_total > WEBHOOK_BOOST_COOLDOWN_PENDING)
    ):
        return {
            'current_rate_per_min': min(WEBHOOK_RATE_BOOST_PER_MIN, WEBHOOK_RATE_HARD_MAX_PER_MIN),
            'mode': 'boost',
            'reason': 'pending backlog is high',
            'processing_mode': processing_mode,
            'hard_max_per_min': WEBHOOK_RATE_HARD_MAX_PER_MIN,
        }
    return {
        'current_rate_per_min': min(WEBHOOK_RATE_NORMAL_PER_MIN, WEBHOOK_RATE_HARD_MAX_PER_MIN),
        'mode': 'normal',
        'reason': 'default controlled processing',
        'processing_mode': processing_mode,
        'hard_max_per_min': WEBHOOK_RATE_HARD_MAX_PER_MIN,
    }


def _webhook_now_text():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _webhook_worker_alive():
    thread = globals().get('_webhook_worker_thread')
    return bool(thread and thread.is_alive())


def _webhook_auto_should_run(pending_total=None):
    try:
        pending_total = _webhook_pending_count() if pending_total is None else int(pending_total or 0)
    except Exception:
        pending_total = 0
    mode = str(_webhook_state.get('processing_mode') or 'manual').strip().lower()
    return mode == 'auto' and not bool(_webhook_state.get('paused')) and pending_total > 0


def _webhook_worker_status_fields(pending_total=None):
    worker_alive = _webhook_worker_alive()
    auto_should = _webhook_auto_should_run(pending_total)
    return {
        'worker_alive': worker_alive,
        'worker_thread_name': str(_webhook_state.get('worker_thread_name') or ''),
        'worker_thread_id': str(_webhook_state.get('worker_thread_id') or ''),
        'worker_last_heartbeat_at': str(_webhook_state.get('worker_last_heartbeat_at') or ''),
        'worker_last_loop_at': str(_webhook_state.get('worker_last_loop_at') or ''),
        'worker_last_exception': str(_webhook_state.get('worker_last_exception') or ''),
        'auto_should_be_running': auto_should,
        'auto_recovery_needed': bool(auto_should and not worker_alive),
        'auto_recovery_action': 'restart_worker' if auto_should and not worker_alive else '',
        'auto_recovery_last_at': str(_webhook_state.get('auto_recovery_last_at') or ''),
        'auto_recovery_reason': str(_webhook_state.get('auto_recovery_reason') or ''),
    }


def _webhook_maybe_restart_auto_worker(pending_total=None, reason='status'):
    if not _webhook_auto_should_run(pending_total):
        return False
    if _webhook_worker_alive():
        return False
    _webhook_state.update({
        'auto_recovery_last_at': _webhook_now_text(),
        'auto_recovery_reason': str(reason or 'auto'),
        'worker_last_exception': '',
        'message': 'auto recovery restarting webhook worker',
    })
    _start_webhook_worker()
    return True


def _webhook_status_snapshot(limit=20):
    try:
        limit = int(limit or 20)
    except Exception:
        limit = 20
    limit = max(1, min(limit, 100))
    counts = {'pending': 0, 'retry_pending': 0, 'processing': 0, 'done': 0, 'failed': 0, 'ignored': 0}
    recent = []
    resource_map = {}
    duplicate_summary = []
    retry_duplicate_summary = []
    stale_processing = 0
    conn = _sales_readonly_conn()
    if conn is None:
        policy = _webhook_rate_policy(0)
        processing_mode = str(_webhook_state.get('processing_mode') or 'manual').strip().lower()
        estimate = _webhook_backlog_estimate(0)
        return {
            'counts': counts,
            'recent': recent,
            'resource_summary': [],
            'by_resource': [],
            'duplicate_summary': [],
            'retry_duplicate_summary': [],
            'stale_processing': 0,
            'backlog_estimate': estimate,
            'backlog_estimate_normal': estimate,
            'backlog_estimate_retry': estimate,
            'current_rate_policy': policy,
            'processing_mode': processing_mode,
            'auto_enabled': processing_mode == 'auto' and not _webhook_state.get('paused'),
            'manual_processing_required': processing_mode != 'auto',
            'pending_count': 0,
            'retry_pending_count': 0,
            'current_rate_limit': policy.get('current_rate_per_min') or 0,
            'estimated_minutes_300': estimate.get('eta_minutes_normal') or 0,
            'estimated_minutes_450': estimate.get('eta_minutes_boost') or 0,
            **_webhook_worker_status_fields(0),
            'notes': ['本地缓存库不存在，尚无 Webhook 事件记录'],
        }
    try:
        for row in conn.execute('SELECT status, COUNT(*) AS c FROM webhook_events GROUP BY status'):
            status = str(row['status'] or '')
            counts[status] = int(row['c'] or 0)

        for row in conn.execute('''
            SELECT resource, status, COUNT(*) AS c
            FROM webhook_events
            GROUP BY resource, status
            ORDER BY resource, status
        '''):
            resource = str(row['resource'] or 'unknown')
            status = str(row['status'] or '')
            item = resource_map.setdefault(resource, {
                'resource': resource,
                'total': 0,
                'pending': 0,
                'retry_pending': 0,
                'processing': 0,
                'done': 0,
                'failed': 0,
                'ignored': 0,
            })
            count = int(row['c'] or 0)
            item['total'] += count
            item[status] = count

        rows = conn.execute('''
            SELECT id, account, biz_type, resource, bill_no, action, status,
                   attempts, error, created_at, processed_at, payload_json
            FROM webhook_events
            ORDER BY id DESC
            LIMIT ?
        ''', (limit,)).fetchall()
        for row in rows:
            payload = str(row['payload_json'] or '')
            recent.append({
                'id': row['id'],
                'account': row['account'] or '',
                'biz_type': row['biz_type'] or '',
                'resource': row['resource'] or '',
                'bill_no': row['bill_no'] or '',
                'action': row['action'] or '',
                'status': row['status'] or '',
                'attempts': int(row['attempts'] or 0),
                'error': row['error'] or '',
                'created_at': row['created_at'] or '',
                'processed_at': row['processed_at'] or '',
                'payload_preview': payload[:200],
            })
        for row in conn.execute('''
            SELECT resource, bill_no, COUNT(*) AS c, GROUP_CONCAT(id) AS ids
            FROM webhook_events
            WHERE status = 'pending'
              AND COALESCE(resource, '') <> ''
              AND COALESCE(bill_no, '') <> ''
            GROUP BY resource, bill_no
            HAVING c > 1
            ORDER BY c DESC, resource, bill_no
            LIMIT 20
        '''):
            duplicate_summary.append({
                'resource': row['resource'] or 'unknown',
                'bill_no': row['bill_no'] or '',
                'count': int(row['c'] or 0),
                'pending_count': int(row['c'] or 0),
                'ids': [
                    int(x) for x in str(row['ids'] or '').split(',')
                    if str(x or '').strip().isdigit()
                ],
            })
        for row in conn.execute('''
            SELECT resource, bill_no, COUNT(*) AS c, GROUP_CONCAT(id) AS ids
            FROM webhook_events
            WHERE status = 'retry_pending'
              AND COALESCE(resource, '') <> ''
              AND COALESCE(bill_no, '') <> ''
            GROUP BY resource, bill_no
            HAVING c > 1
            ORDER BY c DESC, resource, bill_no
            LIMIT 20
        '''):
            retry_duplicate_summary.append({
                'resource': row['resource'] or 'unknown',
                'bill_no': row['bill_no'] or '',
                'count': int(row['c'] or 0),
                'retry_pending_count': int(row['c'] or 0),
                'ids': [
                    int(x) for x in str(row['ids'] or '').split(',')
                    if str(x or '').strip().isdigit()
                ],
            })
        stale_cutoff = datetime.now() - timedelta(minutes=30)
        stale_rows = conn.execute('''
            SELECT created_at, processed_at
            FROM webhook_events
            WHERE status = 'processing'
        ''').fetchall()
        for row in stale_rows:
            raw_ts = str(row['processed_at'] or row['created_at'] or '').strip()
            if not raw_ts:
                continue
            try:
                parsed_ts = datetime.fromisoformat(raw_ts.replace('Z', '').replace('T', ' '))
            except Exception:
                continue
            if parsed_ts <= stale_cutoff:
                stale_processing += 1
    finally:
        conn.close()

    pending_total = counts.get('pending', 0)
    retry_pending_total = counts.get('retry_pending', 0)
    policy = _webhook_rate_policy(pending_total)
    processing_mode = str(_webhook_state.get('processing_mode') or 'manual').strip().lower()
    estimate = _webhook_backlog_estimate(pending_total)
    retry_estimate = _webhook_backlog_estimate(retry_pending_total)
    runtime_settings = _load_webhook_runtime_settings() or {}
    notes = [
        'inventory/product/supplier/purchase_inbound 当前只轻记录',
        '请在 JDY 后台确认消息订阅地址是否配置为 http://gongdashuai.top:5008/jdy-webhook',
    ]
    if not counts.get('done'):
        notes.insert(0, '当前没有真实业务推送成功记录')
    if counts.get('failed'):
        notes.append('存在失败事件，请查看最近事件里的 error 字段')
    return {
        'counts': counts,
        'recent': recent,
        'resource_summary': sorted(resource_map.values(), key=lambda x: x['resource']),
        'by_resource': sorted(resource_map.values(), key=lambda x: x['resource']),
        'duplicate_summary': duplicate_summary,
        'retry_duplicate_summary': retry_duplicate_summary,
        'stale_processing': stale_processing,
        'backlog_estimate': estimate,
        'backlog_estimate_normal': estimate,
        'backlog_estimate_retry': retry_estimate,
        'current_rate_policy': policy,
        'processing_mode': processing_mode,
        'auto_enabled': processing_mode == 'auto' and not _webhook_state.get('paused'),
        'manual_processing_required': processing_mode != 'auto',
        'pending_count': pending_total,
        'retry_pending_count': retry_pending_total,
        'current_rate_limit': policy.get('current_rate_per_min') or 0,
        'estimated_minutes_300': estimate.get('eta_minutes_normal') or 0,
        'estimated_minutes_450': estimate.get('eta_minutes_boost') or 0,
        'runtime_settings': runtime_settings,
        **_webhook_worker_status_fields(pending_total),
        'notes': notes,
    }


def _load_local_json_cached(path):
    try:
        if not os.path.exists(path):
            return {}
        st = os.stat(path)
        key = os.path.abspath(path)
        cached = _LOCAL_JSON_CACHE.get(key)
        sig = (st.st_mtime, st.st_size)
        if cached and cached.get('sig') == sig:
            return cached.get('data') or {}
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        _LOCAL_JSON_CACHE[key] = {'sig': sig, 'data': data}
        return data
    except Exception as e:
        print(f'[LOCAL CACHE] JSON 读取失败 {path}: {e}')
        return {}


def _local_product_cache_path(account):
    account = str(account or '')
    if '箱包' in account:
        return os.path.join(_DATA_BASE, '_cache', 'product_account2.json')
    return os.path.join(_DATA_BASE, '_cache', 'product_account1.json')


def _local_supplier_cache_path(account):
    account = str(account or '')
    if '箱包' in account:
        return os.path.join(_DATA_BASE, '_cache', 'supplier_account2.json')
    return os.path.join(_DATA_BASE, '_cache', 'supplier_account1.json')


def _local_product_index(account):
    return _load_local_product_index(account)


def _local_supplier_index(account):
    conn = _sales_readonly_conn()
    try:
        if conn and _cache_sqlite_table_count(conn, 'jdy_suppliers').get('exists'):
            rows = conn.execute('SELECT * FROM jdy_suppliers WHERE account = ?', (account or '',)).fetchall()
            if rows:
                by_id = {}
                by_number = {}
                for row in rows:
                    item = _supplier_row_to_dict(row)
                    sid = str(item.get('id') or item.get('supplierId') or '').strip()
                    number = str(item.get('number') or item.get('supplierNumber') or '').strip()
                    normalized = {
                        **item,
                        'supplierNumber': number,
                        'supplierName': str(item.get('name') or item.get('supplierName') or '').strip(),
                        'source': 'jdy_suppliers',
                    }
                    if sid:
                        by_id[sid] = normalized
                    if number:
                        by_number[number] = normalized
                return {'by_id': by_id, 'by_number': by_number}
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass
    data = _load_local_json_cached(_local_supplier_cache_path(account))
    items = data.get('items') if isinstance(data, dict) else data
    by_id = {}
    by_number = {}
    for item in items or []:
        if not isinstance(item, dict):
            continue
        sid = str(item.get('id') or '').strip()
        number = str(item.get('number') or '').strip()
        row = {
            'supplierNumber': number,
            'supplierName': str(item.get('name') or '').strip(),
        }
        if sid:
            by_id[sid] = row
        if number:
            by_number[number] = row
    return {'by_id': by_id, 'by_number': by_number}


def _local_json_cache_status(path):
    info = {'path': path, 'exists': os.path.exists(path), 'count': 0, 'sync_time': '', 'account': '', 'error': ''}
    if not info['exists']:
        return info
    try:
        data = _load_local_json_cached(path)
        items = data.get('items') if isinstance(data, dict) else data
        info['count'] = len(items or [])
        if isinstance(data, dict):
            info['sync_time'] = data.get('sync_time') or ''
            info['account'] = data.get('account') or ''
    except Exception as e:
        info['error'] = str(e)
    return info


def _product_first_image(product):
    imgs = product.get('multiImg') if isinstance(product, dict) else []
    if isinstance(imgs, list):
        for img in imgs:
            if isinstance(img, dict):
                url = str(_first_value(img, ['url', 'fileUrl', 'downloadUrl', 'src', 'path'], '') or '').strip()
                if url:
                    return url
            elif img:
                return str(img)
    for sku in (product.get('invSku') or []) if isinstance(product, dict) else []:
        if isinstance(sku, dict):
            url = str(_first_value(sku, ['skuImg', 'imageUrl', 'imgUrl', 'picUrl'], '') or '').strip()
            if url:
                return url
    return str(_first_value(product or {}, ['imageUrl', 'imgUrl', 'picture', 'pictureUrl', 'pic', 'picUrl'], '') or '').strip()


def _normalize_product_cache_item(product, account='', source='jdy_products'):
    product = product or {}
    code = str(_first_value(product, ['productNumber', 'number', 'code'], '') or '').strip()
    if not code:
        return None
    name = str(_first_value(product, ['productName', 'name'], '') or '').strip()
    category_name = str(_first_value(product, ['categoryName', 'category'], '') or '').strip()
    unit_name = str(_first_value(product, ['unitName', 'unit', 'baseUnitName'], '') or '').strip()
    supplier_number = str(_first_value(product, ['defaultSupplierNumber', 'supplierNumber', 'supplierNo', 'vendorNumber'], '') or '').strip()
    supplier_name = str(_first_value(product, ['defaultSupplierName', 'supplierName', 'vendorName'], '') or '').strip()
    image_url = _product_first_image(product)
    return {
        'account': account or str(product.get('account') or ''),
        'code': code,
        'number': code,
        'productNumber': code,
        'name': name,
        'productName': name,
        'spec': str(_first_value(product, ['spec', 'model'], '') or '').strip(),
        'barcode': str(_first_value(product, ['barcode', 'barCode'], '') or '').strip(),
        'category_id': str(_first_value(product, ['categoryId'], '') or '').strip(),
        'category_name': category_name,
        'categoryName': category_name,
        'unit_id': str(_first_value(product, ['unitId'], '') or '').strip(),
        'unit_name': unit_name,
        'unitName': unit_name,
        'default_supplier_id': str(_first_value(product, ['defaultSupplierId', 'supplierId'], '') or '').strip(),
        'default_supplier_number': supplier_number,
        'default_supplier_name': supplier_name,
        'supplierNumber': supplier_number,
        'supplierName': supplier_name,
        'image_url': image_url,
        'imageUrl': image_url,
        'status': str(_first_value(product, ['status', 'statusName', 'enable', 'enabled'], '') or '').strip(),
        'source': source,
        '_raw': product,
    }


def _cache_upsert_jdy_product(conn, account, product, now=None):
    item = _normalize_product_cache_item(product, account=account, source='jdy_products')
    if not item:
        return None, 'skipped'
    now = now or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    existing = conn.execute(
        'SELECT id FROM jdy_products WHERE account = ? AND product_number = ?',
        (account or '', item['code']),
    ).fetchone()
    values = (
        account or '',
        str(_first_value(product or {}, ['id', 'productId'], '') or ''),
        item['code'],
        item['name'],
        item['spec'],
        item['barcode'],
        item['category_id'],
        item['category_name'],
        item['unit_id'],
        item['unit_name'],
        item['default_supplier_id'],
        item['default_supplier_number'],
        item['default_supplier_name'],
        item['image_url'],
        item['status'],
        json.dumps(product or {}, ensure_ascii=False),
        now,
        now,
    )
    if existing:
        conn.execute('''
            UPDATE jdy_products
               SET product_id=?, product_name=?, spec=?, barcode=?, category_id=?, category_name=?,
                   unit_id=?, unit_name=?, default_supplier_id=?, default_supplier_number=?,
                   default_supplier_name=?, image_url=?, status=?, data_json=?, updated_at=?, last_seen_at=?
             WHERE account=? AND product_number=?
        ''', (
            values[1], values[3], values[4], values[5], values[6], values[7],
            values[8], values[9], values[10], values[11], values[12], values[13],
            values[14], values[15], now, now, account or '', item['code'],
        ))
        return item, 'updated'
    conn.execute('''
        INSERT INTO jdy_products
        (account, product_id, product_number, product_name, spec, barcode, category_id, category_name,
         unit_id, unit_name, default_supplier_id, default_supplier_number, default_supplier_name,
         image_url, status, data_json, created_at, updated_at, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (*values, now))
    return item, 'inserted'


def _load_local_product_index(account=None):
    account_filter = str(account or '').strip()
    use_account_filter = bool(account_filter and account_filter != 'all')
    index = {}

    def add_item(item):
        if not item:
            return
        code = str(item.get('code') or item.get('productNumber') or '').strip()
        if code and code not in index:
            index[code] = item

    conn = _sales_readonly_conn()
    try:
        if conn and _cache_sqlite_table_count(conn, 'jdy_products').get('exists'):
            clauses, params = [], []
            if use_account_filter:
                clauses.append('account = ?')
                params.append(account_filter)
            where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
            rows = conn.execute(f'''
                SELECT account, data_json FROM jdy_products
                {where}
                ORDER BY updated_at DESC, id DESC
            ''', params).fetchall()
            for row in rows:
                try:
                    raw = json.loads(row['data_json'] or '{}')
                except Exception:
                    raw = {}
                add_item(_normalize_product_cache_item(raw, row['account'] or account_filter, 'jdy_products'))
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

    json_accounts = [account_filter] if use_account_filter else ['account1', '箱包']
    for acct in json_accounts:
        data = _load_local_json_cached(_local_product_cache_path(acct))
        items = data.get('items') if isinstance(data, dict) else data
        for product in items or []:
            if isinstance(product, dict):
                add_item(_normalize_product_cache_item(product, acct, 'product_json'))

    conn = _sales_readonly_conn()
    try:
        if conn and _cache_sqlite_table_count(conn, 'sales_details').get('exists'):
            clauses, params = [], []
            if use_account_filter:
                clauses.append('account = ?')
                params.append(account_filter)
            where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
            rows = conn.execute(f'SELECT account, data_json FROM sales_details {where}', params).fetchall()
            for row in rows:
                try:
                    order = json.loads(row['data_json'] or '{}')
                except Exception:
                    continue
                acct = order.get('account') or row['account'] or account_filter
                for entry in (order.get('entries') or []):
                    if isinstance(entry, dict):
                        add_item(_normalize_product_cache_item(entry, acct, 'sales_details'))
        if conn and _cache_sqlite_table_count(conn, 'sales_product_quantities').get('exists'):
            clauses, params = [], []
            if use_account_filter:
                clauses.append('account = ?')
                params.append(account_filter)
            where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
            rows = conn.execute(f'SELECT account, code FROM sales_product_quantities {where}', params).fetchall()
            for row in rows:
                add_item(_normalize_product_cache_item({'productNumber': row['code'], 'account': row['account']}, row['account'] or account_filter, 'sales_product_quantities'))
        if conn and _cache_sqlite_table_count(conn, 'accessory_products').get('exists'):
            clauses, params = [], []
            if use_account_filter:
                clauses.append('account = ?')
                params.append(account_filter)
            where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
            rows = conn.execute(f'SELECT account, data_json FROM accessory_products {where}', params).fetchall()
            for row in rows:
                try:
                    raw = json.loads(row['data_json'] or '{}')
                except Exception:
                    raw = {}
                add_item(_normalize_product_cache_item(raw, row['account'] or account_filter, 'accessory_products'))
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass
    return index


def _cache_sqlite_table_count(conn, table):
    result = {'exists': False, 'count': 0}
    if not conn:
        return result
    try:
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table,),
        ).fetchone()
        if not row:
            return result
        result['exists'] = True
        count_row = conn.execute(f'SELECT COUNT(*) AS c FROM {table}').fetchone()
        result['count'] = int(count_row['c'] if isinstance(count_row, sqlite3.Row) else count_row[0])
    except Exception as e:
        result['error'] = str(e)
    return result


def _bounded_query_int(value, default, min_value, max_value):
    try:
        value = int(value)
    except Exception:
        value = default
    return max(min_value, min(value, max_value))


def _local_products_response():
    page = _bounded_query_int(request.args.get('page'), 1, 1, 1000000)
    page_size = _bounded_query_int(
        request.args.get('page_size') or request.args.get('limit'),
        50,
        1,
        200,
    )
    offset = (page - 1) * page_size
    account = str(request.args.get('account') or 'all').strip() or 'all'
    q = str(request.args.get('q') or request.args.get('search') or '').strip().lower()
    has_image = str(request.args.get('has_image') or '').strip().lower()
    supplier = str(request.args.get('supplier') or '').strip().lower()
    category = str(request.args.get('category') or '').strip().lower()

    base = {
        'success': True,
        'local_only': True,
        'live_lookup': False,
        'called_jdy': False,
        'would_call_jdy': False,
        'source': 'jdy_products',
        'total': 0,
        'page': page,
        'page_size': page_size,
        'items': [],
    }
    conn = _sales_readonly_conn()
    if not conn:
        return {
            **base,
            'success': False,
            'error': '本地销售缓存数据库不存在',
            'message': '本地商品主档为空，请先在设置页补齐商品缓存。',
        }, 200
    try:
        products_state = _cache_sqlite_table_count(conn, 'jdy_products')
        if not products_state.get('exists'):
            return {
                **base,
                'success': False,
                'error': '本地表 jdy_products 不存在',
                'message': '本地商品主档为空，请先在设置页补齐商品缓存。',
            }, 200
        qty_state = _cache_sqlite_table_count(conn, 'sales_product_quantities')
        joins = ''
        select_qty = '''
            0 AS stock_new, 0 AS stock_transit, 0 AS stock_local,
            0 AS stock_factory, 0 AS factory_qty
        '''
        if qty_state.get('exists'):
            joins = '''
                LEFT JOIN sales_product_quantities q
                  ON q.account = p.account AND q.code = p.product_number
            '''
            select_qty = '''
                COALESCE(q.stock_new, 0) AS stock_new,
                COALESCE(q.stock_transit, 0) AS stock_transit,
                COALESCE(q.stock_local, 0) AS stock_local,
                COALESCE(q.stock_factory, 0) AS stock_factory,
                COALESCE(q.factory_qty, 0) AS factory_qty
            '''
        clauses = ['COALESCE(p.product_number, "") != ""']
        params = []
        if account and account != 'all':
            clauses.append('p.account = ?')
            params.append(account)
        if q:
            clauses.append('''
                (
                    LOWER(COALESCE(p.product_number, '')) LIKE ?
                    OR LOWER(COALESCE(p.product_name, '')) LIKE ?
                    OR LOWER(COALESCE(p.spec, '')) LIKE ?
                    OR LOWER(COALESCE(p.barcode, '')) LIKE ?
                    OR LOWER(COALESCE(p.default_supplier_name, '')) LIKE ?
                )
            ''')
            kw = f'%{q}%'
            params.extend([kw, kw, kw, kw, kw])
        if supplier:
            clauses.append('''
                (
                    LOWER(COALESCE(p.default_supplier_number, '')) LIKE ?
                    OR LOWER(COALESCE(p.default_supplier_name, '')) LIKE ?
                )
            ''')
            kw = f'%{supplier}%'
            params.extend([kw, kw])
        if category and category != 'all':
            clauses.append('LOWER(COALESCE(p.category_name, "")) LIKE ?')
            params.append(f'%{category}%')
        if has_image in ('1', 'true', 'yes', 'y', 'has'):
            clauses.append("COALESCE(p.image_url, '') != ''")
        elif has_image in ('0', 'false', 'no', 'n', 'none', 'missing'):
            clauses.append("COALESCE(p.image_url, '') = ''")
        where = ' AND '.join(clauses)
        total_row = conn.execute(f'''
            SELECT COUNT(*) AS c
            FROM jdy_products p
            {joins}
            WHERE {where}
        ''', params).fetchone()
        total = int(total_row['c'] if total_row else 0)
        rows = conn.execute(f'''
            SELECT
                p.id, p.account, p.product_number, p.product_name, p.spec, p.barcode,
                p.unit_name, p.category_name, p.default_supplier_number,
                p.default_supplier_name, p.image_url, p.status,
                {select_qty},
                p.data_json
            FROM jdy_products p
            {joins}
            WHERE {where}
            ORDER BY p.account ASC, p.product_number ASC, p.id ASC
            LIMIT ? OFFSET ?
        ''', [*params, page_size, offset]).fetchall()
        items = []
        for row in rows:
            item = {
                'id': row['id'],
                'account': row['account'] or '',
                'product_number': row['product_number'] or '',
                'product_name': row['product_name'] or '',
                'spec': row['spec'] or '',
                'barcode': row['barcode'] or '',
                'unit_name': row['unit_name'] or '',
                'category_name': row['category_name'] or '',
                'default_supplier_number': row['default_supplier_number'] or '',
                'default_supplier_name': row['default_supplier_name'] or '',
                'image_url': row['image_url'] or '',
                'status': row['status'] or '',
                'stock_new': _num(row['stock_new']),
                'stock_transit': _num(row['stock_transit']),
                'stock_local': _num(row['stock_local']),
                'stock_factory': _num(row['stock_factory']),
                'factory_qty': _num(row['factory_qty']),
                'data_source': 'jdy_products',
            }
            item.update({
                'code': item['product_number'],
                'name': item['product_name'],
                'unit': item['unit_name'],
                'supplier_number': item['default_supplier_number'],
                'supplier_name': item['default_supplier_name'],
                'imageUrl': item['image_url'],
                'warehouses': [
                    {'name': '新大仓库', 'qty': item['stock_new']},
                    {'name': '在途', 'qty': item['stock_transit']},
                    {'name': '金华/本地', 'qty': item['stock_local']},
                    {'name': '工厂', 'qty': item['factory_qty'] or item['stock_factory']},
                ],
            })
            items.append(item)
        message = ''
        if products_state.get('count', 0) <= 0:
            message = '本地商品主档为空，请先在设置页补齐商品缓存。'
        return {
            **base,
            'success': True,
            'total': total,
            'page': page,
            'page_size': page_size,
            'items': items,
            'message': message,
        }, 200
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _catalog_cache_status_payload():
    product_json = {
        'account1': _local_json_cache_status(_local_product_cache_path('account1')),
        'account2': _local_json_cache_status(_local_product_cache_path('箱包')),
    }
    supplier_json = {
        'account1': _local_json_cache_status(_local_supplier_cache_path('account1')),
        'account2': _local_json_cache_status(_local_supplier_cache_path('箱包')),
    }
    conn = _sales_readonly_conn()
    try:
        jdy_products = _cache_sqlite_table_count(conn, 'jdy_products')
        if conn and jdy_products.get('exists'):
            try:
                row = conn.execute('SELECT MAX(updated_at) AS updated_at_max, MAX(last_seen_at) AS last_seen_at_max FROM jdy_products').fetchone()
                jdy_products['updated_at_max'] = row['updated_at_max'] or ''
                jdy_products['last_seen_at_max'] = row['last_seen_at_max'] or ''
            except Exception as e:
                jdy_products['stats_error'] = str(e)
        sales_details = _cache_sqlite_table_count(conn, 'sales_details')
        sales_product_quantities = _cache_sqlite_table_count(conn, 'sales_product_quantities')
        accessory_products = _cache_sqlite_table_count(conn, 'accessory_products')
        jdy_suppliers = _cache_sqlite_table_count(conn, 'jdy_suppliers')
        supplier_per_account = {}
        if conn and jdy_suppliers.get('exists'):
            try:
                rows = conn.execute('SELECT account, COUNT(*) AS c FROM jdy_suppliers GROUP BY account').fetchall()
                supplier_per_account = {
                    str(row['account'] or '未识别账套'): int(row['c'])
                    for row in rows
                }
            except Exception as e:
                jdy_suppliers['per_account_error'] = str(e)
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass
    jdy_suppliers['per_account'] = supplier_per_account if 'supplier_per_account' in locals() else {}
    product_sqlite_ready = bool(jdy_products.get('exists') and jdy_products.get('count', 0) > 0)
    supplier_sqlite_ready = bool(jdy_suppliers.get('exists') and jdy_suppliers.get('count', 0) > 0)
    return {
        'success': True,
        'local_only': True,
        'live_lookup': False,
        'called_jdy': False,
        'would_call_jdy': False,
        'product': {
            'primary_cache': 'jdy_products',
            'primary_cache_ready': product_sqlite_ready,
            'sqlite_master_table_ready': product_sqlite_ready,
            'jdy_products': jdy_products,
            'sqlite_master_table': jdy_products,
            'product_json': product_json,
            'fallbacks': {
                'sales_details': sales_details,
                'sales_product_quantities': sales_product_quantities,
                'accessory_products': accessory_products,
            },
            'impact': [
                '商品主档缺失时，商品图片、规格、条码、默认供应商补齐可能不完整',
                '返单导入仍可使用销售明细 payload、库存数量缓存和辅料商品缓存作为 fallback',
            ],
        },
        'supplier': {
            'primary_cache': 'jdy_suppliers',
            'primary_cache_ready': supplier_sqlite_ready,
            'sqlite': jdy_suppliers,
            'json_export': supplier_json,
            'note': 'SQLite jdy_suppliers 是供应商主缓存；JSON 是兼容导出/启用子集，数量不一致不一定是错误',
        },
        'recommendations': [
            '普通业务页面继续默认 local-only，不自动刷新商品主档',
            '管理员可在设置页手动只读预览 JDY 启用商品数量',
            '确认刷新后仅写入本地 SQLite jdy_products，不调用 JDY 商品写接口',
        ],
    }


SYNC_COVERAGE_TABLES = [
    'sales_orders',
    'sales_details',
    'sales_product_quantities',
    'jdy_products',
    'jdy_suppliers',
    'transfer_orders',
    'transfer_details',
    'accessory_products',
    'accessory_purchase_orders',
    'purchase_inbounds',
    'purchase_order_prices',
    'purchase_history_items',
    'bill_attachments',
    'purchase_attachment_items',
    'webhook_events',
    'webhook_runtime_settings',
    'reorder_items',
]


def _sync_coverage_table_info(conn, table):
    info = {
        'exists': False,
        'count': 0,
        'columns': [],
        'last_updated_at': '',
    }
    if conn is None:
        return info
    table = str(table or '').strip()
    if table not in SYNC_COVERAGE_TABLES:
        info['error'] = 'table is not allowed'
        return info
    try:
        row = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (table,),
        ).fetchone()
        if not row:
            return info
        cols = [str(r['name']) for r in conn.execute(f'PRAGMA table_info({table})').fetchall()]
        info['exists'] = True
        info['columns'] = cols
        count_row = conn.execute(f'SELECT COUNT(*) AS c FROM {table}').fetchone()
        info['count'] = int(count_row['c'] if count_row else 0)
        for col in ('updated_at', 'last_seen_at', 'processed_at', 'created_at', 'synced_at', 'date', 'order_date'):
            if col in cols:
                try:
                    max_row = conn.execute(f'SELECT MAX({col}) AS v FROM {table}').fetchone()
                    value = str(max_row['v'] or '') if max_row else ''
                    if value:
                        info.setdefault('max_fields', {})[col] = value
                        if not info['last_updated_at']:
                            info['last_updated_at'] = value
                except Exception as e:
                    info.setdefault('stats_errors', []).append(f'{col}: {e}')
        return info
    except Exception as e:
        info['error'] = str(e)
        return info


def _sync_coverage_tables_snapshot(conn):
    return {table: _sync_coverage_table_info(conn, table) for table in SYNC_COVERAGE_TABLES}


def _sync_coverage_counts(tables, names):
    return {
        name: int((tables.get(name) or {}).get('count') or 0)
        for name in names
    }


def _sync_coverage_last_updated(tables, names):
    values = [
        str((tables.get(name) or {}).get('last_updated_at') or '')
        for name in names
        if str((tables.get(name) or {}).get('last_updated_at') or '')
    ]
    return max(values) if values else ''


def _sync_coverage_webhook_mode(conn):
    default = {
        'processing_mode': str(_webhook_state.get('processing_mode') or 'manual'),
        'auto_enabled': bool(
            str(_webhook_state.get('processing_mode') or 'manual').strip().lower() == 'auto'
            and not _webhook_state.get('paused')
        ),
        'paused': bool(_webhook_state.get('paused')),
        'updated_at': '',
        'source': 'runtime_state',
    }
    if conn is None:
        return default
    try:
        row = conn.execute('''
            SELECT processing_mode, auto_enabled, paused, updated_at
            FROM webhook_runtime_settings
            WHERE id = 1
        ''').fetchone()
        if row:
            mode = _normalize_webhook_processing_mode(row['processing_mode'])
            return {
                'processing_mode': mode,
                'auto_enabled': bool(int(row['auto_enabled'] or 0)),
                'paused': bool(int(row['paused'] or 0)),
                'updated_at': row['updated_at'] or '',
                'source': 'webhook_runtime_settings',
            }
    except sqlite3.OperationalError:
        return default
    except Exception as e:
        default['error'] = str(e)
    return default


def _sync_coverage_resource_definitions():
    return [
        {
            'resource': 'sales / sales_order / sal_bill / delivery',
            'business_name': '销售单',
            'status': 'realtime',
            'webhook_behavior': '实时同步',
            'webhook_auto': True,
            'daily_compare': True,
            'manual_sync': True,
            'local_tables': ['sales_orders', 'sales_details', 'sales_product_quantities'],
            'risk_level': 'low',
            'note': '新建/修改销售单已能通过 Webhook 触发本地缓存更新；每日整体比对继续兜底。',
        },
        {
            'resource': 'transfer',
            'business_name': '调拨单',
            'status': 'realtime',
            'webhook_behavior': '实时同步',
            'webhook_auto': True,
            'daily_compare': True,
            'manual_sync': True,
            'local_tables': ['transfer_orders', 'transfer_details'],
            'risk_level': 'low',
            'note': 'Webhook 和每日整体比对均可写入调拨单缓存。',
        },
        {
            'resource': 'purchase_order / accessory_purchase',
            'business_name': '辅料采购订单',
            'status': 'partial_realtime',
            'webhook_behavior': '部分实时',
            'webhook_auto': True,
            'daily_compare': True,
            'manual_sync': True,
            'local_tables': ['accessory_purchase_orders'],
            'risk_level': 'medium',
            'note': '仅满足辅料供应商和单号命中条件时进入辅料采购订单缓存；内部 ID 或非辅料供应商可能只记录。',
        },
        {
            'resource': 'product',
            'business_name': '商品主档',
            'status': 'record_only',
            'webhook_behavior': '仅记录',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': True,
            'local_tables': ['jdy_products'],
            'risk_level': 'medium',
            'note': 'Webhook 已接收但暂不自动拉取商品主档；管理员可确认后手动刷新本地商品缓存。',
        },
        {
            'resource': 'supplier',
            'business_name': '供应商',
            'status': 'record_only',
            'webhook_behavior': '仅记录',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': True,
            'local_tables': ['jdy_suppliers'],
            'risk_level': 'medium',
            'note': 'Webhook 已接收但暂不自动刷新供应商主档；辅料链路可能按需读取单个供应商。',
        },
        {
            'resource': 'inventory',
            'business_name': '库存',
            'status': 'record_only',
            'webhook_behavior': '仅记录',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': True,
            'local_tables': ['sales_product_quantities', 'accessory_products'],
            'risk_level': 'high',
            'note': '库存事件不会直接刷新库存缓存；销售数量和辅料库存仅随销售/辅料相关流程刷新。',
        },
        {
            'resource': 'purchase_inbound',
            'business_name': '购货入库',
            'status': 'record_only',
            'webhook_behavior': '仅记录',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': True,
            'local_tables': ['purchase_inbounds'],
            'risk_level': 'high',
            'note': '目前只有单张核验入口可写 purchase_inbounds，且该 GET 入口会调用 JDY 并写库。',
        },
        {
            'resource': 'ar_creditbill / ar_othercreditbill / ap_otherpaybill',
            'business_name': '应收/收款/其他往来',
            'status': 'not_supported',
            'webhook_behavior': '暂无落库',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': False,
            'local_tables': [],
            'risk_level': 'high',
            'note': '当前无本地缓存表和业务落库流程，仅保留 Webhook 事件记录。',
        },
        {
            'resource': 'bd_customer',
            'business_name': '客户',
            'status': 'not_supported',
            'webhook_behavior': '暂无落库',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': False,
            'local_tables': [],
            'risk_level': 'high',
            'note': '当前无客户主档本地缓存方案。',
        },
        {
            'resource': 'attachments',
            'business_name': '附件',
            'status': 'manual_confirm',
            'webhook_behavior': '手动确认',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': True,
            'local_tables': ['bill_attachments', 'purchase_attachment_items'],
            'risk_level': 'medium',
            'note': '附件元数据和历史附件库需要管理员确认后低频执行，不进入实时 Webhook 主链路。',
        },
        {
            'resource': 'purchase_history',
            'business_name': '采购历史',
            'status': 'manual_confirm',
            'webhook_behavior': '手动确认',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': True,
            'local_tables': ['purchase_order_prices', 'purchase_inbounds', 'accessory_purchase_orders', 'purchase_attachment_items'],
            'risk_level': 'medium',
            'note': '采购历史主要从辅料采购、购货入库、价格表和历史附件索引拼出，尚无完整独立历史表。',
        },
        {
            'resource': 'reorder_items',
            'business_name': '返单',
            'status': 'local_only',
            'webhook_behavior': '纯本地业务',
            'webhook_auto': False,
            'daily_compare': False,
            'manual_sync': False,
            'local_tables': ['reorder_items'],
            'risk_level': 'low',
            'note': '返单由本地销售/商品缓存导入生成，不是 JDY 入站同步资源；同步到 JDY 功能暂未开放。',
        },
        {
            'resource': 'customs / transfer detail',
            'business_name': '报关/调拨明细',
            'status': 'local_only',
            'webhook_behavior': '本地/缓存读取',
            'webhook_auto': False,
            'daily_compare': True,
            'manual_sync': True,
            'local_tables': ['transfer_orders', 'transfer_details'],
            'risk_level': 'medium',
            'note': '报关流程不由 Webhook 驱动；调拨明细来自调拨缓存，报关文件生成仍是独立手动流程。',
        },
    ]


def _sync_coverage_risk_endpoints():
    return [
        {
            'endpoint': '/purchase-inbound-verify',
            'method': 'GET',
            'risk': '会调用 JDY 并写 purchase_inbounds，不应当作普通只读 GET。',
        },
        {
            'endpoint': '/sales-sync-run',
            'method': 'POST',
            'risk': '调用 JDY，写销售缓存，并可能触发保留期 cleanup。',
        },
        {
            'endpoint': '/sales-cache-refresh',
            'method': 'POST',
            'risk': '调用 JDY，写销售缓存或 sales_product_quantities。',
        },
        {
            'endpoint': '/accessory-orders-sync-run',
            'method': 'POST',
            'risk': '调用 JDY，写入并可能删除 accessory_purchase_orders 中未再命中的旧记录。',
        },
        {
            'endpoint': '/accessory-products-sync-run',
            'method': 'POST',
            'risk': '调用 JDY 商品、分类、库存接口，写 accessory_products。',
        },
        {
            'endpoint': '/cache/product-refresh',
            'method': 'POST',
            'risk': 'confirm 后调用 JDY product/list 并写 jdy_products。',
        },
        {
            'endpoint': '/supplier-cache/refresh',
            'method': 'POST',
            'risk': 'confirm 后调用 JDY supplier/list 并写 jdy_suppliers 和本地 JSON。',
        },
        {
            'endpoint': '/attachments/history-backfill',
            'method': 'POST',
            'risk': 'confirm 后低频调用 JDY，可下载附件并写附件索引表。',
        },
        {
            'endpoint': '/attachments/refresh',
            'method': 'POST',
            'risk': 'confirm 后调用 JDY，仅写附件元数据，不下载文件。',
        },
    ]


REALTIME_HEALTH_SAFE_GROUPS = [
    {
        'key': 'sales',
        'name': '销售单',
        'resource_keys': ['sales', 'sales_order', 'sal_bill', 'sal_bill_outbound', 'delivery'],
        'cache_tables': ['sales_orders', 'sales_details', 'sales_product_quantities'],
        'landing_table': 'sales_orders',
        'note_ok': 'Webhook done 后销售缓存已有本地更新时间。',
    },
    {
        'key': 'transfer',
        'name': '调拨单',
        'resource_keys': ['transfer'],
        'cache_tables': ['transfer_orders', 'transfer_details'],
        'landing_table': 'transfer_orders',
        'note_ok': 'Webhook done 后调拨缓存已有本地更新时间。',
    },
    {
        'key': 'accessory_purchase',
        'name': '辅料采购订单',
        'resource_keys': ['purchase_order', 'accessory_purchase', 'pur_bill_order'],
        'cache_tables': ['accessory_purchase_orders'],
        'landing_table': 'accessory_purchase_orders',
        'note_ok': '命中辅料条件的采购订单会落到辅料采购订单缓存。',
    },
]

REALTIME_HEALTH_SUSPICIOUS_TERMS = [
    '未落库',
    '未写入',
    '没有落库',
    'recorded only',
    '仅记录',
    'no match',
    'no exact match',
    'window_no_exact_match',
    'not found',
    'skipped',
    'unsupported',
]


def _realtime_health_table_exists(conn, table):
    if conn is None:
        return False
    try:
        row = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (table,),
        ).fetchone()
        return bool(row)
    except Exception:
        return False


def _realtime_health_table_columns(conn, table):
    if conn is None:
        return []
    try:
        if not _realtime_health_table_exists(conn, table):
            return []
        return [str(row['name']) for row in conn.execute(f'PRAGMA table_info({table})').fetchall()]
    except Exception:
        return []


def _realtime_health_parse_ts(value):
    text = str(value or '').strip()
    if not text:
        return None
    text = text.replace('T', ' ').replace('Z', '').strip()
    if '.' in text:
        text = text.split('.', 1)[0]
    try:
        return datetime.fromisoformat(text)
    except Exception:
        pass
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(text[:len(fmt)], fmt)
        except Exception:
            continue
    return None


def _realtime_health_event_time(row):
    for key in ('updated_at', 'processed_at', 'created_at'):
        value = str(row.get(key) or '').strip()
        if value:
            return value
    return ''


def _realtime_health_message(row):
    return str(row.get('message') or row.get('error') or '').strip()


def _realtime_health_suspicious_message(message):
    lower = str(message or '').lower()
    return any(term.lower() in lower for term in REALTIME_HEALTH_SUSPICIOUS_TERMS)


def _realtime_health_webhook_summary(conn):
    counts = {'pending': 0, 'retry_pending': 0, 'processing': 0, 'done': 0, 'failed': 0, 'ignored': 0}
    if conn is None or not _realtime_health_table_exists(conn, 'webhook_events'):
        return counts
    try:
        for row in conn.execute('SELECT status, COUNT(*) AS c FROM webhook_events GROUP BY status'):
            status = str(row['status'] or '')
            counts[status] = int(row['c'] or 0)
    except Exception as e:
        counts['error'] = str(e)
    return counts


def _realtime_health_recent_events(conn, resource_keys, limit=50):
    resource_keys = [str(x or '').strip().lower() for x in (resource_keys or []) if str(x or '').strip()]
    if conn is None or not resource_keys or not _realtime_health_table_exists(conn, 'webhook_events'):
        return []
    cols = _realtime_health_table_columns(conn, 'webhook_events')
    wanted = [
        'id', 'event_id', 'account', 'biz_type', 'resource', 'bill_no', 'number',
        'action', 'status', 'attempts', 'error', 'message',
        'created_at', 'processed_at', 'updated_at',
    ]
    select_cols = [col for col in wanted if col in cols]
    if 'id' not in select_cols:
        return []
    marks = ','.join('?' for _ in resource_keys)
    try:
        rows = conn.execute(f'''
            SELECT {', '.join(select_cols)}
            FROM webhook_events
            WHERE LOWER(COALESCE(resource, '')) IN ({marks})
            ORDER BY id DESC
            LIMIT ?
        ''', [*resource_keys, max(1, min(int(limit or 50), 200))]).fetchall()
    except Exception:
        return []
    events = []
    for row in rows:
        item = {col: row[col] for col in select_cols}
        item['resource'] = str(item.get('resource') or '')
        item['status'] = str(item.get('status') or '')
        item['bill_no'] = str(item.get('bill_no') or item.get('number') or '')
        item['message'] = _realtime_health_message(item)
        item['event_time'] = _realtime_health_event_time(item)
        events.append(item)
    return events


def _realtime_health_status_counts(events):
    counts = {'pending': 0, 'retry_pending': 0, 'processing': 0, 'done': 0, 'failed': 0, 'ignored': 0}
    for event in events or []:
        status = str(event.get('status') or '')
        counts[status] = counts.get(status, 0) + 1
    return counts


def _realtime_health_table_max_updated(conn, table):
    info = {
        'exists': False,
        'count': 0,
        'last_updated_at': '',
        'columns': [],
    }
    if conn is None:
        return info
    try:
        if not _realtime_health_table_exists(conn, table):
            return info
        cols = _realtime_health_table_columns(conn, table)
        info['exists'] = True
        info['columns'] = cols
        count_row = conn.execute(f'SELECT COUNT(*) AS c FROM {table}').fetchone()
        info['count'] = int(count_row['c'] or 0) if count_row else 0
        for col in ('updated_at', 'last_seen_at', 'processed_at', 'created_at', 'synced_at'):
            if col not in cols:
                continue
            row = conn.execute(f'SELECT MAX({col}) AS v FROM {table}').fetchone()
            value = str(row['v'] or '') if row else ''
            if value:
                info.setdefault('max_fields', {})[col] = value
                if not info['last_updated_at'] or value > info['last_updated_at']:
                    info['last_updated_at'] = value
        return info
    except Exception as e:
        info['error'] = str(e)
        return info


def _realtime_health_cache_snapshot(conn, table_names):
    return {name: _realtime_health_table_max_updated(conn, name) for name in (table_names or [])}


def _realtime_health_latest_cache_time(cache_tables):
    values = [
        str(row.get('last_updated_at') or '')
        for row in (cache_tables or {}).values()
        if str(row.get('last_updated_at') or '')
    ]
    return max(values) if values else ''


def _realtime_health_landing_exists(conn, table, account, bill_no):
    if conn is None or not table or not _realtime_health_table_exists(conn, table):
        return None
    bill_no = str(bill_no or '').strip()
    if not bill_no or _webhook_bill_no_kind(bill_no) != 'visible_number':
        return None
    cols = _realtime_health_table_columns(conn, table)
    if 'number' not in cols:
        return None
    account = str(account or '').strip()
    try:
        if account and 'account' in cols:
            row = conn.execute(
                f'SELECT 1 FROM {table} WHERE account = ? AND number = ? LIMIT 1',
                (account, bill_no),
            ).fetchone()
            if row:
                return True
        row = conn.execute(f'SELECT 1 FROM {table} WHERE number = ? LIMIT 1', (bill_no,)).fetchone()
        return bool(row)
    except Exception:
        return None


def _realtime_health_event_public(event):
    return {
        'id': event.get('id'),
        'event_id': event.get('event_id') or '',
        'account': event.get('account') or '',
        'resource': event.get('resource') or '',
        'bill_no': event.get('bill_no') or '',
        'action': event.get('action') or '',
        'status': event.get('status') or '',
        'event_time': event.get('event_time') or '',
        'message': event.get('message') or '',
        'landing_status': event.get('landing_status') or '',
    }


def _realtime_health_resource_card(conn, group):
    events = _realtime_health_recent_events(conn, group.get('resource_keys') or [], 50)
    status_counts = _realtime_health_status_counts(events)
    cache_tables = _realtime_health_cache_snapshot(conn, group.get('cache_tables') or [])
    last_cache_updated_at = _realtime_health_latest_cache_time(cache_tables)
    last_webhook_at = ''
    for event in events:
        ts = str(event.get('event_time') or '')
        if ts and ts > last_webhook_at:
            last_webhook_at = ts

    done_messages = []
    suspicious = []
    possible_done_without_landing = 0
    for event in events:
        if event.get('status') != 'done':
            continue
        message = str(event.get('message') or '')
        if message and message not in done_messages:
            done_messages.append(message)
        event_suspicious = False
        landed = _realtime_health_landing_exists(
            conn,
            group.get('landing_table') or '',
            event.get('account') or '',
            event.get('bill_no') or '',
        )
        if landed is True:
            event['landing_status'] = 'landed'
        elif landed is False:
            event['landing_status'] = 'visible_bill_not_found'
            event_suspicious = True
            suspicious.append({
                **_realtime_health_event_public(event),
                'reason': 'done 但本地主表未找到该单号',
            })
        else:
            event['landing_status'] = 'not_verifiable'
        if _realtime_health_suspicious_message(message):
            event_suspicious = True
            suspicious.append({
                **_realtime_health_event_public(event),
                'reason': 'done message 含未确认落库语义',
            })
        if event_suspicious:
            possible_done_without_landing += 1

    warning_reasons = []
    if status_counts.get('pending') or status_counts.get('processing') or status_counts.get('retry_pending'):
        warning_reasons.append('存在待处理或处理中事件')
    if status_counts.get('failed'):
        warning_reasons.append('存在失败事件')
    if possible_done_without_landing:
        warning_reasons.append('存在 done 但未确认落库的可疑事件')
    if last_webhook_at and last_cache_updated_at:
        webhook_dt = _realtime_health_parse_ts(last_webhook_at)
        cache_dt = _realtime_health_parse_ts(last_cache_updated_at)
        if webhook_dt and cache_dt and webhook_dt > cache_dt + timedelta(minutes=1):
            warning_reasons.append('最近 webhook 时间晚于本地缓存更新时间')
    elif last_webhook_at and not last_cache_updated_at:
        warning_reasons.append('已有 webhook 但本地缓存表暂无更新时间')

    if not events and not last_cache_updated_at:
        health_status = 'unknown'
        note = '尚未看到该类 webhook 或本地缓存更新时间。'
    elif status_counts.get('failed') or possible_done_without_landing:
        health_status = 'needs_attention'
        note = '；'.join(warning_reasons) or '存在需要处理的异常。'
    elif warning_reasons:
        health_status = 'warning'
        note = '；'.join(warning_reasons)
    else:
        health_status = 'ok'
        note = group.get('note_ok') or '最近实时同步健康。'

    recent_done_count = status_counts.get('done', 0)
    return {
        'key': group.get('key') or '',
        'name': group.get('name') or '',
        'resource_keys': group.get('resource_keys') or [],
        'health_status': health_status,
        'last_webhook_at': last_webhook_at,
        'last_cache_updated_at': last_cache_updated_at,
        'cache_tables': group.get('cache_tables') or [],
        'cache_table_stats': cache_tables,
        'recent_status_counts': status_counts,
        'recent_done_count': recent_done_count,
        'recent_failed_count': status_counts.get('failed', 0),
        'possible_done_without_landing_count': possible_done_without_landing,
        'done_message_samples': done_messages[:5],
        'suspicious_events': suspicious[:10],
        'recent_events': [_realtime_health_event_public(event) for event in events[:10]],
        'note': note,
    }


def _sync_realtime_health_payload():
    conn = _sales_readonly_conn()
    try:
        tables = {}
        webhook_mode = _sync_coverage_webhook_mode(conn)
        summary = _realtime_health_webhook_summary(conn)
        safe_resources = [_realtime_health_resource_card(conn, group) for group in REALTIME_HEALTH_SAFE_GROUPS]
        for card in safe_resources:
            for name, info in (card.get('cache_table_stats') or {}).items():
                tables[name] = info
        warnings = []
        for card in safe_resources:
            if card.get('health_status') in ('warning', 'needs_attention'):
                warnings.append({
                    'resource': card.get('key'),
                    'name': card.get('name'),
                    'status': card.get('health_status'),
                    'note': card.get('note'),
                    'possible_done_without_landing_count': card.get('possible_done_without_landing_count') or 0,
                })
        recent_events = []
        for card in safe_resources:
            for event in card.get('recent_events') or []:
                recent_events.append({**event, 'group': card.get('key')})
        recent_events.sort(key=lambda x: (str(x.get('event_time') or ''), int(x.get('id') or 0)), reverse=True)
        return {
            'success': True,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'db_path': _SALES_CACHE_DB,
            'webhook_mode': webhook_mode,
            'summary': summary,
            'safe_resources': safe_resources,
            'cache_tables': tables,
            'warnings': warnings,
            'recent_events': recent_events[:20],
            'notes': [
                '本面板只读取本地 webhook_events 与缓存表，用于判断实时同步是否健康；不会调用 JDY，也不会触发同步。',
                'Webhook 收到消息不等于一定自动拉取；销售单、调拨单、辅料采购订单属于安全实时同步资源。',
                '商品、供应商、库存目前仅记录事件，不作为失败；后续 dirty 队列与限速补拉另行实现。',
            ],
        }
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


def _sync_coverage_status_payload():
    conn = _sales_readonly_conn()
    try:
        tables = _sync_coverage_tables_snapshot(conn)
        webhook_mode = _sync_coverage_webhook_mode(conn)
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

    resources = []
    for item in _sync_coverage_resource_definitions():
        local_tables = item.get('local_tables') or []
        resources.append({
            **item,
            'counts': _sync_coverage_counts(tables, local_tables),
            'last_updated_at': _sync_coverage_last_updated(tables, local_tables),
        })

    return {
        'success': True,
        'local_only': True,
        'live_lookup': False,
        'called_jdy': False,
        'would_call_jdy': False,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'db_path': _SALES_CACHE_DB,
        'realtime_health_endpoint': '/sync/realtime-health',
        'webhook_mode': webhook_mode,
        'tables': tables,
        'resources': resources,
        'risk_endpoints': _sync_coverage_risk_endpoints(),
        'notes': [
            'Webhook 接收到不代表一定会自动拉取；不同资源按安全等级进入实时同步、每日兜底、手动确认或仅记录。',
            'recorded-only 表示已接收并留痕，暂不自动调用 JDY 拉取，不等同于失败。',
            '本接口只读取本地 SQLite 统计和静态覆盖配置，不调用 JDY，不写本地库。',
        ],
    }


@app.route('/sync/coverage-status', methods=['GET'])
def sync_coverage_status():
    try:
        return jsonify(_sync_coverage_status_payload())
    except Exception as e:
        return jsonify({
            'success': False,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'error': str(e),
        }), 500


@app.route('/sync/realtime-health', methods=['GET'])
def sync_realtime_health():
    try:
        return jsonify(_sync_realtime_health_payload())
    except Exception as e:
        return jsonify({
            'success': False,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'error': str(e),
        }), 500


def _product_image_url_from_cache(product):
    imgs = product.get('multiImg') if isinstance(product, dict) else []
    if isinstance(imgs, list):
        for img in imgs:
            if isinstance(img, dict):
                url = str(_first_value(img, ['url', 'fileUrl', 'downloadUrl', 'src', 'path'], '')).strip()
                if url:
                    return url
            elif img:
                return str(img)
    return str(_first_value(product or {}, ['imageUrl', 'imgUrl', 'pictureUrl', 'picUrl'], '') or '')


def _sales_summary_supplier_from_cache(account, code, product, code_cache, supplier_index):
    supplier_number = str(_first_value(product or {}, ['supplierNumber', 'supplierNo', 'vendorNumber'], '') or '').strip()
    supplier_name = str(_first_value(product or {}, ['supplierName', 'vendorName'], '') or '').strip()
    default_supplier_id = str((product or {}).get('defaultSupplierId') or '').strip()
    if default_supplier_id and not supplier_name:
        row = (supplier_index.get('by_id') or {}).get(default_supplier_id) or {}
        supplier_number = supplier_number or row.get('supplierNumber') or ''
        supplier_name = supplier_name or row.get('supplierName') or ''
    if supplier_number and not supplier_name:
        row = (supplier_index.get('by_number') or {}).get(supplier_number) or {}
        supplier_name = row.get('supplierName') or ''
    if not (supplier_number or supplier_name):
        cached = (code_cache or {}).get(code) or {}
        po = cached.get('po_result') if isinstance(cached, dict) else {}
        if isinstance(po, dict):
            supplier_number = str(po.get('supplierNumber') or '').strip()
            supplier_name = str(po.get('supplierName') or '').strip()
    return supplier_number, supplier_name


def _read_cached_sales_summary_products(start_date, end_date, account='all', q='', limit=500, offset=0, sort='qty_desc'):
    conn = _sales_readonly_conn()
    if not conn:
        return {'items': [], 'total': 0, 'error_count': 0, 'summary': {'qty': 0, 'amount': 0}}
    q = str(q or '').strip().lower()
    buckets = {}
    error_count = 0
    try:
        clauses = ['date >= ?', 'date <= ?']
        params = [start_date, end_date]
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        rows = conn.execute(f'''
            SELECT account, number, date, data_json
            FROM sales_details
            WHERE {' AND '.join(clauses)}
            ORDER BY date DESC, number DESC
        ''', params).fetchall()
    finally:
        conn.close()

    for row in rows:
        try:
            order = json.loads(row['data_json'] or '{}')
        except Exception:
            error_count += 1
            continue
        order_account = order.get('account') or row['account'] or ''
        order_no = order.get('number') or row['number'] or ''
        order_date = str(order.get('date') or row['date'] or '')[:10]
        customer = str(order.get('customerName') or '').strip()
        for entry in (order.get('entries') or []):
            if not isinstance(entry, dict):
                continue
            code = str(_first_value(entry, ['code', 'productNumber', 'number'], '') or '').strip()
            if not code:
                continue
            key = (order_account, code)
            item = buckets.setdefault(key, {
                'account': order_account,
                'code': code,
                'name': '',
                'spec': '',
                'barcode': '',
                'unit': '',
                'image_url': '',
                'category': '',
                'supplier_number': '',
                'supplier_name': '',
                'sales_qty': 0,
                'sales_amount': 0,
                'orders': set(),
                'customers': set(),
                'latest_sale_date': '',
            })
            item['name'] = item['name'] or str(_first_value(entry, ['name', 'productName', 'goodsName'], '') or '')
            item['spec'] = item['spec'] or str(_first_value(entry, ['spec', 'specification', 'model'], '') or '')
            item['barcode'] = item['barcode'] or str(_first_value(entry, ['barcode', 'barCode', 'productBarcode'], '') or '')
            item['unit'] = item['unit'] or str(_first_value(entry, ['unit', 'unitName', 'baseUnitName'], '') or '')
            item['image_url'] = item['image_url'] or str(_first_value(entry, ['imageUrl', 'imgUrl', 'pictureUrl'], '') or '')
            item['sales_qty'] += _num(_first_value(entry, ['qty', 'quantity', 'baseQty', 'mainQty'], 0))
            item['sales_amount'] += _num(_first_value(entry, ['amount', 'taxAmount', 'totalAmount'], 0))
            if order_no:
                item['orders'].add(order_no)
            if customer:
                item['customers'].add(customer)
            if order_date and order_date > item['latest_sale_date']:
                item['latest_sale_date'] = order_date

    product_indexes = {}
    supplier_indexes = {}
    code_cache = _load_code_cache()
    for (acct, code), item in list(buckets.items()):
        if acct not in product_indexes:
            product_indexes[acct] = _local_product_index(acct)
        if acct not in supplier_indexes:
            supplier_indexes[acct] = _local_supplier_index(acct)
        product = product_indexes[acct].get(code) or {}
        item['name'] = item['name'] or str(product.get('productName') or '未缓存')
        item['spec'] = item['spec'] or str(product.get('spec') or '')
        item['barcode'] = item['barcode'] or str(product.get('barcode') or '')
        item['unit'] = item['unit'] or str(product.get('unitName') or '')
        item['image_url'] = item['image_url'] or _product_image_url_from_cache(product)
        item['category'] = str(product.get('categoryName') or '') if product else ''
        supplier_number, supplier_name = _sales_summary_supplier_from_cache(
            acct, code, product, code_cache, supplier_indexes[acct]
        )
        item['supplier_number'] = supplier_number
        item['supplier_name'] = supplier_name or ''

    if buckets:
        conn = _sales_readonly_conn()
        if conn:
            try:
                by_account = {}
                for acct, code in buckets:
                    by_account.setdefault(acct, []).append(code)
                for acct, codes in by_account.items():
                    for i in range(0, len(codes), 800):
                        batch = codes[i:i + 800]
                        placeholders = ','.join('?' for _ in batch)
                        qrows = conn.execute(f'''
                            SELECT code, stock_new, stock_transit, stock_local, stock_factory, factory_qty, error
                            FROM sales_product_quantities
                            WHERE account = ? AND code IN ({placeholders})
                        ''', [acct, *batch]).fetchall()
                        for qr in qrows:
                            item = buckets.get((acct, qr['code']))
                            if not item:
                                continue
                            item['stock_new'] = _num(qr['stock_new'])
                            item['stock_transit'] = _num(qr['stock_transit'])
                            item['stock_local'] = _num(qr['stock_local'])
                            item['stock_factory'] = _num(qr['stock_factory'])
                            item['factory_qty'] = _num(qr['factory_qty'])
                            item['quantity_error'] = qr['error'] or ''
            finally:
                conn.close()

    items = []
    for item in buckets.values():
        item['order_count'] = len(item.pop('orders', set()))
        item['customer_count'] = len(item.pop('customers', set()))
        for key in ['stock_new', 'stock_transit', 'stock_local', 'stock_factory', 'factory_qty']:
            item.setdefault(key, 0)
        item.setdefault('quantity_error', '')
        item['warehouses'] = [
            {'name': '新大仓库', 'qty': item['stock_new']},
            {'name': '在途', 'qty': item['stock_transit']},
            {'name': '金华/本地', 'qty': item['stock_local']},
            {'name': '工厂', 'qty': item['factory_qty']},
        ]
        item['purchase'] = {'ordered': item['factory_qty'], 'received': 0, 'pending': item['factory_qty']}
        if q:
            hay = ' '.join(str(item.get(k) or '').lower() for k in [
                'account', 'code', 'name', 'spec', 'barcode', 'category',
                'supplier_number', 'supplier_name'
            ])
            if q not in hay:
                continue
        items.append(item)

    if sort == 'amount_desc':
        items.sort(key=lambda x: (_num(x.get('sales_amount')), _num(x.get('sales_qty'))), reverse=True)
    elif sort == 'latest_desc':
        items.sort(key=lambda x: (x.get('latest_sale_date') or '', _num(x.get('sales_qty'))), reverse=True)
    elif sort == 'code_asc':
        items.sort(key=lambda x: (x.get('account') or '', x.get('code') or ''))
    else:
        items.sort(key=lambda x: (_num(x.get('sales_qty')), _num(x.get('sales_amount'))), reverse=True)

    total = len(items)
    offset = max(0, int(offset or 0))
    limit = max(1, min(int(limit or 500), 1000))
    page_items = items[offset:offset + limit]
    return {
        'items': page_items,
        'total': total,
        'error_count': error_count,
        'summary': {
            'qty': sum(_num(x.get('sales_qty')) for x in items),
            'amount': sum(_num(x.get('sales_amount')) for x in items),
        },
    }


def _reorder_entry_identity(entry):
    return str(
        entry.get('code') or entry.get('productNumber') or entry.get('number') or ''
    ).strip()


def _reorder_entry_name(entry):
    return str(
        entry.get('name') or entry.get('productName') or entry.get('materialName') or ''
    ).strip()


def _reorder_purchase_entry_matches(entry, code):
    code = str(code or '').strip().lower()
    if not code:
        return False
    values = [
        entry.get('productNumber'), entry.get('code'), entry.get('number'),
        entry.get('materialNumber'), entry.get('productNo'), entry.get('skuNumber'),
    ]
    return any(str(v or '').strip().lower() == code for v in values)


def _reorder_sales_qty_from_cache(account, code, begin_date='', end_date=''):
    clauses = ['account = ?']
    params = [account or '']
    if begin_date:
        clauses.append('date >= ?')
        params.append(begin_date)
    if end_date:
        clauses.append('date <= ?')
        params.append(end_date)
    total = 0
    with _sales_cache_conn() as conn:
        rows = conn.execute(f'''
            SELECT data_json FROM sales_details
            WHERE {' AND '.join(clauses)}
        ''', params).fetchall()
    target = str(code or '').strip()
    for row in rows:
        try:
            order = json.loads(row['data_json'])
        except Exception:
            continue
        for entry in (order.get('entries') or []):
            if _reorder_entry_identity(entry) == target:
                total += _num(_first_value(entry, ['qty', 'quantity', 'baseQty', 'mainQty'], 0))
    return total


def _reorder_recent_sales_qty(account, code, days=60):
    end = datetime.now().date()
    begin = end - timedelta(days=max(1, int(days or 60)) - 1)
    return _reorder_sales_qty_from_cache(
        account or '',
        code,
        begin.strftime('%Y-%m-%d'),
        end.strftime('%Y-%m-%d'),
    )


def _reorder_supplier_from_cached_sources(account, code):
    account = account or ''
    code = str(code or '').strip()
    if not code:
        return {}
    with _sales_cache_conn() as conn:
        rows = conn.execute('''
            SELECT data_json FROM accessory_purchase_orders
            WHERE account = ? AND data_json LIKE ?
            ORDER BY date DESC, number DESC
            LIMIT 1
        ''', (account, f'%{code}%')).fetchall()
    for row in rows:
        try:
            order = json.loads(row['data_json'])
        except Exception:
            continue
        for entry in (order.get('entries') or []):
            if _reorder_purchase_entry_matches(entry, code):
                return {
                    'supplierNumber': order.get('supplierNumber') or '',
                    'supplierName': order.get('supplierName') or '',
                    'source': 'cached_purchase_order',
                    'orderNumber': order.get('number') or '',
                    'orderDate': order.get('date') or '',
                }
    return {}


def _reorder_supplier_from_live_purchase(cli, code):
    if not cli or not code:
        return {}
    try:
        result = cli.get_purchase_orders_by_product(code, page=1, page_size=10)
        rows = result.get('list') or []
        for row in rows:
            return {
                'supplierNumber': _first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], ''),
                'supplierName': _first_value(row, ['supplierName', 'vendorName'], ''),
                'source': 'purchase_inbound',
                'orderNumber': _first_value(row, ['number', 'billNo', 'id'], ''),
                'orderDate': str(_first_value(row, ['date', 'billDate', 'createTime'], ''))[:10],
            }
    except Exception as e:
        print(f'[REORDER] supplier live lookup failed {code}: {_short_sync_error(e)}')
    return {}


def _extract_purchase_attachments(obj):
    found = []
    seen = set()
    attachment_key_markers = ('attach', 'attachment', 'annex', 'file')
    image_only_key_markers = ('image', 'imageurl', 'pic', 'picture', 'photo', 'img')

    def add(item, hint=''):
        if not isinstance(item, dict):
            return
        url = str(_first_value(item, [
            'url', 'fileUrl', 'downloadUrl', 'attachmentUrl', 'path', 'src'
        ], '') or '').strip()
        name = str(_first_value(item, [
            'name', 'fileName', 'attachmentName', 'title', 'originalName'
        ], '') or '').strip()
        fid = str(_first_value(item, [
            'id', 'fileId', 'attachmentId', 'uid', 'key'
        ], '') or '').strip()
        mime = str(_first_value(item, [
            'file_mime', 'mime', 'mimeType', 'contentType'
        ], '') or '').strip()
        size = _num(_first_value(item, [
            'file_size', 'size', 'fileSize', 'contentLength'
        ], 0))
        hint_text = str(hint or '').lower()
        raw_keys = ' '.join(str(k).lower() for k in item.keys())
        looks_like_attachment = any(x in hint_text for x in attachment_key_markers) or any(
            x in raw_keys for x in attachment_key_markers
        )
        looks_like_product_image = any(x in hint_text for x in image_only_key_markers) or any(
            x in raw_keys for x in image_only_key_markers
        )
        if looks_like_product_image and not looks_like_attachment:
            return
        if not looks_like_attachment:
            return
        if not (url or name or fid):
            return
        key = fid or url or name
        if key in seen:
            return
        seen.add(key)
        found.append({
            'id': fid,
            'name': name or hint or fid or '附件',
            'url': url,
            'file_mime': mime,
            'file_size': size,
            'raw': item,
        })

    def walk(value, key_hint=''):
        if isinstance(value, dict):
            lowered_keys = ' '.join(str(k).lower() for k in value.keys())
            if any(x in lowered_keys for x in attachment_key_markers):
                add(value, key_hint)
            for k, v in value.items():
                walk(v, str(k))
        elif isinstance(value, list):
            for item in value:
                if any(x in str(key_hint).lower() for x in attachment_key_markers):
                    add(item if isinstance(item, dict) else {}, key_hint)
                walk(item, key_hint)

    walk(obj or {})
    return found


ATTACHMENT_SWITCH_DATE = '2026-05-28'
ATTACHMENT_SOURCE_RULE = (
    'source_date <= 2026-05-28 使用购货单附件；'
    'source_date > 2026-05-28 使用采购订单附件'
)


def _attachment_table_info(conn, table_name, include_count=True):
    table_name = str(table_name or '').strip()
    if not table_name:
        return {'exists': False, 'count': 0, 'columns': []}
    exists = bool(conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone())
    info = {'exists': exists, 'count': 0, 'columns': []}
    if not exists:
        return info
    cols = conn.execute(f'PRAGMA table_info({table_name})').fetchall()
    info['columns'] = [row['name'] if isinstance(row, sqlite3.Row) else row[1] for row in cols]
    if include_count:
        try:
            info['count'] = int(conn.execute(f'SELECT COUNT(*) AS c FROM {table_name}').fetchone()['c'] or 0)
        except Exception:
            info['count'] = 0
    return info


def _attachment_has_table(conn, table_name):
    return bool(conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (str(table_name or '').strip(),),
    ).fetchone())


def _attachment_config_dates():
    cfg = _load_ai_config()
    begin = str(cfg.get('sales_factory_purchase_begin_date') or '').strip()
    new_logic = str(cfg.get('sales_factory_new_logic_begin_date') or '').strip()
    return begin, new_logic


def _attachment_config_matches_rule(begin, new_logic):
    return begin == ATTACHMENT_SWITCH_DATE and new_logic == ATTACHMENT_SWITCH_DATE


def _attachment_preferred_source(source_date):
    source_date = str(source_date or '').strip()[:10]
    if not source_date:
        return ''
    return 'purchase_inbound' if source_date <= ATTACHMENT_SWITCH_DATE else 'purchase_order'


def _attachment_is_image(att):
    text = ' '.join(str(att.get(k) or '') for k in ('name', 'url', 'local_path', 'file_mime')).lower()
    return any(x in text for x in ('.png', '.jpg', '.jpeg', '.webp', '.gif', 'image/'))


def _attachment_payload(att, source_type, source_number, source_date, source_account, bill_type=''):
    raw = att.get('raw') if isinstance(att.get('raw'), dict) else att
    local_path = str(att.get('local_path') or att.get('localPath') or '').strip()
    local_url = _attachment_public_url(local_path)
    url = str(att.get('url') or '').strip()
    name = str(att.get('name') or att.get('fileName') or att.get('attachmentName') or '附件').strip()
    file_mime = str(att.get('file_mime') or att.get('mime') or att.get('mimeType') or '').strip()
    file_size = _num(att.get('file_size') or att.get('size') or 0)
    return {
        'id': str(att.get('id') or att.get('attachment_key') or att.get('key') or url or name),
        'name': name,
        'url': url,
        'local_path': local_path,
        'localPath': local_path,
        'local_url': local_url,
        'localUrl': local_url,
        'source_type': source_type,
        'sourceType': source_type,
        'source_number': source_number or '',
        'sourceNumber': source_number or '',
        'source_date': source_date or '',
        'sourceDate': source_date or '',
        'source_account': source_account or '',
        'sourceAccount': source_account or '',
        'bill_type': bill_type or source_type,
        'billType': bill_type or source_type,
        'file_mime': file_mime,
        'fileMime': file_mime,
        'file_size': file_size,
        'fileSize': file_size,
        'download_status': str(att.get('download_status') or ('cached' if local_path else 'not_cached')),
        'downloadStatus': str(att.get('download_status') or ('cached' if local_path else 'not_cached')),
        'download_error': str(att.get('download_error') or ''),
        'downloadError': str(att.get('download_error') or ''),
        'cached': bool(local_path),
        'is_image': _attachment_is_image({
            'name': name,
            'url': url,
            'local_path': local_path,
            'file_mime': file_mime,
        }),
        'isImage': _attachment_is_image({
            'name': name,
            'url': url,
            'local_path': local_path,
            'file_mime': file_mime,
        }),
        'raw': raw,
    }


def _read_local_bill_attachments(conn, account, source_type, source_number):
    source_number = str(source_number or '').strip()
    source_type = str(source_type or '').strip()
    if not source_number or not source_type:
        return []
    clauses = ['source_type = ?', 'source_number = ?']
    params = [source_type, source_number]
    if account:
        clauses.append('account = ?')
        params.append(account)
    rows = conn.execute(f'''
        SELECT account, source_type, source_number, source_date, bill_type,
               attachment_key, name, url, local_path, file_mime, file_size,
               download_status, download_error, data_json, updated_at
        FROM bill_attachments
        WHERE {' AND '.join(clauses)}
        ORDER BY updated_at DESC, id DESC
    ''', params).fetchall()
    items = []
    for row in rows:
        try:
            data = json.loads(row['data_json'] or '{}')
        except Exception:
            data = {}
        data.update({
            'id': row['attachment_key'] or data.get('id') or '',
            'attachment_key': row['attachment_key'] or '',
            'name': row['name'] or data.get('name') or '',
            'url': row['url'] or data.get('url') or '',
            'local_path': row['local_path'] or '',
            'file_mime': row['file_mime'] or data.get('file_mime') or data.get('mimeType') or '',
            'file_size': row['file_size'] or 0,
            'download_status': row['download_status'] or '',
            'download_error': row['download_error'] or '',
        })
        items.append(_attachment_payload(
            data,
            row['source_type'] or source_type,
            row['source_number'] or source_number,
            row['source_date'] or '',
            row['account'] or '',
            row['bill_type'] or source_type,
        ))
    return items


def _read_local_purchase_inbound_attachments(conn, account, source_number):
    source_number = str(source_number or '').strip()
    if not source_number:
        return []
    clauses = ['number = ?']
    params = [source_number]
    if account:
        clauses.append('account = ?')
        params.append(account)
    rows = conn.execute(f'''
        SELECT account, number, attachment_key, name, url, local_path, data_json, updated_at
        FROM purchase_inbound_attachments
        WHERE {' AND '.join(clauses)}
        ORDER BY updated_at DESC, attachment_key
    ''', params).fetchall()
    items = []
    for row in rows:
        try:
            data = json.loads(row['data_json'] or '{}')
        except Exception:
            data = {}
        data.update({
            'attachment_key': row['attachment_key'],
            'name': row['name'] or data.get('name') or '',
            'url': row['url'] or data.get('url') or '',
            'local_path': row['local_path'] or '',
        })
        items.append(_attachment_payload(
            data,
            'purchase_inbound',
            row['number'],
            '',
            row['account'],
            'purchase_inbound',
        ))
    return items


def _read_local_purchase_order_attachments(conn, account, source_number):
    source_number = str(source_number or '').strip()
    if not source_number:
        return []
    clauses = ['number = ?']
    params = [source_number]
    if account:
        clauses.append('account = ?')
        params.append(account)
    rows = conn.execute(f'''
        SELECT account, number, date, data_json
        FROM accessory_purchase_orders
        WHERE {' AND '.join(clauses)}
        ORDER BY date DESC, number DESC
    ''', params).fetchall()
    items = []
    for row in rows:
        try:
            order = json.loads(row['data_json'] or '{}')
        except Exception:
            order = {}
        source_date = str(order.get('date') or row['date'] or '')[:10]
        for att in _extract_purchase_attachments(order):
            items.append(_attachment_payload(
                att,
                'purchase_order',
                row['number'],
                source_date,
                row['account'],
                'purchase_order',
            ))
    return items


def _attachment_date_window(source_date, days=21):
    source_date = str(source_date or '').strip()[:10]
    if not source_date:
        return '', ''
    try:
        center = datetime.strptime(source_date, '%Y-%m-%d')
    except Exception:
        return '', ''
    return (
        (center - timedelta(days=int(days or 21))).strftime('%Y-%m-%d'),
        (center + timedelta(days=int(days or 21))).strftime('%Y-%m-%d'),
    )


def _attachment_entries_contain_code(order, code):
    code = str(code or '').strip()
    if not code:
        return True
    return any(
        isinstance(entry, dict) and _reorder_purchase_entry_matches(entry, code)
        for entry in _reorder_entries_from_order(order)
    )


def _attachment_row_value(row, key, default=''):
    try:
        return row[key]
    except Exception:
        return default


def _attachment_order_supplier(order, row=None):
    return {
        'supplier_number': str(
            _first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], '') or
            _attachment_row_value(row, 'supplier_number') or ''
        ).strip(),
        'supplier_name': str(
            _first_value(order, ['supplierName', 'vendorName'], '') or
            _attachment_row_value(row, 'supplier_name') or ''
        ).strip(),
    }


def _attachment_source_obj(account='', source_type='', source_number='', source_date='',
                           product_code='', supplier_number='', supplier_name='', **extra):
    return {
        'account': str(account or '').strip(),
        'source_type': str(source_type or '').strip(),
        'source_number': str(source_number or '').strip(),
        'source_date': str(source_date or '').strip()[:10],
        'product_code': str(product_code or '').strip(),
        'code': str(product_code or '').strip(),
        'supplier_number': str(supplier_number or '').strip(),
        'supplier_name': str(supplier_name or '').strip(),
    }


def _attachment_normalized_source_type(source_type):
    source_type = str(source_type or '').strip().lower()
    purchase_order_types = {'purchase_order', 'pur_bill_order', 'accessory_purchase', 'purchase_order_request'}
    purchase_inbound_types = {'purchase_inbound', 'purchase', 'purchase_bill', 'old_purchase'}
    if source_type in purchase_order_types:
        return 'purchase_order'
    if source_type in purchase_inbound_types:
        return 'purchase_inbound'
    return ''


def _attachment_source_confidence(item, preferred_source):
    source_type = str(item.get('source_type') or item.get('sourceType') or '').strip().lower()
    source_number = str(item.get('source_number') or item.get('sourceNumber') or '').strip()
    if not source_number:
        return 'unknown'
    sales_source_markers = ('sales', 'summary', 'reorder', 'transfer', 'allot')
    if any(marker in source_type for marker in sales_source_markers):
        return 'unknown'
    normalized_source_type = _attachment_normalized_source_type(source_type)
    if preferred_source == 'purchase_order' and normalized_source_type == 'purchase_order':
        return 'direct'
    if preferred_source == 'purchase_inbound' and normalized_source_type == 'purchase_inbound':
        return 'direct'
    upper_number = source_number.upper()
    if preferred_source == 'purchase_inbound' and upper_number.startswith(('GH', 'PI', 'PURIN', 'JH')):
        return 'direct'
    if preferred_source == 'purchase_order' and upper_number.startswith(('PO', 'CG', 'PUR', 'DD')):
        return 'direct'
    return 'candidate'


def _attachment_candidate_sources(conn, item, preferred_source, limit=8):
    account = str(item.get('account') or '').strip()
    source_number = str(item.get('source_number') or item.get('sourceNumber') or '').strip()
    source_date = str(item.get('source_date') or item.get('sourceDate') or '').strip()[:10]
    source_type = str(item.get('source_type') or item.get('sourceType') or '').strip()
    code = str(item.get('code') or item.get('product_code') or '').strip()
    supplier_number = str(item.get('supplier_number') or item.get('supplierNumber') or '').strip()
    supplier_name = str(item.get('supplier_name') or item.get('supplierName') or '').strip()
    candidates = []
    seen = set()

    def add_candidate(stype, number, date='', reason='', confidence='candidate', account_name='',
                      product_code='', supplier_no='', supplier='', attachment_count=None):
        number = str(number or '').strip()
        if not number:
            return
        key = (stype, number, account_name or account or '')
        if key in seen:
            return
        seen.add(key)
        candidates.append({
            'source_type': stype,
            'source_number': number,
            'source_date': str(date or '')[:10],
            'account': account_name or account or '',
            'product_code': product_code or code or '',
            'supplier_number': supplier_no or supplier_number or '',
            'supplier_name': supplier or supplier_name or '',
            'preferred_source': preferred_source or '',
            'attachment_count': attachment_count,
            'local_only': True,
            'live_lookup': False,
            'reason': reason,
            'confidence': confidence,
        })

    confidence = _attachment_source_confidence(item, preferred_source)
    if source_number:
        add_candidate(
            preferred_source or source_type or 'unknown',
            source_number,
            source_date,
            'source_number 看起来是采购/购货来源单号' if confidence == 'direct' else 'source_number 可能来自销售单或销售汇总，不保证是采购/购货单号',
            confidence,
            account,
        )

    try:
        date_from, date_to = _attachment_date_window(source_date, 21)
        if preferred_source == 'purchase_order' and _attachment_has_table(conn, 'accessory_purchase_orders') and code:
            clauses = ['data_json LIKE ?']
            params = [f'%{code}%']
            if account:
                clauses.append('account = ?')
                params.append(account)
            if date_from and date_to:
                clauses.append('(date = "" OR date IS NULL OR date BETWEEN ? AND ?)')
                params.extend([date_from, date_to])
            if supplier_number:
                clauses.append('(supplier_number = ? OR data_json LIKE ?)')
                params.extend([supplier_number, f'%{supplier_number}%'])
            elif supplier_name:
                clauses.append('(supplier_name LIKE ? OR data_json LIKE ?)')
                params.extend([f'%{supplier_name}%', f'%{supplier_name}%'])
            rows = conn.execute(f'''
                SELECT account, number, date, supplier_name, supplier_number, data_json
                FROM accessory_purchase_orders
                WHERE {' AND '.join(clauses)}
                ORDER BY date DESC, updated_at DESC
                LIMIT 50
            ''', params).fetchall()
            for row in rows:
                try:
                    order = json.loads(row['data_json'] or '{}')
                except Exception:
                    order = {}
                if not _attachment_entries_contain_code(order, code):
                    continue
                attachments = _extract_purchase_attachments(order)
                add_candidate(
                    'purchase_order',
                    row['number'],
                    row['date'],
                    '本地采购订单缓存中匹配到同账套/供应商/日期窗口/商品编号，仅供人工判断',
                    'candidate',
                    row['account'],
                    code,
                    row['supplier_number'],
                    row['supplier_name'],
                    len(attachments),
                )
                if len(candidates) >= int(limit or 8):
                    break
        elif preferred_source == 'purchase_inbound' and _attachment_has_table(conn, 'purchase_inbounds') and code:
            clauses = ['data_json LIKE ?']
            params = [f'%{code}%']
            if account:
                clauses.append('account = ?')
                params.append(account)
            if date_from and date_to:
                clauses.append('(date = "" OR date IS NULL OR date BETWEEN ? AND ?)')
                params.extend([date_from, date_to])
            if supplier_number:
                clauses.append('(supplier_number = ? OR data_json LIKE ?)')
                params.extend([supplier_number, f'%{supplier_number}%'])
            elif supplier_name:
                clauses.append('(supplier_name LIKE ? OR data_json LIKE ?)')
                params.extend([f'%{supplier_name}%', f'%{supplier_name}%'])
            rows = conn.execute(f'''
                SELECT account, number, date, supplier_name, supplier_number, data_json
                FROM purchase_inbounds
                WHERE {' AND '.join(clauses)}
                ORDER BY date DESC, updated_at DESC
                LIMIT 50
            ''', params).fetchall()
            for row in rows:
                try:
                    order = json.loads(row['data_json'] or '{}')
                except Exception:
                    order = {}
                if not _attachment_entries_contain_code(order, code):
                    continue
                attachments = _extract_purchase_attachments(order)
                add_candidate(
                    'purchase_inbound',
                    row['number'],
                    row['date'],
                    '本地旧购货缓存中匹配到同账套/供应商/日期窗口/商品编号，仅供人工判断',
                    'candidate',
                    row['account'],
                    code,
                    row['supplier_number'],
                    row['supplier_name'],
                    len(attachments),
                )
                if len(candidates) >= int(limit or 8):
                    break
    except Exception as e:
        candidates.append({
            'source_type': preferred_source or '',
            'source_number': '',
            'source_date': '',
            'account': account,
            'reason': f'candidate lookup failed: {_short_sync_error(e)}',
            'confidence': 'unknown',
        })
    return candidates


def _read_local_attachment_source(conn, source):
    source = _attachment_source_obj(**source)
    source_date = source['source_date']
    preferred_source = _attachment_normalized_source_type(source['source_type']) or _attachment_preferred_source(source_date)
    diagnostics = [
        '本接口只读取本地 SQLite 元数据，不调用 JDY，不下载附件，不写数据库',
    ]
    attachments = []
    if not source['source_number']:
        diagnostics.append('缺少 source_number，无法直接匹配本地附件')
    if not preferred_source:
        diagnostics.append('缺少 source_type 且 source_date 为空，无法判断首选附件来源')

    confidence = _attachment_source_confidence({
        'source_type': source['source_type'],
        'source_number': source['source_number'],
    }, preferred_source)
    candidate_sources = _attachment_candidate_sources(conn, source, preferred_source)

    if source['source_number'] and preferred_source:
        if _attachment_has_table(conn, 'bill_attachments'):
            try:
                attachments = _read_local_bill_attachments(
                    conn, source['account'], preferred_source, source['source_number']
                )
            except Exception as e:
                diagnostics.append(f'bill_attachments 读取失败：{_short_sync_error(e)}')
        else:
            diagnostics.append('bill_attachments 表不存在')
        if attachments:
            diagnostics.append('已优先从 bill_attachments 命中统一附件缓存')

        if not attachments and preferred_source == 'purchase_inbound':
            diagnostics.append('bill_attachments 未命中，尝试旧购货附件 fallback')
            if _attachment_has_table(conn, 'purchase_inbound_attachments'):
                try:
                    attachments = _read_local_purchase_inbound_attachments(
                        conn, source['account'], source['source_number']
                    )
                except Exception as e:
                    diagnostics.append(f'purchase_inbound_attachments 读取失败：{_short_sync_error(e)}')
            else:
                diagnostics.append('purchase_inbound_attachments 表不存在')
            if not attachments:
                diagnostics.append('旧购货附件 fallback 未命中')

        if not attachments and preferred_source == 'purchase_order':
            diagnostics.append('bill_attachments 未命中，尝试采购订单 data_json fallback')
            if _attachment_has_table(conn, 'accessory_purchase_orders'):
                try:
                    attachments = _read_local_purchase_order_attachments(
                        conn, source['account'], source['source_number']
                    )
                except Exception as e:
                    diagnostics.append(f'accessory_purchase_orders 读取失败：{_short_sync_error(e)}')
            else:
                diagnostics.append('accessory_purchase_orders 表不存在')
            if not attachments:
                diagnostics.append('采购订单 data_json 中未发现明确附件字段')

    if not attachments:
        diagnostics.append('本地未缓存附件；这不代表 JDY 没有附件，只代表本地还没有补齐附件元数据')
    if candidate_sources:
        diagnostics.append('候选来源仅供人工判断，不会自动认定，也不会联网补齐')
    if confidence != 'direct':
        diagnostics.append('source_number 不一定是采购/购货单号，可能来自销售单、销售汇总或返单来源')

    return {
        'success': True,
        'local_only': True,
        'live_lookup': False,
        'source': source,
        'preferred_source': preferred_source,
        'source_rule': ATTACHMENT_SOURCE_RULE,
        'source_confidence': confidence,
        'candidate_sources': candidate_sources,
        'attachments': attachments,
        'diagnostics': diagnostics,
        'message': '已读取本地附件' if attachments else '本地未缓存附件',
    }


def _attachment_order_candidates(conn, table_name, source_type, filters, limit, warnings):
    if not _attachment_has_table(conn, table_name):
        warnings.append(f'{table_name} 表不存在，已跳过')
        return []
    account = str(filters.get('account') or '').strip()
    date_from = str(filters.get('date_from') or '').strip()[:10]
    date_to = str(filters.get('date_to') or '').strip()[:10]
    rows_limit = max(1, min(int(limit or 50) * 4, 500))
    clauses = []
    params = []
    if account and account != 'all':
        clauses.append('account = ?')
        params.append(account)
    if date_from:
        clauses.append('date >= ?')
        params.append(date_from)
    if date_to:
        clauses.append('date <= ?')
        params.append(date_to)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ''
    try:
        rows = conn.execute(f'''
            SELECT account, number, date, supplier_name, supplier_number, data_json
            FROM {table_name}
            {where}
            ORDER BY date DESC, updated_at DESC
            LIMIT ?
        ''', [*params, rows_limit]).fetchall()
    except Exception as e:
        warnings.append(f'{table_name} 读取失败，可能缺少必要字段：{_short_sync_error(e)}')
        return []
    candidates = []
    for row in rows:
        try:
            order = json.loads(row['data_json'] or '{}')
        except Exception:
            order = {}
        supplier = _attachment_order_supplier(order, row)
        entries = [entry for entry in _reorder_entries_from_order(order) if isinstance(entry, dict)]
        attachments = _extract_purchase_attachments(order)
        entry_codes = []
        for entry in entries[:5]:
            code = str(_first_value(entry, [
                'productNumber', 'code', 'number', 'materialNumber', 'productNo', 'skuNumber'
            ], '') or '').strip()
            if code and code not in entry_codes:
                entry_codes.append(code)
        candidates.append({
            'account': row['account'] or '',
            'source_type': source_type,
            'source_number': row['number'] or '',
            'source_date': str(row['date'] or '')[:10],
            'product_code': entry_codes[0] if entry_codes else '',
            'product_codes': entry_codes,
            'supplier_number': supplier['supplier_number'],
            'supplier_name': supplier['supplier_name'],
            'preferred_source': source_type,
            'attachment_count': len(attachments),
            'diagnostics': [
                '本地候选来源，仅供 dry-run 预览',
                '不会调用 JDY，不会下载附件，不会写入 bill_attachments',
                '订单 data_json 中发现明确附件字段' if attachments else '订单 data_json 中暂未发现明确附件字段',
            ],
        })
        if len(candidates) >= int(limit or 50):
            break
    return candidates


def _read_local_attachment_candidates(conn, filters):
    limit = max(1, min(int(filters.get('limit') or 50), 200))
    source_type = str(filters.get('source_type') or '').strip()
    if source_type == 'all':
        source_type = ''
    warnings = [
        'dry-run 仅扫描本地 SQLite 候选来源，不调用 JDY，不下载附件，不写数据库',
    ]
    candidates = []
    if source_type in ('', 'purchase_order'):
        candidates.extend(_attachment_order_candidates(
            conn, 'accessory_purchase_orders', 'purchase_order', filters, limit - len(candidates), warnings
        ))
    if len(candidates) < limit and source_type in ('', 'purchase_inbound'):
        candidates.extend(_attachment_order_candidates(
            conn, 'purchase_inbounds', 'purchase_inbound', filters, limit - len(candidates), warnings
        ))
    candidates = candidates[:limit]
    return {
        'success': True,
        'local_only': True,
        'live_lookup': False,
        'would_call_jdy': False,
        'would_download': False,
        'candidates': candidates,
        'summary': {
            'total_candidates': len(candidates),
            'purchase_order': sum(1 for x in candidates if x.get('source_type') == 'purchase_order'),
            'purchase_inbound': sum(1 for x in candidates if x.get('source_type') == 'purchase_inbound'),
            'with_local_attachment_fields': sum(1 for x in candidates if int(x.get('attachment_count') or 0) > 0),
            'limit': limit,
        },
        'warnings': warnings,
    }


def _attachment_limit(value, default=20, maximum=50):
    try:
        limit = int(value or default)
    except Exception:
        limit = default
    return max(1, min(limit, maximum))


def _ensure_purchase_attachment_items_table(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS purchase_attachment_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account TEXT,
            source_type TEXT,
            source_number TEXT,
            source_date TEXT,
            supplier_number TEXT,
            supplier_name TEXT,
            product_code TEXT,
            product_name TEXT,
            spec TEXT,
            qty REAL,
            unit TEXT,
            price REAL,
            amount REAL,
            attachment_key TEXT,
            attachment_id INTEGER,
            data_json TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_purchase_attachment_items_product
        ON purchase_attachment_items(account, product_code)
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_purchase_attachment_items_supplier
        ON purchase_attachment_items(account, supplier_number)
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_purchase_attachment_items_source
        ON purchase_attachment_items(account, source_type, source_number)
    ''')
    conn.execute('''
        CREATE INDEX IF NOT EXISTS idx_purchase_attachment_items_date
        ON purchase_attachment_items(source_date)
    ''')
    conn.execute('''
        CREATE UNIQUE INDEX IF NOT EXISTS idx_purchase_attachment_items_unique
        ON purchase_attachment_items(account, source_type, source_number, product_code, attachment_key)
    ''')


def _history_backfill_write_conn():
    if not os.path.exists(_SALES_CACHE_DB):
        return None, '本地销售缓存不存在'
    conn = sqlite3.connect(_SALES_CACHE_DB, timeout=30)
    conn.row_factory = sqlite3.Row
    if not _attachment_has_table(conn, 'bill_attachments'):
        conn.close()
        return None, 'bill_attachments 表不存在，已停止写入'
    _ensure_purchase_attachment_items_table(conn)
    conn.commit()
    return conn, ''


def _attachment_stable_key(source_number, name='', url=''):
    raw = '|'.join(str(x or '').strip() for x in (source_number, name, url))
    return hashlib.sha1(raw.encode('utf-8', errors='ignore')).hexdigest()[:24]


def _normalize_attachment_meta(att, source):
    att = att or {}
    source = source or {}
    url = str(att.get('url') or att.get('fileUrl') or att.get('downloadUrl') or '').strip()
    name = str(att.get('name') or att.get('fileName') or att.get('attachmentName') or att.get('title') or '').strip()
    key = str(att.get('id') or att.get('attachment_key') or att.get('fileId') or att.get('attachmentId') or att.get('key') or '').strip()
    if not key:
        key = _attachment_stable_key(source.get('source_number'), name, url)
    file_mime = str(att.get('file_mime') or att.get('mime') or att.get('mimeType') or '').strip()
    file_size = int(_num(att.get('file_size') or att.get('size') or 0))
    raw = att.get('raw') if isinstance(att.get('raw'), dict) else att
    return {
        'account': source.get('account') or '',
        'source_type': source.get('source_type') or '',
        'source_number': source.get('source_number') or '',
        'source_date': source.get('source_date') or '',
        'bill_type': source.get('source_type') or '',
        'attachment_key': key,
        'name': name or key or '附件',
        'url': url,
        'local_path': '',
        'file_mime': file_mime,
        'file_size': file_size,
        'download_status': 'metadata_only',
        'download_error': '',
        'data_json': json.dumps(raw or {}, ensure_ascii=False),
    }


def _safe_attachment_filename(text, default='attachment'):
    text = str(text or '').strip() or default
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', text)
    text = re.sub(r'\s+', ' ', text).strip(' .')
    return (text or default)[:120]


def _attachment_download_root():
    return os.path.join(_DATA_BASE, '_attachments', 'purchase_documents')


def _attachment_public_url(local_path):
    local_path = str(local_path or '').strip()
    if not local_path:
        return ''
    root = os.path.abspath(_attachment_download_root())
    full = os.path.abspath(local_path)
    try:
        rel = os.path.relpath(full, root)
    except Exception:
        return ''
    if rel.startswith('..') or os.path.isabs(rel):
        return ''
    return '/attachments/local-file/' + rel.replace('\\', '/')


def _attachment_local_rel_path(source, meta):
    source_date = str(source.get('source_date') or '')[:10]
    year = source_date[:4] if len(source_date) >= 4 else datetime.now().strftime('%Y')
    month = source_date[5:7] if len(source_date) >= 7 else datetime.now().strftime('%m')
    label = '采购订单' if source.get('source_type') == 'purchase_order' else '购货单'
    name = _safe_attachment_filename(meta.get('name') or meta.get('attachment_key') or 'attachment')
    base, ext = os.path.splitext(name)
    if not ext:
        parsed_ext = os.path.splitext(urlparse(meta.get('url') or '').path)[1]
        ext = parsed_ext[:12] if parsed_ext else ''
    filename = _safe_attachment_filename(f"{label}_{source.get('source_number') or ''}_{base}")[:150]
    suffix = hashlib.sha1('|'.join([
        source.get('account') or '', source.get('source_type') or '',
        source.get('source_number') or '', meta.get('attachment_key') or '',
    ]).encode('utf-8', errors='ignore')).hexdigest()[:8]
    return os.path.join(year, month, f'{filename}_{suffix}{ext}')


def _download_attachment_file(meta, source, timeout=30, max_bytes=20 * 1024 * 1024):
    url = str(meta.get('url') or '').strip()
    if not url:
        return '', 'skipped_no_url', '附件没有明确 URL，未下载'
    rel_path = _attachment_local_rel_path(source, meta)
    abs_path = os.path.join(_attachment_download_root(), rel_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'QihangJDY-AttachmentBackfill/1.0'})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            total = 0
            with open(abs_path, 'wb') as f:
                while True:
                    chunk = resp.read(1024 * 64)
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > max_bytes:
                        raise RuntimeError('附件超过 20MB，已停止下载')
                    f.write(chunk)
        return abs_path, 'downloaded', ''
    except Exception as e:
        try:
            if os.path.exists(abs_path):
                os.remove(abs_path)
        except Exception:
            pass
        return '', 'failed', _short_sync_error(e)


def _order_entry_product_code(entry):
    return str(_first_value(entry, [
        'productNumber', 'code', 'number', 'materialNumber', 'productNo', 'skuNumber'
    ], '') or '').strip()


def _order_entry_projection(entry):
    return {
        'product_code': _order_entry_product_code(entry),
        'product_name': str(_first_value(entry, ['name', 'productName', 'materialName'], '') or '').strip(),
        'spec': str(_first_value(entry, ['spec', 'specification', 'model'], '') or '').strip(),
        'qty': _num(_first_value(entry, ['qty', 'quantity', 'baseQty', 'actualQty'], 0)),
        'unit': str(_first_value(entry, ['unit', 'unitName'], '') or '').strip(),
        'price': _num(_first_value(entry, ['price', 'taxPrice', 'unitPrice'], 0)),
        'amount': _num(_first_value(entry, ['amount', 'taxAmount', 'totalAmount'], 0)),
    }


def _upsert_purchase_attachment_item(conn, source, entry, attachment_key, attachment_id=None):
    data = _order_entry_projection(entry)
    if not data['product_code']:
        return False
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT OR REPLACE INTO purchase_attachment_items
        (id, account, source_type, source_number, source_date, supplier_number, supplier_name,
         product_code, product_name, spec, qty, unit, price, amount, attachment_key, attachment_id,
         data_json, created_at, updated_at)
        VALUES (
          COALESCE((SELECT id FROM purchase_attachment_items
                    WHERE account=? AND source_type=? AND source_number=? AND product_code=? AND attachment_key=?), NULL),
          ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
          ?,
          COALESCE((SELECT created_at FROM purchase_attachment_items
                    WHERE account=? AND source_type=? AND source_number=? AND product_code=? AND attachment_key=?), ?),
          ?
        )
    ''', (
        source.get('account') or '', source.get('source_type') or '', source.get('source_number') or '',
        data['product_code'], attachment_key,
        source.get('account') or '', source.get('source_type') or '', source.get('source_number') or '',
        source.get('source_date') or '', source.get('supplier_number') or '', source.get('supplier_name') or '',
        data['product_code'], data['product_name'], data['spec'], data['qty'], data['unit'], data['price'], data['amount'],
        attachment_key, attachment_id, json.dumps(entry or {}, ensure_ascii=False),
        source.get('account') or '', source.get('source_type') or '', source.get('source_number') or '',
        data['product_code'], attachment_key, now, now,
    ))
    return True


def _bill_attachments_existing_write_conn():
    if not os.path.exists(_SALES_CACHE_DB):
        return None, '本地销售缓存不存在'
    conn = sqlite3.connect(_SALES_CACHE_DB, timeout=30)
    conn.row_factory = sqlite3.Row
    if not _attachment_has_table(conn, 'bill_attachments'):
        conn.close()
        return None, 'bill_attachments 表不存在，已停止写入'
    return conn, ''


def _upsert_bill_attachment_metadata(conn, meta):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    existing = conn.execute('''
        SELECT id FROM bill_attachments
        WHERE account = ? AND source_type = ? AND source_number = ? AND attachment_key = ?
        LIMIT 1
    ''', (
        meta['account'], meta['source_type'], meta['source_number'], meta['attachment_key']
    )).fetchone()
    conn.execute('''
        INSERT OR REPLACE INTO bill_attachments
        (id, account, source_type, source_number, source_date, bill_type, attachment_key,
         name, url, local_path, file_mime, file_size, download_status, download_error,
         data_json, created_at, updated_at)
        VALUES (
          COALESCE((SELECT id FROM bill_attachments WHERE account=? AND source_type=? AND source_number=? AND attachment_key=?), NULL),
          ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
          COALESCE((SELECT created_at FROM bill_attachments WHERE account=? AND source_type=? AND source_number=? AND attachment_key=?), ?),
          ?
        )
    ''', (
        meta['account'], meta['source_type'], meta['source_number'], meta['attachment_key'],
        meta['account'], meta['source_type'], meta['source_number'], meta['source_date'], meta['bill_type'],
        meta['attachment_key'], meta['name'], meta['url'], meta.get('local_path') or '',
        meta['file_mime'], meta['file_size'], meta.get('download_status') or 'metadata_only',
        meta.get('download_error') or '', meta['data_json'],
        meta['account'], meta['source_type'], meta['source_number'], meta['attachment_key'], now,
        now,
    ))
    return 'updated' if existing else 'inserted'


def _attachment_account_clients(account=''):
    clients = []
    for cli_fn, name in _sales_sources_for_account(account or 'all'):
        if account and account != 'all' and name != account:
            continue
        try:
            cli = cli_fn()
        except Exception:
            cli = None
        if cli:
            clients.append((name, cli))
    return clients


def _find_live_order_for_attachment_candidate(candidate):
    account = candidate.get('account') or ''
    source_type = candidate.get('source_type') or ''
    number = candidate.get('source_number') or ''
    warnings = []
    if not number:
        return None, False, ['候选来源缺少单号，已跳过']
    clients = _attachment_account_clients(account)
    if not clients:
        return None, False, [f'{account or "全部账套"} 未配置可用 JDY 客户端，已跳过']
    called = False
    for name, cli in clients:
        try:
            if source_type == 'purchase_order':
                if not hasattr(cli, 'get_purchase_order_requests'):
                    warnings.append('当前项目未确认官方采购订单 list 查询能力，已跳过该来源。')
                    continue
                called = True
                result = cli.get_purchase_order_requests(page=1, page_size=10, search=number, check_status=2)
            elif source_type == 'purchase_inbound':
                if not hasattr(cli, 'get_purchase_orders'):
                    warnings.append('当前项目未确认官方购货/采购 list 查询能力，已跳过该来源。')
                    continue
                called = True
                result = cli.get_purchase_orders(page=1, page_size=10, search=number)
            else:
                warnings.append(f'{source_type or "未知来源"} 暂不支持官方查询，已跳过')
                continue
            rows = result.get('list') or []
            for row in rows:
                row_no = str(_first_value(row, ['number', 'billNo', 'billNumber', 'id'], '') or '').strip()
                if row_no == number:
                    row['_account'] = name
                    return row, called, warnings
            warnings.append(f'{name} 已查询 {number}，但没有精确命中候选单号')
        except Exception as e:
            called = True
            warnings.append(f'{account or name} 查询 {number} 失败：{_short_sync_error(e)}')
    return None, called, warnings


def _refresh_attachment_metadata_for_candidates(filters):
    limit = _attachment_limit(filters.get('limit'), default=20, maximum=50)
    warnings = []
    conn_ro = _sales_readonly_conn()
    if not conn_ro:
        return {
            'success': True, 'metadata_only': True, 'would_download': False,
            'local_file_download': False, 'called_jdy': False,
            'inserted': 0, 'updated': 0, 'skipped': 0, 'sources_checked': 0,
            'attachments_found': 0, 'warnings': ['本地销售缓存不存在，未查询 JDY，未写库'], 'results': [],
        }
    try:
        dry = _read_local_attachment_candidates(conn_ro, {
            'account': filters.get('account') or '',
            'source_type': filters.get('source_type') or '',
            'date_from': filters.get('date_from') or '',
            'date_to': filters.get('date_to') or '',
            'limit': limit,
        })
    finally:
        conn_ro.close()
    candidates = (dry.get('candidates') or [])[:limit]
    warnings.extend(dry.get('warnings') or [])
    if not candidates:
        return {
            'success': True, 'metadata_only': True, 'would_download': False,
            'local_file_download': False, 'called_jdy': False,
            'inserted': 0, 'updated': 0, 'skipped': 0, 'sources_checked': 0,
            'attachments_found': 0, 'warnings': warnings + ['本地 dry-run 候选为空，未调用 JDY，未写库'], 'results': [],
        }
    conn_wr, err = _bill_attachments_existing_write_conn()
    if not conn_wr:
        return {
            'success': False, 'metadata_only': True, 'would_download': False,
            'local_file_download': False, 'called_jdy': False,
            'inserted': 0, 'updated': 0, 'skipped': len(candidates), 'sources_checked': 0,
            'attachments_found': 0, 'warnings': warnings + [err], 'results': [],
            'error': err,
        }
    inserted = updated = skipped = sources_checked = attachments_found = 0
    called_jdy = False
    results = []
    try:
        for candidate in candidates:
            sources_checked += 1
            order, called, source_warnings = _find_live_order_for_attachment_candidate(candidate)
            called_jdy = called_jdy or called
            warnings.extend(source_warnings)
            if not order:
                skipped += 1
                results.append({**candidate, 'status': 'skipped', 'warnings': source_warnings})
                continue
            attachments = _extract_purchase_attachments(order)
            attachments_found += len(attachments)
            if not attachments:
                skipped += 1
                results.append({**candidate, 'status': 'no_attachments', 'attachments_found': 0, 'warnings': source_warnings})
                continue
            source = {
                'account': candidate.get('account') or order.get('_account') or '',
                'source_type': candidate.get('source_type') or '',
                'source_number': candidate.get('source_number') or '',
                'source_date': str(_first_value(order, ['date', 'billDate', 'orderDate'], candidate.get('source_date') or '') or '')[:10],
            }
            row_inserted = row_updated = 0
            for att in attachments:
                meta = _normalize_attachment_meta(att, source)
                action = _upsert_bill_attachment_metadata(conn_wr, meta)
                if action == 'inserted':
                    inserted += 1
                    row_inserted += 1
                else:
                    updated += 1
                    row_updated += 1
            conn_wr.commit()
            results.append({
                **candidate,
                'status': 'metadata_saved',
                'attachments_found': len(attachments),
                'inserted': row_inserted,
                'updated': row_updated,
                'warnings': source_warnings,
            })
    finally:
        conn_wr.close()
    return {
        'success': True,
        'metadata_only': True,
        'would_download': False,
        'local_file_download': False,
        'called_jdy': called_jdy,
        'inserted': inserted,
        'updated': updated,
        'skipped': skipped,
        'sources_checked': sources_checked,
        'attachments_found': attachments_found,
        'warnings': warnings,
        'results': results,
    }


def _history_backfill_plan(data, sample_jdy=False):
    today = datetime.now().strftime('%Y-%m-%d')
    mode = str(data.get('mode') or 'all').strip() or 'all'
    if mode not in ('old_only', 'new_only', 'all'):
        mode = 'all'
    limit = _attachment_limit(data.get('limit'), default=50 if not sample_jdy else 5, maximum=5 if sample_jdy else 200)
    page_size = _attachment_limit(data.get('page_size'), default=20, maximum=50)
    sleep_ms = max(300, int(_num(data.get('sleep_ms') or 500) or 500))
    old_from = str(data.get('old_date_from') or '2025-01-01').strip()[:10]
    old_to = str(data.get('old_date_to') or ATTACHMENT_SWITCH_DATE).strip()[:10]
    new_from = str(data.get('new_date_from') or '2026-05-29').strip()[:10]
    new_to = str(data.get('new_date_to') or today).strip()[:10]
    flows = []
    if mode in ('old_only', 'all'):
        flows.append({
            'source_type': 'purchase_inbound',
            'label': '旧流程购货/采购单',
            'date_from': old_from,
            'date_to': old_to,
            'endpoint': 'purchase/list',
        })
    if mode in ('new_only', 'all'):
        flows.append({
            'source_type': 'purchase_order',
            'label': '新流程采购订单',
            'date_from': new_from,
            'date_to': new_to,
            'endpoint': 'purchaseOrder/list',
        })
    return {
        'account': str(data.get('account') or '').strip(),
        'mode': mode,
        'limit': limit,
        'page_size': page_size,
        'sleep_ms': sleep_ms,
        'download_files': bool(data.get('download_files', True)),
        'flows': flows,
    }


def _history_fetch_flow(cli, flow, page, page_size):
    if flow['source_type'] == 'purchase_order':
        if not hasattr(cli, 'get_purchase_order_requests'):
            raise RuntimeError('当前项目未确认 purchaseOrder/list 查询能力')
        return cli.get_purchase_order_requests(
            page=page,
            page_size=page_size,
            begin_date=flow['date_from'],
            end_date=flow['date_to'],
            check_status=2,
        )
    if not hasattr(cli, 'get_purchase_orders'):
        raise RuntimeError('当前项目未确认 purchase/list 查询能力')
    return cli.get_purchase_orders(
        page=page,
        page_size=page_size,
        begin_date=flow['date_from'],
        end_date=flow['date_to'],
    )


def _history_order_source(account, source_type, order):
    supplier = _attachment_order_supplier(order)
    return {
        'account': account or '',
        'source_type': source_type,
        'source_number': str(_first_value(order, ['number', 'billNo', 'billNumber', 'id'], '') or '').strip(),
        'source_date': str(_first_value(order, ['date', 'billDate', 'orderDate', 'createTime'], '') or '')[:10],
        'supplier_number': supplier['supplier_number'],
        'supplier_name': supplier['supplier_name'],
    }


def _history_process_order(conn, account, source_type, order, download_files=True):
    source = _history_order_source(account, source_type, order)
    if not source['source_number']:
        return {'status': 'skipped', 'warning': '单据缺少单号', 'attachments_found': 0, 'items_indexed': 0}
    attachments = _extract_purchase_attachments(order)
    if not attachments:
        return {'status': 'no_attachments', 'attachments_found': 0, 'items_indexed': 0, **source}
    entries = [e for e in _reorder_entries_from_order(order) if isinstance(e, dict) and _order_entry_product_code(e)]
    inserted = updated = files_downloaded = items_indexed = skipped = 0
    warnings = []
    for att in attachments:
        meta = _normalize_attachment_meta(att, source)
        if download_files:
            local_path, status, err = _download_attachment_file(meta, source)
            meta['local_path'] = local_path
            meta['download_status'] = status
            meta['download_error'] = err
            if status == 'downloaded':
                files_downloaded += 1
            elif status in ('failed', 'skipped_no_url'):
                warnings.append(f"{meta['name']}: {err or status}")
        else:
            meta['download_status'] = 'metadata_only'
        action = _upsert_bill_attachment_metadata(conn, meta)
        if action == 'inserted':
            inserted += 1
        else:
            updated += 1
        attach_row = conn.execute('''
            SELECT id FROM bill_attachments
            WHERE account=? AND source_type=? AND source_number=? AND attachment_key=?
            LIMIT 1
        ''', (meta['account'], meta['source_type'], meta['source_number'], meta['attachment_key'])).fetchone()
        attachment_id = attach_row['id'] if attach_row else None
        for entry in entries:
            if _upsert_purchase_attachment_item(conn, source, entry, meta['attachment_key'], attachment_id):
                items_indexed += 1
    return {
        'status': 'indexed',
        **source,
        'attachments_found': len(attachments),
        'entries_count': len(entries),
        'inserted': inserted,
        'updated': updated,
        'files_downloaded': files_downloaded,
        'items_indexed': items_indexed,
        'skipped': skipped,
        'warnings': warnings,
    }


def _run_history_backfill(data):
    plan = _history_backfill_plan(data)
    try:
        backup_path = _backup_sales_cache_db('before_purchase_attachment_backfill')
    except Exception as e:
        return {
            'success': False,
            'called_jdy': False,
            'download_files': plan['download_files'],
            'metadata_written': False,
            'sources_checked': 0,
            'sources_with_attachments': 0,
            'attachments_found': 0,
            'files_downloaded': 0,
            'items_indexed': 0,
            'inserted': 0,
            'updated': 0,
            'skipped': 0,
            'warnings': [f'备份 sales_cache.sqlite3 失败，已禁止继续写库：{_short_sync_error(e)}'],
            'results': [],
        }
    conn, err = _history_backfill_write_conn()
    if not conn:
        return {
            'success': False,
            'called_jdy': False,
            'download_files': plan['download_files'],
            'metadata_written': False,
            'sources_checked': 0,
            'sources_with_attachments': 0,
            'attachments_found': 0,
            'files_downloaded': 0,
            'items_indexed': 0,
            'inserted': 0,
            'updated': 0,
            'skipped': 0,
            'warnings': [err],
            'results': [],
            'backup_path': backup_path,
        }
    warnings = []
    results = []
    called_jdy = False
    counters = {
        'sources_checked': 0,
        'sources_with_attachments': 0,
        'attachments_found': 0,
        'files_downloaded': 0,
        'items_indexed': 0,
        'inserted': 0,
        'updated': 0,
        'skipped': 0,
    }
    try:
        for account_name, cli in _attachment_account_clients(plan['account'] or 'all'):
            for flow in plan['flows']:
                page = 1
                while counters['sources_checked'] < plan['limit']:
                    try:
                        called_jdy = True
                        res = _history_fetch_flow(cli, flow, page, plan['page_size'])
                    except Exception as e:
                        warnings.append(f"{account_name} {flow['label']} 查询失败：{_short_sync_error(e)}")
                        break
                    rows = res.get('list') or []
                    if not rows:
                        break
                    for order in rows:
                        if counters['sources_checked'] >= plan['limit']:
                            break
                        counters['sources_checked'] += 1
                        summary = _history_process_order(conn, account_name, flow['source_type'], order, plan['download_files'])
                        conn.commit()
                        if summary.get('attachments_found'):
                            counters['sources_with_attachments'] += 1
                        counters['attachments_found'] += int(summary.get('attachments_found') or 0)
                        counters['files_downloaded'] += int(summary.get('files_downloaded') or 0)
                        counters['items_indexed'] += int(summary.get('items_indexed') or 0)
                        counters['inserted'] += int(summary.get('inserted') or 0)
                        counters['updated'] += int(summary.get('updated') or 0)
                        counters['skipped'] += int(summary.get('skipped') or (1 if summary.get('status') in ('skipped', 'no_attachments') else 0))
                        warnings.extend(summary.get('warnings') or [])
                        results.append(summary)
                    if len(rows) < plan['page_size']:
                        break
                    page += 1
                    time.sleep(plan['sleep_ms'] / 1000.0)
    finally:
        conn.close()
    return {
        'success': True,
        'called_jdy': called_jdy,
        'download_files': plan['download_files'],
        'metadata_written': True,
        **counters,
        'warnings': warnings,
        'results': results,
        'backup_path': backup_path,
        'plan': plan,
    }


def _read_historical_purchase_attachments(conn, item, limit=50):
    code = str(item.get('code') or item.get('product_code') or '').strip()
    account = str(item.get('account') or '').strip()
    supplier_number = str(item.get('supplier_number') or item.get('supplierNumber') or '').strip()
    supplier_name = str(item.get('supplier_name') or item.get('supplierName') or '').strip()
    diagnostics = [
        '返单附件主逻辑按商品编号查询本地历史采购附件库，不再把返单 source_number 当采购/购货单号主匹配',
    ]
    if not code:
        diagnostics.append('返单商品缺少商品编号，无法查询历史采购附件库')
        return [], diagnostics
    if not _attachment_has_table(conn, 'purchase_attachment_items'):
        diagnostics.append('purchase_attachment_items 表不存在，请先执行历史采购附件补齐')
        return [], diagnostics
    clauses = ['i.product_code = ?']
    params = [code]
    if account:
        clauses.append('i.account = ?')
        params.append(account)
    rows = conn.execute(f'''
        SELECT i.*, b.name AS attachment_name, b.url, b.local_path, b.file_mime, b.file_size,
               b.download_status, b.download_error, b.data_json AS attachment_data_json
        FROM purchase_attachment_items i
        LEFT JOIN bill_attachments b
          ON b.account = i.account
         AND b.source_type = i.source_type
         AND b.source_number = i.source_number
         AND b.attachment_key = i.attachment_key
        WHERE {' AND '.join(clauses)}
        ORDER BY
          CASE
            WHEN i.supplier_number = ? AND ? != '' THEN 0
            WHEN i.supplier_name = ? AND ? != '' THEN 0
            ELSE 1
          END,
          CASE WHEN COALESCE(b.local_path, '') != '' THEN 0 ELSE 1 END,
          CASE WHEN COALESCE(b.url, '') != '' THEN 0 ELSE 1 END,
          i.source_date DESC,
          i.source_number DESC
        LIMIT ?
    ''', [*params, supplier_number, supplier_number, supplier_name, supplier_name, int(limit or 50)]).fetchall()
    attachments = []
    for row in rows:
        same_supplier = bool(
            (supplier_number and row['supplier_number'] == supplier_number) or
            (supplier_name and row['supplier_name'] == supplier_name)
        )
        match_reason = '同商品同供应商历史采购附件' if same_supplier else '同商品其他供应商历史参考'
        payload = _attachment_payload({
            'id': row['attachment_key'],
            'attachment_key': row['attachment_key'],
            'name': row['attachment_name'] or row['attachment_key'] or '历史采购附件',
            'url': row['url'] or '',
            'local_path': row['local_path'] or '',
            'file_mime': row['file_mime'] or '',
            'file_size': row['file_size'] or 0,
            'download_status': row['download_status'] or '',
            'download_error': row['download_error'] or '',
        }, row['source_type'], row['source_number'], row['source_date'], row['account'], row['source_type'])
        payload.update({
            'supplier_number': row['supplier_number'] or '',
            'supplier_name': row['supplier_name'] or '',
            'product_code': row['product_code'] or '',
            'product_name': row['product_name'] or '',
            'spec': row['spec'] or '',
            'qty': _num(row['qty']),
            'unit': row['unit'] or '',
            'price': _num(row['price']),
            'amount': _num(row['amount']),
            'is_same_supplier': same_supplier,
            'match_reason': match_reason,
        })
        attachments.append(payload)
    if not attachments:
        diagnostics.append('本地历史采购附件库中暂未找到该商品的采购附件')
    return attachments, diagnostics


def _cache_upsert_purchase_inbound(conn, account, order):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    number = _first_value(order, ['number', 'billNo', 'id'], '')
    conn.execute('''
        INSERT OR REPLACE INTO purchase_inbounds
        (account, number, date, supplier_number, supplier_name, total_qty, total_amount, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        account or '',
        number or '',
        str(_first_value(order, ['date', 'billDate', 'createTime'], ''))[:10],
        _first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], ''),
        _first_value(order, ['supplierName', 'vendorName'], ''),
        _num(_first_value(order, ['totalQty', 'qty', 'totalQuantity'], 0)),
        _num(_first_value(order, ['totalAmount', 'amount'], 0)),
        json.dumps(order, ensure_ascii=False),
        now,
    ))
    for idx, att in enumerate(_extract_purchase_attachments(order)):
        key = att.get('id') or att.get('url') or att.get('name') or str(idx)
        conn.execute('''
            INSERT OR REPLACE INTO purchase_inbound_attachments
            (account, number, attachment_key, name, url, local_path, data_json, updated_at)
            VALUES (?, ?, ?, ?, ?, COALESCE((SELECT local_path FROM purchase_inbound_attachments WHERE account=? AND number=? AND attachment_key=?), ''), ?, ?)
        ''', (
            account or '', number or '', key, att.get('name') or '', att.get('url') or '',
            account or '', number or '', key,
            json.dumps(att, ensure_ascii=False), now,
        ))


def _read_cached_purchase_history(account, code, limit=20, date_from='', date_to=''):
    code = str(code or '').strip()
    if not code:
        return []
    limit = max(1, min(int(limit or 20), 100))
    records = []
    conn = _sales_readonly_conn()
    if not conn:
        return records
    def table_exists(name):
        row = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (name,),
        ).fetchone()
        return bool(row)
    def date_ok(value):
        value = str(value or '')[:10]
        if date_from and value and value < date_from:
            return False
        if date_to and value and value > date_to:
            return False
        return True
    def append_order_rows(table, source_type):
        if not table_exists(table):
            return
        clauses = ['data_json LIKE ?']
        params = [f'%{code}%']
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if date_from:
            clauses.append('date >= ?')
            params.append(date_from)
        if date_to:
            clauses.append('date <= ?')
            params.append(date_to)
        rows = conn.execute(f'''
            SELECT account, number, date, supplier_name, supplier_number, data_json
            FROM {table}
            WHERE {' AND '.join(clauses)}
            ORDER BY date DESC, number DESC
            LIMIT ?
        ''', [*params, limit]).fetchall()
        for row in rows:
            try:
                order = json.loads(row['data_json'] or '{}')
            except Exception:
                continue
            order_date = str(_first_value(order, ['date', 'billDate', 'orderDate', 'createTime'], row['date'] or ''))[:10]
            if not date_ok(order_date):
                continue
            matched_entries = [
                e for e in _reorder_entries_from_order(order)
                if isinstance(e, dict) and _reorder_purchase_entry_matches(e, code)
            ]
            if not matched_entries:
                continue
            for entry in matched_entries:
                qty = _purchase_entry_order_qty(entry, prefer_actual=True) or _purchase_entry_order_qty(entry)
                price = _num(_first_value(entry, ['price', 'unitPrice', 'taxPrice', 'discountPrice'], 0))
                amount = _num(_first_value(entry, ['amount', 'taxAmount', 'totalAmount'], 0))
                if not amount and price and qty:
                    amount = price * qty
                records.append({
                    'account': row['account'] or '',
                    'source_type': source_type,
                    'source_number': _first_value(order, ['number', 'billNo', 'id'], row['number']),
                    'source_date': order_date,
                    'number': _first_value(order, ['number', 'billNo', 'id'], row['number']),
                    'date': order_date,
                    'supplier_number': _first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], row['supplier_number'] or ''),
                    'supplier_name': _first_value(order, ['supplierName', 'vendorName'], row['supplier_name'] or ''),
                    'supplierNumber': _first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], row['supplier_number'] or ''),
                    'supplierName': _first_value(order, ['supplierName', 'vendorName'], row['supplier_name'] or ''),
                    'qty': qty,
                    'price': price,
                    'amount': amount,
                    'from_cache': True,
                })
    try:
        append_order_rows('purchase_inbounds', 'purchase_inbound')
        append_order_rows('accessory_purchase_orders', 'purchase_order')
        if table_exists('purchase_order_prices'):
            clauses = ['product_number = ?']
            params = [code]
            if account and account != 'all':
                clauses.append('account = ?')
                params.append(account)
            if date_from:
                clauses.append('order_date >= ?')
                params.append(date_from)
            if date_to:
                clauses.append('order_date <= ?')
                params.append(date_to)
            rows = conn.execute(f'''
                SELECT account, order_number, order_date, supplier_name, supplier_number,
                       product_number, price, qty
                FROM purchase_order_prices
                WHERE {' AND '.join(clauses)}
                ORDER BY order_date DESC, order_number DESC
                LIMIT ?
            ''', [*params, limit]).fetchall()
            for row in rows:
                qty = _num(row['qty'])
                price = _num(row['price'])
                records.append({
                    'account': row['account'] or '',
                    'source_type': 'purchase_order_price',
                    'source_number': row['order_number'] or '',
                    'source_date': row['order_date'] or '',
                    'number': row['order_number'] or '',
                    'date': row['order_date'] or '',
                    'supplier_number': row['supplier_number'] or '',
                    'supplier_name': row['supplier_name'] or '',
                    'supplierNumber': row['supplier_number'] or '',
                    'supplierName': row['supplier_name'] or '',
                    'qty': qty,
                    'price': price,
                    'amount': qty * price if qty and price else 0,
                    'from_cache': True,
                })
    finally:
        conn.close()
    records.sort(key=lambda x: (x.get('source_date') or '', x.get('source_number') or ''), reverse=True)
    return records[:limit]


def _current_reorder_user():
    username = (
        session.get('auth_user')
        or session.get('auth_name')
        or session.get('username')
        or ''
    )
    display = (
        session.get('auth_name')
        or session.get('auth_user')
        or session.get('username')
        or username
        or '未知添加人'
    )
    return str(username or display or '').strip(), str(display or username or '未知添加人').strip()


def _reorder_collect_codes(items, account=''):
    by_account = {}
    for raw in items or []:
        if not isinstance(raw, dict):
            continue
        acct = raw.get('account') or account or ''
        code = _reorder_entry_identity(raw)
        if code:
            by_account.setdefault(acct, set()).add(code)
    return by_account


def _reorder_batch_local_context(items, account=''):
    by_account = _reorder_collect_codes(items, account)
    quantities = {}
    products = {}
    suppliers = {}
    supplier_by_product = {}
    code_cache = _load_code_cache()
    for acct, codes in by_account.items():
        code_list = sorted(codes)
        for code, row in _read_cached_sales_product_quantities(acct, code_list).items():
            quantities[(acct, code)] = row
        product_index = _local_product_index(acct)
        supplier_index = _local_supplier_index(acct)
        products[acct] = product_index
        suppliers[acct] = supplier_index
        for code in code_list:
            product = product_index.get(code) or {}
            supplier_number, supplier_name = _sales_summary_supplier_from_cache(
                acct, code, product, code_cache, supplier_index
            )
            supplier_by_product[(acct, code)] = {
                'supplierNumber': supplier_number,
                'supplierName': supplier_name,
                'source': 'local_product_supplier',
            }
    return {
        'quantities': quantities,
        'products': products,
        'suppliers': suppliers,
        'supplier_by_product': supplier_by_product,
    }


def _reorder_item_payload_from_entry(entry, account='', source=None, local_ctx=None, created_by='', created_by_name=''):
    source = source or {}
    account = entry.get('account') or account or ''
    code = _reorder_entry_identity(entry)
    if str(source.get('type') or '').strip() == 'product_catalog':
        stock_new = entry.get('stock_new', entry.get('stockNew', 0))
        stock_transit = entry.get('stock_transit', entry.get('stockTransit', 0))
        stock_local = entry.get('stock_local', entry.get('stockLocal', 0))
        stock_factory = entry.get('stock_factory', entry.get('stockFactory', 0))
        factory_qty = entry.get('factory_qty', entry.get('factoryQty', stock_factory))
        return {
            'account': account or '',
            'supplierNumber': entry.get('supplier_number') or entry.get('supplierNumber') or entry.get('default_supplier_number') or '',
            'supplierName': entry.get('supplier_name') or entry.get('supplierName') or entry.get('default_supplier_name') or '未识别供应商',
            'code': code,
            'name': _reorder_entry_name(entry) or entry.get('product_name') or '',
            'spec': entry.get('spec') or entry.get('specification') or '',
            'barcode': entry.get('barcode') or entry.get('barCode') or '',
            'unit': entry.get('unit') or entry.get('unitName') or entry.get('unit_name') or '',
            'imageUrl': entry.get('imageUrl') or entry.get('image_url') or '',
            'sourceType': 'product_catalog',
            'sourceNumber': source.get('number') or 'product_catalog',
            'sourceDate': source.get('date') or '',
            'sourceCustomer': source.get('customer') or '本地商品',
            'sourceUser': source.get('user') or '',
            'salesQty60': 0,
            'stockNew': _num(stock_new),
            'stockTransit': _num(stock_transit),
            'stockLocal': _num(stock_local),
            'stockFactory': _num(stock_factory),
            'factoryQty': _num(factory_qty),
            'suggestedQty': 0,
            'confirmedQty': 0,
            'reorderPrice': _num(entry.get('reorder_price') or entry.get('reorderPrice') or 0),
            'createdBy': created_by or '',
            'createdByName': created_by_name or '未知添加人',
            'raw': entry,
            'supplierSource': {'source': 'product_catalog_payload'},
            'localOnly': True,
            'liveLookup': False,
        }
    warehouses = entry.get('warehouses') or []
    stock = {str(w.get('name') or ''): _num(w.get('qty')) for w in warehouses if isinstance(w, dict)}
    local_ctx = local_ctx or _reorder_batch_local_context([entry], account)
    cached_q = (local_ctx.get('quantities') or {}).get((account, code)) or {}
    cached_stock = cached_q.get('stock') or {}
    product = ((local_ctx.get('products') or {}).get(account) or {}).get(code) or {}
    supplier = {
        'supplierNumber': entry.get('supplier_number') or entry.get('supplierNumber') or '',
        'supplierName': entry.get('supplier_name') or entry.get('supplierName') or '',
    }
    if not (supplier.get('supplierNumber') or supplier.get('supplierName')):
        supplier = (local_ctx.get('supplier_by_product') or {}).get((account, code)) or {}
    image_url = entry.get('imageUrl') or entry.get('image_url') or _product_image_url_from_cache(product)
    factory_qty = _num(cached_q.get('factory_qty') or (entry.get('purchase') or {}).get('pending') or 0)
    return {
        'account': account or '',
        'supplierNumber': supplier.get('supplierNumber') or '',
        'supplierName': supplier.get('supplierName') or '未识别供应商',
        'code': code,
        'name': _reorder_entry_name(entry) or str(product.get('productName') or ''),
        'spec': entry.get('spec') or entry.get('specification') or str(product.get('spec') or ''),
        'barcode': entry.get('barcode') or entry.get('barCode') or str(product.get('barcode') or ''),
        'unit': entry.get('unit') or entry.get('unitName') or str(product.get('unitName') or ''),
        'imageUrl': image_url,
        'sourceType': source.get('type') or 'sales',
        'sourceNumber': source.get('number') or '',
        'sourceDate': source.get('date') or '',
        'sourceCustomer': source.get('customer') or '',
        'sourceUser': source.get('user') or '',
        'salesQty60': 0,
        'stockNew': _num(cached_stock.get('新大仓库') or cached_stock.get('鏂板ぇ浠撳簱') or stock.get('新大仓库') or stock.get('鏂板ぇ浠撳簱')),
        'stockTransit': _num(cached_stock.get('在途') or cached_stock.get('鍦ㄩ€?') or stock.get('在途') or stock.get('鍦ㄩ€?')),
        'stockLocal': _num(cached_stock.get('金华/本地') or cached_stock.get('閲戝崕/鏈湴') or stock.get('金华/本地') or stock.get('閲戝崕/鏈湴')),
        'stockFactory': _num(cached_stock.get('工厂订单') or cached_stock.get('宸ュ巶璁㈠崟') or stock.get('工厂') or stock.get('宸ュ巶')),
        'factoryQty': factory_qty,
        'suggestedQty': 0,
        'confirmedQty': 0,
        'reorderPrice': _num(entry.get('reorder_price') or entry.get('reorderPrice') or 0),
        'createdBy': created_by or '',
        'createdByName': created_by_name or '未知添加人',
        'raw': entry,
        'supplierSource': supplier,
        'localOnly': True,
        'liveLookup': False,
    }


def _cache_insert_reorder_item(conn, item):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    existing = conn.execute('''
        SELECT id FROM reorder_items
        WHERE status = 'pending'
          AND COALESCE(account, '') = ?
          AND code = ?
          AND COALESCE(source_number, '') = ?
        LIMIT 1
    ''', (item.get('account') or '', item.get('code') or '', item.get('sourceNumber') or '')).fetchone()
    data_json = json.dumps(item, ensure_ascii=False)
    if existing:
        conn.execute('''
            UPDATE reorder_items
               SET supplier_number=?, supplier_name=?, name=?, spec=?, barcode=?, unit=?,
                   image_url=?, sales_qty_60=?, stock_new=?, stock_transit=?, stock_local=?,
                   stock_factory=?, factory_qty=?,
                   created_by=COALESCE(NULLIF(created_by, ''), ?),
                   created_by_name=COALESCE(NULLIF(created_by_name, ''), ?),
                   data_json=?, updated_at=?
             WHERE id=?
        ''', (
            item.get('supplierNumber') or '', item.get('supplierName') or '',
            item.get('name') or '', item.get('spec') or '', item.get('barcode') or '',
            item.get('unit') or '', item.get('imageUrl') or '',
            _num(item.get('salesQty60')), _num(item.get('stockNew')),
            _num(item.get('stockTransit')), _num(item.get('stockLocal')),
            _num(item.get('stockFactory')), _num(item.get('factoryQty')),
            item.get('createdBy') or '', item.get('createdByName') or '',
            data_json, now, existing['id'],
        ))
        return existing['id'], False
    cur = conn.execute('''
        INSERT INTO reorder_items
        (account, supplier_number, supplier_name, code, name, spec, barcode, unit,
         image_url, source_type, source_number, source_date, source_customer, source_user,
         sales_qty_60, stock_new, stock_transit, stock_local, stock_factory, factory_qty,
         suggested_qty, confirmed_qty, status, note, created_by, created_by_name,
         reorder_price, data_json, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', '', ?, ?, ?, ?, ?, ?)
    ''', (
        item.get('account') or '', item.get('supplierNumber') or '', item.get('supplierName') or '',
        item.get('code') or '', item.get('name') or '', item.get('spec') or '',
        item.get('barcode') or '', item.get('unit') or '', item.get('imageUrl') or '',
        item.get('sourceType') or '', item.get('sourceNumber') or '', item.get('sourceDate') or '',
        item.get('sourceCustomer') or '', item.get('sourceUser') or '',
        _num(item.get('salesQty60')), _num(item.get('stockNew')), _num(item.get('stockTransit')),
        _num(item.get('stockLocal')), _num(item.get('stockFactory')), _num(item.get('factoryQty')),
        _num(item.get('suggestedQty')), _num(item.get('confirmedQty')),
        item.get('createdBy') or '', item.get('createdByName') or '未知添加人',
        _num(item.get('reorderPrice')), data_json, now, now,
    ))
    return cur.lastrowid, True


def _reorder_row_to_item(row):
    try:
        data = json.loads(row['data_json'] or '{}')
    except Exception:
        data = {}
    keys = set(row.keys()) if hasattr(row, 'keys') else set()
    def rv(name, default=''):
        return row[name] if name in keys else default
    data.update({
        'id': row['id'],
        'account': row['account'] or '',
        'supplierNumber': row['supplier_number'] or '',
        'supplier_number': row['supplier_number'] or '',
        'supplierName': row['supplier_name'] or '',
        'supplier_name': row['supplier_name'] or '',
        'code': row['code'] or '',
        'name': row['name'] or '',
        'spec': row['spec'] or '',
        'barcode': row['barcode'] or '',
        'unit': row['unit'] or '',
        'imageUrl': row['image_url'] or '',
        'image_url': row['image_url'] or '',
        'sourceType': row['source_type'] or '',
        'source_type': row['source_type'] or '',
        'sourceNumber': row['source_number'] or '',
        'source_number': row['source_number'] or '',
        'sourceDate': row['source_date'] or '',
        'source_date': row['source_date'] or '',
        'sourceCustomer': row['source_customer'] or '',
        'source_customer': row['source_customer'] or '',
        'salesQty60': _num(row['sales_qty_60']),
        'sales_qty_60': _num(row['sales_qty_60']),
        'stockNew': _num(row['stock_new']),
        'stock_new': _num(row['stock_new']),
        'stockTransit': _num(row['stock_transit']),
        'stock_transit': _num(row['stock_transit']),
        'stockLocal': _num(row['stock_local']),
        'stock_local': _num(row['stock_local']),
        'stockFactory': _num(row['stock_factory']),
        'stock_factory': _num(row['stock_factory']),
        'factoryQty': _num(row['factory_qty']),
        'factory_qty': _num(row['factory_qty']),
        'suggestedQty': _num(row['suggested_qty']),
        'suggested_qty': _num(row['suggested_qty']),
        'confirmedQty': _num(row['confirmed_qty']),
        'confirmed_qty': _num(row['confirmed_qty']),
        'status': row['status'] or '',
        'note': row['note'] or '',
        'createdBy': rv('created_by') or '',
        'created_by': rv('created_by') or '',
        'createdByName': rv('created_by_name') or '未知添加人',
        'created_by_name': rv('created_by_name') or '未知添加人',
        'reorderPrice': _num(rv('reorder_price', 0)),
        'reorder_price': _num(rv('reorder_price', 0)),
        'generatedBatchNo': rv('generated_batch_no') or '',
        'generated_batch_no': rv('generated_batch_no') or '',
        'generatedAt': rv('generated_at') or '',
        'generated_at': rv('generated_at') or '',
        'syncStatus': rv('sync_status') or 'not_synced',
        'sync_status': rv('sync_status') or 'not_synced',
        'jdyOrderNo': rv('jdy_order_no') or '',
        'jdy_order_no': rv('jdy_order_no') or '',
        'syncedAt': rv('synced_at') or '',
        'synced_at': rv('synced_at') or '',
        'syncError': rv('sync_error') or '',
        'sync_error': rv('sync_error') or '',
        'createdAt': row['created_at'] or '',
        'created_at': row['created_at'] or '',
        'updatedAt': row['updated_at'] or '',
        'updated_at': row['updated_at'] or '',
        'warehouses': [
            {'name': '新大仓库', 'qty': _num(row['stock_new'])},
            {'name': '在途', 'qty': _num(row['stock_transit'])},
            {'name': '金华/本地', 'qty': _num(row['stock_local'])},
            {'name': '工厂', 'qty': _num(row['factory_qty'])},
        ],
    })
    return data


def _reorder_stock_from_cache(conn, account, code):
    if not code:
        return {}
    params = [code]
    where = 'code = ?'
    if account:
        where += ' AND account = ?'
        params.append(account)
    row = conn.execute(f'''
        SELECT stock_new, stock_transit, stock_local, stock_factory, factory_qty
        FROM sales_product_quantities
        WHERE {where}
        ORDER BY updated_at DESC
        LIMIT 1
    ''', params).fetchone()
    if not row:
        return {}
    return {
        'stockNew': _num(row['stock_new']),
        'stockTransit': _num(row['stock_transit']),
        'stockLocal': _num(row['stock_local']),
        'stockFactory': _num(row['stock_factory']),
        'factoryQty': _num(row['factory_qty']),
    }


def _reorder_entries_from_order(order):
    if not isinstance(order, dict):
        return []
    entries = (
        order.get('entries') or order.get('details') or order.get('items') or
        order.get('lines') or order.get('materialEntries') or order.get('billEntry') or
        order.get('billEntries') or order.get('entry') or []
    )
    if isinstance(entries, dict):
        entries = [entries]
    return entries if isinstance(entries, list) else []


def _reorder_recent_style_from_entry(entry, account, supplier, source_order, conn):
    code = _reorder_entry_identity(entry)
    if not code:
        return None
    stock = _reorder_stock_from_cache(conn, account, code)
    return {
        'account': account or '',
        'supplierNumber': supplier.get('supplierNumber') or '',
        'supplierName': supplier.get('supplierName') or '未识别供应商',
        'code': code,
        'name': _reorder_entry_name(entry),
        'spec': _first_value(entry, ['spec', 'model', 'specification', 'materialModel', 'materialSpec'], ''),
        'barcode': _first_value(entry, ['barcode', 'barCode', 'number', 'materialNumber'], ''),
        'unit': _first_value(entry, ['unit', 'unitName', 'baseUnitName'], ''),
        'imageUrl': _first_value(entry, ['imageUrl', 'image', 'picUrl', 'picture'], ''),
        'sourceNumber': source_order.get('number') or source_order.get('billNo') or '',
        'sourceDate': str(source_order.get('date') or source_order.get('billDate') or '')[:10],
        'salesQty60': _reorder_recent_sales_qty(account, code),
        'warehouses': [
            {'name': '新大仓库', 'qty': stock.get('stockNew', 0)},
            {'name': '在途', 'qty': stock.get('stockTransit', 0)},
            {'name': '金华/本地', 'qty': stock.get('stockLocal', 0)},
            {'name': '工厂', 'qty': stock.get('factoryQty', stock.get('stockFactory', 0))},
        ],
    }


def _normalize_transfer_order(row, account):
    entries = row.get('entries') or row.get('details') or row.get('items') or row.get('lines') or []
    out_locs = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        loc = (
            entry.get('outLocationName')
            or entry.get('outWarehouseName')
            or entry.get('outStockName')
            or ''
        )
        if loc and loc not in out_locs:
            out_locs.append(loc)
    number = _first_value(row, ['number', 'billNo', 'id'], '')
    date_str = str(_first_value(row, ['date', 'billDate', 'createTime'], ''))[:10]
    norm = dict(row)
    norm.update({
        'number': number,
        'date': date_str,
        '_account': account,
        '_outLocNames': out_locs,
        '_outLocName': out_locs[0] if out_locs else '',
    })
    return norm


_TRANSFER_ORDER_HASH_VERSION = 'transfer-order-v1'


def _transfer_order_signature(order):
    keep = dict(order or {})
    keep.pop('cacheHash', None)
    keep.pop('cacheHashVersion', None)
    raw = json.dumps(keep, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def _cached_transfer_hash(conn, account, number):
    row = conn.execute('''
        SELECT data_json FROM transfer_details
        WHERE account = ? AND number = ?
    ''', (account or '', number or '')).fetchone()
    if not row:
        return None
    try:
        cached = json.loads(row['data_json'])
        if cached.get('cacheHashVersion') == _TRANSFER_ORDER_HASH_VERSION and cached.get('cacheHash'):
            return cached.get('cacheHash')
        return _transfer_order_signature(cached)
    except Exception:
        return ''


def _cache_upsert_transfer_order(conn, order):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    order = dict(order or {})
    order['cacheHash'] = _transfer_order_signature(order)
    order['cacheHashVersion'] = _TRANSFER_ORDER_HASH_VERSION
    account = order.get('_account') or order.get('account') or ''
    number = order.get('number') or order.get('billNo') or order.get('id') or ''
    date_str = str(order.get('date') or order.get('billDate') or order.get('createTime') or '')[:10]
    conn.execute('''
        INSERT OR REPLACE INTO transfer_orders
        (account, number, date, out_location, check_status, remark, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        account,
        number,
        date_str,
        order.get('_outLocName') or '',
        str(order.get('checkStatusName') or order.get('statusName') or order.get('checkStatus') or ''),
        order.get('remark') or '',
        json.dumps(order, ensure_ascii=False),
        now,
    ))
    conn.execute('''
        INSERT OR REPLACE INTO transfer_details
        (account, number, date, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        account,
        number,
        date_str,
        json.dumps(order, ensure_ascii=False),
        now,
    ))


def _cache_one_sales_order_from_jdy(cli, account, number, include_quantities=True):
    if not cli or not number:
        return _webhook_process_result(
            failed=True,
            reason='missing_bill_no',
            message='sales cache failed: missing JDY client or bill number',
        )
    detail = cli.get_sales_order_detail(number)
    result = _cache_sales_rows_from_jdy_rows(
        cli, account, [detail], include_quantities=include_quantities, source='webhook_exact'
    )
    if result.get('cached'):
        return _webhook_process_result(
            ok=True,
            cached=True,
            reason='cached_exact_sales_order',
            message=f'cached exact sales order: {number}; orders={result.get("cached")}, quantities={result.get("quantity_codes")}',
        )
    return _webhook_process_result(
        failed=True,
        reason='sales_order_not_found',
        message=f'sales order not found or empty: {number}',
    )


def _cache_one_transfer_order_from_jdy(cli, account, number):
    if not cli or not number:
        return False
    detail = cli.get_transfer_order_detail(number)
    order = _normalize_transfer_order(detail, account)
    if not (order.get('number') or order.get('billNo') or order.get('id')):
        return False
    with _sales_cache_conn() as conn:
        _cache_upsert_transfer_order(conn, order)
        conn.commit()
    return True


def _webhook_process_result(ok=False, cached=False, reason='', message='', failed=False):
    return {
        'ok': bool(ok),
        'cached': bool(cached),
        'reason': str(reason or ''),
        'message': str(message or reason or ''),
        'failed': bool(failed),
    }


def _accessory_webhook_payload(event):
    if not isinstance(event, dict):
        return {}
    try:
        payload = json.loads(event.get('payload_json') or '{}')
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def _accessory_webhook_visible_numbers(payload):
    return _extract_webhook_values(payload or {}, [
        'number', 'billNo', 'billNumber', 'bill_no',
        'orderNo', 'orderNumber', 'sourceBillNo', 'srcBillNo',
    ])


def _accessory_webhook_internal_ids(payload, fallback_number=''):
    values = _extract_webhook_values(payload or {}, ['id', 'billId', 'bill_id', 'billID', 'dataId'])
    fallback_number = str(fallback_number or '').strip()
    if fallback_number and _webhook_bill_no_kind(fallback_number) == 'internal_id':
        values.append(fallback_number)
    return [x for x in dict.fromkeys(str(v).strip() for v in values if str(v or '').strip()) if x]


def _accessory_webhook_should_try_window(event, number):
    payload = _accessory_webhook_payload(event)
    return bool(_accessory_webhook_internal_ids(payload, number)) and not _accessory_webhook_visible_numbers(payload)


def _parse_webhook_window_time(value):
    text = str(value or '').strip()
    if not text:
        return None
    text = text.replace('T', ' ').replace('/', '-')
    if '+' in text:
        text = text.split('+', 1)[0].strip()
    if text.endswith('Z'):
        text = text[:-1].strip()
    text = text[:19]
    formats = (
        ('%Y-%m-%d %H:%M:%S', 19),
        ('%Y-%m-%d %H:%M', 16),
        ('%Y-%m-%d', 10),
    )
    for fmt, size in formats:
        try:
            return datetime.strptime(text[:size], fmt)
        except Exception:
            pass
    return None


def _accessory_purchase_row_internal_ids(row):
    values = []
    if isinstance(row, dict):
        for key in ('id', 'billId', 'bill_id', 'billID', 'dataId'):
            val = row.get(key)
            if val not in (None, ''):
                values.append(str(val).strip())
        raw = row.get('_raw')
        if isinstance(raw, dict):
            values.extend(_accessory_purchase_row_internal_ids(raw))
    return [x for x in dict.fromkeys(values) if x]


def _accessory_purchase_row_time(row):
    if not isinstance(row, dict):
        return None
    for key in ('updTime', 'updateTime', 'updatedAt', 'modifyTime', 'createTime', 'createdAt', 'billDate', 'date'):
        parsed = _parse_webhook_window_time(row.get(key))
        if parsed:
            return parsed
    return None


def _cache_accessory_purchase_window_candidate(cli, account, row, internal_id):
    number = str(_first_value(row or {}, ['number', 'billNo', 'billNumber', 'id'], '')).strip()
    try:
        with _sales_cache_conn() as conn:
            supplier_no = str(_first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], '')).strip()
            supplier = _read_cached_jdy_supplier(conn, account, supplier_no) if supplier_no else {}
            if supplier_no and supplier is None:
                supplier = cli.get_supplier_by_number(supplier_no, status=2) or {}
                if supplier:
                    _cache_upsert_jdy_supplier(conn, account, supplier)
            order = _normalize_accessory_purchase_order(row, account, supplier or {})
            if not _supplier_matches_accessory(supplier or {}, row):
                supplier_name = order.get('supplierName') or supplier_no or '未知供应商'
                return _webhook_process_result(
                    reason='accessory_supplier_not_matched',
                    message=f'已收到精斗云推送，但对应采购订单不是辅料供应商，未保存：{supplier_name}（精斗云内部编号：{internal_id}）',
                )
            if not order.get('entries'):
                return _webhook_process_result(
                    reason='accessory_order_empty_entries',
                    message=f'已收到精斗云推送，但对应采购订单没有商品明细，未保存：{order.get("number") or number or "未知单号"}（精斗云内部编号：{internal_id}）',
                )
            order = _enrich_accessory_purchase_order(order)
            _cache_upsert_accessory_purchase_order(conn, order)
            conn.commit()
            return _webhook_process_result(
                ok=True,
                cached=True,
                reason='accessory_window_cached',
                message=f'已通过小时间窗口匹配并保存辅料采购订单：{order.get("number") or number}（精斗云内部编号：{internal_id}）',
            )
    except Exception as e:
        return _webhook_process_result(
            failed=True,
            reason='accessory_window_save_error',
            message=f'已收到精斗云推送，但保存辅料采购订单失败，未写入空数据：{_short_sync_error(e)}（精斗云内部编号：{internal_id}）',
        )


def _cache_accessory_purchase_order_for_webhook_window(cli, account, payload, internal_id, event_timestamp):
    internal_id = str(internal_id or '').strip()
    if not cli or not account or not internal_id:
        return _webhook_process_result(
            reason='accessory_window_missing_context',
            message=f'已收到精斗云推送，但推送内容只有内部编号，暂无法确认采购订单（精斗云内部编号：{internal_id or "空"}）',
        )
    event_dt = _parse_webhook_window_time(event_timestamp) or datetime.now()
    window_minutes = 10
    window_start = event_dt - timedelta(minutes=window_minutes)
    window_end = event_dt + timedelta(minutes=window_minutes)
    page_size = 25
    max_pages = 2
    max_rows = 50
    rows = []
    try:
        for page in range(1, max_pages + 1):
            result = cli.get_purchase_order_requests(
                page=page,
                page_size=page_size,
                bill_status=None,
                check_status=2,
                begin_date=window_start.strftime('%Y-%m-%d'),
                end_date=window_end.strftime('%Y-%m-%d'),
            )
            batch = result.get('list') or []
            rows.extend(batch[:max(0, max_rows - len(rows))])
            total = result.get('total') or 0
            if not batch or len(batch) < page_size or len(rows) >= max_rows or (total and len(rows) >= int(total)):
                break
        exact_matches = [
            row for row in rows
            if internal_id in _accessory_purchase_row_internal_ids(row)
        ]
        if len(exact_matches) == 1:
            return _cache_accessory_purchase_window_candidate(cli, account, exact_matches[0], internal_id)
        if len(exact_matches) > 1:
            return _webhook_process_result(
                reason='accessory_window_multiple_exact',
                message=f'已收到精斗云推送，但窗口内匹配到多条候选，未自动保存，避免误写（精斗云内部编号：{internal_id}）',
            )
        window_rows = []
        for row in rows:
            row_time = _accessory_purchase_row_time(row)
            if row_time and window_start <= row_time <= window_end:
                window_rows.append(row)
        if len(window_rows) == 1:
            return _cache_accessory_purchase_window_candidate(cli, account, window_rows[0], internal_id)
        if len(window_rows) > 1:
            return _webhook_process_result(
                reason='accessory_window_multiple_candidates',
                message=f'已收到精斗云推送，但窗口内匹配到多条候选，未自动保存，避免误写（精斗云内部编号：{internal_id}）',
            )
        return _webhook_process_result(
            reason='accessory_window_not_found',
            message=f'已收到精斗云推送，但推送内容只有内部编号，窗口内未匹配到辅料采购订单（精斗云内部编号：{internal_id}）',
        )
    except Exception as e:
        return _webhook_process_result(
            failed=True,
            reason='accessory_window_query_error',
            message=f'已收到精斗云推送，但小时间窗口查询采购订单失败，未保存：{_short_sync_error(e)}（精斗云内部编号：{internal_id}）',
        )


def _cache_one_accessory_purchase_order_from_jdy(cli, account, number, event=None):
    if not cli or not number:
        return _webhook_process_result(
            reason='missing_bill_no',
            message='缺少采购订单单号，无法更新辅料页面',
        )
    payload = _accessory_webhook_payload(event)
    internal_ids = _accessory_webhook_internal_ids(payload, number)
    should_try_window = _accessory_webhook_should_try_window(event, number)
    result = cli.get_purchase_order_requests(
        page=1, page_size=1, search=number, bill_status=0
    )
    rows = result.get('list') or []
    if not rows:
        if should_try_window:
            return _cache_accessory_purchase_order_for_webhook_window(
                cli, account, payload, internal_ids[0] if internal_ids else number, (event or {}).get('created_at')
            )
        return _webhook_process_result(
            reason='purchase_order_not_found',
            message='采购订单未查到，可能 webhook 传的是内部 ID，或 bill_status=0 状态过滤导致未命中',
        )
    exact_mismatch = False
    with _sales_cache_conn() as conn:
        for row in rows:
            row_no = str(_first_value(row, ['number', 'billNo', 'id'], '')).strip()
            if row_no and row_no != str(number).strip():
                exact_mismatch = True
                continue
            supplier_no = str(_first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], '')).strip()
            supplier = _read_cached_jdy_supplier(conn, account, supplier_no) if supplier_no else {}
            if supplier_no and supplier is None:
                supplier = cli.get_supplier_by_number(supplier_no, status=2) or {}
                if supplier:
                    _cache_upsert_jdy_supplier(conn, account, supplier)
            order = _normalize_accessory_purchase_order(row, account, supplier or {})
            if not _supplier_matches_accessory(supplier or {}, row):
                if supplier_no and not supplier:
                    return _webhook_process_result(
                        reason='supplier_missing',
                        message=f'供应商资料缺失，无法判断是否为辅料供应商：{supplier_no}',
                    )
                supplier_name = order.get('supplierName') or supplier_no or '未知供应商'
                return _webhook_process_result(
                    reason='supplier_not_accessory',
                    message=f'供应商不是辅料供应商，未进入辅料页面：{supplier_name}',
                )
            if not order.get('entries'):
                return _webhook_process_result(
                    reason='empty_entries',
                    message=f'采购订单无商品明细，未进入辅料页面：{order.get("number") or number}',
                )
            order = _enrich_accessory_purchase_order(order)
            _cache_upsert_accessory_purchase_order(conn, order)
            conn.commit()
            return _webhook_process_result(
                ok=True,
                cached=True,
                reason='cached',
                message=f'已缓存辅料采购订单：{order.get("number") or number}',
            )
    if exact_mismatch:
        if should_try_window:
            return _cache_accessory_purchase_order_for_webhook_window(
                cli, account, payload, internal_ids[0] if internal_ids else number, (event or {}).get('created_at')
            )
        return _webhook_process_result(
            reason='purchase_order_not_exact_match',
            message=f'采购订单未精确匹配：webhook 单号 {number} 与查询结果单号不一致',
        )
    return _webhook_process_result(
        reason='recorded_only',
        message='仅记录，未处理：采购订单未满足辅料缓存条件',
    )


def _webhook_result_to_finish(result):
    if isinstance(result, dict):
        success = not bool(result.get('failed'))
        message = str(result.get('message') or result.get('reason') or '')
        return success, message
    ok = bool(result)
    return True, '' if ok else 'recorded only'


def _webhook_stale_cutoff(older_than_minutes=30):
    try:
        minutes = int(older_than_minutes or 30)
    except Exception:
        minutes = 30
    minutes = max(1, min(minutes, 24 * 60))
    return minutes, (datetime.now() - timedelta(minutes=minutes)).strftime('%Y-%m-%d %H:%M:%S')


def _webhook_stale_processing_rows(conn, older_than_minutes=30, limit=100):
    minutes, cutoff = _webhook_stale_cutoff(older_than_minutes)
    try:
        limit = max(1, min(int(limit or 100), 500))
    except Exception:
        limit = 100
    rows = conn.execute('''
        SELECT id, account, biz_type, resource, bill_no, action, status, attempts, error, created_at, processed_at
        FROM webhook_events
        WHERE status = 'processing'
          AND (
            (processed_at IS NOT NULL AND processed_at != '' AND processed_at <= ?)
            OR ((processed_at IS NULL OR processed_at = '') AND created_at <= ?)
          )
        ORDER BY id ASC
        LIMIT ?
    ''', (cutoff, cutoff, limit)).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        err = str(item.get('error') or '').lower()
        item['recover_to_status'] = 'retry_pending' if 'recovered for retry' in err else 'pending'
        result.append(item)
    return result, minutes, cutoff


def _webhook_int_arg(value, default=0, min_value=0, max_value=1000):
    try:
        num = int(value)
    except Exception:
        num = default
    return max(min_value, min(num, max_value))


def _webhook_failure_params(source):
    source = source or {}
    resource = str(source.get('resource') or '').strip()
    error_keyword = str(source.get('error_keyword') or source.get('error') or '').strip()
    order = str(source.get('order') or 'oldest').strip().lower()
    if order not in ('oldest', 'newest'):
        order = 'oldest'
    limit = _webhook_int_arg(source.get('limit'), 20, 1, 500)
    offset = _webhook_int_arg(source.get('offset'), 0, 0, 100000)
    return {
        'resource': resource,
        'error_keyword': error_keyword,
        'order': order,
        'limit': limit,
        'offset': offset,
    }


def _webhook_failed_where(params, recorded_only=False):
    clauses = ["status = 'failed'"]
    values = []
    resource = str((params or {}).get('resource') or '').strip()
    error_keyword = str((params or {}).get('error_keyword') or '').strip()
    if resource:
        clauses.append('resource = ?')
        values.append(resource)
    if error_keyword:
        clauses.append('LOWER(COALESCE(error, "")) LIKE ?')
        values.append(f'%{error_keyword.lower()}%')
    if recorded_only:
        clauses.append('(LOWER(COALESCE(error, "")) LIKE ? OR COALESCE(error, "") LIKE ?)')
        values.extend(['%recorded only%', '%仅记录%'])
    return ' AND '.join(clauses), values


def _webhook_failed_rows(conn, params, recorded_only=False):
    params = _webhook_failure_params(params)
    where, values = _webhook_failed_where(params, recorded_only=recorded_only)
    direction = 'ASC' if params['order'] == 'oldest' else 'DESC'
    rows = conn.execute(f'''
        SELECT id, account, biz_type, resource, bill_no, action, status, attempts, error, created_at, processed_at
        FROM webhook_events
        WHERE {where}
        ORDER BY id {direction}
        LIMIT ? OFFSET ?
    ''', values + [params['limit'], params['offset']]).fetchall()
    return [dict(row) for row in rows], params


def _webhook_failures_snapshot(params):
    params = _webhook_failure_params(params)
    conn = _sales_readonly_conn()
    empty = {
        'total': 0,
        'by_resource': [],
        'by_error': [],
        'items': [],
        'params': params,
    }
    if conn is None:
        return empty
    try:
        where, values = _webhook_failed_where(params)
        total = conn.execute(f'SELECT COUNT(*) AS c FROM webhook_events WHERE {where}', values).fetchone()['c']
        by_resource = [
            dict(row) for row in conn.execute(f'''
                SELECT COALESCE(resource, '') AS resource, COUNT(*) AS count
                FROM webhook_events
                WHERE {where}
                GROUP BY resource
                ORDER BY count DESC, resource
                LIMIT 30
            ''', values).fetchall()
        ]
        by_error = [
            dict(row) for row in conn.execute(f'''
                SELECT COALESCE(error, '') AS error, COUNT(*) AS count
                FROM webhook_events
                WHERE {where}
                GROUP BY error
                ORDER BY count DESC, error
                LIMIT 30
            ''', values).fetchall()
        ]
        items, _ = _webhook_failed_rows(conn, params)
        return {
            'total': int(total or 0),
            'by_resource': by_resource,
            'by_error': by_error,
            'items': items,
            'params': params,
        }
    finally:
        conn.close()


def _flatten_webhook_items(value):
    if value is None:
        return []
    if isinstance(value, list):
        rows = []
        for item in value:
            rows.extend(_flatten_webhook_items(item))
        return rows
    if isinstance(value, dict):
        rows = [value]
        for key in ('data', 'items', 'list', 'detail', 'details'):
            if isinstance(value.get(key), (list, dict)):
                rows.extend(_flatten_webhook_items(value.get(key)))
        return rows
    return []


def _extract_webhook_bill_numbers(payload):
    payload = payload or {}
    numbers = []
    keys = [
        'number', 'billNo', 'billNumber', 'bill_no', 'orderNo', 'orderNumber',
        'sourceBillNo', 'srcBillNo', 'billId', 'id',
    ]
    list_keys = ['numbers', 'billNos', 'billNoList', 'ids', 'idList']
    for item in _flatten_webhook_items(payload):
        for key in keys:
            val = item.get(key)
            if val not in (None, ''):
                numbers.append(str(val).strip())
        for key in list_keys:
            val = item.get(key)
            if isinstance(val, list):
                numbers.extend(str(x).strip() for x in val if x not in (None, ''))
            elif val not in (None, ''):
                numbers.extend(x.strip() for x in str(val).replace(';', ',').split(',') if x.strip())
    return [x for x in dict.fromkeys(numbers) if x]


def _extract_webhook_values(payload, keys):
    values = []
    keys = list(keys or [])
    for item in _flatten_webhook_items(payload or {}):
        for key in keys:
            val = item.get(key)
            if val not in (None, ''):
                values.append(str(val).strip())
    return [x for x in dict.fromkeys(values) if x]


def _looks_like_internal_id(value):
    text = str(value or '').strip()
    return bool(text) and text.isdigit() and len(text) >= 8


def _looks_like_visible_bill_no(value):
    text = str(value or '').strip()
    return any(ch.isalpha() for ch in text) and any(ch.isdigit() for ch in text)


def _webhook_bill_no_kind(value):
    if _looks_like_visible_bill_no(value):
        return 'visible_number'
    if _looks_like_internal_id(value):
        return 'internal_id'
    return 'unknown'


def _norm_account_token(value):
    return str(value or '').strip().lower().replace(' ', '').replace('_', '').replace('-', '')


def _account_aliases_for_cfg(cfg, index):
    name = str(cfg.get('name') or '').strip()
    app_key = str(cfg.get('app_key') or '').strip()
    db_id = str(cfg.get('db_id') or '').strip()
    aliases = {str(index), f'account{index}', f'账套{index}', app_key, db_id, name}
    if index == 1:
        aliases.update({'饰品', '祺航饰品', 'jdy1', 'accountone'})
    elif index == 2:
        aliases.update({'箱包', '祺航箱包', 'jdy2', 'accounttwo'})
    return {_norm_account_token(x) for x in aliases if str(x or '').strip()}


def _account_matches_cfg(account, cfg, index):
    token = _norm_account_token(account)
    return bool(token) and token in _account_aliases_for_cfg(cfg, index)


def _jdy_account_from_payload(payload):
    payload = payload or {}
    cfg1 = _load_jdy_config()
    cfg2 = _load_jdy_config2()
    candidates = [
        payload.get('account'), payload.get('accountName'), payload.get('account_name'),
        payload.get('appKey'), payload.get('app_key'),
        payload.get('accountId'), payload.get('dbId'), payload.get('dbid'),
        payload.get('outerInstanceId'), payload.get('outer_instance_id'),
    ]
    for candidate in candidates:
        if _account_matches_cfg(candidate, cfg2, 2):
            return cfg2.get('name') or '祺航箱包'
        if _account_matches_cfg(candidate, cfg1, 1):
            return cfg1.get('name') or '祺航饰品'
    return ''


def _jdy_resource_from_event(biz_type, payload):
    text = ' '.join([
        str(biz_type or ''),
        str(payload.get('bizType') or ''),
        str(payload.get('resource') or ''),
        str(payload.get('type') or ''),
        str(payload.get('billType') or ''),
        str(payload.get('formId') or ''),
        str(payload.get('entity') or ''),
    ]).lower()
    mapping = [
        ('transfer', ('transfer', 'inv_transfer', '调拨')),
        ('sales', ('sal_bill_outbound', 'delivery', 'saleout', 'sales', '销货单')),
        ('sales_order', ('sal_bill_order', 'saleorder', '销货订单')),
        ('accessory_purchase', ('pur_bill_order', 'purchaseorder', 'purchase_order', '购货订单')),
        ('purchase_inbound', ('pur_bill_inbound', 'purchase/list', '购货单')),
        ('inventory', ('inv_inventory', 'inventory', '库存')),
        ('product', ('bd_material', 'product', '商品')),
        ('supplier', ('bd_supplier', 'supplier', '供应商')),
    ]
    for resource, needles in mapping:
        if any(needle.lower() in text for needle in needles):
            return resource
    return str(payload.get('resource') or biz_type or 'unknown')


def _webhook_event_id_from_payload(payload):
    payload = payload or {}
    for key in ('msgId', 'msg_id', 'eventId', 'event_id', 'id'):
        val = payload.get(key)
        if val not in (None, ''):
            return str(val).strip()
    return ''


def _webhook_payload_hash(payload):
    raw = json.dumps(payload or {}, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def _webhook_duplicate_exists(conn, event_id, payload_hash, account, resource, bill_no):
    if event_id:
        row = conn.execute('''
            SELECT id FROM webhook_events
            WHERE event_id = ?
              AND COALESCE(account, '') = ?
              AND COALESCE(resource, '') = ?
              AND COALESCE(bill_no, '') = ?
            LIMIT 1
        ''', (event_id, account or '', resource or '', bill_no or '')).fetchone()
        if row:
            return True
    if payload_hash:
        row = conn.execute('''
            SELECT id FROM webhook_events
            WHERE payload_hash = ?
              AND COALESCE(account, '') = ?
              AND COALESCE(resource, '') = ?
              AND COALESCE(bill_no, '') = ?
            LIMIT 1
        ''', (payload_hash, account or '', resource or '', bill_no or '')).fetchone()
        return bool(row)
    return False


def _enqueue_jdy_webhook_events(biz_type, payload):
    payload = payload or {}
    account = _jdy_account_from_payload(payload)
    resource = _jdy_resource_from_event(biz_type, payload)
    action = str(payload.get('operation') or payload.get('action') or payload.get('event') or biz_type or '')
    numbers = _extract_webhook_bill_numbers(payload) or ['']
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    event_id = _webhook_event_id_from_payload(payload)
    payload_hash = _webhook_payload_hash(payload)
    payload_json = json.dumps(payload, ensure_ascii=False)
    safe_resource = _is_webhook_safe_auto_resource(resource)
    initial_status = 'pending' if safe_resource else 'done'
    initial_error = '' if safe_resource else f'recorded only unsafe resource: {resource or "unknown"}'
    processed_at = '' if safe_resource else now
    inserted = 0
    with _sales_cache_conn() as conn:
        for number in numbers:
            number = str(number or '').strip()
            if _webhook_duplicate_exists(conn, event_id, payload_hash, account, resource, number):
                continue
            conn.execute('''
                INSERT INTO webhook_events
                (account, biz_type, resource, bill_no, action, status, attempts,
                 payload_json, error, created_at, processed_at, event_id, payload_hash, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                account, str(biz_type or ''), resource, number, action,
                initial_status, payload_json, initial_error, now, processed_at, event_id, payload_hash, now,
            ))
            inserted += 1
        conn.commit()
    message = f'queued {inserted} event(s)' if safe_resource else f'recorded only {inserted} unsafe event(s)'
    _webhook_state.update({'last_event_at': now, 'pending': _webhook_pending_count(), 'message': message})
    if safe_resource:
        _start_webhook_worker()
    return inserted


def _webhook_queue_name(queue='normal'):
    q = str(queue or 'normal').strip().lower()
    return q if q in ('normal', 'retry') else 'normal'


def _webhook_status_for_queue(queue='normal'):
    return 'retry_pending' if _webhook_queue_name(queue) == 'retry' else 'pending'


def _is_webhook_safe_auto_resource(resource):
    resource = str(resource or '').strip().lower()
    return (
        resource in WEBHOOK_SAFE_AUTO_RESOURCES
        or 'sal_bill' in resource
        or 'delivery' in resource
        or 'transfer' in resource
        or 'pur_bill_order' in resource
        or 'purchaseorder' in resource
    )


def _webhook_pending_count(queue='normal'):
    try:
        status = _webhook_status_for_queue(queue)
        with _sales_cache_conn() as conn:
            row = conn.execute("SELECT COUNT(*) AS c FROM webhook_events WHERE status = ?", (status,)).fetchone()
        return int(row['c'] or 0)
    except Exception:
        return 0


def _claim_webhook_event(queue='normal'):
    with _webhook_queue_lock:
        with _sales_cache_conn() as conn:
            resource_filter = str(_webhook_state.get('resource_filter') or '').strip()
            queue_name = _webhook_queue_name(queue)
            status = _webhook_status_for_queue(queue_name)
            processing_mode = _normalize_webhook_processing_mode(_webhook_state.get('processing_mode') or 'manual')
            safe_auto_only = processing_mode == 'auto'
            if resource_filter:
                row = conn.execute('''
                    SELECT * FROM webhook_events
                    WHERE status = ? AND resource = ?
                    ORDER BY id ASC
                    LIMIT 1
                ''', (status, resource_filter)).fetchone()
            elif safe_auto_only:
                row = conn.execute('''
                    SELECT * FROM webhook_events
                    WHERE status = ?
                      AND (
                        resource IN ('sales', 'sales_order', 'sal_bill', 'sal_bill_outbound', 'delivery', 'transfer', 'purchase_order', 'accessory_purchase')
                        OR resource LIKE '%sal_bill%'
                        OR resource LIKE '%delivery%'
                        OR resource LIKE '%transfer%'
                        OR resource LIKE '%pur_bill_order%'
                        OR resource LIKE '%purchaseorder%'
                      )
                    ORDER BY id ASC
                    LIMIT 1
                ''', (status,)).fetchone()
            else:
                row = conn.execute('''
                    SELECT * FROM webhook_events
                    WHERE status = ?
                    ORDER BY id ASC
                    LIMIT 1
                ''', (status,)).fetchone()
            if not row:
                return None
            conn.execute(
                "UPDATE webhook_events SET status = 'processing', attempts = attempts + 1 WHERE id = ?",
                (row['id'],),
            )
            conn.commit()
            event = dict(row)
            event['_webhook_queue'] = queue_name
            return event


def _finish_webhook_event(event_id, success=True, error=''):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    status = 'done' if success else 'failed'
    with _sales_cache_conn() as conn:
        conn.execute(
            "UPDATE webhook_events SET status = ?, error = ?, processed_at = ?, updated_at = ? WHERE id = ?",
            (status, str(error or '')[:500], now, now, event_id),
        )
        conn.commit()
    _webhook_state.update({
        'last_processed_at': now,
        'last_error': '' if success else str(error or '')[:300],
        'pending': _webhook_pending_count(),
        'retry_pending': _webhook_pending_count('retry'),
    })


def _extract_visible_sales_number(payload):
    """从 webhook payload 提取可见销售单编号（非内部 id）。"""
    payload = payload or {}
    visible_keys = ['number', 'billNo', 'billNumber', 'orderNo', 'orderNumber', 'sourceBillNo', 'srcBillNo']
    for item in _flatten_webhook_items(payload):
        for key in visible_keys:
            val = item.get(key)
            if val not in (None, ''):
                return str(val).strip()
    return ''


def _number_looks_like_internal_id(number, payload):
    """判断 bill_no 是否可能是 JDY 内部 data.id 而非可见编号。"""
    payload = payload or {}
    number = str(number or '').strip()
    if not number or len(number) < 6:
        return False
    for item in _flatten_webhook_items(payload):
        raw_id = item.get('id')
        if raw_id is not None and str(raw_id).strip() == number:
            return True
    return False


def _refresh_sales_for_webhook_window(cli, account_name, event):
    """Webhook 仅有内部 id 时，用 updTime 小窗口刷新 sales cache。"""
    number = str(event.get('bill_no') or '').strip()
    payload = json.loads(event.get('payload_json') or '{}')
    if not cli or not account_name:
        return _webhook_process_result(
            reason='sales_window_no_client',
            message='sales window refresh skipped: no JDY client; daily compare covers',
        )
    try:
        created_str = str(event.get('created_at') or '').strip()
        if created_str and len(created_str) >= 16:
            try:
                event_dt = datetime.strptime(created_str[:16], '%Y-%m-%d %H:%M')
            except ValueError:
                event_dt = datetime.now()
        else:
            event_dt = datetime.now()
        window_minutes = max(1, min(int(WEBHOOK_SALES_WINDOW_MINUTES), 30))
        window_start = (event_dt - timedelta(minutes=window_minutes)).strftime('%Y-%m-%d %H:%M:%S')
        window_end = (event_dt + timedelta(minutes=window_minutes)).strftime('%Y-%m-%d %H:%M:%S')

        _log_event('JDY_WEBHOOK', f'sales window refresh: {account_name} updTime=[{window_start} ~ {window_end}] reason=internal_id_only bill_no={number}')
        orders = []
        page = 1
        page_size = max(1, min(int(WEBHOOK_SALES_WINDOW_PAGE_SIZE), 100))
        max_pages = max(1, min(int(WEBHOOK_SALES_WINDOW_MAX_PAGES), 5))
        max_rows = max(1, min(int(WEBHOOK_SALES_WINDOW_MAX_ROWS), 500))
        while page <= max_pages and len(orders) < max_rows:
            result = cli.get_sales_orders(
                page=page, page_size=page_size,
                upd_time_begin=window_start, upd_time_end=window_end,
            )
            batch = result.get('list') or []
            if batch:
                orders.extend(batch[:max(0, max_rows - len(orders))])
            total = result.get('total') or 0
            if not batch or len(batch) < page_size or (total and len(orders) >= total):
                break
            page += 1
            time.sleep(0.5)

        if orders:
            target_ids = _webhook_sales_internal_ids(payload, number)
            matched = [row for row in orders if _sales_row_matches_internal_id(row, target_ids)]
            if matched:
                result = _cache_sales_rows_from_jdy_rows(
                    cli, account_name, matched, include_quantities=True, source='webhook_window_exact'
                )
                return _webhook_process_result(
                    ok=True,
                    cached=bool(result.get('cached')),
                    reason='sales_window_exact_match',
                    message=(
                        f'sales window refresh exact match: {result.get("cached")} order(s) cached; '
                        f'quantities={result.get("quantity_codes")}; internal_id={number}'
                    ),
                )
            cache_limit = max(1, min(int(WEBHOOK_SALES_WINDOW_MAX_CACHE), len(orders)))
            result = _cache_sales_rows_from_jdy_rows(
                cli, account_name, orders[:cache_limit], include_quantities=True, source='webhook_window_no_exact_match'
            )
            return _webhook_process_result(
                ok=True,
                cached=bool(result.get('cached')),
                reason='sales_window_no_exact_match',
                message=(
                    f'sales window refresh no exact match: cached {result.get("cached")} of {len(orders)} returned order(s); '
                    f'quantities={result.get("quantity_codes")}; internal_id={number}; window_no_exact_match'
                ),
            )
        else:
            return _webhook_process_result(
                ok=True,
                reason='sales_window_no_rows',
                message=f'sales window refresh found no rows; webhook only provided internal id (bill_no={number}), daily compare covers',
            )
    except Exception as e:
        err = _short_sync_error(e)
        return _webhook_process_result(
            failed=True,
            reason='sales_window_error',
            message=f'sales window refresh error: {err}; daily compare covers',
        )


def _process_webhook_event(event):
    account = event.get('account') or ''
    resource = (event.get('resource') or '').lower()
    number = (event.get('bill_no') or '').strip()
    if not account:
        payload = json.loads(event.get('payload_json') or '{}')
        account = _jdy_account_from_payload(payload)
    is_sales_resource = resource in ('sales', 'sales_order') or 'sal_bill' in resource or 'delivery' in resource
    is_transfer_resource = resource == 'transfer' or 'transfer' in resource
    is_purchase_resource = (
        resource in ('accessory_purchase', 'purchase_order')
        or 'pur_bill_order' in resource
        or 'purchaseorder' in resource
    )
    is_safe_resource = is_sales_resource or is_transfer_resource or is_purchase_resource
    if not is_safe_resource:
        _log_event('JDY_WEBHOOK', f'recorded only unsafe resource: resource={resource or "unknown"} account={account or "-"} number={number}')
        return _webhook_process_result(
            ok=True,
            reason='recorded_only_unsafe_resource',
            message=f'recorded only unsafe resource: {resource or "unknown"}',
        )
    cli, account_name = _sales_client_for_account(account)
    if not cli:
        raise RuntimeError(f'account not configured: {account or "-"}')
    if not number:
        _log_event('JDY_WEBHOOK', f'event has no bill number, queued only: resource={resource} account={account_name}')
        return _webhook_process_result(
            reason='missing_bill_no',
            message=f'recorded only: missing bill_no, daily compare covers',
        )

    if is_sales_resource:
        payload = json.loads(event.get('payload_json') or '{}')
        biz_type = str(event.get('biz_type') or '').lower()
        action = str(payload.get('operation') or payload.get('action') or event.get('action') or '').lower()

        # delete events: don't try to fetch, daily compare covers
        if action == 'delete' or 'delete' in biz_type:
            _log_event('JDY_WEBHOOK', f'sales delete event recorded: resource={resource} bill_no={number} account={account_name}')
            return _webhook_process_result(
                ok=True,
                reason='sales_delete_recorded',
                message=f'sales delete event recorded: daily compare covers (bill_no={number})',
            )

        # check if we have a visible sales order number (not just internal data.id)
        visible_number = _extract_visible_sales_number(payload) or number
        # if number came from data.id but payload has no visible number, use window refresh
        has_visible = bool(_extract_visible_sales_number(payload))
        if not has_visible and _number_looks_like_internal_id(number, payload):
            return _refresh_sales_for_webhook_window(cli, account_name, event)
        if not has_visible:
            return _refresh_sales_for_webhook_window(cli, account_name, event)

        return _cache_one_sales_order_from_jdy(cli, account_name, visible_number)
    if is_transfer_resource:
        return _cache_one_transfer_order_from_jdy(cli, account_name, number)
    if is_purchase_resource:
        return _cache_one_accessory_purchase_order_from_jdy(cli, account_name, number, event=event)
    return _webhook_process_result(
        ok=True,
        reason='recorded_only_unsafe_resource',
        message=f'recorded only unsafe resource: {resource or "unknown"}',
    )


def _webhook_wait_for_rate_slot():
    global _webhook_last_claim_at
    policy = _webhook_rate_policy()
    _webhook_state.update({
        'mode': policy.get('mode'),
        'current_rate_per_min': policy.get('current_rate_per_min'),
        'rate_reason': policy.get('reason'),
    })
    rate = max(1, min(int(policy.get('current_rate_per_min') or 1), WEBHOOK_RATE_HARD_MAX_PER_MIN))
    interval = 60.0 / rate
    now = time.time()
    wait_for = (_webhook_last_claim_at + interval) - now
    if wait_for > 0:
        _webhook_worker_stop.wait(wait_for)
    _webhook_last_claim_at = time.time()


def _webhook_record_backoff(error):
    msg = str(error or '')
    lowered = msg.lower()
    _webhook_state['consecutive_failures'] = int(_webhook_state.get('consecutive_failures') or 0) + 1
    should_backoff = (
        '429' in lowered
        or 'rate limit' in lowered
        or 'too many requests' in lowered
        or 'frequency' in lowered
        or '频率' in msg
        or int(_webhook_state.get('consecutive_failures') or 0) >= 3
    )
    if should_backoff:
        _webhook_state['backoff_until'] = time.time() + WEBHOOK_ERROR_BACKOFF_SECONDS
        _webhook_state['backoff_reason'] = msg[:160] or 'consecutive webhook processing failures'


def _webhook_worker_loop():
    _webhook_state.update({
        'running': True,
        'worker_thread_name': threading.current_thread().name,
        'worker_thread_id': str(threading.get_ident()),
        'worker_last_exception': '',
    })
    try:
        while not _webhook_worker_stop.is_set():
            heartbeat = _webhook_now_text()
            _webhook_state.update({
                'worker_last_loop_at': heartbeat,
                'worker_last_heartbeat_at': heartbeat,
            })
            try:
                if _webhook_state.get('paused'):
                    processing_mode = str(_webhook_state.get('processing_mode') or 'manual').strip().lower()
                    _webhook_state.update({
                        'running': False,
                        'mode': 'paused' if processing_mode == 'paused' else 'manual',
                        'pending': _webhook_pending_count(),
                        'retry_pending': _webhook_pending_count('retry'),
                        'message': _webhook_state.get('pause_reason') or 'manual processing required',
                    })
                    _webhook_worker_stop.wait(5)
                    continue
                max_items = _webhook_state.get('max_items_remaining')
                if max_items is not None and int(max_items or 0) <= 0:
                    _webhook_state.update({
                        'paused': True,
                        'pause_reason': 'max_items reached',
                        'running': False,
                        'mode': 'paused',
                    })
                    _webhook_worker_stop.wait(5)
                    continue
                processing_mode = str(_webhook_state.get('processing_mode') or 'manual').strip().lower()
                queue_name = 'normal' if processing_mode == 'auto' else _webhook_queue_name(_webhook_state.get('queue') or 'normal')
                _webhook_state.update({'queue': queue_name, 'running': True})
                _webhook_wait_for_rate_slot()
                event = _claim_webhook_event(queue_name)
                if not event:
                    _webhook_state.update({
                        'running': False,
                        'pending': _webhook_pending_count(),
                        'retry_pending': _webhook_pending_count('retry'),
                    })
                    _webhook_worker_stop.wait(5)
                    _webhook_state['running'] = True
                    continue
                try:
                    event_queue = event.get('_webhook_queue') or queue_name
                    _webhook_state.update({
                        'message': f"processing {event_queue} {event.get('resource')} {event.get('bill_no')}",
                        'pending': _webhook_pending_count(),
                        'retry_pending': _webhook_pending_count('retry'),
                        'queue': event_queue,
                    })
                    result = _process_webhook_event(event)
                    success, message = _webhook_result_to_finish(result)
                    _finish_webhook_event(event['id'], success, message)
                    if not success:
                        if event_queue != 'retry':
                            _webhook_record_backoff(message)
                        else:
                            _webhook_state['retry_last_error'] = str(message or '')[:300]
                        _webhook_worker_stop.wait(3)
                        continue
                    _webhook_state['processed'] = int(_webhook_state.get('processed') or 0) + 1
                    if event_queue != 'retry':
                        _webhook_state['consecutive_failures'] = 0
                    if _webhook_state.get('max_items_remaining') is not None:
                        _webhook_state['max_items_remaining'] = max(0, int(_webhook_state.get('max_items_remaining') or 0) - 1)
                    if _webhook_state.get('stop_after_current'):
                        _webhook_state.update({
                            'paused': True,
                            'pause_reason': 'stopped after current item',
                            'stop_after_current': False,
                        })
                except Exception as e:
                    err = _short_sync_error(e)
                    _finish_webhook_event(event['id'], False, err)
                    _log_event('JDY_WEBHOOK_ERROR', f"{event.get('resource')} {event.get('bill_no')}: {err}")
                    _webhook_state.update({
                        'last_error': err,
                        'worker_last_exception': err,
                    })
                    if (event.get('_webhook_queue') or queue_name) != 'retry':
                        _webhook_record_backoff(err)
                    else:
                        _webhook_state['retry_last_error'] = err
                    _webhook_worker_stop.wait(3)
            except Exception as e:
                err = _short_sync_error(e)
                _webhook_state.update({
                    'running': False,
                    'last_error': err,
                    'worker_last_exception': err,
                    'message': f'webhook worker loop error: {err}',
                })
                _log_event('JDY_WEBHOOK_WORKER_ERROR', err)
                queue_name = _webhook_queue_name(_webhook_state.get('queue') or 'normal')
                if queue_name != 'retry':
                    _webhook_record_backoff(err)
                else:
                    _webhook_state['retry_last_error'] = err
                _webhook_worker_stop.wait(3)
    finally:
        _webhook_state.update({
            'running': False,
            'worker_last_heartbeat_at': _webhook_now_text(),
        })


def _start_webhook_worker():
    global _webhook_worker_thread
    if _webhook_worker_thread and _webhook_worker_thread.is_alive():
        return
    _webhook_state.update({
        'worker_thread_name': 'jdy-webhook-worker',
        'worker_thread_id': '',
        'worker_last_exception': '',
        'worker_last_heartbeat_at': _webhook_now_text(),
        'worker_last_loop_at': '',
    })
    _webhook_worker_thread = threading.Thread(target=_webhook_worker_loop, daemon=True, name='jdy-webhook-worker')
    _webhook_worker_thread.start()


def _read_cached_transfer_orders(begin_date='', end_date='', account='all', search=''):
    with _sales_cache_conn() as conn:
        clauses = []
        params = []
        if begin_date:
            clauses.append('date >= ?')
            params.append(begin_date)
        if end_date:
            clauses.append('date <= ?')
            params.append(end_date)
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if search:
            clauses.append('(LOWER(number) LIKE ? OR LOWER(data_json) LIKE ?)')
            kw = f'%{search.lower()}%'
            params.extend([kw, kw])
        where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
        rows = conn.execute(f'''
            SELECT data_json FROM transfer_orders
            {where}
            ORDER BY date DESC, number DESC
        ''', params).fetchall()
    return [json.loads(r['data_json']) for r in rows]


def _read_cached_transfer_detail(order_no, account=''):
    with _sales_cache_conn() as conn:
        if account and account != 'all':
            row = conn.execute('''
                SELECT data_json FROM transfer_details
                WHERE account = ? AND number = ?
            ''', (account, order_no)).fetchone()
        else:
            row = conn.execute('''
                SELECT data_json FROM transfer_details
                WHERE number = ?
                ORDER BY updated_at DESC
                LIMIT 1
            ''', (order_no,)).fetchone()
    return json.loads(row['data_json']) if row else None


def _transfer_cache_stats():
    with _sales_cache_conn() as conn:
        row = conn.execute('''
            SELECT COUNT(*) AS orders_count,
                   MIN(date) AS min_date,
                   MAX(date) AS max_date,
                   MAX(updated_at) AS updated_at
            FROM transfer_orders
        ''').fetchone()
        details = conn.execute('SELECT COUNT(*) AS details_count FROM transfer_details').fetchone()
    return {
        'orders_count': row['orders_count'] or 0,
        'details_count': details['details_count'] or 0,
        'min_date': row['min_date'] or '',
        'max_date': row['max_date'] or '',
        'updated_at': row['updated_at'] or '',
        'db_path': _SALES_CACHE_DB,
    }


def _sales_cache_stats():
    with _sales_cache_conn() as conn:
        row = conn.execute('''
            SELECT COUNT(*) AS orders_count,
                   MIN(date) AS min_date,
                   MAX(date) AS max_date,
                   MAX(updated_at) AS updated_at
            FROM sales_orders
        ''').fetchone()
        details = conn.execute('SELECT COUNT(*) AS details_count FROM sales_details').fetchone()
        qty = conn.execute('''
            SELECT COUNT(*) AS product_quantity_count,
                   MAX(updated_at) AS quantity_updated_at
            FROM sales_product_quantities
        ''').fetchone()
    return {
        'orders_count': row['orders_count'] or 0,
        'details_count': details['details_count'] or 0,
        'product_quantity_count': qty['product_quantity_count'] or 0,
        'min_date': row['min_date'] or '',
        'max_date': row['max_date'] or '',
        'updated_at': row['updated_at'] or '',
        'quantity_updated_at': qty['quantity_updated_at'] or '',
        'db_path': _SALES_CACHE_DB,
    }


_sales_sync_lock = threading.Lock()
_sales_sync_stop = threading.Event()
_sales_sync_thread = None
_sales_sync_state = {
    'running': False,
    'last_started_at': '',
    'last_finished_at': '',
    'last_success': False,
    'last_error': '',
    'last_reason': '',
    'next_run_at': '',
    'phase': '',
    'account': '',
    'current': 0,
    'total': 0,
    'message': '',
    'summary': {},
}
_accessory_sync_lock = threading.Lock()
_accessory_sync_state = {
    'running': False,
    'last_started_at': '',
    'last_finished_at': '',
    'last_success': False,
    'last_error': '',
    'last_reason': '',
    'next_run_at': '',
    'summary': {},
}
_accessory_product_sync_lock = threading.Lock()
_accessory_product_sync_state = {
    'running': False,
    'last_started_at': '',
    'last_finished_at': '',
    'last_success': False,
    'last_error': '',
    'last_reason': '',
    'summary': {},
}
_product_cache_refresh_lock = threading.Lock()
_product_cache_refresh_state = {
    'running': False,
    'last_started_at': '',
    'last_finished_at': '',
    'last_success': False,
    'last_error': '',
    'last_result': {},
    'message': '当前商品刷新为同步执行，暂无后台任务',
}

ACCESSORY_MATERIAL_CATEGORIES = ['绳子', '卡片', '泡壳', '塑料盒子', '塑料桶', '贴纸', '吸塑卡片', '纸盒', '纸箱']
_SALES_ORDER_HASH_VERSION = 'sales-order-v2'

_webhook_queue_lock = threading.Lock()
_webhook_worker_thread = None
_webhook_worker_stop = threading.Event()
_webhook_last_claim_at = 0.0
WEBHOOK_SAFE_AUTO_RESOURCES = {
    'sales',
    'sales_order',
    'sal_bill',
    'sal_bill_outbound',
    'delivery',
    'transfer',
    'purchase_order',
    'accessory_purchase',
}
WEBHOOK_RECORDED_ONLY_RESOURCES = {
    'product',
    'supplier',
    'inventory',
    'purchase_inbound',
    'ar_creditbill',
    'unknown',
}
WEBHOOK_SALES_WINDOW_MINUTES = 10
WEBHOOK_SALES_WINDOW_PAGE_SIZE = 100
WEBHOOK_SALES_WINDOW_MAX_PAGES = 3
WEBHOOK_SALES_WINDOW_MAX_ROWS = 200
WEBHOOK_SALES_WINDOW_MAX_CACHE = 20
_webhook_state = {
    'running': False,
    'paused': True,
    'pause_reason': 'manual processing required',
    'stop_after_current': False,
    'max_items_remaining': None,
    'resource_filter': '',
    'queue': 'normal',
    'mode': 'paused',
    'processing_mode': 'manual',
    'last_event_at': '',
    'last_processed_at': '',
    'last_error': '',
    'retry_last_error': '',
    'backoff_until': 0,
    'backoff_reason': '',
    'consecutive_failures': 0,
    'pending': 0,
    'processed': 0,
    'worker_thread_name': '',
    'worker_thread_id': '',
    'worker_last_heartbeat_at': '',
    'worker_last_loop_at': '',
    'worker_last_exception': '',
    'auto_recovery_last_at': '',
    'auto_recovery_reason': '',
    'message': '',
}
_daily_compare_state = {
    'last_date': '',
    'last_started_at': '',
    'last_finished_at': '',
    'last_error': '',
}


def _normalize_webhook_processing_mode(mode):
    mode = str(mode or 'manual').strip().lower()
    return mode if mode in ('paused', 'manual', 'auto') else 'manual'


def _webhook_runtime_settings_from_state():
    mode = _normalize_webhook_processing_mode(_webhook_state.get('processing_mode') or 'manual')
    paused = bool(_webhook_state.get('paused'))
    return {
        'processing_mode': mode,
        'auto_enabled': bool(mode == 'auto' and not paused),
        'paused': paused,
    }


def _save_webhook_runtime_settings():
    settings = _webhook_runtime_settings_from_state()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _sales_cache_conn() as conn:
        conn.execute('''
            INSERT INTO webhook_runtime_settings
            (id, processing_mode, auto_enabled, paused, updated_at)
            VALUES (1, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                processing_mode = excluded.processing_mode,
                auto_enabled = excluded.auto_enabled,
                paused = excluded.paused,
                updated_at = excluded.updated_at
        ''', (
            settings['processing_mode'],
            1 if settings['auto_enabled'] else 0,
            1 if settings['paused'] else 0,
            now,
        ))
        conn.commit()
    return settings


def _load_webhook_runtime_settings():
    conn = _sales_readonly_conn()
    if conn is None:
        return None
    try:
        try:
            row = conn.execute('''
                SELECT processing_mode, auto_enabled, paused, updated_at
                FROM webhook_runtime_settings
                WHERE id = 1
            ''').fetchone()
        except sqlite3.OperationalError:
            return None
        return dict(row) if row else None
    except Exception as e:
        _log_event('JDY_WEBHOOK_SETTINGS_ERROR', f'load runtime settings failed: {_short_sync_error(e)}')
        return None
    finally:
        conn.close()


def _apply_webhook_runtime_settings(settings):
    if not settings:
        return False
    mode = _normalize_webhook_processing_mode(settings.get('processing_mode') or 'manual')
    paused = bool(int(settings.get('paused') or 0))
    if mode == 'paused':
        paused = True
    if mode == 'manual':
        paused = True
    _webhook_state.update({
        'processing_mode': mode,
        'paused': paused,
        'pause_reason': '' if mode == 'auto' and not paused else ('manual pause' if mode == 'paused' else 'manual processing required'),
        'mode': 'auto' if mode == 'auto' and not paused else ('paused' if mode == 'paused' else 'manual'),
        'message': f'webhook runtime mode restored: {mode}',
    })
    return True


def _restore_webhook_runtime_settings():
    return _apply_webhook_runtime_settings(_load_webhook_runtime_settings())


def _sales_sync_config():
    cfg = _load_ai_config()
    def _first_text(*names):
        for name in names:
            value = str(cfg.get(name) or '').strip()
            if value:
                return value
        return ''

    def _int(name, default, min_value=1, max_value=365):
        try:
            value = int(cfg.get(name, default))
        except Exception:
            value = default
        return max(min_value, min(max_value, value))
    expire_action = str(cfg.get('sales_cache_expire_action') or 'delete').strip().lower()
    if expire_action not in ('delete', 'archive'):
        expire_action = 'delete'
    manual_days = _int(
        'sales_cache_manual_days',
        _int('sales_incremental_lookback_days', 3, 1, 60),
        1,
        60,
    )
    factory_begin = _first_text('sales_factory_purchase_begin_date', 'factory_purchase_begin_date')
    factory_new_begin = _first_text(
        'sales_factory_new_logic_begin_date',
        'factory_new_logic_begin_date',
        'sales_factory_purchase_begin_date',
        'factory_purchase_begin_date',
    )
    accessory_begin = _first_text(
        'accessory_purchase_begin_date',
        'sales_accessory_purchase_begin_date',
        'sales_factory_purchase_begin_date',
        'factory_purchase_begin_date',
    )
    accessory_terms = _first_text('accessory_supplier_terms', 'accessory_supplier_keywords')
    return {
        'enabled': bool(cfg.get('sales_auto_sync_enabled', False)),
        'interval_minutes': _int('sales_auto_sync_interval_minutes', 5, 1, 1440),
        'keep_days': _int('sales_cache_keep_days', 30, 1, 365),
        'lookback_days': manual_days,
        'manual_days': manual_days,
        'daily_compare_time': _first_text('sales_cache_daily_time', 'jdy_daily_compare_time'),
        'expire_action': expire_action,
        'archive_path': str(cfg.get('sales_cache_archive_path') or '').strip(),
        'factory_purchase_begin_date': factory_begin,
        'factory_new_logic_begin_date': factory_new_begin,
        'factory_disabled_suppliers': _first_text('sales_factory_disabled_suppliers', 'factory_disabled_suppliers'),
        'accessory_enabled': bool(cfg.get('accessory_auto_sync_enabled', False)),
        'accessory_interval_minutes': _int('accessory_auto_sync_interval_minutes', 5, 1, 1440),
        'accessory_purchase_begin_date': accessory_begin,
        'accessory_supplier_terms': accessory_terms or '辅料供应商\n辅料',
        'accessory_supplier_cache_hours': _int('accessory_supplier_cache_hours', 24, 1, 720),
    }


def _split_sales_factory_disabled_suppliers(text):
    if isinstance(text, list):
        raw = []
        for item in text:
            raw.extend(str(item or '').splitlines())
    else:
        raw = str(text or '').replace('，', ',').replace('；', ';').splitlines()
    terms = []
    for line in raw:
        for part in str(line).replace(';', ',').split(','):
            val = part.strip()
            if val:
                terms.append(val)
    return [x for x in dict.fromkeys(terms)]


def _sales_factory_purchase_filter_config():
    cfg = _sales_sync_config()
    return {
        'begin_date': cfg.get('factory_purchase_begin_date') or '',
        'new_logic_begin_date': cfg.get('factory_purchase_begin_date') or '',
        'disabled_suppliers': _split_sales_factory_disabled_suppliers(
            cfg.get('factory_disabled_suppliers') or ''
        ),
    }


def _purchase_order_allowed_for_factory(order, filter_conf=None):
    filter_conf = filter_conf or _sales_factory_purchase_filter_config()
    begin_date = filter_conf.get('begin_date') or ''
    if begin_date:
        order_date = str(_first_value(order, ['date', 'billDate', 'createTime', 'checkTime'], ''))[:10]
        if order_date and order_date < begin_date:
            return False
    supplier_no = str(_first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], '')).strip()
    supplier_name = str(_first_value(order, ['supplierName', 'vendorName', 'contactName'], '')).strip()
    supplier_no_l = supplier_no.lower()
    supplier_name_l = supplier_name.lower()
    for term in filter_conf.get('disabled_suppliers') or []:
        t = str(term or '').strip().lower()
        if not t:
            continue
        if t == supplier_no_l or t == supplier_name_l or (supplier_name_l and t in supplier_name_l):
            return False
    return True


def _fetch_factory_purchase_orders(cli, code, filter_conf=None, page_size=100, max_pages=10):
    filter_conf = filter_conf or _sales_factory_purchase_filter_config()
    rows = []
    for page in range(1, max_pages + 1):
        po = cli.get_purchase_order_requests(
            page=page,
            page_size=page_size,
            bill_status=0,
            begin_date=filter_conf.get('begin_date') or '',
            product_number=code,
        )
        batch = po.get('list') or []
        if not batch:
            break
        rows.extend([row for row in batch if _purchase_order_allowed_for_factory(row, filter_conf)])
        total = po.get('total') or 0
        if len(batch) < page_size:
            break
        if total and page * page_size >= int(total):
            break
    return rows


def _fetch_checked_factory_purchase_orders(cli, code, filter_conf=None, page_size=100, max_pages=10):
    filter_conf = filter_conf or _sales_factory_purchase_filter_config()
    rows = []
    for page in range(1, max_pages + 1):
        po = cli.get_purchase_order_requests(
            page=page,
            page_size=page_size,
            check_status=1,
            begin_date=filter_conf.get('begin_date') or '',
            product_number=code,
        )
        batch = po.get('list') or []
        if not batch:
            break
        rows.extend([row for row in batch if _purchase_order_allowed_for_factory(row, filter_conf)])
        total = po.get('total') or 0
        if len(batch) < page_size:
            break
        if total and page * page_size >= int(total):
            break
    return rows


def _sales_order_signature(order):
    keep = {
        'number': order.get('number') or '',
        'date': order.get('date') or '',
        'customerName': order.get('customerName') or '',
        'account': order.get('account') or '',
        'totalQty': _num(order.get('totalQty')),
        'totalAmount': _num(order.get('totalAmount')),
        'checkStatusName': order.get('checkStatusName') or '',
        'entries': [
            {
                'code': e.get('code') or '',
                'name': e.get('name') or '',
                'qty': _num(e.get('qty')),
                'unit': e.get('unit') or '',
                'amount': _num(e.get('amount')),
            }
            for e in (order.get('entries') or [])
        ],
    }
    raw = json.dumps(keep, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def _cached_sales_hash(conn, account, number):
    row = conn.execute('''
        SELECT data_json FROM sales_details
        WHERE account = ? AND number = ?
    ''', (account or '', number or '')).fetchone()
    if not row:
        return ''
    try:
        cached = json.loads(row['data_json'])
        if cached.get('cacheHashVersion') == _SALES_ORDER_HASH_VERSION and cached.get('cacheHash'):
            return cached.get('cacheHash')
        return _sales_order_signature(cached)
    except Exception:
        return ''


def _sales_sources_for_account(account='all'):
    cfg1 = _load_jdy_config();  name1 = cfg1.get('name', '饰品')
    cfg2 = _load_jdy_config2(); name2 = cfg2.get('name', '箱包')
    sources = []
    if account in ('', 'all', name1):
        sources.append((_ensure_jdy_client, name1))
    if account in ('', 'all', name2):
        sources.append((_ensure_jdy_client2, name2))
    return sources


def _collect_sales_quantity_refs(rows):
    refs = set()
    for row in rows:
        try:
            order = json.loads(row['data_json'])
        except Exception:
            continue
        account = order.get('account') or row['account'] or ''
        for entry in (order.get('entries') or []):
            code = str(entry.get('code') or '').strip()
            if code:
                refs.add((account, code))
    return refs


def _cleanup_orphan_sales_quantities(conn):
    rows = conn.execute('SELECT account, data_json FROM sales_details').fetchall()
    refs = _collect_sales_quantity_refs(rows)
    all_rows = conn.execute('SELECT account, code FROM sales_product_quantities').fetchall()
    removed = 0
    for row in all_rows:
        key = (row['account'] or '', row['code'] or '')
        if key not in refs:
            conn.execute('DELETE FROM sales_product_quantities WHERE account = ? AND code = ?', key)
            removed += 1
    return removed


def _cleanup_sales_cache_range(begin_date, end_date, action='delete', archive_path=''):
    action = (action or 'delete').lower()
    archived_file = ''
    with _sales_cache_conn() as conn:
        order_rows = conn.execute('''
            SELECT account, number, date, data_json, updated_at
            FROM sales_orders
            WHERE date >= ? AND date <= ?
            ORDER BY date, number
        ''', (begin_date, end_date)).fetchall()
        detail_rows = conn.execute('''
            SELECT account, number, date, data_json, updated_at
            FROM sales_details
            WHERE date >= ? AND date <= ?
            ORDER BY date, number
        ''', (begin_date, end_date)).fetchall()
        refs = _collect_sales_quantity_refs(detail_rows)
        qty_rows = []
        for account, code in refs:
            row = conn.execute('''
                SELECT account, code, stock_new, stock_transit, stock_local, factory_qty, updated_at, error
                FROM sales_product_quantities
                WHERE account = ? AND code = ?
            ''', (account, code)).fetchone()
            if row:
                qty_rows.append(dict(row))

        if action == 'archive' and archive_path:
            os.makedirs(archive_path, exist_ok=True)
            safe_begin = begin_date.replace('-', '')
            safe_end = end_date.replace('-', '')
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archived_file = os.path.join(
                archive_path,
                f'sales_cache_archive_{safe_begin}_{safe_end}_{stamp}.json'
            )
            payload = {
                'begin_date': begin_date,
                'end_date': end_date,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'sales_orders': [dict(r) for r in order_rows],
                'sales_details': [dict(r) for r in detail_rows],
                'sales_product_quantities': qty_rows,
            }
            with open(archived_file, 'w', encoding='utf-8') as fp:
                json.dump(payload, fp, ensure_ascii=False, indent=2)

        conn.execute('DELETE FROM sales_orders WHERE date >= ? AND date <= ?', (begin_date, end_date))
        conn.execute('DELETE FROM sales_details WHERE date >= ? AND date <= ?', (begin_date, end_date))
        orphan_quantities = _cleanup_orphan_sales_quantities(conn)
        conn.commit()

    return {
        'deleted_orders': len(order_rows),
        'deleted_details': len(detail_rows),
        'deleted_orphan_quantities': orphan_quantities,
        'archived_file': archived_file,
        'stats': _sales_cache_stats(),
    }


def _cleanup_sales_cache_retention(conf):
    today = datetime.now().date()
    keep_start = today - timedelta(days=int(conf.get('keep_days') or 30))
    end_date = (keep_start - timedelta(days=1)).strftime('%Y-%m-%d')
    begin_date = '1900-01-01'
    with _sales_cache_conn() as conn:
        row = conn.execute('SELECT MIN(date) AS min_date FROM sales_orders').fetchone()
    if not row or not row['min_date'] or row['min_date'] > end_date:
        return {'deleted_orders': 0, 'deleted_details': 0, 'deleted_orphan_quantities': 0, 'archived_file': ''}
    return _cleanup_sales_cache_range(
        begin_date,
        end_date,
        conf.get('expire_action') or 'delete',
        conf.get('archive_path') or '',
    )


def _run_sales_incremental_sync(reason='manual', account='all'):
    if not _sales_sync_lock.acquire(blocking=False):
        return {'running': True, 'message': 'sales sync already running'}
    started_at = datetime.now()
    _sales_sync_state.update({
        'running': True,
        'last_started_at': started_at.strftime('%Y-%m-%d %H:%M:%S'),
        'last_finished_at': '',
        'last_success': False,
        'last_error': '',
        'last_reason': reason,
        'phase': 'starting',
        'account': '',
        'current': 0,
        'total': 0,
        'message': '准备同步',
    })
    try:
        conf = _sales_sync_config()
        today = datetime.now().date()
        keep_start = today - timedelta(days=int(conf.get('keep_days') or 30))
        lookback_start = today - timedelta(days=int(conf.get('lookback_days') or 3) - 1)
        begin_date = max(keep_start, lookback_start).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

        total_seen = 0
        total_changed = 0
        total_cached = 0
        errors = []
        _log_event('SALES_SYNC', f'start reason={reason} account={account} begin={begin_date} end={end_date}')
        with _sales_cache_conn() as conn:
            sources = _sales_sources_for_account(account)
            _sales_sync_state.update({'total': len(sources), 'current': 0})
            for idx, (cli_fn, name) in enumerate(sources, 1):
                _sales_sync_state.update({
                    'phase': 'fetching',
                    'account': name,
                    'current': idx,
                    'message': f'{name}: 读取 {begin_date}~{end_date}',
                })
                try:
                    cli = _client_for_sync_account(name, cli_fn)
                    if not cli:
                        err = f'{name}: JDY API not configured'
                        errors.append(err)
                        _log_event('SALES_SYNC_ERROR', err)
                        continue
                    raw_rows = _fetch_live_sales_orders_by_day(cli, begin_date, end_date)
                    orders = [_normalize_sales_order(row, name) for row in raw_rows]
                    total_seen += len(orders)
                    changed_orders = []
                    for order in orders:
                        order_hash = _sales_order_signature(order)
                        order['cacheHash'] = order_hash
                        if _cached_sales_hash(conn, name, order.get('number')) != order_hash:
                            changed_orders.append(order)
                    if changed_orders:
                        _sales_sync_state.update({'phase': 'enriching', 'message': f'{name}: 补全 {len(changed_orders)} 张改动单'})
                        changed_orders = _enrich_sales_orders_batch(
                            cli, changed_orders, include_quantities=True, account=name
                        )
                        for order in changed_orders:
                            order['cacheHash'] = _sales_order_signature(order)
                            order['cacheHashVersion'] = _SALES_ORDER_HASH_VERSION
                            _cache_upsert_sales_order(conn, order)
                            _cache_upsert_sales_detail(conn, order)
                        conn.commit()
                    total_changed += len(changed_orders)
                    total_cached += len(orders)
                    print(f'[SALES SYNC] {name}: seen={len(orders)} changed={len(changed_orders)} window={begin_date}~{end_date}')
                    _log_event('SALES_SYNC', f'{name}: seen={len(orders)} changed={len(changed_orders)} window={begin_date}~{end_date}')
                except Exception as e:
                    err = f'{name}: {e}'
                    errors.append(err)
                    print(f'[SALES SYNC ERROR] {err}')
                    _log_event('SALES_SYNC_ERROR', err)
                time.sleep(1.0)

        _sales_sync_state.update({'phase': 'cleanup', 'message': '清理过期本地缓存'})
        cleanup = _cleanup_sales_cache_retention(conf)
        finished_at = datetime.now()
        summary = {
            'begin_date': begin_date,
            'end_date': end_date,
            'seen': total_seen,
            'changed': total_changed,
            'cached': total_cached,
            'errors': errors,
            'cleanup': cleanup,
            'stats': _sales_cache_stats(),
            'duration_seconds': round((finished_at - started_at).total_seconds(), 2),
        }
        _sales_sync_state.update({
            'running': False,
            'last_finished_at': finished_at.strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': not bool(errors),
            'last_error': '; '.join(errors[:5]),
            'phase': 'done',
            'message': f'完成：扫描 {total_seen} 张，更新 {total_changed} 张',
            'summary': summary,
        })
        _log_event('SALES_SYNC', f'finish seen={total_seen} changed={total_changed} cached={total_cached} errors={errors}')
        return summary
    except Exception as e:
        finished_at = datetime.now()
        _sales_sync_state.update({
            'running': False,
            'last_finished_at': finished_at.strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': False,
            'last_error': str(e),
            'phase': 'error',
            'message': str(e),
            'summary': {},
        })
        raise
    finally:
        _sales_sync_lock.release()


def _daily_compare_due(conf, now):
    daily_time = str(conf.get('daily_compare_time') or '').strip()
    if not daily_time:
        return False
    try:
        hh, mm = daily_time.split(':')[:2]
        target = now.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)
    except Exception:
        return False
    today_key = now.strftime('%Y-%m-%d')
    if _daily_compare_state.get('last_date') == today_key:
        return False
    return now >= target


def _sales_sync_worker():
    last_sales_run_at = None
    last_accessory_run_at = None
    _start_webhook_worker()
    while not _sales_sync_stop.is_set():
        try:
            conf = _sales_sync_config()
            sales_interval_seconds = int(conf.get('interval_minutes') or 5) * 60
            accessory_interval_seconds = int(conf.get('accessory_interval_minutes') or 5) * 60
            now = datetime.now()
            # High-frequency polling is disabled. JDY changes should arrive via webhook;
            # manual sync and daily compare remain available for reconciliation.
            sales_due = False
            accessory_due = False
            if sales_due:
                last_sales_run_at = now
                _run_sales_incremental_sync(reason='auto')
            if _daily_compare_due(conf, now):
                _daily_compare_state.update({
                    'last_date': now.strftime('%Y-%m-%d'),
                    'last_started_at': now.strftime('%Y-%m-%d %H:%M:%S'),
                    'last_error': '',
                })
                try:
                    _run_sales_incremental_sync(reason='daily-compare')
                    _run_accessory_incremental_sync(reason='daily-compare')
                    begin = (now.date() - timedelta(days=int(conf.get('lookback_days') or 3) - 1)).strftime('%Y-%m-%d')
                    end = now.strftime('%Y-%m-%d')
                    _refresh_transfer_cache(begin, end, 'all', '')
                    _daily_compare_state['last_finished_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                except Exception as e:
                    _daily_compare_state.update({
                        'last_error': _short_sync_error(e),
                        'last_finished_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    })
            if accessory_due:
                last_accessory_run_at = now
                _run_accessory_incremental_sync(reason='auto')
                _run_accessory_product_sync(reason='auto')
            if False and conf.get('enabled') and last_sales_run_at:
                next_at = last_sales_run_at + timedelta(seconds=sales_interval_seconds)
                _sales_sync_state['next_run_at'] = next_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                _sales_sync_state['next_run_at'] = ''
            if False and conf.get('accessory_enabled') and last_accessory_run_at:
                next_at = last_accessory_run_at + timedelta(seconds=accessory_interval_seconds)
                _accessory_sync_state['next_run_at'] = next_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                _accessory_sync_state['next_run_at'] = ''
        except Exception as e:
            _sales_sync_state.update({
                'running': False,
                'last_success': False,
                'last_error': str(e),
                'last_finished_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            })
            print(f'[SALES SYNC WORKER ERROR] {e}')
        _sales_sync_stop.wait(10)


def _start_sales_sync_worker():
    global _sales_sync_thread
    if _sales_sync_thread and _sales_sync_thread.is_alive():
        return
    _restore_webhook_runtime_settings()
    _sales_sync_thread = threading.Thread(target=_sales_sync_worker, daemon=True, name='sales-sync-worker')
    _sales_sync_thread.start()


def _num(v, default=0):
    try:
        if v in (None, ''):
            return default
        return float(str(v).replace(',', ''))
    except Exception:
        return default


def _first_value(obj, keys, default=''):
    for key in keys:
        val = obj.get(key)
        if val not in (None, ''):
            return val
    return default


def _accessory_supplier_terms():
    cfg = _load_ai_config()
    raw = str(cfg.get('accessory_supplier_terms') or '辅料供应商\n辅料\n辅材供应商\n辅材').strip()
    terms = []
    for line in raw.replace(';', '\n').replace(',', '\n').splitlines():
        val = line.strip()
        if val:
            terms.append(val)
            if '辅料' in val:
                terms.append(val.replace('辅料', '辅材'))
            if '辅材' in val:
                terms.append(val.replace('辅材', '辅料'))
    return [x for x in dict.fromkeys(terms)]


def _supplier_payload(supplier):
    supplier = dict(supplier or {})
    raw = supplier.get('data_json')
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                merged = dict(parsed)
                merged.update({k: v for k, v in supplier.items() if k != 'data_json'})
                return merged
        except Exception:
            pass
    return supplier


def _supplier_status_parts(supplier, forced_status_code=None):
    supplier = _supplier_payload(supplier)
    raw = forced_status_code if forced_status_code is not None else _first_value(
        supplier,
        ['status_code', 'statusCode', 'status', 'enableStatus', 'forbidStatus', 'disabled'],
        ''
    )
    text = str(raw if raw is not None else '').strip().lower()
    status_name = str(_first_value(supplier, ['status_name', 'statusName', 'statusText', 'enableStatusName'], '') or '').strip()
    if text in ('0', 'enabled', 'enable', 'active', 'true', '启用', '正常'):
        return '0', 'enabled', status_name or '启用'
    if text in ('1', 'disabled', 'disable', 'inactive', 'false', '禁用', '停用'):
        return '1', 'disabled', status_name or '禁用'
    if status_name:
        low_name = status_name.lower()
        if '禁用' in status_name or '停用' in status_name or 'disable' in low_name:
            return '1', 'disabled', status_name
        if '启用' in status_name or '正常' in status_name or 'enable' in low_name:
            return '0', 'enabled', status_name
    return '', 'unknown', status_name


def _supplier_contact_text(supplier):
    supplier = _supplier_payload(supplier)
    return str(_first_value(supplier, ['contat', 'contact', 'contactName', 'linkMan', 'linkman'], '') or '').strip()


def _supplier_phone_text(supplier):
    supplier = _supplier_payload(supplier)
    return str(_first_value(supplier, ['phone', 'mobile', 'tel', 'telephone', 'contactPhone'], '') or '').strip()


def _supplier_accessory_match_info(supplier, order=None):
    supplier = _supplier_payload(supplier or {})
    order = order or {}
    terms = [t.lower() for t in _accessory_supplier_terms() if t]
    result = {
        'is_accessory': False,
        'matched_terms': [],
        'source': '',
        'status': _supplier_status_parts(supplier)[1],
        'override': str(supplier.get('accessory_override') or '').strip().lower(),
    }
    if not terms:
        return result
    override = result['override']
    if override in ('no', 'false', '0', 'n', '否', '不是'):
        result['source'] = 'accessory_override_no'
        return result
    if override in ('yes', 'true', '1', 'y', '是', '辅料', '輔料'):
        result.update({'is_accessory': True, 'matched_terms': ['accessory_override'], 'source': 'accessory_override_yes'})
        return result
    if result['status'] == 'disabled':
        result['source'] = 'disabled_supplier'
        return result

    manual_values = [
        str(supplier.get('manual_category_text') or ''),
        str(supplier.get('manual_tags') or ''),
        str(supplier.get('manual_note') or ''),
    ]
    manual_matches = [term for term in terms if term in '\n'.join(manual_values).lower()]
    if manual_matches:
        result.update({'is_accessory': True, 'matched_terms': manual_matches, 'source': 'manual_fields'})
        return result

    fields = [
        'supplierCategoryName', 'supplierCategory', 'categoryName', 'category',
        'typeName', 'supplierTypeName', 'className', 'groupName',
        'newRecTypeName', 'recTypeName', 'remark', 'description',
    ]
    values = []
    if supplier.get('category_text'):
        values.append(str(supplier.get('category_text') or ''))
    for src in (supplier, order):
        for key in fields:
            val = src.get(key)
            if val not in (None, ''):
                values.append(str(val))
    matches = [term for term in terms if term in '\n'.join(values).lower()]
    if matches:
        result.update({'is_accessory': True, 'matched_terms': matches, 'source': 'category_or_remark'})
        return result

    values.extend([
        str(_first_value(order, ['supplierName', 'vendorName'], '')),
        str(_first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], '')),
        str(_first_value(supplier, ['name', 'supplierName'], '')),
        str(_first_value(supplier, ['number', 'supplierNumber', 'supplierNo'], '')),
    ])
    matches = [term for term in terms if term in '\n'.join(values).lower()]
    if matches:
        result.update({'is_accessory': True, 'matched_terms': matches, 'source': 'name_or_number'})
    return result


def _supplier_matches_accessory(supplier, order=None):
    return bool(_supplier_accessory_match_info(supplier, order).get('is_accessory'))


def _supplier_category_text(supplier):
    supplier = _supplier_payload(supplier)
    fields = [
        'supplierCategoryName', 'supplierCategory', 'categoryName', 'category',
        'typeName', 'supplierTypeName', 'className', 'groupName',
        'newRecTypeName', 'recTypeName', 'remark', 'description',
    ]
    values = []
    for key in fields:
        val = supplier.get(key)
        if val not in (None, ''):
            values.append(str(val))
    return '\n'.join(values)


def _cache_upsert_jdy_supplier(conn, account, supplier, status_code=None, now=None):
    number = str(_first_value(supplier, ['number', 'supplierNumber', 'supplierNo'], '')).strip()
    if not number:
        return
    name = str(_first_value(supplier, ['name', 'supplierName'], '')).strip()
    now = now or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    columns = set(_jdy_supplier_columns(conn))
    if 'status_code' not in columns:
        conn.execute('''
            INSERT INTO jdy_suppliers
            (account, number, name, category_text, data_json, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(account, number) DO UPDATE SET
                name = excluded.name,
                category_text = excluded.category_text,
                data_json = excluded.data_json,
                updated_at = excluded.updated_at
        ''', (
            account or '',
            number,
            name,
            _supplier_category_text(supplier),
            json.dumps(supplier, ensure_ascii=False),
            now,
        ))
        return
    status_code_val, status_val, status_name = _supplier_status_parts(supplier, status_code)
    last_seen_enabled_at = now if status_val == 'enabled' else ''
    disabled_at = '' if status_val == 'enabled' else (now if status_val == 'disabled' else '')
    conn.execute('''
        INSERT INTO jdy_suppliers
        (account, number, name, category_text, data_json, updated_at,
         status_code, status, status_name, contact, phone,
         last_seen_at, last_seen_enabled_at, disabled_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(account, number) DO UPDATE SET
            name = excluded.name,
            category_text = excluded.category_text,
            data_json = excluded.data_json,
            updated_at = excluded.updated_at,
            status_code = excluded.status_code,
            status = excluded.status,
            status_name = excluded.status_name,
            contact = excluded.contact,
            phone = excluded.phone,
            last_seen_at = excluded.last_seen_at,
            last_seen_enabled_at = CASE
                WHEN excluded.last_seen_enabled_at != '' THEN excluded.last_seen_enabled_at
                ELSE jdy_suppliers.last_seen_enabled_at
            END,
            disabled_at = CASE
                WHEN excluded.status = 'enabled' THEN ''
                WHEN excluded.disabled_at != '' THEN excluded.disabled_at
                ELSE jdy_suppliers.disabled_at
            END
    ''', (
        account or '',
        number,
        name,
        _supplier_category_text(supplier),
        json.dumps(supplier, ensure_ascii=False),
        now,
        status_code_val,
        status_val,
        status_name,
        _supplier_contact_text(supplier),
        _supplier_phone_text(supplier),
        now,
        last_seen_enabled_at,
        disabled_at,
    ))


def _cache_mark_jdy_supplier_disabled(conn, account, supplier, now=None):
    number = str(_first_value(supplier, ['number', 'supplierNumber', 'supplierNo'], '')).strip()
    if not number:
        return False
    existing = conn.execute('SELECT number FROM jdy_suppliers WHERE account = ? AND number = ?', (account or '', number)).fetchone()
    if not existing:
        return False
    now = now or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    name = str(_first_value(supplier, ['name', 'supplierName'], '')).strip()
    conn.execute('''
        UPDATE jdy_suppliers
        SET name = COALESCE(NULLIF(?, ''), name),
            category_text = ?,
            data_json = ?,
            updated_at = ?,
            status_code = '1',
            status = 'disabled',
            status_name = ?,
            contact = ?,
            phone = ?,
            last_seen_at = ?,
            disabled_at = COALESCE(NULLIF(disabled_at, ''), ?)
        WHERE account = ? AND number = ?
    ''', (
        name,
        _supplier_category_text(supplier),
        json.dumps(supplier, ensure_ascii=False),
        now,
        _supplier_status_parts(supplier, 1)[2],
        _supplier_contact_text(supplier),
        _supplier_phone_text(supplier),
        now,
        now,
        account or '',
        number,
    ))
    return True


def _read_cached_jdy_supplier(conn, account, number):
    number = str(number or '').strip()
    if not number:
        return None
    row = conn.execute('''
        SELECT * FROM jdy_suppliers
        WHERE account = ? AND number = ?
    ''', (account or '', number)).fetchone()
    if not row:
        return None
    try:
        data = json.loads(row['data_json']) if row['data_json'] else {}
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}
    for key in row.keys():
        if key != 'data_json':
            data[key] = row[key]
    return data


def _supplier_cache_stats(account=''):
    with _sales_cache_conn() as conn:
        clauses = []
        params = []
        if account:
            clauses.append('account = ?')
            params.append(account)
        where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
        row = conn.execute(f'''
            SELECT COUNT(*) AS suppliers_count, MAX(updated_at) AS updated_at
            FROM jdy_suppliers
            {where}
        ''', params).fetchone()
    return {
        'suppliers_count': row['suppliers_count'] or 0,
        'updated_at': row['updated_at'] or '',
    }


def _supplier_cache_fresh(account, max_age_hours):
    stats = _supplier_cache_stats(account)
    if not stats.get('suppliers_count') or not stats.get('updated_at'):
        return False
    try:
        ts = datetime.strptime(stats['updated_at'], '%Y-%m-%d %H:%M:%S')
    except Exception:
        return False
    return datetime.now() - ts < timedelta(hours=int(max_age_hours or 24))


def _refresh_supplier_cache_for_account(cli, account_name, max_age_hours=24):
    if _supplier_cache_fresh(account_name, max_age_hours):
        return {'skipped': True, **_supplier_cache_stats(account_name)}
    page = 1
    page_size = 20
    total_seen = 0
    with _sales_cache_conn() as conn:
        while True:
            result = cli.get_suppliers(page=page, page_size=page_size, status=2)
            rows = result.get('list') or []
            for supplier in rows:
                _cache_upsert_jdy_supplier(conn, account_name, supplier)
            conn.commit()
            total_seen += len(rows)
            total = result.get('total') or total_seen
            if not rows or len(rows) < page_size:
                break
            if total and total_seen >= int(total):
                break
            page += 1
            if page > 200:
                break
    stats = _supplier_cache_stats(account_name)
    stats.update({'skipped': False, 'seen': total_seen})
    return stats


def _supplier_cache_account_specs(account='all'):
    cfg1 = _load_jdy_config()
    cfg2 = _load_jdy_config2()
    name1 = cfg1.get('name') or '饰品'
    name2 = cfg2.get('name') or '箱包'
    raw = str(account or 'all').strip()
    specs = [
        {'idx': '1', 'account': name1, 'label': name1, 'client_factory': _ensure_jdy_client},
        {'idx': '2', 'account': name2, 'label': name2, 'client_factory': _ensure_jdy_client2},
    ]
    if raw.lower() in ('', 'all', '全部'):
        return specs
    selected = []
    for spec in specs:
        if raw in (spec['idx'], spec['account'], spec['label'], f'account{spec["idx"]}'):
            selected.append(spec)
    return selected


def _supplier_cache_json_info(account):
    path = _local_supplier_cache_path(account)
    info = {'path': path, 'exists': os.path.exists(path), 'items_count': 0, 'sync_time': '', 'account': ''}
    if not info['exists']:
        return info
    data = _load_local_json_cached(path)
    items = data.get('items') if isinstance(data, dict) else data
    info['items_count'] = len(items or [])
    if isinstance(data, dict):
        info['sync_time'] = data.get('sync_time') or ''
        info['account'] = data.get('account') or ''
    return info


def _supplier_row_to_dict(row):
    if not row:
        return {}
    data = {}
    try:
        raw = row['data_json']
        if raw:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                data.update(parsed)
    except Exception:
        pass
    try:
        keys = row.keys()
    except Exception:
        keys = []
    for key in keys:
        if key != 'data_json':
            data[key] = row[key]
    return data


def _supplier_cache_status_from_conn(conn):
    columns = _jdy_supplier_columns(conn) if conn else []
    missing = _missing_jdy_supplier_columns(columns) if columns else list(JDY_SUPPLIER_EXTRA_COLUMNS.keys())
    cfg1 = _load_jdy_config()
    cfg2 = _load_jdy_config2()
    accounts = [cfg1.get('name') or '饰品', cfg2.get('name') or '箱包']
    per_account = {}
    if not conn or 'account' not in columns:
        for account in accounts:
            per_account[account] = {
                'total': 0, 'enabled': 0, 'disabled': 0, 'unknown': 0,
                'accessory_matched': 0, 'last_updated_at': '',
                'last_seen_at': '', 'last_seen_enabled_at': '',
            }
        return columns, missing, per_account
    for account in accounts:
        rows = conn.execute('SELECT * FROM jdy_suppliers WHERE account = ?', (account,)).fetchall()
        enabled = disabled = unknown = accessory = 0
        updated_values = []
        seen_values = []
        enabled_seen_values = []
        for row in rows:
            supplier = _supplier_row_to_dict(row)
            status = _supplier_status_parts(supplier)[1]
            if status == 'enabled':
                enabled += 1
            elif status == 'disabled':
                disabled += 1
            else:
                unknown += 1
            if _supplier_matches_accessory(supplier):
                accessory += 1
            if supplier.get('updated_at'):
                updated_values.append(str(supplier.get('updated_at')))
            if supplier.get('last_seen_at'):
                seen_values.append(str(supplier.get('last_seen_at')))
            if supplier.get('last_seen_enabled_at'):
                enabled_seen_values.append(str(supplier.get('last_seen_enabled_at')))
        per_account[account] = {
            'total': len(rows),
            'enabled': enabled,
            'disabled': disabled,
            'unknown': unknown,
            'accessory_matched': accessory,
            'last_updated_at': max(updated_values) if updated_values else '',
            'last_seen_at': max(seen_values) if seen_values else '',
            'last_seen_enabled_at': max(enabled_seen_values) if enabled_seen_values else '',
        }
    return columns, missing, per_account


def _supplier_existing_numbers(conn, account):
    if not conn:
        return set()
    try:
        rows = conn.execute('SELECT number FROM jdy_suppliers WHERE account = ?', (account or '',)).fetchall()
        return {str(row['number']).strip() for row in rows if str(row['number'] or '').strip()}
    except Exception:
        return set()


def _normalize_supplier_cache_request(payload, default_max_pages):
    payload = payload or {}
    account = str(payload.get('account') or 'all').strip() or 'all'
    mode = str(payload.get('mode') or 'enabled_only').strip()
    if mode not in ('enabled_only', 'status_reconcile', 'full'):
        raise ValueError('unsupported supplier refresh mode')
    page_size = max(1, min(int(payload.get('page_size') or 500), 500))
    max_pages = max(1, min(int(payload.get('max_pages') or default_max_pages), 500))
    specs = _supplier_cache_account_specs(account)
    if not specs:
        raise ValueError('unknown account')
    return account, mode, page_size, max_pages, specs


def _supplier_list_result_rows(result):
    rows = result.get('list') or result.get('items') or []
    total = result.get('total') or result.get('records') or result.get('totalsize') or len(rows)
    try:
        total = int(total or 0)
    except Exception:
        total = len(rows)
    return rows, total


def _run_supplier_cache_mode(conn, cli, account_name, mode, page_size, max_pages, apply_changes=False, now=None):
    status_code = 0 if mode == 'enabled_only' else 1
    page = 1
    seen = 0
    summary = {
        'mode': mode,
        'status_code': status_code,
        'pages': [],
        'would_insert': 0,
        'would_update': 0,
        'inserted': 0,
        'updated': 0,
        'would_mark_disabled': 0,
        'marked_disabled': 0,
        'new_disabled_ignored': 0,
        'ignored': 0,
    }
    existing = _supplier_existing_numbers(conn, account_name)
    now = now or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    while page <= max_pages:
        result = cli.get_suppliers(page=page, page_size=page_size, status=status_code)
        rows, total = _supplier_list_result_rows(result)
        page_info = {'page': page, 'rows': len(rows), 'total': total}
        if mode == 'enabled_only':
            for supplier in rows:
                number = str(_first_value(supplier, ['number', 'supplierNumber', 'supplierNo'], '')).strip()
                if not number:
                    summary['ignored'] += 1
                    continue
                if number in existing:
                    summary['would_update'] += 1
                    if apply_changes:
                        _cache_upsert_jdy_supplier(conn, account_name, supplier, status_code=0, now=now)
                        summary['updated'] += 1
                else:
                    summary['would_insert'] += 1
                    if apply_changes:
                        _cache_upsert_jdy_supplier(conn, account_name, supplier, status_code=0, now=now)
                        summary['inserted'] += 1
                        existing.add(number)
        else:
            for supplier in rows:
                number = str(_first_value(supplier, ['number', 'supplierNumber', 'supplierNo'], '')).strip()
                if not number:
                    summary['ignored'] += 1
                    continue
                if number in existing:
                    summary['would_mark_disabled'] += 1
                    if apply_changes and _cache_mark_jdy_supplier_disabled(conn, account_name, supplier, now=now):
                        summary['marked_disabled'] += 1
                else:
                    summary['new_disabled_ignored'] += 1
        summary['pages'].append(page_info)
        seen += len(rows)
        if not rows or len(rows) < page_size:
            break
        if total and seen >= total:
            break
        page += 1
        if apply_changes:
            time.sleep(0.15)
    summary['seen'] = seen
    summary['truncated_by_max_pages'] = page > max_pages
    return summary


def _export_supplier_json_cache(conn, account_name, idx):
    path = _local_supplier_cache_path(account_name)
    rows = conn.execute('SELECT * FROM jdy_suppliers WHERE account = ?', (account_name or '',)).fetchall()
    items = []
    for row in rows:
        supplier = _supplier_row_to_dict(row)
        if _supplier_status_parts(supplier)[1] == 'disabled':
            continue
        item = _supplier_payload(supplier)
        item.setdefault('number', supplier.get('number') or '')
        item.setdefault('name', supplier.get('name') or '')
        items.append(item)
    payload = {
        'items': items,
        'total': len(items),
        'sync_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'account': f'account{idx}',
    }
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    _LOCAL_JSON_CACHE.pop(os.path.abspath(path), None)
    return {'path': path, 'total': len(items), 'account': payload['account'], 'sync_time': payload['sync_time']}


def _run_supplier_cache_refresh(conn, specs, mode, page_size, max_pages, apply_changes=False):
    modes = ['enabled_only', 'status_reconcile'] if mode == 'full' else [mode]
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    results = []
    for spec in specs:
        cli = spec.get('client') or spec['client_factory']()
        account_name = spec['account']
        account_result = {'account': account_name, 'idx': spec['idx'], 'modes': []}
        for submode in modes:
            account_result['modes'].append(_run_supplier_cache_mode(
                conn, cli, account_name, submode, page_size, max_pages,
                apply_changes=apply_changes, now=now
            ))
        results.append(account_result)
    return results


@app.route('/supplier-cache/status')
def supplier_cache_status():
    conn = _sales_readonly_conn()
    try:
        columns, missing, per_account = _supplier_cache_status_from_conn(conn)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    cfg1 = _load_jdy_config()
    cfg2 = _load_jdy_config2()
    name1 = cfg1.get('name') or '饰品'
    name2 = cfg2.get('name') or '箱包'
    return jsonify({
        'success': True,
        'schema_columns': columns,
        'missing_columns': missing,
        'per_account': per_account,
        'json_cache': {
            name1: _supplier_cache_json_info(name1),
            name2: _supplier_cache_json_info(name2),
        },
        'notes': [
            '此接口只读本地供应商缓存，不调用 JDY。',
            '缺失字段会在管理员确认刷新时幂等补齐。',
        ],
    })


@app.route('/supplier-cache/refresh-dry-run', methods=['POST'])
def supplier_cache_refresh_dry_run():
    try:
        _, mode, page_size, max_pages, specs = _normalize_supplier_cache_request(request.get_json(silent=True) or {}, 5)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    conn = _sales_readonly_conn()
    if conn is None:
        conn = sqlite3.connect(':memory:')
        conn.row_factory = sqlite3.Row
        conn.execute('''
            CREATE TABLE jdy_suppliers (
                account TEXT NOT NULL,
                number TEXT NOT NULL,
                name TEXT,
                category_text TEXT,
                data_json TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (account, number)
            )
        ''')
        _ensure_jdy_suppliers_schema(conn)
    try:
        results = _run_supplier_cache_refresh(conn, specs, mode, page_size, max_pages, apply_changes=False)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return jsonify({
        'success': True,
        'dry_run': True,
        'mode': mode,
        'page_size': page_size,
        'max_pages': max_pages,
        'results': results,
        'message': 'dry-run 只查询 JDY 并估算，不写 SQLite，不写 JSON。',
    })


@app.route('/supplier-cache/refresh', methods=['POST'])
def supplier_cache_refresh():
    payload = request.get_json(silent=True) or {}
    if str(payload.get('confirm') or '') != 'REFRESH_SUPPLIERS':
        return jsonify({'success': False, 'error': '缺少确认码 REFRESH_SUPPLIERS，未调用 JDY。'}), 400
    try:
        _, mode, page_size, max_pages, specs = _normalize_supplier_cache_request(payload, 20)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    backup_path = _backup_sales_cache_db('before_supplier_refresh')
    exports = []
    with _sales_cache_conn() as conn:
        _ensure_jdy_suppliers_schema(conn)
        results = _run_supplier_cache_refresh(conn, specs, mode, page_size, max_pages, apply_changes=True)
        conn.commit()
        for spec in specs:
            exports.append(_export_supplier_json_cache(conn, spec['account'], spec['idx']))
    return jsonify({
        'success': True,
        'dry_run': False,
        'mode': mode,
        'page_size': page_size,
        'max_pages': max_pages,
        'backup_path': backup_path,
        'results': results,
        'json_exports': exports,
        'message': '供应商缓存刷新完成；未调用任何 JDY 写接口。',
    })


def _normalize_accessory_purchase_entry(entry):
    return {
        'code': _first_value(entry, ['productNumber', 'productCode', 'number', 'code'], ''),
        'name': _first_value(entry, ['productName', 'name', 'goodsName'], ''),
        'spec': _first_value(entry, ['specification', 'spec', 'model', 'skuName'], ''),
        'barcode': _first_value(entry, ['barCode', 'barcode', 'productBarcode'], ''),
        'qty': _num(_first_value(entry, ['qty', 'mainQty', 'quantity', 'baseQty'], 0)),
        'unit': _first_value(entry, ['unitName', 'unit', 'baseUnitName'], ''),
        'price': _num(_first_value(entry, ['price', 'taxPrice'], 0)),
        'amount': _num(_first_value(entry, ['amount', 'taxAmount', 'totalAmount'], 0)),
        'location': _first_value(entry, ['location', 'locationName', 'warehouseName'], ''),
    }


def _normalize_accessory_purchase_order(row, account, supplier=None):
    entries = [
        _normalize_accessory_purchase_entry(e)
        for e in (row.get('entries') or row.get('items') or row.get('details') or [])
        if isinstance(e, dict)
    ]
    supplier = supplier or {}
    return {
        'number': _first_value(row, ['number', 'billNo', 'id'], ''),
        'date': str(_first_value(row, ['date', 'billDate', 'createTime'], ''))[:10],
        'supplierName': _first_value(row, ['supplierName', 'vendorName'], '') or supplier.get('name') or '',
        'supplierNumber': _first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], '') or supplier.get('number') or '',
        'account': account,
        'totalQty': _num(_first_value(row, ['totalQty', 'qty', 'quantity', 'totalQuantity'], 0)),
        'totalAmount': _num(_first_value(row, ['totalAmount', 'amount', 'taxAmount'], 0)),
        'billStatus': _first_value(row, ['billStatus', 'status'], ''),
        'billStatusName': _first_value(row, ['billStatusName', 'statusName'], ''),
        'checkStatus': row.get('checkStatus'),
        'description': _first_value(row, ['description', 'remark'], ''),
        'entries': entries,
        '_product_codes': [e.get('code') for e in entries if e.get('code')],
        '_raw': row,
        '_supplier': supplier,
    }


def _cache_upsert_accessory_purchase_order(conn, order):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT OR REPLACE INTO accessory_purchase_orders
        (account, number, date, supplier_name, supplier_number, total_qty,
         total_amount, bill_status, bill_status_name, entries_count, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        order.get('account') or '',
        order.get('number') or '',
        order.get('date') or '',
        order.get('supplierName') or '',
        order.get('supplierNumber') or '',
        _num(order.get('totalQty')),
        _num(order.get('totalAmount')),
        str(order.get('billStatus') or ''),
        str(order.get('billStatusName') or ''),
        len(order.get('entries') or []),
        json.dumps(order, ensure_ascii=False),
        now,
    ))


def _accessory_purchase_order_projection(order):
    return {
        'number': order.get('number') or '',
        'date': order.get('date') or '',
        'supplierName': order.get('supplierName') or '',
        'supplierNumber': order.get('supplierNumber') or '',
        'account': order.get('account') or '',
        'totalQty': order.get('totalQty') or 0,
        'totalAmount': order.get('totalAmount') or 0,
        'billStatusName': order.get('billStatusName') or '',
        'entriesCount': len(order.get('entries') or []),
        'productCodes': order.get('_product_codes') or [],
    }


def _read_cached_accessory_purchase_orders(account='all', search=''):
    with _sales_cache_conn() as conn:
        clauses = []
        params = []
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if search:
            clauses.append('(LOWER(number) LIKE ? OR LOWER(supplier_name) LIKE ? OR LOWER(supplier_number) LIKE ? OR LOWER(data_json) LIKE ?)')
            kw = f'%{search.lower()}%'
            params.extend([kw, kw, kw, kw])
        where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
        rows = conn.execute(f'''
            SELECT data_json FROM accessory_purchase_orders
            {where}
            ORDER BY date DESC, number DESC
        ''', params).fetchall()
    return [_accessory_purchase_order_projection(json.loads(r['data_json'])) for r in rows]


def _read_cached_accessory_purchase_order_detail(order_no, account=''):
    with _sales_cache_conn() as conn:
        if account and account != 'all':
            row = conn.execute('''
                SELECT data_json FROM accessory_purchase_orders
                WHERE number = ? AND account = ?
                LIMIT 1
            ''', (order_no, account)).fetchone()
        else:
            row = conn.execute('''
                SELECT data_json FROM accessory_purchase_orders
                WHERE number = ?
                ORDER BY updated_at DESC
                LIMIT 1
            ''', (order_no,)).fetchone()
    return json.loads(row['data_json']) if row else None


def _enrich_accessory_purchase_order(order):
    entries = order.get('entries') or []
    codes = [str(x.get('code') or '').strip() for x in entries if x.get('code')]
    product_map = {}
    needs_product = any(not (e.get('imageUrl') or e.get('barcode')) for e in entries)
    if needs_product and codes:
        cli, _ = _sales_client_for_account(order.get('account') or '')
        if cli:
            product_map = _product_map_by_codes(cli, codes)

    for entry in entries:
        code = str(entry.get('code') or '').strip()
        product = product_map.get(code) or {}
        if not product:
            continue
        entry['name'] = entry.get('name') or product.get('name') or product.get('productName') or ''
        entry['spec'] = product.get('spec') or entry.get('spec') or ''
        entry['barcode'] = product.get('barcode') or entry.get('barcode') or ''
        entry['unit'] = product.get('unitName') or entry.get('unit') or ''
        entry['license'] = product.get('proLicense') or entry.get('license') or ''
        imgs = product.get('multiImg') or []
        if imgs and isinstance(imgs, list):
            entry['imageUrl'] = imgs[0].get('url') or entry.get('imageUrl') or ''
    order['_product_codes'] = [e.get('code') for e in entries if e.get('code')]
    return order


def _enrich_accessory_purchase_orders_batch(cli, orders):
    codes = []
    for order in orders or []:
        for entry in (order.get('entries') or []):
            code = str(entry.get('code') or '').strip()
            if code:
                codes.append(code)
    product_map = _product_map_by_codes(cli, list(dict.fromkeys(codes))) if codes else {}

    for order in orders or []:
        for entry in (order.get('entries') or []):
            code = str(entry.get('code') or '').strip()
            product = product_map.get(code) or {}
            if not product:
                continue
            entry['name'] = entry.get('name') or product.get('name') or product.get('productName') or ''
            entry['spec'] = product.get('spec') or entry.get('spec') or ''
            entry['barcode'] = product.get('barcode') or entry.get('barcode') or ''
            entry['unit'] = product.get('unitName') or entry.get('unit') or ''
            entry['license'] = product.get('proLicense') or entry.get('license') or ''
            imgs = product.get('multiImg') or []
            if imgs and isinstance(imgs, list):
                entry['imageUrl'] = imgs[0].get('url') or entry.get('imageUrl') or ''
        order['_product_codes'] = [e.get('code') for e in (order.get('entries') or []) if e.get('code')]
    return orders


def _normalize_accessory_product(product, account):
    imgs = product.get('multiImg') or []
    image_url = ''
    if imgs and isinstance(imgs, list):
        image_url = imgs[0].get('url') or ''
    return {
        'account': account,
        'code': product.get('productNumber') or product.get('number') or '',
        'name': product.get('productName') or product.get('name') or '',
        'spec': product.get('spec') or product.get('specification') or '',
        'barcode': product.get('barcode') or product.get('barCode') or '',
        'category': product.get('categoryName') or product.get('category') or '',
        'unit': product.get('unitName') or product.get('unit') or '',
        'imageUrl': image_url,
        'license': product.get('proLicense') or '',
    }


def _fetch_accessory_products(cli, account, category='耗材', search='', page_size=100):
    category_kw = str(category or '').strip()
    search_kw = str(search or '').strip()

    def _read(use_category_filter=True, max_pages=20):
        page = 1
        rows = []
        total = None
        while True:
            result = cli.get_products(
                page=page,
                page_size=page_size,
                product_number=search_kw,
                category_name=(category_kw if (use_category_filter and not search_kw) else ''),
            )
            batch = result.get('list') or []
            for item in batch:
                norm = _normalize_accessory_product(item, account)
                if category_kw and not search_kw and category_kw not in str(norm.get('category') or ''):
                    continue
                if search_kw and search_kw.lower() not in str(norm.get('code') or '').lower():
                    continue
                rows.append(norm)
            total = result.get('total') or total
            if not batch or len(batch) < page_size:
                break
            if total and page * page_size >= int(total):
                break
            page += 1
            if page > max_pages:
                break
        return rows

    rows = _read(use_category_filter=True, max_pages=20)
    if category_kw and not search_kw and not rows:
        rows = _read(use_category_filter=False, max_pages=5)
    return rows


def _accessory_material_stock_from_inventory(cli, code):
    buckets = {'厂房物料': 0, '新村物料': 0}
    try:
        inv = cli.get_inventory_by_product(code, page_size=100)
        for row in (inv.get('list') or []):
            if str(row.get('productNumber') or '').strip() != str(code).strip():
                continue
            name = str(row.get('locationName') or row.get('location') or '').strip()
            qty = _num(row.get('qty') or row.get('quantity') or row.get('availableQty') or 0)
            if '厂房物料' in name or ('厂房' in name and '物料' in name):
                buckets['厂房物料'] += qty
            elif '新村物料' in name or ('新村' in name and '物料' in name):
                buckets['新村物料'] += qty
    except Exception as e:
        print(f'[ACCESSORY PRODUCT] inventory failed {code}: {_short_sync_error(e)}')
    return buckets


def _accessory_material_stock_from_product(product):
    buckets = {'厂房物料': 0, '新村物料': 0}
    for prop in (product.get('propertys') or []):
        name = str(prop.get('locationName') or prop.get('location') or '').strip()
        qty = _num(prop.get('quantity') or prop.get('qty') or prop.get('availableQty') or 0)
        if '厂房物料' in name or ('厂房' in name and '物料' in name):
            buckets['厂房物料'] += qty
        elif '新村物料' in name or ('新村' in name and '物料' in name):
            buckets['新村物料'] += qty
    return buckets


def _refresh_accessory_product_stocks_from_inventory(cli, account, conn):
    product_rows = conn.execute(
        'SELECT code, data_json FROM accessory_products WHERE account = ?',
        (account,),
    ).fetchall()
    product_map = {}
    for row in product_rows:
        code = row['code'] or ''
        if not code:
            continue
        try:
            product_map[code] = json.loads(row['data_json'])
        except Exception:
            product_map[code] = {'account': account, 'code': code}
    if not product_map:
        return {'scanned_inventory': 0, 'updated_stock': 0}

    stock_map = {code: {'厂房物料': 0, '新村物料': 0} for code in product_map}
    page = 1
    page_size = 500
    scanned = 0
    while True:
        result = cli._request(
            'GET',
            '/jdyscm/inventory/list',
            body=None,
            query=cli._api_query({'page': page, 'pageSize': page_size}),
            timeout=120,
        )
        rows = result.get('items') or result.get('list') or []
        scanned += len(rows)
        for item in rows:
            code = str(item.get('productNumber') or '').strip()
            if code not in stock_map:
                continue
            name = str(item.get('locationName') or item.get('location') or '').strip()
            qty = _num(item.get('qty') or item.get('quantity') or item.get('availableQty') or 0)
            if '厂房物料' in name or ('厂房' in name and '物料' in name):
                stock_map[code]['厂房物料'] += qty
            elif '新村物料' in name or ('新村' in name and '物料' in name):
                stock_map[code]['新村物料'] += qty
        total = result.get('records') or result.get('totalsize') or 0
        if not rows or len(rows) < page_size:
            break
        if total and page * page_size >= int(total):
            break
        page += 1
        if page > 300:
            break

    updated = 0
    for code, stock in stock_map.items():
        product = product_map.get(code) or {}
        product['stock'] = stock
        _cache_upsert_accessory_product(conn, product)
        updated += 1
    conn.commit()
    return {'scanned_inventory': scanned, 'updated_stock': updated}


def _cache_upsert_accessory_product(conn, product):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('''
        INSERT OR REPLACE INTO accessory_products
        (account, code, name, category, data_json, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (
        product.get('account') or '',
        product.get('code') or '',
        product.get('name') or '',
        product.get('category') or '',
        json.dumps(product, ensure_ascii=False),
        now,
    ))


def _accessory_product_cache_stats():
    with _sales_cache_conn() as conn:
        row = conn.execute('''
            SELECT COUNT(*) AS products_count,
                   MAX(updated_at) AS updated_at
            FROM accessory_products
        ''').fetchone()
    return {
        'products_count': row['products_count'] or 0,
        'updated_at': row['updated_at'] or '',
        'categories': ACCESSORY_MATERIAL_CATEGORIES,
        'db_path': _SALES_CACHE_DB,
    }


def _read_cached_accessory_products(search='', category='', supplier=''):
    search = str(search or '').strip().lower()
    category = str(category or '').strip()
    supplier_account = ''
    supplier_number = ''
    if '|' in str(supplier or ''):
        supplier_account, supplier_number = [x.strip() for x in str(supplier).split('|', 1)]
    elif supplier:
        supplier_number = str(supplier).strip()

    clauses = []
    params = []
    if supplier_account:
        clauses.append('account = ?')
        params.append(supplier_account)
    if category and category != 'all':
        clauses.append('category = ?')
        params.append(category)
    if search:
        clauses.append('(LOWER(code) LIKE ? OR LOWER(name) LIKE ? OR LOWER(data_json) LIKE ?)')
        kw = f'%{search}%'
        params.extend([kw, kw, kw])
    where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
    with _sales_cache_conn() as conn:
        rows = conn.execute(f'''
            SELECT data_json FROM accessory_products
            {where}
            ORDER BY account, category, code
        ''', params).fetchall()

    items = []
    categories = set()
    for row in rows:
        try:
            product = json.loads(row['data_json'])
        except Exception:
            continue
        code = product.get('code') or ''
        factory_qty, factory_orders = _accessory_factory_qty_from_cache(
            product.get('account') or '', code, supplier_number
        )
        if supplier_number and factory_qty <= 0:
            continue
        product['factoryQty'] = factory_qty
        product['factoryOrders'] = factory_orders[:8]
        product['factoryOrderCount'] = len(factory_orders)
        categories.add(product.get('category') or '')
        items.append(product)

    summary = {
        'qty': sum(_num(x.get('factoryQty')) for x in items),
        'factoryStock': sum(_num((x.get('stock') or {}).get('厂房物料')) for x in items),
        'xincunStock': sum(_num((x.get('stock') or {}).get('新村物料')) for x in items),
    }
    return items, summary, sorted(x for x in categories if x)


def _run_accessory_product_sync(reason='manual', account='all'):
    if not _accessory_product_sync_lock.acquire(blocking=False):
        return {'running': True, 'message': 'accessory product sync already running'}
    started_at = datetime.now()
    _accessory_product_sync_state.update({
        'running': True,
        'last_started_at': started_at.strftime('%Y-%m-%d %H:%M:%S'),
        'last_finished_at': '',
        'last_success': False,
        'last_error': '',
        'last_reason': reason,
    })
    try:
        scanned = 0
        cached = 0
        inventory_scanned = 0
        stock_updated = 0
        errors = []
        warnings = []
        category_set = set(ACCESSORY_MATERIAL_CATEGORIES)
        with _sales_cache_conn() as conn:
            for cli_fn, name in _sales_sources_for_account(account):
                try:
                    cli = cli_fn()
                    if not cli:
                        errors.append(f'{name}: 请先配置 JDY API')
                        continue
                    cats = []
                    try:
                        cat_result = cli.get_product_categories(page=1, page_size=300)
                        cats = cat_result.get('list') or []
                    except Exception as e:
                        errors.append(f'{name}: 商品分类读取失败 {_short_sync_error(e)}')
                    target_cats = [
                        c for c in cats
                        if str(c.get('name') or '').strip() in category_set
                    ]
                    if not target_cats:
                        warnings.append(f'{name}: 未找到耗材子分类')
                        continue
                    for cat in target_cats:
                        page = 1
                        page_size = 200
                        while True:
                            result = cli.get_products(
                                page=page,
                                page_size=page_size,
                                category_id=cat.get('id') or '',
                            )
                            rows = result.get('list') or []
                            rows = [
                                r for r in rows
                                if str(r.get('categoryName') or '').strip() == str(cat.get('name') or '').strip()
                            ]
                            scanned += len(rows)
                            for item in rows:
                                product = _normalize_accessory_product(item, name)
                                if product.get('category') not in category_set:
                                    continue
                                product['stock'] = _accessory_material_stock_from_product(item)
                                _cache_upsert_accessory_product(conn, product)
                                cached += 1
                            conn.commit()
                            if len(rows) < page_size:
                                break
                            page += 1
                            if page > 200:
                                break
                        # 精斗云该接口有时 records 仍返回全商品数，以上以本页实际分类命中数判断结束。
                    try:
                        stock_result = _refresh_accessory_product_stocks_from_inventory(cli, name, conn)
                        inventory_scanned += stock_result.get('scanned_inventory') or 0
                        stock_updated += stock_result.get('updated_stock') or 0
                    except Exception as e:
                        warnings.append(f'{name}: 库存同步失败 {_short_sync_error(e)}')
                    print(f'[ACCESSORY PRODUCT SYNC] {name}: scanned={scanned} cached={cached}')
                except Exception as e:
                    err = f'{name}: {_short_sync_error(e)}'
                    errors.append(err)
                    print(f'[ACCESSORY PRODUCT SYNC ERROR] {err}')
            conn.commit()
        finished_at = datetime.now()
        summary = {
            'scanned': scanned,
            'cached': cached,
            'inventory_scanned': inventory_scanned,
            'stock_updated': stock_updated,
            'categories': ACCESSORY_MATERIAL_CATEGORIES,
            'errors': errors,
            'warnings': warnings,
            'stats': _accessory_product_cache_stats(),
            'duration_seconds': round((finished_at - started_at).total_seconds(), 2),
        }
        _accessory_product_sync_state.update({
            'running': False,
            'last_finished_at': finished_at.strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': not bool(errors),
            'last_error': '; '.join(errors[:5]),
            'summary': summary,
        })
        return summary
    except Exception as e:
        _accessory_product_sync_state.update({
            'running': False,
            'last_finished_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': False,
            'last_error': str(e),
            'summary': {},
        })
        raise
    finally:
        _accessory_product_sync_lock.release()


def _accessory_factory_qty_from_cache(account, code, supplier_number=''):
    qty = 0
    orders = []
    with _sales_cache_conn() as conn:
        clauses = ['data_json LIKE ?']
        params = [f'%{code}%']
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if supplier_number:
            clauses.append('supplier_number = ?')
            params.append(supplier_number)
        rows = conn.execute(f'''
            SELECT data_json FROM accessory_purchase_orders
            WHERE {' AND '.join(clauses)}
            ORDER BY date DESC, number DESC
        ''', params).fetchall()
    for row in rows:
        try:
            order = json.loads(row['data_json'])
        except Exception:
            continue
        hit_qty = 0
        for entry in (order.get('entries') or []):
            if str(entry.get('code') or '').strip() == str(code).strip():
                hit_qty += _num(entry.get('qty') or 0)
        if hit_qty:
            qty += hit_qty
            orders.append({
                'number': order.get('number') or '',
                'date': order.get('date') or '',
                'supplierName': order.get('supplierName') or '',
                'supplierNumber': order.get('supplierNumber') or '',
                'qty': hit_qty,
            })
    return qty, orders


def _read_accessory_supplier_options(account='all'):
    with _sales_cache_conn() as conn:
        params = []
        clauses = []
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
        rows = conn.execute(f'''
            SELECT *
            FROM jdy_suppliers
            {where}
            ORDER BY account, number
        ''', params).fetchall()
    result = []
    for row in rows:
        supplier = _supplier_row_to_dict(row)
        if not _supplier_matches_accessory(supplier):
            continue
        result.append({
            'account': row['account'] or '',
            'number': row['number'] or '',
            'name': row['name'] or '',
            'category': row['category_text'] or '',
        })
    return result


def _accessory_purchase_cache_stats():
    with _sales_cache_conn() as conn:
        row = conn.execute('''
            SELECT COUNT(*) AS orders_count,
                   MIN(date) AS min_date,
                   MAX(date) AS max_date,
                   MAX(updated_at) AS updated_at
            FROM accessory_purchase_orders
        ''').fetchone()
    return {
        'orders_count': row['orders_count'] or 0,
        'min_date': row['min_date'] or '',
        'max_date': row['max_date'] or '',
        'updated_at': row['updated_at'] or '',
        'db_path': _SALES_CACHE_DB,
    }


def _normalize_sales_entry(entry):
    qty = _num(_first_value(entry, ['qty', 'quantity', 'baseQty', 'unitQty', 'number'], 0))
    amount = _num(_first_value(entry, ['taxAmount', 'amount', 'totalAmount', 'allAmount'], 0))
    price = _num(_first_value(entry, [
        'taxPrice', 'price', 'unitPrice', 'discountPrice', 'salePrice'
    ], 0))
    if not price and qty:
        price = round(amount / qty, 6)
    wh_name = _first_value(entry, [
        'stockName', 'warehouseName', 'locationName', 'outLocationName',
        'inLocationName', 'storageName'
    ], '')
    warehouses = []
    if wh_name:
        warehouses.append({'name': wh_name, 'qty': qty})
    return {
        'code': _first_value(entry, ['productNumber', 'productCode', 'number', 'code'], ''),
        'name': _first_value(entry, ['productName', 'name', 'goodsName'], ''),
        'spec': _first_value(entry, ['specification', 'spec', 'model', 'skuName'], ''),
        'barcode': _first_value(entry, ['barCode', 'barcode', 'productBarcode'], ''),
        'qty': qty,
        'unit': _first_value(entry, ['unitName', 'unit', 'baseUnitName'], ''),
        'price': price,
        'amount': amount,
        'warehouses': warehouses,
        # 真实采购订单待收数量后续接采购订单接口聚合；这里保持只读占位。
        'purchase': {'ordered': 0, 'received': 0, 'pending': 0},
    }


def _product_map_by_codes(cli, codes):
    """批量读取商品档案，用于补规格、条码、仓库库存。只读。"""
    codes = [c for c in dict.fromkeys([str(c).strip() for c in codes if c])]
    result = {}
    for i in range(0, len(codes), 30):
        batch = codes[i:i + 30]
        try:
            for item in (cli.get_products_by_codes(batch) or []):
                no = str(item.get('productNumber') or '').strip()
                if no:
                    result[no] = item
        except Exception as e:
            print(f'[SALES] 批量商品档案读取失败: {e}')
        missing = [c for c in batch if c not in result]
        for code in missing:
            try:
                item = cli.get_product_by_code(code)
                if item:
                    result[code] = item
            except Exception as e:
                print(f'[SALES] 商品档案读取失败 {code}: {e}')
    return result


def _is_local_stock_location(name):
    name = str(name or '').strip()
    if not name:
        return False
    if any(x in name for x in ['金华', '本地', '义乌', '国内']):
        return True
    if any(x in name for x in ['新大', '在途', '工厂', '厂家', '厂房物料', '新村物料']):
        return False
    return False


def _empty_sales_stock_buckets():
    return {'新大仓库': 0, '在途': 0, '金华/本地': 0, '工厂订单': 0}


def _is_factory_order_location(name):
    name = str(name or '').strip()
    return bool(name and ('工厂订单' in name or '厂家订单' in name))


def _stock_from_product(product):
    """从商品档案 propertys 提取仓库库存：新大仓库、在途、金华/本地。"""
    buckets = _empty_sales_stock_buckets()
    for prop in (product.get('propertys') or []):
        name = str(prop.get('locationName') or prop.get('location') or '').strip()
        qty = _num(prop.get('quantity') or prop.get('qty') or prop.get('availableQty') or 0)
        if not name:
            continue
        if '新大' in name:
            buckets['新大仓库'] += qty
        elif '在途' in name:
            buckets['在途'] += qty
        elif _is_factory_order_location(name):
            buckets['工厂订单'] += qty
        elif _is_local_stock_location(name):
            buckets['金华/本地'] += qty
    return buckets


def _stock_from_inventory(cli, code):
    """从即时库存接口提取新大仓库、在途、金华/本地数量。只读。"""
    buckets = _empty_sales_stock_buckets()
    try:
        inv = cli.get_inventory_by_product(code, page_size=80)
        for row in (inv.get('list') or []):
            if str(row.get('productNumber') or '').strip() != str(code).strip():
                continue
            name = str(row.get('locationName') or '').strip()
            qty = _num(row.get('qty') or 0)
            if '新大' in name:
                buckets['新大仓库'] += qty
            elif '在途' in name:
                buckets['在途'] += qty
            elif _is_factory_order_location(name):
                buckets['工厂订单'] += qty
            elif _is_local_stock_location(name):
                buckets['金华/本地'] += qty
    except Exception as e:
        print(f'[SALES] 即时库存读取失败 {code}: {e}')
    return buckets


def _stock_from_inventory_strict(cli, code):
    """从即时库存接口提取新大仓库、在途、金华/本地数量。失败时抛出异常，避免把失败缓存成 0。"""
    buckets = _empty_sales_stock_buckets()
    inv = cli.get_inventory_by_product(code, page_size=80)
    for row in (inv.get('list') or []):
        if str(row.get('productNumber') or '').strip() != str(code).strip():
            continue
        name = str(row.get('locationName') or '').strip()
        qty = _num(row.get('qty') or 0)
        if '新大' in name:
            buckets['新大仓库'] += qty
        elif '在途' in name:
            buckets['在途'] += qty
        elif _is_factory_order_location(name):
            buckets['工厂订单'] += qty
        elif _is_local_stock_location(name):
            buckets['金华/本地'] += qty
    return buckets


def _sales_order_uses_new_factory_logic(order_date=''):
    begin_date = (_sales_sync_config().get('factory_purchase_begin_date') or '').strip()
    if not begin_date:
        return False
    order_date = str(order_date or '')[:10]
    return bool(order_date and order_date >= begin_date)


def _sales_quantity_logic_key(order_date=''):
    conf = _sales_sync_config()
    begin_date = (conf.get('factory_purchase_begin_date') or '').strip()
    if begin_date and str(order_date or '')[:10] >= begin_date:
        return f'factory-new:{begin_date}'
    return f'factory-warehouse-before:{begin_date or ""}'


def _purchase_entry_order_qty(entry, prefer_actual=False):
    if prefer_actual:
        actual = _num(_first_value(entry, [
            'receiveQty', 'receivedQty', 'actualQty', 'actualReceiveQty',
            'stockInQty', 'inQty', 'realQty', 'associatedQty'
        ], 0))
        if actual:
            return actual
    return _num(_first_value(entry, ['qty', 'mainQty', 'quantity', 'baseQty'], 0))


def _factory_qty_from_purchase(cli, code):
    """
    从未转购货单的购货订单里聚合工厂订单数量。只读。
    """
    ordered = 0
    try:
        filter_conf = _sales_factory_purchase_filter_config()
        for order in _fetch_factory_purchase_orders(cli, code, filter_conf):
            for entry in (order.get('entries') or []):
                if str(entry.get('productNumber') or '').strip() != str(code).strip():
                    continue
                ordered += _num(entry.get('qty') or entry.get('mainQty') or 0)
    except Exception as e:
        print(f'[SALES] 工厂采购数量读取失败 {code}: {e}')
    return ordered


def _factory_qty_from_checked_purchase_strict(cli, code, stock=None):
    ordered = 0
    filter_conf = _sales_factory_purchase_filter_config()
    for order in _fetch_checked_factory_purchase_orders(cli, code, filter_conf):
        linked = bool(_first_value(order, [
            'sourceBillNo', 'sourceNumber', 'linkNumber', 'relationNumber',
            'associatedNumber', 'convertNumber', 'purchaseNumber'
        ], ''))
        for entry in (order.get('entries') or []):
            if str(entry.get('productNumber') or '').strip() != str(code).strip():
                continue
            ordered += _purchase_entry_order_qty(entry, prefer_actual=linked)
    transit = _num((stock or {}).get('在途'))
    local = _num((stock or {}).get('金华/本地'))
    return max(ordered - transit - local, 0)


def _factory_qty_from_purchase_strict(cli, code, stock=None, order_date=''):
    """按切换日期读取工厂数量：切换日前读仓库，切换日及以后读采购订单。"""
    if _sales_order_uses_new_factory_logic(order_date):
        return _factory_qty_from_checked_purchase_strict(cli, code, stock)
    if stock is None:
        stock = _stock_from_inventory_strict(cli, code)
    return _num((stock or {}).get('工厂订单'))


def _apply_sales_quantities(order, quantity_map):
    missing = 0
    for entry in (order.get('entries') or []):
        code = str(entry.get('code') or '').strip()
        q = quantity_map.get(code)
        if not q:
            missing += 1
            entry.setdefault('warehouses', [])
            entry.setdefault('purchase', {'ordered': 0, 'received': 0, 'pending': 0})
            entry['quantityStatus'] = 'missing'
            continue
        stock = q.get('stock') or {}
        factory_qty = _num(q.get('factory_qty'))
        entry['warehouses'] = [
            {'name': '新大仓库', 'qty': _num(stock.get('新大仓库'))},
            {'name': '在途', 'qty': _num(stock.get('在途'))},
            {'name': '金华/本地', 'qty': _num(stock.get('金华/本地'))},
            {'name': '工厂', 'qty': factory_qty},
        ]
        entry['purchase'] = {'ordered': factory_qty, 'received': 0, 'pending': factory_qty}
        entry['quantityStatus'] = 'ok' if not q.get('error') else 'stale'
    order['quantityMissingCount'] = missing
    return order


def _enrich_sales_order(cli, order):
    """给销货单详情补充商品档案、仓库和工厂采购数量。"""
    entries = order.get('entries') or []
    codes = [x.get('code') for x in entries if x.get('code')]
    product_map = _product_map_by_codes(cli, codes)
    factory_cache = {}
    stock_cache = {}
    unique_codes = [c for c in dict.fromkeys(codes) if c]
    if unique_codes:
        workers = min(8, len(unique_codes))
        with ThreadPoolExecutor(max_workers=workers) as exe:
            futures = {exe.submit(_factory_qty_from_purchase, cli, code): code for code in unique_codes}
            for fut in as_completed(futures):
                code = futures[fut]
                try:
                    factory_cache[code] = fut.result()
                except Exception as e:
                    print(f'[SALES] 工厂采购数量并发读取失败 {code}: {e}')
                    factory_cache[code] = 0
        with ThreadPoolExecutor(max_workers=workers) as exe:
            futures = {exe.submit(_stock_from_inventory, cli, code): code for code in unique_codes}
            for fut in as_completed(futures):
                code = futures[fut]
                try:
                    stock_cache[code] = fut.result()
                except Exception as e:
                    print(f'[SALES] 即时库存并发读取失败 {code}: {e}')
                    stock_cache[code] = _empty_sales_stock_buckets()
    for entry in entries:
        code = entry.get('code')
        product = product_map.get(code) or {}
        if product:
            entry['spec'] = product.get('spec') or entry.get('spec') or ''
            entry['barcode'] = product.get('barcode') or entry.get('barcode') or ''
            entry['unit'] = product.get('unitName') or entry.get('unit') or ''
            entry['license'] = product.get('proLicense') or ''
            imgs = product.get('multiImg') or []
            if imgs and isinstance(imgs, list):
                entry['imageUrl'] = imgs[0].get('url') or ''
            stock = stock_cache.get(code) or _stock_from_product(product)
        else:
            stock = stock_cache.get(code) or _empty_sales_stock_buckets()
        factory_qty = factory_cache.get(code, 0)
        if _sales_order_uses_new_factory_logic(order.get('date') or ''):
            try:
                factory_qty = _factory_qty_from_purchase_strict(
                    cli, code, stock=stock, order_date=order.get('date') or ''
                )
            except Exception as e:
                print(f'[SALES] 新工厂逻辑读取失败 {code}: {e}')
        entry['warehouses'] = [
            {'name': '新大仓库', 'qty': stock.get('新大仓库', 0)},
            {'name': '在途', 'qty': stock.get('在途', 0)},
            {'name': '金华/本地', 'qty': stock.get('金华/本地', 0)},
            {'name': '工厂', 'qty': factory_qty},
        ]
        entry['purchase'] = {
            'ordered': factory_qty,
            'received': 0,
            'pending': factory_qty,
        }
    return order


def _normalize_sales_order(row, account):
    entries = row.get('entries') or row.get('details') or row.get('items') or row.get('lines') or []
    norm_entries = [_normalize_sales_entry(e) for e in entries if isinstance(e, dict)]
    total_qty = _num(_first_value(row, ['totalQty', 'qty', 'quantity', 'totalQuantity'], 0))
    if not total_qty and norm_entries:
        total_qty = sum(_num(x.get('qty')) for x in norm_entries)
    return {
        'number': _first_value(row, ['number', 'billNo', 'id'], ''),
        'date': str(_first_value(row, ['date', 'billDate', 'createTime'], ''))[:10],
        'customerName': _first_value(row, ['customerName', 'clientName', 'contactName', 'customer'], ''),
        'account': account,
        'totalQty': total_qty,
        'totalAmount': _num(_first_value(row, ['amount', 'taxAmount', 'totalAmount', 'allAmount', 'discountAmount'], 0)),
        'checkStatusName': _first_value(row, ['checkStatusName', 'statusName', 'status'], ''),
        'entries': norm_entries,
        '_raw': row,
    }


def _query_live_sales_orders(date_str, account, search):
    cfg1 = _load_jdy_config();  name1 = cfg1.get('name', '饰品')
    cfg2 = _load_jdy_config2(); name2 = cfg2.get('name', '箱包')
    sources = []
    if account in ('', 'all', name1):
        sources.append((_ensure_jdy_client, name1))
    if account in ('', 'all', name2):
        sources.append((_ensure_jdy_client2, name2))

    all_items, errors = [], []
    for cli_fn, name in sources:
        try:
            cli = cli_fn()
            if not cli:
                errors.append(f'{name}: 请先配置 JDY API')
                continue
            result = cli.get_sales_orders(page=1, page_size=200, search=search,
                                          begin_date=date_str, end_date=date_str)
            for row in (result.get('list') or []):
                all_items.append(_normalize_sales_order(row, name))
        except Exception as e:
            errors.append(f'{name}: {e}')
    all_items.sort(key=lambda x: x.get('date') or '', reverse=True)
    return all_items, errors


def _fetch_live_sales_orders_all(cli, begin_date, end_date, search=''):
    page = 1
    page_size = 50
    all_rows = []
    total = None
    while True:
        result = cli.get_sales_orders(page=page, page_size=page_size, search=search,
                                      begin_date=begin_date, end_date=end_date)
        rows = result.get('list') or []
        all_rows.extend(rows)
        total = result.get('total') or total
        if not rows or len(rows) < page_size:
            break
        if total and len(all_rows) >= int(total):
            break
        page += 1
        if page > 50:
            break
        time.sleep(0.6)
    return all_rows


def _iter_date_strings(begin_date, end_date):
    try:
        start = datetime.strptime(str(begin_date)[:10], '%Y-%m-%d').date()
        end = datetime.strptime(str(end_date)[:10], '%Y-%m-%d').date()
    except Exception:
        return [begin_date] if begin_date else []
    if end < start:
        start, end = end, start
    days = []
    cur = start
    while cur <= end:
        days.append(cur.strftime('%Y-%m-%d'))
        cur += timedelta(days=1)
    return days


def _fetch_live_sales_orders_by_day(cli, begin_date, end_date, search=''):
    rows = []
    seen_numbers = set()
    for day in _iter_date_strings(begin_date, end_date):
        day_rows = _fetch_live_sales_orders_all(cli, day, day, search=search)
        print(f'[SALES FETCH] {day}: {len(day_rows)} rows')
        _log_event('SALES_SYNC', f'{day}: fetched {len(day_rows)} rows')
        for row in day_rows:
            key = str(row.get('number') or row.get('billNo') or row.get('id') or '')
            if key and key in seen_numbers:
                continue
            if key:
                seen_numbers.add(key)
            rows.append(row)
    return rows


def _short_sync_error(error):
    text = str(error or '').replace('\r', ' ').replace('\n', ' ').strip()
    if 'HTTP 504' in text or 'Gateway Time-out' in text:
        return '精斗云接口 504 超时，请稍后重试或调晚最早日期'
    return text[:240]


def _fetch_live_accessory_purchase_orders_all(cli, account_name, search='', begin_date=''):
    page = 1
    page_size = 20
    all_rows = []
    total = None
    supplier_cache = {}
    while True:
        result = cli.get_purchase_order_requests(
            page=page,
            page_size=page_size,
            search=search,
            bill_status=0,
            begin_date=begin_date,
        )
        rows = result.get('list') or []
        with _sales_cache_conn() as conn:
            for row in rows:
                if begin_date:
                    order_date = str(_first_value(row, ['date', 'billDate', 'createTime'], ''))[:10]
                    if order_date and order_date < begin_date:
                        continue
                supplier_no = str(_first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], '')).strip()
                supplier = supplier_cache.get(supplier_no)
                if supplier is None:
                    supplier = _read_cached_jdy_supplier(conn, account_name, supplier_no) if supplier_no else {}
                if supplier_no and supplier is None:
                    try:
                        supplier = cli.get_supplier_by_number(supplier_no, status=2) or {}
                        if supplier:
                            _cache_upsert_jdy_supplier(conn, account_name, supplier)
                            conn.commit()
                    except Exception as e:
                        print(f'[ACCESSORY] supplier read failed {account_name}/{supplier_no}: {_short_sync_error(e)}')
                        supplier = {}
                supplier_cache[supplier_no] = supplier or {}
                if _supplier_matches_accessory(supplier or {}, row):
                    all_rows.append(_normalize_accessory_purchase_order(row, account_name, supplier or {}))
        total = result.get('total') or total
        if not rows or len(rows) < page_size:
            break
        if total and page * page_size >= int(total):
            break
        page += 1
        if page > 80:
            break
    return all_rows


def _run_accessory_incremental_sync(reason='manual', account='all'):
    if not _accessory_sync_lock.acquire(blocking=False):
        return {'running': True, 'message': 'accessory sync already running'}
    started_at = datetime.now()
    _accessory_sync_state.update({
        'running': True,
        'last_started_at': started_at.strftime('%Y-%m-%d %H:%M:%S'),
        'last_finished_at': '',
        'last_success': False,
        'last_error': '',
        'last_reason': reason,
    })
    try:
        conf = _sales_sync_config()
        begin_date = conf.get('accessory_purchase_begin_date') or ''
        supplier_cache_results = {}
        total_seen = 0
        total_cached = 0
        errors = []
        warnings = []
        seen_keys = set()
        successful_accounts = set()
        with _sales_cache_conn() as conn:
            for cli_fn, name in _sales_sources_for_account(account):
                try:
                    cli = cli_fn()
                    if not cli:
                        errors.append(f'{name}: JDY API not configured')
                        continue
                    try:
                        supplier_cache_results[name] = _refresh_supplier_cache_for_account(
                            cli, name, conf.get('accessory_supplier_cache_hours') or 24
                        )
                    except Exception as e:
                        warn = f'{name}: 供应商缓存刷新失败，已改用订单供应商逐个缓存（{_short_sync_error(e)}）'
                        warnings.append(warn)
                        print(f'[ACCESSORY SYNC WARN] {warn}')
                    rows = _fetch_live_accessory_purchase_orders_all(
                        cli, name, begin_date=begin_date
                    )
                    rows = _enrich_accessory_purchase_orders_batch(cli, rows)
                    total_seen += len(rows)
                    for order in rows:
                        key = (order.get('account') or '', order.get('number') or '')
                        seen_keys.add(key)
                        _cache_upsert_accessory_purchase_order(conn, order)
                    total_cached += len(rows)
                    successful_accounts.add(name)
                    print(f'[ACCESSORY SYNC] {name}: cached={len(rows)} begin={begin_date or "-"}')
                except Exception as e:
                    err = f'{name}: {_short_sync_error(e)}'
                    errors.append(err)
                    print(f'[ACCESSORY SYNC ERROR] {err}')

            if account in ('', 'all'):
                current_rows = conn.execute('SELECT account, number FROM accessory_purchase_orders').fetchall()
                for row in current_rows:
                    key = (row['account'] or '', row['number'] or '')
                    if key[0] in successful_accounts and key not in seen_keys:
                        conn.execute(
                            'DELETE FROM accessory_purchase_orders WHERE account = ? AND number = ?',
                            key,
                        )
            else:
                if account in successful_accounts:
                    current_rows = conn.execute(
                        'SELECT account, number FROM accessory_purchase_orders WHERE account = ?',
                        (account,),
                    ).fetchall()
                    for row in current_rows:
                        key = (row['account'] or '', row['number'] or '')
                        if key not in seen_keys:
                            conn.execute(
                                'DELETE FROM accessory_purchase_orders WHERE account = ? AND number = ?',
                                key,
                            )
            conn.commit()

        finished_at = datetime.now()
        summary = {
            'begin_date': begin_date,
            'seen': total_seen,
            'cached': total_cached,
            'errors': errors,
            'warnings': warnings,
            'supplier_cache': supplier_cache_results,
            'stats': _accessory_purchase_cache_stats(),
            'duration_seconds': round((finished_at - started_at).total_seconds(), 2),
        }
        _accessory_sync_state.update({
            'running': False,
            'last_finished_at': finished_at.strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': not bool(errors),
            'last_error': '; '.join(errors[:5]),
            'summary': summary,
        })
        return summary
    except Exception as e:
        _accessory_sync_state.update({
            'running': False,
            'last_finished_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': False,
            'last_error': str(e),
            'summary': {},
        })
        raise
    finally:
        _accessory_sync_lock.release()


def _fetch_sales_quantity_map(cli, account, codes, order_date=''):
    """批量读取并缓存商品库存/工厂数量。失败的商品沿用旧缓存，不写入假 0。"""
    codes = [c for c in dict.fromkeys([str(c).strip() for c in codes if c])]
    cached = _read_cached_sales_product_quantities(account, codes)
    if not codes:
        return cached, []

    errors = []
    workers = min(16, len(codes))

    def _fetch_one(code):
        stock = _stock_from_inventory_strict(cli, code)
        factory_qty = _factory_qty_from_purchase_strict(cli, code, stock=stock, order_date=order_date)
        return code, {'stock': stock, 'factory_qty': factory_qty, 'error': ''}

    with ThreadPoolExecutor(max_workers=workers) as exe:
        futures = {exe.submit(_fetch_one, code): code for code in codes}
        with _sales_cache_conn() as conn:
            done = 0
            for fut in as_completed(futures):
                code = futures[fut]
                done += 1
                try:
                    _, row = fut.result()
                    cached[code] = row
                    _cache_upsert_sales_product_quantity(
                        conn, account, code, row['stock'], row['factory_qty'], ''
                    )
                except Exception as e:
                    err = str(e)
                    errors.append(f'{code}: {err}')
                    if code in cached:
                        cached[code]['error'] = err
                    print(f'[SALES CACHE] 数量读取失败 {account}/{code}: {err}')
                if done % 100 == 0 or done == len(codes):
                    print(f'[SALES CACHE] {account} 数量进度 {done}/{len(codes)}')
            conn.commit()
    return cached, errors


def _enrich_sales_orders_batch(cli, orders, include_quantities=False, account=''):
    """批量补全一批销货单。

    月度缓存只批量补商品档案，库存和工厂采购数量在打开单据详情时懒加载，
    避免一次同步把每个商品都逐个查库存/采购而卡很久。
    """
    all_codes = []
    for order in orders:
        for entry in (order.get('entries') or []):
            code = entry.get('code')
            if code:
                all_codes.append(code)
    unique_codes = [c for c in dict.fromkeys(all_codes) if c]
    product_map = _product_map_by_codes(cli, unique_codes)
    quantity_map = {}
    if include_quantities and unique_codes:
        quantity_map, _ = _fetch_sales_quantity_map(cli, account, unique_codes)

    for order in orders:
        for entry in (order.get('entries') or []):
            code = entry.get('code')
            product = product_map.get(code) or {}
            if product:
                entry['spec'] = product.get('spec') or entry.get('spec') or ''
                entry['barcode'] = product.get('barcode') or entry.get('barcode') or ''
                entry['unit'] = product.get('unitName') or entry.get('unit') or ''
                entry['license'] = product.get('proLicense') or ''
                imgs = product.get('multiImg') or []
                if imgs and isinstance(imgs, list):
                    entry['imageUrl'] = imgs[0].get('url') or ''
            if include_quantities:
                _apply_sales_quantities({'entries': [entry]}, quantity_map)
            else:
                entry.setdefault('warehouses', [])
                entry.setdefault('purchase', {'ordered': 0, 'received': 0, 'pending': 0})
        if include_quantities:
            order['quantityLogicVersion'] = _sales_quantity_logic_key(order.get('date') or '')
    return orders


def _cache_sales_rows_from_jdy_rows(cli, account, rows, include_quantities=True, source='webhook'):
    raw_rows = [row for row in (rows or []) if isinstance(row, dict)]
    normalized = []
    for row in raw_rows:
        order = _normalize_sales_order(row, account)
        if order.get('number') or order.get('entries'):
            normalized.append(order)
    if not normalized:
        return {'seen': len(raw_rows), 'cached': 0, 'orders': [], 'quantity_codes': 0, 'source': source}
    enriched = _enrich_sales_orders_batch(
        cli, normalized, include_quantities=include_quantities, account=account
    )
    quantity_codes = set()
    with _sales_cache_conn() as conn:
        for order in enriched:
            order['cacheHash'] = _sales_order_signature(order)
            order['cacheHashVersion'] = _SALES_ORDER_HASH_VERSION
            _cache_upsert_sales_order(conn, order)
            _cache_upsert_sales_detail(conn, order)
            if include_quantities:
                for entry in (order.get('entries') or []):
                    code = str(entry.get('code') or '').strip()
                    if code:
                        quantity_codes.add(code)
        conn.commit()
    return {
        'seen': len(raw_rows),
        'cached': len(enriched),
        'orders': enriched,
        'numbers': [order.get('number') or '' for order in enriched],
        'quantity_codes': len(quantity_codes),
        'source': source,
    }


def _webhook_sales_internal_ids(payload, event_number=''):
    values = []
    for item in _flatten_webhook_items(payload or {}):
        for key in ('id', 'billId', 'billID', 'dataId', 'invoiceId'):
            val = item.get(key)
            if val not in (None, ''):
                values.append(str(val).strip())
    if event_number:
        values.append(str(event_number).strip())
    return [x for x in dict.fromkeys(values) if x]


def _sales_row_internal_id_values(row):
    values = []
    if isinstance(row, dict):
        for key in ('id', 'billId', 'billID', 'dataId', 'invoiceId'):
            val = row.get(key)
            if val not in (None, ''):
                values.append(str(val).strip())
        raw = row.get('_raw')
        if isinstance(raw, dict):
            values.extend(_sales_row_internal_id_values(raw))
    return [x for x in dict.fromkeys(values) if x]


def _sales_row_matches_internal_id(row, target_ids):
    targets = {str(x).strip() for x in (target_ids or []) if str(x or '').strip()}
    if not targets:
        return False
    return bool(targets.intersection(_sales_row_internal_id_values(row)))


def _sales_client_for_account(account):
    cfg1 = _load_jdy_config();  name1 = cfg1.get('name', '饰品')
    cfg2 = _load_jdy_config2(); name2 = cfg2.get('name', '箱包')
    if _account_matches_cfg(account, cfg2, 2):
        return _ensure_jdy_client2(), name2
    if _account_matches_cfg(account, cfg1, 1):
        return _ensure_jdy_client(), name1
    return None, account


def _sales_detail_needs_quantity_enrich(order):
    entries = order.get('entries') or []
    if not entries:
        return False
    required = {'新大仓库', '在途', '金华/本地', '工厂'}
    for entry in entries:
        names = {str(w.get('name') or '') for w in (entry.get('warehouses') or [])}
        if not required.issubset(names):
            return True
    return False


def _refresh_cached_sales_detail_quantities(order):
    account = order.get('account') or ''
    cli, account_name = _sales_client_for_account(account)
    if not cli:
        return order, '当前账套未配置 JDY API，无法刷新库存和工厂数量。'
    codes = [
        str(entry.get('code') or '').strip()
        for entry in (order.get('entries') or [])
        if str(entry.get('code') or '').strip()
    ]
    if not codes:
        return order, ''
    quantity_map, errors = _fetch_sales_quantity_map(
        cli,
        account_name or account,
        codes,
        order_date=order.get('date') or '',
    )
    _apply_sales_quantities(order, quantity_map)
    order['quantityLogicVersion'] = _sales_quantity_logic_key(order.get('date') or '')
    with _sales_cache_conn() as conn:
        _cache_upsert_sales_detail(conn, order)
        _cache_upsert_sales_order(conn, order)
        conn.commit()
    return order, '；'.join(errors[:5])


def _refresh_sales_cache(begin_date, end_date, account='all'):
    cfg1 = _load_jdy_config();  name1 = cfg1.get('name', '饰品')
    cfg2 = _load_jdy_config2(); name2 = cfg2.get('name', '箱包')
    sources = []
    if account in ('', 'all', name1):
        sources.append((_ensure_jdy_client, name1))
    if account in ('', 'all', name2):
        sources.append((_ensure_jdy_client2, name2))

    total_orders = 0
    total_details = 0
    errors = []
    with _sales_cache_conn() as conn:
        for cli_fn, name in sources:
            try:
                cli = cli_fn()
                if not cli:
                    errors.append(f'{name}: 请先配置 JDY API')
                    continue
                raw_rows = _fetch_live_sales_orders_by_day(cli, begin_date, end_date)
                orders = [_normalize_sales_order(row, name) for row in raw_rows]
                orders = _enrich_sales_orders_batch(cli, orders, include_quantities=True, account=name)
                for order in orders:
                    _cache_upsert_sales_order(conn, order)
                    _cache_upsert_sales_detail(conn, order)
                conn.commit()
                total_orders += len(orders)
                total_details += len(orders)
                print(f'[SALES CACHE] {name} {begin_date}~{end_date}: {len(orders)} 单')
            except Exception as e:
                errors.append(f'{name}: {e}')
                print(f'[SALES CACHE ERROR] {name}: {e}')
    return {
        'orders': total_orders,
        'details': total_details,
        'errors': errors,
        'stats': _sales_cache_stats(),
    }


def _refresh_sales_quantities_from_cache(begin_date, end_date, account='all'):
    cfg1 = _load_jdy_config();  name1 = cfg1.get('name', '饰品')
    cfg2 = _load_jdy_config2(); name2 = cfg2.get('name', '箱包')
    clients = {}
    if account in ('', 'all', name1):
        clients[name1] = _ensure_jdy_client()
    if account in ('', 'all', name2):
        clients[name2] = _ensure_jdy_client2()

    total_orders = 0
    total_codes = 0
    errors = []
    with _sales_cache_conn() as conn:
        clauses = ['date >= ?', 'date <= ?']
        params = [begin_date, end_date]
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        rows = conn.execute(f'''
            SELECT account, number, data_json
            FROM sales_details
            WHERE {' AND '.join(clauses)}
            ORDER BY date, number
        ''', params).fetchall()

    by_account = {}
    for row in rows:
        by_account.setdefault(row['account'], []).append(json.loads(row['data_json']))

    with _sales_cache_conn() as conn:
        for acct, orders in by_account.items():
            cli = clients.get(acct)
            if not cli:
                errors.append(f'{acct}: 请先配置 JDY API')
                continue
            codes = []
            for order in orders:
                for entry in (order.get('entries') or []):
                    if entry.get('code'):
                        codes.append(entry.get('code'))
            unique_codes = [c for c in dict.fromkeys(codes) if c]
            total_codes += len(unique_codes)
            quantity_map, qty_errors = _fetch_sales_quantity_map(cli, acct, unique_codes)
            if qty_errors:
                errors.extend([f'{acct}/{x}' for x in qty_errors[:50]])
                if len(qty_errors) > 50:
                    errors.append(f'{acct}: 另有 {len(qty_errors) - 50} 个商品数量读取失败')
            for order in orders:
                _apply_sales_quantities(order, quantity_map)
                _cache_upsert_sales_order(conn, order)
                _cache_upsert_sales_detail(conn, order)
            conn.commit()
            total_orders += len(orders)
            print(f'[SALES CACHE] {acct} 本地明细数量补全: {len(orders)} 单, {len(unique_codes)} 款')

    return {
        'orders': total_orders,
        'codes': total_codes,
        'errors': errors,
        'stats': _sales_cache_stats(),
    }


@app.route('/sales-orders', methods=['GET'])
def sales_orders_list():
    """
    销货单接口。source=live 只读精斗云；source=mock 使用本地模拟数据。
    """
    try:
        date_str = request.args.get('date', '') or datetime.now().strftime('%Y-%m-%d')
        account = request.args.get('account', 'all')
        search = request.args.get('search', '').strip().lower()
        source = request.args.get('source', 'cache')
        date_mode = (request.args.get('date_mode') or 'bill_date').strip().lower()
        if date_mode not in ('bill_date', 'updated'):
            date_mode = 'bill_date'
        if date_mode == 'updated':
            source = 'cache'
        errors = []
        if source == 'cache':
            if date_mode == 'updated':
                updated_from = (request.args.get('updated_from') or '').strip()
                updated_to = (request.args.get('updated_to') or '').strip()
                try:
                    limit = int(request.args.get('limit') or request.args.get('page_size') or 100)
                except Exception:
                    limit = 100
                items = _read_cached_sales_orders_by_updated(
                    account=account,
                    search=search,
                    updated_from=updated_from,
                    updated_to=updated_to,
                    limit=limit,
                )
            else:
                items = _read_cached_sales_orders(date_str, account, search)
        elif source == 'live':
            items, errors = _query_live_sales_orders(date_str, account, search)
            if not items and errors:
                return jsonify({'success': False, 'error': '；'.join(errors), 'list': [], 'mock': False})
        else:
            items = _mock_sales_orders(date_str)
            if account and account != 'all':
                items = [x for x in items if x.get('account') == account]
            if search:
                items = [
                    x for x in items
                    if search in (x.get('number', '') + x.get('customerName', '')).lower()
                ]
        return jsonify({
            'success': True,
            'list': items,
            'total': len(items),
            'summary': {
                'qty': sum(float(x.get('totalQty') or 0) for x in items),
                'amount': sum(float(x.get('totalAmount') or 0) for x in items),
            },
            'mock': source == 'mock',
            'cache': source == 'cache',
            'source': source,
            'date_mode': date_mode,
            'local_only': source == 'cache',
            'called_jdy': source == 'live',
            'errors': errors,
            'cache_stats': _sales_cache_stats() if source == 'cache' else None,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/local-products', methods=['GET'])
def local_products():
    """Local product catalog. Reads only jdy_products and local quantity cache."""
    try:
        payload, status = _local_products_response()
        return jsonify(payload), status
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] local_products: {tb}')
        return jsonify({
            'success': False,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'source': 'jdy_products',
            'total': 0,
            'page': _bounded_query_int(request.args.get('page'), 1, 1, 1000000),
            'page_size': _bounded_query_int(request.args.get('page_size'), 50, 1, 200),
            'items': [],
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/customs-products', methods=['GET'])
def customs_products():
    """报关商品资料列表。只读取本地 SQLite，不调用 JDY。"""
    try:
        payload, status = _customs_product_list_response()
        return jsonify(payload), status
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] customs_products: {tb}')
        return jsonify({
            'success': False,
            'local_only': True,
            'called_jdy': False,
            'error': str(e),
            'traceback': tb,
            'items': [],
        }), 500


@app.route('/customs-products/status', methods=['GET'])
def customs_products_status():
    """只读统计本地报关商品资料状态，不调用 JDY。"""
    try:
        payload, status = _customs_product_status_response()
        return jsonify(payload), status
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] customs_products_status: {tb}')
        return jsonify({
            'success': False,
            'local_only': True,
            'called_jdy': False,
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/customs-products/import-dry-run', methods=['POST'])
def customs_products_import_dry_run():
    """只读预览智谱 Excel 导入，不写 SQLite，不修改 Excel。"""
    try:
        body = request.get_json(silent=True) or {}
        requested_excel_path = str(body.get('excel_path') or body.get('path') or '').strip()
        if requested_excel_path and not _is_admin():
            return jsonify({
                'success': False,
                'dry_run': True,
                'would_write': False,
                'called_jdy': False,
                'error': '只有管理员可以指定自定义Excel路径',
            }), 403
        excel_path = _customs_product_source_path(requested_excel_path)
        if not excel_path:
            return jsonify({
                'success': False,
                'dry_run': True,
                'would_write': False,
                'called_jdy': False,
                'error': '未找到报关产品基础资料（智谱）.xlsx，请在设置中配置 Excel 路径或传入 excel_path',
            }), 404
        conn = _sales_readonly_conn()
        try:
            result = _customs_import_preview(excel_path, conn=conn)
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass
        return jsonify(result)
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] customs_products_import_dry_run: {tb}')
        return jsonify({
            'success': False,
            'dry_run': True,
            'would_write': False,
            'called_jdy': False,
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/customs-products/import-confirm', methods=['POST'])
def customs_products_import_confirm():
    """确认从智谱 Excel 导入本地 customs_product_master。不会修改 Excel，不调用 JDY。"""
    body = request.get_json(silent=True) or {}
    if str(body.get('confirm') or '').strip() != 'IMPORT_CUSTOMS_PRODUCTS':
        return jsonify({
            'success': False,
            'dry_run': False,
            'would_write': True,
            'called_jdy': False,
            'error': 'missing confirm token IMPORT_CUSTOMS_PRODUCTS',
        }), 400
    try:
        excel_path = _customs_product_source_path(body.get('excel_path') or body.get('path') or '')
        if not excel_path:
            return jsonify({
                'success': False,
                'dry_run': False,
                'would_write': True,
                'called_jdy': False,
                'error': '未找到报关产品基础资料（智谱）.xlsx，请在设置中配置 Excel 路径或传入 excel_path',
            }), 404
        parsed = _customs_parse_excel(excel_path)
        conn = _sales_cache_conn()
        user = _current_user() or {}
        updated_by = user.get('username') or user.get('name') or 'admin'
        inserted = updated = unchanged = skipped = 0
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conflict_codes = {item['product_code'] for item in parsed['conflict_rows']}
        try:
            for item in parsed['rows']:
                code = item.get('product_code')
                if not code or code in conflict_codes:
                    skipped += 1
                    continue
                action = _customs_upsert_product(
                    conn,
                    item,
                    updated_by=updated_by,
                    change_source='zhipu_excel',
                    now=now,
                )
                if action == 'inserted':
                    inserted += 1
                elif action == 'updated':
                    updated += 1
                elif action == 'unchanged':
                    unchanged += 1
                else:
                    skipped += 1
            conn.commit()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({
            'success': True,
            'dry_run': False,
            'would_write': True,
            'called_jdy': False,
            'written_table': 'customs_product_master',
            'excel_path': excel_path,
            'sheet_name': parsed['sheet_name'],
            'total_rows': parsed['total_rows'],
            'valid_code_count': parsed['valid_code_count'],
            'blank_code_count': parsed['blank_code_count'],
            'duplicate_code_count': parsed['duplicate_code_count'],
            'duplicate_codes': parsed['duplicate_codes'],
            'conflict_count': len(parsed['conflict_rows']),
            'inserted': inserted,
            'updated': updated,
            'unchanged': unchanged,
            'skipped': skipped,
            'field_missing_counts': parsed['missing_field_counts'],
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] customs_products_import_confirm: {tb}')
        return jsonify({
            'success': False,
            'dry_run': False,
            'would_write': True,
            'called_jdy': False,
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/customs-products/<path:product_code>', methods=['PATCH'])
def customs_product_patch(product_code):
    """单商品编辑。只写 customs_product_master 和 history，不改 jdy_products/JDY/Excel。"""
    try:
        code = str(product_code or '').strip()
        if not code:
            return jsonify({'success': False, 'error': '缺少 product_code'}), 400
        body = request.get_json(force=True) or {}
        changes = body.get('customs') if isinstance(body.get('customs'), dict) else body
        if not isinstance(changes, dict):
            return jsonify({'success': False, 'error': '无效请求体'}), 400
        conn = _sales_cache_conn()
        user = _current_user() or {}
        updated_by = user.get('username') or user.get('name') or 'admin'
        try:
            old = conn.execute(
                'SELECT * FROM customs_product_master WHERE product_code = ?',
                (code,),
            ).fetchone()
            item = _customs_product_row_to_dict(old) if old else {'product_code': code}
            for field in CUSTOMS_PRODUCT_MUTABLE_FIELDS:
                if field in changes:
                    val = changes.get(field)
                    if field in CUSTOMS_PRODUCT_NUMERIC_FIELDS:
                        val = _clean_excel_number(val)
                    else:
                        val = _clean_excel_text(val)
                    item[field] = val
            item['product_code'] = code
            item['source'] = 'local_edit'
            action = _customs_upsert_product(
                conn,
                item,
                updated_by=updated_by,
                change_source='local_edit',
            )
            conn.commit()
            row = conn.execute(
                'SELECT * FROM customs_product_master WHERE product_code = ?',
                (code,),
            ).fetchone()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        data = _customs_product_row_to_dict(row)
        return jsonify({
            'success': True,
            'called_jdy': False,
            'modified_jdy': False,
            'modified_excel': False,
            'action': action,
            'product_code': code,
            'customs': data,
            'status': _customs_product_status(data),
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] customs_product_patch: {tb}')
        return jsonify({'success': False, 'called_jdy': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/customs-products/precheck', methods=['POST'])
def customs_products_precheck():
    """只读预检查商品报关资料状态，不改变报关生成结果。"""
    try:
        body = request.get_json(force=True) or {}
        codes = body.get('codes') or body.get('items') or []
        normalized = []
        for item in codes:
            code = item.get('code') if isinstance(item, dict) else item
            code = str(code or '').strip()
            if code:
                normalized.append(code)
        counts = Counter(normalized)
        conn = _sales_readonly_conn()
        customs_map = {}
        try:
            if conn and _cache_sqlite_table_count(conn, 'customs_product_master').get('exists'):
                marks = ','.join(['?'] * len(counts))
                if marks:
                    rows = conn.execute(
                        f'SELECT * FROM customs_product_master WHERE product_code IN ({marks})',
                        list(counts.keys()),
                    ).fetchall()
                    customs_map = {
                        str(row['product_code'] or '').strip(): _customs_product_row_to_dict(row)
                        for row in rows
                    }
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass
        ready, missing, incomplete, duplicate = [], [], [], []
        for code in counts:
            if counts[code] > 1:
                duplicate.append({'product_code': code, 'count': counts[code]})
            data = customs_map.get(code)
            if not data:
                missing.append({'product_code': code})
                continue
            miss = _customs_missing_fields(data)
            if miss:
                incomplete.append({'product_code': code, 'missing_fields': miss})
            else:
                ready.append({'product_code': code})
        return jsonify({
            'success': True,
            'called_jdy': False,
            'ready': ready,
            'missing': missing,
            'incomplete': incomplete,
            'duplicate': duplicate,
            'summary': {
                'input_count': len(normalized),
                'unique_count': len(counts),
                'ready_count': len(ready),
                'missing_count': len(missing),
                'incomplete_count': len(incomplete),
                'duplicate_count': len(duplicate),
            },
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] customs_products_precheck: {tb}')
        return jsonify({'success': False, 'called_jdy': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/sales-summary-products', methods=['GET'])
def sales_summary_products():
    """销售汇总按商品：只读取本地 SQLite/JSON 缓存，不调用 JDY。"""
    try:
        today = datetime.now().date()
        start_default = (today - timedelta(days=59)).strftime('%Y-%m-%d')
        end_default = today.strftime('%Y-%m-%d')
        start_date = (request.args.get('start_date') or request.args.get('begin_date') or start_default)[:10]
        end_date = (request.args.get('end_date') or end_default)[:10]
        account = request.args.get('account', 'all') or 'all'
        q = request.args.get('q', '') or request.args.get('search', '')
        sort = request.args.get('sort', 'qty_desc') or 'qty_desc'
        try:
            limit = int(request.args.get('limit', 500))
        except Exception:
            limit = 500
        try:
            offset = int(request.args.get('offset', 0))
        except Exception:
            offset = 0
        result = _read_cached_sales_summary_products(
            start_date, end_date, account=account, q=q, limit=limit, offset=offset, sort=sort
        )
        return jsonify({
            'success': True,
            'source': 'local_cache',
            'start_date': start_date,
            'end_date': end_date,
            'account': account,
            'total': result['total'],
            'limit': max(1, min(limit, 1000)),
            'offset': max(0, offset),
            'error_count': result['error_count'],
            'items': result['items'],
            'summary': result['summary'],
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] sales_summary_products: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/sales-order/<order_no>', methods=['GET'])
def sales_order_detail(order_no):
    """
    销货单详情接口。source=cache 读本地缓存；source=live 只读精斗云；source=mock 使用本地模拟数据。
    """
    try:
        date_str = request.args.get('date', '') or datetime.now().strftime('%Y-%m-%d')
        account = request.args.get('account', 'all')
        source = request.args.get('source', 'cache')
        allow_live = str(request.args.get('allow_live') or '').lower() in ('1', 'true', 'yes', 'y')
        refresh_quantities = str(request.args.get('refresh_quantities') or '').lower() in ('1', 'true', 'yes', 'y')
        if source == 'cache':
            cached = _read_cached_sales_detail(order_no, account)
            if cached:
                warning = ''
                logic_key = _sales_quantity_logic_key(cached.get('date') or date_str)
                uses_new_logic = _sales_order_uses_new_factory_logic(cached.get('date') or date_str)
                should_refresh_qty = (
                    _sales_detail_needs_quantity_enrich(cached)
                    or (uses_new_logic and cached.get('quantityLogicVersion') != logic_key)
                )
                quantity_enrichment_skipped = bool(should_refresh_qty)
                live_lookup = False
                called_jdy = False
                if should_refresh_qty and allow_live and refresh_quantities:
                    try:
                        cached, warning = _refresh_cached_sales_detail_quantities(cached)
                        quantity_enrichment_skipped = False
                        live_lookup = True
                        called_jdy = True
                    except Exception as e:
                        warning = f'本单库存/工厂数量刷新失败，已显示本地旧缓存：{e}'
                return jsonify({
                    'success': True,
                    'data': cached,
                    'mock': False,
                    'cache': True,
                    'local_only': not live_lookup,
                    'live_lookup': live_lookup,
                    'called_jdy': called_jdy,
                    'would_call_jdy': bool(should_refresh_qty and allow_live and refresh_quantities),
                    'cache_source': 'sales_details',
                    'quantity_enrichment_skipped': quantity_enrichment_skipped,
                    'warning': warning,
                    'message': (
                        '已按显式参数刷新销售单数量字段。'
                        if live_lookup else
                        '本地缓存数量字段可能需要补齐；默认未联网刷新。'
                        if quantity_enrichment_skipped else '读取本地销售单缓存。'
                    ),
                })
            return jsonify({
                'success': False,
                'error': '本地缓存未找到这张销货单，请先同步本月数据或切换实际读取',
                'local_only': True,
                'live_lookup': False,
                'called_jdy': False,
                'would_call_jdy': False,
                'cache_source': 'sales_details',
                'quantity_enrichment_skipped': True,
            }), 404
        if source == 'live':
            cfg2 = _load_jdy_config2()
            cfg1 = _load_jdy_config()
            name1 = cfg1.get('name', '饰品')
            name2 = cfg2.get('name', '箱包')
            candidates = []
            if account == name2:
                candidates = [(_ensure_jdy_client2, name2)]
            elif account == name1:
                candidates = [(_ensure_jdy_client, name1)]
            else:
                candidates = [(_ensure_jdy_client, name1), (_ensure_jdy_client2, name2)]

            errors = []
            for cli_fn, name in candidates:
                try:
                    cli = cli_fn()
                    if not cli:
                        errors.append(f'{name}: 请先配置 JDY API')
                        continue
                    detail = cli.get_sales_order_detail(order_no)
                    if not isinstance(detail, dict) or not (detail.get('entries') or detail.get('number')):
                        errors.append(f'{name}: 未找到销货单')
                        continue
                    if detail.get('number') and str(detail.get('number')) != str(order_no):
                        errors.append(f'{name}: 未找到销货单')
                        continue
                    order = _normalize_sales_order(detail, name)
                    return jsonify({
                        'success': True,
                        'data': _enrich_sales_order(cli, order),
                        'mock': False,
                        'cache': False,
                        'local_only': False,
                        'live_lookup': True,
                        'called_jdy': True,
                        'would_call_jdy': True,
                        'cache_source': 'jdy_live',
                        'quantity_enrichment_skipped': False,
                    })
                except Exception as e:
                    errors.append(f'{name}: {e}')
            return jsonify({'success': False, 'error': '；'.join(errors) or '未找到销货单'}), 404
        orders = _mock_sales_orders(date_str)
        for item in orders:
            if item.get('number') == order_no and (account in ('', 'all') or item.get('account') == account):
                return jsonify({'success': True, 'data': item, 'mock': True})
        return jsonify({'success': False, 'error': '未找到销货单'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/sales-cache/status', methods=['GET'])
def sales_cache_status():
    """查看销货单本地缓存状态。"""
    return jsonify({'success': True, 'stats': _sales_cache_stats()})


def _product_status_label(status):
    status = str(status or '').strip()
    if status == '0':
        return 'enabled'
    if status == '1':
        return 'disabled'
    if status in ('', 'all'):
        return 'all'
    return status


def _normalize_product_refresh_request(payload):
    payload = payload or {}
    account = str(payload.get('account') or 'all').strip() or 'all'
    status = str(payload.get('status') or 'enabled').strip()
    if status in ('enabled', '启用', '仅启用'):
        status = '0'
    elif status in ('disabled', '禁用', '仅禁用'):
        status = '1'
    elif status in ('all', '全部', ''):
        status = 'all'

    def _bounded_int(value, default, min_value, max_value):
        try:
            value = int(value)
        except Exception:
            value = default
        return max(min_value, min(value, max_value))

    page_size = _bounded_int(payload.get('page_size'), 100, 1, 500)
    try:
        limit = int(payload.get('limit') if payload.get('limit') not in (None, '') else 0)
    except Exception:
        limit = 0
    limit = max(0, min(limit, 1000000))
    sleep_ms = _bounded_int(payload.get('sleep_ms'), 300, 0, 5000)
    return {
        'account': account,
        'status': status,
        'status_filter': _product_status_label(status),
        'upd_time_begin': str(payload.get('updTimeBegin') or payload.get('begin') or payload.get('begin_date') or '').strip(),
        'upd_time_end': str(payload.get('updTimeEnd') or payload.get('end') or payload.get('end_date') or '').strip(),
        'page_size': page_size,
        'limit': limit,
        'sleep_ms': sleep_ms,
    }


def _product_refresh_sources(account):
    return _sales_sources_for_account(account or 'all')


def _jdy_product_existing_numbers():
    result = {'total': 0, 'by_account': {}}
    conn = _sales_readonly_conn()
    if not conn:
        return result
    try:
        if not _cache_sqlite_table_count(conn, 'jdy_products').get('exists'):
            return result
        rows = conn.execute('SELECT account, product_number FROM jdy_products').fetchall()
        result['total'] = len(rows)
        for row in rows:
            account = str(row['account'] or '')
            code = str(row['product_number'] or '').strip()
            bucket = result['by_account'].setdefault(account, set())
            if code:
                bucket.add(code)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return result


def _run_product_cache_refresh(payload, dry_run=True):
    opts = _normalize_product_refresh_request(payload)
    started_at = datetime.now()
    existing = _jdy_product_existing_numbers()
    result = {
        'success': True,
        'dry_run': bool(dry_run),
        'local_only': False,
        'live_lookup': True,
        'called_jdy': True,
        'would_call_jdy': True,
        'would_write': not dry_run,
        'would_download': False,
        'options': opts,
        'status_filter': opts['status_filter'],
        'written_table': 'jdy_products' if not dry_run else '',
        'remote_total': 0,
        'jdy_enabled_total': 0,
        'local_total': existing['total'],
        'local_total_before': existing['total'],
        'local_total_after': existing['total'],
        'estimated_insert': 0,
        'estimated_update': 0,
        'estimated_skip': 0,
        'account_summary': {},
        'scanned': 0,
        'valid': 0,
        'inserted': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'estimated_pages': 0,
        'sample_products': [],
        'fields': [
            'account', 'product_id', 'product_number', 'product_name', 'spec', 'barcode',
            'category_id', 'category_name', 'unit_id', 'unit_name', 'default_supplier_id',
            'default_supplier_number', 'default_supplier_name', 'image_url', 'status',
            'data_json', 'updated_at', 'last_seen_at',
        ],
    }
    sources = _product_refresh_sources(opts['account'])
    if not sources:
        result['success'] = False
        result['errors'].append(f"unknown account: {opts['account']}")
        return result

    def status_arg():
        return None if opts['status'] in ('', 'all') else opts['status']

    now = started_at.strftime('%Y-%m-%d %H:%M:%S')
    conn = None
    if not dry_run:
        conn = _sales_cache_conn()
    try:
        remaining = opts['limit']
        unlimited = remaining <= 0
        for cli_fn, account_name in sources:
            if not unlimited and remaining <= 0:
                break
            summary = result['account_summary'].setdefault(account_name, {
                'remote_total': 0,
                'local_total': len(existing['by_account'].get(account_name, set())),
                'scanned': 0,
                'valid': 0,
                'estimated_insert': 0,
                'estimated_update': 0,
                'estimated_skip': 0,
                'inserted': 0,
                'updated': 0,
                'skipped': 0,
            })
            cli = cli_fn()
            if not cli:
                result['errors'].append(f'{account_name}: 请先配置 JDY API')
                continue
            page = 1
            seen_total = False
            while unlimited or remaining > 0:
                page_size = opts['page_size'] if unlimited else min(opts['page_size'], remaining)
                try:
                    resp = cli.get_products(
                        page=page,
                        page_size=page_size,
                        status=status_arg(),
                        upd_time_begin=opts['upd_time_begin'],
                        upd_time_end=opts['upd_time_end'],
                    )
                except Exception as e:
                    result['errors'].append(f'{account_name} page {page}: {_short_sync_error(e)}')
                    break
                rows = resp.get('list') or []
                total = int(resp.get('total') or len(rows) or 0)
                if not seen_total:
                    summary['remote_total'] = total
                    result['remote_total'] += total
                    seen_total = True
                if total and opts['page_size']:
                    result['estimated_pages'] = max(result['estimated_pages'], int(math.ceil(total / opts['page_size'])))
                if not rows:
                    break
                for product in rows:
                    result['scanned'] += 1
                    summary['scanned'] += 1
                    item = _normalize_product_cache_item(product, account_name, 'jdy_products')
                    if not item:
                        result['skipped'] += 1
                        result['estimated_skip'] += 1
                        summary['skipped'] += 1
                        summary['estimated_skip'] += 1
                        continue
                    result['valid'] += 1
                    summary['valid'] += 1
                    is_existing = item['code'] in existing['by_account'].get(account_name, set())
                    if is_existing:
                        result['estimated_update'] += 1
                        summary['estimated_update'] += 1
                    else:
                        result['estimated_insert'] += 1
                        summary['estimated_insert'] += 1
                    if len(result['sample_products']) < 10:
                        result['sample_products'].append({
                            'account': account_name,
                            'code': item['code'],
                            'name': item['name'],
                            'spec': item['spec'],
                            'category_name': item['category_name'],
                            'unit_name': item['unit_name'],
                            'default_supplier_name': item['default_supplier_name'],
                            'image_url': item['image_url'],
                        })
                    if not dry_run:
                        _, action = _cache_upsert_jdy_product(conn, account_name, product, now=now)
                        if action == 'inserted':
                            result['inserted'] += 1
                            summary['inserted'] += 1
                        elif action == 'updated':
                            result['updated'] += 1
                            summary['updated'] += 1
                        else:
                            result['skipped'] += 1
                            summary['skipped'] += 1
                if not unlimited:
                    remaining -= len(rows)
                if len(rows) < page_size:
                    break
                page += 1
                if opts['sleep_ms']:
                    time.sleep(opts['sleep_ms'] / 1000.0)
        if conn:
            conn.commit()
            local_after = conn.execute('SELECT COUNT(*) AS c FROM jdy_products').fetchone()
            result['local_total_after'] = int(local_after['c'] if isinstance(local_after, sqlite3.Row) else local_after[0])
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass
    result['duration_seconds'] = round((datetime.now() - started_at).total_seconds(), 2)
    result['success'] = not bool(result['errors'])
    if opts['status_filter'] == 'enabled':
        result['jdy_enabled_total'] = result['remote_total']
    return result


@app.route('/cache/catalog-status', methods=['GET'])
def cache_catalog_status():
    """只读查看商品/供应商缓存口径状态，不调用 JDY，不写本地库。"""
    try:
        return jsonify(_catalog_cache_status_payload())
    except Exception as e:
        return jsonify({
            'success': False,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'error': str(e),
        }), 500


@app.route('/cache/product-refresh-dry-run', methods=['POST'])
def cache_product_refresh_dry_run():
    """管理员手动预览商品主档刷新。只调用 JDY product/list，不写本地库。"""
    try:
        payload = request.get_json(silent=True) or {}
        result = _run_product_cache_refresh(payload, dry_run=True)
        status = 200 if result.get('success') else 500
        return jsonify(result), status
    except Exception as e:
        return jsonify({
            'success': False,
            'dry_run': True,
            'local_only': False,
            'live_lookup': True,
            'called_jdy': True,
            'would_call_jdy': True,
            'would_write': False,
            'would_download': False,
            'error': str(e),
        }), 500


@app.route('/cache/product-refresh', methods=['POST'])
def cache_product_refresh():
    """管理员确认后刷新商品主档 SQLite 缓存。仅调用 JDY product/list 读接口。"""
    payload = request.get_json(silent=True) or {}
    if str(payload.get('confirm') or '').strip() != 'REFRESH_PRODUCT_CACHE':
        return jsonify({
            'success': False,
            'dry_run': False,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'would_write': True,
            'would_download': False,
            'status_filter': 'enabled',
            'written_table': 'jdy_products',
            'error': 'missing confirm token REFRESH_PRODUCT_CACHE',
            'message': '请先预览商品数量，再确认拉取启用商品到本地。',
        }), 400
    if not _product_cache_refresh_lock.acquire(blocking=False):
        return jsonify({
            'success': False,
            'running': True,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'would_write': True,
            'message': '商品缓存刷新正在执行，请稍后查看状态。',
            'state': _product_cache_refresh_state,
        }), 409
    started = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _product_cache_refresh_state.update({
        'running': True,
        'last_started_at': started,
        'last_finished_at': '',
        'last_success': False,
        'last_error': '',
        'message': '商品缓存刷新执行中',
    })
    try:
        result = _run_product_cache_refresh(payload, dry_run=False)
        _product_cache_refresh_state.update({
            'running': False,
            'last_finished_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': bool(result.get('success')),
            'last_error': '; '.join(result.get('errors') or [])[:800],
            'last_result': result,
            'message': '商品缓存刷新已完成' if result.get('success') else '商品缓存刷新完成但存在错误',
        })
        status = 200 if result.get('success') else 500
        return jsonify({**result, 'state': _product_cache_refresh_state}), status
    except Exception as e:
        _product_cache_refresh_state.update({
            'running': False,
            'last_finished_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_success': False,
            'last_error': str(e),
            'message': '商品缓存刷新失败',
        })
        return jsonify({
            'success': False,
            'dry_run': False,
            'local_only': False,
            'live_lookup': True,
            'called_jdy': True,
            'would_call_jdy': True,
            'would_write': True,
            'would_download': False,
            'error': str(e),
            'state': _product_cache_refresh_state,
        }), 500
    finally:
        _product_cache_refresh_lock.release()


@app.route('/cache/product-refresh/status', methods=['GET'])
def cache_product_refresh_status():
    return jsonify({
        'success': True,
        'local_only': True,
        'live_lookup': False,
        'called_jdy': False,
        'would_call_jdy': False,
        'state': _product_cache_refresh_state,
    })


@app.route('/cache/product-refresh/cancel', methods=['POST'])
def cache_product_refresh_cancel():
    return jsonify({
        'success': False,
        'local_only': True,
        'live_lookup': False,
        'called_jdy': False,
        'would_call_jdy': False,
        'cancelled': False,
        'message': '当前商品缓存刷新为同步执行，暂无可取消的后台任务。',
        'state': _product_cache_refresh_state,
    }), 409


@app.route('/sales-cache-refresh', methods=['POST'])
def sales_cache_refresh():
    """只读精斗云并刷新本地销货单缓存。"""
    try:
        payload = request.get_json(silent=True) or {}
        today = datetime.now()
        begin_default = today.replace(day=1).strftime('%Y-%m-%d')
        end_default = today.strftime('%Y-%m-%d')
        begin_date = (
            request.args.get('begin_date')
            or payload.get('begin_date')
            or begin_default
        )
        end_date = (
            request.args.get('end_date')
            or payload.get('end_date')
            or end_default
        )
        account = request.args.get('account') or payload.get('account') or 'all'
        quantities_only = str(
            request.args.get('quantities_only')
            or payload.get('quantities_only')
            or ''
        ).lower() in ('1', 'true', 'yes', 'y')
        if quantities_only:
            result = _refresh_sales_quantities_from_cache(begin_date, end_date, account)
        else:
            result = _refresh_sales_cache(begin_date, end_date, account)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/sales-sync-config', methods=['GET'])
def sales_sync_config_get():
    return jsonify({
        'success': True,
        'config': _sales_sync_config(),
        'state': _sales_sync_state,
        'config_debug': _runtime_config_debug(),
    })


@app.route('/sales-sync-config', methods=['POST'])
def sales_sync_config_set():
    try:
        data = request.get_json(force=True) or {}

        def _bounded_int(name, default, min_value, max_value):
            try:
                value = int(data.get(name, default))
            except Exception:
                value = default
            return max(min_value, min(max_value, value))

        expire_action = str(data.get('expire_action') or 'delete').strip().lower()
        if expire_action not in ('delete', 'archive'):
            expire_action = 'delete'
        daily_compare_time = str(data.get('daily_compare_time') or '').strip()
        archive_path = str(data.get('archive_path') or '').strip()
        factory_switch_date = str(data.get('factory_purchase_begin_date') or '').strip()
        factory_disabled_suppliers = str(data.get('factory_disabled_suppliers') or '').strip()
        accessory_purchase_begin_date = str(data.get('accessory_purchase_begin_date') or '').strip()
        accessory_supplier_terms = str(data.get('accessory_supplier_terms') or '').strip()
        updates = {
            'sales_auto_sync_enabled': False,
            'sales_auto_sync_interval_minutes': _bounded_int('interval_minutes', 5, 1, 1440),
            'sales_cache_keep_days': _bounded_int('keep_days', 30, 1, 365),
            'sales_incremental_lookback_days': _bounded_int('lookback_days', 3, 1, 60),
            'sales_cache_manual_days': _bounded_int('lookback_days', 3, 1, 60),
            'sales_cache_expire_action': expire_action,
            'accessory_auto_sync_enabled': False,
            'accessory_auto_sync_interval_minutes': _bounded_int('accessory_interval_minutes', 5, 1, 1440),
            'accessory_supplier_terms': str(data.get('accessory_supplier_terms') or '').strip() or '辅料供应商\n辅料',
        }
        if not accessory_supplier_terms:
            updates.pop('accessory_supplier_terms', None)
        if daily_compare_time:
            updates['jdy_daily_compare_time'] = daily_compare_time
            updates['sales_cache_daily_time'] = daily_compare_time
        if archive_path:
            updates['sales_cache_archive_path'] = archive_path
        if factory_switch_date:
            updates['sales_factory_purchase_begin_date'] = factory_switch_date
            updates['sales_factory_new_logic_begin_date'] = factory_switch_date
            updates['factory_purchase_begin_date'] = factory_switch_date
            updates['factory_new_logic_begin_date'] = factory_switch_date
        if factory_disabled_suppliers:
            updates['sales_factory_disabled_suppliers'] = factory_disabled_suppliers
            updates['factory_disabled_suppliers'] = factory_disabled_suppliers
        if accessory_purchase_begin_date:
            updates['accessory_purchase_begin_date'] = accessory_purchase_begin_date
        if accessory_supplier_terms:
            updates['accessory_supplier_terms'] = accessory_supplier_terms
            updates['accessory_supplier_keywords'] = accessory_supplier_terms
        _save_jdy_config(updates)
        return jsonify({'success': True, 'config': _sales_sync_config()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/sales-sync/status', methods=['GET'])
def sales_sync_status():
    _webhook_state['pending'] = _webhook_pending_count()
    return jsonify({
        'success': True,
        'config': _sales_sync_config(),
        'state': _sales_sync_state,
        'webhook_state': _webhook_state,
        'daily_compare_state': _daily_compare_state,
    })


@app.route('/jdy-webhook/accessory-diagnose', methods=['GET'])
def jdy_webhook_accessory_diagnose():
    event_id = str(request.args.get('id') or '').strip()
    bill_no = str(request.args.get('bill_no') or '').strip()
    if not event_id and not bill_no:
        return jsonify({'success': False, 'error': '请提供 id 或 bill_no'}), 400

    with _sales_cache_conn() as conn:
        if event_id:
            row = conn.execute('''
                SELECT * FROM webhook_events WHERE id = ? LIMIT 1
            ''', (event_id,)).fetchone()
        else:
            row = conn.execute('''
                SELECT * FROM webhook_events
                WHERE bill_no = ?
                ORDER BY id DESC
                LIMIT 1
            ''', (bill_no,)).fetchone()
        if not row:
            return jsonify({
                'success': False,
                'error': '未找到对应 webhook 事件',
                'query': {'id': event_id, 'bill_no': bill_no},
            }), 404

        event = {k: row[k] for k in row.keys() if k != 'payload_json'}
        payload_raw = row['payload_json'] or ''
        try:
            payload = json.loads(payload_raw) if payload_raw else {}
        except Exception as e:
            payload = {}
            payload_error = str(e)
        else:
            payload_error = ''

        visible_numbers = _extract_webhook_values(payload, [
            'number', 'billNo', 'billNumber', 'bill_no',
            'orderNo', 'orderNumber', 'sourceBillNo', 'srcBillNo',
        ])
        internal_ids = _extract_webhook_values(payload, ['id', 'billId'])
        all_numbers = _extract_webhook_bill_numbers(payload)
        current_bill_no = str(row['bill_no'] or '').strip()
        candidates = [current_bill_no] + visible_numbers + internal_ids + all_numbers
        candidates = [x for x in dict.fromkeys(candidates) if x]

        cached_orders = []
        if candidates:
            placeholders = ','.join('?' for _ in candidates)
            params = list(candidates)
            account = str(row['account'] or '').strip()
            account_clause = ''
            if account:
                account_clause = ' AND account = ?'
                params.append(account)
            cached_rows = conn.execute(f'''
                SELECT account, number, date, supplier_name, supplier_number,
                       entries_count, data_json, updated_at
                FROM accessory_purchase_orders
                WHERE number IN ({placeholders}){account_clause}
                ORDER BY updated_at DESC
            ''', params).fetchall()
            for cached in cached_rows:
                cached_orders.append({
                    'account': cached['account'],
                    'number': cached['number'],
                    'date': cached['date'],
                    'supplier_name': cached['supplier_name'],
                    'supplier_number': cached['supplier_number'],
                    'entries_count': cached['entries_count'],
                    'updated_at': cached['updated_at'],
                })

        supplier_no_values = _extract_webhook_values(payload, [
            'supplierNumber', 'supplierNo', 'supplier_number',
            'vendorNumber', 'vendorNo', 'supplierCode',
        ])
        supplier_name_values = _extract_webhook_values(payload, [
            'supplierName', 'supplier_name', 'vendorName', 'vendor',
        ])
        if cached_orders:
            supplier_no_values = [cached_orders[0].get('supplier_number') or ''] + supplier_no_values
            supplier_name_values = [cached_orders[0].get('supplier_name') or ''] + supplier_name_values
        supplier_no_values = [x for x in dict.fromkeys(supplier_no_values) if x]
        supplier_name_values = [x for x in dict.fromkeys(supplier_name_values) if x]

        supplier = None
        supplier_source = ''
        account = str(row['account'] or '').strip()
        if supplier_no_values:
            params = [account, supplier_no_values[0]]
            supplier_row = conn.execute('''
                SELECT data_json FROM jdy_suppliers
                WHERE account = ? AND number = ?
                LIMIT 1
            ''', params).fetchone()
            if supplier_row:
                try:
                    supplier = json.loads(supplier_row['data_json'] or '{}')
                    supplier_source = 'jdy_suppliers.number'
                except Exception:
                    supplier = None
        if supplier is None and supplier_name_values:
            supplier_row = conn.execute('''
                SELECT data_json FROM jdy_suppliers
                WHERE account = ? AND name = ?
                LIMIT 1
            ''', (account, supplier_name_values[0])).fetchone()
            if supplier_row:
                try:
                    supplier = json.loads(supplier_row['data_json'] or '{}')
                    supplier_source = 'jdy_suppliers.name'
                except Exception:
                    supplier = None

    supplier_order_hint = {
        'supplierNumber': supplier_no_values[0] if supplier_no_values else '',
        'supplierName': supplier_name_values[0] if supplier_name_values else '',
    }
    has_supplier_hint = bool(supplier_no_values or supplier_name_values)
    supplier_is_accessory = None
    if supplier or has_supplier_hint:
        supplier_is_accessory = _supplier_matches_accessory(supplier or {}, supplier_order_hint)

    bill_kind = _webhook_bill_no_kind(current_bill_no)
    payload_analysis = {
        'has_payload_json': bool(payload_raw),
        'payload_parse_error': payload_error,
        'visible_numbers': visible_numbers,
        'internal_ids': internal_ids,
        'all_extracted_numbers': all_numbers,
        'current_bill_no': current_bill_no,
        'current_bill_no_kind': bill_kind,
        'looks_like_internal_id': bill_kind == 'internal_id',
        'looks_like_visible_number': bill_kind == 'visible_number',
        'suspected_internal_id_used_as_number': bill_kind == 'internal_id' and not visible_numbers,
    }

    local_cache_check = {
        'candidate_numbers_checked': candidates,
        'exists': bool(cached_orders),
        'orders': cached_orders,
        'message': '本地辅料采购订单缓存中已找到' if cached_orders else '本地辅料采购订单缓存中未找到',
    }

    supplier_check = {
        'supplier_number_candidates': supplier_no_values,
        'supplier_name_candidates': supplier_name_values,
        'cached_supplier_found': supplier is not None,
        'cached_supplier_source': supplier_source,
        'is_accessory_supplier': supplier_is_accessory,
        'terms': _accessory_supplier_terms(),
        'message': '',
    }
    if supplier_is_accessory is True:
        supplier_check['message'] = '本地信息判断该供应商像辅料供应商'
    elif supplier_is_accessory is False:
        supplier_check['message'] = '本地信息判断该供应商不是辅料供应商'
    else:
        supplier_check['message'] = '本地信息不足，无法判断供应商是否为辅料供应商'

    likely = []
    resource = str(row['resource'] or '')
    error_text = str(row['error'] or '')
    if resource != 'accessory_purchase':
        likely.append('这条 webhook 不是 accessory_purchase 类型，当前辅料诊断只做参考')
    if cached_orders:
        likely.append('本地已存在对应辅料采购订单；如果页面未显示，优先检查辅料页面筛选条件')
    else:
        likely.append('本地 accessory_purchase_orders 未找到该单据')
    if 'recorded only' in error_text.lower() or '仅记录' in error_text:
        likely.append('该记录是在旧逻辑下处理的 recorded only，历史记录无法追溯当时的精确原因')
    if payload_analysis['suspected_internal_id_used_as_number']:
        likely.append('当前 bill_no 看起来像 JDY 内部 id，且 payload 中没有可见采购订单编号')
    elif bill_kind == 'internal_id':
        likely.append('当前 bill_no 看起来像 JDY 内部 id，可能需要按内部 id 做一次只读 JDY 诊断')
    if not visible_numbers:
        likely.append('payload 中没有提取到可见采购订单编号')
    if supplier_is_accessory is False:
        likely.append('本地供应商判断结果不是辅料供应商')
    elif supplier_is_accessory is None:
        likely.append('本地信息不足，无法判断供应商是否为辅料供应商')
    if not likely:
        likely.append('本地信息未发现明显原因；需要人工提供可见采购订单编号或单独确认只读 JDY 查询')

    recommended = []
    if payload_analysis['looks_like_internal_id']:
        recommended.append('如果这是内部 id，建议后续单独确认后增加“按内部 id 只读查询采购订单”的诊断')
    if not visible_numbers:
        recommended.append('建议从 JDY 页面复制可见采购订单编号，再用 bill_no 诊断')
    if supplier_is_accessory is False:
        recommended.append('检查供应商分类或设置页 accessory_supplier_terms 是否覆盖该供应商')
    if cached_orders:
        recommended.append('检查辅料页面账套、搜索词或筛选条件是否隐藏了该单据')
    if 'recorded only' in error_text.lower() or '仅记录' in error_text:
        recommended.append('历史 recorded only 不会自动重跑；以后新事件会写入更明确原因')

    return jsonify({
        'success': True,
        'event': event,
        'payload_analysis': payload_analysis,
        'local_cache_check': local_cache_check,
        'supplier_check_local': supplier_check,
        'expected_logic': [
            '采购订单 webhook 会被识别为 accessory_purchase',
            'worker 处理时会按 bill_no 拉取采购订单详情',
            '系统判断供应商是否属于辅料供应商',
            '匹配则写入 accessory_purchase_orders 供辅料页面读取',
            '本诊断接口不联网、不重试、不写库，只解释本地已有信息',
        ],
        'likely_reason': likely,
        'recommended_next_action': recommended,
    })


def _webhook_event_by_id_or_bill_no(conn, event_id='', bill_no=''):
    event_id = str(event_id or '').strip()
    bill_no = str(bill_no or '').strip()
    if event_id:
        return conn.execute('''
            SELECT * FROM webhook_events WHERE id = ? LIMIT 1
        ''', (event_id,)).fetchone()
    if bill_no:
        return conn.execute('''
            SELECT * FROM webhook_events
            WHERE bill_no = ?
            ORDER BY id DESC
            LIMIT 1
        ''', (bill_no,)).fetchone()
    return None


def _webhook_event_public(row):
    return {k: row[k] for k in row.keys() if k != 'payload_json'}


def _parse_webhook_payload(raw):
    try:
        payload = json.loads(raw) if raw else {}
        return payload, ''
    except Exception as e:
        return {}, str(e)


def _analyze_webhook_payload(payload, current_bill_no):
    visible_numbers = _extract_webhook_values(payload, [
        'number', 'billNo', 'billNumber', 'bill_no',
        'orderNo', 'orderNumber', 'sourceBillNo', 'srcBillNo',
    ])
    internal_ids = _extract_webhook_values(payload, ['id', 'billId'])
    all_numbers = _extract_webhook_bill_numbers(payload)
    bill_kind = _webhook_bill_no_kind(current_bill_no)
    return {
        'visible_numbers': visible_numbers,
        'internal_ids': internal_ids,
        'all_extracted_numbers': all_numbers,
        'current_bill_no': str(current_bill_no or '').strip(),
        'current_bill_no_kind': bill_kind,
        'looks_like_internal_id': bill_kind == 'internal_id',
        'looks_like_visible_number': bill_kind == 'visible_number',
        'suspected_internal_id_used_as_number': bill_kind == 'internal_id' and not visible_numbers,
    }


def _accessory_supplier_match_terms(supplier, order=None):
    supplier = supplier or {}
    order = order or {}
    terms = [str(t or '').strip() for t in _accessory_supplier_terms() if str(t or '').strip()]
    fields = [
        'supplierCategoryName', 'supplierCategory', 'categoryName', 'category',
        'typeName', 'supplierTypeName', 'className', 'groupName',
        'newRecTypeName', 'recTypeName', 'remark', 'description',
    ]
    values = []
    for src in (supplier, order):
        for key in fields:
            val = src.get(key)
            if val not in (None, ''):
                values.append(str(val))
    values.extend([
        str(_first_value(order, ['supplierName', 'vendorName'], '')),
        str(_first_value(supplier, ['name', 'supplierName'], '')),
    ])
    haystack = '\n'.join(values).lower()
    return [term for term in terms if term.lower() in haystack]


def _purchase_order_candidate_preview(rows, limit=5):
    preview = []
    for row in (rows or [])[:limit]:
        entries = (
            row.get('entries') or row.get('items') or row.get('details') or
            row.get('entryList') or row.get('goods') or []
        )
        preview.append({
            'visible_number': str(_first_value(row, ['number', 'billNo', 'billNumber'], '') or ''),
            'internal_id': str(_first_value(row, ['id', 'billId'], '') or ''),
            'date': str(_first_value(row, ['date', 'billDate', 'orderDate'], '') or '')[:10],
            'check_status': row.get('checkStatus'),
            'bill_status': row.get('billStatus'),
            'bill_status_name': _first_value(row, ['billStatusName', 'statusName', 'status'], ''),
            'supplier_name': str(_first_value(row, ['supplierName', 'vendorName'], '') or ''),
            'supplier_number': str(_first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], '') or ''),
            'entries_count': len(entries) if isinstance(entries, list) else 0,
        })
    return preview


def _find_purchase_order_by_number_list_readonly(cli, number, method='visible_number',
                                                note='', page_size=10,
                                                allow_internal_match=False):
    attempts = []
    number = str(number or '').strip()
    if not cli or not number:
        return None, attempts, []
    attempt = {
        'method': method,
        'query': number,
        'found': False,
        'note': note or '使用 purchaseOrder/list 的 filter.number 只读查询',
    }
    result = cli.get_purchase_order_requests(
        page=1,
        page_size=page_size,
        search=number,
        bill_status=None,
        check_status=2,
    )
    rows = result.get('list') or []
    attempt['result_count'] = len(rows)
    attempt['candidate_count'] = len(rows)
    exact = None
    matched_by = ''
    for row in rows:
        row_no = str(_first_value(row, ['number', 'billNo', 'billNumber'], '')).strip()
        row_id = str(_first_value(row, ['id', 'billId'], '')).strip()
        if row_no == number:
            exact = row
            matched_by = 'visible_number'
            break
        if allow_internal_match and row_id and row_id == number:
            exact = row
            matched_by = method
            break
    if exact:
        attempt['found'] = True
        attempt['matched_by'] = matched_by or method
        if method == 'internal_id_as_number_list_search':
            attempt['note'] = '把内部 id 作为 filter.number 做受控只读尝试并命中候选；这不等同于官方内部 ID 精确查询'
        else:
            attempt['note'] = '按可见采购订单编号精确命中'
    elif rows:
        attempt['note'] = '查询有返回，但没有精确匹配；需要人工确认可见采购订单编号'
    else:
        if method == 'internal_id_as_number_list_search':
            attempt['note'] = '当前 list 查询方式未命中；不能证明订单不存在'
        else:
            attempt['note'] = '当前查询方式未命中'
    attempts.append(attempt)
    return exact, attempts, rows


def _find_purchase_order_by_visible_number_readonly(cli, number):
    return _find_purchase_order_by_number_list_readonly(
        cli,
        number,
        method='visible_number',
        note='按可见采购订单编号使用 purchaseOrder/list filter.number 查询；checkStatus=2 查询全部审核状态',
        page_size=10,
        allow_internal_match=False,
    )


def _purchase_order_live_preview(order, account, supplier=None):
    normalized = _normalize_accessory_purchase_order(order or {}, account, supplier or {})
    preview = []
    for entry in (normalized.get('entries') or [])[:20]:
        preview.append({
            'code': entry.get('code') or '',
            'name': entry.get('name') or '',
            'spec': entry.get('spec') or '',
            'qty': entry.get('qty') or 0,
            'unit': entry.get('unit') or '',
            'price': entry.get('price') or 0,
            'amount': entry.get('amount') or 0,
        })
    return normalized, preview


@app.route('/jdy-webhook/accessory-diagnose-live', methods=['POST'])
def jdy_webhook_accessory_diagnose_live():
    data = request.get_json(silent=True) or {}
    if str(data.get('confirm') or '').strip() != 'READ_ONLY_JDY':
        return jsonify({
            'success': False,
            'error': '必须确认 confirm=READ_ONLY_JDY 才能执行 JDY 只读诊断',
        }), 400

    event_id = str(data.get('id') or '').strip()
    bill_no = str(data.get('bill_no') or '').strip()
    allow_internal_id_search = data.get('allow_internal_id_search') is True
    internal_id_confirm = str(data.get('internal_id_confirm') or '').strip()
    if not event_id and not bill_no:
        return jsonify({'success': False, 'error': '请提供 id 或 bill_no'}), 400

    with _sales_cache_conn() as conn:
        row = _webhook_event_by_id_or_bill_no(conn, event_id, bill_no)
    if not row:
        return jsonify({
            'success': False,
            'error': '未找到对应 webhook 事件',
            'mode': 'read_only_jdy',
            'query': {'id': event_id, 'bill_no': bill_no},
            'jdy_query': {'called': False, 'reason': 'event_not_found'},
        }), 404

    payload_raw = row['payload_json'] or ''
    payload, payload_error = _parse_webhook_payload(payload_raw)
    event = _webhook_event_public(row)
    payload_analysis = _analyze_webhook_payload(payload, row['bill_no'] or '')
    payload_analysis.update({
        'has_payload_json': bool(payload_raw),
        'payload_parse_error': payload_error,
    })

    cfg1 = _load_jdy_config()
    cfg2 = _load_jdy_config2()
    account = str(row['account'] or '').strip() or _jdy_account_from_payload(payload)
    cli = None
    account_name = account
    if account and account == str(cfg2.get('name') or ''):
        cli = _ensure_jdy_client2()
        account_name = cfg2.get('name') or account
    elif account and account == str(cfg1.get('name') or ''):
        cli = _ensure_jdy_client()
        account_name = cfg1.get('name') or account

    local_diag = {
        'resource': row['resource'] or '',
        'historical_error': row['error'] or '',
        'note': '本接口只做 JDY 单条只读诊断，不重试 webhook，不写本地缓存',
    }
    live_reasons = []
    recommended = []
    jdy_query = {
        'called': False,
        'account': account_name or '',
        'search_attempts': [],
        'found': False,
        'matched_by': '',
        'order_candidates_count': 0,
        'order_candidates_preview': [],
        'order': {},
        'product_lines_preview': [],
    }

    if not account_name or not cli:
        live_reasons.append('无法确定账套或账套 JDY 客户端未配置，未调用 JDY')
        recommended.append('请检查 webhook 的 account/appKey/dbId 与设置页账套配置是否匹配')
        return jsonify({
            'success': True,
            'mode': 'read_only_jdy',
            'event': event,
            'payload_analysis': payload_analysis,
            'local_diagnosis': local_diag,
            'jdy_query': jdy_query,
            'supplier_check': {
                'supplier_found': False,
                'supplier_source': '',
                'is_accessory_supplier': None,
                'accessory_supplier_terms': _accessory_supplier_terms(),
                'message': '未调用 JDY，无法判断供应商',
            },
            'live_likely_reason': live_reasons,
            'recommended_next_action': recommended,
        })

    visible_numbers = payload_analysis.get('visible_numbers') or []
    internal_ids = payload_analysis.get('internal_ids') or []
    current_bill_no = str(row['bill_no'] or '').strip()
    if not visible_numbers and _webhook_bill_no_kind(current_bill_no) == 'visible_number':
        visible_numbers = [current_bill_no]

    order = None
    if visible_numbers:
        try:
            order, attempts, candidates = _find_purchase_order_by_visible_number_readonly(cli, visible_numbers[0])
        except Exception as e:
            return jsonify({
                'success': False,
                'mode': 'read_only_jdy',
                'event': event,
                'payload_analysis': payload_analysis,
                'local_diagnosis': local_diag,
                'jdy_query': {
                    **jdy_query,
                    'called': True,
                    'account': account_name,
                    'error': str(e),
                },
                'live_likely_reason': ['JDY 只读查询报错'],
                'recommended_next_action': ['稍后再试，或检查 JDY 网关/账号授权状态'],
            }), 502
        jdy_query['called'] = True
        jdy_query['search_attempts'] = attempts
        jdy_query['order_candidates_count'] = len(candidates or [])
        jdy_query['order_candidates_preview'] = _purchase_order_candidate_preview(candidates or [])
        if order:
            jdy_query['found'] = True
            jdy_query['matched_by'] = 'visible_number'
        else:
            live_reasons.append('当前查询方式未查到订单，不能证明订单不存在')
            recommended.append('请确认 webhook payload 中的可见采购订单编号是否正确')
    else:
        internal_query = (internal_ids[0] if internal_ids else current_bill_no).strip()
        if internal_query and (internal_ids or _webhook_bill_no_kind(current_bill_no) == 'internal_id'):
            if not allow_internal_id_search:
                live_reasons.append('只有内部 id，没有可见采购订单编号；默认不调用 JDY')
                recommended.append('如需受控只读尝试，请传 allow_internal_id_search=true 且 internal_id_confirm=TRY_INTERNAL_ID_SEARCH')
                return jsonify({
                    'success': True,
                    'mode': 'read_only_jdy',
                    'event': event,
                    'payload_analysis': payload_analysis,
                    'local_diagnosis': local_diag,
                    'jdy_query': jdy_query,
                    'supplier_check': {
                        'supplier_found': False,
                        'supplier_source': '',
                        'is_accessory_supplier': None,
                        'accessory_supplier_terms': _accessory_supplier_terms(),
                        'message': '只有内部 id，未显式允许受控只读查询，未调用 JDY',
                    },
                    'live_likely_reason': live_reasons,
                    'recommended_next_action': recommended,
                })
            if internal_id_confirm != 'TRY_INTERNAL_ID_SEARCH':
                return jsonify({
                    'success': False,
                    'mode': 'read_only_jdy',
                    'event': event,
                    'payload_analysis': payload_analysis,
                    'local_diagnosis': local_diag,
                    'jdy_query': jdy_query,
                    'error': '内部 id 受控只读尝试需要 internal_id_confirm=TRY_INTERNAL_ID_SEARCH',
                }), 400
            try:
                order, attempts, candidates = _find_purchase_order_by_number_list_readonly(
                    cli,
                    internal_query,
                    method='internal_id_as_number_list_search',
                    note='把内部 id 当作 purchaseOrder/list 的 filter.number 做受控只读尝试；不等同于官方内部 ID 精确查询',
                    page_size=10,
                    allow_internal_match=True,
                )
            except Exception as e:
                return jsonify({
                    'success': False,
                    'mode': 'read_only_jdy',
                    'event': event,
                    'payload_analysis': payload_analysis,
                    'local_diagnosis': local_diag,
                    'jdy_query': {
                        **jdy_query,
                        'called': True,
                        'account': account_name,
                        'error': str(e),
                    },
                    'live_likely_reason': ['JDY 内部 id 受控只读尝试报错'],
                    'recommended_next_action': ['稍后再试，或检查 JDY 网关/账号授权状态'],
                }), 502
            jdy_query['called'] = True
            jdy_query['search_attempts'] = attempts
            jdy_query['order_candidates_count'] = len(candidates or [])
            jdy_query['order_candidates_preview'] = _purchase_order_candidate_preview(candidates or [])
            if order:
                jdy_query['found'] = True
                jdy_query['matched_by'] = 'internal_id_as_number_list_search'
                live_reasons.append('通过内部 id 作为 filter.number 的受控只读尝试找到了候选采购订单')
                recommended.append('这不是官方内部 ID 精确查询；如需补缓存，建议先人工核对可见采购订单编号')
            elif candidates:
                live_reasons.append('通过内部 id 作为 filter.number 查询到了候选订单，但没有精确匹配，需要人工确认')
                recommended.append('请从候选列表或 JDY 页面确认可见采购订单编号后再决定是否补缓存')
            else:
                live_reasons.append('当前 list 查询方式未命中，不能证明订单不存在')
                recommended.append('请从 JDY 页面复制可见采购订单编号，或向官方确认是否有内部 ID 查询接口')
        else:
            live_reasons.append('payload 中没有可见采购订单编号，未调用 JDY')
            recommended.append('建议从 JDY 页面复制可见采购订单编号后再诊断')
            return jsonify({
                'success': True,
                'mode': 'read_only_jdy',
                'event': event,
                'payload_analysis': payload_analysis,
                'local_diagnosis': local_diag,
                'jdy_query': jdy_query,
                'supplier_check': {
                    'supplier_found': False,
                    'supplier_source': '',
                    'is_accessory_supplier': None,
                    'accessory_supplier_terms': _accessory_supplier_terms(),
                    'message': '没有可见采购订单编号，未调用 JDY',
                },
                'live_likely_reason': live_reasons,
                'recommended_next_action': recommended,
            })

    supplier = None
    supplier_source = ''
    supplier_error = ''
    order_summary = {}
    preview = []
    supplier_check = {
        'supplier_found': False,
        'supplier_source': '',
        'supplier_number': '',
        'supplier_name': '',
        'is_accessory_supplier': None,
        'matched_terms': [],
        'accessory_supplier_terms': _accessory_supplier_terms(),
        'message': '未查到采购订单，无法判断供应商',
    }
    if jdy_query.get('found'):
        supplier_no = str(_first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], '')).strip()
        supplier_name = str(_first_value(order, ['supplierName', 'vendorName'], '')).strip()
        with _sales_cache_conn() as conn:
            supplier = _read_cached_jdy_supplier(conn, account_name, supplier_no) if supplier_no else None
        if supplier:
            supplier_source = 'local_jdy_suppliers'
        elif supplier_no and hasattr(cli, 'get_supplier_by_number'):
            try:
                supplier = cli.get_supplier_by_number(supplier_no, status=2) or None
                supplier_source = 'jdy_supplier_readonly' if supplier else ''
            except Exception as e:
                supplier_error = str(e)
                supplier = None

        normalized, preview = _purchase_order_live_preview(order, account_name, supplier or {})
        order_summary = {
            'visible_number': normalized.get('number') or '',
            'internal_id': str(_first_value(order, ['id', 'billId'], '') or ''),
            'date': normalized.get('date') or '',
            'bill_status': normalized.get('billStatus') or '',
            'bill_status_name': normalized.get('billStatusName') or '',
            'check_status': normalized.get('checkStatus'),
            'supplier_name': normalized.get('supplierName') or supplier_name,
            'supplier_number': normalized.get('supplierNumber') or supplier_no,
            'entries_count': len(normalized.get('entries') or []),
            'total_qty': normalized.get('totalQty') or 0,
            'total_amount': normalized.get('totalAmount') or 0,
            'would_cache': bool(normalized.get('entries')),
        }
        supplier_hint = {
            'supplierNumber': order_summary['supplier_number'],
            'supplierName': order_summary['supplier_name'],
        }
        is_accessory = _supplier_matches_accessory(supplier or {}, {**order, **supplier_hint})
        matched_terms = _accessory_supplier_match_terms(supplier or {}, {**order, **supplier_hint})
        supplier_check.update({
            'supplier_found': bool(supplier) or bool(supplier_no or supplier_name),
            'supplier_source': supplier_source or ('purchase_order_fields' if (supplier_no or supplier_name) else ''),
            'supplier_number': order_summary['supplier_number'],
            'supplier_name': order_summary['supplier_name'],
            'is_accessory_supplier': is_accessory,
            'matched_terms': matched_terms,
            'message': '',
        })
        if supplier_error:
            supplier_check['supplier_error'] = supplier_error
        if is_accessory and order_summary['entries_count']:
            live_reasons.append('JDY 查询到了订单，供应商像辅料供应商，理论上可以单条补缓存')
            recommended.append('如确认无误，下一步可单独做“管理员确认后单条补缓存”')
            supplier_check['message'] = '供应商匹配辅料关键词'
        elif is_accessory and not order_summary['entries_count']:
            live_reasons.append('JDY 查询到了订单，供应商像辅料供应商，但采购订单无商品明细')
            recommended.append('请在 JDY 检查该采购订单明细是否为空')
            supplier_check['message'] = '供应商匹配辅料关键词，但订单无明细'
        elif not is_accessory:
            live_reasons.append('JDY 查询到了订单，但供应商不是辅料供应商')
            recommended.append('不建议进入辅料页面；如判断错误，请检查供应商分类或辅料关键词')
            supplier_check['message'] = '供应商未匹配辅料关键词'
        else:
            live_reasons.append('JDY 查询到了订单，但供应商资料不足，无法判断是否为辅料供应商')
            recommended.append('建议先补供应商缓存或检查供应商分类/关键词')
            supplier_check['message'] = '供应商资料不足，无法判断'

    jdy_query.update({
        'account': account_name,
        'order': order_summary,
        'product_lines_preview': preview,
    })

    return jsonify({
        'success': True,
        'mode': 'read_only_jdy',
        'event': event,
        'payload_analysis': payload_analysis,
        'local_diagnosis': local_diag,
        'jdy_query': jdy_query,
        'supplier_check': supplier_check,
        'live_likely_reason': live_reasons,
        'recommended_next_action': recommended,
    })


@app.route('/jdy-webhook/status', methods=['GET'])
def jdy_webhook_status():
    snapshot = _webhook_status_snapshot(request.args.get('limit') or 20)
    restarted = _webhook_maybe_restart_auto_worker(snapshot.get('pending_count') or 0, 'status')
    if restarted:
        snapshot = _webhook_status_snapshot(request.args.get('limit') or 20)
    state = dict(_webhook_state)
    state['pending'] = snapshot['counts'].get('pending', 0)
    state['retry_pending'] = snapshot['counts'].get('retry_pending', 0)
    return jsonify({
        'success': True,
        'endpoint': {
            'local_ready': True,
            'path': '/jdy-webhook',
            'public': True,
            'method': 'GET/POST',
            'remote_url': 'http://gongdashuai.top:5008/jdy-webhook',
        },
        'state': state,
        'counts': snapshot['counts'],
        'recent': snapshot['recent'],
        'resource_summary': snapshot['resource_summary'],
        'by_resource': snapshot['by_resource'],
        'duplicate_summary': snapshot['duplicate_summary'],
        'retry_duplicate_summary': snapshot.get('retry_duplicate_summary') or [],
        'stale_processing': snapshot['stale_processing'],
        'backlog_estimate': snapshot['backlog_estimate'],
        'backlog_estimate_normal': snapshot.get('backlog_estimate_normal') or snapshot['backlog_estimate'],
        'backlog_estimate_retry': snapshot.get('backlog_estimate_retry') or {},
        'current_rate_policy': snapshot['current_rate_policy'],
        'processing_mode': snapshot.get('processing_mode'),
        'auto_enabled': snapshot.get('auto_enabled'),
        'manual_processing_required': snapshot.get('manual_processing_required'),
        'pending_count': snapshot.get('pending_count'),
        'retry_pending_count': snapshot.get('retry_pending_count'),
        'current_rate_limit': snapshot.get('current_rate_limit'),
        'estimated_minutes_300': snapshot.get('estimated_minutes_300'),
        'estimated_minutes_450': snapshot.get('estimated_minutes_450'),
        'worker_alive': snapshot.get('worker_alive'),
        'worker_thread_name': snapshot.get('worker_thread_name'),
        'worker_thread_id': snapshot.get('worker_thread_id'),
        'worker_last_heartbeat_at': snapshot.get('worker_last_heartbeat_at'),
        'worker_last_loop_at': snapshot.get('worker_last_loop_at'),
        'worker_last_exception': snapshot.get('worker_last_exception'),
        'auto_should_be_running': snapshot.get('auto_should_be_running'),
        'auto_recovery_needed': snapshot.get('auto_recovery_needed'),
        'auto_recovery_action': 'restarted worker from status check' if restarted else snapshot.get('auto_recovery_action'),
        'auto_recovery_last_at': snapshot.get('auto_recovery_last_at'),
        'auto_recovery_reason': snapshot.get('auto_recovery_reason'),
        'notes': snapshot['notes'],
    })


@app.route('/jdy-webhook/worker-control', methods=['POST'])
def jdy_webhook_worker_control():
    try:
        data = request.get_json(silent=True) or {}
        action = str(data.get('action') or '').strip().lower()
        dry_run = str(data.get('dry_run') or '').strip().lower() in ('1', 'true', 'yes', 'y', 'on')
        max_items = data.get('max_items')
        resource_filter = str(data.get('resource_filter') or '').strip()
        requested_mode = str(data.get('mode') or '').strip().lower()
        requested_queue = str(data.get('queue') or 'normal').strip().lower()
        if requested_queue not in ('normal', 'retry'):
            return jsonify({'success': False, 'error': 'queue must be normal or retry'}), 400
        snapshot = _webhook_status_snapshot(20)

        if dry_run:
            return jsonify({
                'success': True,
                'dry_run': True,
                'action': action or 'status',
                'mode': requested_mode or snapshot.get('processing_mode'),
                'would_change_state': action in ('start', 'resume', 'pause', 'stop_after_current', 'set_mode'),
                'counts': snapshot['counts'],
                'backlog_estimate': snapshot['backlog_estimate'],
                'current_rate_policy': snapshot['current_rate_policy'],
                'processing_mode': snapshot.get('processing_mode'),
                'auto_enabled': snapshot.get('auto_enabled'),
                'manual_processing_required': snapshot.get('manual_processing_required'),
                'duplicate_summary': snapshot['duplicate_summary'],
                'retry_duplicate_summary': snapshot.get('retry_duplicate_summary') or [],
                'stale_processing': snapshot['stale_processing'],
                'queue': requested_queue,
            })

        if action not in ('start', 'resume', 'pause', 'stop_after_current', 'set_mode'):
            return jsonify({'success': False, 'error': 'unsupported action; use set_mode/start/resume/pause/stop_after_current or dry_run=true'}), 400

        if action == 'set_mode':
            if str(data.get('confirm') or '') != 'WEBHOOK_MODE':
                return jsonify({'success': False, 'error': '缺少确认码 WEBHOOK_MODE，未改变 Webhook 处理模式。'}), 400
            if requested_mode not in ('paused', 'manual', 'auto'):
                return jsonify({'success': False, 'error': 'mode must be paused/manual/auto'}), 400
            if requested_mode == 'paused':
                _webhook_state.update({
                    'paused': True,
                    'processing_mode': 'paused',
                    'pause_reason': 'manual pause',
                    'stop_after_current': False,
                    'max_items_remaining': None,
                    'queue': 'normal',
                    'message': 'webhook processing mode set to paused',
                })
                _save_webhook_runtime_settings()
            elif requested_mode == 'manual':
                _webhook_state.update({
                    'paused': True,
                    'processing_mode': 'manual',
                    'pause_reason': 'manual processing required',
                    'stop_after_current': False,
                    'max_items_remaining': None,
                    'queue': 'normal',
                    'message': 'webhook processing mode set to manual',
                })
                _save_webhook_runtime_settings()
            else:
                _webhook_state.update({
                    'paused': False,
                    'processing_mode': 'auto',
                    'pause_reason': '',
                    'stop_after_current': False,
                    'max_items_remaining': None,
                    'resource_filter': '',
                    'queue': 'normal',
                    'message': 'webhook processing mode set to auto',
                })
                _save_webhook_runtime_settings()
                _start_webhook_worker()

        if action in ('start', 'resume'):
            try:
                max_items_value = int(max_items) if max_items not in (None, '') else None
            except Exception:
                max_items_value = None
            _webhook_state.update({
                'paused': False,
                'processing_mode': 'manual',
                'pause_reason': '',
                'stop_after_current': False,
                'max_items_remaining': max_items_value,
                'resource_filter': resource_filter,
                'queue': requested_queue,
                'message': f'manual webhook processing enabled ({requested_queue})',
            })
            _save_webhook_runtime_settings()
            _start_webhook_worker()
        elif action == 'pause':
            _webhook_state.update({
                'paused': True,
                'processing_mode': 'paused',
                'pause_reason': 'manual pause',
                'message': 'manual pause',
            })
            _save_webhook_runtime_settings()
        elif action == 'stop_after_current':
            _webhook_state.update({
                'stop_after_current': True,
                'message': 'will stop after current item',
            })

        snapshot = _webhook_status_snapshot(20)
        state = dict(_webhook_state)
        state['pending'] = snapshot['counts'].get('pending', 0)
        state['retry_pending'] = snapshot['counts'].get('retry_pending', 0)
        return jsonify({
            'success': True,
            'dry_run': False,
            'action': action,
            'state': state,
            'counts': snapshot['counts'],
            'backlog_estimate': snapshot['backlog_estimate'],
            'current_rate_policy': snapshot['current_rate_policy'],
            'processing_mode': snapshot.get('processing_mode'),
            'auto_enabled': snapshot.get('auto_enabled'),
            'manual_processing_required': snapshot.get('manual_processing_required'),
            'pending_count': snapshot.get('pending_count'),
            'retry_pending_count': snapshot.get('retry_pending_count'),
            'queue': state.get('queue') or 'normal',
            'current_rate_limit': snapshot.get('current_rate_limit'),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/recover-stale', methods=['POST'])
def jdy_webhook_recover_stale():
    try:
        data = request.get_json(silent=True) or {}
        action = str(data.get('action') or 'dry_run').strip().lower()
        older_than_minutes = data.get('older_than_minutes') or 30
        limit = data.get('limit') or 100

        if action == 'dry_run':
            conn = _sales_readonly_conn()
            if conn is None:
                return jsonify({
                    'success': True,
                    'dry_run': True,
                    'older_than_minutes': older_than_minutes,
                    'cutoff': '',
                    'count': 0,
                    'items': [],
                    'message': 'sales cache database not found',
                })
            try:
                rows, minutes, cutoff = _webhook_stale_processing_rows(conn, older_than_minutes, limit)
            finally:
                conn.close()
            return jsonify({
                'success': True,
                'dry_run': True,
                'older_than_minutes': minutes,
                'cutoff': cutoff,
                'count': len(rows),
                'items': rows,
            })

        if action != 'recover_to_pending':
            return jsonify({'success': False, 'error': 'unsupported action; use dry_run or recover_to_pending'}), 400
        if str(data.get('confirm') or '') != 'RECOVER_STALE_WEBHOOK':
            return jsonify({'success': False, 'error': '缺少确认码 RECOVER_STALE_WEBHOOK，未恢复卡住事件。'}), 400

        with _sales_cache_conn() as conn:
            rows, minutes, cutoff = _webhook_stale_processing_rows(conn, older_than_minutes, limit)
            normal_ids = [int(row['id']) for row in rows if row.get('recover_to_status') != 'retry_pending']
            retry_ids = [int(row['id']) for row in rows if row.get('recover_to_status') == 'retry_pending']
            if normal_ids:
                marks = ','.join('?' for _ in normal_ids)
                conn.execute(f'''
                    UPDATE webhook_events
                    SET status = 'pending',
                        error = 'recovered from stale processing',
                        processed_at = ''
                    WHERE id IN ({marks}) AND status = 'processing'
                ''', normal_ids)
            if retry_ids:
                marks = ','.join('?' for _ in retry_ids)
                conn.execute(f'''
                    UPDATE webhook_events
                    SET status = 'retry_pending',
                        error = 'recovered for retry',
                        processed_at = ''
                    WHERE id IN ({marks}) AND status = 'processing'
                ''', retry_ids)
            ids = normal_ids + retry_ids
            if ids:
                conn.commit()
        _webhook_state.update({
            'pending': _webhook_pending_count(),
            'retry_pending': _webhook_pending_count('retry'),
            'message': f'recovered {len(normal_ids)} stale webhook event(s) to pending and {len(retry_ids)} to retry_pending',
        })
        return jsonify({
            'success': True,
            'dry_run': False,
            'older_than_minutes': minutes,
            'cutoff': cutoff,
            'recovered': len(ids),
            'normal_recovered': len(normal_ids),
            'retry_recovered': len(retry_ids),
            'ids': ids,
            'message': 'stale processing events recovered to their original queue; worker was not started',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/failures', methods=['GET'])
def jdy_webhook_failures():
    try:
        snapshot = _webhook_failures_snapshot(request.args)
        return jsonify({
            'success': True,
            'total': snapshot['total'],
            'by_resource': snapshot['by_resource'],
            'by_error': snapshot['by_error'],
            'items': snapshot['items'],
            'params': snapshot['params'],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/retry-failed-dry-run', methods=['POST'])
def jdy_webhook_retry_failed_dry_run():
    try:
        data = request.get_json(silent=True) or {}
        conn = _sales_readonly_conn()
        if conn is None:
            return jsonify({'success': True, 'dry_run': True, 'count': 0, 'items': []})
        try:
            rows, params = _webhook_failed_rows(conn, data)
        finally:
            conn.close()
        return jsonify({
            'success': True,
            'dry_run': True,
            'count': len(rows),
            'items': rows,
            'params': params,
            'message': 'dry run only; no failed events were changed',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/retry-failed', methods=['POST'])
def jdy_webhook_retry_failed():
    try:
        data = request.get_json(silent=True) or {}
        if str(data.get('confirm') or '') != 'RETRY_FAILED_WEBHOOK':
            return jsonify({'success': False, 'error': '缺少确认码 RETRY_FAILED_WEBHOOK，未恢复失败事件。'}), 400
        with _sales_cache_conn() as conn:
            rows, params = _webhook_failed_rows(conn, data)
            ids = [int(row['id']) for row in rows]
            if ids:
                marks = ','.join('?' for _ in ids)
                conn.execute(f'''
                    UPDATE webhook_events
                    SET status = 'retry_pending',
                        error = 'recovered for retry',
                        processed_at = ''
                    WHERE id IN ({marks}) AND status = 'failed'
                ''', ids)
                conn.commit()
        _webhook_state.update({
            'pending': _webhook_pending_count(),
            'retry_pending': _webhook_pending_count('retry'),
            'message': f'recovered {len(ids)} failed webhook event(s) to retry_pending',
        })
        return jsonify({
            'success': True,
            'dry_run': False,
            'recovered': len(ids),
            'ids': ids,
            'params': params,
            'message': 'failed events recovered to retry_pending; worker was not started',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _webhook_archive_reason(data):
    raw = str((data or {}).get('reason') or '').strip()
    aliases = {
        'historical_account_mapping_failure_daily_compare_covers': 'historical account mapping failure; covered by daily full compare',
        'unresolved_account_missing_daily_compare_covers': 'unresolved account missing; covered by daily full compare or manual review',
        'unsupported_resource_daily_compare_covers': 'unsupported historical resource; covered by daily full compare where applicable',
        'recorded_only_historical_cleanup': 'historical recorded only failure archived; no replay needed',
    }
    if raw in aliases:
        return aliases[raw]
    if raw:
        return raw[:500]
    return aliases['historical_account_mapping_failure_daily_compare_covers']


@app.route('/jdy-webhook/archive-failed-dry-run', methods=['POST'])
def jdy_webhook_archive_failed_dry_run():
    try:
        data = request.get_json(silent=True) or {}
        conn = _sales_readonly_conn()
        if conn is None:
            return jsonify({
                'success': True,
                'dry_run': True,
                'count': 0,
                'items': [],
                'archive_reason': _webhook_archive_reason(data),
            })
        try:
            rows, params = _webhook_failed_rows(conn, data)
        finally:
            conn.close()
        return jsonify({
            'success': True,
            'dry_run': True,
            'count': len(rows),
            'items': rows,
            'params': params,
            'archive_reason': _webhook_archive_reason(data),
            'message': 'dry run only; no failed events were archived',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/archive-failed', methods=['POST'])
def jdy_webhook_archive_failed():
    try:
        data = request.get_json(silent=True) or {}
        if str(data.get('confirm') or '') != 'ARCHIVE_FAILED_WEBHOOK':
            return jsonify({'success': False, 'error': 'missing confirm ARCHIVE_FAILED_WEBHOOK; failed events were not archived'}), 400
        archive_reason = _webhook_archive_reason(data)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with _sales_cache_conn() as conn:
            rows, params = _webhook_failed_rows(conn, data)
            ids = [int(row['id']) for row in rows]
            if ids:
                marks = ','.join('?' for _ in ids)
                conn.execute(f'''
                    UPDATE webhook_events
                    SET status = 'ignored',
                        error = ?,
                        processed_at = ?
                    WHERE id IN ({marks}) AND status = 'failed'
                ''', [archive_reason, now] + ids)
                conn.commit()
        _webhook_state.update({
            'pending': _webhook_pending_count(),
            'retry_pending': _webhook_pending_count('retry'),
            'message': f'archived {len(ids)} failed webhook event(s) to ignored',
        })
        return jsonify({
            'success': True,
            'dry_run': False,
            'archived': len(ids),
            'ids': ids,
            'params': params,
            'archive_reason': archive_reason,
            'message': 'failed events archived to ignored; worker was not started',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _webhook_reclassify_retry_rows(conn, params):
    params = _webhook_failure_params(params)
    direction = 'ASC' if params['order'] == 'oldest' else 'DESC'
    rows = conn.execute(f'''
        SELECT id, account, biz_type, resource, bill_no, action, status, attempts, error, created_at, processed_at
        FROM webhook_events
        WHERE status = 'pending'
          AND LOWER(COALESCE(error, '')) LIKE '%recovered for retry%'
        ORDER BY id {direction}
        LIMIT ? OFFSET ?
    ''', (params['limit'], params['offset'])).fetchall()
    return [dict(row) for row in rows], params


@app.route('/jdy-webhook/reclassify-retry-pending-dry-run', methods=['POST'])
def jdy_webhook_reclassify_retry_pending_dry_run():
    try:
        data = request.get_json(silent=True) or {}
        conn = _sales_readonly_conn()
        if conn is None:
            return jsonify({'success': True, 'dry_run': True, 'count': 0, 'items': []})
        try:
            rows, params = _webhook_reclassify_retry_rows(conn, data)
        finally:
            conn.close()
        return jsonify({
            'success': True,
            'dry_run': True,
            'count': len(rows),
            'items': rows,
            'params': params,
            'message': 'dry run only; no pending retry events were reclassified',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/reclassify-retry-pending', methods=['POST'])
def jdy_webhook_reclassify_retry_pending():
    try:
        data = request.get_json(silent=True) or {}
        if str(data.get('confirm') or '') != 'RECLASSIFY_RETRY_PENDING':
            return jsonify({'success': False, 'error': '缺少确认码 RECLASSIFY_RETRY_PENDING，未移动重试事件。'}), 400
        with _sales_cache_conn() as conn:
            rows, params = _webhook_reclassify_retry_rows(conn, data)
            ids = [int(row['id']) for row in rows]
            if ids:
                marks = ','.join('?' for _ in ids)
                conn.execute(f'''
                    UPDATE webhook_events
                    SET status = 'retry_pending'
                    WHERE id IN ({marks})
                      AND status = 'pending'
                      AND LOWER(COALESCE(error, '')) LIKE '%recovered for retry%'
                ''', ids)
                conn.commit()
        _webhook_state.update({
            'pending': _webhook_pending_count(),
            'retry_pending': _webhook_pending_count('retry'),
            'message': f'reclassified {len(ids)} pending retry event(s) to retry_pending',
        })
        return jsonify({
            'success': True,
            'dry_run': False,
            'reclassified': len(ids),
            'ids': ids,
            'params': params,
            'message': 'pending retry events moved to retry_pending; worker was not started',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/resolve-recorded-only-dry-run', methods=['POST'])
def jdy_webhook_resolve_recorded_only_dry_run():
    try:
        data = request.get_json(silent=True) or {}
        conn = _sales_readonly_conn()
        if conn is None:
            return jsonify({'success': True, 'dry_run': True, 'count': 0, 'items': []})
        try:
            rows, params = _webhook_failed_rows(conn, data, recorded_only=True)
        finally:
            conn.close()
        return jsonify({
            'success': True,
            'dry_run': True,
            'count': len(rows),
            'items': rows,
            'params': params,
            'message': 'dry run only; no recorded-only failures were changed',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/jdy-webhook/resolve-recorded-only', methods=['POST'])
def jdy_webhook_resolve_recorded_only():
    try:
        data = request.get_json(silent=True) or {}
        if str(data.get('confirm') or '') != 'RESOLVE_RECORDED_ONLY':
            return jsonify({'success': False, 'error': '缺少确认码 RESOLVE_RECORDED_ONLY，未归档 recorded only 历史失败。'}), 400
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with _sales_cache_conn() as conn:
            rows, params = _webhook_failed_rows(conn, data, recorded_only=True)
            ids = [int(row['id']) for row in rows]
            if ids:
                marks = ','.join('?' for _ in ids)
                conn.execute(f'''
                    UPDATE webhook_events
                    SET status = 'done',
                        error = 'recorded only resolved',
                        processed_at = ?
                    WHERE id IN ({marks}) AND status = 'failed'
                ''', [now] + ids)
                conn.commit()
        return jsonify({
            'success': True,
            'dry_run': False,
            'resolved': len(ids),
            'ids': ids,
            'params': params,
            'message': 'recorded-only failed events resolved to done; worker was not started',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/sales-sync-run', methods=['POST'])
def sales_sync_run():
    try:
        wait = str(request.args.get('wait') or '').lower() in ('1', 'true', 'yes', 'y')
        if wait:
            result = _run_sales_incremental_sync(reason='manual')
            return jsonify({'success': True, 'started': True, 'result': result, 'state': _sales_sync_state})
        if _sales_sync_lock.locked() or _sales_sync_state.get('running'):
            return jsonify({'success': True, 'started': False, 'running': True, 'state': _sales_sync_state})
        t = threading.Thread(
            target=_run_sales_incremental_sync,
            kwargs={'reason': 'manual'},
            daemon=True,
            name='sales-sync-manual',
        )
        t.start()
        return jsonify({'success': True, 'started': True, 'running': True, 'state': _sales_sync_state})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'state': _sales_sync_state}), 500


@app.route('/sales-cache-cleanup', methods=['POST'])
def sales_cache_cleanup():
    try:
        data = request.get_json(force=True) or {}
        begin_date = str(data.get('begin_date') or '').strip()
        end_date = str(data.get('end_date') or '').strip()
        if not begin_date or not end_date:
            return jsonify({'success': False, 'error': 'begin_date and end_date are required'}), 400
        action = str(data.get('action') or 'delete').strip().lower()
        if action not in ('delete', 'archive'):
            action = 'delete'
        archive_path = str(data.get('archive_path') or '').strip()
        result = _cleanup_sales_cache_range(begin_date, end_date, action, archive_path)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reorder-items/import', methods=['POST'])
def reorder_items_import():
    try:
        started = time.perf_counter()
        data = request.get_json(force=True) or {}
        items = data.get('items') or []
        source = data.get('source') or {}
        order_no = str(data.get('order_number') or source.get('number') or '').strip()
        account = str(data.get('account') or source.get('account') or '').strip()
        if not items and order_no:
            order = _read_cached_sales_detail(order_no, account)
            if not order:
                return jsonify({'success': False, 'error': '本地没有这张销售单详情，请先同步销售单'}), 404
            source = {
                'type': 'sales',
                'number': order.get('number') or order_no,
                'date': order.get('date') or '',
                'customer': order.get('customerName') or '',
                'user': session.get('auth_name') or session.get('auth_user') or '',
            }
            account = order.get('account') or account
            wanted = {str(x).strip() for x in (data.get('codes') or []) if str(x).strip()}
            items = [
                e for e in (order.get('entries') or [])
                if not wanted or _reorder_entry_identity(e) in wanted
            ]
        if not items:
            return jsonify({'success': False, 'error': '请选择需要返单的商品'}), 400

        source_type = str(source.get('type') or '').strip()
        load_started = time.perf_counter()
        if source_type == 'product_catalog':
            local_ctx = {'quantities': {}, 'products': {}, 'suppliers': {}, 'supplier_by_product': {}}
        else:
            local_ctx = _reorder_batch_local_context(items, account)
        load_ms = int((time.perf_counter() - load_started) * 1000)
        build_ms = 0
        inserted = 0
        updated = 0
        skipped = 0
        created_ids = []
        created_by, created_by_name = _current_reorder_user()
        upsert_started = time.perf_counter()
        with _sales_cache_conn() as conn:
            for raw in items:
                if not isinstance(raw, dict):
                    skipped += 1
                    continue
                item_account = raw.get('account') or account
                build_item_started = time.perf_counter()
                item = _reorder_item_payload_from_entry(
                    raw, item_account, source, local_ctx=local_ctx,
                    created_by=created_by, created_by_name=created_by_name,
                )
                build_ms += int((time.perf_counter() - build_item_started) * 1000)
                if not item.get('code'):
                    skipped += 1
                    continue
                rid, is_new = _cache_insert_reorder_item(conn, item)
                created_ids.append(rid)
                if is_new:
                    inserted += 1
                else:
                    updated += 1
            conn.commit()
        db_ms = int((time.perf_counter() - upsert_started) * 1000)
        total_ms = int((time.perf_counter() - started) * 1000)
        perf = {
            'items_count': len(items),
            'load_local_cache_ms': load_ms,
            'build_payload_ms': build_ms,
            'db_upsert_ms': db_ms,
            'total_ms': total_ms,
            'live_lookup': False,
        }
        print(
            '[REORDER IMPORT] '
            f"items_count={perf['items_count']} load_local_cache_ms={load_ms} "
            f"build_payload_ms={build_ms} db_upsert_ms={db_ms} total_ms={total_ms} live_lookup=false"
        )
        return jsonify({
            'success': True,
            'inserted': inserted,
            'updated': updated,
            'skipped': skipped,
            'ids': created_ids,
            'local_only': True,
            'live_lookup': False,
            'performance': perf,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_items_import: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/reorder-suppliers', methods=['GET'])
def reorder_suppliers():
    try:
        view_mode = (request.args.get('view_mode') or 'supplier').strip().lower()
        if view_mode not in ('supplier', 'picker'):
            view_mode = 'supplier'
        status = request.args.get('status', 'pending').strip() or 'pending'
        account = request.args.get('account', '').strip()
        search = request.args.get('search', '').strip().lower()
        clauses = []
        params = []
        if status and status != 'all':
            clauses.append('status = ?')
            params.append(status)
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if search:
            clauses.append('''
                (
                    LOWER(COALESCE(supplier_name, '')) LIKE ?
                    OR LOWER(COALESCE(supplier_number, '')) LIKE ?
                    OR LOWER(COALESCE(code, '')) LIKE ?
                    OR LOWER(COALESCE(name, '')) LIKE ?
                    OR LOWER(COALESCE(created_by_name, '')) LIKE ?
                    OR LOWER(COALESCE(created_by, '')) LIKE ?
                )
            ''')
            kw = f'%{search}%'
            params.extend([kw, kw, kw, kw, kw, kw])
        where = 'WHERE ' + ' AND '.join(clauses) if clauses else ''
        with _sales_cache_conn() as conn:
            if view_mode == 'picker':
                rows = conn.execute(f'''
                    SELECT COALESCE(NULLIF(created_by, ''), 'unknown') AS created_by,
                           COALESCE(NULLIF(created_by_name, ''), '未知添加人') AS created_by_name,
                           COUNT(*) AS count,
                           SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) AS pending_count,
                           SUM(CASE WHEN status = 'confirmed' THEN 1 ELSE 0 END) AS confirmed_count,
                           SUM(CASE WHEN status IN ('completed', 'done', 'ordered') THEN 1 ELSE 0 END) AS completed_count,
                           SUM(CASE WHEN status = 'ignored' THEN 1 ELSE 0 END) AS ignored_count,
                           SUM(sales_qty_60) AS sales_qty_60,
                           SUM(stock_new) AS stock_new,
                           SUM(stock_transit) AS stock_transit,
                           SUM(stock_local) AS stock_local,
                           SUM(factory_qty) AS factory_qty,
                           SUM(suggested_qty) AS suggested_qty,
                           SUM(confirmed_qty) AS confirmed_qty,
                           SUM(confirmed_qty * COALESCE(reorder_price, 0)) AS confirmed_amount,
                           MIN(created_at) AS first_created_at,
                           MAX(created_at) AS latest_created_at,
                           MAX(updated_at) AS updated_at,
                           GROUP_CONCAT(DISTINCT account) AS accounts
                    FROM reorder_items
                    {where}
                    GROUP BY COALESCE(NULLIF(created_by, ''), 'unknown'), COALESCE(NULLIF(created_by_name, ''), '未知添加人')
                    ORDER BY latest_created_at DESC, updated_at DESC, created_by_name, created_by
                ''', params).fetchall()
            else:
                rows = conn.execute(f'''
                    SELECT supplier_number, supplier_name,
                           COUNT(*) AS count,
                           SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) AS pending_count,
                           SUM(CASE WHEN status = 'confirmed' THEN 1 ELSE 0 END) AS confirmed_count,
                           SUM(CASE WHEN status IN ('completed', 'done', 'ordered') THEN 1 ELSE 0 END) AS completed_count,
                           SUM(CASE WHEN status = 'ignored' THEN 1 ELSE 0 END) AS ignored_count,
                           SUM(sales_qty_60) AS sales_qty_60,
                           SUM(stock_new) AS stock_new,
                           SUM(stock_transit) AS stock_transit,
                           SUM(stock_local) AS stock_local,
                           SUM(factory_qty) AS factory_qty,
                           SUM(suggested_qty) AS suggested_qty,
                           SUM(confirmed_qty) AS confirmed_qty,
                           SUM(confirmed_qty * COALESCE(reorder_price, 0)) AS confirmed_amount,
                           MIN(created_at) AS first_created_at,
                           MAX(created_at) AS latest_created_at,
                           MAX(updated_at) AS updated_at,
                           GROUP_CONCAT(DISTINCT account) AS accounts
                    FROM reorder_items
                    {where}
                    GROUP BY COALESCE(supplier_number, ''), COALESCE(supplier_name, '')
                    ORDER BY latest_created_at DESC, updated_at DESC, supplier_name, supplier_number
                ''', params).fetchall()
        items = []
        for row in rows:
            if view_mode == 'picker':
                created_by = row['created_by'] or 'unknown'
                created_by_name = row['created_by_name'] or '未知添加人'
                items.append({
                    'viewMode': 'picker',
                    'groupKey': f'picker:{created_by}',
                    'supplierKey': f'picker:{created_by}',
                    'pickerKey': created_by,
                    'pickerName': created_by_name,
                    'createdBy': created_by,
                    'created_by': created_by,
                    'createdByName': created_by_name,
                    'created_by_name': created_by_name,
                    'supplierName': created_by_name,
                    'supplier_name': created_by_name,
                    'supplierNumber': '',
                    'supplier_number': '',
                    'count': row['count'] or 0,
                    'itemCount': row['count'] or 0,
                    'item_count': row['count'] or 0,
                    'pendingCount': row['pending_count'] or 0,
                    'confirmedCount': row['confirmed_count'] or 0,
                    'completedCount': row['completed_count'] or 0,
                    'ignoredCount': row['ignored_count'] or 0,
                    'salesQty60': _num(row['sales_qty_60']),
                    'stockNew': _num(row['stock_new']),
                    'stockTransit': _num(row['stock_transit']),
                    'stockLocal': _num(row['stock_local']),
                    'factoryQty': _num(row['factory_qty']),
                    'suggestedQty': _num(row['suggested_qty']),
                    'confirmedQty': _num(row['confirmed_qty']),
                    'confirmedAmount': _num(row['confirmed_amount']),
                    'latestCreatedAt': row['latest_created_at'] or '',
                    'createdAt': row['first_created_at'] or '',
                    'updatedAt': row['updated_at'] or '',
                    'accounts': row['accounts'] or '',
                })
                continue
            supplier_number = row['supplier_number'] or ''
            supplier_name = row['supplier_name'] or '未识别供应商'
            items.append({
                'viewMode': 'supplier',
                'groupKey': supplier_number or supplier_name,
                'supplierKey': supplier_number or supplier_name,
                'supplierNumber': supplier_number,
                'supplier_number': supplier_number,
                'supplierName': supplier_name,
                'supplier_name': supplier_name,
                'count': row['count'] or 0,
                'itemCount': row['count'] or 0,
                'item_count': row['count'] or 0,
                'pendingCount': row['pending_count'] or 0,
                'confirmedCount': row['confirmed_count'] or 0,
                'completedCount': row['completed_count'] or 0,
                'ignoredCount': row['ignored_count'] or 0,
                'salesQty60': _num(row['sales_qty_60']),
                'stockNew': _num(row['stock_new']),
                'stockTransit': _num(row['stock_transit']),
                'stockLocal': _num(row['stock_local']),
                'factoryQty': _num(row['factory_qty']),
                'suggestedQty': _num(row['suggested_qty']),
                'confirmedQty': _num(row['confirmed_qty']),
                'confirmedAmount': _num(row['confirmed_amount']),
                'latestCreatedAt': row['latest_created_at'] or '',
                'createdAt': row['first_created_at'] or '',
                'updatedAt': row['updated_at'] or '',
                'accounts': row['accounts'] or '',
            })
        return jsonify({'success': True, 'view_mode': view_mode, 'list': items, 'total': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reorder-items', methods=['GET'])
def reorder_items():
    try:
        view_mode = (request.args.get('view_mode') or 'supplier').strip().lower()
        if view_mode not in ('supplier', 'picker'):
            view_mode = 'supplier'
        status = request.args.get('status', 'pending').strip() or 'pending'
        account = request.args.get('account', '').strip()
        supplier = request.args.get('supplier', '').strip()
        supplier_number = request.args.get('supplier_number', '').strip()
        supplier_name = request.args.get('supplier_name', '').strip()
        picker_key = request.args.get('picker_key', '').strip()
        created_by = request.args.get('created_by', '').strip()
        created_by_name = request.args.get('created_by_name', '').strip()
        search = request.args.get('search', '').strip().lower()
        clauses = []
        params = []
        if status and status != 'all':
            clauses.append('status = ?')
            params.append(status)
        if account and account != 'all':
            clauses.append('account = ?')
            params.append(account)
        if view_mode == 'picker':
            picker = created_by or picker_key
            if picker.startswith('picker:'):
                picker = picker.split(':', 1)[1]
            if picker and picker != 'unknown':
                clauses.append('created_by = ?')
                params.append(picker)
            elif created_by_name:
                clauses.append('created_by_name = ?')
                params.append(created_by_name)
            elif picker == 'unknown':
                clauses.append("(COALESCE(created_by, '') = '' OR created_by = 'unknown')")
        else:
            if supplier_number:
                clauses.append('supplier_number = ?')
                params.append(supplier_number)
            elif supplier_name:
                clauses.append('supplier_name = ?')
                params.append(supplier_name)
            elif supplier:
                clauses.append('(supplier_number = ? OR supplier_name = ?)')
                params.extend([supplier, supplier])
        if search:
            clauses.append('''
                (
                    LOWER(COALESCE(code, '')) LIKE ?
                    OR LOWER(COALESCE(name, '')) LIKE ?
                    OR LOWER(COALESCE(spec, '')) LIKE ?
                    OR LOWER(COALESCE(barcode, '')) LIKE ?
                    OR LOWER(COALESCE(supplier_name, '')) LIKE ?
                    OR LOWER(COALESCE(supplier_number, '')) LIKE ?
                    OR LOWER(COALESCE(created_by_name, '')) LIKE ?
                )
            ''')
            kw = f'%{search}%'
            params.extend([kw, kw, kw, kw, kw, kw, kw])
        where = 'WHERE ' + ' AND '.join(clauses) if clauses else ''
        with _sales_cache_conn() as conn:
            rows = conn.execute(f'''
                SELECT * FROM reorder_items
                {where}
                ORDER BY supplier_name, supplier_number, updated_at DESC, id DESC
            ''', params).fetchall()
        items = [_reorder_row_to_item(row) for row in rows]
        summary = {
            'count': len(items),
            'salesQty60': sum(_num(x.get('salesQty60')) for x in items),
            'stockNew': sum(_num(x.get('stockNew')) for x in items),
            'stockTransit': sum(_num(x.get('stockTransit')) for x in items),
            'stockLocal': sum(_num(x.get('stockLocal')) for x in items),
            'factoryQty': sum(_num(x.get('factoryQty')) for x in items),
            'suggestedQty': sum(_num(x.get('suggestedQty')) for x in items),
            'confirmedQty': sum(_num(x.get('confirmedQty')) for x in items),
            'confirmedAmount': sum(_num(x.get('confirmedQty')) * _num(x.get('reorderPrice')) for x in items),
        }
        return jsonify({'success': True, 'view_mode': view_mode, 'list': items, 'summary': summary})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reorder-recent-styles', methods=['GET'])
def reorder_recent_styles():
    supplier = (request.args.get('supplier') or '').strip()
    account = (request.args.get('account') or '').strip()
    limit = max(1, min(int(request.args.get('limit') or 50), 100))
    if not supplier:
        return jsonify({'success': False, 'error': '缺少 supplier 参数'}), 400
    seen = set()
    items = []
    with _sales_cache_conn() as conn:
        queries = [
            ('accessory_purchase_orders', '''
                SELECT account, number, data_json FROM accessory_purchase_orders
                WHERE supplier_name = ? {account_where}
                ORDER BY date DESC, updated_at DESC
                LIMIT 200
            '''),
            ('purchase_inbounds', '''
                SELECT account, number, data_json FROM purchase_inbounds
                WHERE supplier_name = ? {account_where}
                ORDER BY date DESC, updated_at DESC
                LIMIT 200
            '''),
        ]
        for table, sql_tpl in queries:
            account_where = 'AND account = ?' if account else ''
            params = [supplier] + ([account] if account else [])
            rows = conn.execute(sql_tpl.format(account_where=account_where), params).fetchall()
            for row in rows:
                try:
                    order = json.loads(row['data_json'] or '{}')
                except Exception:
                    continue
                supplier_info = {
                    'supplierName': supplier,
                    'supplierNumber': _first_value(order, ['supplierNumber', 'supplierNo', 'vendorNumber'], ''),
                }
                for entry in _reorder_entries_from_order(order):
                    if not isinstance(entry, dict):
                        continue
                    item = _reorder_recent_style_from_entry(entry, row['account'] or account, supplier_info, order, conn)
                    if not item:
                        continue
                    key = (item.get('account') or '', item.get('code') or '')
                    if key in seen:
                        continue
                    seen.add(key)
                    item['sourceTable'] = table
                    items.append(item)
                    if len(items) >= limit:
                        return jsonify({'success': True, 'list': items, 'total': len(items), 'supplier': supplier})
    return jsonify({'success': True, 'list': items, 'total': len(items), 'supplier': supplier})


def _normalize_reorder_status(status):
    status = str(status or '').strip()
    aliases = {'done': 'completed', 'ordered': 'completed'}
    return aliases.get(status, status)


@app.route('/reorder-items/<int:item_id>', methods=['POST', 'PATCH'])
def reorder_item_update(item_id):
    try:
        data = request.get_json(force=True) or {}
        allowed = {'pending', 'confirmed', 'completed', 'ignored'}
        status = _normalize_reorder_status(data.get('status'))
        if status and status not in allowed:
            return jsonify({'success': False, 'error': '状态无效'}), 400
        confirmed_qty = _num(data.get('confirmed_qty'))
        reorder_price = _num(data.get('reorder_price'))
        note = str(data.get('note') or '').strip()
        updates = []
        params = []
        if status:
            updates.append('status = ?')
            params.append(status)
        if 'confirmed_qty' in data:
            updates.append('confirmed_qty = ?')
            params.append(confirmed_qty)
        if 'reorder_price' in data:
            updates.append('reorder_price = ?')
            params.append(reorder_price)
        if 'note' in data:
            updates.append('note = ?')
            params.append(note)
        if not updates:
            return jsonify({'success': True, 'updated': False})
        updates.append('updated_at = ?')
        params.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        params.append(item_id)
        with _sales_cache_conn() as conn:
            cur = conn.execute(f"UPDATE reorder_items SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
            row = conn.execute('SELECT * FROM reorder_items WHERE id = ?', (item_id,)).fetchone()
        return jsonify({
            'success': True,
            'updated': bool(cur.rowcount),
            'item': _reorder_row_to_item(row) if row else None,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reorder-item-attachments/<int:item_id>', methods=['GET'])
def reorder_item_attachments(item_id):
    try:
        config_begin, config_new_logic = _attachment_config_dates()
        conn = _sales_readonly_conn()
        if not conn:
            return jsonify({
                'success': False,
                'error': '本地销售缓存不存在',
                'local_only': True,
                'live_lookup': False,
                'attachment_switch_date': ATTACHMENT_SWITCH_DATE,
                'source_rule': ATTACHMENT_SOURCE_RULE,
                'config_factory_purchase_begin_date': config_begin,
                'config_factory_new_logic_begin_date': config_new_logic,
                'config_matches_attachment_rule': _attachment_config_matches_rule(config_begin, config_new_logic),
            }), 404
        try:
            if not _attachment_has_table(conn, 'reorder_items'):
                return jsonify({
                    'success': False,
                    'error': '本地返单表不存在',
                    'local_only': True,
                    'live_lookup': False,
                    'attachment_switch_date': ATTACHMENT_SWITCH_DATE,
                    'source_rule': ATTACHMENT_SOURCE_RULE,
                    'config_factory_purchase_begin_date': config_begin,
                    'config_factory_new_logic_begin_date': config_new_logic,
                    'config_matches_attachment_rule': _attachment_config_matches_rule(config_begin, config_new_logic),
                }), 404
            row = conn.execute('SELECT * FROM reorder_items WHERE id = ?', (item_id,)).fetchone()
            if not row:
                return jsonify({
                    'success': False,
                    'error': '返单商品不存在',
                    'attachment_switch_date': ATTACHMENT_SWITCH_DATE,
                    'source_rule': ATTACHMENT_SOURCE_RULE,
                    'config_factory_purchase_begin_date': config_begin,
                    'config_factory_new_logic_begin_date': config_new_logic,
                    'config_matches_attachment_rule': _attachment_config_matches_rule(config_begin, config_new_logic),
                }), 404
            item = _reorder_row_to_item(row)
            account = str(item.get('account') or '').strip()
            source_number = str(item.get('source_number') or item.get('sourceNumber') or '').strip()
            source_date = str(item.get('source_date') or item.get('sourceDate') or '').strip()[:10]
            legacy_query_payload = _read_local_attachment_source(conn, {
                'account': account,
                'source_type': item.get('source_type') or item.get('sourceType') or '',
                'source_number': source_number,
                'source_date': source_date,
                'product_code': item.get('code') or '',
                'supplier_number': item.get('supplier_number') or item.get('supplierNumber') or '',
                'supplier_name': item.get('supplier_name') or item.get('supplierName') or '',
            })
            attachments, history_diagnostics = _read_historical_purchase_attachments(conn, item, limit=80)
            preferred_source = 'historical_purchase_documents'
            diagnostics = history_diagnostics + [
                '原返单 source_number 仅用于高级诊断，不作为附件主匹配依据',
                *legacy_query_payload.get('diagnostics', []),
            ]
            candidate_sources = legacy_query_payload.get('candidate_sources', [])
            source_confidence = 'historical_product_match' if attachments else 'unknown'
            message = '已读取历史采购附件' if attachments else '本地历史采购附件库中暂未找到该商品的采购附件。请先在 设置 -> 附件设置 中执行历史采购附件补齐。'
        finally:
            conn.close()

        return jsonify({
            'success': True,
            'local_only': True,
            'live_lookup': False,
            'item': {
                'id': item.get('id'),
                'account': item.get('account') or '',
                'code': item.get('code') or '',
                'name': item.get('name') or '',
                'supplier_number': item.get('supplier_number') or item.get('supplierNumber') or '',
                'supplier_name': item.get('supplier_name') or item.get('supplierName') or '',
                'source_type': item.get('source_type') or item.get('sourceType') or '',
                'source_number': source_number,
                'source_date': source_date,
            },
            'attachment_switch_date': ATTACHMENT_SWITCH_DATE,
            'config_factory_purchase_begin_date': config_begin,
            'config_factory_new_logic_begin_date': config_new_logic,
            'config_matches_attachment_rule': _attachment_config_matches_rule(config_begin, config_new_logic),
            'preferred_source': preferred_source,
            'source_rule': ATTACHMENT_SOURCE_RULE,
            'source_confidence': source_confidence,
            'candidate_sources': candidate_sources,
            'attachments': attachments,
            'diagnostics': diagnostics,
            'message': message or ('已读取本地附件' if attachments else '本地未缓存附件'),
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_item_attachments: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/attachments/source', methods=['GET'])
def attachments_source():
    try:
        conn = _sales_readonly_conn()
        if not conn:
            return jsonify({
                'success': True,
                'local_only': True,
                'live_lookup': False,
                'source': _attachment_source_obj(
                    account=request.args.get('account', ''),
                    source_type=request.args.get('source_type', ''),
                    source_number=request.args.get('source_number', ''),
                    source_date=request.args.get('source_date', ''),
                    product_code=request.args.get('product_code', ''),
                    supplier_number=request.args.get('supplier_number', ''),
                    supplier_name=request.args.get('supplier_name', ''),
                ),
                'preferred_source': '',
                'source_rule': ATTACHMENT_SOURCE_RULE,
                'source_confidence': 'unknown',
                'candidate_sources': [],
                'attachments': [],
                'diagnostics': [
                    '本地销售缓存不存在，无法查询附件元数据',
                    '本接口只读本地 SQLite，不调用 JDY，不下载附件，不写数据库',
                ],
                'message': '本地未缓存附件',
            })
        try:
            payload = _read_local_attachment_source(conn, {
                'account': request.args.get('account', ''),
                'source_type': request.args.get('source_type', ''),
                'source_number': request.args.get('source_number', ''),
                'source_date': request.args.get('source_date', ''),
                'product_code': request.args.get('product_code', ''),
                'supplier_number': request.args.get('supplier_number', ''),
                'supplier_name': request.args.get('supplier_name', ''),
            })
        finally:
            conn.close()
        return jsonify(payload)
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] attachments_source: {tb}')
        return jsonify({'success': False, 'local_only': True, 'live_lookup': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/attachments/refresh-dry-run', methods=['POST'])
def attachments_refresh_dry_run():
    try:
        data = request.get_json(silent=True) or {}
        mode = str(data.get('mode') or 'local_candidates_only').strip()
        if mode != 'local_candidates_only':
            return jsonify({
                'success': False,
                'local_only': True,
                'live_lookup': False,
                'would_call_jdy': False,
                'would_download': False,
                'error': '当前只支持 local_candidates_only dry-run',
            }), 400
        conn = _sales_readonly_conn()
        if not conn:
            return jsonify({
                'success': True,
                'local_only': True,
                'live_lookup': False,
                'would_call_jdy': False,
                'would_download': False,
                'candidates': [],
                'summary': {'total_candidates': 0, 'purchase_order': 0, 'purchase_inbound': 0, 'with_local_attachment_fields': 0, 'limit': int(data.get('limit') or 50)},
                'warnings': ['本地销售缓存不存在，dry-run 未扫描任何候选来源'],
            })
        try:
            payload = _read_local_attachment_candidates(conn, {
                'account': data.get('account') or '',
                'source_type': data.get('source_type') or '',
                'date_from': data.get('date_from') or '',
                'date_to': data.get('date_to') or '',
                'limit': data.get('limit') or 50,
            })
        finally:
            conn.close()
        return jsonify(payload)
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] attachments_refresh_dry_run: {tb}')
        return jsonify({
            'success': False,
            'local_only': True,
            'live_lookup': False,
            'would_call_jdy': False,
            'would_download': False,
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/attachments/refresh', methods=['POST'])
def attachments_refresh():
    try:
        data = request.get_json(silent=True) or {}
        if str(data.get('confirm') or '').strip() != 'REFRESH_ATTACHMENT_METADATA_ONLY':
            return jsonify({
                'success': False,
                'metadata_only': True,
                'would_download': False,
                'local_file_download': False,
                'called_jdy': False,
                'error': '必须确认 confirm = REFRESH_ATTACHMENT_METADATA_ONLY',
            }), 400
        payload = _refresh_attachment_metadata_for_candidates({
            'account': data.get('account') or '',
            'source_type': data.get('source_type') or '',
            'date_from': data.get('date_from') or '',
            'date_to': data.get('date_to') or '',
            'limit': _attachment_limit(data.get('limit'), default=20, maximum=50),
        })
        status = 200 if payload.get('success') else 400
        return jsonify(payload), status
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] attachments_refresh: {tb}')
        return jsonify({
            'success': False,
            'metadata_only': True,
            'would_download': False,
            'local_file_download': False,
            'called_jdy': False,
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/attachments/local-file/<path:relpath>', methods=['GET'])
def attachments_local_file(relpath):
    try:
        root = os.path.abspath(_attachment_download_root())
        full = os.path.abspath(os.path.join(root, relpath or ''))
        if not full.startswith(root + os.sep) and full != root:
            return jsonify({'success': False, 'error': '附件路径无效'}), 400
        if not os.path.exists(full) or not os.path.isfile(full):
            return jsonify({'success': False, 'error': '本地附件文件不存在'}), 404
        return send_file(full)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/attachments/history-backfill-dry-run', methods=['POST'])
def attachments_history_backfill_dry_run():
    try:
        data = request.get_json(silent=True) or {}
        sample_jdy = bool(data.get('sample_jdy'))
        plan = _history_backfill_plan(data, sample_jdy=sample_jdy)
        warnings = [
            'dry-run 默认不写库、不下载附件',
            '历史附件补齐会按商品编号建立 purchase_attachment_items，用于返单附件匹配',
        ]
        sample = []
        called_jdy = False
        if sample_jdy:
            for account_name, cli in _attachment_account_clients(plan['account'] or 'all')[:1]:
                for flow in plan['flows'][:1]:
                    try:
                        called_jdy = True
                        res = _history_fetch_flow(cli, flow, page=1, page_size=min(plan['page_size'], 5))
                        for row in (res.get('list') or [])[:5]:
                            sample.append({
                                'account': account_name,
                                'source_type': flow['source_type'],
                                'source_number': str(_first_value(row, ['number', 'billNo', 'billNumber', 'id'], '') or ''),
                                'source_date': str(_first_value(row, ['date', 'billDate', 'orderDate'], '') or '')[:10],
                                'attachment_count': len(_extract_purchase_attachments(row)),
                                'entry_count': len(_reorder_entries_from_order(row)),
                                'raw_keys': sorted(str(k) for k in row.keys())[:40],
                            })
                    except Exception as e:
                        warnings.append(f'抽样检查失败：{_short_sync_error(e)}')
                    break
                break
        return jsonify({
            'success': True,
            'dry_run': True,
            'would_call_jdy': bool(sample_jdy),
            'called_jdy': called_jdy,
            'would_download': False,
            'plan': plan,
            'sample': sample,
            'warnings': warnings,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] attachments_history_backfill_dry_run: {tb}')
        return jsonify({'success': False, 'dry_run': True, 'error': str(e), 'traceback': tb}), 500


@app.route('/attachments/history-backfill', methods=['POST'])
def attachments_history_backfill():
    try:
        data = request.get_json(silent=True) or {}
        if str(data.get('confirm') or '').strip() != 'BACKFILL_PURCHASE_ATTACHMENTS':
            return jsonify({
                'success': False,
                'called_jdy': False,
                'download_files': bool(data.get('download_files', True)),
                'metadata_written': False,
                'error': '必须确认 confirm = BACKFILL_PURCHASE_ATTACHMENTS',
            }), 400
        payload = _run_history_backfill(data)
        return jsonify(payload), (200 if payload.get('success') else 400)
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] attachments_history_backfill: {tb}')
        return jsonify({
            'success': False,
            'called_jdy': False,
            'download_files': bool((request.get_json(silent=True) or {}).get('download_files', True)),
            'metadata_written': False,
            'error': str(e),
            'traceback': tb,
        }), 500


@app.route('/attachments/status', methods=['GET'])
def attachments_status():
    try:
        tables = {
            'bill_attachments': {'exists': False, 'count': 0, 'columns': []},
            'purchase_attachment_items': {'exists': False, 'count': 0, 'columns': []},
            'purchase_inbound_attachments': {'exists': False, 'count': 0, 'columns': []},
            'purchase_inbounds': {'exists': False, 'count': 0, 'columns': []},
            'accessory_purchase_orders': {'exists': False, 'count': 0, 'columns': []},
        }
        conn = _sales_readonly_conn()
        db_exists = bool(conn)
        if conn:
            try:
                for name in list(tables.keys()):
                    tables[name] = _attachment_table_info(conn, name)
            finally:
                conn.close()
        has_local_path_fields = any(
            'local_path' in (info.get('columns') or [])
            for info in tables.values()
        )
        return jsonify({
            'success': True,
            'local_only': True,
            'live_lookup': False,
            'db_exists': db_exists,
            'db_path': _SALES_CACHE_DB,
            'switch_date': ATTACHMENT_SWITCH_DATE,
            'source_rule': ATTACHMENT_SOURCE_RULE,
            'tables': tables,
            'download': {
                'has_local_path_fields': has_local_path_fields,
                'download_function_ready': False,
                'message': '当前仅支持本地元数据查看，附件下载需后续确认官方接口',
            },
            'notes': [
                '普通页面查看附件只读本地缓存，不会自动调用 JDY',
                '采购订单附件仅从本地 accessory_purchase_orders.data_json 中提取明确附件字段',
                '商品 imageUrl 不会被当作采购订单附件',
            ],
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] attachments_status: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/reorder-product-history/<path:code>', methods=['GET'])
def reorder_product_history(code):
    try:
        account = request.args.get('account', '').strip()
        date_from = (request.args.get('date_from') or '').strip()[:10]
        date_to = (request.args.get('date_to') or '').strip()[:10]
        local_only = request.args.get('local_only', '1').lower() not in ('0', 'false', 'no')
        items = _read_cached_purchase_history(account, code, limit=30, date_from=date_from, date_to=date_to)
        year = datetime.now().strftime('%Y')
        year_items = [x for x in items if str(x.get('source_date') or x.get('date') or '').startswith(year)]
        return jsonify({
            'success': True,
            'code': code,
            'local_only': True,
            'live_lookup': False,
            'would_call_jdy': False,
            'date_from': date_from,
            'date_to': date_to,
            'count': len(items),
            'records': items,
            'list': items,
            'history': items,
            'yearCount': len(year_items),
            'latest': items[0] if items else None,
            'source': 'local_cache',
            'message': '' if items else '本地采购历史未缓存',
            'liveRefreshSkipped': True,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_product_history: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


def _reorder_generated_dir():
    path = os.path.join(_SALES_CACHE_DIR, 'reorder_generated')
    os.makedirs(path, exist_ok=True)
    return path


def _reorder_generate_html(batch_no, mode, items):
    title = f'返单 {batch_no}'
    groups = {}
    for item in items:
        if mode == 'picker':
            key = item.get('createdByName') or '未知添加人'
        else:
            key = item.get('supplierName') or item.get('supplier_name') or '未识别供应商'
        groups.setdefault(key, []).append(item)
    group_html = []
    for group_name, rows in groups.items():
        total_amount = sum(
            (_num(row.get('confirmedQty')) or _num(row.get('suggestedQty'))) * _num(row.get('reorderPrice'))
            for row in rows
        )
        body = []
        for idx, row in enumerate(rows, 1):
            qty = _num(row.get('confirmedQty')) or _num(row.get('suggestedQty'))
            price = _num(row.get('reorderPrice'))
            amount = qty * price
            img = row.get('imageUrl') or ''
            img_html = f'<img src="{html.escape(img)}" alt="">' if img else '<span class="no-img">无图</span>'
            body.append(f'''
                <tr>
                  <td>{idx}</td>
                  <td class="img-cell">{img_html}</td>
                  <td>{html.escape(str(row.get('code') or ''))}</td>
                  <td>{html.escape(str(row.get('name') or ''))}</td>
                  <td>{html.escape(str(row.get('spec') or ''))}</td>
                  <td>{html.escape(str(row.get('barcode') or ''))}</td>
                  <td>{html.escape(str(row.get('supplierName') or row.get('supplier_name') or ''))}</td>
                  <td class="num">{qty:g}</td>
                  <td class="num">{price:.2f}</td>
                  <td class="num">{amount:.2f}</td>
                  <td>{html.escape(str(row.get('note') or ''))}</td>
                  <td>{html.escape(str(row.get('createdByName') or '未知添加人'))}</td>
                </tr>
            ''')
        group_html.append(f'''
          <section>
            <h2>{html.escape(str(group_name))}</h2>
            <div class="meta">共 {len(rows)} 款，金额 {total_amount:.2f}</div>
            <table>
              <thead>
                <tr>
                  <th>#</th><th>图片</th><th>商品编号</th><th>商品名称</th><th>规格</th><th>条码</th>
                  <th>供应商</th><th>数量</th><th>单价</th><th>金额</th><th>备注</th><th>添加人</th>
                </tr>
              </thead>
              <tbody>{''.join(body)}</tbody>
            </table>
          </section>
        ''')
    return f'''<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>{html.escape(title)}</title>
  <style>
    body {{ font-family: "Microsoft YaHei", Arial, sans-serif; margin: 24px; color: #202124; }}
    h1 {{ font-size: 22px; margin: 0 0 8px; }}
    h2 {{ font-size: 18px; margin: 24px 0 6px; }}
    .meta {{ color: #5f6368; font-size: 13px; margin-bottom: 10px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; page-break-inside: auto; }}
    th, td {{ border: 1px solid #d9dee8; padding: 6px 7px; vertical-align: middle; }}
    th {{ background: #f5f7fb; text-align: left; }}
    .num {{ text-align: right; white-space: nowrap; }}
    .img-cell {{ width: 64px; text-align: center; }}
    .img-cell img {{ max-width: 56px; max-height: 56px; object-fit: contain; }}
    .no-img {{ color: #98a1b2; }}
    @media print {{ body {{ margin: 10mm; }} .no-print {{ display: none; }} }}
  </style>
</head>
<body>
  <button class="no-print" onclick="window.print()">打印</button>
  <h1>{html.escape(title)}</h1>
  <div class="meta">模式：{html.escape(mode)}；生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}；本文件仅为本地返单打印页，不会提交精斗云采购订单。</div>
  {''.join(group_html)}
</body>
</html>'''


def _normalize_reorder_sync_status(value):
    value = str(value or '').strip()
    return value or 'not_synced'


def _reorder_batch_summary(batch_no, items):
    sync_values = [_normalize_reorder_sync_status(x.get('syncStatus') or x.get('sync_status')) for x in items]
    sync_status = 'synced' if sync_values and all(x == 'synced' for x in sync_values) else (
        'failed' if any(x == 'failed' for x in sync_values) else 'not_synced'
    )
    suppliers = sorted({
        str(x.get('supplierName') or x.get('supplier_name') or '未识别供应商').strip() or '未识别供应商'
        for x in items
    })
    created_by_names = sorted({
        str(x.get('createdByName') or x.get('created_by_name') or '未知添加人').strip() or '未知添加人'
        for x in items
    })
    total_qty = sum((_num(x.get('confirmedQty')) or _num(x.get('suggestedQty'))) for x in items)
    total_amount = sum(
        (_num(x.get('confirmedQty')) or _num(x.get('suggestedQty'))) * _num(x.get('reorderPrice'))
        for x in items
    )
    generated_at = max((str(x.get('generatedAt') or x.get('generated_at') or '') for x in items), default='')
    synced_at = max((str(x.get('syncedAt') or x.get('synced_at') or '') for x in items), default='')
    jdy_order_no = ', '.join(sorted({
        str(x.get('jdyOrderNo') or x.get('jdy_order_no') or '').strip()
        for x in items if str(x.get('jdyOrderNo') or x.get('jdy_order_no') or '').strip()
    }))
    preview_items = [{
        'id': x.get('id'),
        'code': x.get('code') or '',
        'name': x.get('name') or '',
        'supplier_name': x.get('supplierName') or x.get('supplier_name') or '',
        'qty': _num(x.get('confirmedQty')) or _num(x.get('suggestedQty')),
        'price': _num(x.get('reorderPrice')),
    } for x in items[:5]]
    return {
        'batch_no': batch_no,
        'generated_at': generated_at,
        'sync_status': sync_status,
        'jdy_order_no': jdy_order_no,
        'synced_at': synced_at,
        'supplier_count': len(suppliers),
        'item_count': len(items),
        'total_qty': total_qty,
        'total_amount': total_amount,
        'created_by_names': created_by_names,
        'suppliers': suppliers,
        'preview_items': preview_items,
    }


def _read_reorder_batch_items(batch_no='', args=None):
    args = args or {}
    clauses = ["COALESCE(generated_batch_no, '') <> ''"]
    params = []
    if batch_no:
        clauses.append('generated_batch_no = ?')
        params.append(batch_no)
    sync_status = str(args.get('sync_status') or 'all').strip().lower()
    if sync_status == 'not_synced':
        clauses.append("COALESCE(NULLIF(sync_status, ''), 'not_synced') = 'not_synced'")
    elif sync_status in ('synced', 'failed'):
        clauses.append("COALESCE(NULLIF(sync_status, ''), 'not_synced') = ?")
        params.append(sync_status)
    date_from = str(args.get('date_from') or '').strip()[:10]
    date_to = str(args.get('date_to') or '').strip()[:10]
    if date_from:
        clauses.append('generated_at >= ?')
        params.append(date_from)
    if date_to:
        clauses.append('generated_at <= ?')
        params.append(date_to + ' 23:59:59')
    search = str(args.get('search') or '').strip().lower()
    if search:
        clauses.append('''
            (
                LOWER(COALESCE(generated_batch_no, '')) LIKE ?
                OR LOWER(COALESCE(jdy_order_no, '')) LIKE ?
                OR LOWER(COALESCE(supplier_name, '')) LIKE ?
                OR LOWER(COALESCE(code, '')) LIKE ?
                OR LOWER(COALESCE(name, '')) LIKE ?
                OR LOWER(COALESCE(created_by_name, '')) LIKE ?
            )
        ''')
        kw = f'%{search}%'
        params.extend([kw, kw, kw, kw, kw, kw])
    where = ' AND '.join(clauses)
    with _sales_cache_conn() as conn:
        rows = conn.execute(f'''
            SELECT * FROM reorder_items
            WHERE {where}
            ORDER BY generated_at DESC, generated_batch_no DESC, supplier_name, id
        ''', params).fetchall()
    return [_reorder_row_to_item(row) for row in rows]


@app.route('/reorder-batches', methods=['GET'])
def reorder_batches():
    try:
        items = _read_reorder_batch_items(args=request.args)
        grouped = {}
        for item in items:
            batch_no = item.get('generatedBatchNo') or item.get('generated_batch_no') or ''
            if not batch_no:
                continue
            grouped.setdefault(batch_no, []).append(item)
        batches = [_reorder_batch_summary(batch_no, rows) for batch_no, rows in grouped.items()]
        batches.sort(key=lambda x: (x.get('generated_at') or '', x.get('batch_no') or ''), reverse=True)
        return jsonify({'success': True, 'batches': batches, 'total': len(batches)})
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_batches: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/reorder-batches/<path:batch_no>', methods=['GET'])
def reorder_batch_detail(batch_no):
    try:
        items = _read_reorder_batch_items(batch_no=batch_no, args={'sync_status': 'all'})
        if not items:
            return jsonify({'success': False, 'error': '本地返单批次不存在'}), 404
        supplier_groups = {}
        for item in items:
            key = item.get('supplierName') or item.get('supplier_name') or '未识别供应商'
            supplier_groups.setdefault(key, []).append(item)
        groups = [
            {
                'supplier_name': key,
                'item_count': len(rows),
                'total_qty': sum((_num(x.get('confirmedQty')) or _num(x.get('suggestedQty'))) for x in rows),
                'total_amount': sum(
                    (_num(x.get('confirmedQty')) or _num(x.get('suggestedQty'))) * _num(x.get('reorderPrice'))
                    for x in rows
                ),
                'items': rows,
            }
            for key, rows in supplier_groups.items()
        ]
        batch = _reorder_batch_summary(batch_no, items)
        return jsonify({
            'success': True,
            'batch': batch,
            'items': items,
            'supplier_groups': groups,
            'total_amount': batch['total_amount'],
            'sync_status': batch['sync_status'],
            'jdy_order_no': batch['jdy_order_no'],
            'generated_at': batch['generated_at'],
            'created_by_names': batch['created_by_names'],
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_batch_detail: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/reorder-batches/<path:batch_no>/print', methods=['GET'])
def reorder_batch_print(batch_no):
    try:
        items = _read_reorder_batch_items(batch_no=batch_no, args={'sync_status': 'all'})
        if not items:
            return '<h2>本地返单批次不存在</h2>', 404, {'Content-Type': 'text/html; charset=utf-8'}
        mode = request.args.get('mode') or 'supplier'
        batch = _reorder_batch_summary(batch_no, items)
        html_text = _reorder_generate_html(batch_no, mode, items)
        html_text = html_text.replace(
            '本文件仅为本地返单打印页，不会提交精斗云采购订单。',
            f"同步状态：{batch['sync_status']}；JDY单号：{batch['jdy_order_no'] or '未同步'}；本文件仅为本地返单打印页，不会提交精斗云采购订单。",
            1,
        )
        return html_text, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_batch_print: {tb}')
        return f'<pre>{html.escape(str(e))}</pre>', 500, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/reorder-batches/<path:batch_no>/sync-jdy', methods=['POST'])
def reorder_batch_sync_jdy(batch_no):
    return jsonify({
        'success': False,
        'batch_no': batch_no,
        'called_jdy': False,
        'would_call_jdy': False,
        'error': '同步到 JDY 功能暂未开放',
    }), 400


@app.route('/reorder-generate', methods=['POST'])
def reorder_generate():
    try:
        data = request.get_json(force=True) or {}
        ids = [int(x) for x in (data.get('item_ids') or []) if str(x).isdigit()]
        mode = str(data.get('mode') or 'supplier').strip().lower()
        if mode not in ('supplier', 'picker'):
            mode = 'supplier'
        if not ids:
            return jsonify({'success': False, 'error': '请先勾选要生成返单的商品'}), 400
        marks = ','.join('?' for _ in ids)
        with _sales_cache_conn() as conn:
            rows = conn.execute(f'SELECT * FROM reorder_items WHERE id IN ({marks})', ids).fetchall()
            items = [_reorder_row_to_item(row) for row in rows]
            missing_qty = [
                f"{item.get('code') or item.get('name') or item.get('id')}"
                for item in items
                if (_num(item.get('confirmedQty')) or _num(item.get('suggestedQty'))) <= 0
            ]
            missing_price = [
                f"{item.get('code') or item.get('name') or item.get('id')}"
                for item in items
                if _num(item.get('reorderPrice')) <= 0
            ]
            if missing_qty or missing_price:
                return jsonify({
                    'success': False,
                    'error': '生成返单前请补齐数量和价格',
                    'missing_qty': missing_qty,
                    'missing_price': missing_price,
                    'called_jdy': False,
                }), 400
            batch_no = 'RO' + datetime.now().strftime('%Y%m%d%H%M%S')
            generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            html_text = _reorder_generate_html(batch_no, mode, items)
            filename = f'{batch_no}.html'
            out_path = os.path.join(_reorder_generated_dir(), filename)
            with open(out_path, 'w', encoding='utf-8') as fp:
                fp.write(html_text)
            conn.execute(
                f"UPDATE reorder_items SET generated_batch_no=?, generated_at=?, sync_status='not_synced', status='confirmed', updated_at=? WHERE id IN ({marks})",
                [batch_no, generated_at, generated_at, *ids],
            )
            conn.commit()
        return jsonify({
            'success': True,
            'batch_no': batch_no,
            'output_type': 'html',
            'url': f'/reorder-generate/download/{filename}',
            'print_url': f'/reorder-batches/{batch_no}/print',
            'detail_url': f'/jdy#reorder-batches?batch_no={batch_no}',
            'sync_status': 'not_synced',
            'called_jdy': False,
            'would_call_jdy': False,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] reorder_generate: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb, 'called_jdy': False}), 500


@app.route('/reorder-generate/download/<path:filename>', methods=['GET'])
def reorder_generate_download(filename):
    safe = os.path.basename(filename)
    return send_file(os.path.join(_reorder_generated_dir(), safe), mimetype='text/html')


@app.route('/purchase-inbound-verify', methods=['GET'])
def purchase_inbound_verify():
    try:
        number = request.args.get('number', '').strip()
        code = request.args.get('code', '').strip()
        account = request.args.get('account', 'all').strip() or 'all'
        if not number:
            return jsonify({'success': False, 'error': '缺少购货单号'}), 400
        results = []
        errors = []
        for cli_fn, name in _sales_sources_for_account(account):
            cli = cli_fn()
            if not cli:
                errors.append(f'{name}: 未配置 JDY')
                continue
            try:
                result = cli.get_purchase_orders(page=1, page_size=20, search=number)
                rows = result.get('list') or []
                for row in rows:
                    row_no = str(_first_value(row, ['number', 'billNo', 'id'], '')).strip()
                    if row_no and row_no != number:
                        continue
                    entries = row.get('entries') or []
                    matched = [e for e in entries if _reorder_purchase_entry_matches(e, code)] if code else entries
                    attachments = _extract_purchase_attachments(row)
                    with _sales_cache_conn() as conn:
                        _cache_upsert_purchase_inbound(conn, name, row)
                        conn.commit()
                    results.append({
                        'account': name,
                        'number': row_no or number,
                        'date': str(_first_value(row, ['date', 'billDate', 'createTime'], ''))[:10],
                        'supplierNumber': _first_value(row, ['supplierNumber', 'supplierNo', 'vendorNumber'], ''),
                        'supplierName': _first_value(row, ['supplierName', 'vendorName'], ''),
                        'containsProduct': bool(matched),
                        'matchedEntries': matched,
                        'attachments': attachments,
                        'attachmentCount': len(attachments),
                        'rawKeys': sorted(str(k) for k in row.keys()),
                    })
            except Exception as e:
                errors.append(f'{name}: {_short_sync_error(e)}')
        return jsonify({
            'success': True,
            'number': number,
            'code': code,
            'list': results,
            'errors': errors,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(f'[ERROR] purchase_inbound_verify: {tb}')
        return jsonify({'success': False, 'error': str(e), 'traceback': tb}), 500


@app.route('/accessory-orders', methods=['GET'])
def accessory_orders_list():
    try:
        account = request.args.get('account', 'all')
        search = request.args.get('search', '').strip().lower()
        items = _read_cached_accessory_purchase_orders(account, search)
        return jsonify({
            'success': True,
            'list': items,
            'total': len(items),
            'summary': {
                'qty': sum(float(x.get('totalQty') or 0) for x in items),
                'amount': sum(float(x.get('totalAmount') or 0) for x in items),
            },
            'cache': True,
            'cache_stats': _accessory_purchase_cache_stats(),
            'state': _accessory_sync_state,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/accessory-order/<order_no>', methods=['GET'])
def accessory_order_detail(order_no):
    try:
        account = request.args.get('account', '').strip()
        allow_live = str(request.args.get('allow_live') or '').lower() in ('1', 'true', 'yes', 'y')
        refresh = str(request.args.get('refresh') or '').lower() in ('1', 'true', 'yes', 'y')
        order = _read_cached_accessory_purchase_order_detail(order_no, account)
        if not order:
            return jsonify({
                'success': False,
                'error': '未找到本地辅料订单',
                'local_only': True,
                'live_lookup': False,
                'called_jdy': False,
                'would_call_jdy': False,
                'cache_source': 'accessory_purchase_orders',
                'enrich_skipped': True,
            }), 404
        enrich_skipped = True
        live_lookup = False
        called_jdy = False
        if allow_live and refresh:
            order = _enrich_accessory_purchase_order(order)
            with _sales_cache_conn() as conn:
                _cache_upsert_accessory_purchase_order(conn, order)
                conn.commit()
            enrich_skipped = False
            live_lookup = True
            called_jdy = True
        return jsonify({
            'success': True,
            'data': order,
            'cache': True,
            'local_only': not live_lookup,
            'live_lookup': live_lookup,
            'called_jdy': called_jdy,
            'would_call_jdy': bool(allow_live and refresh),
            'cache_source': 'accessory_purchase_orders',
            'enrich_skipped': enrich_skipped,
            'message': (
                '默认仅读取本地辅料订单缓存，未联网补齐商品信息。'
                if enrich_skipped else '已按显式参数刷新辅料订单商品信息。'
            ),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/accessory-suppliers', methods=['GET'])
def accessory_suppliers():
    try:
        account = request.args.get('account', 'all').strip()
        items = _read_accessory_supplier_options(account)
        return jsonify({'success': True, 'list': items, 'total': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/accessory-products', methods=['GET'])
def accessory_products():
    try:
        category = request.args.get('category', 'all').strip() or 'all'
        search = request.args.get('search', '').strip()
        supplier = request.args.get('supplier', '').strip()
        items, summary, categories = _read_cached_accessory_products(search, category, supplier)
        stats = _accessory_product_cache_stats()
        cache_empty = not stats.get('products_count')
        return jsonify({
            'success': True,
            'list': items,
            'total': len(items),
            'summary': summary,
            'categories': ACCESSORY_MATERIAL_CATEGORIES,
            'matched_categories': categories,
            'cache_stats': stats,
            'sync_state': _accessory_product_sync_state,
            'local_only': True,
            'live_lookup': False,
            'called_jdy': False,
            'would_call_jdy': False,
            'cache_source': 'accessory_products',
            'auto_sync_started': False,
            'message': '本地辅料商品缓存为空，请在设置页手动同步。' if cache_empty else '读取本地辅料商品缓存。',
            'errors': [],
            'error': '',
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/accessory-products-sync/status', methods=['GET'])
def accessory_products_sync_status():
    return jsonify({
        'success': True,
        'state': _accessory_product_sync_state,
        'stats': _accessory_product_cache_stats(),
    })


@app.route('/accessory-products-sync-run', methods=['POST'])
def accessory_products_sync_run():
    try:
        wait = str(request.args.get('wait') or '').lower() in ('1', 'true', 'yes', 'y')
        if wait:
            result = _run_accessory_product_sync(reason='manual')
            return jsonify({'success': True, 'started': True, 'result': result, 'state': _accessory_product_sync_state})
        if _accessory_product_sync_lock.locked() or _accessory_product_sync_state.get('running'):
            return jsonify({'success': True, 'started': False, 'running': True, 'state': _accessory_product_sync_state})
        t = threading.Thread(
            target=_run_accessory_product_sync,
            kwargs={'reason': 'manual'},
            daemon=True,
            name='accessory-product-sync-manual',
        )
        t.start()
        return jsonify({'success': True, 'started': True, 'running': True, 'state': _accessory_product_sync_state})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'state': _accessory_product_sync_state}), 500


@app.route('/accessory-orders-sync/status', methods=['GET'])
def accessory_orders_sync_status():
    return jsonify({
        'success': True,
        'config': _sales_sync_config(),
        'state': _accessory_sync_state,
        'stats': _accessory_purchase_cache_stats(),
        'supplier_stats': _supplier_cache_stats(),
    })


@app.route('/accessory-orders-sync-run', methods=['POST'])
def accessory_orders_sync_run():
    try:
        wait = str(request.args.get('wait') or '').lower() in ('1', 'true', 'yes', 'y')
        account = request.args.get('account') or 'all'
        if wait:
            result = _run_accessory_incremental_sync(reason='manual', account=account)
            return jsonify({'success': True, 'started': True, 'result': result, 'state': _accessory_sync_state})
        if _accessory_sync_lock.locked() or _accessory_sync_state.get('running'):
            return jsonify({'success': True, 'started': False, 'running': True, 'state': _accessory_sync_state})
        t = threading.Thread(
            target=_run_accessory_incremental_sync,
            kwargs={'reason': 'manual', 'account': account},
            daemon=True,
            name='accessory-sync-manual',
        )
        t.start()
        return jsonify({'success': True, 'started': True, 'running': True, 'state': _accessory_sync_state})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'state': _accessory_sync_state}), 500


@app.route('/jdy')
def jdy_ui():
    """加载 PyWebView 桌面 UI（jdy.html）"""
    import os
    html_path = os.path.join(_BASE, 'templates', 'jdy.html')
    if not os.path.exists(html_path):
        # 生产包中查 exe 同目录
        exe_dir = _EXE_DIR if getattr(sys, 'frozen', False) else _BASE
        html_path = os.path.join(exe_dir, 'templates', 'jdy.html')
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return f.read(), 200, {'Content-Type': 'text/html; charset=utf-8'}
    except FileNotFoundError:
        return '<h2>jdy.html 未找到</h2>', 404


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    # 启动前初始化两个 JDY 账套客户端（如已配置）
    cfg = _load_jdy_config()
    if cfg['client_id'] and cfg['app_key'] and cfg['app_secret'] and cfg['db_id']:
        jdy_api.init_client(
            cfg['client_id'], cfg['app_key'], cfg['app_secret'],
            cfg['db_id'], cfg.get('domain', ''), cfg.get('app_signature', ''),
            cfg.get('client_secret', '')
        )
        print(f'[JDY] 账套1({cfg.get("name","饰品")}) 已初始化，dbId={cfg["db_id"]}')

    cfg2 = _load_jdy_config2()
    if cfg2['client_id'] and cfg2['app_key'] and cfg2['app_secret'] and cfg2['db_id']:
        jdy_api.init_client_2(
            cfg2['client_id'], cfg2['app_key'], cfg2['app_secret'],
            cfg2['db_id'], cfg2.get('domain', ''), cfg2.get('app_signature', ''),
            cfg2.get('client_secret', '')
        )
        print(f'[JDY] 账套2({cfg2.get("name","箱包")}) 已初始化，dbId={cfg2["db_id"]}')

    print(f'✅ 报关服务已启动：http://localhost:{PORT}')
    print(f'   按 Ctrl+C 停止服务')
    if _workers_disabled():
        print('[WORKERS] workers disabled by QIHANG_DISABLE_WORKERS')
    else:
        _start_sales_sync_worker()
    try:
        from waitress import serve
        serve(app, host='0.0.0.0', port=PORT, threads=16,
              channel_timeout=300, cleanup_interval=30)
    except ImportError:
        # waitress 未安装时降级为 Flask 开发服务器（仅单线程，不推荐多用户）
        print('[WARN] waitress 未安装，使用 Flask 开发服务器（单线程）')
        print('       建议运行: pip install waitress')
        app.run(host='0.0.0.0', port=PORT, debug=False, threaded=True)
