"""
新版序号分析表 v2（完整版）
- Sheet1 主表    ：有默认供应商 + 分类匹配，同前缀只保留最大序号那行
                   末尾附所有启用供应商中无产品的行
- Sheet2 无供应商：defaultSupplierId=0 或找不到，同前缀保留最大序号
- Sheet3 分类对不上：有供应商但 JDY categoryName 不在映射表，同前缀保留最大序号

列（主表）：
  账套 | 供应商编号 | 供应商名称 | 档口号 | 商品编号(最大序号) | 商品名称
  | JDY原始分类 | 大类代码 | 大类名称 | 中类代码 | 中类名称
  | 小类代码 | 小类名称 | 前缀商品数 | 当前最大序号
  | 前缀和新编号的前缀对比 | 模拟编号规则的新号码 | 选择（我来选）

运行：python3 build_seq_analysis_v2.py
输出：../../序号分析表_v2.xlsx
"""
import json, sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from jdy_cache import JDYCache

# ── 路径 ────────────────────────────────────────────────────────────────────
_BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CFG_PATH = os.path.join(_BASE, '..', 'ai_config.json')
CAT_FILE = os.path.join(_BASE, '..', '钱姝包分类_整理版.xlsx')
OUT_FILE = os.path.join(_BASE, '..', '序号分析表_v2.xlsx')

# ── 箱包账套：JDY categoryName → (大类代码, 大类名称, 中类代码, 中类名称) ─────
BAG_CAT_MAP = {
    'Lady Bag':            ('L', "Women's Bags 女包", 'F', 'Fashion Bags 时尚包'),
    'Lady Bag(CR)':        ('L', "Women's Bags 女包", 'F', 'Fashion Bags 时尚包'),
    'Lady Bag(JMG)':       ('L', "Women's Bags 女包", 'F', 'Fashion Bags 时尚包'),
    'Lady Bag(QZHG)':      ('L', "Women's Bags 女包", 'F', 'Fashion Bags 时尚包'),
    'Lady Bag(Tote bag)':  ('L', "Women's Bags 女包", 'F', 'Fashion Bags 时尚包'),
    'Lady Backpack':       ('L', "Women's Bags 女包", 'F', 'Fashion Bags 时尚包'),
    'Lady Wallet':         ('L', "Women's Bags 女包", 'W', 'Wallets 钱包'),
    'Party Bag':           ('L', "Women's Bags 女包", 'P', 'Party Bags 派对包'),
    "Children's bag":      ('C', "Children's 儿童",   '',  'Small Handbags 手提小包'),
    'Man Bag':             ('M', "Men's Bags 男包",    'M', 'Crossbody Bags 斜挎包'),
    'Man Wallet':          ('M', "Men's Bags 男包",    'W', 'Wallets 钱包'),
    'Passport bag':        ('M', "Men's Bags 男包",    'W', 'Wallets 钱包'),
    'Laptop bag':          ('M', "Men's Bags 男包",    'C', 'Laptop Bags 电脑包'),
    'Lunch Bag':           ('O', 'Other Bags 其他包',  'L', 'Lunch Bags 午餐包'),
    'Wool Bag':            ('O', 'Other Bags 其他包',  'H', 'Fur Bags 毛包'),
    'Wash Bag':            ('O', 'Other Bags 其他包',  'W', 'Toiletry Bags 洗漱包'),
    'Handbag、Canvas Bag': ('O', 'Other Bags 其他包',  'P', 'Canvas Bags 帆布包'),
    'Woven bag':           ('O', 'Other Bags 其他包',  'S', 'Straw Bags 草编织包'),
    'Wood bag':            ('O', 'Other Bags 其他包',  'C', 'Coconut Shell Bags 椰壳包'),
}

# ── 从分类编码表读取 饰品/箱包 小类数据 ──────────────────────────────────────
def _build_cat_maps(cat_file):
    """
    返回两个字典：
      acc_cat_map: 饰品  英文小类名 → (l1c, l1n, l2c, l2n, l3c, l3n)
      bag_l3_map:  箱包  (l1c, l2c) → [(l3c, l3n)]（用于推断小类）
    """
    wb = openpyxl.load_workbook(cat_file)
    ws = wb.active
    acc_cat_map = {}
    bag_l3_map  = {}

    cur_l1c, cur_l1n = None, None
    cur_l2c, cur_l2n = None, None
    cur_acc          = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        acc, l1c, l1n, l2c, l2n, l3c, l3n = row
        if acc:
            cur_acc = str(acc)
        if l1c:
            cur_l1c, cur_l1n = l1c, l1n
        if l2c:
            cur_l2c, cur_l2n = l2c, l2n
        if not l3c:
            continue
        l3c_str = str(l3c).zfill(2)
        l3n_str = str(l3n) if l3n else ''

        if cur_acc and '饰品' in cur_acc:
            # English part before first CJK char / bracket
            en = re.sub(r'[\u4e00-\u9fff（【\(（].*', '', l3n_str).strip()
            if en:
                acc_cat_map[en] = (cur_l1c, cur_l1n, cur_l2c, cur_l2n, l3c_str, l3n_str)
        elif cur_acc and '箱包' in cur_acc:
            key = (cur_l1c, cur_l2c)
            bag_l3_map.setdefault(key, []).append((l3c_str, l3n_str))

    return acc_cat_map, bag_l3_map

ACC_CAT_MAP, BAG_L3_MAP = _build_cat_maps(CAT_FILE)

# ── 档口号转换规则 ────────────────────────────────────────────────────────────
_D2L = {'1':'A','2':'B','3':'C','4':'D','5':'E','6':'F','7':'G','8':'H','9':'I','0':'J'}

def transform_stall(number):
    """
    将档口号（supplier.number）转换为新编号的前缀部分。
    字母开头 → 数字开头；数字开头 → 字母开头。
    第2位+1，若为9则变0（不进位不加位）。
    特殊：空/0/0001等非真实档口号原样返回空字符串。
    """
    if not number:
        return ''
    s = str(number).strip().upper()
    if not s or s in ('0',):
        return ''
    if len(s) < 2:
        return s  # 太短，无法变换，原样

    first  = s[0]
    second = s[1]
    rest   = s[2:]

    def _inc(ch):
        return str((int(ch) + 1) % 10) if ch.isdigit() else ch

    if first.isalpha():
        # Letter → position number (A=1 … Z=26)
        num1 = str(ord(first) - ord('A') + 1)
        return num1 + _inc(second) + rest
    else:
        # Digit → letter (1=A … 9=I, 0=J)
        letter1 = _D2L.get(first, first)
        return letter1 + _inc(second) + rest

# ── 前缀提取 ─────────────────────────────────────────────────────────────────
def _extract_prefix(pno):
    """
    从商品编号提取旧前缀（供比较用）。
    新饰品格式  X.XXXXXL-n  → 返回 XXXXX（去掉末尾字母）
    旧格式      XXXXXL-n    → 返回 XXXXX（去掉末尾字母）
    其他返回 None。
    """
    pno = str(pno or '').strip().upper()
    # 新饰品格式：单字母 . (内容以字母结尾) - 数字
    m = re.match(r'^[A-Z]\.(.*[A-Z])-\d+$', pno)
    if m:
        core = m.group(1)           # e.g. "B8029T"
        return re.sub(r'[A-Z]+$', '', core)  # strip trailing letters → "B8029"
    # 旧格式：内容以字母结尾 - 数字
    m = re.match(r'^(.*[A-Z])-\d+$', pno)
    if m:
        core = m.group(1)
        return re.sub(r'[A-Z]+$', '', core)
    return None

def _parse_seq(pno):
    m = re.search(r'-(\d+)$', str(pno or ''))
    return int(m.group(1)) if m else 0

# ── 前缀比较 ─────────────────────────────────────────────────────────────────
def _compare_prefix(old_prefix, new_prefix):
    """
    返回 ('same'|'similar'|'different'|'no_stall', 说明字符串)
    same:     完全相同 → 旧最大序号 + 1
    similar:  共同后缀 ≥ 4 位 → 标黄，手动确认
    different: 其他 → 从 1 开始
    no_stall: taxPayerNo 为空，无法比较
    """
    if not new_prefix:
        return 'no_stall', f'旧前缀: {old_prefix} | 档口号未设置'
    if not old_prefix:
        return 'different', f'（无法提取旧前缀） → {new_prefix}'
    if old_prefix == new_prefix:
        return 'same', f'✓ 相同 {old_prefix}'
    # 计算公共后缀长度
    common = 0
    for a, b in zip(reversed(old_prefix), reversed(new_prefix)):
        if a == b:
            common += 1
        else:
            break
    if common >= 4:
        return 'similar', f'⚠ 相似({common}位) {old_prefix} / {new_prefix}'
    return 'different', f'✗ 不同 {old_prefix} → {new_prefix}'

# ── 模拟新编号 ────────────────────────────────────────────────────────────────
def _sim_new_no(acc, l1c, l2c, l3c, new_prefix, next_seq):
    """
    构造模拟新编号字符串。
    箱包：{大类}{中类}.{转换档口号}{小类}-{序号}
    饰品：{中类}.{转换档口号}{小类}-{序号}
    l3c 为 None 时返回空（箱包多小类无法确定时）。
    """
    if not new_prefix or not l3c:
        return ''
    l3c2 = str(l3c).zfill(2)
    if acc == 'account2':   # 箱包
        if not l1c or not l2c:
            return ''
        return f'{l1c}{l2c}.{new_prefix}{l3c2}-{next_seq}'
    else:                   # 饰品
        if not l2c:
            return ''
        return f'{l2c}.{new_prefix}{l3c2}-{next_seq}'

# ── 主逻辑 ────────────────────────────────────────────────────────────────────
cache = JDYCache(cfg_path=CFG_PATH)

main_rows    = []
no_sup_rows  = []
no_cat_rows  = []

for acc, label, cat_map in [
    ('account1', '祺航饰品', ACC_CAT_MAP),
    ('account2', '祺航箱包', BAG_CAT_MAP),
]:
    prods = cache.get_products(acc)
    sups  = cache.get_suppliers(acc)
    sup_by_id = {str(s['id']): s for s in sups}
    if not prods:
        print(f'[WARN] {label} 商品缓存为空，跳过')
        continue

    # 按前缀聚合
    prefix_map = {}
    for p in prods:
        pno = str(p.get('productNumber') or '')
        if not pno:
            continue
        prefix = _extract_prefix(pno)
        if prefix is None:
            continue
        seq    = _parse_seq(pno)
        key    = prefix
        if key not in prefix_map:
            prefix_map[key] = {
                'acc': acc, 'label': label,
                'max_seq': -1, 'max_pno': '', 'max_name': '',
                'cat': '', 'sup_id': '0', 'count': 0,
            }
        g = prefix_map[key]
        g['count'] += 1
        if seq > g['max_seq']:
            g['max_seq'] = seq
            g['max_pno'] = pno
            g['max_name'] = p.get('productName') or ''
            g['cat']      = p.get('categoryName') or ''
            g['sup_id']   = str(p.get('defaultSupplierId') or 0)

    # 记录有产品的供应商ID集合（用于后面列出无产品供应商）
    sup_ids_with_product = set()

    for prefix, g in prefix_map.items():
        sid = g['sup_id']
        sup = sup_by_id.get(sid)
        cat = g['cat']

        # 无供应商
        if sid == '0' or not sup:
            no_sup_rows.append({
                'label': g['label'], 'pno': g['max_pno'],
                'name': g['max_name'], 'cat': cat,
                'prefix': prefix, 'max_seq': g['max_seq'], 'count': g['count'],
            })
            continue

        sup_ids_with_product.add(sid)
        sup_name    = sup.get('name', '')
        sup_jdy_no  = sup.get('number', '')        # B列：JDY 供应商编号（内部码）
        sup_stall   = sup.get('taxPayerNo', '')    # D列：档口号 = 纳税人识别号
        new_prefix  = transform_stall(sup_stall)
        cmp_type, cmp_str = _compare_prefix(prefix, new_prefix)

        # 分类匹配
        match = cat_map.get(cat)
        if not match:
            no_cat_rows.append({
                'label': g['label'], 'pno': g['max_pno'],
                'name': g['max_name'], 'cat': cat,
                'sup_name': sup_name, 'sup_jdy_no': sup_jdy_no, 'sup_stall': sup_stall,
                'prefix': prefix, 'max_seq': g['max_seq'], 'count': g['count'],
                'new_prefix': new_prefix, 'cmp_type': cmp_type, 'cmp_str': cmp_str,
            })
            continue

        l1c, l1n, l2c, l2n = match[0], match[1], match[2], match[3]

        # 小类推断
        if acc == 'account1':
            # 饰品：ACC_CAT_MAP 已包含所有3级 (6-tuple)
            l3c = match[4] if len(match) > 4 else ''
            l3n = match[5] if len(match) > 5 else ''
        else:
            # 箱包：查 BAG_L3_MAP，只有1个小类时自动填
            l3_options = BAG_L3_MAP.get((l1c, l2c), [])
            if len(l3_options) == 1:
                l3c, l3n = l3_options[0]
            else:
                l3c, l3n = '', ''   # 多选项，留空

        # 序号推算
        if cmp_type == 'same':
            next_seq = g['max_seq'] + 1
        elif cmp_type == 'similar':
            next_seq = g['max_seq'] + 1   # 标黄，需人工确认
        elif cmp_type == 'no_stall':
            next_seq = 1                  # 档口号未设置，模拟编号留空
        else:
            next_seq = 1

        sim_no = _sim_new_no(acc, l1c, l2c, l3c, new_prefix, next_seq)

        main_rows.append({
            'label': label,
            'sup_jdy_no': sup_jdy_no,   # B列：JDY编号
            'sup_stall':  sup_stall,     # D列：档口号
            'sup_name':   sup_name,
            'pno':   g['max_pno'],
            'name':  g['max_name'],
            'cat':   cat,
            'l1c': l1c, 'l1n': l1n,
            'l2c': l2c, 'l2n': l2n,
            'l3c': l3c, 'l3n': l3n,
            'count':    g['count'],
            'max_seq':  g['max_seq'],
            'cmp_type': cmp_type,
            'cmp_str':  cmp_str,
            'sim_no':   sim_no,
        })

    # ── 无产品的启用供应商（追加在末尾）────────────────────────────────────────
    for sup in sups:
        sid = str(sup.get('id', ''))
        if sup.get('isDeleted'):
            continue
        if sid in sup_ids_with_product:
            continue
        main_rows.append({
            'label': label,
            'sup_jdy_no': sup.get('number', ''),
            'sup_stall':  sup.get('taxPayerNo', ''),
            'sup_name':   sup.get('name', ''),
            'pno': '', 'name': '', 'cat': '',
            'l1c': '', 'l1n': '', 'l2c': '', 'l2n': '',
            'l3c': '', 'l3n': '',
            'count': 0, 'max_seq': 0,
            'cmp_type': '', 'cmp_str': '', 'sim_no': '',
            '_no_product': True,
        })

    print(f'{label}: 前缀组 {len(prefix_map)} 个 → '
          f'主表 {sum(1 for r in main_rows if r["label"]==label and not r.get("_no_product"))} '
          f'| 无供应商 {sum(1 for r in no_sup_rows if r["label"]==label)} '
          f'| 分类对不上 {sum(1 for r in no_cat_rows if r["label"]==label)}')

# ── 排序 ──────────────────────────────────────────────────────────────────────
def _main_sort_key(r):
    # 无产品的放最后，其余按账套→大类→中类→供应商
    return (r['label'], 1 if r.get('_no_product') else 0,
            r['l1c'], r['l2c'], r['sup_name'])

main_rows.sort(key=_main_sort_key)
no_sup_rows.sort(key=lambda x: (x['label'], x['cat']))
no_cat_rows.sort(key=lambda x: (x['label'], x['cat'], x['sup_name']))

# ── Excel 样式 ────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
FILL1     = PatternFill('solid', fgColor='D6E4F0')
FILL2     = PatternFill('solid', fgColor='FFFFFF')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')   # 黄：相似前缀
NO_PROD   = PatternFill('solid', fgColor='F2F2F2')   # 灰：无产品供应商
DATA_FONT = Font(name='Calibri', size=10)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=False)
thin  = Side(style='thin',   color='B8CCE4')
outer = Side(style='medium', color='1F4E79')
INNER_BDR = Border(left=thin, right=thin, top=thin, bottom=thin)
OUTER_BDR = Border(left=outer, right=outer, top=outer, bottom=outer)


def _apply_cell(cell, fill, center=False):
    cell.fill      = fill
    cell.font      = DATA_FONT
    cell.border    = INNER_BDR
    cell.alignment = CENTER if center else LEFT


# ── 写主表（含供应商合并 B/C 列）────────────────────────────────────────────
MAIN_HEADERS = [
    '账套', '供应商编号', '供应商名称', '档口号',
    '商品编号(最大序号)', '商品名称', 'JDY原始分类',
    '大类代码', '大类名称', '中类代码', '中类名称',
    '小类代码', '小类名称',
    '前缀商品数', '当前最大序号',
    '前缀和新编号的前缀对比',
    '模拟编号规则的新号码',
    '选择（我来选）',
]
MAIN_WIDTHS = [10, 14, 20, 14, 22, 22, 20,
               8, 26, 8, 26,
               8, 26,
               10, 12,
               36, 24, 14]
MAIN_CENTER = {1, 2, 4, 8, 10, 12, 14, 15}

def _write_main(ws, rows):
    ws.append(MAIN_HEADERS)
    for c in range(1, len(MAIN_HEADERS)+1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = CENTER
        cell.border    = OUTER_BDR

    row_idx = 2
    # track merges for B/C columns (same supplier consecutive rows)
    merge_start   = None   # row where current supplier block starts
    cur_sup_key   = None   # (label, sup_jdy_no, sup_name)

    def _flush_merge(end_row):
        if merge_start and merge_start < end_row:
            ws.merge_cells(start_row=merge_start, start_column=2,
                           end_row=end_row-1,   end_column=2)
            ws.merge_cells(start_row=merge_start, start_column=3,
                           end_row=end_row-1,   end_column=3)
            for col in (2, 3):
                c = ws.cell(merge_start, col)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    for r in rows:
        sup_key = (r['label'], r['sup_jdy_no'], r['sup_name'])
        if sup_key != cur_sup_key:
            _flush_merge(row_idx)
            merge_start = row_idx
            cur_sup_key = sup_key

        is_no_prod  = r.get('_no_product', False)
        cmp_type    = r.get('cmp_type', '')

        if is_no_prod:
            fill = NO_PROD
        elif cmp_type == 'similar':
            fill = WARN_FILL
        elif row_idx % 2 == 0:
            fill = FILL1
        else:
            fill = FILL2

        data = [
            r['label'], r['sup_jdy_no'], r['sup_name'], r['sup_stall'],   # A-D
            r['pno'],   r['name'],  r['cat'],                              # E-G
            r['l1c'],   r['l1n'],   r['l2c'],  r['l2n'],                   # H-K
            r['l3c'],   r['l3n'],                                           # L-M
            r['count'] or '', r['max_seq'] if not is_no_prod else '',       # N-O
            r['cmp_str'], r['sim_no'], '',                                  # P-R
        ]
        ws.append(data)
        for col, val in enumerate(data, 1):
            _apply_cell(ws.cell(row_idx, col), fill, col in MAIN_CENTER)

        row_idx += 1

    # flush last supplier merge
    _flush_merge(row_idx)

    for i, w in enumerate(MAIN_WIDTHS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'


# ── 写其他 sheet ─────────────────────────────────────────────────────────────
def _write_sheet(ws, headers, rows, col_widths, center_cols, row_fn, warn_fn=None):
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = CENTER
        cell.border    = OUTER_BDR
    for i, row_data in enumerate(rows):
        ws.append(row_fn(row_data))
        fill = WARN_FILL if (warn_fn and warn_fn(row_data)) else (FILL1 if i%2==0 else FILL2)
        for c in range(1, len(headers)+1):
            cell = ws.cell(i+2, c)
            cell.fill      = fill
            cell.font      = DATA_FONT
            cell.border    = INNER_BDR
            cell.alignment = CENTER if c in center_cols else LEFT
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'


# ── 生成 Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Sheet1 主表
ws1 = wb.active
ws1.title = '主表'
_write_main(ws1, main_rows)

# Sheet2 无供应商
ws2 = wb.create_sheet('无供应商')
_write_sheet(
    ws2,
    headers=['账套', '商品编号(最大序号)', '商品名称', 'JDY原始分类', '前缀', '前缀商品数', '当前最大序号'],
    rows=no_sup_rows,
    col_widths=[10, 22, 22, 22, 18, 10, 12],
    center_cols={1, 6, 7},
    row_fn=lambda r: [r['label'], r['pno'], r['name'], r['cat'], r['prefix'], r['count'], r['max_seq']],
)

# Sheet3 分类对不上
ws3 = wb.create_sheet('分类对不上')
_write_sheet(
    ws3,
    headers=['账套', '商品编号(最大序号)', '商品名称', 'JDY原始分类',
             '供应商名称', '供应商编号', '前缀商品数', '当前最大序号',
             '前缀和新编号的前缀对比'],
    rows=no_cat_rows,
    col_widths=[10, 22, 22, 22, 20, 12, 10, 12, 36],
    center_cols={1, 7, 8},
    row_fn=lambda r: [
        r['label'], r['pno'], r['name'], r['cat'],
        r['sup_name'], r['sup_jdy_no'], r['count'], r['max_seq'],
        r.get('cmp_str', ''),
    ],
)

wb.save(OUT_FILE)
no_prod_cnt = sum(1 for r in main_rows if r.get('_no_product'))
has_prod_cnt = len(main_rows) - no_prod_cnt
print(f'\n✅ 已生成: {OUT_FILE}')
print(f'   主表: {has_prod_cnt} 行（含产品）+ {no_prod_cnt} 行（无产品供应商）'
      f' | 无供应商: {len(no_sup_rows)} 行 | 分类对不上: {len(no_cat_rows)} 行')
