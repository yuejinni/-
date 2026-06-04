"""
A2: 生成序号分析表 Excel
需要先跑完商品缓存（jdy_cache.py）才有完整数据。
运行: python3 build_seq_analysis.py
输出: ../../序号分析表.xlsx
"""
import json, sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from jdy_cache import JDYCache
from code_gen import transform_vendor_code

cache = JDYCache(cfg_path='../../ai_config.json')

def _parse_seq(pno):
    """从商品编码末尾提取序号"""
    m = re.search(r'-(\d+)$', str(pno or ''))
    return int(m.group(1)) if m else 0

def _extract_prefix(pno):
    """
    提取商品编码中的'档口号+分类字母'前缀部分
    格式一: C.A12629C-0008 → A12629C（新格式）
    格式二: 102584C-7      → 102584C（旧格式）
    """
    # 新格式：{cat}.{code}{cat}-{seq}
    m = re.match(r'^[A-Z]\.(.*[A-Z])-\d+$', str(pno or '').upper())
    if m:
        return m.group(1), 'new'
    # 旧格式：{code}{cat}-{seq}
    m = re.match(r'^(.*[A-Z])-\d+$', str(pno or '').upper())
    if m:
        return m.group(1), 'old'
    return None, None

# ── 构建 supplier_id → {taxPayerNo, name, transformed} ──────────────────────
all_rows = []   # 最终分析结果

for acc, label in [('account1', '祺航饰品'), ('account2', '祺航箱包')]:
    print(f'\n处理 {label}...')

    sups = cache.get_suppliers(acc)
    sup_by_id = {str(s['id']): s for s in sups}
    sup_by_id.update({str(s.get('number','')): s for s in sups if s.get('number')})

    print(f'  供应商: {len(sups)} 个')

    prods = cache.get_products(acc)
    if not prods:
        print(f'  ⚠️  {acc} 商品缓存为空，跳过（请先运行 python3 jdy_cache.py）')
        continue

    print(f'  商品: {len(prods)} 个')

    # 聚合：transformed_code + cat_letter → {max_seq, sup_info, examples, count}
    group_map = {}   # (transformed, cat_letter) → {}

    for p in prods:
        pno = str(p.get('productNumber') or '')
        if not pno:
            continue

        # 获取默认供应商
        sup_id = str(p.get('defaultSupplierId') or '')
        sup = sup_by_id.get(sup_id)
        if not sup:
            # 尝试 category name 的供应商（backup）
            continue

        tax_no = (sup.get('taxPayerNo') or '').strip()
        if not tax_no:
            continue

        transformed, err = transform_vendor_code(tax_no)
        if err:
            transformed = tax_no  # 无法变换，用原始值

        # 提取前缀和序号
        prefix, fmt = _extract_prefix(pno)
        if not prefix:
            continue

        seq = _parse_seq(pno)
        # 从前缀末尾取分类字母
        cat_letter = prefix[-1] if prefix and prefix[-1].isalpha() else '?'
        vendor_part = prefix[:-1] if cat_letter != '?' else prefix

        key = (transformed, cat_letter)
        if key not in group_map:
            group_map[key] = {
                'acc': acc,
                'label': label,
                'sup_name': sup.get('name', ''),
                'sup_number': sup.get('number', ''),
                'tax_no': tax_no,
                'transformed': transformed,
                'transform_err': err or '',
                'cat_letter': cat_letter,
                'max_seq': 0,
                'old_prefix_example': '',
                'count': 0,
                'examples': [],
            }
        g = group_map[key]
        g['count'] += 1
        if seq > g['max_seq']:
            g['max_seq'] = seq
        if len(g['examples']) < 3:
            g['examples'].append(pno)
        if fmt == 'old' and not g['old_prefix_example']:
            g['old_prefix_example'] = vendor_part

    for g in group_map.values():
        g['examples_str'] = ' / '.join(g['examples'])
        all_rows.append(g)

    print(f'  找到 {len(group_map)} 个(变换码, 分类字母)组合')

if not all_rows:
    print('\n⚠️  没有数据，请先确认商品缓存已完成')
    sys.exit(1)

# ── 生成 Excel ────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '序号分析表'

    headers = [
        '账套', '供应商名称', '供应商编号', 'taxPayerNo(档口号)',
        '变换后编码', '变换错误', '分类字母',
        '现有商品数', '现有最大序号', '示例商品编码',
        '建议初始值', '你的决定（旧/新）', '备注',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='366092')

    yellow = PatternFill('solid', fgColor='FFFF99')
    green  = PatternFill('solid', fgColor='C6EFCE')
    red_   = PatternFill('solid', fgColor='FF9999')

    all_rows.sort(key=lambda x: (x['acc'], x['sup_name'], x['cat_letter']))

    for g in all_rows:
        suggest = f'从 {g["max_seq"]+1:04d} 开始（接旧序号）' if g['max_seq'] > 0 else '从 0001 开始（新建）'
        row = [
            g['label'], g['sup_name'], g['sup_number'], g['tax_no'],
            g['transformed'], g['transform_err'], g['cat_letter'],
            g['count'], g['max_seq'],
            g['examples_str'],
            g['max_seq'] if g['max_seq'] > 0 else 0,
            '旧序号+1' if g['max_seq'] > 0 else '新建0001',
            '',
        ]
        ws.append(row)
        lr = ws.max_row
        if g['transform_err']:
            ws.cell(lr, 5).fill = red_
            ws.cell(lr, 12).fill = red_
        elif g['max_seq'] > 0:
            ws.cell(lr, 11).fill = yellow
            ws.cell(lr, 12).fill = yellow
        else:
            ws.cell(lr, 12).fill = green

    col_widths = [10,22,12,16,14,20,10,10,12,30,14,14,16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', '序号分析表.xlsx')
    wb.save(out)
    print(f'\n✅ 已生成: {out}')
    print(f'总分析组合: {len(all_rows)} 个')
    has_old = sum(1 for g in all_rows if g['max_seq'] > 0)
    print(f'  有旧序号（接续）: {has_old}  新建0001: {len(all_rows)-has_old}')
except ImportError:
    print('需要 openpyxl: pip install openpyxl')
