"""
箱包账套产品分类脚本
- 按产品名称 + JDY分类 自动推断新分类路径
- 输出 JDY 商品导入格式 Excel（商品类别 + 商品编号）
- 分高/中/低置信度，低置信度标红提醒人工核查
运行：python3 classify_bag_products.py
输出：../../箱包分类导入.xlsx
"""
import json, re, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

_BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAT_FILE = os.path.join(_BASE, '..', '钱姝包分类_整理版.xlsx')
OUT_FILE = os.path.join(_BASE, '..', '箱包分类导入.xlsx')
CACHE    = os.path.join(_BASE, '_cache', 'product_account2.json')

# ── 1. 构建分类路径表 ─────────────────────────────────────────────────────────
wb_cat = openpyxl.load_workbook(CAT_FILE)
ws_cat = wb_cat.active
cat_paths = {}   # (l1c, l2c, l3c_str) → "大类名/中类名/小类名"
cur_acc = cur_l1c = cur_l1n = cur_l2c = cur_l2n = None

for row in ws_cat.iter_rows(min_row=2, values_only=True):
    acc, l1c, l1n, l2c, l2n, l3c, l3n = row
    if acc: cur_acc = str(acc)
    if l1c: cur_l1c, cur_l1n = l1c, l1n
    if l2c: cur_l2c, cur_l2n = l2c, l2n
    if not l3c: continue
    if '箱包' not in str(cur_acc): continue
    l3str = str(l3c).zfill(2)
    # 中类代码可能为 None（如儿童/Small Handbags），用空串代替
    l2key = cur_l2c if cur_l2c else ''
    key   = (cur_l1c, l2key, l3str)
    cat_paths[key] = f'{cur_l1n}/{cur_l2n}/{l3n}'

def p(l1, l2, l3):
    return cat_paths.get((l1, l2 or '', str(l3).zfill(2)), '')

# ── 2. 关键词分类函数 ─────────────────────────────────────────────────────────
def classify_lf(name):
    """L-F 时尚包 10个小类"""
    n = str(name or '')
    kw = [
        (['斜挎'], '04'), (['腋下'], '05'), (['水桶'], '06'),
        (['妈咪'], '07'), (['菜篮'], '08'), (['枕头'], '09'),
        (['贝壳'], '10'), (['托特', 'tote', 'Tote'], '01'),
    ]
    for keys, code in kw:
        if any(k in n for k in keys):
            return code, 'high'
    if '双' in n and ('手挽' in n or '手腕' in n):
        return '02', 'mid'
    if '手挽' in n or '手腕' in n:
        return '01', 'mid'
    return '01', 'low'   # 名称无关键词，默认定型单手腕，需核查

def classify_lw(name):
    """L-W 女士钱包"""
    n = str(name or '')
    if '单拉' in n: return '01', 'high'
    if '双拉' in n: return '02', 'high'
    if '三折' in n: return '04', 'high'
    if '双折' in n: return '03', 'high'
    if '钢夹' in n: return '05', 'high'
    if '卡包' in n or 'card' in n.lower(): return '06', 'high'
    if '长款' in n: return '03', 'mid'
    if '短款' in n: return '04', 'mid'
    return '03', 'low'

def classify_ow(name):
    """O-W 洗漱包"""
    n = str(name or '')
    if '套装' in n and '大' in n: return '01', 'high'
    if '套装' in n and '小' in n: return '02', 'high'
    if '套装' in n:               return '01', 'mid'
    if '单个' in n and '小' in n: return '03', 'high'
    if '单个' in n and '大' in n: return '04', 'high'
    if '小' in n:                 return '03', 'mid'
    if '大' in n:                 return '04', 'mid'
    return '01', 'low'

def classify_lp(name):
    """L-P 派对包"""
    n = str(name or '')
    if '满钻' in n or 'diamond' in n.lower(): return '01', 'high'
    if '硬壳' in n or '硬' in n:              return '02', 'high'
    return '01', 'low'

def classify_mw(name):
    """M-W 男士钱包"""
    n = str(name or '')
    if '护照' in n: return '02', 'high'
    if '卡包' in n: return '01', 'high'
    if '长款' in n: return '03', 'high'
    if '短款' in n: return '04', 'high'
    return '03', 'low'

def classify_children(name):
    """儿童包：返回 (l2key, l3code, conf)"""
    n = str(name or '')
    if '书包' in n:
        if any(k in n for k in ['毛', '绒', '布绒']): return 'S', '01', 'high'
        return 'S', '02', 'mid'
    if '编织' in n: return '', '02', 'mid'    # Woven Handbag
    return '', '01', 'low'                     # Small Handbag 默认

BAG_JDY_MAP = {
    'Lady Bag':            ('L', 'F'), "Lady Bag(CR)":       ('L', 'F'),
    'Lady Bag(JMG)':       ('L', 'F'), 'Lady Bag(QZHG)':     ('L', 'F'),
    'Lady Bag(Tote bag)':  ('L', 'F'), 'Lady Backpack':       ('L', 'S'),  # → School Bags
    'Lady Wallet':         ('L', 'W'), 'Party Bag':           ('L', 'P'),
    'Man Bag':             ('M', 'M'), 'Man Wallet':          ('M', 'W'),
    'Passport bag':        ('M', 'W'), 'Laptop bag':          ('M', 'C'),
    'Lunch Bag':           ('O', 'L'), 'Wool Bag':            ('O', 'H'),
    'Wash Bag':            ('O', 'W'), 'Handbag\u3001Canvas Bag': ('O', 'P'),
    'Woven bag':           ('O', 'S'), 'Wood bag':            ('O', 'C'),
}
AUTO_L3 = {
    ('L','S'):'01', ('M','M'):'01', ('M','C'):'01', ('M','S'):'01',
    ('M','O'):'01', ('M','B'):'01', ('O','L'):'01', ('O','H'):'01',
    ('O','P'):'01', ('O','S'):'01', ('O','C'):'01',
}

# 非箱包关键词（命中则跳过）
NON_BAG_KW = ['指甲', '梳子', '发夹', '发饰', '耳环', '项链', '戒指', '手链',
              '爪子', '汗带', '头绳', '发圈', '发卡', '配件', '钥匙扣', '香水',
              '口红', '眼镜', '手表', '镜子', '针', '线', 'Null', 'null']

def is_non_bag(name, jcat):
    n = str(name or '')
    for kw in NON_BAG_KW:
        if kw in n:
            return True
    # 空分类 + 名称不像包 → 跳过
    if not jcat and not any(k in n for k in ['包', 'bag', 'Bag', '袋', '囊', '夹']):
        return True
    return False

# ── 3. 主分类逻辑 ─────────────────────────────────────────────────────────────
with open(CACHE, encoding='utf-8') as f:
    prods = [p_ for p_ in json.load(f)['items'] if not p_.get('isDeleted')]

rows_out = []   # (pno, name, jcat, path, conf, note)
skipped  = 0

# 修正撇号：JDY 里可能用的是花引号
def norm_cat(c):
    return str(c or '').replace('\u2019', "'").replace('\u2018', "'").strip()

for prod in prods:
    pno   = prod.get('productNumber', '')
    name  = prod.get('productName', '')
    jcat  = norm_cat(prod.get('categoryName', ''))

    # 跳过明显非箱包
    if is_non_bag(name, jcat):
        skipped += 1
        continue

    path = conf = note = ''

    if jcat in BAG_JDY_MAP:
        l1, l2 = BAG_JDY_MAP[jcat]
        if (l1, l2) in AUTO_L3:
            l3 = AUTO_L3[(l1, l2)]
            path = p(l1, l2, l3)
            conf = 'high'
            note = ''
        elif (l1, l2) == ('L', 'F'):
            l3, conf = classify_lf(name)
            path = p(l1, l2, l3)
        elif (l1, l2) == ('L', 'W'):
            l3, conf = classify_lw(name)
            path = p(l1, l2, l3)
        elif (l1, l2) == ('L', 'P'):
            l3, conf = classify_lp(name)
            path = p(l1, l2, l3)
        elif (l1, l2) == ('M', 'W'):
            l3, conf = classify_mw(name)
            path = p(l1, l2, l3)
        elif (l1, l2) == ('O', 'W'):
            l3, conf = classify_ow(name)
            path = p(l1, l2, l3)
        else:
            path = p(l1, l2, '01')
            conf = 'mid'
        if conf == 'low':
            note = '名称无关键词，请核查'

    elif jcat == "Children's bag":
        l2k, l3, conf = classify_children(name)
        path = p('C', l2k, l3)
        if conf == 'low':
            note = '儿童包类型不明，默认手提小包'

    else:
        # YJ BAG 等其他分类，或空分类剩余
        # 按名称猜
        n = str(name or '')
        if '背包' in n or '双肩' in n:
            path, conf = p('M','S','01'), 'mid'
        elif '钱包' in n or '零钱' in n:
            path, conf = p('L','W','03'), 'mid'
        elif '腰包' in n:
            path, conf = p('M','O','01'), 'mid'
        elif '午餐' in n or '饭盒' in n:
            path, conf = p('O','L','01'), 'mid'
        elif '毛包' in n or '毛绒包' in n:
            path, conf = p('O','H','01'), 'mid'
        elif '草编' in n or '藤' in n:
            path, conf = p('O','S','01'), 'mid'
        elif '帆布' in n:
            path, conf = p('O','P','01'), 'mid'
        elif '洗漱' in n:
            path, conf = p('O','W','01'), 'mid'
        elif '书包' in n:
            path, conf = p('L','S','01'), 'mid'
        elif any(k in n for k in ['包包', '女包', '手挽', '手腕', '斜挎', '腋下']):
            l3, conf = classify_lf(name)
            path = p('L', 'F', l3)
        else:
            skipped += 1
            continue

    if not path:
        skipped += 1
        continue

    rows_out.append((pno, name, jcat, path, conf, note))

print(f'分类完成: {len(rows_out)} 条，跳过: {skipped} 条')

# ── 4. 输出 Excel ─────────────────────────────────────────────────────────────
FILL_HIGH = PatternFill('solid', fgColor='E2EFDA')   # 绿：高置信
FILL_MID  = PatternFill('solid', fgColor='FFFFFF')   # 白：中置信
FILL_LOW  = PatternFill('solid', fgColor='FCE4D6')   # 红：低置信需核查
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
DATA_FONT = Font(name='Calibri', size=10)
CENTER    = Alignment(horizontal='center', vertical='center')
LEFT      = Alignment(horizontal='left',   vertical='center')
thin = Side(style='thin', color='BDD7EE')
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# Sheet1：JDY 导入格式（直接可导入）
ws1 = wb.active
ws1.title = '导入格式'
headers1 = ['商品类别', '*商品编号', '商品名称', '置信度', '备注']
ws1.append(headers1)
for c in range(1, 6):
    cell = ws1.cell(1, c)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BDR

fill_map = {'high': FILL_HIGH, 'mid': FILL_MID, 'low': FILL_LOW}
conf_label = {'high': '高', 'mid': '中', 'low': '低⚠'}

for i, (pno, name, jcat, path, conf, note) in enumerate(rows_out):
    ws1.append([path, pno, name, conf_label.get(conf, conf), note])
    fill = fill_map.get(conf, FILL_MID)
    for c in range(1, 6):
        cell = ws1.cell(i + 2, c)
        cell.fill = fill
        cell.font = DATA_FONT
        cell.border = BDR
        cell.alignment = CENTER if c in (4,) else LEFT

col_widths = [60, 22, 28, 8, 20]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[ws1.cell(1, i).column_letter].width = w
ws1.row_dimensions[1].height = 20
ws1.freeze_panes = 'A2'

# Sheet2：分类统计
ws2 = wb.create_sheet('分类统计')
from collections import Counter
stats = Counter(r[4] for r in rows_out)
ws2.append(['置信度', '数量', '说明'])
ws2.append(['high', stats['high'], '自动填写，名称明确'])
ws2.append(['mid',  stats['mid'],  '名称推断，大概率正确'])
ws2.append(['low',  stats['low'],  '名称无关键词，请核查（红底行）'])
ws2.append(['跳过', skipped, '非箱包类或无法分类'])

wb.save(OUT_FILE)
print(f'\n✅ 已生成: {OUT_FILE}')
print(f'   high={stats["high"]}, mid={stats["mid"]}, low={stats["low"]}, 跳过={skipped}')
print(f'\n使用说明：')
print(f'  1. 检查红底行（低置信度），必要时手动修改商品类别列')
print(f'  2. 在JDY 箱包账套先导入/创建分类树（按钱姝包分类_整理版.xlsx）')
print(f'  3. 再用本文件做商品导入（只更新分类，其余字段可留空）')
