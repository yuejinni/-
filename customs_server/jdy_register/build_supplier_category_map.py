"""
从供应商缓存生成 supplier_category_map.json 完整初始模板
运行: python3 build_supplier_category_map.py
"""
import json, sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from jdy_cache import JDYCache

DIGIT_TO_LETTER = {'1':'A','2':'B','3':'C','4':'D','5':'E',
                   '6':'F','7':'G','8':'H','9':'I','0':'J'}
LETTER_TO_DIGIT = {chr(ord('A')+i): str(i+1) for i in range(26)}


def transform_vendor_code(code):
    s = str(code).strip().upper()
    if len(s) < 2:
        return s, None
    c1, c2 = s[0], s[1]
    rest = s[2:]
    if not c1.isalpha() and not c2.isalpha():
        nc1 = DIGIT_TO_LETTER.get(c1, 'J')
        nc2 = str((int(c2) + 1) % 10)
        return nc1 + nc2 + rest, None
    elif c1.isalpha():
        nc1 = LETTER_TO_DIGIT.get(c1, '?')
        if c2.isdigit():
            nc2 = str((int(c2) + 1) % 10)
            return nc1 + nc2 + rest, None
        else:
            return s, f'第二位是字母({c2})，需人工处理'
    else:
        return s, f'第一位数字第二位字母({s[:2]})，需人工处理'


cache = JDYCache(cfg_path='../../ai_config.json')

result = {
    '_说明': [
        'key = 变换后档口号（根据 taxPayerNo 变换）',
        'categories: 该供应商做哪些分类（填 letter，如 ["A","B"]）',
        'note: 备注',
    ],
    'account1': {},
    'account2': {},
}

for acc, label in [('account1', '祺航饰品'), ('account2', '祺航箱包')]:
    sups = cache.get_suppliers(acc)
    print(f'\n{label}: {len(sups)} 个供应商')
    skipped = 0
    for s in sups:
        tax_no = (s.get('taxPayerNo') or '').strip()
        if not tax_no:
            skipped += 1
            continue
        trans, err = transform_vendor_code(tax_no)
        entry = {
            'sup_name': s.get('name', ''),
            'sup_number': s.get('number', ''),
            'tax_no': tax_no,
            'transformed': trans,
            'categories': [],
            'note': err or '',
        }
        # 若变换后编码重复（两个供应商相同档口号），合并
        if trans in result[acc]:
            existing = result[acc][trans]
            existing['sup_name'] += f' / {entry["sup_name"]}'
        else:
            result[acc][trans] = entry

    print(f'  有 taxPayerNo: {len(result[acc])} 个，无 taxPayerNo(跳过): {skipped} 个')
    if skipped > 0:
        print(f'  ⚠️  {skipped} 个供应商无 taxPayerNo，不纳入建档体系')

# 保存
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'supplier_category_map.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print(f'\n✅ 已生成: {out_path}')

# 同时生成 Excel 供用户填写
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    for acc, label in [('account1', '祺航饰品'), ('account2', '祺航箱包')]:
        ws = wb.create_sheet(title=label)
        ws.append(['变换后档口号', '供应商名称', '供应商编号', 'taxPayerNo', '分类字母（填写）', '备注/错误'])
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='366092')
        for trans, v in result[acc].items():
            ws.append([trans, v['sup_name'], v.get('sup_number',''), v['tax_no'], '', v['note']])
            if v['note']:
                ws.cell(ws.max_row, 6).fill = PatternFill('solid', fgColor='FF9999')
        for col, w in enumerate([16,22,14,14,16,24], 1):
            ws.column_dimensions[ws.cell(1,col).column_letter].width = w

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    xls_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', '供应商大类映射表.xlsx')
    wb.save(xls_path)
    print(f'✅ Excel: {xls_path}')
except ImportError:
    print('openpyxl 未安装，跳过 Excel 生成')
