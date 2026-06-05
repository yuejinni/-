"""
构建/更新 category_code_map.json（新结构）

新结构字段说明:
  prefix:     产品编号点前部分（中类字母，如 'C' 或 'LF'）
  small_code: Excel 小类顺序码（如 '02'），产品编号后缀
  code_2d:    EAN-13 专用码（全局唯一 01-99，按账套分配）

运行方式:
    python build_category_map.py [--dry-run] [--excel 路径] [--out 路径]

依赖:
    pip install openpyxl
    ai_config.json 中配置了 jdy_client_id 等

分类 Excel 格式要求（钱姝包分类_整理版.xlsx）:
  - Sheet 名称随意
  - 必须有列: '分类路径'（如 Accessories 饰品/Jewelry 首饰/Bracelet）
  - 必须有列: 'prefix'（如 C 或 LF）
  - 必须有列: 'small_code'（如 02）
  - 必须有列: 'code_2d'（如 11，EAN-13 专用，全局唯一）
  - 可选列: 'account'（如 account1/account2，默认 account1）

注意:
  - 运行前先备份 category_code_map.json
  - 第一次运行建议加 --dry-run 验证输出
  - account2 的 code_2d 需要单独分配，不能与 account1 重复
"""

import json
import os
import sys
import argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

_HERE = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_EXCEL = '/Users/yuejin/Desktop/钱姝包分类_整理版.xlsx'
_DEFAULT_CFG = os.path.join(_HERE, '..', '..', 'ai_config.json')
_OUT_FILE = os.path.join(_HERE, '..', 'category_code_map.json')


def _load_existing(path) -> dict:
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _load_excel(excel_path: str) -> list[dict]:
    """
    读取 Excel 分类表，返回 [{path, prefix, small_code, code_2d, account}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        print('❌ 请先安装 openpyxl: pip install openpyxl')
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [str(c.value or '').strip() for c in ws[1]]
    print(f'[Excel] 读取到列: {headers}')

    required = ['分类路径', 'prefix', 'small_code', 'code_2d']
    for col in required:
        if col not in headers:
            print(f'❌ Excel 缺少列: {col!r}，请检查表头')
            sys.exit(1)

    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        path = str(row[idx['分类路径']] or '').strip()
        prefix = str(row[idx['prefix']] or '').strip()
        small_code = str(row[idx['small_code']] or '').strip().zfill(2)
        code_2d = str(row[idx['code_2d']] or '').strip().zfill(2)
        account = str(row[idx.get('account', -1)] or 'account1').strip() if 'account' in idx else 'account1'
        if path and prefix and small_code and code_2d:
            rows.append({'path': path, 'prefix': prefix,
                         'small_code': small_code, 'code_2d': code_2d,
                         'account': account})
    print(f'[Excel] 读取到 {len(rows)} 条有效分类行')
    return rows


def _fetch_jdy_categories(account: str, cfg_path: str) -> list[dict]:
    """
    从 JDY API 拉取分类列表。
    返回 [{id, name, fullName/path}, ...]
    """
    try:
        import jdy_api
        with open(cfg_path, encoding='utf-8') as f:
            cfg = json.load(f)

        if account == 'account1':
            cli = jdy_api.JDYClient(
                cfg['jdy_client_id'], cfg['jdy_app_key'],
                cfg['jdy_app_secret'], cfg['jdy_db_id'],
                cfg['jdy_domain'], client_secret=cfg['jdy_client_secret'],
            )
        else:
            cli = jdy_api.JDYClient(
                cfg['jdy2_client_id'], cfg['jdy2_app_key'],
                cfg['jdy2_app_secret'], cfg['jdy2_db_id'],
                cfg['jdy2_domain'], client_secret=cfg['jdy2_client_secret'],
            )

        result = cli._request('POST', '/jdyscm/category/list',
                              body={'filter': {'page': 1, 'pageSize': 500}},
                              query=cli._api_query(), timeout=15)
        items = result.get('items') or []
        print(f'[JDY] {account} 拉取到 {len(items)} 个分类')
        return items
    except Exception as e:
        print(f'[WARN] {account} 拉取分类失败: {e}，将跳过')
        return []


def _match_category(jdy_cats: list[dict], excel_path: str) -> dict | None:
    """
    按路径最后一段（叶子名称）在 JDY 分类列表中匹配。
    返回匹配到的 JDY 分类 dict 或 None。
    """
    leaf = excel_path.split('/')[-1].strip().lower()
    for c in jdy_cats:
        # JDY 分类可能有 name / fullName / path 字段
        names = [
            str(c.get('name') or '').strip().lower(),
            str(c.get('fullName') or '').strip().lower(),
        ]
        if any(leaf in n or n in leaf for n in names if n):
            return c
    return None


def build_map(excel_path: str, cfg_path: str, existing: dict,
              dry_run: bool = False) -> dict:
    """
    构建新版 category_code_map.json。

    逻辑:
      1. 读 Excel → 按 account 分组
      2. 对每个 account，从 JDY API 拉分类列表
      3. 对每个 Excel 行，按分类路径叶子名称匹配 JDY 分类 ID
      4. 已有 jdy_cat_id 的保留，未匹配到的打印警告
      5. 保留已有 category_code_map.json 中其他 account 数据（兼容性）
    """
    excel_rows = _load_excel(excel_path)

    # 按 account 分组
    by_account: dict[str, list] = {}
    for row in excel_rows:
        acc = row['account']
        by_account.setdefault(acc, []).append(row)

    new_map: dict = {}
    # 先复制已有数据（兼容性）
    for k, v in existing.items():
        if k != '_note':
            new_map[k] = dict(v) if isinstance(v, dict) else v

    for account, rows in by_account.items():
        print(f'\n[{account}] 处理 {len(rows)} 条分类...')
        jdy_cats = _fetch_jdy_categories(account, cfg_path)

        acc_map = new_map.get(account, {})
        matched = 0
        unmatched = []

        for row in rows:
            jdy_cat = _match_category(jdy_cats, row['path'])
            if not jdy_cat:
                unmatched.append(row['path'])
                continue

            cat_id = str(jdy_cat.get('id') or jdy_cat.get('categoryId') or '')
            if not cat_id:
                unmatched.append(row['path'])
                continue

            acc_map[cat_id] = {
                'path':       row['path'],
                'prefix':     row['prefix'],
                'small_code': row['small_code'],
                'code_2d':    row['code_2d'],
            }
            matched += 1

        new_map[account] = acc_map
        print(f'[{account}] 匹配成功 {matched} 个，未匹配 {len(unmatched)} 个')
        if unmatched:
            print(f'[{account}] 未匹配分类（请手动补充 JDY 分类 ID）:')
            for p in unmatched:
                print(f'  - {p}')

    new_map['_note'] = (f'自动生成于 {datetime.now().strftime("%Y-%m-%d %H:%M")}，'
                        '字段说明: prefix=产品编号点前部分, small_code=小类顺序码, '
                        'code_2d=EAN-13专用码')

    if dry_run:
        print('\n[dry-run] 输出预览（前3个 account1 条目）:')
        for k, v in list(new_map.get('account1', {}).items())[:3]:
            print(f'  {k}: {json.dumps(v, ensure_ascii=False)}')
    else:
        # 备份
        bak = _OUT_FILE + '.bak'
        if os.path.exists(_OUT_FILE):
            import shutil
            shutil.copy2(_OUT_FILE, bak)
            print(f'\n[备份] {bak}')

        with open(_OUT_FILE, 'w', encoding='utf-8') as f:
            json.dump(new_map, f, ensure_ascii=False, indent=2)
        print(f'\n✅ 已写入 {_OUT_FILE}')

    return new_map


def migrate_letter_to_prefix(existing: dict, out_path: str, dry_run: bool = False) -> dict:
    """
    迁移旧 category_code_map.json（letter/code_2d）到新格式（prefix/small_code/code_2d）。
    - letter → prefix
    - code_2d → 保留为 code_2d（EAN 码）
    - small_code → 若无则从 code_2d 复制（过渡方案，后续可用 Excel 覆盖）

    用法：仅在没有 Excel 文件时，用于快速迁移旧数据。
    """
    new_map = {}
    for k, v in existing.items():
        if k == '_note':
            continue
        new_acc = {}
        for cat_id, info in (v.items() if isinstance(v, dict) else {}.items()):
            if not isinstance(info, dict):
                continue
            new_info = dict(info)
            # 迁移 letter → prefix
            if 'letter' in new_info and 'prefix' not in new_info:
                new_info['prefix'] = new_info.pop('letter')
            # 若无 small_code，用 code_2d 过渡
            if 'small_code' not in new_info and 'code_2d' in new_info:
                new_info['small_code'] = new_info['code_2d']
            new_acc[cat_id] = new_info
        new_map[k] = new_acc

    new_map['_note'] = (f'从旧格式迁移于 {datetime.now().strftime("%Y-%m-%d %H:%M")}，'
                        'prefix 从 letter 迁移，small_code 暂用 code_2d 填充')

    if dry_run:
        count = sum(len(v) for v in new_map.values() if isinstance(v, dict))
        print(f'[dry-run] 迁移完成，共 {count} 条（未写入文件）')
    else:
        bak = out_path + '.bak'
        if os.path.exists(out_path):
            import shutil
            shutil.copy2(out_path, bak)
            print(f'[备份] {bak}')
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(new_map, f, ensure_ascii=False, indent=2)
        print(f'✅ 迁移完成，已写入 {out_path}')
    return new_map


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='构建/更新 category_code_map.json')
    parser.add_argument('--excel', default=_DEFAULT_EXCEL, help='Excel 文件路径')
    parser.add_argument('--cfg',   default=_DEFAULT_CFG,   help='ai_config.json 路径')
    parser.add_argument('--out',   default=_OUT_FILE,       help='输出文件路径')
    parser.add_argument('--dry-run', action='store_true',   help='仅预览，不写入文件')
    parser.add_argument('--migrate-only', action='store_true',
                        help='仅做旧格式迁移（letter→prefix），不需要 Excel 文件')
    args = parser.parse_args()

    existing = _load_existing(args.out)

    if args.migrate_only:
        print('=== 旧格式迁移模式（letter → prefix）===')
        migrate_letter_to_prefix(existing, args.out, dry_run=args.dry_run)
    else:
        if not os.path.exists(args.excel):
            print(f'❌ Excel 文件不存在: {args.excel}')
            print('提示：如果只想迁移旧格式，使用 --migrate-only 参数')
            sys.exit(1)
        print(f'=== 构建 category_code_map.json ===')
        print(f'Excel: {args.excel}')
        build_map(args.excel, args.cfg, existing, dry_run=args.dry_run)
