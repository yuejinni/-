"""
Excel 生成器
生成：出口报关单、Proforma Invoice、装箱单
"""
import os
import re
import sys
import shutil
import json
import random
import ssl
import threading
import urllib.request
import urllib.parse
from datetime import datetime, date, timedelta
from collections import defaultdict

# 基础资料文件写锁（多线程并发时保护 BASE_DATA xlsx）
_base_data_lock = threading.Lock()

import openpyxl
from openpyxl.styles import Font as _Font
import xlrd
import zipfile
import io
try:
    from lxml import etree as _etree
    _LXML = True
except ImportError:
    import xml.etree.ElementTree as _etree
    _etree.register_namespace('xdr', 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing')
    _LXML = False

_XDR_NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'


def _shift_drawing_rows(xlsx_path, insert_before_row_1based, n_inserted):
    """
    将 xlsx 文件中所有浮动图片锚点，若其 xdr:from/xdr:to 的行号
    >= insert_before_row_1based (1-based)，则加上 n_inserted。
    直接修改 xlsx 文件。
    """
    if n_inserted <= 0:
        return
    threshold = insert_before_row_1based - 1   # XML 行号是 0-based

    buf = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
         zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if (item.filename.startswith('xl/drawings/drawing')
                    and item.filename.endswith('.xml')):
                tree = _etree.fromstring(data)
                for tag in ('from', 'to'):
                    for row_el in tree.iter(f'{{{_XDR_NS}}}{tag}'):
                        r_el = row_el.find(f'{{{_XDR_NS}}}row')
                        if r_el is not None and r_el.text:
                            try:
                                rv = int(r_el.text)
                                if rv >= threshold:
                                    r_el.text = str(rv + n_inserted)
                            except ValueError:
                                pass
                if _LXML:
                    data = _etree.tostring(
                        tree, xml_declaration=True,
                        encoding='UTF-8', standalone=True)
                else:
                    data = _etree.tostring(tree, encoding='unicode').encode('utf-8')
            zout.writestr(item, data)

    with open(xlsx_path, 'wb') as f:
        f.write(buf.getvalue())

# ── 路径配置 ──────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJ = os.path.join(_HERE, '..')

# PyInstaller 打包后：可写文件放 exe 同目录
if getattr(sys, 'frozen', False):
    _EXE_DIR   = os.path.dirname(sys.executable)
    _DATA_BASE = _EXE_DIR
else:
    _EXE_DIR   = None
    _DATA_BASE = _PROJ

# items.json 统一存储目录（所有批次共享，供 Stage 2 列表选择）
_ITEMS_STORE = os.path.join(_DATA_BASE, '_items_store')


def _find_tpl(filename):
    """
    模板文件查找顺序：
      1. exe 同目录的外部文件（放在旁边可随时替换，无需重新打包）
      2. PyInstaller 内嵌文件（sys._MEIPASS）
      3. 开发时的项目目录（../）
    """
    if _EXE_DIR:
        ext = os.path.join(_EXE_DIR, filename)
        if os.path.exists(ext):
            return ext
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(_PROJ, filename)


BASE_DATA      = _find_tpl('报关产品基础资料（智谱）.xlsx')
TPL_CUSTOMS    = _find_tpl('出口报关单.xlsx')
TPL_INVOICE    = _find_tpl('Proforma Invoice 形式发票.xlsx')
TPL_PACKING    = _find_tpl('装箱单 Packing List.xlsx')
TPL_SALES      = _find_tpl('QIHANG销售合同.xlsx')
TPL_PURCHASE   = _find_tpl('工厂采购合同用.xlsx')
TPL_INV_IMPORT = _find_tpl('发票开具项目信息导入模板.xlsx')
TPL_LINGHANG   = _find_tpl('副本金华市凌航国际贸易有限发票模板excel.xlsx')
_PI_COUNTER    = os.path.join(_DATA_BASE, '.pi_counter.json')  # 每日 PI 流水计数器

# ── 固定信息 ──────────────────────────────────────────────────────────────
BUYER_NAME  = 'QI HANG GENERAL TRADING L.L.C'
BUYER_ADDR  = 'SHOP 14.DUBAI WHOLESALE PLAZA,AL RAS,DEIRA,DUBAI,UAE'
BUYER_TEL   = '+971508691698'
DEST_COUNTRY_CN = '阿联酋'
DEST_COUNTRY_EN = 'UAE'
PORT_LOAD   = 'NINGBO'
PORT_DISC   = 'JEBEL ALI, UAE'
ORIGIN_CN   = '中国'


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. 读取基础资料
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_base_data():
    wb = openpyxl.load_workbook(BASE_DATA, data_only=True)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[1]
        if not code:
            continue
        # 新表：N(13)=长(箱), O(14)=宽(箱), P(15)=高(箱), Q(16)=体积
        dim_l = row[13]
        dim_w = row[14]
        dim_h = row[15]
        dim_vol = row[16]   # Q列体积（cm³）

        # N/O/P 都为空则 dimensions 留空
        if dim_l or dim_w or dim_h:
            dimensions = {
                'l':   float(dim_l) if dim_l else 0.0,
                'w':   float(dim_w) if dim_w else 0.0,
                'h':   float(dim_h) if dim_h else 0.0,
                'vol': float(dim_vol) if dim_vol else None,
            }
        else:
            dimensions = None

        # S列(row[18]) = 商品备注，用作 PCS 换算系数（每包/箱含多少个 PCS）
        pcs_per_pkg = 1  # 默认 1（不换算）
        if len(row) > 18 and row[18] is not None:
            try:
                v = float(row[18])
                if v > 0:
                    pcs_per_pkg = v
            except (ValueError, TypeError):
                pass

        # R列(index 17) = 是否退税（'是'/'否'，空默认'是'）
        is_tax_refund_raw = row[17] if len(row) > 17 else None
        is_tax_refund = str(is_tax_refund_raw).strip() if is_tax_refund_raw else '是'

        # V列(index 21) = 商品和服务税收编码
        tax_code = row[21] if len(row) > 21 else None

        # W列(index 22) = 英文报关品名
        en_customs_name_raw = row[22] if len(row) > 22 else None

        # X列(index 23) = 项目名称（凌航发票用）
        project_name_raw = row[23] if len(row) > 23 else None
        project_name = str(project_name_raw).strip() if project_name_raw else ''

        data[code] = {
            'code':            code,
            'cn_name':         row[2],
            'category':        row[3] or '',
            'hs_code':         str(int(row[4])) if row[4] else '',
            'customs_name':    row[5] or '',
            'unit':            row[6] or 'PAC',
            'origin':          row[7] or '义乌',
            'material':        row[8] or '',
            'usage':           row[9] or '',
            'gw_per_pkg':      float(row[11]) if row[11] else None,   # L列 毛重(KGS/包)
            'nw_per_pkg':      float(row[12]) if row[12] else None,   # M列 净重(KGS/包)
            'gross_weight':    float(row[11]) if row[11] else 0.0,    # 向后兼容
            'net_weight':      float(row[12]) if row[12] else 0.0,    # 向后兼容
            'dimensions':      dimensions,
            'pcs_per_pkg':     pcs_per_pkg,  # S列换算系数
            'tax_code':        str(tax_code) if tax_code else '',      # V列 税收编码
            'en_customs_name': str(en_customs_name_raw).strip() if en_customs_name_raw else '',  # W列
            'is_tax_refund':   is_tax_refund,   # R列 是否退税
            'project_name':    project_name,    # X列 项目名称（凌航发票用）
        }
    return data


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 数据处理
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _en_category(cat_str):
    """从商品类别提取末段英文名，如 'Foundation'"""
    if not cat_str:
        return ''
    parts = cat_str.split('/')
    last = parts[-1].strip()
    # 取空格前的英文（或整体）
    return last.split(' ')[0] if ' ' in last else last


def _calc_cbm(dimensions, qty):
    """根据 dimensions dict {'l','w','h'} 和数量计算体积 m³
    如果 dimensions 为 None 或 l/w/h 全为 0，返回 0
    注意：qty 应为原始包数（非 PCS 换算后的数量），因为体积按箱计算
    """
    if not dimensions:
        return 0.0
    try:
        l = dimensions.get('l', 0) or 0
        w = dimensions.get('w', 0) or 0
        h = dimensions.get('h', 0) or 0
        if l == 0 and w == 0 and h == 0:
            return 0.0
        return round(float(l) * float(w) * float(h) / 1_000_000 * qty, 4)
    except Exception:
        return 0.0


def process_items(raw_items, exchange_rate, target_total_usd=None,
                  weighed_total_gw=None, api_dims_map=None, origin_map=None, total_cbm=None):
    """
    raw_items: [{'code': str, 'unit': str, 'qty': float, 'rmb_price': float,
                 'remark': str, 'boxes': int|None}]
    weighed_total_gw: 货物过磅总重量(KGS)，可选，用于分配 G.W.
    api_dims_map: {code: {'l','w','h'}} 来自精斗云 registrationNo
    origin_map:   {code: city_str} 来自供应商联系人城市
    返回处理好的 items 列表
    """
    base = load_base_data()
    if api_dims_map is None:
        api_dims_map = {}
    if origin_map is None:
        origin_map = {}
    result = []

    for item in raw_items:
        code = item['code']
        qty  = float(item.get('qty', 1))
        rmb  = float(item.get('rmb_price', 0))
        unit = item.get('unit', '').strip()

        # 若 rmb_price 为 0，从备注中解析 ￥ 后的价格（格式：{箱数}*...￥{价格} 或 商品备注直接为数字）
        remark = item.get('remark', '') or ''
        if rmb == 0 and remark:
            m_price = re.search(r'[¥￥]([\d]+(?:\.[\d]+)?)', remark)
            if m_price:
                try:
                    rmb = float(m_price.group(1))
                except (ValueError, TypeError):
                    pass
            elif re.fullmatch(r'[\d]+(?:\.[\d]+)?', remark.strip()):
                # 商品备注本身就是一个纯数字价格
                try:
                    rmb = float(remark.strip())
                except (ValueError, TypeError):
                    pass

        bd = base.get(code)
        if not bd:
            # 未在基础资料中找到的商品，跳过并记录
            result.append({
                'code': code, 'found': False,
                'qty': qty, 'unit': unit, 'rmb_price': rmb,
                'usd_price': 0, 'total_usd': 0,
            })
            continue

        # S列换算系数：数量×系数，价格÷系数，单位→PCS
        pcs_per_pkg = bd.get('pcs_per_pkg', 1)
        converted_qty = qty * pcs_per_pkg
        # 保留4位小数，减少累积误差；Excel 显示格式由模板单元格格式控制
        usd_price = round(rmb / exchange_rate / pcs_per_pkg, 4) if exchange_rate > 0 and rmb > 0 else 0.0
        total_usd = round(usd_price * converted_qty, 4)

        # 解析装箱总件数：优先使用前端直接传入的 boxes，否则从备注解析
        boxes = item.get('boxes')  # 前端直接传入的箱数（优先）
        if boxes is None or boxes == '' or boxes == 0:
            boxes = 0
            m = re.match(r'^(\d+)\*', remark.strip())
            if m:
                boxes = int(m.group(1))

        result.append({
            'found':           True,
            'code':            code,
            'qty':             qty,                     # 原始包数（未换算）
            'unit':            unit or bd['unit'] or 'PCS',   # G列为空时默认 PCS
            'customs_unit':    bd['unit'] or 'PCS',         # G列为空时默认 PCS
            'rmb_price':       rmb,
            'usd_price':       usd_price,              # 已除以 pcs_per_pkg
            'total_usd':       total_usd,
            'hs_code':         bd['hs_code'],
            'customs_name':    bd['customs_name'],
            'cn_name':         bd['cn_name'],
            'en_category':     _en_category(bd['category']),
            'en_customs_name': bd.get('en_customs_name', ''),   # W列 英文报关品名
            'origin':          origin_map.get(code) or bd['origin'],
            'material':        bd['material'],
            'usage':           bd['usage'],
            'gw_per_pkg':      bd['gw_per_pkg'],   # L列 毛重(KGS/包)，None表示无数据
            'nw_per_pkg':      bd['nw_per_pkg'],   # M列 净重(KGS/包)，None表示无数据
            'gross_weight':    bd['gross_weight'],
            'net_weight':      bd['net_weight'],
            'dimensions':      bd['dimensions'],
            'api_dimensions':  api_dims_map.get(code),   # API 尺寸（优先级高于基础资料）
            'boxes':           boxes,
            'pcs_per_pkg':     pcs_per_pkg,         # 保留系数供后续使用
            'tax_code':        bd.get('tax_code', ''),  # V列 商品税收编码
            'is_tax_refund':   bd.get('is_tax_refund', '是'),  # R列 是否退税
            'project_name':    bd.get('project_name', ''),     # X列 项目名称
        })

    found = [i for i in result if i.get('found')]

    # ── 计算毛重 G.W. / 净重 N.W. ──────────────────────────────
    # G.W. 规则：
    #   1) L列有毛重 → gw = gw_per_pkg × 原始包数(qty)
    #   2) L列无毛重 + 有过磅总重量 → 先扣除有L列数据的商品重量，
    #      剩余重量按包数平均分配给无L列数据的商品
    #   3) L列无毛重 + 无过磅 → gw = 0
    # N.W. 规则：
    #   1) M列有净重 → nw = nw_per_pkg × 原始包数
    #   2) M列无净重 → nw = gw × 0.95
    if found:
        # 先计算有 L 列数据的商品总毛重
        known_gw_total = 0.0
        needs_estimate = []
        for i in found:
            pkg_qty = i['qty'] / i.get('pcs_per_pkg', 1)  # 换回原始包数
            if i['gw_per_pkg'] is not None:
                known_gw_total += i['gw_per_pkg'] * pkg_qty
            else:
                needs_estimate.append(i)

        # 对需要估算的商品，按包数分配剩余重量
        if needs_estimate and weighed_total_gw and float(weighed_total_gw) > 0:
            remaining_weight = float(weighed_total_gw) - known_gw_total
            estimate_pkg_qty = sum(i['qty'] / i.get('pcs_per_pkg', 1) for i in needs_estimate)
            weight_per_pkg = remaining_weight / estimate_pkg_qty if estimate_pkg_qty > 0 else 0

        for i in found:
            pkg_qty = i['qty'] / i.get('pcs_per_pkg', 1)  # 换回原始包数
            # ── G.W. 计算 ──
            if i['gw_per_pkg'] is not None:
                gw_val = round(i['gw_per_pkg'] * pkg_qty, 2)
            elif needs_estimate and weighed_total_gw and float(weighed_total_gw) > 0:
                gw_val = round(weight_per_pkg * pkg_qty, 2)
            else:
                gw_val = 0.0

            # ── N.W. 计算 ──
            if i['nw_per_pkg'] is not None:
                nw_val = round(i['nw_per_pkg'] * pkg_qty, 2)
            else:
                nw_val = round(gw_val * 0.95, 2)

            i['gw_val'] = gw_val
            i['nw_val'] = nw_val

    # ── 按包数比例分配总 CBM ────────────────────────────────
    if total_cbm and float(total_cbm) > 0 and found:
        total_boxes_all = sum(i.get('boxes', 0) for i in found)
        if total_boxes_all > 0:
            for i in found:
                i['cbm_val'] = round(i.get('boxes', 0) / total_boxes_all * float(total_cbm), 4)

    # ── 按目标总价等比例调整单价 ──────────────────────────────
    if target_total_usd and float(target_total_usd) > 0 and found:
        current = sum(i['total_usd'] for i in found)
        if current > 0:
            # rmb_price 有效 → 按相对价值等比缩放 USD
            ratio = float(target_total_usd) / current
            for i in found:
                i['usd_price'] = round(i['usd_price'] * ratio, 4)
                i['total_usd'] = round(i['total_usd'] * ratio, 4)  # 按比例缩放，不依赖 qty 单位
                # 同步更新 rmb_price（反算），保证采购合同/凌航发票金额一致
                i['rmb_price'] = round(i['usd_price'] * exchange_rate, 4)
        else:
            # rmb_price 全为 0（调拨单无价格）→ 按包数等比分配目标 USD 总额
            total_qty_pkg = sum(i['qty'] for i in found)
            if total_qty_pkg > 0 and exchange_rate > 0:
                usd_per_pkg = float(target_total_usd) / total_qty_pkg
                for i in found:
                    i['usd_price'] = round(usd_per_pkg, 4)
                    i['total_usd'] = round(usd_per_pkg * i['qty'], 4)
                    i['rmb_price'] = round(usd_per_pkg * exchange_rate, 4)

    return result


def merge_for_customs(items):
    """
    相同海关编码 + 相同报关产品单位 → 合并为一行
    海关编码相同但单位不同 → 不合并
    """
    groups = {}
    for item in items:
        if not item.get('found'):
            continue
        key = (item['hs_code'], item['customs_name'], item['customs_unit'], item['origin'], item['material'])
        groups.setdefault(key, []).append(item)

    def _dedup_join(values):
        """把多个可能含顿号的字符串拆开，统一去重后拼回。"""
        seen, result = set(), []
        for v in values:
            for part in (v or '').split('、'):
                part = part.strip()
                if part and part not in seen:
                    seen.add(part)
                    result.append(part)
        return '、'.join(result)

    merged = []
    for idx, (hs, grp) in enumerate(groups.items()):
        total_qty  = sum(i['qty'] * i.get('pcs_per_pkg', 1) for i in grp)  # PCS
        pkg_count  = sum(i['qty'] for i in grp)                            # 原始包数（装箱/CBM用）
        total_usd  = sum(i['total_usd'] for i in grp)
        avg_price  = round(total_usd / total_qty, 4) if total_qty else 0

        hs_code, customs_name, customs_unit, origin, material = hs
        usages    = _dedup_join(i['usage'] for i in grp)
        spec_desc = f"0|2|{material}|{usages}|无品牌|无型号"
        merged.append({
            'hs_code':         hs_code,
            'customs_unit':    customs_unit,
            'customs_name':    customs_name,
            'spec_desc':       spec_desc,
            'origin':          origin,
            'material':        material,
            'qty':             total_qty,    # PCS（报关单/发票/销售合同用）
            'pkg_count':       pkg_count,    # 原始包数（装箱单用）
            'item_no':         idx + 1,      # 与报关单项号一致
            'usd_price':       avg_price,
            'total_usd':       total_usd,
            'gross_weight':    round(sum(i.get('gw_val', i.get('gross_weight', 0) * i['qty']) for i in grp), 3),
            'net_weight':      round(sum(i.get('nw_val', i.get('net_weight', 0) * i['qty'])   for i in grp), 3),
            'boxes':           sum(i.get('boxes', 0) for i in grp),
            # 装箱单需要的额外字段（同组取第一条）
            'en_category':     grp[0].get('en_category', ''),
            'en_customs_name': grp[0].get('en_customs_name', ''),
            'cn_name':         grp[0].get('cn_name', ''),
            'dimensions':      grp[0].get('dimensions', ''),
            'api_dimensions':  grp[0].get('api_dimensions'),
            'pcs_per_pkg':     grp[0].get('pcs_per_pkg', 1),  # 保留换算系数
        })
    return merged


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 出口报关单
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 模板结构（每个商品占 3 行）：
#   行 A+0 : B=项号  C=商品编号  E=商品名称  P=数量   T=单价  W=中国     Y=阿联酋   AB=境内货源地
#   行 A+1 : E=规格型号(0|2|材料|用途|无品牌|无型号)  T=总价  W=(CN)    Y=(ARE)
#   行 A+2 : P=单位  T=币制
#
# 第 1 页：6 个商品，A 起始：23, 26, 29, 32, 35, 38
# 第 2+ 页：从 模板 (2) 复制追加，每页 11 个商品，产品区偏移 +7 行

_CUSTOMS_ROWS = [23, 26, 29, 32, 35, 38]   # 第 1 页每组起始行

# E:O 合并格列宽合计（从模板读取：5.875+5.625+4.125+4.375+4+4.25+3.625+13+13+13+5.875）
_EO_CELL_WIDTH = 76.75


def _auto_font_size(text):
    """
    估算使文字在 E:O 合并单元格内不换行所需的最小字号。
    CJK 字符占 2 单位，ASCII 占 1 单位；10pt 基准，按比例缩放。
    """
    dw = sum(
        2 if ('\u4e00' <= c <= '\u9fff' or '\u3000' <= c <= '\u303f' or '\uff00' <= c <= '\uffef')
        else 1
        for c in (text or '')
    )
    max10 = _EO_CELL_WIDTH * 0.7   # ≈53.7，含字间距/边距经验修正
    for size in [10, 9, 8, 7, 6]:
        if dw <= max10 * (10.0 / size):
            return size
    return 6


def _set_font_size(cell, size):
    """仅修改单元格字号，保留原字体名称及其他属性。"""
    f = cell.font
    cell.font = _Font(name=f.name or '宋体', size=size,
                      bold=f.bold, italic=f.italic,
                      color=f.color, underline=f.underline, strike=f.strike)


def _write_customs_row(ws, row_a, item_no, item):
    """向 ws 写入一个报关商品行（3 行）。"""
    ws.cell(row_a,     2).value = item_no              # B  = 项号
    ws.cell(row_a,     3).value = item['hs_code']      # C  = 商品编号
    c = ws.cell(row_a, 5);     c.value = item['customs_name']; _set_font_size(c, _auto_font_size(item['customs_name']))   # E  = 商品名称
    c = ws.cell(row_a + 1, 5); c.value = item['spec_desc'];    _set_font_size(c, _auto_font_size(item['spec_desc']))      # E+1= 规格型号
    qty_str = str(int(item['qty'])) if item['qty'] == int(item['qty']) else str(item['qty'])
    ws.cell(row_a,    16).value = qty_str + ' ' + item['customs_unit']  # P   = 数量 单位
    ws.cell(row_a + 2,16).value = str(item['net_weight']) + ' 千克'      # P+2 = 净重
    ws.cell(row_a,    20).value = item['usd_price']    # T  = 单价
    ws.cell(row_a + 1,20).value = item['total_usd']   # T+1= 总价
    ws.cell(row_a + 2,20).value = '美元'               # T+2= 币制
    ws.cell(row_a,    23).value = '中国'               # W  = 原产国
    ws.cell(row_a + 1,23).value = '(CN)'               # W+1= 原产国代码
    ws.cell(row_a,    25).value = '阿联酋'             # Y  = 目的国
    ws.cell(row_a + 1,25).value = '(ARE)'              # Y+1= 目的国代码
    ws.cell(row_a,    28).value = item['origin']       # AB = 境内货源地


def _copy_template2(src_ws, dest_ws, copy_start_row):
    """
    将 src_ws（模板 (2)）全部行复制到 dest_ws，
    第 1 行对齐到 dest_ws 的 copy_start_row。
    复制内容：单元格值、样式、合并区域、行高。
    """
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter

    row_delta = copy_start_row - 1

    for row in src_ws.iter_rows():
        for cell in row:
            dest = dest_ws.cell(cell.row + row_delta, cell.column)
            dest.value = cell.value
            if cell.has_style:
                dest.font        = _copy(cell.font)
                dest.border      = _copy(cell.border)
                dest.fill        = _copy(cell.fill)
                dest.alignment   = _copy(cell.alignment)
                dest.number_format = cell.number_format

    for mc in list(src_ws.merged_cells.ranges):
        new_rng = (
            f"{get_column_letter(mc.min_col)}{mc.min_row + row_delta}"
            f":{get_column_letter(mc.max_col)}{mc.max_row + row_delta}"
        )
        dest_ws.merged_cells.add(new_rng)

    for r_idx, rd in src_ws.row_dimensions.items():
        if rd.height:
            dest_ws.row_dimensions[r_idx + row_delta].height = rd.height


def generate_customs(items, invoice_no, output_path):
    wb  = openpyxl.load_workbook(TPL_CUSTOMS)
    ws  = wb['模板']
    ws2 = wb['模板 (2)']

    merged = merge_for_customs(items)
    found  = [i for i in items if i.get('found')]

    # ── 汇总行 ────────────────────────────────────────────────
    total_gw    = round(sum(i.get('gw_val', i.get('gross_weight', 0) * i['qty']) for i in found), 3)
    total_nw    = round(sum(i.get('nw_val', i.get('net_weight', 0) * i['qty'])   for i in found), 3)
    total_boxes = sum(i.get('boxes', 0) for i in found)

    ws.cell(14, 2).value  = invoice_no   # B14 = 合同协议号
    ws.cell(16, 12).value = total_gw     # L16 = 毛重（千克）
    ws.cell(16, 16).value = total_nw     # P16 = 净重（千克）
    if total_boxes > 0:
        ws.cell(16, 10).value = total_boxes  # J16 = 件数

    # ── 第 1 页：最多 6 个商品 ─────────────────────────────────
    for idx, item in enumerate(merged[:6]):
        _write_customs_row(ws, _CUSTOMS_ROWS[idx], idx + 1, item)

    # ── 第 2+ 页：从 模板 (2) 追加，每页 11 个商品 ────────────
    # 模板 (2) 共 44 行，产品区从第 8 行开始
    # 追加到主 sheet 时：copy_start=45,89,133,...
    # 产品起始行 = copy_start + 7（第 8 行偏移 7）
    remaining = merged[6:]
    page_num  = 0
    while remaining:
        page_items = remaining[:11]
        remaining  = remaining[11:]

        copy_start    = 45 + page_num * 44
        product_start = copy_start + 7       # 模板2 第 8 行 = copy_start+7

        _copy_template2(ws2, ws, copy_start)

        for pidx, item in enumerate(page_items):
            row_a   = product_start + pidx * 3
            item_no = 7 + page_num * 11 + pidx
            _write_customs_row(ws, row_a, item_no, item)

        page_num += 1

    out = os.path.join(output_path, f'出口报关单_{invoice_no}.xlsx')
    wb.save(out)
    return out


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. Proforma Invoice 流水号
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _number_to_english(amount):
    """将金额数字转为英文大写书写，如 12345.67 →
    TWELVE THOUSAND THREE HUNDRED AND FORTY-FIVE AND CENTS SIXTY-SEVEN
    """
    _ONES = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX',
             'SEVEN', 'EIGHT', 'NINE', 'TEN', 'ELEVEN', 'TWELVE',
             'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN',
             'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
    _TENS = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY',
             'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']

    def _below_1000(n):
        parts = []
        if n >= 100:
            parts.append(f'{_ONES[n // 100]} HUNDRED')
            n %= 100
            if n:
                parts.append('AND')
        if n >= 20:
            parts.append(_TENS[n // 10])
            if n % 10:
                parts[-1] += f'-{_ONES[n % 10]}'
        elif n > 0:
            parts.append(_ONES[n])
        return ' '.join(parts)

    if amount < 0:
        return 'NEGATIVE ' + _number_to_english(-amount)

    # 分离整数和小数
    integer = int(amount)
    cents   = round((amount - integer) * 100)

    parts = []
    if integer == 0:
        parts.append('ZERO')
    else:
        billions = integer // 1_000_000_000
        millions = (integer // 1_000_000) % 1000
        thousands = (integer // 1000) % 1000
        remainder = integer % 1000

        if billions:
            parts.append(f'{_below_1000(billions)} BILLION')
        if millions:
            parts.append(f'{_below_1000(millions)} MILLION')
        if thousands:
            parts.append(f'{_below_1000(thousands)} THOUSAND')
        if remainder:
            r_text = _below_1000(remainder)
            if parts and 'AND' not in r_text:
                parts.append('AND')
            parts.append(r_text)

    result = ' '.join(parts)
    if cents:
        result += f' AND CENTS {_below_1000(cents)}'
    return result


def rmb_to_chinese(amount):
    """人民币大写，精确到分。例：12345.67 → 壹万贰仟叁佰肆拾伍元陆角柒分"""
    digits = '零壹贰叁肆伍陆柒捌玖'

    def _section(n):
        """将 0-9999 转为汉字，不含末尾组别单位（万/亿）。"""
        if n == 0:
            return ''
        su = ['仟', '佰', '拾', '']
        vs = [n // 1000, (n % 1000) // 100, (n % 100) // 10, n % 10]
        res, prev_zero = '', False
        for i, v in enumerate(vs):
            if v == 0:
                prev_zero = True
            else:
                if prev_zero and res:
                    res += '零'
                res += digits[v] + su[i]
                prev_zero = False
        return res

    if amount == 0:
        return '零元整'
    neg = amount < 0
    amount = abs(round(float(amount), 2))
    integer = int(amount)
    decimal = round((amount - integer) * 100)
    jiao, fen = decimal // 10, decimal % 10

    yi   = integer // 100_000_000
    wan  = (integer % 100_000_000) // 10_000
    yuan = integer % 10_000

    int_str = ''
    if yi:
        int_str += _section(yi) + '亿'
    if wan:
        if yi and wan < 1000:
            int_str += '零'
        int_str += _section(wan) + '万'
        if 0 < yuan < 1000:
            int_str += '零'
    if yuan:
        int_str += _section(yuan)

    result = int_str + '元' if int_str else ''

    if jiao == 0 and fen == 0:
        result += '整'
    elif jiao == 0:
        if int_str:
            result += '零'
        else:
            result = '零'
        result += digits[fen] + '分'
    elif fen == 0:
        result += digits[jiao] + '角整'
    else:
        result += digits[jiao] + '角' + digits[fen] + '分'

    return ('负' if neg else '') + result


def _random_workday(base_date, min_d, max_d, direction='forward'):
    """从 base_date 前/后随机偏移 min_d~max_d 个自然日，结果调整到工作日（跳过周六日）"""
    offset = random.randint(min_d, max_d)
    sign = 1 if direction == 'forward' else -1
    target = base_date + timedelta(days=sign * offset)
    while target.weekday() >= 5:   # 5=周六, 6=周日
        target += timedelta(days=sign)
    return target


def fetch_first_workday_usd_rate(year, month):
    """
    从国家外汇管理局(SAFE)网站获取指定月第一个工作日的 USD 中间价。
    返回 (date_str, rate_float)，如 ('2026-05-06', 6.8562)
    100USD = X人民币，除以100得1USD汇率。
    """
    start = f'{year:04d}-{month:02d}-01'
    end   = f'{year:04d}-{month:02d}-10'
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    payload = urllib.parse.urlencode({
        'startDate': start, 'endDate': end, 'queryYN': 'true'
    }).encode()
    req = urllib.request.Request(
        'https://www.safe.gov.cn/AppStructured/hlw/RMBQuery.do',
        data=payload, method='POST'
    )
    with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
        html = resp.read().decode('utf-8', errors='ignore')
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.S)
    data_rows = []
    for row in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.S)
        cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        if len(cells) >= 3 and re.match(r'\d{4}-\d{2}-\d{2}', cells[0]):
            data_rows.append(cells)
    if not data_rows:
        raise ValueError(f'SAFE 汇率查询无结果 ({start} ~ {end})')
    first = data_rows[-1]   # 结果倒序排列，最后一条 = 最早日期 = 第一个工作日
    return first[0], round(float(first[2]) / 100, 6)


def _next_pi_no():
    """返回当天下一个 PI 编号，格式 PI{MMDD}{NNN}，如 PI0507001。"""
    import json
    now   = datetime.now()
    today = now.strftime('%Y-%m-%d')
    mmdd  = now.strftime('%m%d')

    counter = {}
    if os.path.exists(_PI_COUNTER):
        try:
            with open(_PI_COUNTER, 'r', encoding='utf-8') as f:
                counter = json.load(f)
        except Exception:
            pass

    n = counter.get(today, 0) + 1
    counter[today] = n

    try:
        with open(_PI_COUNTER, 'w', encoding='utf-8') as f:
            json.dump(counter, f)
    except Exception:
        pass

    return f'PI{mmdd}{n:03d}'


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. Proforma Invoice
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 模板：
#   行 4  : F=PI No. → invoice_no
#   行 5  : F=PO No. → invoice_no
#   行 7  : F=Date
#   行 12 : 表头（Item No. / Description / Quantity / Unit / Unit Price / Total）
#   行 13+: 商品行（模板有 2 行占位）
#   Total 行：C=Total  D=SUM(C列qty)  F=SUM(total)
#   TOTAL AMOUNT 行：A="TOTAL AMOUNT: SAY US DOLLARS XXXXX ONLY."

def generate_invoice(items, invoice_no, output_path, date_str=None, pi_no=None):
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(TPL_INVOICE)
    ws = wb.active

    now = datetime.now()
    if not date_str:
        date_str = now.strftime('%d %b., %Y').upper()

    # 与报关单保持一致，使用合并后的数据
    merged = merge_for_customs(items)
    n = len(merged)

    # ── 表头 ──────────────────────────────────────────────────
    ws['F4'] = invoice_no              # PI No. → 取插件发票号
    ws['F5'] = invoice_no              # PO No. → 取插件发票号
    ws['F7'] = date_str                 # 当天日期

    # ── 保存第一数据行（row 13）的样式和行高 ─────────────────
    start = 13

    # ── 保存表头区域（行 1 ~ start-1）样式，防止 insert_rows 污染 ─
    hdr_styles = {}
    for r in range(1, start):
        for cell in ws[r]:
            hdr_styles[(r, cell.column)] = {
                'font':          _copy(cell.font),
                'border':        _copy(cell.border),
                'fill':          _copy(cell.fill),
                'alignment':     _copy(cell.alignment),
                'number_format': cell.number_format,
            }
    hdr_heights = {r: ws.row_dimensions[r].height for r in range(1, start)}

    tpl_height = ws.row_dimensions[start].height or 57.5
    tpl_styles = []
    for cell in ws[start]:
        tpl_styles.append({
            'font':          _copy(cell.font),
            'border':        _copy(cell.border),
            'fill':          _copy(cell.fill),
            'alignment':     _copy(cell.alignment),
            'number_format': cell.number_format,
        })

    # ── 插入前：保存所有合并区域 + start 以下的自定义行高 ────────
    # （insert_rows 不会自动移合并区域，且可能导致双重偏移；行高也不自动移）
    from openpyxl.utils import get_column_letter

    # 只保存/取消 start 行及以下的合并区域；
    # start 以上的行（公司名/标题等）insert_rows 不影响，完全不动
    saved_merges = [
        (mc.min_row, mc.max_row, mc.min_col, mc.max_col)
        for mc in list(ws.merged_cells.ranges)
        if mc.min_row >= start
    ]
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start:
            ws.unmerge_cells(str(mc))

    saved_heights = {
        r: rd.height
        for r, rd in ws.row_dimensions.items()
        if r >= start and rd.height
    }

    # ── 先插入 n 行，再删除 2 个占位行，净偏移 = n-2 ────────────
    ws.insert_rows(start, n)
    ws.delete_rows(start + n, 2)
    net = n - 2   # footer 行（原 start+2 以下）的最终净偏移量

    # ── 还原行高（占位行 start/start+1 已删，跳过；footer 偏移 net） ─
    for old_r in sorted(saved_heights.keys(), reverse=True):
        if old_r < start + 2:        # 原占位行已删，跳过
            continue
        ws.row_dimensions[old_r + net].height = saved_heights[old_r]

    # ── 还原合并区域（同上逻辑，delete_rows 不会自动移合并区域） ──
    for min_r, max_r, min_c, max_c in saved_merges:
        if start <= min_r < start + 2:   # 原占位行，跳过
            continue
        if min_r >= start + 2:           # footer 行，偏移 net
            min_r += net
            max_r += net
        # start 之前的行：不偏移
        ws.merge_cells(
            f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        )

    for idx, item in enumerate(merged):
        r = start + idx
        ws.row_dimensions[r].height = tpl_height
        for col_idx, s in enumerate(tpl_styles, 1):
            c = ws.cell(r, col_idx)
            c.font          = _copy(s['font'])
            c.border        = _copy(s['border'])
            c.fill          = _copy(s['fill'])
            c.alignment     = _copy(s['alignment'])
            c.number_format = s['number_format']

        ws.cell(r, 1).value = item['hs_code']       # Item No.       = 海关编码
        ws.cell(r, 2).value = item['customs_name']  # Description    = 商品名称
        ws.cell(r, 3).value = item['qty']            # C列 Quantity   = 数量（原 Picture 位置）
        ws.cell(r, 4).value = item['customs_unit']   # D列 Unit       = 报关产品单位（跟随 Unit，来自基础资料 G 列）
        ws.cell(r, 5).value = item['usd_price']      # E列 Unit Price = 单价
        ws.cell(r, 5).number_format = '0.0000'       # 保留 4 位小数
        ws.cell(r, 6).value = item['total_usd']       # F列 Total      = 总价（直接写入，与报关单一致）

    # ── Total 行 + TOTAL AMOUNT 行（此时在 start+n） ────────
    tr = start + n

    ws.cell(tr, 1).value = 'Total'
    ws.cell(tr, 3).value = f'=SUM(C{start}:C{tr - 1})'   # C列 = 合计数量（Quantity 之和）
    ws.cell(tr, 6).value = f'=SUM(F{start}:F{tr - 1})'

    # TOTAL AMOUNT: 英文金额书写
    total_usd_value = sum(item['total_usd'] for item in merged)
    amount_eng = _number_to_english(round(total_usd_value, 2))
    total_amount_text = f'TOTAL AMOUNT: SAY US DOLLARS {amount_eng} ONLY.'
    ws.cell(tr + 1, 1).value = total_amount_text

    # 自适应行高：根据文本长度估算所需行数
    # 合并区域 A:F 列宽合计 ≈ 115 个字符宽度单位
    _merged_width = sum(
        ws.column_dimensions[get_column_letter(c)].width or 10
        for c in range(1, 7)
    )
    # 估算文本显示宽度（ASCII=1, CJK=2）
    _text_width = sum(
        2 if ('\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef')
        else 1
        for ch in total_amount_text
    )
    _font_size = ws.cell(tr + 1, 1).font.size or 10
    _chars_per_line = int(_merged_width * (_font_size / 10) * 0.95)  # 经验系数
    _num_lines = max(1, (_text_width + _chars_per_line - 1) // _chars_per_line)
    _line_height = _font_size * 1.5  # 每行高度(pt)
    ws.row_dimensions[tr + 1].height = max(24.5, _num_lines * _line_height + 6)  # +6为上下边距

    # ── 还原表头区域样式（insert_rows/delete_rows 可能污染表头）──
    for (r, col), s in hdr_styles.items():
        c = ws.cell(r, col)
        c.font          = _copy(s['font'])
        c.border        = _copy(s['border'])
        c.fill          = _copy(s['fill'])
        c.alignment     = _copy(s['alignment'])
        c.number_format = s['number_format']
    for r, h in hdr_heights.items():
        if h:
            ws.row_dimensions[r].height = h

    # ── 更新表头：C列 Picture→Quantity, D列 Quantity→Unit ────────
    ws.cell(12, 3).value = 'Quantity'
    ws.cell(12, 4).value = 'Unit'

    out = os.path.join(output_path, f'Proforma_Invoice_{invoice_no}.xlsx')
    wb.save(out)
    return out


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5. 装箱单 Packing List
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# .xlsx 模板结构（27行 10列，行1为新增空行）：
#   行 1    : （新增行）
#   行 2    : 公司名 (A2:J2 合并)
#   行 3    : PACKING LIST (A3:J3 合并)
#   行 4-5  : From 区 (A4:E5) / Invoice NO. (F4:G4) / 值 (H4:J4) ← PI号
#             S/C NO. (F5:G5) / 值 (H5:J5) ← 日期
#   行 6-7  : To 区 (A6:E7) / L/C (F6:J6) / Vessel (F7:J7)
#   行 8    : Port of Loading (A8:B8) / Discharge (C8:E8) / Destination (F8:J8)
#   行 9    : 列头，Dimension 占 G9:I9 合并 (= L×W×H 三列), J=CBM
#   行 10-11: 占位商品行（无合并）
#   行 12   : TOTAL (A12:B12) / qty / GW / NW / (G12:I12) / CBM
#   行 13   : Total Quantity (A13:J13 合并)
#   行 14   : Remark (A14:E14) / SIGNATURE & DATE (F14:J22 合并，跨9行)
#   行 15-22: Declaration (A15:E22 合并，跨8行)
# 列映射：col3=Quantity, col4=Package(件数/boxes), col5=G.W.(净重),
#          col6=N.W.(毛重), col7=L, col8=W, col9=H, col10=CBM

def generate_packing(items, invoice_no, output_path, pi_no='', date_str=None):
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.cell.cell import MergedCell

    now = datetime.now()
    if not date_str:
        date_str = now.strftime('%d %b., %Y').upper()

    # 与报关单、发票保持一致，使用合并后的数据
    merged = merge_for_customs(items)
    n      = len(merged)

    # ⚠️ TPL_PACKING 须为 .xlsx（用 Excel/WPS 将原 .xls 另存一次）
    wb = openpyxl.load_workbook(TPL_PACKING)
    ws = wb.active

    # ── H4 = 合同编号(发票号)\n日期（wrap），H5 = 日期 ──────────────
    c = ws.cell(4, 8)
    c.value     = f'{invoice_no}\n{date_str}'
    c.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
    ws.cell(5, 8).value = date_str

    # ── 保存第一占位行（row 10）的样式，作为所有产品行的模板 ───────
    start = 10

    # ── 保存表头区域（行 1 ~ start-1）样式，防止 insert_rows 污染 ─
    # 跳过合并区域的从属格（MergedCell），只保存主格的样式
    hdr_styles = {}
    for r in range(1, start):
        for cell in ws[r]:
            if isinstance(cell, MergedCell):
                continue
            hdr_styles[(r, cell.column)] = {
                'font':          _copy(cell.font),
                'border':        _copy(cell.border),
                'fill':          _copy(cell.fill),
                'alignment':     _copy(cell.alignment),
                'number_format': cell.number_format,
            }
    hdr_heights = {r: ws.row_dimensions[r].height for r in range(1, start)}

    tpl_height = ws.row_dimensions[start].height or 15
    tpl_styles = []
    for cell in ws[start]:
        tpl_styles.append({
            'font':          _copy(cell.font),
            'border':        _copy(cell.border),
            'fill':          _copy(cell.fill),
            'alignment':     _copy(cell.alignment),
            'number_format': cell.number_format,
        })

    # 只保存/取消 start 行及以下的合并区域；
    # start 以上（行 1-9，公司名/PACKING LIST 标题等）完全不动
    saved_merges = [
        (mc.min_row, mc.max_row, mc.min_col, mc.max_col)
        for mc in list(ws.merged_cells.ranges)
        if mc.min_row >= start
    ]
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start:
            ws.unmerge_cells(str(mc))

    saved_heights = {r: rd.height for r, rd in ws.row_dimensions.items() if rd.height}

    # ── 插入 n 行，再删除原来的 2 个占位行 ───────────────────────
    # 净偏移：原行 11+（= start+2+）最终位移 (n-2) 行
    ws.insert_rows(start, n)
    ws.delete_rows(start + n, 2)
    net = n - 2

    # ── 还原行高 ──────────────────────────────────────────────────
    for old_r, h in sorted(saved_heights.items(), reverse=True):
        if old_r < start:              # 行 1-9：原位不动
            ws.row_dimensions[old_r].height = h
        elif old_r < start + 2:        # 行 10-11（占位行，已删）：跳过
            continue
        else:                          # 行 12+：偏移 net
            ws.row_dimensions[old_r + net].height = h

    # ── 还原合并区域 ──────────────────────────────────────────────
    for min_r, max_r, min_c, max_c in saved_merges:
        if start <= min_r < start + 2:   # 行 10-11：已删，跳过
            continue
        if min_r >= start + 2:           # 行 12+：偏移 net
            min_r += net
            max_r += net
        # 行 1-9：不偏移
        ws.merge_cells(
            f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        )

    # ── 写入产品行（样式 + 数据） ────────────────────────────────
    tot_qty = tot_boxes = 0
    tot_gw = tot_nw = tot_cbm = 0.0

    for i, item in enumerate(merged):
        r = start + i
        ws.row_dimensions[r].height = tpl_height
        for col_idx, s in enumerate(tpl_styles, 1):
            c = ws.cell(r, col_idx)
            c.font          = _copy(s['font'])
            c.border        = _copy(s['border'])
            c.fill          = _copy(s['fill'])
            c.alignment     = _copy(s['alignment'])
            c.number_format = s['number_format']

        # qty 现在是 PCS；pkg_count 是原始包数（装箱/CBM用）
        pcs_qty = item['qty']
        pkg_cnt = item.get('pkg_count', pcs_qty)  # 原始包数
        boxes   = item.get('boxes', 0)
        gw_val  = item.get('gross_weight', 0)   # 合并后总毛重
        nw_val  = item.get('net_weight', 0)     # 合并后总净重

        # G/H/I：优先用 api_dimensions，其次基础资料 dimensions
        dims = item.get('api_dimensions') or item.get('dimensions')
        cbm  = item.get('cbm_val') or _calc_cbm(dims, pkg_cnt)  # 优先用比例分配值
        if dims:
            dim_l = dims.get('l', 0) or ''
            dim_w = dims.get('w', 0) or ''
            dim_h = dims.get('h', 0) or ''
            dim_l = dim_l if dim_l else ''
            dim_w = dim_w if dim_w else ''
            dim_h = dim_h if dim_h else ''
        else:
            dim_l = dim_w = dim_h = ''

        # B列 = 英文报关品名 + 中文报关商品名称
        en_name = item.get('en_customs_name') or item.get('en_category', '')
        ws.cell(r,  1).value = ''
        ws.cell(r,  2).value = f"{en_name}\n{item['customs_name']}"
        ws.cell(r,  3).value = pcs_qty   # PCS 数量
        ws.cell(r,  4).value = boxes
        ws.cell(r,  5).value = gw_val
        ws.cell(r,  6).value = nw_val
        ws.cell(r,  7).value = dim_l
        ws.cell(r,  8).value = dim_w
        ws.cell(r,  9).value = dim_h
        ws.cell(r, 10).value = round(cbm, 4)

        tot_qty   += pcs_qty
        tot_boxes += boxes
        tot_gw    += gw_val
        tot_nw    += nw_val
        tot_cbm   += cbm

    # ── TOTAL 行（在 start+n） ────────────────────────────────────
    tr = start + n
    ws.cell(tr,  3).value = tot_qty
    ws.cell(tr,  4).value = tot_boxes
    ws.cell(tr,  5).value = round(tot_gw, 2)
    ws.cell(tr,  6).value = round(tot_nw, 2)
    ws.cell(tr, 10).value = round(tot_cbm, 4)

    # ── 还原表头区域样式（insert_rows/delete_rows 可能污染表头）──
    for (r, col), s in hdr_styles.items():
        c = ws.cell(r, col)
        if isinstance(c, MergedCell):
            continue   # 从属格不可写样式，跳过
        c.font          = _copy(s['font'])
        c.border        = _copy(s['border'])
        c.fill          = _copy(s['fill'])
        c.alignment     = _copy(s['alignment'])
        c.number_format = s['number_format']
    for r, h in hdr_heights.items():
        if h:
            ws.row_dimensions[r].height = h

    out = os.path.join(output_path, f'Packing_List_{invoice_no}.xlsx')
    wb.save(out)
    return out


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 6. 统一入口
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_merge_detail(items, invoice_no, output_path, supplier_map=None, exchange_rate=0):
    """
    生成合并明细 Excel：展示相同海关编码商品的合并情况
    Sheet1: 合并汇总 — 每个合并组一行
    Sheet2: 合并明细 — 每个商品一行（标注所属合并组）
    Sheet3: 数量对比 — 合并前商品数 vs 合并后行数
    Sheet4: 发票核对 — 按项号展示报关单合计 vs 各供应商分项金额
    """
    from openpyxl.styles import Font as XlFont, Alignment, Border, Side, PatternFill

    found = [i for i in items if i.get('found')]

    # 使用 merge_for_customs 结果，与报关单顺序一致（item_no 对应报关单项号）
    merged_list = merge_for_customs(found)

    # 同时保留原始分组，供 Sheet2 per-item 明细使用
    groups = {}
    for item in found:
        key = (item['hs_code'], item['customs_name'], item['customs_unit'],
               item['origin'], item['material'])
        groups.setdefault(key, []).append(item)

    def _dedup_join(values):
        seen, result = set(), []
        for v in values:
            for part in (v or '').split('、'):
                part = part.strip()
                if part and part not in seen:
                    seen.add(part)
                    result.append(part)
        return '、'.join(result)

    wb = openpyxl.Workbook()

    # ── 样式 ──
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_font = XlFont(bold=True, size=11, color='FFFFFF')
    multi_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    single_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    border = Border(left=Side('thin'), right=Side('thin'),
                    top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center')
    wrap = Alignment(wrap_text=True, vertical='center')

    # ━━━ Sheet1: 合并汇总 ━━━
    ws1 = wb.active
    ws1.title = '合并汇总'
    h1 = ['组号', '海关编码', '报关商品名称', '单位', '境内货源地', '材料',
          '用途(合并后)', '规格型号(合并后)', '商品数量', '是否合并',
          '合并前总数量(PCS)', '合并后总数量(PCS)',
          '包含商品编号(前10个)']
    for c, h in enumerate(h1, 1):
        cell = ws1.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border; cell.alignment = center

    for m in merged_list:
        key = (m['hs_code'], m['customs_name'], m['customs_unit'], m['origin'],
               m.get('material', ''))
        grp   = groups.get(key, [])
        count = len(grp)
        is_merged  = '是' if count > 1 else '否'
        merged_usage = _dedup_join(i['usage'] for i in grp)
        spec_desc    = m.get('spec_desc', '')
        before_qty   = sum(i['qty'] * i.get('pcs_per_pkg', 1) for i in grp)
        after_qty    = m['qty']  # PCS
        codes_preview = ', '.join(i['code'] for i in grp[:10])
        if count > 10:
            codes_preview += f' ... 等{count}个'

        item_no = m['item_no']
        vals = [item_no, m['hs_code'], m['customs_name'], m['customs_unit'],
                m['origin'], m.get('material', ''),
                merged_usage, spec_desc, count, is_merged,
                before_qty, after_qty, codes_preview]
        fill = multi_fill if count > 1 else single_fill
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(item_no + 1, c, v)
            cell.border = border; cell.alignment = wrap; cell.fill = fill

    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 8
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 20
    ws1.column_dimensions['H'].width = 30
    ws1.column_dimensions['I'].width = 10
    ws1.column_dimensions['J'].width = 8
    ws1.column_dimensions['K'].width = 16
    ws1.column_dimensions['L'].width = 16
    ws1.column_dimensions['M'].width = 40

    # ━━━ Sheet2: 合并明细 ━━━
    ws2 = wb.create_sheet('合并明细')
    h2 = ['组号', '海关编码', '报关商品名称', '材料', '商品编号',
          '中文名称', '商品类别', '用途', '换算系数', '数量(PCS)',
          '单价(USD)', '总金额(USD)', '毛重(KGS)', '净重(KGS)']
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border; cell.alignment = center

    r = 2
    first_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    for m in merged_list:
        key = (m['hs_code'], m['customs_name'], m['customs_unit'], m['origin'],
               m.get('material', ''))
        grp = groups.get(key, [])
        for j, item in enumerate(grp):
            vals = [m['item_no'], m['hs_code'], m['customs_name'],
                    m.get('material', ''), item['code'],
                    item.get('cn_name', ''), item.get('en_category', ''),
                    item['usage'], item.get('pcs_per_pkg', 1),
                    item['qty'] * item.get('pcs_per_pkg', 1),  # PCS
                    item.get('usd_price', 0),
                    item.get('total_usd', 0),
                    item.get('gw_val', 0), item.get('nw_val', 0)]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(r, c, v)
                cell.border = border; cell.alignment = wrap
                if j == 0 and len(grp) > 1:
                    cell.fill = first_fill
            r += 1

    # 未找到的商品
    not_found = [i for i in items if not i.get('found')]
    if not_found:
        r += 1
        ws2.cell(r, 1, '未在基础资料中找到的商品：').font = XlFont(bold=True, color='FF0000')
        r += 1
        nohs_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        for item in not_found:
            vals = ['', '', '', '', item['code'], '', '', '', '',
                    item.get('qty', 0), '', '', '', '']
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(r, c, v)
                cell.border = border; cell.fill = nohs_fill
            r += 1

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 18
    ws2.column_dimensions['H'].width = 14
    ws2.column_dimensions['I'].width = 10
    ws2.column_dimensions['J'].width = 12
    ws2.column_dimensions['K'].width = 12
    ws2.column_dimensions['L'].width = 14
    ws2.column_dimensions['M'].width = 12
    ws2.column_dimensions['N'].width = 12

    # ━━━ Sheet3: 数量对比 ━━━
    ws3 = wb.create_sheet('数量对比')

    total_items = len(found)
    total_groups = len(merged_list)
    multi_groups = sum(1 for g in groups.values() if len(g) > 1)
    single_groups = total_groups - multi_groups
    multi_items = sum(len(g) for g in groups.values() if len(g) > 1)
    single_items = total_items - multi_items

    stats = [
        ['统计项', '数量'],
        ['调拨单商品总数', total_items + len(not_found)],
        ['未在基础资料中找到', len(not_found)],
        ['有效商品数(在基础资料中找到)', total_items],
        ['', ''],
        ['合并后报关单行数', total_groups],
        ['需合并组数(含2个以上商品)', multi_groups],
        ['不需合并组数(仅1个商品)', single_groups],
        ['', ''],
        ['需合并的商品数', multi_items],
        ['不需合并的商品数', single_items],
        ['', ''],
        ['合并前行数(=有效商品数)', total_items],
        ['合并后行数(=合并组数)', total_groups],
        ['减少行数', total_items - total_groups],
    ]
    if total_items > 0:
        stats.append(['合并率', f'{(total_items - total_groups) / total_items * 100:.1f}%'])

    for ri, row_data in enumerate(stats, 1):
        for c, v in enumerate(row_data, 1):
            cell = ws3.cell(ri, c, v)
            cell.border = border
            if ri == 1:
                cell.font = hdr_font; cell.fill = hdr_fill
            elif isinstance(v, str) and v.endswith('%'):
                cell.font = XlFont(bold=True, color='FF0000', size=12)

    # 按海关编码维度
    hs_map = {}
    for key, grp in groups.items():
        hs_code = key[0]
        hs_map.setdefault(hs_code, []).extend(grp)

    start_row = len(stats) + 3
    ws3.cell(start_row, 1, '按海关编码维度统计').font = XlFont(bold=True, size=12)
    start_row += 1
    h3b = ['海关编码', '商品数量', '合并组数', '减少行数', '主要报关名称']
    for c, h in enumerate(h3b, 1):
        cell = ws3.cell(start_row, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border; cell.alignment = center

    sorted_hs = sorted(hs_map.items(), key=lambda x: len(x[1]), reverse=True)
    for i, (hs, hs_items) in enumerate(sorted_hs):
        r = start_row + 1 + i
        sub_groups = defaultdict(list)
        for item in hs_items:
            key = (item['customs_name'], item['customs_unit'], item['origin'], item['material'])
            sub_groups[key].append(item)
        name_counts = defaultdict(int)
        for item in hs_items:
            name_counts[item['customs_name']] += 1
        main_name = max(name_counts, key=name_counts.get) if name_counts else ''
        vals = [hs, len(hs_items), len(sub_groups), len(hs_items) - len(sub_groups), main_name]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(r, c, v)
            cell.border = border; cell.alignment = wrap

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 30

    # ━━━ Sheet4: 发票核对 ━━━
    ws4 = wb.create_sheet('发票核对')
    sup_map = supplier_map or {}

    item_fill   = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # 黄：报关合计行
    sup_fill    = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')   # 灰：供应商明细行
    ok_fill     = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')   # 绿：核对通过
    err_fill    = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')   # 红：核对差异
    bill_fill   = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')   # 橙：开票建议列
    bold_font   = XlFont(bold=True, size=11)
    red_font    = XlFont(bold=True, color='C0392B')
    green_font  = XlFont(bold=True, color='27AE60')

    # 列：项号/报关名/供应商名/供应商编码/商品编码/数量(PCS)/单价(USD)/金额(USD)/发票单价(RMB)/发票金额(RMB)/核对(USD)/核对(采购RMB)/采购合同(RMB)/开票数量/开票单价(RMB)/开票金额(RMB)
    h4 = ['项号', '报关商品名称', '供应商名称', '供应商编码', '商品编码',
          '数量(PCS)', '单价(USD)', '金额(USD)', '发票单价(RMB)', '发票金额(RMB)',
          '核对(USD)', '核对(采购RMB)', '采购合同(RMB)',
          '开票数量', '开票单价(RMB)', '开票金额(RMB)']
    CTR = {1, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}   # 居中列索引
    for c, h in enumerate(h4, 1):
        cell = ws4.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border; cell.alignment = center

    r4 = 2
    for m in merged_list:
        key = (m['hs_code'], m['customs_name'], m['customs_unit'], m['origin'], m.get('material', ''))
        grp = groups.get(key, [])
        item_no   = m['item_no']
        cust_name = m['customs_name']
        cust_qty  = m['qty']                   # PCS
        cust_usd  = round(m['total_usd'], 2)
        # 报关单项的 RMB 合计 = 各商品 qty × rmb_price（与采购合同一致）
        cust_rmb  = round(sum(i['qty'] * i.get('rmb_price', 0) for i in grp), 2)

        # 报关单合计行（黄色）
        rate = float(exchange_rate) if exchange_rate else 0
        avg_usd = round(cust_usd / cust_qty, 4) if cust_qty else ''
        avg_rmb = round(avg_usd * rate, 4) if (avg_usd != '' and rate) else ''
        inv_rmb_total = round(cust_usd * rate, 2) if rate else ''
        row_vals = [item_no, cust_name, '【报关单合计】', '', '', cust_qty,
                    avg_usd, cust_usd, avg_rmb, inv_rmb_total, '', '', cust_rmb,
                    cust_qty, avg_rmb, inv_rmb_total]   # N/O/P 开票汇总
        for c, v in enumerate(row_vals, 1):
            cell = ws4.cell(r4, c, v)
            cell.fill = bill_fill if c >= 14 else item_fill   # N/O/P 橙色
            cell.border = border; cell.font = bold_font
            cell.alignment = center if c in CTR else wrap
        r4 += 1

        # 按供应商汇总：{sup_name: {sup_num, qty, usd, rmb, codes}}
        sup_agg = {}
        for item in grp:
            code     = item['code']
            info     = sup_map.get(code, {})
            sup_name = (info.get('supplierName') or '未知供应商').strip()
            sup_num  = (info.get('supplierNumber') or '').strip()
            pcs_per  = item.get('pcs_per_pkg', 1) or 1
            qty_pcs  = round(item['qty'] * pcs_per, 2)
            usd      = round(item.get('total_usd', 0), 2)
            rmb      = round(item['qty'] * item.get('rmb_price', 0), 2)   # 采购合同同口径
            if sup_name not in sup_agg:
                sup_agg[sup_name] = {'sup_num': sup_num, 'qty': 0.0, 'usd': 0.0, 'rmb': 0.0, 'codes': []}
            sup_agg[sup_name]['qty'] += qty_pcs
            sup_agg[sup_name]['usd'] += usd
            sup_agg[sup_name]['rmb'] += rmb
            sup_agg[sup_name]['codes'].append(code)

        # 第一步：算出每个供应商的 raw bill_qty，再统一修正合计误差
        sup_bill = {}   # {sup_name: {'qty': int, 'raw': float, 'rmb': float}}
        for sup_name, sd in sup_agg.items():
            susd = round(sd['usd'], 2)
            bill_rmb = round(susd * rate, 2) if rate else susd
            if avg_usd and susd:
                raw = susd / avg_usd
                qty = max(1, round(raw))
            else:
                raw = round(sd['qty'], 2)
                qty = max(1, int(round(raw)))
            sup_bill[sup_name] = {'qty': qty, 'raw': raw, 'rmb': bill_rmb}

        # 修正：若合计不等于 cust_qty，调整误差最大那个供应商
        bill_total = sum(v['qty'] for v in sup_bill.values())
        diff = bill_total - int(cust_qty)
        if diff != 0 and sup_bill:
            worst = max(sup_bill, key=lambda s: abs(sup_bill[s]['raw'] - round(sup_bill[s]['raw'])))
            sup_bill[worst]['qty'] = max(1, sup_bill[worst]['qty'] - diff)

        for sup_name, sd in sup_agg.items():
            sq   = round(sd['qty'], 2)
            susd = round(sd['usd'], 2)
            srmb = round(sd['rmb'], 2)
            usd_unit     = round(susd / sq, 4) if sq else 0
            inv_rmb_unit = round(usd_unit * rate, 4) if rate else ''
            inv_rmb_sup  = round(susd * rate, 2) if rate else ''
            bill_qty  = sup_bill[sup_name]['qty']
            bill_rmb  = sup_bill[sup_name]['rmb']
            bill_unit = round(bill_rmb / bill_qty, 4) if bill_qty else ''
            codes_str = ', '.join(sd['codes'])
            row_vals = [item_no, cust_name, sup_name, sd['sup_num'],
                        codes_str, sq, usd_unit, susd, inv_rmb_unit, inv_rmb_sup, '', '', srmb,
                        bill_qty, bill_unit, bill_rmb]   # N/O/P 开票建议
            for c, v in enumerate(row_vals, 1):
                cell = ws4.cell(r4, c, v)
                cell.fill = bill_fill if c >= 14 else sup_fill   # N/O/P 橙色
                cell.border = border
                cell.alignment = center if c in CTR else wrap
            r4 += 1

        # 核对行
        sup_usd_total = round(sum(sd['usd'] for sd in sup_agg.values()), 2)
        sup_rmb_total = round(sum(sd['rmb'] for sd in sup_agg.values()), 2)  # 采购合同RMB合计
        sup_inv_rmb   = round(sup_usd_total * rate, 2) if rate else ''
        diff_usd = round(cust_usd - sup_usd_total, 2)
        diff_rmb = round(cust_rmb - sup_rmb_total, 2)

        chk_usd = '✅' if abs(diff_usd) < 0.02 else f'❌ {diff_usd:+.2f}'
        chk_rmb = '✅' if abs(diff_rmb) < 0.02 else f'❌ {diff_rmb:+.2f}'
        chk_fill = ok_fill if (abs(diff_usd) < 0.02 and abs(diff_rmb) < 0.02) else err_fill
        chk_font = green_font if chk_fill == ok_fill else red_font

        row_vals = [item_no, cust_name, '供应商合计', '', '', '', '', sup_usd_total,
                    '', sup_inv_rmb, chk_usd, chk_rmb, sup_rmb_total, '', '', '']
        for c, v in enumerate(row_vals, 1):
            cell = ws4.cell(r4, c, v)
            cell.fill = chk_fill; cell.border = border; cell.font = chk_font
            cell.alignment = center if c in CTR else wrap
        r4 += 2   # +1 空行

    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 14
    ws4.column_dimensions['E'].width = 30
    ws4.column_dimensions['F'].width = 12
    ws4.column_dimensions['G'].width = 12
    ws4.column_dimensions['H'].width = 14
    ws4.column_dimensions['I'].width = 14
    ws4.column_dimensions['J'].width = 14
    ws4.column_dimensions['K'].width = 12
    ws4.column_dimensions['L'].width = 14
    ws4.column_dimensions['M'].width = 14
    ws4.column_dimensions['N'].width = 12
    ws4.column_dimensions['O'].width = 14
    ws4.column_dimensions['P'].width = 14

    out = os.path.join(output_path, f'合并明细_{invoice_no}.xlsx')
    wb.save(out)
    return out


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 阶段一：销售合同 / 采购合同 / 存档JSON
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_items_json(invoice_no, processed, exchange_rate, output_path, order_no='', supplier_map=None):
    """存档本批次商品数据，供阶段二使用。"""
    items_list = []
    for item in processed:
        if not item.get('found'):
            continue
        pcs_per_pkg = item.get('pcs_per_pkg', 1) or 1
        qty_original = item.get('qty', 0)                  # 原始包数（process_items 保留的）
        qty_pcs = round(qty_original * pcs_per_pkg, 4)     # PCS 数量（发票/凌航用）
        items_list.append({
            'code':         item['code'],
            'cn_name':      item.get('cn_name', ''),
            'hs_code':      item.get('hs_code', ''),
            'customs_name': item.get('customs_name', ''),
            'unit':         item.get('unit', ''),
            'qty':          qty_original,      # 原始箱/包数
            'qty_pcs':      qty_pcs,           # PCS 数量（发票/凌航用）
            'qty_original': qty_original,      # 兼容字段
            'pcs_per_pkg':  pcs_per_pkg,
            'rmb_price':    item.get('rmb_price', 0),
            'usd_price':    item.get('usd_price', 0),
            'total_usd':    item.get('total_usd', 0),
            'tax_code':     item.get('tax_code', ''),
            'customs_unit': item.get('customs_unit', ''),
            'origin':       item.get('origin', ''),
            'material':     item.get('material', ''),
            'gw_val':       item.get('gw_val', 0),    # 毛重(KGS)，供阶段二对比
            'project_name': item.get('project_name', ''),  # X列 项目名称（凌航发票用）
        })
    data = {
        'invoice_no':    invoice_no,
        'order_no':      order_no,
        'output_path':   output_path,   # 报关单子文件夹，Stage 2 输出用
        'exchange_rate': float(exchange_rate),
        'generated_at':  datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'items':         items_list,
        'supplier_map':  supplier_map or {},   # 供阶段二直接使用，避免重复调 API
    }
    os.makedirs(_ITEMS_STORE, exist_ok=True)
    file_key = order_no or invoice_no or 'items'
    path = os.path.join(_ITEMS_STORE, f'{file_key}_items.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def _apply_row_style(ws, r, tpl_styles, tpl_height):
    """将模板行样式应用到指定行（跳过 MergedCell）。"""
    from copy import copy as _copy
    ws.row_dimensions[r].height = tpl_height
    for col_idx, s in tpl_styles.items():
        c = ws.cell(r, col_idx)
        if hasattr(c, 'column'):  # 非 MergedCell
            c.font        = _copy(s['font'])
            c.border      = _copy(s['border'])
            c.fill        = _copy(s['fill'])
            c.alignment   = _copy(s['alignment'])
            c.number_format = s['number_format']


def _save_row_style(ws, r):
    """保存一行的样式（跳过 MergedCell），返回 {col_idx: style_dict}。"""
    from copy import copy as _copy
    from openpyxl.cell.cell import MergedCell
    styles = {}
    for cell in ws[r]:
        if isinstance(cell, MergedCell):
            continue
        styles[cell.column] = {
            'font':          _copy(cell.font),
            'border':        _copy(cell.border),
            'fill':          _copy(cell.fill),
            'alignment':     _copy(cell.alignment),
            'number_format': cell.number_format,
        }
    return styles


def _insert_rows_in_template(ws, start, n_items, n_tpl):
    """
    通用的合同模板行插入辅助：
    - 保存 start 行样式作为数据行模板
    - 取消 start 行以下的所有合并
    - insert_rows(start, n_items), delete_rows(start+n_items, n_tpl)
    - 返回 (tpl_styles, tpl_height, saved_merges, net)
    """
    from openpyxl.utils import get_column_letter

    tpl_height = ws.row_dimensions[start].height or 25.0
    tpl_styles = _save_row_style(ws, start)

    # 保存 ≥ start 的合并区域
    saved_merges = [
        (mc.min_row, mc.max_row, mc.min_col, mc.max_col)
        for mc in list(ws.merged_cells.ranges)
        if mc.min_row >= start
    ]
    # 取消这些合并
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start:
            ws.unmerge_cells(str(mc))

    # 保存行高
    saved_heights = {}
    for r_idx, rd in ws.row_dimensions.items():
        if r_idx >= start and rd.height:
            saved_heights[r_idx] = rd.height

    ws.insert_rows(start, n_items)
    ws.delete_rows(start + n_items, n_tpl)
    net = n_items - n_tpl

    # 还原行高（footer 行）
    for old_r in sorted(saved_heights.keys(), reverse=True):
        if start <= old_r < start + n_tpl:
            continue  # 占位行已删除
        new_r = old_r + net
        if saved_heights[old_r]:
            ws.row_dimensions[new_r].height = saved_heights[old_r]

    # 还原合并区域（footer 行）
    from openpyxl.utils import get_column_letter as gcl
    for min_r, max_r, min_c, max_c in saved_merges:
        if start <= min_r < start + n_tpl:
            continue  # 占位行已删除
        new_min_r = min_r + net
        new_max_r = max_r + net
        ws.merge_cells(
            start_row=new_min_r, start_column=min_c,
            end_row=new_max_r,   end_column=max_c
        )

    return tpl_styles, tpl_height, net


def generate_sales_contract(processed, invoice_no, output_path, today=None):
    """
    生成 QIHANG 销售合同。
    返回 (output_path, sc_date, ship_date) — sc_date/ship_date 供采购合同使用。
    """
    if today is None:
        today = date.today()

    # 使用合并后数据（与报关单行数一致）
    items = merge_for_customs([i for i in processed if i.get('found')])

    wb = openpyxl.load_workbook(TPL_SALES)
    ws = wb['销售合同']

    # 合同编号 H13, 合同日期 H14
    sc_date     = _random_workday(today, 48, 55, 'backward')
    ship_date   = _random_workday(today, 5, 10, 'forward')
    ship_str    = ship_date.strftime('Before %b %d, %Y')  # e.g. 'Before May 28, 2026'

    ws.cell(13, 8).value = invoice_no
    ws.cell(14, 8).value = sc_date.strftime('%Y-%m-%d')

    if not items:
        out = os.path.join(output_path, f'QIHANG销售合同_{invoice_no}.xlsx')
        wb.save(out)
        return out, sc_date, ship_date

    start  = 19   # 第一数据行
    n_tpl  = 5    # 占位行 19-23
    n      = len(items)

    tpl_styles, tpl_height, net = _insert_rows_in_template(ws, start, n, n_tpl)

    for idx, item in enumerate(items):
        r = start + idx
        _apply_row_style(ws, r, tpl_styles, tpl_height)
        en_name = item.get('en_customs_name') or item.get('en_category', '')
        ws.cell(r, 3).value = idx + 1                                       # C 序号
        ws.cell(r, 4).value = item.get('hs_code', '')                       # D HS编码
        ws.cell(r, 5).value = f"{en_name}\n{item['customs_name']}"          # E 英文+中文报关名
        ws.cell(r, 6).value = item.get('usd_price', 0)                      # F 单价
        ws.cell(r, 7).value = item.get('qty', 0)                            # G 数量（PCS）
        ws.cell(r, 8).value = item.get('customs_unit', 'PCS')               # H 单位
        ws.cell(r, 9).value = item.get('total_usd', 0)                      # I 总价（避免单价四舍五入造成误差）

    # 更新 footer 公式（旧行号 + net）
    last = start + n - 1
    total_r = 25 + net   # 合计行
    tax_r   = 27 + net   # 增值税行
    combo_r = 29 + net   # 价税合计行
    ship_r  = 33 + net   # 装船期行

    ws.cell(total_r, 8).value = f'=SUM(I{start}:I{last})'
    ws.cell(tax_r,   8).value = f'=H{total_r}*0.13'
    ws.cell(combo_r, 8).value = f'=H{total_r}+H{tax_r}'
    ws.cell(ship_r,  5).value = ship_str
    ws.cell(37 + net, 8).value = None   # 签名区日期留空，手动填写

    out = os.path.join(output_path, f'QIHANG销售合同_{invoice_no}.xlsx')
    wb.save(out)
    # 图片锚点随插入行下移（模板商品区从 start=19 行开始，净插入 net 行）
    if net != 0:
        try:
            _shift_drawing_rows(out, start, net)
        except Exception as e:
            print(f'[WARN] 销售合同图片位移失败: {e}')
    return out, sc_date, ship_date


def generate_purchase_contracts(processed, supplier_map, invoice_no, output_path,
                                 sc_date=None, ship_date=None):
    """
    按供应商生成工厂采购合同，每供应商一份。
    supplier_map: {code: {'supplierName': str, 'supplierIdx': int}}
    返回生成文件路径列表。
    """
    if sc_date is None:
        sc_date = date.today()

    items = [i for i in processed if i.get('found')]

    # 按供应商分组
    groups = {}   # supplierName -> {'items': [...], 'idx': int, 'origin_city': str}
    for item in items:
        info  = supplier_map.get(item['code'], {})
        name  = info.get('supplierName', '未知供应商') or '未知供应商'
        idx   = info.get('supplierIdx', 99)
        if name not in groups:
            groups[name] = {'items': [], 'idx': idx,
                            'origin_city': info.get('origin_city', ''),
                            'supplier_number': info.get('supplierNumber', '')}
        groups[name]['items'].append(item)

    results = []
    for supplier_name, gdata in groups.items():
        supplier_idx   = gdata['idx']
        supplier_items = gdata['items']
        n              = len(supplier_items)

        wb = openpyxl.load_workbook(TPL_PURCHASE)
        ws = wb.active

        # 合同日期 B6 / 合同编号 G6 / 供方名称 B7
        purchase_date = _random_workday(sc_date, 1, 5, 'forward')
        contract_no   = f'{invoice_no}-{supplier_idx:02d}'
        origin_city   = gdata.get('origin_city') or '义乌市'
        print(f'[DEBUG] 采购合同 invoice_no={invoice_no!r}, contract_no={contract_no!r}')
        ws.cell(6, 2).value = f'签订时间/地点: {purchase_date.strftime("%Y-%m-%d")} / {origin_city}'
        ws.cell(6, 7).value = contract_no
        ws.cell(7, 2).value = f'供方：{supplier_name}'

        start = 13    # 第一数据行
        n_tpl = 3     # 占位行 13-15

        tpl_styles, tpl_height, net = _insert_rows_in_template(ws, start, n, n_tpl)

        total_rmb = 0.0
        for idx2, item in enumerate(supplier_items):
            r = start + idx2
            _apply_row_style(ws, r, tpl_styles, tpl_height)

            dims = item.get('dimensions')
            d_str = (f"{dims.get('l',0)}×{dims.get('w',0)}×{dims.get('h',0)}cm"
                     if dims else '')
            boxes = item.get('boxes') or ''
            qty_original = item.get('qty', 0)   # qty 已是原始包数
            unit  = item.get('unit', '')
            rmb   = item.get('rmb_price', 0)
            amount = round(qty_original * rmb, 2)
            total_rmb += amount

            ws.cell(r, 2).value = item.get('code', '')
            ws.cell(r, 3).value = item.get('customs_name', '')
            ws.cell(r, 4).value = d_str
            ws.cell(r, 5).value = boxes
            ws.cell(r, 6).value = f'{int(qty_original) if qty_original == int(qty_original) else qty_original} {unit}'
            ws.cell(r, 7).value = rmb
            ws.cell(r, 8).value = amount

        # 更新 footer
        last       = start + n - 1
        total_row  = 16 + net   # 合计行
        daxie_row  = 17 + net   # 大写行
        ws.cell(total_row, 8).value = f'=SUM(H{start}:H{last})'
        ws.cell(daxie_row, 3).value = rmb_to_chinese(total_rmb)

        # B18 交货期限：按销售合同装船日期提前 2 天同步。
        if ship_date:
            from datetime import timedelta
            delivery_date = ship_date - timedelta(days=2)
            delivery_str  = f'{delivery_date.year}年{delivery_date.month}月{delivery_date.day}日前'
            b18_row = 18 + net
            old_val = ws.cell(b18_row, 2).value or ''
            ws.cell(b18_row, 2).value = re.sub(
                r'\d{4}年\d{1,2}月\d{1,2}日前', delivery_str, old_val)

        # 清理文件名非法字符
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', supplier_name)
        supplier_number = gdata.get('supplier_number', '')
        safe_number = re.sub(r'[\\/:*?"<>|]', '_', supplier_number) if supplier_number else ''
        fname = f'工厂采购合同_{safe_number+"_" if safe_number else ""}{safe_name}_{invoice_no}.xlsx'
        out   = os.path.join(output_path, fname)
        wb.save(out)
        # 图片锚点随插入行下移（商品区从 start=13 行开始，净插入 net 行）
        if net != 0:
            try:
                _shift_drawing_rows(out, start, net)
            except Exception as e:
                print(f'[WARN] 采购合同图片位移失败: {e}')
        results.append(out)

    return results


def generate_purchase_contracts_check(processed, supplier_map, invoice_no, output_path,
                                       sc_date=None, exchange_rate=0, ship_date=None):
    """
    核对用采购合同（对账用）：使用报关数量(PCS) + customs_unit + 发票金额口径。
    模板与工厂采购合同相同（TPL_PURCHASE）。
    """
    if sc_date is None:
        sc_date = date.today()
    items = [i for i in processed if i.get('found')]

    # 按供应商分组
    groups = {}
    for item in items:
        info  = supplier_map.get(item['code'], {})
        name  = info.get('supplierName', '未知供应商') or '未知供应商'
        idx   = info.get('supplierIdx', 99)
        if name not in groups:
            groups[name] = {'items': [], 'idx': idx,
                            'origin_city': info.get('origin_city', ''),
                            'supplier_number': info.get('supplierNumber', '')}
        groups[name]['items'].append(item)

    results = []
    for supplier_name, gdata in groups.items():
        supplier_idx = gdata['idx']
        # 在该供应商范围内按 hs_code+customs_name 合并，得到 PCS 数量
        merged_items = merge_for_customs(gdata['items'])
        n = len(merged_items)
        if n == 0:
            continue

        wb = openpyxl.load_workbook(TPL_PURCHASE)
        ws = wb.active

        purchase_date = _random_workday(sc_date, 1, 5, 'forward')
        contract_no   = f'{invoice_no}-{supplier_idx:02d}'
        origin_city   = gdata.get('origin_city') or '义乌市'
        ws.cell(6, 2).value = f'签订时间/地点: {purchase_date.strftime("%Y-%m-%d")} / {origin_city}'
        ws.cell(6, 7).value = contract_no
        ws.cell(7, 2).value = f'供方：{supplier_name}'

        start = 13; n_tpl = 3
        tpl_styles, tpl_height, net = _insert_rows_in_template(ws, start, n, n_tpl)

        total_rmb = 0.0
        for idx2, item in enumerate(merged_items):
            r = start + idx2
            _apply_row_style(ws, r, tpl_styles, tpl_height)
            qty_pcs   = item.get('qty', 0)
            cunit     = item.get('customs_unit', 'PCS')
            usd_price = item.get('usd_price', 0)
            usd_amt   = item.get('total_usd', 0)
            rmb_price = round(float(usd_price) * float(exchange_rate), 4) if exchange_rate else 0.0
            rmb_amt   = round(float(usd_amt) * float(exchange_rate), 2) if exchange_rate else 0.0
            total_rmb += rmb_amt

            ws.cell(r, 2).value = item.get('hs_code', '')
            ws.cell(r, 3).value = item.get('customs_name', '')
            ws.cell(r, 4).value = ''
            ws.cell(r, 5).value = ''
            ws.cell(r, 6).value = f'{int(qty_pcs) if qty_pcs == int(qty_pcs) else qty_pcs} {cunit}'
            ws.cell(r, 7).value = rmb_price
            ws.cell(r, 8).value = rmb_amt

        last = start + n - 1
        total_row = 16 + net; daxie_row = 17 + net
        ws.cell(total_row, 8).value = f'=SUM(H{start}:H{last})'
        ws.cell(daxie_row, 3).value = rmb_to_chinese(total_rmb)

        # B18 交货期限：按销售合同装船日期提前 2 天同步。
        if ship_date:
            from datetime import timedelta
            delivery_date = ship_date - timedelta(days=2)
            delivery_str  = f'{delivery_date.year}年{delivery_date.month}月{delivery_date.day}日前'
            b18_row = 18 + net
            old_val = ws.cell(b18_row, 2).value or ''
            ws.cell(b18_row, 2).value = re.sub(
                r'\d{4}年\d{1,2}月\d{1,2}日前', delivery_str, old_val)

        safe_name   = re.sub(r'[\\/:*?"<>|]', '_', supplier_name)
        supplier_number = gdata.get('supplier_number', '')
        safe_number = re.sub(r'[\\/:*?"<>|]', '_', supplier_number) if supplier_number else ''
        fname = f'核对用采购合同_{safe_number+"_" if safe_number else ""}{safe_name}_{invoice_no}.xlsx'
        out   = os.path.join(output_path, fname)
        wb.save(out)
        if net != 0:
            try:
                _shift_drawing_rows(out, start, net)
            except Exception as e:
                print(f'[WARN] 核对用采购合同图片位移失败: {e}')
        results.append(out)

    return results


def generate_all(data):
    """
    data = {
        items: [{'code', 'unit', 'qty', 'rmb_price'}, ...],
        exchange_rate: float,
        invoice_no: str,
        target_total_usd: float | None,
        weighed_total_gw: float | None,   # 货物过磅总重量(KGS)
        output_path: str,
    }
    """
    items            = data['items']
    rate             = float(data['exchange_rate'])
    invoice_no       = data.get('invoice_no', 'INV001')
    target_usd       = data.get('target_total_usd')
    weighed_total_gw = data.get('weighed_total_gw')
    output_path      = data['output_path']
    supplier_map     = data.get('supplier_map', {})

    # ── 子文件夹：{order_no}_{remark}_{invoice_no} ────────────────────────
    def _safe_name(s):
        return re.sub(r'[\\/:*?"<>|\n\r\t\x00-\x1f]', '_', s)

    order_no     = data.get('order_no', '')
    order_remark = data.get('order_remark', '')
    if order_no or order_remark:
        folder_name = '_'.join(filter(None, [
            _safe_name(order_no), _safe_name(order_remark), _safe_name(invoice_no)
        ]))
        output_path = os.path.join(output_path, folder_name)

    os.makedirs(output_path, exist_ok=True)

    # 采购合同单独放子目录
    purchase_sub = os.path.join(output_path, '采购合同')
    os.makedirs(purchase_sub, exist_ok=True)

    # origin_map: code → origin_city（来自供应商联系人）
    origin_map = {c: v.get('origin_city', '') for c, v in supplier_map.items()}

    processed = process_items(
        items, rate, target_usd, weighed_total_gw,
        api_dims_map=data.get('api_dims_map'),
        origin_map=origin_map,
        total_cbm=data.get('total_cbm'),
    )

    # ── 退税分流：R列='否' 的商品单独生成报关单，不参与其他步骤 ──
    tax_items    = [i for i in processed if i.get('is_tax_refund', '是') != '否']
    no_tax_items = [i for i in processed if i.get('is_tax_refund', '是') == '否']

    pi_no = _next_pi_no()   # 同一批次共用同一个 PI 号
    files = {
        'customs': generate_customs(tax_items, invoice_no, output_path),
        'invoice': generate_invoice(tax_items, invoice_no, output_path, pi_no=pi_no),
        'packing': generate_packing(tax_items, invoice_no, output_path, pi_no=pi_no),
        'merge':   generate_merge_detail(tax_items, invoice_no, output_path,
                                         supplier_map=data.get('supplier_map', {}),
                                         exchange_rate=rate),
    }

    # 非退税商品单独生成报关单
    if no_tax_items:
        try:
            no_tax_customs = generate_customs(no_tax_items, invoice_no + '_非退税', output_path)
            files['customs_no_tax'] = no_tax_customs
        except Exception as e:
            print(f'[WARN] 非退税报关单生成失败: {e}')

    # ── 阶段一：存档 JSON + 销售合同 + 采购合同 ──────────────────────────
    items_json_path = save_items_json(invoice_no, tax_items, rate, output_path, order_no=order_no,
                                      supplier_map=supplier_map)
    files['items_json'] = items_json_path

    today = date.today()
    ship_date = None
    try:
        sales_path, sc_date, ship_date = generate_sales_contract(tax_items, invoice_no, output_path, today)
        files['sales_contract'] = sales_path
    except Exception as e:
        print(f'[WARN] 销售合同生成失败: {e}')
        sc_date = today
        files['sales_contract'] = None

    try:
        purchase_paths = generate_purchase_contracts(
            tax_items, supplier_map, invoice_no, purchase_sub, sc_date=sc_date, ship_date=ship_date)
        files['purchase_contracts'] = purchase_paths
    except Exception as e:
        print(f'[WARN] 采购合同生成失败: {e}')
        files['purchase_contracts'] = []

    # 核对用采购合同（独立子目录）
    check_sub = os.path.join(output_path, '采购合同+核对用')
    os.makedirs(check_sub, exist_ok=True)
    try:
        check_paths = generate_purchase_contracts_check(
            tax_items, supplier_map, invoice_no, check_sub,
            sc_date=sc_date, exchange_rate=rate, ship_date=ship_date)
        files['purchase_contracts_check'] = check_paths
    except Exception as e:
        print(f'[WARN] 核对用采购合同生成失败: {e}')
        files['purchase_contracts_check'] = []

    # 汇总信息返回给前端
    found = [i for i in tax_items if i.get('found')]
    not_found = [i['code'] for i in processed if not i.get('found')]

    # 计算实际毛重/净重合计
    total_gw = round(sum(i.get('gw_val', 0) for i in found), 2)
    total_nw = round(sum(i.get('nw_val', 0) for i in found), 2)

    return {
        'files':     files,
        'total_usd': round(sum(i['total_usd'] for i in found), 2),
        'total_gw':  total_gw,
        'total_nw':  total_nw,
        'not_found': not_found,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 阶段二：退税联 PDF 处理
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_tax_pdf(pdf_path):
    """
    解析出口退税联 PDF，提取 HS 编码 → 单位/数量/金额映射，以及表头元数据。

    退税联 PDF 单元格格式（打包多行）：
      商品行: '1 9603290090 化妆刷 法定数量/法定单位 2496支 ... 2465.28 USD\n申报数量/申报单位 2496支'
      表头行: '件数\n731'  '毛重（千克）\n13350'  '提运单号\n292002260512035595'  等

    返回:
      hs_data   {hs_code: {'unit': str, 'qty': float, 'amount_usd': float}}
      meta      {'件数': int, '毛重': float, '净重': float, '提运单号': str,
                 '出口日期': str, '合同协议号': str}
      raw_rows  原始行（调试用）
    """
    import pdfplumber

    hs_data  = {}
    meta     = {}
    raw_rows = []

    _META_PATTERNS = [
        ('件数',      r'件数[：:\s]*\n?([0-9,]+)'),
        ('毛重',      r'毛重[（(]千克[）)][：:\s]*\n?([0-9,]+(?:\.[0-9]+)?)'),
        ('净重',      r'净重[（(]千克[）)][：:\s]*\n?([0-9,]+(?:\.[0-9]+)?)'),
        ('提运单号',  r'提运单号[：:\s]*\n?([A-Za-z0-9]+)'),
        ('出口日期',  r'出口日期[：:\s]*\n?(\d{8})'),
        ('合同协议号', r'合同协议号[：:\s]*\n?(\S+)'),
    ]

    def _try_meta(text):
        for key, pat in _META_PATTERNS:
            if key in meta:
                continue
            m = re.search(pat, text)
            if m:
                val = m.group(1).replace(',', '').replace(' ', '')  # 去逗号和空格
                if key == '件数':
                    try:
                        val = int(val)
                    except ValueError:
                        pass
                elif key in ('毛重', '净重'):
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                meta[key] = val

    # 海关编号在 PDF 文字层（表格外），单独用 extract_text 提取
    _CUSTOMS_NO_PAT = re.compile(r'海关编号[：:]\s*([\d\s]{10,25})')

    with pdfplumber.open(pdf_path) as pdf:
        # 先从第一页文字层提取海关编号
        first_text = pdf.pages[0].extract_text() or ''
        m_cno = _CUSTOMS_NO_PAT.search(first_text)
        if m_cno:
            meta['海关编号'] = m_cno.group(1).replace(' ', '').strip()

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row:
                        continue
                    row_clean = [str(c).strip() if c else '' for c in row]
                    raw_rows.append(row_clean)

                    # 合并整行文本，用于正则匹配（换行保留，便于多行单元格）
                    full_text = '\n'.join(row_clean)

                    # 先从全行文本尝试提取表头元数据
                    _try_meta(full_text)

                    # 查找 10 位 HS 编码（商品行）
                    hs_match = re.search(r'\b(\d{10})\b', full_text)
                    if not hs_match:
                        continue
                    hs = hs_match.group(1)

                    # 提取申报数量/申报单位
                    # 格式: 申报数量/申报单位 2496支
                    qty  = 0.0
                    unit = ''
                    m_shen = re.search(
                        r'申报数量/申报单位\s+([0-9,]+(?:\.[0-9]+)?)([\u4e00-\u9fa5A-Za-z]{1,5})',
                        full_text)
                    if m_shen:
                        try:
                            qty  = float(m_shen.group(1).replace(',', ''))
                        except ValueError:
                            pass
                        unit = m_shen.group(2).strip()
                    else:
                        # 备用：法定数量/法定单位（有时申报行缺失）
                        m_fa = re.search(
                            r'法定数量/法定单位\s+([0-9,]+(?:\.[0-9]+)?)([\u4e00-\u9fa5A-Za-z]{1,5})',
                            full_text)
                        if m_fa:
                            try:
                                qty  = float(m_fa.group(1).replace(',', ''))
                            except ValueError:
                                pass
                            unit = m_fa.group(2).strip()

                    # 提取 USD 总金额（行内最后一个"数字 USD"组合；金额可能是整数，如 8142 USD）
                    amount_usd = 0.0
                    usd_matches = re.findall(r'([0-9,]+(?:\.[0-9]+)?)\s*USD', full_text)
                    if usd_matches:
                        try:
                            amount_usd = float(usd_matches[-1].replace(',', ''))
                        except ValueError:
                            pass

                    if hs not in hs_data:
                        hs_data[hs] = {'unit': unit, 'qty': qty, 'amount_usd': amount_usd}
                    else:
                        hs_data[hs]['qty']        += qty
                        hs_data[hs]['amount_usd'] += amount_usd
                        if not hs_data[hs]['unit'] and unit:
                            hs_data[hs]['unit'] = unit

    return hs_data, meta, raw_rows


def generate_invoice_import(invoice_no, pdf_hs_data, items_json_path, output_path, extra_data=None):
    """
    生成发票开具项目信息导入模板（按 HS 编码聚合）。
    """
    with open(items_json_path, encoding='utf-8') as f:
        saved = json.load(f)

    items = saved.get('items', [])

    # 按 (HS编码 + 报关名称 + 报关单位 + 产地 + 材质) 聚合，与报关单 merge_for_customs 分组逻辑完全一致
    hs_map = {}   # (hs_code, customs_name, customs_unit, origin, material) -> {...}
    for item in items:
        hs    = item.get('hs_code', '')
        cname = item.get('customs_name', '')
        if not hs:
            continue
        qty_pcs   = float(item.get('qty_pcs', 0) or item.get('qty', 0))  # PCS，与报关单一致
        rmb_price = float(item.get('rmb_price', 0))
        pcs_per_pkg = float(item.get('pcs_per_pkg', 1) or 1)
        amount    = round(qty_pcs * rmb_price / pcs_per_pkg, 2)

        cunit  = item.get('customs_unit', '')
        origin = item.get('origin', '')
        mat    = item.get('material', '')
        key = (hs, cname, cunit, origin, mat)
        if key not in hs_map:
            # PDF 中该 HS 的单位优先，其次报关单位 customs_unit，再次基础资料 unit
            pdf_unit = (pdf_hs_data.get(hs) or {}).get('unit', '')
            unit = pdf_unit or item.get('customs_unit', '') or item.get('unit', '')
            hs_map[key] = {
                'customs_name': cname,
                'tax_code':     item.get('tax_code', ''),
                'qty_sum':      0.0,
                'amount_sum':   0.0,
                'usd_sum':      0.0,
                'unit':         unit,
            }
        hs_map[key]['qty_sum']    += qty_pcs
        hs_map[key]['amount_sum'] += amount
        hs_map[key]['usd_sum']    += float(item.get('total_usd', 0))

    wb = openpyxl.load_workbook(TPL_INV_IMPORT)
    ws = wb.active

    # 清除已有数据行（从第 4 行开始）
    for r in range(4, ws.max_row + 1):
        for c in range(1, 12):
            ws.cell(r, c).value = None

    # 发票汇率：优先用 items.json 回写的发票汇率，其次传入参数，再次 Stage 1 存档汇率
    ed = extra_data or {}
    inv_rate = (float(saved.get('invoice_exchange_rate') or 0)
                or float(ed.get('invoice_exchange_rate') or 0)
                or float(saved.get('exchange_rate') or 0))

    # 写入数据（从第 4 行开始）
    row = 4
    for (hs, _cname, _cunit, _origin, _mat), agg in hs_map.items():
        # G 列金额：有汇率时用 USD × 汇率，否则退回 qty × rmb_price
        if inv_rate > 0:
            amount_rmb = round(agg['usd_sum'] * inv_rate, 2)
        else:
            amount_rmb = round(agg['amount_sum'], 2)
        ws.cell(row, 1).value = agg['customs_name']      # A 项目名称
        ws.cell(row, 2).value = agg['tax_code']           # B 商品税收分类编码
        ws.cell(row, 3).value = ''                         # C 规格型号（留空）
        ws.cell(row, 4).value = agg['unit']               # D 单位
        ws.cell(row, 5).value = round(agg['qty_sum'], 4)  # E 数量
        ws.cell(row, 6).value = ''                         # F 单价（留空）
        ws.cell(row, 7).value = amount_rmb                 # G 金额（RMB）
        ws.cell(row, 8).value = 0                          # H 税率
        row += 1

    # ── L1：出口业务备注行 ──────────────────────────────────────────────────
    # 报关单号：UI 优先；UI 未填时用 PDF 文字层提取的海关编号（18位）作为备用
    # 注意：绝对不能用 合同协议号 代替报关单号
    export_date = ed.get('export_date') or ed.get('pdf_export_date', '')
    customs_no  = ed.get('customs_no') or ed.get('pdf_customs_no', '')
    bl_no       = ed.get('bl_no')       or ed.get('pdf_bl_no', '')
    # L1 汇率用外管局汇率（inv_rate），不用 UI 汇率
    l1_rate_str = str(inv_rate).rstrip('0').rstrip('.') if inv_rate else ''

    if export_date or customs_no or bl_no:
        # USD 金额按表体行顺序列出，与 hs_map 写入顺序完全一致
        usd_parts = [f'USD{round(agg["usd_sum"], 2)}' for agg in hs_map.values()]
        total_usd = round(sum(agg['usd_sum'] for agg in hs_map.values()), 2)

        l1_text = (
            f'出口业务 一般贸易 出口退税 '
            f'出口日期:{export_date} '
            f'报关单号:{customs_no} '
            f'合同号:{invoice_no} '
            f'成交方式:FOB 金额:'
            f'{"；".join(usd_parts)} 合计USD{total_usd} '
            f'汇率:{l1_rate_str} 提运单号:'
            f'{bl_no}'
        )
        ws.cell(1, 12).value = l1_text   # L1

    out = os.path.join(output_path, f'发票开具项目信息导入_{invoice_no}.xlsx')
    wb.save(out)
    return out


def update_base_data_units(invoice_no, pdf_hs_data, items_json_path, output_path):
    """
    根据 PDF 中的 HS→单位映射更新基础资料 G列（报关产品单位）。
    差异记录到 diff_log，更新前自动备份原文件。
    返回 diff_log 路径（若无差异返回 None）。
    """
    with open(items_json_path, encoding='utf-8') as f:
        saved = json.load(f)

    items = saved.get('items', [])

    # 建立 code → hs_code 映射
    code_to_hs = {it['code']: it.get('hs_code', '') for it in items}

    # 备份基础资料文件
    bak_path = BASE_DATA.replace('.xlsx', f'_bak_{date.today().strftime("%Y%m%d")}.xlsx')
    with _base_data_lock:   # 同一时刻只允许一个线程读写基础资料文件
        if not os.path.exists(bak_path):
            shutil.copy2(BASE_DATA, bak_path)
            print(f'[INFO] 基础资料已备份到: {bak_path}')

        wb = openpyxl.load_workbook(BASE_DATA)
        ws = wb.active

        diff_lines = [f'差异日志 {invoice_no} — {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
                      f'备份文件: {bak_path}', '']

        updated = 0
        for row in ws.iter_rows(min_row=2):
            code_cell = row[1]   # B列
            if not code_cell.value:
                continue
            code = str(code_cell.value).strip()
            hs   = code_to_hs.get(code)
            if not hs:
                continue
            pdf_info = pdf_hs_data.get(hs)
            if not pdf_info or not pdf_info.get('unit'):
                continue

            pdf_unit = pdf_info['unit'].strip()
            cur_unit = str(row[6].value or '').strip()   # G列 = col index 6

            if cur_unit != pdf_unit:
                diff_lines.append(
                    f'code={code}  HS={hs}  当前G列={repr(cur_unit)}  PDF单位={repr(pdf_unit)}  → 已更新'
                )
                row[6].value = pdf_unit
                updated += 1

        if updated > 0:
            wb.save(BASE_DATA)
            diff_lines.append(f'\n共更新 {updated} 条。')
        else:
            diff_lines.append('无差异，未更新。')

    log_path = os.path.join(output_path, f'{invoice_no}_diff_log.txt')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(diff_lines))
    if updated > 0:
        print(f'[INFO] 基础资料 G列更新 {updated} 条，diff 记录: {log_path}')
    return log_path


def compare_pdf_vs_items(pdf_hs_data, meta, items_json_path, output_path, invoice_no):
    """
    将 PDF 退税联数据与 items.json 全面对比，写入 {invoice_no}_pdf_compare.txt。
    对比维度：
      1. HS 码：PDF 有但系统无 / 系统有但 PDF 无
      2. 各 HS 申报数量 vs 系统 PCS 数量
      3. 各 HS 申报金额(USD) vs 系统金额(USD)
      4. 总件数（箱/包）vs PDF 件数
      5. 总毛重 vs PDF 毛重
    返回 (log_path, summary_lines)
    """
    with open(items_json_path, encoding='utf-8') as f:
        saved = json.load(f)
    items = saved.get('items', [])

    # 系统侧：按 HS 汇总 PCS 数量、USD 金额、箱数、毛重
    sys_map         = {}    # {hs: {'pcs': float, 'usd': float, 'boxes': float, 'gw': float, 'unit': str}}
    sys_total_boxes = 0.0
    sys_total_gw    = 0.0
    for it in items:
        hs    = it.get('hs_code', '')
        if not hs:
            continue
        pcs   = float(it.get('qty_pcs', 0) or it.get('qty', 0))
        usd   = float(it.get('total_usd', 0))
        boxes = float(it.get('qty', 0))
        gw    = float(it.get('gw_val', 0))
        unit  = it.get('customs_unit', '') or it.get('unit', '')
        if hs not in sys_map:
            sys_map[hs] = {'pcs': 0.0, 'usd': 0.0, 'boxes': 0.0, 'gw': 0.0, 'unit': unit}
        sys_map[hs]['pcs']   += pcs
        sys_map[hs]['usd']   += usd
        sys_map[hs]['boxes'] += boxes
        sys_map[hs]['gw']    += gw
        sys_total_boxes += boxes
        sys_total_gw    += gw

    pdf_hs_set = set(pdf_hs_data.keys())
    sys_hs_set = set(sys_map.keys())
    only_pdf   = pdf_hs_set - sys_hs_set
    only_sys   = sys_hs_set - pdf_hs_set
    common     = pdf_hs_set & sys_hs_set

    lines = [
        f'=== PDF退税联 vs 系统数据 对比报告 ===',
        f'单号: {invoice_no}    时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        '',
    ]

    # 1. HS 码集合对比
    lines.append('【1】HS编码集合核对')
    if only_pdf:
        lines.append(f'  ⚠️  PDF有但系统无 ({len(only_pdf)}个): {", ".join(sorted(only_pdf))}')
    else:
        lines.append('  ✅  PDF中的HS码均在系统中')
    if only_sys:
        lines.append(f'  ⚠️  系统有但PDF无 ({len(only_sys)}个): {", ".join(sorted(only_sys))}')
    else:
        lines.append('  ✅  系统中的HS码均在PDF中')
    lines.append('')

    # 2. 各 HS 申报单位
    lines.append('【2】各HS申报单位核对')
    for hs in sorted(common):
        pdf_unit = (pdf_hs_data[hs].get('unit') or '').strip()
        sys_unit = (sys_map[hs].get('unit') or '').strip()
        if not pdf_unit:
            lines.append(f'  ⚠️  HS={hs}  PDF单位=（未提取到）  系统单位={sys_unit}')
        elif pdf_unit == sys_unit:
            lines.append(f'  ✅  HS={hs}  申报单位={pdf_unit}')
        else:
            lines.append(f'  ⚠️  HS={hs}  PDF单位={pdf_unit}  系统单位={sys_unit}  （已在diff_log中处理）')
    lines.append('')

    # 3. 各 HS 申报数量 vs 系统 PCS 数量
    lines.append('【3】各HS申报数量 vs 系统PCS数量')
    qty_ok = True
    for hs in sorted(common):
        pdf_qty = round(pdf_hs_data[hs].get('qty', 0), 4)
        sys_pcs = round(sys_map[hs]['pcs'], 4)
        delta   = round(pdf_qty - sys_pcs, 4)
        status  = '✅' if abs(delta) <= 0.01 else '❌'
        lines.append(f'  {status} HS={hs}  PDF申报={pdf_qty}  系统PCS={sys_pcs}  差={delta}')
        if abs(delta) > 0.01:
            qty_ok = False
    if not common:
        lines.append('  （无共同HS码）')
    lines.append('')

    # 4. 各 HS 申报金额 vs 系统 USD 金额
    lines.append('【4】各HS申报金额(USD) vs 系统金额(USD)')
    for hs in sorted(common):
        pdf_amt = round(pdf_hs_data[hs].get('amount_usd', 0), 2)
        sys_amt = round(sys_map[hs]['usd'], 2)
        if pdf_amt == 0:
            lines.append(f'  ⚠️  HS={hs}  PDF金额=（未提取到）  系统={sys_amt}')
            continue
        delta   = round(pdf_amt - sys_amt, 2)
        pct     = abs(delta / sys_amt * 100) if sys_amt else 0.0
        status  = '✅' if pct < 1 else ('⚠️' if pct < 5 else '❌')
        lines.append(f'  {status} HS={hs}  PDF={pdf_amt}  系统={sys_amt}  差={delta} ({pct:.1f}%)')
    if not common:
        lines.append('  （无共同HS码）')
    lines.append('')

    # 5. 总件数（箱/包）vs PDF 件数
    lines.append('【5】总件数（箱/包）核对')
    pdf_boxes = meta.get('件数')
    if pdf_boxes is not None:
        sys_boxes_int = int(round(sys_total_boxes))
        delta_b = int(pdf_boxes) - sys_boxes_int
        status  = '✅' if delta_b == 0 else '❌'
        lines.append(f'  {status} PDF件数={pdf_boxes}  系统件数={sys_boxes_int}  差={delta_b}')
    else:
        lines.append('  ⚠️  PDF未提取到件数，无法对比')
    lines.append('')

    # 6. 总毛重 vs PDF 毛重
    lines.append('【6】总毛重(KGS)核对')
    pdf_gw = meta.get('毛重')
    if pdf_gw is not None:
        if sys_total_gw > 0:
            delta_gw = round(float(pdf_gw) - sys_total_gw, 1)
            pct_gw   = abs(delta_gw / float(pdf_gw) * 100) if pdf_gw else 0.0
            status   = '✅' if pct_gw < 2 else ('⚠️' if pct_gw < 10 else '❌')
            lines.append(f'  {status} PDF毛重={pdf_gw}  系统估算={round(sys_total_gw, 1)}  差={delta_gw} ({pct_gw:.1f}%)')
        else:
            lines.append(f'  ℹ️  PDF毛重={pdf_gw}  系统侧无毛重数据（items.json 未含 gw_val）')
    else:
        lines.append('  ⚠️  PDF未提取到毛重，无法对比')

    # PDF 其他元数据展示
    lines.append('')
    lines.append('【PDF元数据】')
    for k, v in meta.items():
        lines.append(f'  {k}: {v}')

    # ── 写 txt ──────────────────────────────────────────────────────────
    txt_path = os.path.join(output_path, f'{invoice_no}_pdf_compare.txt')
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    # ── 写 Excel ─────────────────────────────────────────────────────────
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    _thin = Side(style='thin', color='CCCCCC')
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _fill_green  = PatternFill('solid', fgColor='C6EFCE')
    _fill_yellow = PatternFill('solid', fgColor='FFEB9C')
    _fill_red    = PatternFill('solid', fgColor='FFC7CE')
    _fill_blue   = PatternFill('solid', fgColor='BDD7EE')
    _fill_header = PatternFill('solid', fgColor='4472C4')
    _font_header = Font(bold=True, color='FFFFFF', size=10)
    _font_bold   = Font(bold=True, size=10)
    _font_normal = Font(size=10)
    _align_c     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _align_l     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    wb_cmp = openpyxl.Workbook()
    ws     = wb_cmp.active
    ws.title = 'PDF对比'

    # 标题行
    ws.merge_cells('A1:G1')
    ws['A1'].value     = f'PDF退税联 vs 系统数据 对比报告 — {invoice_no}  生成时间:{datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A1'].font      = Font(bold=True, size=12)
    ws['A1'].alignment = _align_l
    ws.row_dimensions[1].height = 20

    # PDF 元数据行
    meta_row = 2
    meta_str = '  '.join(f'{k}:{v}' for k, v in meta.items())
    ws.merge_cells(f'A{meta_row}:G{meta_row}')
    ws[f'A{meta_row}'].value     = f'PDF元数据：{meta_str}'
    ws[f'A{meta_row}'].font      = Font(italic=True, size=10, color='444444')
    ws[f'A{meta_row}'].alignment = _align_l
    ws.row_dimensions[meta_row].height = 16

    # 表头
    hdr_row = 4
    headers = ['类别', 'HS编码', '核对项目', 'PDF值', '系统值', '差异', '状态']
    col_widths = [12, 14, 18, 16, 16, 12, 8]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(hdr_row, ci)
        c.value     = h
        c.font      = _font_header
        c.fill      = _fill_header
        c.alignment = _align_c
        c.border    = _border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[hdr_row].height = 18

    r = hdr_row + 1

    def _write_row(category, hs, item, pdf_val, sys_val, delta, status):
        nonlocal r
        vals = [category, hs, item, pdf_val, sys_val, delta, status]
        if status == '✅':
            fill = _fill_green
        elif status == '❌':
            fill = _fill_red
        elif status == '⚠️':
            fill = _fill_yellow
        else:
            fill = None
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci)
            c.value     = v
            c.font      = _font_normal
            c.alignment = _align_c if ci != 3 else _align_l
            c.border    = _border
            if fill:
                c.fill = fill
        ws.row_dimensions[r].height = 16
        r += 1

    def _section_title(title):
        nonlocal r
        ws.merge_cells(f'A{r}:G{r}')
        c = ws.cell(r, 1)
        c.value     = title
        c.font      = _font_bold
        c.fill      = _fill_blue
        c.alignment = _align_l
        c.border    = _border
        ws.row_dimensions[r].height = 16
        r += 1

    # ──────────────────────────────────────────────────────
    # 【0】汇总行（总件数 + 总毛重）
    _section_title('汇总核对')
    pdf_boxes = meta.get('件数')
    if pdf_boxes is not None:
        sys_boxes_int = int(round(sys_total_boxes))
        delta_b = int(pdf_boxes) - sys_boxes_int
        status  = '✅' if delta_b == 0 else '❌'
        _write_row('汇总', '', '总件数(箱/包)', pdf_boxes, sys_boxes_int, delta_b, status)
    else:
        _write_row('汇总', '', '总件数(箱/包)', '未提取', int(round(sys_total_boxes)), '', '⚠️')

    pdf_gw = meta.get('毛重')
    if pdf_gw is not None:
        if sys_total_gw > 0:
            delta_gw = round(float(pdf_gw) - sys_total_gw, 1)
            pct_gw   = abs(delta_gw / float(pdf_gw) * 100) if pdf_gw else 0.0
            status   = '✅' if pct_gw < 2 else ('⚠️' if pct_gw < 10 else '❌')
            _write_row('汇总', '', f'总毛重KGS ({pct_gw:.1f}%差)', pdf_gw, round(sys_total_gw, 1), delta_gw, status)
        else:
            _write_row('汇总', '', '总毛重KGS', pdf_gw, '无数据', '', '⚠️')
    else:
        _write_row('汇总', '', '总毛重KGS', '未提取', round(sys_total_gw, 1), '', '⚠️')

    # ──────────────────────────────────────────────────────
    # 【1】HS 码集合
    _section_title('HS编码集合核对')
    for hs in sorted(only_pdf):
        _write_row('HS集合', hs, 'PDF有但系统无', hs, '', '', '⚠️')
    for hs in sorted(only_sys):
        _write_row('HS集合', hs, '系统有但PDF无', '', hs, '', '⚠️')
    if not only_pdf and not only_sys:
        _write_row('HS集合', '', '两侧HS码完全一致', '', '', '', '✅')

    # ──────────────────────────────────────────────────────
    # 【2】单位 + 【3】数量 + 【4】金额（合并到一张按HS排列的表）
    _section_title('各HS编码详细核对（单位 / 申报数量 / 申报金额）')
    for hs in sorted(common):
        # 单位
        pdf_unit = (pdf_hs_data[hs].get('unit') or '').strip()
        sys_unit = (sys_map[hs].get('unit') or '').strip()
        if not pdf_unit:
            _write_row('单位', hs, '申报单位', '未提取', sys_unit, '', '⚠️')
        elif pdf_unit == sys_unit:
            _write_row('单位', hs, '申报单位', pdf_unit, sys_unit, '', '✅')
        else:
            _write_row('单位', hs, '申报单位', pdf_unit, sys_unit, f'{pdf_unit}≠{sys_unit}', '⚠️')
        # 数量
        pdf_qty = round(pdf_hs_data[hs].get('qty', 0), 4)
        sys_pcs = round(sys_map[hs]['pcs'], 4)
        delta_q = round(pdf_qty - sys_pcs, 4)
        status  = '✅' if abs(delta_q) <= 0.01 else '❌'
        _write_row('数量', hs, '申报数量(PCS)', pdf_qty, sys_pcs, delta_q, status)
        # 金额
        pdf_amt = round(pdf_hs_data[hs].get('amount_usd', 0), 2)
        sys_amt = round(sys_map[hs]['usd'], 2)
        if pdf_amt == 0:
            _write_row('金额', hs, '申报金额(USD)', '未提取', sys_amt, '', '⚠️')
        else:
            delta_a = round(pdf_amt - sys_amt, 2)
            pct_a   = abs(delta_a / sys_amt * 100) if sys_amt else 0.0
            status  = '✅' if pct_a < 1 else ('⚠️' if pct_a < 5 else '❌')
            _write_row('金额', hs, f'申报金额USD ({pct_a:.1f}%差)', pdf_amt, sys_amt, delta_a, status)

    # 冻结首行
    ws.freeze_panes = 'A5'

    xlsx_path = os.path.join(output_path, f'{invoice_no}_pdf_compare.xlsx')
    wb_cmp.save(xlsx_path)
    print(f'[COMPARE] PDF对比Excel写入: {xlsx_path}')
    return xlsx_path, lines


def generate_linghang_invoices(invoice_no, pdf_hs_data, supplier_map,
                                items_json_path, output_path, exchange_rate=0):
    """
    按供应商生成凌航 VAT 发票，每供应商一份（只填表体，不填供应商抬头）。
    分组逻辑：供应商 × (hs_code + 报关商品名称) = 一行
    金额 = 该供应商该组所有 code 的 total_usd 之和 × exchange_rate
    数量 = qty_pcs（PCS），单价 = 金额 ÷ 数量（保留 13 位小数）
    返回生成文件路径列表。
    """
    with open(items_json_path, encoding='utf-8') as f:
        saved = json.load(f)

    # 凌航发票用 UI 填写的汇率（Stage 1 存档），只有发票导入模板才用外管局汇率
    exchange_rate = (float(saved.get('exchange_rate') or 0)
                     or exchange_rate
                     or float(saved.get('invoice_exchange_rate') or 0))

    items = saved.get('items', [])

    # 预先计算每个 (hs_code, customs_name) 全组的平均USD单价，用于发票整数取量
    _grp_totals = {}
    for _it in items:
        _k = (_it.get('hs_code', ''), _it.get('customs_name', ''))
        _pcs = float(_it.get('qty_pcs', 0)) or float(_it.get('qty', 0)) * float(_it.get('pcs_per_pkg', 1) or 1)
        if _k not in _grp_totals:
            _grp_totals[_k] = {'usd': 0.0, 'pcs': 0.0}
        _grp_totals[_k]['usd'] += float(_it.get('total_usd', 0))
        _grp_totals[_k]['pcs'] += _pcs
    group_avg_usd = {k: v['usd'] / v['pcs'] if v['pcs'] else 0
                     for k, v in _grp_totals.items()}

    # 按 供应商 → (hs_code, customs_name) 双层分组，汇总 usd_sum / pcs_sum
    supplier_groups = {}   # {sname: {'idx': int, 'rows': {(hs,cname): {...}}}}
    for item in items:
        code  = item.get('code', '')
        info  = supplier_map.get(code, {})
        sname = (info.get('supplierName') or '未知供应商').strip()
        sidx  = info.get('supplierIdx', 99)
        hs    = item.get('hs_code', '')
        cname = item.get('customs_name', '')
        key   = (hs, cname)

        if sname not in supplier_groups:
            snumber = (info.get('supplierNumber') or '').strip()
            supplier_groups[sname] = {'idx': sidx, 'number': snumber, 'rows': {}}
        rows = supplier_groups[sname]['rows']
        if key not in rows:
            rows[key] = {
                'customs_name': cname,
                'customs_unit': item.get('customs_unit', 'PCS'),   # G列 报关单位
                'project_name': item.get('project_name', ''),       # X列 项目名称
                'usd_sum':      0.0,
                'pcs_sum':      0.0,
                'rmb_sum':      0.0,   # 采购合同RMB（qty × rmb_price，与采购合同口径一致）
            }
        rows[key]['usd_sum'] += float(item.get('total_usd', 0))
        rows[key]['pcs_sum'] += float(item.get('qty_pcs', item.get('qty', 0)))
        rows[key]['rmb_sum'] += float(item.get('qty', 0)) * float(item.get('rmb_price', 0))

    # 预计算每个 (hs, cname) 组在各供应商的 bill_qty，并修正合计误差
    # {(hs,cname): {sname: int_qty}}
    _bill_qty_map = {}
    for _sname, _gdata in supplier_groups.items():
        for _key, _rd in _gdata['rows'].items():
            _avg = group_avg_usd.get(_key, 0)
            _raw = _rd['usd_sum'] / _avg if _avg and _rd['usd_sum'] else _rd['pcs_sum']
            _qty = max(1, round(_raw))
            if _key not in _bill_qty_map:
                _bill_qty_map[_key] = {}
            _bill_qty_map[_key][_sname] = {'qty': _qty, 'raw': _raw}
    # 逐组检查合计，调整误差最大的供应商
    for _key, _sup_qtys in _bill_qty_map.items():
        _target = int(round(_grp_totals[_key]['pcs']))
        _total  = sum(v['qty'] for v in _sup_qtys.values())
        _diff   = _total - _target
        if _diff != 0 and _sup_qtys:
            _worst = max(_sup_qtys, key=lambda s: abs(_sup_qtys[s]['raw'] - round(_sup_qtys[s]['raw'])))
            _sup_qtys[_worst]['qty'] = max(1, _sup_qtys[_worst]['qty'] - _diff)

    results = []
    for supplier_name, gdata in supplier_groups.items():
        supplier_idx  = gdata['idx']
        row_items     = list(gdata['rows'].items())   # [(key, rd), ...]
        row_list      = [rd for _, rd in row_items]
        n             = len(row_list)

        wb = openpyxl.load_workbook(TPL_LINGHANG)
        ws = wb.active

        start = 6     # 第一数据行
        n_tpl = 3     # 占位行 6-8

        tpl_styles, tpl_height, net = _insert_rows_in_template(ws, start, n, n_tpl)

        total_amount = 0.0
        for idx2, (row_key, rd) in enumerate(row_items):
            r = start + idx2
            _apply_row_style(ws, r, tpl_styles, tpl_height)

            # 重新建立数据行内的合并
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)   # A:D
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)  # H:J
            ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12) # K:L

            amount_rmb_calc = round(rd['usd_sum'] * exchange_rate, 2) if exchange_rate > 0 else 0.0
            contract_rmb    = round(rd.get('rmb_sum', 0), 2)

            # 与采购合同金额比对，决定实际使用的RMB金额
            CLOSE_THRESHOLD = 0.05   # 5%以内视为"相差不大"
            if contract_rmb > 0:
                diff_pct = abs(amount_rmb_calc - contract_rmb) / contract_rmb
                if diff_pct < CLOSE_THRESHOLD:
                    amount_rmb = contract_rmb   # 直接用采购合同金额
                else:
                    amount_rmb = contract_rmb   # 差异大时仍用合同金额，但输出警告
                    print(f'[WARN][凌航发票] {supplier_name} · {rd["customs_name"]}: '
                          f'换算金额 {amount_rmb_calc} vs 采购合同 {contract_rmb}, '
                          f'差异 {diff_pct:.1%}，请核查')
            else:
                amount_rmb = amount_rmb_calc   # 无采购合同数据，退回换算值

            # 使用预计算（已修正合计误差）的整数开票数量
            bill_qty   = _bill_qty_map.get(row_key, {}).get(supplier_name, {}).get('qty',
                             max(1, round(rd['pcs_sum'])))
            unit_price = round(amount_rmb / bill_qty, 13) if bill_qty > 0 else 0.0
            total_amount += amount_rmb

            pname = rd.get('project_name', '')
            if not pname:
                raise ValueError(
                    f'凌航发票：商品「{rd["customs_name"]}」X列（项目名称）为空，'
                    f'请先在基础资料表X列填写项目名称')
            ws.cell(r, 1).value  = f'*日用杂品*{pname}{rd["customs_name"]}'  # A: 项目名称
            ws.cell(r, 5).value  = ''                    # E: 规格型号（留空）
            ws.cell(r, 6).value  = rd.get('customs_unit', 'PCS')             # F: 单位
            ws.cell(r, 7).value  = bill_qty              # G: 数量（整数，以报关均价取整）
            ws.cell(r, 8).value  = unit_price            # H: 单价（微调保证金额精确）
            ws.cell(r, 11).value = amount_rmb            # K: 金额（直接填，不用公式）
            ws.cell(r, 13).value = ''                    # M: 税率（留空）
            ws.cell(r, 14).value = ''                    # N: 税额（留空）

        # 合计行（原 row 9）
        total_row = 9 + net
        ws.cell(total_row, 6).value  = rmb_to_chinese(total_amount)
        ws.cell(total_row, 14).value = total_amount

        # 采购合同号（原 row 12，B12:N12 merged）+ 合同号（D12）
        contract_row = 12 + net
        contract_no  = f'{invoice_no}-{supplier_idx:02d}'
        ws.cell(contract_row, 2).value = f'采购合同号：{contract_no}'
        ws.cell(contract_row, 4).value = contract_no

        safe_name   = re.sub(r'[\\/:*?"<>|]', '_', supplier_name)
        supplier_number = gdata.get('number', '')
        safe_number = re.sub(r'[\\/:*?"<>|]', '_', supplier_number) if supplier_number else ''
        fname = f'凌航发票_{safe_number+"_" if safe_number else ""}{safe_name}_{invoice_no}.xlsx'
        out   = os.path.join(output_path, fname)
        wb.save(out)
        results.append(out)

    return results


def process_tax_pdf(data):
    """
    阶段二主函数：生成发票导入模板 + 凌航发票。
    pdf_path 可选：
      - 有 PDF → 解析 HS→单位，同时更新基础资料 G列
      - 无 PDF → pdf_hs_data={} 直接用 items.json 里的单位填表，跳过基础资料更新
    data = {
        invoice_no:   str,
        pdf_path:     str | '',   # 可为空
        output_path:  str,
        supplier_map: dict
    }
    """
    invoice_no   = data['invoice_no']
    pdf_path     = (data.get('pdf_path') or '').strip()
    output_path  = data['output_path']
    supplier_map = data.get('supplier_map', {})

    os.makedirs(output_path, exist_ok=True)
    # 优先用调用方传入的路径（来自 _items_store），否则按旧逻辑兼容
    items_json_path = data.get('items_json_path') or ''
    if not items_json_path:
        order_no = data.get('order_no', '') or ''
        file_key = order_no or invoice_no or 'items'
        items_json_path = os.path.join(output_path, f'{file_key}_items.json')

    has_pdf = bool(pdf_path and os.path.exists(pdf_path))

    pdf_meta       = {}
    compare_log    = None
    comparison_lines = []

    if has_pdf:
        pdf_hs_data, pdf_meta, _ = parse_tax_pdf(pdf_path)
        print(f'[PDF] 解析到 {len(pdf_hs_data)} 个 HS 编码: {list(pdf_hs_data.keys())}')
        print(f'[PDF] 表头元数据: {pdf_meta}')
        # 更新基础资料 G列（单位）
        diff_log = update_base_data_units(invoice_no, pdf_hs_data, items_json_path, output_path)
        # 全面对比（件数、毛重、HS码、数量、金额）
        try:
            compare_log, comparison_lines = compare_pdf_vs_items(
                pdf_hs_data, pdf_meta, items_json_path, output_path, invoice_no)
        except Exception as e:
            print(f'[WARN] compare_pdf_vs_items 失败: {e}')
    else:
        pdf_hs_data = {}
        diff_log    = None
        print('[PDF] 无退税联 PDF，直接用 items.json 数据生成')

    # 生成发票导入模板
    extra = {k: data.get(k, '') for k in ('export_date', 'customs_no', 'bl_no')}
    extra['invoice_exchange_rate'] = data.get('invoice_exchange_rate') or 0
    # 补充 PDF 提取的元数据作为 L1 备用（UI 未填时自动使用）
    if pdf_meta:
        raw_dt = str(pdf_meta.get('出口日期', ''))
        if len(raw_dt) == 8:
            raw_dt = f'{raw_dt[:4]}-{raw_dt[4:6]}-{raw_dt[6:8]}'
        extra['pdf_export_date'] = raw_dt
        extra['pdf_bl_no']       = str(pdf_meta.get('提运单号', ''))
        extra['pdf_customs_no']  = str(pdf_meta.get('海关编号', ''))   # 报关单号 = 海关编号（18位）
    inv_import = generate_invoice_import(invoice_no, pdf_hs_data, items_json_path, output_path, extra_data=extra)

    # 生成凌航发票
    linghang = generate_linghang_invoices(invoice_no, pdf_hs_data, supplier_map,
                                          items_json_path, output_path)

    return {
        'invoice_import':    inv_import,
        'diff_log':          diff_log,
        'compare_log':       compare_log,
        'comparison_lines':  comparison_lines,
        'linghang_invoices': linghang,
        'hs_codes_found':    list(pdf_hs_data.keys()),
        'pdf_meta':          pdf_meta,
        'used_pdf':          has_pdf,
    }
