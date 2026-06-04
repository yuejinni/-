"""
AI 识别商品基础信息（材料）
- 支持图片 URL（优先）或 base64
- 调用 GLM（智谱）视觉 API
- 结果写入 报关产品基础资料（智谱）.xlsx
"""
import os
import sys
import json
import ssl
import base64
import http.client
import threading
from urllib.parse import quote as _url_quote, urlparse
import openpyxl

# SSL 上下文：优先用 certifi 证书，否则用系统默认
try:
    import certifi
    _SSL_CTX = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    _SSL_CTX = ssl.create_default_context()

# ── 路径 ─────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    _EXE_DIR = os.path.dirname(sys.executable)
else:
    _EXE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')

_CONFIG_FILE   = os.path.join(_EXE_DIR, 'ai_config.json')
_TAX_TABLE_XLS = os.path.join(_EXE_DIR, '商品和服务税收分类编码表.xls')
_EXCEL_LOCK    = threading.Lock()   # 防并发写 Excel
_tax_table_cache = None             # {分类简称: [(名称, 编码), ...]}

# ── 材料关键词 → HS 编码映射 ─────────────────────────────────────────
# 匹配顺序即优先级；material_name 写 I 列，hs_code 写 E 列
_RULES = [
    # ── 箱包类 ──
    (['帆布', '尼龙', '涤纶', '布', '织物', '化纤', '无纺布',
      'canvas', 'nylon', 'polyester', 'fabric', 'textile'],
     '4202220000'),
    (['塑料', '硬壳', 'abs', 'pp ', ' pc', 'plastic', 'eva', 'tpu'],
     '4202220000'),
    (['皮革', '真皮', '头层皮', 'leather', 'genuine'],
     '4202210000'),
    (['pu', 'pvc', '人造革', '仿皮', '合成革', '超纤', '荔枝纹', '压纹'],
     '4202210000'),
    (['竹', 'bamboo'],
     '4621100000'),
    (['草', '稻草', '藤', '麻', '席', 'straw', 'rattan', 'wicker', 'rush'],
     '4602191000'),
    # ── 饰品类 ──
    (['925', '999银', '足银'],
     '7113110000'),   # 银质首饰
    (['银', 'silver'],
     '7113110000'),   # 银质首饰
    (['珍珠', 'pearl'],
     '7116100000'),   # 珍珠制品
    (['水晶', '玻璃珠', 'crystal', 'glass bead'],
     '7018109090'),   # 玻璃小珠
    (['金属', '合金', '锌合金', '铜', '铁', '铝', '不锈钢', '钢',
      'metal', 'alloy', 'zinc', 'copper', 'iron', 'aluminum', 'steel', 'stainless'],
     '7117190000'),   # 贱金属仿首饰
    (['树脂', '亚克力', '丙烯', 'resin', 'acrylic'],
     '3926200000'),   # 塑料饰品
]


# ── 内部函数 ─────────────────────────────────────────────────────────

def _load_config():
    if os.path.exists(_CONFIG_FILE):
        try:
            with open(_CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _build_prompt(name, category):
    return (
        '请仔细观察图片中商品，分析以下信息：\n'
        f'商品参考名称：{name or "未知"}\n'
        f'商品类别：{category or "未知"}\n\n'
        '请严格按以下格式输出，每行一项，不要输出其他内容：\n'
        '材料：从以下选项中选一个最匹配的实际材料词，'
        '选项：帆布、尼龙、涤纶、塑料、皮革、PU、人造革、竹子、草编、'
        '金属、合金、锌合金、银、不锈钢、铜、珍珠、水晶、树脂、亚克力。'
        '若以上均不符合，直接写出实际材料名称（≤8字，只写材料名，不要写"简短描述"等说明文字）\n'
        '商品名称：根据图片识别该商品的中文品名（≤10字）\n'
        '用途：简述该商品的主要用途（≤10字）'
    )


def _parse_ai_response(ai_text):
    """解析 AI 多行回复，返回 (material_text, product_name, usage)"""
    material_text = None
    product_name  = None
    usage         = None
    for line in ai_text.strip().split('\n'):
        line = line.strip()
        if not line:
            continue
        # 兼容中英文冒号
        if line.startswith('材料') and ('：' in line or ':' in line):
            material_text = line.split('：', 1)[-1].split(':', 1)[-1].strip()
        elif line.startswith('商品名称') and ('：' in line or ':' in line):
            product_name = line.split('：', 1)[-1].split(':', 1)[-1].strip()
        elif line.startswith('用途') and ('：' in line or ':' in line):
            usage = line.split('：', 1)[-1].split(':', 1)[-1].strip()
    # 如果 AI 没有按格式返回，把整段当作材料文本（兼容旧逻辑）
    if not material_text:
        material_text = ai_text.strip()
    return material_text, product_name, usage


def _match_material(ai_text):
    """返回 (material_name, hs_code) 或 (None, None)；material_name 取 AI 原文"""
    t = ai_text.lower().strip()
    for keywords, hs_code in _RULES:
        for kw in keywords:
            if kw.strip().lower() in t:
                return ai_text.strip(), hs_code
    return None, None


def _safe_url(url):
    """将 URL 中的非 ASCII 字符 percent-encode，避免 urllib latin-1 编码错误"""
    try:
        url.encode('ascii')
        return url
    except UnicodeEncodeError:
        return _url_quote(url, safe=':/?=&%#+@!$,;~*()[]@')


def _http_get(url, depth=0):
    """用 http.client 下载 URL，自动跟随最多 3 次跳转，返回 bytes 或 None"""
    if depth > 3:
        return None
    try:
        parsed  = urlparse(_safe_url(url))
        host    = parsed.netloc
        path    = parsed.path + ('?' + parsed.query if parsed.query else '')
        https   = parsed.scheme == 'https'
        ConnCls = http.client.HTTPSConnection if https else http.client.HTTPConnection
        conn    = ConnCls(host, timeout=15, context=(_SSL_CTX if https else None))
        try:
            conn.request('GET', path, headers={'User-Agent': 'Mozilla/5.0'})
            resp = conn.getresponse()
            if resp.status in (301, 302, 303, 307, 308):
                location = resp.getheader('Location', '')
                resp.read()
                return _http_get(location, depth + 1) if location else None
            if resp.status in (200, 206):
                return resp.read()
            return None
        finally:
            conn.close()
    except Exception:
        return None


def _url_to_b64(image_url):
    """在独立线程下载图片（避免 urllib 信号超时污染主线程），返回 base64 或 None"""
    result = [None]

    def _worker():
        data = _http_get(image_url)
        if data:
            result[0] = base64.b64encode(data).decode('ascii')

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=25)   # 线程级超时，不依赖系统信号
    return result[0]


def _call_qianwen(image_url, image_b64, api_key, model, prompt_text):
    # 优先把 URL 下载成 base64，避免千问服务器无法访问 COS 签名链接
    if image_url and not image_b64:
        image_b64 = _url_to_b64(image_url)
        if not image_b64:
            raise ValueError('图片下载失败，请确认图片链接可访问（COS 签名是否已过期）')
        image_url = None

    content = []
    if image_b64:
        content.append({'type': 'image_url',
                        'image_url': {'url': f'data:image/jpeg;base64,{image_b64}'}})
    elif image_url:
        content.append({'type': 'image_url', 'image_url': {'url': _safe_url(image_url)}})
    content.append({'type': 'text', 'text': prompt_text})

    payload = json.dumps({
        'model':    model,
        'messages': [{'role': 'user', 'content': content}],
        'max_tokens': 150,
    }).encode('utf-8')

    # 用 http.client 直接发请求，避免 urllib 在 macOS 上信号级联超时问题
    conn = http.client.HTTPSConnection('dashscope.aliyuncs.com', timeout=60,
                                       context=_SSL_CTX)
    try:
        conn.request('POST', '/compatible-mode/v1/chat/completions',
                     body=payload,
                     headers={
                         'Authorization': f'Bearer {api_key}',
                         'Content-Type':  'application/json',
                     })
        resp   = conn.getresponse()
        result = json.loads(resp.read().decode('utf-8'))
    finally:
        conn.close()
    return result['choices'][0]['message']['content'].strip()


def get_existing_codes(excel_path):
    """一次性读取 Excel B 列所有编码，返回 set"""
    codes = set()
    with _EXCEL_LOCK:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        try:
            for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
                if row[0] is not None:
                    codes.add(str(row[0]).strip())
        finally:
            wb.close()
    return list(codes)


def _find_excel(config):
    """按优先级查找 Excel：config → exe 同目录"""
    path = config.get('excel_path', '').strip()
    if path and os.path.exists(path):
        return path
    guess = os.path.join(_EXE_DIR, '报关产品基础资料（智谱）.xlsx')
    if os.path.exists(guess):
        return guess
    return None


def _check_existing(excel_path, code):
    """查 B 列是否已有该编码，有则返回现有数据，否则返回 None
    如果 F 列（商品名称）或 J 列（用途）为空，视为数据不完整，返回 None 以重新识别
    """
    with _EXCEL_LOCK:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        try:
            # read_only 模式必须用 iter_rows 顺序读取，不能随机访问 ws.cell(r,c)
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if len(row) < 2:
                    continue
                b = row[1]  # B 列（0-indexed）
                if b is not None and str(b).strip() == code:
                    material = row[8]  if len(row) > 8  else None  # I 列
                    hs_raw   = row[4]  if len(row) > 4  else None  # E 列
                    product_name = row[5] if len(row) > 5 else None  # F 列
                    usage        = row[9] if len(row) > 9 else None  # J 列
                    hs_code  = str(int(hs_raw)) if hs_raw else None
                    # F/J 为空，或 I（材料）为空 → 重新识别以补全
                    if not product_name or not usage or not material:
                        return None
                    return {
                        'code':        code,
                        'ai_text':     str(material or ''),
                        'material':    str(material) if material else None,
                        'hs_code':     hs_code,
                        'product_name': str(product_name) if product_name else None,
                        'usage':        str(usage) if usage else None,
                        'row':         row_idx,
                        'matched':     bool(material),
                        'skipped':     True,
                    }
        finally:
            wb.close()
    return None


def _find_existing_row(ws, code):
    """查找 B 列中已有该编码的行号，返回行号或 None"""
    for r in range(2, ws.max_row + 1):
        b_val = ws.cell(r, 2).value
        if b_val is not None and str(b_val).strip() == code:
            return r
    return None


def _write_row(excel_path, code, material_name, hs_code, unknown_text,
               c_value=None, product_name=None, usage=None):
    with _EXCEL_LOCK:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # 优先查找已有同编号的行（补全F/J场景），否则追加新行
        existing_row = _find_existing_row(ws, code)

        if existing_row:
            new_row = existing_row
        else:
            # 找 B 列最后有值的行，追加新行
            last_row = 1
            for r in range(ws.max_row, 0, -1):
                if ws.cell(r, 2).value is not None:
                    last_row = r
                    break
            new_row = last_row + 1

            # A: 项号 = 上一行 +1
            try:
                ws.cell(new_row, 1).value = int(ws.cell(last_row, 1).value or 0) + 1
            except (TypeError, ValueError):
                pass

        ws.cell(new_row, 2).value = code           # B: 商品编号
        if c_value:
            ws.cell(new_row, 3).value = c_value     # C: 名称(空格后下划线前)

        if material_name:
            ws.cell(new_row, 5).value = int(hs_code)  # E: 海关编码
            if product_name:
                ws.cell(new_row, 6).value = product_name  # F: 商品名称(AI识别)
            ws.cell(new_row, 9).value = material_name  # I: 材料
            if usage:
                ws.cell(new_row, 10).value = usage   # J: 用途
        else:
            ws.cell(new_row, 18).value = unknown_text  # R: 未识别材料
            # 即使材料未匹配，仍写入AI识别的商品名称和用途
            if product_name:
                ws.cell(new_row, 6).value = product_name
            if usage:
                ws.cell(new_row, 10).value = usage

        wb.save(excel_path)
        return new_row


# ── 配置读写 ─────────────────────────────────────────────────────────

def get_config_for_ui():
    """返回给前端的配置（API Key 打码）"""
    cfg = _load_config()
    key = cfg.get('qianwen_api_key', '')
    masked = ('*' * max(len(key) - 4, 0) + key[-4:]) if len(key) >= 4 else ('*' * len(key))
    return {
        'qianwen_api_key_masked': masked,
        'has_key':        bool(key),
        'qianwen_model':  cfg.get('qianwen_model', 'qwen-vl-max'),
        'excel_path':     cfg.get('excel_path', ''),
    }


def save_config(updates):
    """
    updates = {
        "qianwen_api_key": "...",   # 空字符串表示不修改
        "qianwen_model":   "...",
        "excel_path":      "...",
    }
    """
    cfg = _load_config()
    if updates.get('qianwen_api_key'):
        cfg['qianwen_api_key'] = updates['qianwen_api_key']
    if 'qianwen_model' in updates and updates['qianwen_model']:
        cfg['qianwen_model'] = updates['qianwen_model']
    if 'excel_path' in updates:
        cfg['excel_path'] = updates['excel_path']
    with open(_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ── 公开入口 ─────────────────────────────────────────────────────────

def identify_and_save(payload):
    """
    payload = {
        "code":      "T.F240T-308",   # 必填
        "name":      "网格领结包_5",   # 可选，C列取空格后下划线前部分
        "category":  "Lady Bag",      # 可选，辅助 AI 判断
        "image_url": "https://...",   # 与 image_b64 二选一（优先用 url）
        "image_b64": "...",
    }
    返回 {
        "code", "ai_text", "material", "hs_code",
        "product_name", "usage", "row", "matched"
    }
    写入 Excel 列：A=项号, B=编码, C=名称, E=海关编码,
                   F=商品名称(AI), I=材料, J=用途, R=未识别材料
    """
    config  = _load_config()
    api_key = config.get('qianwen_api_key', '').strip()
    if not api_key:
        raise ValueError('请在设置面板填入千问 API Key')
    try:
        api_key.encode('ascii')
    except UnicodeEncodeError:
        raise ValueError('API Key 无效，请在设置面板填入真实的千问 API Key（不能含中文）')

    model     = config.get('qianwen_model', 'qwen-vl-max')
    code      = (payload.get('code') or '').strip()
    if not code:
        raise ValueError('缺少 code（商品编码）')

    image_url = (payload.get('image_url') or '').strip()
    image_b64 = (payload.get('image_b64') or '').strip()
    if not image_url and not image_b64:
        raise ValueError('image_url 和 image_b64 至少提供一个')

    excel_path = _find_excel(config)
    if not excel_path:
        raise FileNotFoundError(
            '未找到 报关产品基础资料（智谱）.xlsx\n'
            '请将文件放到 server.exe 同目录，或在 ai_config.json 中配置 excel_path'
        )

    # 已识别过则直接返回，不调用 AI
    existing = _check_existing(excel_path, code)
    if existing:
        return existing

    prompt        = _build_prompt(payload.get('name', ''), payload.get('category', ''))
    ai_text       = _call_qianwen(image_url or None, image_b64 or None, api_key, model, prompt)

    # 解析 AI 多行回复
    material_text, product_name, usage = _parse_ai_response(ai_text)
    material_name, hs_code = _match_material(material_text)

    # C 列：从 name 中取空格后、下划线前的部分
    # 例如 name="网格领结包_5" → c_value="网格领结包"
    raw_name = (payload.get('name') or '').strip()
    c_value = None
    if raw_name:
        idx = raw_name.find('_')
        c_value = raw_name[:idx] if idx > 0 else raw_name

    new_row = _write_row(excel_path, code, material_name, hs_code,
                         material_text if not material_name else None,
                         c_value=c_value, product_name=product_name, usage=usage)
    return {
        'code':        code,
        'ai_text':     ai_text,
        'material':    material_name,
        'hs_code':     hs_code,
        'product_name': product_name,
        'usage':       usage,
        'row':         new_row,
        'matched':     material_name is not None,
    }


# ── 历史对比 & JDY 扩展 ───────────────────────────────────────────────────────

def compare_with_history(material_name, excel_path, name_hint=None):
    """
    在基础资料表中查找 I 列 == material_name 的已有记录，
    统计最常见的组合作为历史参考。
    name_hint：若提供（报关名称），优先选名称相近的行作为参考。
    返回 dict（含 E/G/H/J/V/W 列值）或 None（无历史）。
    """
    if not material_name:
        return None
    counts = {}
    import time as _time
    last_err = None
    for _attempt in range(3):
        try:
            with _EXCEL_LOCK:
                wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                ws = wb.active
                try:
                    for row in ws.iter_rows(values_only=True):
                        if len(row) < 10:
                            continue
                        i_val = row[8]   # I 列（材料）
                        if i_val and str(i_val).strip().lower() == material_name.lower().strip():
                            hs   = str(int(row[4])) if row[4] else None          # E 海关编码
                            pn   = str(row[5]).strip() if row[5] else None        # F 报关名称
                            usg  = str(row[9]).strip() if row[9] else None        # J 用途
                            unit = str(row[6]).strip() if row[6] else None        # G 报关单位
                            orig = str(row[7]).strip() if row[7] else None        # H 境内货源地
                            tv   = str(row[21]).strip() if len(row) > 21 and row[21] else None  # V 税收编码
                            ew   = str(row[22]).strip() if len(row) > 22 and row[22] else None  # W 英文报关品名
                            key  = (hs, pn, usg, unit, orig, tv, ew)
                            counts[key] = counts.get(key, 0) + 1
                finally:
                    wb.close()
            break   # 成功，退出重试
        except (EOFError, Exception) as e:
            last_err = e
            _time.sleep(0.3)
            counts = {}
    else:
        print(f'[WARN] compare_with_history 读取失败（3次）: {last_err}')
        return None
    if not counts:
        return None

    if name_hint:
        # 同材料下，优先选与 name_hint 字符重叠最多的 F 列名称
        hint_chars = set(name_hint)
        def _score(k):
            pn_chars = set(k[1] or '')
            overlap = len(hint_chars & pn_chars) / max(len(hint_chars | pn_chars), 1)
            return (overlap, counts[k])
        best = max(counts, key=_score)
    else:
        best = max(counts, key=counts.get)

    return {
        'hs_code':        best[0],
        'product_name':   best[1],
        'usage':          best[2],
        'customs_unit':   best[3],
        'origin':         best[4],
        'tax_code':       best[5],
        'en_customs_name': best[6],
        'count':          counts[best],
    }


def _load_tax_table():
    """
    懒加载并缓存 商品和服务税收分类编码表.xls。
    返回 {分类简称: [(名称, 编码), ...]}，失败时返回 {}。
    """
    global _tax_table_cache
    if _tax_table_cache is not None:
        return _tax_table_cache
    result = {}
    xls_path = _TAX_TABLE_XLS
    if not os.path.exists(xls_path):
        print(f'[WARN] 未找到税收分类编码表: {xls_path}')
        _tax_table_cache = result
        return result
    try:
        import xlrd
        wb = xlrd.open_workbook(xls_path)
        sh = wb.sheet_by_index(0)
        for r in range(2, sh.nrows):
            code = str(sh.cell_value(r, 11)).strip()
            name = str(sh.cell_value(r, 12)).strip()
            cat  = str(sh.cell_value(r, 13)).strip()
            if code and name and cat:
                result.setdefault(cat, []).append((name, code))
        print(f'[TAX TABLE] 已加载 {len(result)} 个分类简称，共 {sum(len(v) for v in result.values())} 条')
    except ImportError:
        print('[WARN] 缺少 xlrd，无法读取 .xls 文件：pip install xlrd')
    except Exception as e:
        print(f'[WARN] 读取税收分类编码表失败: {e}')
    _tax_table_cache = result
    return result


def lookup_tax_code_from_table(name_part, cat_str):
    """
    在 商品和服务税收分类编码表.xls 中，按分类简称筛选后用名称模糊匹配，
    返回最匹配的税收编码，或 None。

    匹配优先级：
      1. 名称与 name_part 字符重叠最多的行
      2. 名称 == cat_str 的父级行（无具体品名匹配时）
      3. 名称以 '其他' 开头的兜底行
    """
    if not cat_str:
        return None
    table = _load_tax_table()
    entries = table.get(cat_str)
    if not entries:
        return None

    # 优先按 name_part 字符重叠打分
    if name_part:
        hint_chars = set(name_part.strip())
        best_code, best_score = None, -1
        for (ename, ecode) in entries:
            overlap = len(hint_chars & set(ename)) / max(len(hint_chars | set(ename)), 1)
            if overlap > best_score:
                best_score, best_code = overlap, ecode
        if best_score > 0:
            return best_code

    # 名称 == 分类简称 的父级行
    for (ename, ecode) in entries:
        if ename == cat_str:
            return ecode

    # "其他XXX" 兜底行
    for (ename, ecode) in entries:
        if ename.startswith('其他'):
            return ecode

    # 取第一条
    return entries[0][1]


def compare_by_category(cat_str, excel_path):
    """
    在基础资料表中查找 X 列（column 24）== cat_str 的已有记录，
    返回最常见的 V 列 tax_code（商品及服务税收编码）或 None。
    用于：手动输入 品名*材料*分类简称 时，通过分类简称反查 V 列。
    """
    if not cat_str or not excel_path:
        return None
    counts = {}
    try:
        with _EXCEL_LOCK:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
            try:
                for row in ws.iter_rows(values_only=True):
                    if len(row) < 24:
                        continue
                    x_val = row[23]  # X列 (0-indexed)
                    v_val = row[21]  # V列
                    if (x_val and str(x_val).strip() == cat_str.strip() and v_val):
                        tv = str(v_val).strip()
                        counts[tv] = counts.get(tv, 0) + 1
            finally:
                wb.close()
    except Exception as e:
        print(f'[WARN] compare_by_category 失败: {e}')
        return None
    return max(counts, key=counts.get) if counts else None


def _write_jdy_row(excel_path, code, material_name, hs_code, unknown_text,
                   c_value=None, d_value=None, product_name=None, usage=None,
                   h_value=None, dimensions=None, conversion=None,
                   supplier_number=None, account=None, address_detail=None,
                   customs_unit=None, tax_code=None, en_customs_name=None,
                   project_name=None):
    """
    写入 JDY 自动补全数据（含 D/G/H/N/O/P/Q/R/S/T/U/V/W 列）。
    dimensions:      {'l': float, 'w': float, 'h': float, 'vol': float} 或 None
    conversion:      换算系数（数字或 None）
    supplier_number: 供应商编号，写 T 列；已有值则逗号追加（去重）
    account:         账套名称（如"饰品"/"箱包"），写 U 列
    address_detail:  供应商首要联系人详细地址，写 R 列
    customs_unit:    报关单位（G 列），来自同材料历史行
    tax_code:        商品及服务税收编码（V 列），来自同材料历史行
    en_customs_name: 英文报关品名（W 列），来自同材料历史行
    project_name:    商品和服务分类简称（X 列），手动输入的第三部分
    """
    with _EXCEL_LOCK:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        existing_row = _find_existing_row(ws, code)
        if existing_row:
            new_row = existing_row
        else:
            last_row = 1
            for r in range(ws.max_row, 0, -1):
                if ws.cell(r, 2).value is not None:
                    last_row = r
                    break
            new_row = last_row + 1
            try:
                ws.cell(new_row, 1).value = int(ws.cell(last_row, 1).value or 0) + 1
            except (TypeError, ValueError):
                pass

        ws.cell(new_row, 2).value = code                        # B: 商品编号
        if c_value:
            ws.cell(new_row, 3).value = c_value                 # C: 名称
        if d_value:
            ws.cell(new_row, 4).value = d_value                 # D: 商品类别
        if h_value:
            ws.cell(new_row, 8).value = h_value                 # H: 境内货源地
        if material_name:
            if hs_code:
                ws.cell(new_row, 5).value = int(hs_code)        # E: 海关编码
            if product_name:
                ws.cell(new_row, 6).value = product_name        # F: 商品名称(AI)
            ws.cell(new_row, 9).value = material_name           # I: 材料
            if usage:
                ws.cell(new_row, 10).value = usage              # J: 用途
        else:
            if product_name:
                ws.cell(new_row, 6).value = product_name
            if usage:
                ws.cell(new_row, 10).value = usage
        if address_detail:
            ws.cell(new_row, 18).value = address_detail         # R: 详细地址
        if dimensions:
            ws.cell(new_row, 14).value = dimensions.get('l')    # N: 长
            ws.cell(new_row, 15).value = dimensions.get('w')    # O: 宽
            ws.cell(new_row, 16).value = dimensions.get('h')    # P: 高
            ws.cell(new_row, 17).value = dimensions.get('vol')  # Q: 体积
        if conversion is not None:
            ws.cell(new_row, 19).value = conversion             # S: 换算系数
        # G: 报关单位（来自同材料历史行）
        if customs_unit and not ws.cell(new_row, 7).value:
            ws.cell(new_row, 7).value = customs_unit
        # T: 供应商编号 — 追加（逗号分隔，去重）
        if supplier_number:
            existing_sup = str(ws.cell(new_row, 20).value or '').strip()
            existing_set = set(x.strip() for x in existing_sup.split(',') if x.strip())
            existing_set.add(str(supplier_number).strip())
            ws.cell(new_row, 20).value = ','.join(sorted(existing_set))  # T
        # U: 账套
        if account:
            ws.cell(new_row, 21).value = account                         # U
        # V: 商品及服务税收编码（来自同材料历史行）
        if tax_code and not ws.cell(new_row, 22).value:
            ws.cell(new_row, 22).value = tax_code
        # W: 英文报关品名（来自同材料历史行）
        if en_customs_name and not ws.cell(new_row, 23).value:
            ws.cell(new_row, 23).value = en_customs_name
        # X: 商品和服务分类简称（手动输入的第三部分 / 凌航发票项目名）
        if project_name and not ws.cell(new_row, 24).value:
            ws.cell(new_row, 24).value = project_name

        wb.save(excel_path)
        return new_row


def identify_from_jdy(payload, extra_fields=None, confirmed=None):
    """
    JDY 自动补全入口。

    payload 与 identify_and_save 相同格式。
    extra_fields (可选): {
        'category':    str,   # D列 — 来自 JDY categoryName
        'origin':      str,   # H列 — 来自供应商 remark
        'dimensions':  dict,  # N/O/P/Q — parse_dimensions() 结果
        'conversion':  float, # S列 — 来自 JDY remark
        'account':     str,   # U列 — 账套名称（饰品/箱包）
    }
    confirmed (可选): 用户从弹窗选择的覆盖结果 {
        'material':     str,
        'hs_code':      str,
        'product_name': str,
        'usage':        str,
    }

    返回 dict，新增字段:
        'needs_confirm':  bool
        'ai_result':      dict  (AI 识别的原始结果)
        'history_result': dict or None  (历史参考)
    """
    config  = _load_config()
    api_key = config.get('qianwen_api_key', '').strip()
    if not api_key:
        raise ValueError('请在设置面板填入千问 API Key')

    model      = config.get('qianwen_model', 'qwen-vl-max')
    code       = (payload.get('code') or '').strip()
    if not code:
        raise ValueError('缺少 code（商品编码）')

    image_url = (payload.get('image_url') or '').strip()
    image_b64 = (payload.get('image_b64') or '').strip()
    if not image_url and not image_b64:
        raise ValueError('image_url 和 image_b64 至少提供一个')

    excel_path = _find_excel(config)
    if not excel_path:
        raise FileNotFoundError('未找到 报关产品基础资料（智谱）.xlsx')

    extra = extra_fields or {}

    # 已完整识别过：跳过 AI，但仍把 JDY 额外字段（H/S/T/尺寸）补写进去
    existing = _check_existing(excel_path, code)
    if existing:
        c_value = _extract_c_value(payload.get('name', ''))
        _write_jdy_row(
            excel_path, code,
            existing.get('material'),
            existing.get('hs_code'),
            unknown_text=None,
            c_value=c_value,
            d_value=extra.get('category'),
            product_name=existing.get('product_name'),
            usage=existing.get('usage'),
            h_value=extra.get('origin'),
            dimensions=extra.get('dimensions'),
            conversion=extra.get('conversion'),
            supplier_number=extra.get('supplier_number'),
            account=extra.get('account'),
            address_detail=extra.get('address_detail'),
        )
        existing['needs_confirm'] = False
        return existing

    # 如果用户已确认（二次调用），直接写入确认结果
    if confirmed:
        material_name = confirmed.get('material')
        hs_code       = confirmed.get('hs_code')
        product_name  = confirmed.get('product_name')
        usage         = confirmed.get('usage')
        c_value = _extract_c_value(payload.get('name', ''))
        new_row = _write_jdy_row(
            excel_path, code, material_name, hs_code,
            unknown_text=None,
            c_value=c_value,
            d_value=extra.get('category'),
            product_name=product_name,
            usage=usage,
            h_value=extra.get('origin'),
            dimensions=extra.get('dimensions'),
            conversion=extra.get('conversion'),
            supplier_number=extra.get('supplier_number'),
            account=extra.get('account'),
            address_detail=extra.get('address_detail'),
        )
        return {
            'code': code, 'material': material_name, 'hs_code': hs_code,
            'product_name': product_name, 'usage': usage,
            'row': new_row, 'matched': bool(material_name),
            'needs_confirm': False, 'confirmed_by_user': True,
        }

    # 调用 AI
    prompt = _build_prompt(payload.get('name', ''), payload.get('category', ''))
    ai_text = _call_qianwen(image_url or None, image_b64 or None, api_key, model, prompt)

    material_text, product_name, usage = _parse_ai_response(ai_text)
    material_name, hs_code = _match_material(material_text)

    ai_result = {
        'material':     material_name or material_text,
        'hs_code':      hs_code,
        'product_name': product_name,
        'usage':        usage,
    }

    # 与历史对比
    history = compare_with_history(material_name, excel_path) if material_name else None
    needs_confirm = _should_confirm(ai_result, history)

    if not needs_confirm:
        # 直接写入
        c_value = _extract_c_value(payload.get('name', ''))
        new_row = _write_jdy_row(
            excel_path, code, material_name, hs_code,
            unknown_text=material_text if not material_name else None,
            c_value=c_value,
            d_value=extra.get('category'),
            product_name=product_name,
            usage=usage,
            h_value=extra.get('origin'),
            dimensions=extra.get('dimensions'),
            conversion=extra.get('conversion'),
            supplier_number=extra.get('supplier_number'),
            account=extra.get('account'),
            address_detail=extra.get('address_detail'),
        )
        return {
            'code': code, 'ai_text': ai_text,
            'material': material_name, 'hs_code': hs_code,
            'product_name': product_name, 'usage': usage,
            'row': new_row, 'matched': material_name is not None,
            'needs_confirm': False,
            'ai_result': ai_result, 'history_result': history,
        }

    # 需要用户确认：不写入，返回两侧数据
    return {
        'code':  code, 'ai_text': ai_text,
        'needs_confirm': True,
        'ai_result':     ai_result,
        'history_result': history,
    }


def _extract_c_value(raw_name):
    raw_name = (raw_name or '').strip()
    if not raw_name:
        return None
    idx = raw_name.find('_')
    return raw_name[:idx] if idx > 0 else raw_name


def _should_confirm(ai_result, history):
    """对比 AI 识别结果与历史记录，决定是否弹窗确认"""
    if not history:
        return False   # 无历史，直接用 AI 结果
    if ai_result.get('hs_code') != history.get('hs_code'):
        return True
    if ai_result.get('product_name') != history.get('product_name'):
        return True
    if ai_result.get('usage') != history.get('usage'):
        return True
    return False


# ── 档案补全（生产许可证解析）──────────────────────────────────────────────

def parse_prolicense(pro_license):
    """
    解析生产许可证字段，格式: "报关名称*材料"  （* 为分隔符）
    返回 (name_part, material_part) 或 (None, None)
    """
    if not pro_license:
        return None, None
    parts = str(pro_license).split('*', 1)
    if len(parts) == 2 and parts[0].strip() and parts[1].strip():
        return parts[0].strip(), parts[1].strip()
    return None, None


def preview_from_license(code, name_part, material_part, extra_fields=None):
    """
    预览档案补全结果（不写入 Excel），供前端展示可编辑确认表格。
    返回 {'code', 'product_name', 'material', 'hs_code', 'usage', 'matched', ...}
    """
    extra = extra_fields or {}
    material_name, hs_code = _match_material(material_part)
    if not material_name:
        material_name = material_part.strip()

    excel_path = _find_excel(_load_config())
    history = compare_with_history(material_name, excel_path) if (material_name and excel_path) else None
    usage = history.get('usage') if history else None
    if not hs_code and history and history.get('hs_code'):
        hs_code = history['hs_code']

    return {
        'code':         code,
        'product_name': name_part,
        'material':     material_name,
        'hs_code':      hs_code or '',
        'usage':        usage or '',
        'matched':      hs_code is not None,
        'category':     extra.get('category', ''),
        'account':      extra.get('account', ''),
        'dimensions':   extra.get('dimensions'),
        'conversion':   extra.get('conversion'),
    }


def fill_from_license(excel_path, code, name_part, material_part, extra_fields=None, cat_str=None):
    """
    用生产许可证解析出的信息补全 Excel，不调用 AI。
      name_part     → F 列（报关名称）
      material_part → 用于关键词匹配 → E（海关编码）、I（材料）
                      无匹配时直接将原文写入 I 列
      用途          → 从 Excel 历史记录中查同材料最常见的 J 值
      cat_str       → 商品和服务分类简称 → X 列；若 V 列历史无值则从同分类简称行反查

    返回 {'code', 'material', 'hs_code', 'product_name', 'usage', 'row', 'matched'}
    """
    extra = extra_fields or {}

    # 关键词匹配
    material_name, hs_code = _match_material(material_part)
    if not material_name:
        # 无关键词命中：直接写原文到 I，E 留空（或从历史补）
        material_name = material_part.strip()

    # 从历史记录查同材料最常见的组合：E/G/H/J/V/W 列
    # name_hint 让优先选与 name_part 字符重叠多的行
    history = compare_with_history(material_name, excel_path, name_hint=name_part)
    usage        = history.get('usage')          if history else None
    customs_unit = history.get('customs_unit')   if history else None
    origin_hist  = history.get('origin')         if history else None
    tax_code     = history.get('tax_code')       if history else None
    en_cname     = history.get('en_customs_name') if history else None
    # 历史记录（已人工确认过的数据）优先于 _RULES 关键词匹配
    # _RULES 仅作无历史时兜底（原设计只针对箱包账套，用在饰品类会产生错误编码）
    if history and history.get('hs_code'):
        hs_code = history['hs_code']

    # 若提供了分类简称（cat_str）且历史无 tax_code，
    # 先从基础资料 X 列相同行反查，再从税收分类编码表 xls 兜底
    if cat_str and not tax_code:
        tax_code = compare_by_category(cat_str, excel_path)
        if tax_code:
            print(f'[LICENSE] {code} 通过历史 X 列 "{cat_str}" 补全 tax_code: {tax_code}')
        else:
            tax_code = lookup_tax_code_from_table(name_part, cat_str)
            if tax_code:
                print(f'[LICENSE] {code} 通过税收编码表 "{cat_str}/{name_part}" 补全 tax_code: {tax_code}')

    new_row = _write_jdy_row(
        excel_path, code,
        material_name, hs_code,
        unknown_text=None,
        c_value=_extract_c_value(name_part),
        d_value=extra.get('category'),
        product_name=name_part,          # F: 报关名称
        usage=usage,
        h_value=extra.get('origin') or origin_hist,   # 优先传入值，次用历史
        dimensions=extra.get('dimensions'),
        conversion=extra.get('conversion'),
        supplier_number=extra.get('supplier_number'),
        account=extra.get('account'),
        address_detail=extra.get('address_detail'),
        customs_unit=customs_unit,
        tax_code=tax_code,
        en_customs_name=en_cname,
        project_name=cat_str or None,    # X: 分类简称
    )
    return {
        'code':         code,
        'material':     material_name,
        'hs_code':      hs_code,
        'product_name': name_part,
        'usage':        usage,
        'row':          new_row,
        'matched':      hs_code is not None,
    }


def write_confirmed_license(excel_path, code, product_name, material, hs_code, usage, extra_fields=None):
    """
    将用户在确认弹窗中（可能已修改的）值写入 Excel。
    不再重新调用 AI 或匹配规则，直接写入传入的值。
    返回 {'code', 'row', 'success'}
    """
    extra = extra_fields or {}
    new_row = _write_jdy_row(
        excel_path, code,
        material_name=material,
        hs_code=hs_code or None,
        unknown_text=None,
        c_value=_extract_c_value(product_name) if product_name else None,
        d_value=extra.get('category'),
        product_name=product_name,
        usage=usage,
        h_value=extra.get('origin'),
        dimensions=extra.get('dimensions'),
        conversion=extra.get('conversion'),
        supplier_number=extra.get('supplier_number'),
        account=extra.get('account'),
        address_detail=extra.get('address_detail'),
    )
    return {'code': code, 'row': new_row, 'success': True}
